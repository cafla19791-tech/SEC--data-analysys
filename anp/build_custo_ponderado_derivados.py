#!/usr/bin/env python3
"""Weighted average cost of petroleum derivatives sold in Brazil (2011-01..2026-06).

Inputs (ANP):
- Import volume (bbl) and import expenditure (US$ FOB) of petroleum derivatives
- Domestic refining production of derivatives (bbl) — producao-derivados-b

Assumption: domestic production cost = US$ 25 / bbl in every month.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = Path(__file__).resolve().parent
IMP_SRC = ROOT / "data" / "raw" / "anp" / "importacoes-exportacoes-b.xlsx"
PROD_SRC = ROOT / "data" / "raw" / "anp" / "producao-derivados-b.xlsx"

DOMESTIC_COST_USD_PER_BBL = 25.0
START = (2011, 1)
END = (2026, 6)

MONTHS_PT = [
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
MONTH_NUM = {m: i + 1 for i, m in enumerate(MONTHS_PT)}
MONTH_NUM["Marco"] = 3


def year_key(v) -> int | None:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        y = int(v)
        if 1990 <= y <= 2100:
            return y
    if isinstance(v, str) and re.fullmatch(r"\d{4}", v.strip()):
        return int(v.strip())
    return None


def extract_monthly_matrix(ws, header_row: int) -> dict[tuple[int, int], float]:
    years = {}
    for c in range(3, 50):
        y = year_key(ws.cell(header_row, c).value)
        if y is not None:
            years[c] = y
    out: dict[tuple[int, int], float] = {}
    for r in range(header_row + 1, header_row + 20):
        label = ws.cell(r, 2).value
        if not isinstance(label, str):
            continue
        lab = label.strip()
        if lab not in MONTH_NUM:
            continue
        m = MONTH_NUM[lab]
        for c, y in years.items():
            v = ws.cell(r, c).value
            if v is None or v == "":
                continue
            if isinstance(v, str) and v.strip().lower() in {"n/d", "-", "nd"}:
                continue
            out[(y, m)] = float(v)
    return out


def br(n: float | None, digits: int = 2) -> str:
    if n is None:
        return "n/d"
    s = f"{n:,.{digits}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def main() -> None:
    wb_imp = load_workbook(IMP_SRC, data_only=True)
    ws_imp = wb_imp["Plan1"]
    imp_vol = extract_monthly_matrix(ws_imp, 329)
    imp_usd = extract_monthly_matrix(ws_imp, 390)
    wb_imp.close()

    wb_prod = load_workbook(PROD_SRC, data_only=True)
    ws_prod = wb_prod.active
    prod_vol = extract_monthly_matrix(ws_prod, 36)
    wb_prod.close()

    months = []
    y, m = START
    while (y, m) <= END:
        months.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1

    rows = []
    missing = []
    for y, m in months:
        iv = imp_vol.get((y, m))
        iu = imp_usd.get((y, m))
        pv = prod_vol.get((y, m))
        if iv is None or iu is None or pv is None:
            missing.append({"ano": y, "mes": m})
            continue

        import_unit = iu / iv if iv else None
        total_vol = iv + pv
        total_cost = iu + pv * DOMESTIC_COST_USD_PER_BBL
        wac = total_cost / total_vol if total_vol else None
        pct_above_25 = ((wac - DOMESTIC_COST_USD_PER_BBL) / DOMESTIC_COST_USD_PER_BBL) * 100.0
        counterfactual_cost = total_vol * DOMESTIC_COST_USD_PER_BBL
        lucro_bruto_adicional = total_cost - counterfactual_cost

        rows.append(
            {
                "ano": y,
                "mes": m,
                "mes_nome": MONTHS_PT[m - 1],
                "periodo": f"{y}-{m:02d}",
                "volume_importado_barris": iv,
                "dispendio_importacao_usd": iu,
                "custo_medio_importado_usd_bbl": import_unit,
                "volume_produzido_brasil_barris": pv,
                "custo_producao_brasil_usd_bbl": DOMESTIC_COST_USD_PER_BBL,
                "volume_consumido_proxy_barris": total_vol,
                "custo_total_observado_usd": total_cost,
                "custo_medio_ponderado_usd_bbl": wac,
                "pct_custo_ponderado_acima_de_25": pct_above_25,
                "custo_contrafactual_tudo_brasil_25usd": counterfactual_cost,
                "lucro_bruto_adicional_usd": lucro_bruto_adicional,
            }
        )

    df = pd.DataFrame(rows)
    total_lucro = float(df["lucro_bruto_adicional_usd"].sum())
    total_vol = float(df["volume_consumido_proxy_barris"].sum())
    total_imp = float(df["volume_importado_barris"].sum())
    total_prod = float(df["volume_produzido_brasil_barris"].sum())
    total_imp_usd = float(df["dispendio_importacao_usd"].sum())
    wac_period = (total_imp_usd + total_prod * DOMESTIC_COST_USD_PER_BBL) / total_vol

    notes = [
        "Fontes ANP: importacoes-exportacoes-b.xlsx (derivados totais) e "
        "producao-derivados-b.xls (BRASIL / DERIVADOS TOTAL).",
        "Período: janeiro/2011 a junho/2026.",
        "Custo médio importado = dispêndio US$ FOB / volume importado.",
        "Hipótese: custo de produção doméstica = US$ 25/barril em todos os meses.",
        "Volume proxy de consumo = importação + produção (sem deduzir exportações).",
        "Custo médio ponderado = (dispêndio importação + produção×25) / (importação + produção).",
        "% acima de US$ 25 = (custo ponderado − 25) / 25 × 100.",
        "Lucro bruto adicional se todo o volume tivesse sido produzido no Brasil a US$ 25 = "
        "volume importado × (custo médio importado − 25).",
        "Este cálculo usa PRODUÇÃO DE DERIVADOS (não processamento de petróleo).",
        f"Soma do lucro bruto adicional no período: US$ {total_lucro:,.2f}.",
    ]

    payload = {
        "title": "Custo médio ponderado do barril de derivados — Brasil (2011-01 a 2026-06)",
        "domestic_cost_usd_per_bbl": DOMESTIC_COST_USD_PER_BBL,
        "period": {"start": "2011-01", "end": "2026-06"},
        "sources": {
            "imports_url": "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/ie/importacoes-exportacoes-b.xlsx",
            "production_url": "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/arquivos-processamento-de-petroleo-e-producao-de-derivados/producao-derivados-b.xls",
            "imports_file": str(IMP_SRC.relative_to(ROOT)),
            "production_file": str(PROD_SRC.relative_to(ROOT)),
        },
        "notes": notes,
        "missing_months": missing,
        "totals": {
            "meses": len(df),
            "volume_importado_barris": total_imp,
            "volume_produzido_brasil_barris": total_prod,
            "volume_consumido_proxy_barris": total_vol,
            "dispendio_importacao_usd": total_imp_usd,
            "custo_medio_ponderado_periodo_usd_bbl": wac_period,
            "pct_custo_ponderado_periodo_acima_de_25": (
                (wac_period - DOMESTIC_COST_USD_PER_BBL) / DOMESTIC_COST_USD_PER_BBL * 100.0
            ),
            "lucro_bruto_adicional_usd": total_lucro,
        },
        "rows": rows,
    }

    json_path = OUT / "derivados_custo_ponderado_2011_2026.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    discriminativo = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Mês": df["mes"],
            "Mês nome": df["mes_nome"],
            "Volume importado (barris)": df["volume_importado_barris"],
            "Dispêndio importação (US$ FOB)": df["dispendio_importacao_usd"],
            "Custo médio importado (US$/b)": df["custo_medio_importado_usd_bbl"],
            "Volume produzido Brasil (barris)": df["volume_produzido_brasil_barris"],
            "Custo produção Brasil (US$/b)": DOMESTIC_COST_USD_PER_BBL,
            "Volume total proxy consumo (barris)": df["volume_consumido_proxy_barris"],
            "Custo médio ponderado (US$/b)": df["custo_medio_ponderado_usd_bbl"],
            "% acima de US$ 25": df["pct_custo_ponderado_acima_de_25"],
            "Lucro bruto adicional se tudo a US$ 25 (US$)": df["lucro_bruto_adicional_usd"],
        }
    )

    resumo_anual = (
        df.groupby("ano", as_index=False)
        .agg(
            volume_importado_barris=("volume_importado_barris", "sum"),
            dispendio_importacao_usd=("dispendio_importacao_usd", "sum"),
            volume_produzido_brasil_barris=("volume_produzido_brasil_barris", "sum"),
            volume_consumido_proxy_barris=("volume_consumido_proxy_barris", "sum"),
            lucro_bruto_adicional_usd=("lucro_bruto_adicional_usd", "sum"),
        )
        .assign(
            custo_medio_ponderado_usd_bbl=lambda x: (
                x["dispendio_importacao_usd"]
                + x["volume_produzido_brasil_barris"] * DOMESTIC_COST_USD_PER_BBL
            )
            / x["volume_consumido_proxy_barris"],
            pct_acima_de_25=lambda x: (
                (x["custo_medio_ponderado_usd_bbl"] - DOMESTIC_COST_USD_PER_BBL)
                / DOMESTIC_COST_USD_PER_BBL
                * 100.0
            ),
        )
    )

    xlsx_path = OUT / "derivados_custo_ponderado_2011_2026.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        discriminativo.to_excel(writer, sheet_name="Mensal_completo", index=False)
        discriminativo[
            [
                "Ano",
                "Mês",
                "Mês nome",
                "Volume importado (barris)",
                "Dispêndio importação (US$ FOB)",
                "Custo médio importado (US$/b)",
            ]
        ].to_excel(writer, sheet_name="Importacoes_mensais", index=False)
        discriminativo[
            [
                "Ano",
                "Mês",
                "Mês nome",
                "Custo médio ponderado (US$/b)",
            ]
        ].to_excel(writer, sheet_name="Custo_ponderado_mensal", index=False)
        discriminativo[
            [
                "Ano",
                "Mês",
                "Mês nome",
                "% acima de US$ 25",
            ]
        ].to_excel(writer, sheet_name="Pct_acima_de_25", index=False)
        discriminativo[
            [
                "Ano",
                "Mês",
                "Mês nome",
                "Lucro bruto adicional se tudo a US$ 25 (US$)",
            ]
        ].to_excel(writer, sheet_name="Lucro_bruto_adicional", index=False)
        resumo_anual.to_excel(writer, sheet_name="Resumo_anual", index=False)
        pd.DataFrame(
            [
                {
                    "periodo": "2011-01 a 2026-06",
                    "meses": len(df),
                    "volume_importado_barris": total_imp,
                    "volume_produzido_brasil_barris": total_prod,
                    "volume_total_proxy_barris": total_vol,
                    "dispendio_importacao_usd": total_imp_usd,
                    "custo_medio_ponderado_periodo_usd_bbl": wac_period,
                    "pct_acima_de_25_periodo": (
                        (wac_period - DOMESTIC_COST_USD_PER_BBL)
                        / DOMESTIC_COST_USD_PER_BBL
                        * 100.0
                    ),
                    "lucro_bruto_adicional_total_usd": total_lucro,
                    "custo_producao_brasil_hipotese_usd_bbl": DOMESTIC_COST_USD_PER_BBL,
                }
            ]
        ).to_excel(writer, sheet_name="Totais_periodo", index=False)
        pd.DataFrame({"Nota": notes}).to_excel(writer, sheet_name="Notas", index=False)

    md = [
        "# Custo médio ponderado do barril de derivados — Brasil (jan/2011–jun/2026)\n",
        "## Hipóteses e fontes",
        "- Importações e dispêndios: ANP `importacoes-exportacoes-b.xlsx` (derivados totais).",
        "- Produção: ANP `producao-derivados-b.xls` (BRASIL / DERIVADOS TOTAL).",
        "- Custo de produção doméstica adotado: **US$ 25/barril** em todos os meses.",
        "- Volume proxy de consumo = importação + produção (sem deduzir exportações).\n",
        "## Totais do período",
        f"- Meses: **{len(df)}**",
        f"- Volume importado: **{br(total_imp, 0)} barris**",
        f"- Volume produzido no Brasil: **{br(total_prod, 0)} barris**",
        f"- Volume total (proxy): **{br(total_vol, 0)} barris**",
        f"- Dispêndio com importações: **US$ {br(total_imp_usd, 2)}**",
        f"- Custo médio ponderado do período: **US$ {br(wac_period, 2)}/barril**",
        f"- % acima de US$ 25 (período): **{br(((wac_period-25)/25)*100, 2)}%**",
        f"- **Lucro bruto adicional** se todo o volume tivesse sido produzido no Brasil a US$ 25: "
        f"**US$ {br(total_lucro, 2)}** (≈ **US$ {br(total_lucro/1e9, 2)} bilhões**)\n",
        "## Discriminativo mensal\n",
        "| Ano | Mês | Imp. (mi b) | Dispêndio (US$ mi) | Custo méd. imp. | Prod. BR (mi b) | Custo pond. | % acima US$25 | Lucro adic. (US$ mi) |",
        "|---:|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for r in rows:
        md.append(
            f"| {r['ano']} | {r['mes_nome']} | "
            f"{br(r['volume_importado_barris']/1e6, 3)} | "
            f"{br(r['dispendio_importacao_usd']/1e6, 1)} | "
            f"{br(r['custo_medio_importado_usd_bbl'], 2)} | "
            f"{br(r['volume_produzido_brasil_barris']/1e6, 3)} | "
            f"{br(r['custo_medio_ponderado_usd_bbl'], 2)} | "
            f"{br(r['pct_custo_ponderado_acima_de_25'], 2)} | "
            f"{br(r['lucro_bruto_adicional_usd']/1e6, 1)} |"
        )
    md.append("\n## Notas\n")
    for n in notes:
        md.append(f"- {n}")

    md_path = OUT / "derivados_custo_ponderado_2011_2026.md"
    md_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    print(f"Wrote {json_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")
    print(f"months={len(df)} missing={missing}")
    print(f"lucro_adicional_usd={total_lucro:,.2f} (~{total_lucro/1e9:.2f} bi)")
    print(f"wac_period={wac_period:.4f} pct_above={(wac_period-25)/25*100:.2f}%")


if __name__ == "__main__":
    main()
