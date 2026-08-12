#!/usr/bin/env python3
"""Parse ANP Importações e Exportações (barris) workbook into tidy annual tables."""

from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "raw" / "anp" / "importacoes-exportacoes-b.xlsx"
OUT = Path(__file__).resolve().parent

# (title_row_1based, series_id, unit, description)
# title_row points to the section title cell in column B.
SECTIONS = [
    (42, "petroleo_importacao_volume", "barris", "Importação de petróleo"),
    (97, "petroleo_importacao_dispendio", "US$ FOB", "Dispêndio com importação de petróleo"),
    (153, "petroleo_importacao_preco_medio", "US$/b FOB", "Preço médio do barril de petróleo importado"),
    (210, "petroleo_exportacao_volume", "barris", "Exportação de petróleo"),
    (266, "petroleo_exportacao_receita", "US$ FOB", "Receita com exportação de petróleo"),
    (322, "derivados_importacao_volume", "barris", "Importação de derivados (total)"),
    (383, "derivados_importacao_dispendio", "US$ FOB", "Dispêndio com importação de derivados (total)"),
    (443, "derivados_exportacao_volume", "barris", "Exportação de derivados (total)"),
    (504, "derivados_exportacao_receita", "US$ FOB", "Receita com exportação de derivados (total)"),
    (568, "gas_importacao_volume", "bep", "Importação de gás natural"),
    (623, "gas_importacao_dispendio", "US$ FOB", "Dispêndio com importação de gás natural"),
    (680, "etanol_anidro_importacao_volume", "barris", "Importação de etanol anidro"),
    (734, "etanol_anidro_importacao_dispendio", "US$ FOB", "Dispêndio com importação de etanol anidro"),
    (789, "etanol_hidratado_importacao_volume", "barris", "Importação de etanol hidratado"),
    (843, "etanol_hidratado_importacao_dispendio", "US$ FOB", "Dispêndio com importação de etanol hidratado"),
    (899, "etanol_anidro_exportacao_volume", "barris", "Exportação de etanol anidro"),
    (952, "etanol_anidro_exportacao_receita", "US$ FOB", "Receita com exportação de etanol anidro"),
    (1008, "etanol_hidratado_exportacao_volume", "barris", "Exportação de etanol hidratado"),
    (1062, "etanol_hidratado_exportacao_receita", "US$ FOB", "Receita com exportação de etanol hidratado"),
]


def year_key(v) -> int | None:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        y = int(v)
        if 1990 <= y <= 2100:
            return y
    if isinstance(v, str) and re.fullmatch(r"\d{4}", v.strip()):
        return int(v.strip())
    return None


def find_header_and_total(ws, start_row: int, max_scan: int = 40):
    """Locate year header row and 'Total do Ano' within a section."""
    header_row = None
    years = {}
    total_row = None
    for r in range(start_row, start_row + max_scan):
        label = ws.cell(r, 2).value
        if isinstance(label, str) and label.strip().upper() in {"MÊS", "MES", "MESES"}:
            header_row = r
            for c in range(3, 50):
                y = year_key(ws.cell(r, c).value)
                if y is not None:
                    years[c] = y
            break
    if header_row is None:
        raise ValueError(f"No month header near row {start_row}")
    for r in range(header_row + 1, header_row + 20):
        label = ws.cell(r, 2).value
        if not isinstance(label, str):
            continue
        low = label.strip().lower()
        if "total do ano" in low or "média do ano" in low or "media do ano" in low:
            total_row = r
            break
    if total_row is None:
        raise ValueError(f"No Total/Média do Ano near row {start_row}")
    return years, total_row


def extract_annual(ws, start_row: int) -> dict[int, float | None]:
    years, total_row = find_header_and_total(ws, start_row)
    out: dict[int, float | None] = {}
    for c, y in years.items():
        v = ws.cell(total_row, c).value
        if v is None or v == "" or (isinstance(v, str) and v.strip().lower() in {"n/d", "-"}):
            out[y] = None
        else:
            out[y] = float(v)
    return out


def br(n: float | None, digits: int = 1) -> str:
    if n is None:
        return "n/d"
    s = f"{n:,.{digits}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")

    wb = load_workbook(SRC, data_only=True)
    ws = wb["Plan1"]

    updated = None
    for r in range(1, 20):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and "atualizados" in v.lower():
            updated = v.strip()
            break

    series_map: dict[str, dict] = {}
    for start, sid, unit, desc in SECTIONS:
        annual = extract_annual(ws, start)
        series_map[sid] = {
            "id": sid,
            "description": desc,
            "unit": unit,
            "section_start_row": start,
            "annual": {str(y): annual[y] for y in sorted(annual)},
        }

    wb.close()

    # Wide annual table 2000-2025 (2026 may be partial YTD in source)
    years = list(range(2000, 2026))
    rows = []
    for y in years:
        row = {"ano": y}
        for sid, meta in series_map.items():
            row[sid] = meta["annual"].get(str(y))
        # Convenience aggregates
        pet_imp_usd = row["petroleo_importacao_dispendio"]
        pet_exp_usd = row["petroleo_exportacao_receita"]
        der_imp_usd = row["derivados_importacao_dispendio"]
        der_exp_usd = row["derivados_exportacao_receita"]
        gas_imp_usd = row["gas_importacao_dispendio"]
        if None not in (pet_imp_usd, der_imp_usd, gas_imp_usd):
            row["importacao_energia_usd"] = pet_imp_usd + der_imp_usd + gas_imp_usd
        else:
            row["importacao_energia_usd"] = None
        if None not in (pet_exp_usd, der_exp_usd):
            row["exportacao_petroleo_derivados_usd"] = pet_exp_usd + der_exp_usd
        else:
            row["exportacao_petroleo_derivados_usd"] = None
        if (
            row["importacao_energia_usd"] is not None
            and row["exportacao_petroleo_derivados_usd"] is not None
        ):
            row["saldo_comercio_petroleo_derivados_gas_usd"] = (
                row["exportacao_petroleo_derivados_usd"] - row["importacao_energia_usd"]
            )
        else:
            row["saldo_comercio_petroleo_derivados_gas_usd"] = None
        rows.append(row)

    notes = [
        "Fonte: ANP — Importações e Exportações (barris). "
        f"URL: https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/ie/importacoes-exportacoes-b.xlsx",
        updated or "Data de atualização não identificada no cabeçalho.",
        "Valores anuais = linha 'Total do Ano' de cada bloco da planilha.",
        "Derivados: totais agregados da visão 'DERIVADOS TOTAL' / produto '(Tudo)' "
        "quando a planilha usa filtro de produto.",
        "FOB: free on board; dólares correntes. Volume de petróleo/derivados/etanol em barris; "
        "gás natural em barris equivalentes de petróleo (bep).",
        "A partir de nov/2006, exportações de derivados incluem combustíveis para aeronaves e navios "
        "(série revisada desde 2000, conforme nota da ANP).",
        "2026 no arquivo-fonte pode estar parcial (YTD); as tabelas tidy cobrem 2000–2025.",
        "Saldo comércio = receita exportação petróleo+derivados − dispêndio importação "
        "petróleo+derivados+gás (não inclui etanol).",
    ]

    payload = {
        "title": "ANP — Importações e Exportações (barris), séries anuais",
        "source_file": str(SRC.relative_to(ROOT)),
        "source_url": "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-estatisticos/de/ie/importacoes-exportacoes-b.xlsx",
        "updated": updated,
        "notes": notes,
        "series": series_map,
        "rows": rows,
    }

    json_path = OUT / "anp_importacoes_exportacoes_anual_2000_2025.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    df = pd.DataFrame(rows)

    # Human sheets
    petroleo = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Importação (barris)": df["petroleo_importacao_volume"],
            "Dispêndio importação (US$ FOB)": df["petroleo_importacao_dispendio"],
            "Preço médio (US$/b FOB)": df["petroleo_importacao_preco_medio"],
            "Exportação (barris)": df["petroleo_exportacao_volume"],
            "Receita exportação (US$ FOB)": df["petroleo_exportacao_receita"],
        }
    )
    derivados = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Importação (barris)": df["derivados_importacao_volume"],
            "Dispêndio importação (US$ FOB)": df["derivados_importacao_dispendio"],
            "Exportação (barris)": df["derivados_exportacao_volume"],
            "Receita exportação (US$ FOB)": df["derivados_exportacao_receita"],
        }
    )
    gas = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Importação (bep)": df["gas_importacao_volume"],
            "Dispêndio importação (US$ FOB)": df["gas_importacao_dispendio"],
        }
    )
    etanol = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Imp. anidro (b)": df["etanol_anidro_importacao_volume"],
            "Disp. imp. anidro (US$)": df["etanol_anidro_importacao_dispendio"],
            "Imp. hidratado (b)": df["etanol_hidratado_importacao_volume"],
            "Disp. imp. hidratado (US$)": df["etanol_hidratado_importacao_dispendio"],
            "Exp. anidro (b)": df["etanol_anidro_exportacao_volume"],
            "Rec. exp. anidro (US$)": df["etanol_anidro_exportacao_receita"],
            "Exp. hidratado (b)": df["etanol_hidratado_exportacao_volume"],
            "Rec. exp. hidratado (US$)": df["etanol_hidratado_exportacao_receita"],
        }
    )
    saldo = pd.DataFrame(
        {
            "Ano": df["ano"],
            "Imp. petróleo+derivados+gás (US$)": df["importacao_energia_usd"],
            "Exp. petróleo+derivados (US$)": df["exportacao_petroleo_derivados_usd"],
            "Saldo (US$)": df["saldo_comercio_petroleo_derivados_gas_usd"],
        }
    )

    xlsx_path = OUT / "anp_importacoes_exportacoes_anual_2000_2025.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        petroleo.to_excel(writer, sheet_name="Petroleo", index=False)
        derivados.to_excel(writer, sheet_name="Derivados", index=False)
        gas.to_excel(writer, sheet_name="Gas_natural", index=False)
        etanol.to_excel(writer, sheet_name="Etanol", index=False)
        saldo.to_excel(writer, sheet_name="Saldo_comercio", index=False)
        df.to_excel(writer, sheet_name="Todas_series", index=False)
        pd.DataFrame({"Nota": notes}).to_excel(writer, sheet_name="Notas", index=False)

    # Markdown summary
    md = [
        "# ANP — Importações e Exportações (barris), anuais 2000–2025\n",
        f"Fonte: [ANP importacoes-exportacoes-b.xlsx]({payload['source_url']})",
        f"{updated or ''}\n",
        "## Petróleo\n",
        "| Ano | Imp. (mi b) | Dispêndio (US$ bi) | Preço méd. | Exp. (mi b) | Receita (US$ bi) |",
        "|---:|---:|---:|---:|---:|---:|",
    ]
    for r in rows:
        md.append(
            f"| {r['ano']} | {br((r['petroleo_importacao_volume'] or 0)/1e6 if r['petroleo_importacao_volume'] is not None else None)} | "
            f"{br((r['petroleo_importacao_dispendio'] or 0)/1e9 if r['petroleo_importacao_dispendio'] is not None else None, 2)} | "
            f"{br(r['petroleo_importacao_preco_medio'], 2)} | "
            f"{br((r['petroleo_exportacao_volume'] or 0)/1e6 if r['petroleo_exportacao_volume'] is not None else None)} | "
            f"{br((r['petroleo_exportacao_receita'] or 0)/1e9 if r['petroleo_exportacao_receita'] is not None else None, 2)} |"
        )
    md += [
        "\n## Derivados (total)\n",
        "| Ano | Imp. (mi b) | Dispêndio (US$ bi) | Exp. (mi b) | Receita (US$ bi) |",
        "|---:|---:|---:|---:|---:|",
    ]
    for r in rows:
        md.append(
            f"| {r['ano']} | {br((r['derivados_importacao_volume'] or 0)/1e6 if r['derivados_importacao_volume'] is not None else None)} | "
            f"{br((r['derivados_importacao_dispendio'] or 0)/1e9 if r['derivados_importacao_dispendio'] is not None else None, 2)} | "
            f"{br((r['derivados_exportacao_volume'] or 0)/1e6 if r['derivados_exportacao_volume'] is not None else None)} | "
            f"{br((r['derivados_exportacao_receita'] or 0)/1e9 if r['derivados_exportacao_receita'] is not None else None, 2)} |"
        )
    md += [
        "\n## Saldo comércio (petróleo + derivados + gás nas importações)\n",
        "| Ano | Importações energia (US$ bi) | Exportações pét.+deriv. (US$ bi) | Saldo (US$ bi) |",
        "|---:|---:|---:|---:|",
    ]
    for r in rows:
        md.append(
            f"| {r['ano']} | {br((r['importacao_energia_usd'] or 0)/1e9 if r['importacao_energia_usd'] is not None else None, 2)} | "
            f"{br((r['exportacao_petroleo_derivados_usd'] or 0)/1e9 if r['exportacao_petroleo_derivados_usd'] is not None else None, 2)} | "
            f"{br((r['saldo_comercio_petroleo_derivados_gas_usd'] or 0)/1e9 if r['saldo_comercio_petroleo_derivados_gas_usd'] is not None else None, 2)} |"
        )
    md.append("\n## Notas\n")
    for n in notes:
        md.append(f"- {n}")

    md_path = OUT / "anp_importacoes_exportacoes_anual_2000_2025.md"
    md_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    print(f"Wrote {json_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")
    print("\nSample recent years (petróleo):")
    for r in rows[-6:]:
        print(
            f"{r['ano']}: imp={r['petroleo_importacao_volume'] and r['petroleo_importacao_volume']/1e6:.1f} mi b | "
            f"exp={r['petroleo_exportacao_volume'] and r['petroleo_exportacao_volume']/1e6:.1f} mi b | "
            f"saldo US$ bi={r['saldo_comercio_petroleo_derivados_gas_usd'] and r['saldo_comercio_petroleo_derivados_gas_usd']/1e9:.2f}"
        )


if __name__ == "__main__":
    main()
