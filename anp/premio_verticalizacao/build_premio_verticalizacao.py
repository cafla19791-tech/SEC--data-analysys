#!/usr/bin/env python3
"""Prêmio da verticalização: (lifting + opex refino) vs (Brent + opex refino).

Custo caixa integrado (Petrobras verticalizada):
    C_int = lifting + opex_refino

Custo caixa de refinaria que compra petróleo a mercado:
    C_mkt = Brent + opex_refino

Prêmio da verticalização (economia de caixa por barril):
    P = C_mkt - C_int = Brent - lifting

O opex de refino cancela na diferença; entra nos níveis absolutos.

Fontes:
- Brent: EIA Europe Brent Spot FOB (mensal/anual)
- Produção Brasil (ponderação): ANP derivados por ano
- Lifting / opex: hipóteses documentadas (Petrobras Performance Report / plano)
"""

from __future__ import annotations

import json
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent
BRENT_M = ROOT / "data" / "raw" / "eia" / "RBRTEm.xls"
BRENT_A = ROOT / "data" / "raw" / "eia" / "RBRTEa.xls"
VOL_JSON = ROOT / "anp" / "derivados_volumes_produto_refinaria_2011_2026.json"

# Hipóteses de custo caixa (US$/barril)
OPEX_REFINO_SCENARIOS = {
    "baixo_3": 3.0,
    "base_5": 5.0,
    "alto_8": 8.0,
}
LIFTING_SCENARIOS = {
    # lifting operacional (sem leases / sem participações) — faixa recente Petrobras
    "lifting_apenas_6": 6.0,
    # média CE 2016–2022 citada no plano estratégico
    "lifting_media_7_3": 7.3,
    # proxy caixa ampliado: lifting + participações governamentais (~US$15 no CTPP)
    "lifting_mais_gov_21": 21.0,
}
FLAT_DOMESTIC_HYPOTHESIS = 25.0  # hipótese usada no modelo ANP anterior


def load_brent_monthly() -> dict[tuple[int, int], float]:
    wb = xlrd.open_workbook(BRENT_M)
    sh = wb.sheet_by_name("Data 1")
    out: dict[tuple[int, int], float] = {}
    for i in range(3, sh.nrows):
        dt = xlrd.xldate_as_datetime(sh.cell_value(i, 0), wb.datemode)
        val = float(sh.cell_value(i, 1))
        out[(dt.year, dt.month)] = val
    return out


def load_brent_annual() -> dict[int, float]:
    wb = xlrd.open_workbook(BRENT_A)
    sh = wb.sheet_by_name("Data 1")
    out: dict[int, float] = {}
    for i in range(3, sh.nrows):
        dt = xlrd.xldate_as_datetime(sh.cell_value(i, 0), wb.datemode)
        out[dt.year] = float(sh.cell_value(i, 1))
    # completar anos só no mensal (ex.: 2026 YTD)
    monthly = load_brent_monthly()
    by_year: dict[int, list[float]] = {}
    for (y, _m), v in monthly.items():
        by_year.setdefault(y, []).append(v)
    for y, vals in by_year.items():
        if y not in out and vals:
            out[y] = sum(vals) / len(vals)
    return out


def load_production_annual() -> dict[int, float]:
    if VOL_JSON.exists():
        data = json.loads(VOL_JSON.read_text())
        brasil: dict[int, float] = {}
        for r in data["production_by_product"]:
            y = int(r["ano"])
            if 2011 <= y <= 2026:
                brasil[y] = brasil.get(y, 0.0) + float(r["volume_produzido_barris"])
        return brasil
    raise FileNotFoundError(VOL_JSON)


def wavg(pairs: list[tuple[float, float]]) -> float:
    """pairs = (value, weight)."""
    num = sum(v * w for v, w in pairs)
    den = sum(w for _, w in pairs)
    return num / den if den else float("nan")


def build() -> dict:
    brent_m = load_brent_monthly()
    brent_a = load_brent_annual()
    prod = load_production_annual()

    years = sorted(y for y in prod if y in brent_a and 2011 <= y <= 2026)
    annual_rows = []
    for y in years:
        brent = brent_a[y]
        months = [(m, brent_m[(y, m)]) for m in range(1, 13) if (y, m) in brent_m]
        row = {
            "ano": y,
            "brent_usd_bbl": brent,
            "brent_meses_disponiveis": len(months),
            "producao_derivados_barris": prod[y],
            "cenarios": {},
        }
        for lname, lifting in LIFTING_SCENARIOS.items():
            for oname, opex in OPEX_REFINO_SCENARIOS.items():
                c_int = lifting + opex
                c_mkt = brent + opex
                premium = c_mkt - c_int  # == brent - lifting
                row["cenarios"][f"{lname}|{oname}"] = {
                    "lifting_usd_bbl": lifting,
                    "opex_refino_usd_bbl": opex,
                    "custo_integrado_lifting_mais_opex": c_int,
                    "custo_mercado_brent_mais_opex": c_mkt,
                    "premio_verticalizacao_usd_bbl": premium,
                    "hipotese_flat_25": FLAT_DOMESTIC_HYPOTHESIS,
                    "flat25_menos_integrado": FLAT_DOMESTIC_HYPOTHESIS - c_int,
                    "mercado_menos_flat25": c_mkt - FLAT_DOMESTIC_HYPOTHESIS,
                }
        annual_rows.append(row)

    def period_summary(y0: int, y1: int) -> dict:
        subset = [r for r in annual_rows if y0 <= r["ano"] <= y1]
        out: dict = {
            "start": y0,
            "end": y1,
            "brent_ponderado_producao": wavg(
                [(r["brent_usd_bbl"], r["producao_derivados_barris"]) for r in subset]
            ),
            "producao_total_barris": sum(r["producao_derivados_barris"] for r in subset),
            "cenarios": {},
        }
        keys = subset[0]["cenarios"].keys() if subset else []
        for key in keys:
            out["cenarios"][key] = {
                "custo_integrado_ponderado": wavg(
                    [
                        (
                            r["cenarios"][key]["custo_integrado_lifting_mais_opex"],
                            r["producao_derivados_barris"],
                        )
                        for r in subset
                    ]
                ),
                "custo_mercado_ponderado": wavg(
                    [
                        (
                            r["cenarios"][key]["custo_mercado_brent_mais_opex"],
                            r["producao_derivados_barris"],
                        )
                        for r in subset
                    ]
                ),
                "premio_verticalizacao_ponderado": wavg(
                    [
                        (
                            r["cenarios"][key]["premio_verticalizacao_usd_bbl"],
                            r["producao_derivados_barris"],
                        )
                        for r in subset
                    ]
                ),
                "lifting_usd_bbl": subset[0]["cenarios"][key]["lifting_usd_bbl"],
                "opex_refino_usd_bbl": subset[0]["cenarios"][key]["opex_refino_usd_bbl"],
            }
            # economia total implícita = prêmio × produção
            prem = out["cenarios"][key]["premio_verticalizacao_ponderado"]
            out["cenarios"][key]["economia_caixa_implicita_usd"] = (
                prem * out["producao_total_barris"]
            )
        return out

    # série mensal (base: lifting 6 + opex 5)
    monthly_rows = []
    for (y, m), brent in sorted(brent_m.items()):
        if y < 2011 or y > 2026:
            continue
        lifting = LIFTING_SCENARIOS["lifting_apenas_6"]
        opex = OPEX_REFINO_SCENARIOS["base_5"]
        monthly_rows.append(
            {
                "ano": y,
                "mes": m,
                "brent_usd_bbl": brent,
                "lifting_usd_bbl": lifting,
                "opex_refino_usd_bbl": opex,
                "custo_integrado_lifting_mais_opex": lifting + opex,
                "custo_mercado_brent_mais_opex": brent + opex,
                "premio_verticalizacao_usd_bbl": brent - lifting,
            }
        )

    results = {
        "title": "Prêmio da verticalização — (lifting+opex) vs (Brent+opex)",
        "formula": {
            "custo_integrado": "lifting + opex_refino",
            "custo_mercado": "Brent + opex_refino",
            "premio_verticalizacao": "custo_mercado - custo_integrado = Brent - lifting",
        },
        "notes": [
            "Brent: EIA Europe Brent Spot Price FOB (RBRTE).",
            "Ponderação: produção anual de derivados ANP (Brasil).",
            "Lifting US$6 ≈ Performance Report Petrobras 2024 (Brasil, ex-leases/participações).",
            "Lifting US$7,3 ≈ média CE 2016–2022 citada no plano estratégico.",
            "Lifting+gov US$21 ≈ proxy caixa ampliado (CE ~6 + participações ~15 do CTPP); exclui DD&A.",
            "Opex de refino US$3/5/8: faixa ilustrativa de custo caixa de planta (sem feedstock).",
            "US$25 flat: hipótese do modelo ANP de custo doméstico de derivados.",
            "Prêmio = economia de caixa da verticalização frente a comprar Brent; não é lucro contábil do segmento.",
            "Custo econômico/oportunidade do óleo próprio continua próximo do Brent (exportação alternativa).",
            "2026: Brent média dos meses disponíveis (YTD); produção ANP também parcial.",
        ],
        "assumptions": {
            "lifting_scenarios_usd_bbl": LIFTING_SCENARIOS,
            "opex_refino_scenarios_usd_bbl": OPEX_REFINO_SCENARIOS,
            "flat_domestic_hypothesis_usd_bbl": FLAT_DOMESTIC_HYPOTHESIS,
        },
        "periods": {
            "2011_2025": period_summary(2011, 2025),
            "2015_2025": period_summary(2015, 2025),
            "2011_2026_YTD": period_summary(2011, 2026),
        },
        "annual": annual_rows,
        "monthly_base_lifting6_opex5": monthly_rows,
    }
    return results


def write_xlsx(results: dict, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = results["title"]
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "C_int = lifting + opex_refino | C_mkt = Brent + opex_refino | Prêmio = C_mkt − C_int = Brent − lifting"

    ws["A4"] = "Período"
    ws["B4"] = "Cenário lifting"
    ws["C4"] = "Opex refino"
    ws["D4"] = "Brent pond."
    ws["E4"] = "C_int pond."
    ws["F4"] = "C_mkt pond."
    ws["G4"] = "Prêmio vert. pond. US$/b"
    ws["H4"] = "Economia caixa implícita US$"
    ws["I4"] = "US$25 − C_int"
    for col in range(1, 10):
        ws.cell(4, col).font = Font(bold=True)

    r = 5
    for pname, pdata in results["periods"].items():
        for key, c in pdata["cenarios"].items():
            if c["opex_refino_usd_bbl"] != 5.0 and c["lifting_usd_bbl"] not in (6.0, 21.0):
                # manter planilha legível: base opex=5 para todos lifting; e lifting=6 com todos opex
                if not (
                    c["opex_refino_usd_bbl"] == 5.0
                    or c["lifting_usd_bbl"] == 6.0
                ):
                    continue
            ws.cell(r, 1, pname)
            ws.cell(r, 2, c["lifting_usd_bbl"])
            ws.cell(r, 3, c["opex_refino_usd_bbl"])
            ws.cell(r, 4, pdata["brent_ponderado_producao"])
            ws.cell(r, 5, c["custo_integrado_ponderado"])
            ws.cell(r, 6, c["custo_mercado_ponderado"])
            ws.cell(r, 7, c["premio_verticalizacao_ponderado"])
            ws.cell(r, 8, c["economia_caixa_implicita_usd"])
            ws.cell(r, 9, FLAT_DOMESTIC_HYPOTHESIS - c["custo_integrado_ponderado"])
            for col in range(2, 10):
                ws.cell(r, col).number_format = "#,##0.00"
            ws.cell(r, 8).number_format = "#,##0"
            r += 1

    # destaque headline
    base = results["periods"]["2011_2025"]["cenarios"]["lifting_apenas_6|base_5"]
    p = results["periods"]["2011_2025"]
    ws.cell(r + 1, 1, "HEADLINE 2011–2025 | lifting US$6 + opex US$5")
    ws.cell(r + 1, 1).font = Font(bold=True)
    ws.cell(r + 2, 1, "Brent ponderado")
    ws.cell(r + 2, 2, p["brent_ponderado_producao"])
    ws.cell(r + 3, 1, "C_int (lifting+opex)")
    ws.cell(r + 3, 2, base["custo_integrado_ponderado"])
    ws.cell(r + 4, 1, "C_mkt (Brent+opex)")
    ws.cell(r + 4, 2, base["custo_mercado_ponderado"])
    ws.cell(r + 5, 1, "Prêmio verticalização US$/b")
    ws.cell(r + 5, 2, base["premio_verticalizacao_ponderado"])
    ws.cell(r + 5, 2).font = Font(bold=True)

    ws2 = wb.create_sheet("Anual_base")
    ws2.append(
        [
            "ano",
            "brent",
            "producao_barris",
            "lifting",
            "opex_refino",
            "C_int",
            "C_mkt",
            "premio_vert",
            "flat25",
            "flat25_menos_Cint",
            "Cmkt_menos_flat25",
        ]
    )
    for row in results["annual"]:
        c = row["cenarios"]["lifting_apenas_6|base_5"]
        ws2.append(
            [
                row["ano"],
                row["brent_usd_bbl"],
                row["producao_derivados_barris"],
                c["lifting_usd_bbl"],
                c["opex_refino_usd_bbl"],
                c["custo_integrado_lifting_mais_opex"],
                c["custo_mercado_brent_mais_opex"],
                c["premio_verticalizacao_usd_bbl"],
                c["hipotese_flat_25"],
                c["flat25_menos_integrado"],
                c["mercado_menos_flat25"],
            ]
        )

    ws3 = wb.create_sheet("Anual_todos_cenarios")
    ws3.append(
        [
            "ano",
            "brent",
            "producao",
            "cenario",
            "lifting",
            "opex",
            "C_int",
            "C_mkt",
            "premio",
        ]
    )
    for row in results["annual"]:
        for key, c in row["cenarios"].items():
            ws3.append(
                [
                    row["ano"],
                    row["brent_usd_bbl"],
                    row["producao_derivados_barris"],
                    key,
                    c["lifting_usd_bbl"],
                    c["opex_refino_usd_bbl"],
                    c["custo_integrado_lifting_mais_opex"],
                    c["custo_mercado_brent_mais_opex"],
                    c["premio_verticalizacao_usd_bbl"],
                ]
            )

    ws4 = wb.create_sheet("Mensal_base")
    ws4.append(
        [
            "ano",
            "mes",
            "brent",
            "lifting",
            "opex",
            "C_int",
            "C_mkt",
            "premio",
        ]
    )
    for row in results["monthly_base_lifting6_opex5"]:
        ws4.append(
            [
                row["ano"],
                row["mes"],
                row["brent_usd_bbl"],
                row["lifting_usd_bbl"],
                row["opex_refino_usd_bbl"],
                row["custo_integrado_lifting_mais_opex"],
                row["custo_mercado_brent_mais_opex"],
                row["premio_verticalizacao_usd_bbl"],
            ]
        )

    ws5 = wb.create_sheet("Notas")
    for i, n in enumerate(results["notes"], start=1):
        ws5.cell(i, 1, n)
    ws5.column_dimensions["A"].width = 110

    for sheet in wb.worksheets:
        if sheet.title != "Notas":
            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 16

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_md(results: dict, path: Path) -> None:
    p = results["periods"]["2011_2025"]
    p2 = results["periods"]["2015_2025"]
    base = p["cenarios"]["lifting_apenas_6|base_5"]
    gov = p["cenarios"]["lifting_mais_gov_21|base_5"]
    base2 = p2["cenarios"]["lifting_apenas_6|base_5"]

    lines = [
        "# Prêmio da verticalização — `(lifting + opex)` vs `(Brent + opex)`",
        "",
        "## Fórmulas",
        "",
        "- **Custo caixa integrado** (Petrobras verticalizada): `C_int = lifting + opex_refino`",
        "- **Custo caixa a mercado** (compra Brent): `C_mkt = Brent + opex_refino`",
        "- **Prêmio da verticalização**: `P = C_mkt − C_int = Brent − lifting`",
        "",
        "O opex de refino entra nos níveis absolutos e **cancela** na diferença.",
        "",
        "## Headline (2011–2025, ponderado pela produção ANP de derivados)",
        "",
        f"- Brent ponderado: **US$ {p['brent_ponderado_producao']:.2f}/b**",
        f"- Cenário base: lifting **US$ 6** + opex refino **US$ 5**",
        f"- **C_int = US$ {base['custo_integrado_ponderado']:.2f}/b**",
        f"- **C_mkt = US$ {base['custo_mercado_ponderado']:.2f}/b**",
        f"- **Prêmio da verticalização = US$ {base['premio_verticalizacao_ponderado']:.2f}/b**",
        f"- Economia de caixa implícita no período: **US$ {base['economia_caixa_implicita_usd']/1e9:.1f} bi** "
        f"(prêmio × {p['producao_total_barris']/1e9:.2f} bi barris produzidos)",
        "",
        "### Sensibilidade do lifting (opex refino = US$ 5)",
        "",
        "| Lifting | C_int | C_mkt | Prêmio US$/b |",
        "|---:|---:|---:|---:|",
    ]
    for key, lab in [
        ("lifting_apenas_6|base_5", "6,0 (só lifting)"),
        ("lifting_media_7_3|base_5", "7,3 (média CE)"),
        ("lifting_mais_gov_21|base_5", "21,0 (lifting+gov)"),
    ]:
        c = p["cenarios"][key]
        lines.append(
            f"| {lab} | {c['custo_integrado_ponderado']:.2f} | "
            f"{c['custo_mercado_ponderado']:.2f} | "
            f"**{c['premio_verticalizacao_ponderado']:.2f}** |"
        )

    lines += [
        "",
        f"Janela 2015–2025 (base lifting 6 + opex 5): prêmio **US$ {base2['premio_verticalizacao_ponderado']:.2f}/b** "
        f"(Brent pond. US$ {p2['brent_ponderado_producao']:.2f}).",
        "",
        "## Relação com a hipótese US$ 25",
        "",
        f"- Hipótese flat do modelo ANP: **US$ {FLAT_DOMESTIC_HYPOTHESIS:.0f}/b**",
        f"- Contra C_int base (US$ {base['custo_integrado_ponderado']:.2f}): "
        f"US$25 está **{FLAT_DOMESTIC_HYPOTHESIS - base['custo_integrado_ponderado']:+.2f}** "
        "(conservador / margem para logística e itens não modelados).",
        f"- Contra C_int com gov (US$ {gov['custo_integrado_ponderado']:.2f}): "
        f"US$25 está **{FLAT_DOMESTIC_HYPOTHESIS - gov['custo_integrado_ponderado']:+.2f}** "
        "(quase colado no caixa ampliado).",
        f"- Contra C_mkt (US$ {base['custo_mercado_ponderado']:.2f}): "
        f"US$25 está **{base['custo_mercado_ponderado'] - FLAT_DOMESTIC_HYPOTHESIS:.2f}** abaixo "
        "— ou seja, captura grande parte do prêmio da verticalização.",
        "",
        "## Série anual (base: lifting 6 + opex 5)",
        "",
        "| Ano | Brent | C_int | C_mkt | Prêmio |",
        "|---:|---:|---:|---:|---:|",
    ]
    for row in results["annual"]:
        if row["ano"] > 2025:
            continue
        c = row["cenarios"]["lifting_apenas_6|base_5"]
        lines.append(
            f"| {row['ano']} | {row['brent_usd_bbl']:.2f} | "
            f"{c['custo_integrado_lifting_mais_opex']:.2f} | "
            f"{c['custo_mercado_brent_mais_opex']:.2f} | "
            f"{c['premio_verticalizacao_usd_bbl']:.2f} |"
        )

    lines += ["", "## Notas", ""]
    for n in results["notes"]:
        lines.append(f"- {n}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    results = build()
    OUT.mkdir(parents=True, exist_ok=True)
    jp = OUT / "premio_verticalizacao.json"
    xp = OUT / "premio_verticalizacao.xlsx"
    mp = OUT / "premio_verticalizacao.md"
    jp.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    write_xlsx(results, xp)
    write_md(results, mp)

    p = results["periods"]["2011_2025"]
    b = p["cenarios"]["lifting_apenas_6|base_5"]
    g = p["cenarios"]["lifting_mais_gov_21|base_5"]
    print("=== 2011-2025 (prod-weighted) ===")
    print(f"Brent: {p['brent_ponderado_producao']:.2f}")
    print(f"C_int (6+5): {b['custo_integrado_ponderado']:.2f}")
    print(f"C_mkt (Brent+5): {b['custo_mercado_ponderado']:.2f}")
    print(f"Premium: {b['premio_verticalizacao_ponderado']:.2f}")
    print(f"Economy US$ bi: {b['economia_caixa_implicita_usd']/1e9:.2f}")
    print(f"Premium with gov21: {g['premio_verticalizacao_ponderado']:.2f}")
    print("wrote", jp, xp, mp)


if __name__ == "__main__":
    main()
