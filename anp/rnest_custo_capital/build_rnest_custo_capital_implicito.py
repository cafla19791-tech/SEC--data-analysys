#!/usr/bin/env python3
"""Custo implícito de capital da RNEST vs parque nacional / benchmarks.

Métricas:
  1) Capex ÷ produção acumulada de derivados (2015–2025 e 2015–2026 YTD)
  2) Capex ÷ (capacidade × dias × anos × fator de utilização)
  3) Carga de capital anualizada (CRF) ÷ produção média anual

Fontes de produção: ANP producao-derivados-b (série por refinaria no JSON
derivados_volumes_produto_refinaria_2011_2026.json).
Fontes de Capex/capacidade: TCU Acórdão 1839/2018-Plenário e benchmarks
Credit Suisse citados no acórdão; capacidade Brasil ANP ~2,43 mi bpd (2023).
"""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

ROOT = Path(__file__).resolve().parents[2]
VOL_JSON = ROOT / "anp" / "derivados_volumes_produto_refinaria_2011_2026.json"
OUT_DIR = Path(__file__).resolve().parent

# Capacidade de processamento (bpd)
RNEST_CAP_PLANEJADA_BPD = 230_000
RNEST_CAP_FASE1_BPD = 115_000  # ~Trem 1; TCU cita ~100 kbpd efetivamente acrescidos
RNEST_CAP_TCU_ENTREGUE_BPD = 100_000
BRASIL_CAP_BPD = 2_430_000  # ANP ~2023

# Cenários de Capex RNEST (US$)
CAPEX_SCENARIOS = {
    "A_orcamento_atualizado_TCU_US19bi": 19.0e9,  # orçamento ~90% avanço físico
    "B_revisao_publica_US20_1bi": 20.1e9,  # PNG/revisões amplamente citadas
    "C_cenario_base_2009_US13bi": 13.0e9,  # análise de risco Abastecimento (TCU)
    "D_programa_expansao_atribuido_100kbpd_US30bi": 30.0e9,  # desembolso programa ÷ só Trem 1
}

# Benchmarks Capex/bpd (Credit Suisse / TCU item 462–463)
BENCH_CAPEX_PER_BPD = {
    "Jamnagar_Reliance": 9_000,
    "media_mundial_greenfield": 18_000,
    "Premium_I_II_planejado": 35_000,  # faixa ~33–37k no acórdão
    "RNEST_CreditSuisse": 52_000,
}

LIFE_YEARS = 30
WACC = 0.10  # real aproximado para CRF ilustrativo


def days_in_year(y: int) -> int:
    leap = y % 4 == 0 and (y % 100 != 0 or y % 400 == 0)
    return 366 if leap else 365


def crf(r: float, n: int) -> float:
    if r == 0:
        return 1 / n
    return r * (1 + r) ** n / ((1 + r) ** n - 1)


def load_production() -> tuple[dict[int, float], dict[int, float]]:
    """Retorna (rnest_ano->barris, brasil_ano->barris)."""
    if VOL_JSON.exists():
        data = json.loads(VOL_JSON.read_text())
        rnest = {
            r["ano"]: float(r["volume_produzido_barris"])
            for r in data["production_by_refinery"]
            if r["refinaria"] == "RNEST" and 2015 <= r["ano"] <= 2026
        }
        brasil: dict[int, float] = {}
        for r in data["production_by_product"]:
            if 2015 <= r["ano"] <= 2026:
                brasil[r["ano"]] = brasil.get(r["ano"], 0.0) + float(
                    r["volume_produzido_barris"]
                )
        return rnest, brasil

    # fallback (mesmos totais já validados na série ANP)
    rnest = {
        2015: 20549736.0,
        2016: 32285006.0,
        2017: 29659020.0,
        2018: 27327975.0,
        2019: 30804513.0,
        2020: 41624120.0,
        2021: 26491739.0,
        2022: 29094368.0,
        2023: 33480284.0,
        2024: 33843282.0,
        2025: 29358000.0,
        2026: 21251314.0,
    }
    brasil = {
        2015: 745032614.0,
        2016: 697379475.0,
        2017: 666817834.0,
        2018: 656671171.0,
        2019: 657201195.0,
        2020: 682339125.0,
        2021: 697950694.0,
        2022: 747981947.0,
        2023: 781375150.0,
        2024: 792214977.0,
        2025: 777732563.0,
        2026: 396853712.0,
    }
    return rnest, brasil


def cum(series: dict[int, float], y0: int, y1: int) -> float:
    return sum(series[y] for y in range(y0, y1 + 1) if y in series)


def mean_bpd(series: dict[int, float], y0: int, y1: int) -> float:
    barrels = 0.0
    days = 0
    for y in range(y0, y1 + 1):
        if y not in series:
            continue
        barrels += series[y]
        days += days_in_year(y)
    return barrels / days if days else math.nan


def utilization(avg_product_bpd: float, capacity_bpd: float) -> float:
    return avg_product_bpd / capacity_bpd if capacity_bpd else math.nan


def metric_block(
    name: str,
    capex: float,
    prod_cum: float,
    avg_bpd: float,
    capacity_bpd: float,
    years: float,
) -> dict:
    util = utilization(avg_bpd, capacity_bpd)
    denom_cap = capacity_bpd * 365.25 * years * util  # ≈ prod_cum se util=prod/cap
    # Explicit capacity×util identity check uses actual production days approx
    denom_explicit = capacity_bpd * 365.25 * years * util
    annual_cap_charge = capex * crf(WACC, LIFE_YEARS)
    avg_annual_prod = prod_cum / years if years else math.nan
    return {
        "cenario": name,
        "capex_usd": capex,
        "producao_acumulada_barris": prod_cum,
        "anos": years,
        "producao_media_bpd": avg_bpd,
        "capacidade_bpd": capacity_bpd,
        "fator_utilizacao_sobre_capacidade": util,
        "capex_por_bpd_capacidade_usd": capex / capacity_bpd if capacity_bpd else math.nan,
        "custo_implicito_capex_por_barril_produzido_usd": (
            capex / prod_cum if prod_cum else math.nan
        ),
        "custo_implicito_capex_capacidade_x_util_usd": (
            capex / denom_explicit if denom_explicit else math.nan
        ),
        "carga_capital_anualizada_usd": annual_cap_charge,
        "carga_capital_por_barril_usd": (
            annual_cap_charge / avg_annual_prod if avg_annual_prod else math.nan
        ),
    }


def counterfactual_park(
    capacity_bpd: float,
    prod_cum: float,
    years: float,
    avg_bpd: float,
    label: str,
) -> list[dict]:
    rows = []
    for bench, usd_bpd in BENCH_CAPEX_PER_BPD.items():
        capex = capacity_bpd * usd_bpd
        rows.append(
            {
                "referencia": label,
                "benchmark_capex_por_bpd": bench,
                "capex_por_bpd_usd": usd_bpd,
                **metric_block(
                    f"{label}|{bench}",
                    capex,
                    prod_cum,
                    avg_bpd,
                    capacity_bpd,
                    years,
                ),
            }
        )
    return rows


def fmt(n: float, nd: int = 2) -> str:
    if n is None or (isinstance(n, float) and (math.isnan(n) or math.isinf(n))):
        return "—"
    return f"{n:,.{nd}f}"


def build() -> dict:
    rnest, brasil = load_production()

    # Janelas: anos cheios 2015–2025; e 2015–2026 (2026 parcial/YTD)
    windows = {
        "2015_2025": (2015, 2025),
        "2015_2026_incl_YTD": (2015, 2026),
    }

    results = {
        "title": "Custo implícito de capital — RNEST vs parque nacional",
        "unit_production": "barris de derivados (ANP)",
        "notes": [
            "Produção = soma de derivados ANP por refinaria/ano (não é carga de petróleo).",
            "2026 na série ANP está parcial (YTD); preferir janela 2015–2025 para comparações.",
            "Capex RNEST cenários A/B/C são do empreendimento; D atribui o desembolso do programa de expansão (US$ 30 bi) à única capacidade entregue (~100 kbpd), como no TCU.",
            "Parque nacional: Capex de reposição hipotético = capacidade × benchmark US$/bpd (não é Capex histórico contábil).",
            f"Carga anualizada usa CRF com WACC={WACC:.0%} e vida={LIFE_YEARS} anos (ilustrativo).",
            "Fator de utilização aqui = produção média de derivados (bpd) ÷ capacidade de processamento (bpd); é proxy, pois unidades diferem (produtos vs carga).",
        ],
        "assumptions": {
            "rnest_capacidade_planejada_bpd": RNEST_CAP_PLANEJADA_BPD,
            "rnest_capacidade_fase1_bpd": RNEST_CAP_FASE1_BPD,
            "rnest_capacidade_tcu_entregue_bpd": RNEST_CAP_TCU_ENTREGUE_BPD,
            "brasil_capacidade_bpd": BRASIL_CAP_BPD,
            "capex_scenarios_usd": CAPEX_SCENARIOS,
            "benchmarks_capex_per_bpd_usd": BENCH_CAPEX_PER_BPD,
            "wacc": WACC,
            "life_years": LIFE_YEARS,
        },
        "production_annual": {
            "RNEST": rnest,
            "BRASIL": brasil,
        },
        "windows": {},
    }

    for wname, (y0, y1) in windows.items():
        years = y1 - y0 + 1
        # para 2026 parcial, ainda contamos 1 ano-calendário no denominador de média —
        # mas destacamos YTD; util/média usam dias reais do período presente na série
        r_cum = cum(rnest, y0, y1)
        b_cum = cum(brasil, y0, y1)
        r_avg = mean_bpd(rnest, y0, y1)
        b_avg = mean_bpd(brasil, y0, y1)
        # anos efetivos ponderados por dias
        days = sum(days_in_year(y) for y in range(y0, y1 + 1) if y in rnest)
        years_eff = days / 365.25

        rnest_rows = []
        for scen, capex in CAPEX_SCENARIOS.items():
            for cap_label, cap in [
                ("fase1_115kbpd", RNEST_CAP_FASE1_BPD),
                ("planejada_230kbpd", RNEST_CAP_PLANEJADA_BPD),
                ("tcu_entregue_100kbpd", RNEST_CAP_TCU_ENTREGUE_BPD),
            ]:
                row = metric_block(
                    f"{scen}|{cap_label}",
                    capex,
                    r_cum,
                    r_avg,
                    cap,
                    years_eff,
                )
                row["capex_scenario"] = scen
                row["capacidade_ref"] = cap_label
                rnest_rows.append(row)

        park_rows = counterfactual_park(
            BRASIL_CAP_BPD, b_cum, years_eff, b_avg, "parque_nacional"
        )
        # Contrafactual: o que o Capex da RNEST compraria a benchmarks mundiais
        cf = []
        for scen, capex in CAPEX_SCENARIOS.items():
            for bench, usd_bpd in BENCH_CAPEX_PER_BPD.items():
                if bench == "RNEST_CreditSuisse":
                    continue
                bpd_bought = capex / usd_bpd
                cf.append(
                    {
                        "capex_scenario": scen,
                        "benchmark": bench,
                        "capex_usd": capex,
                        "capacidade_compravel_bpd": bpd_bought,
                        "multiplo_vs_fase1_115k": bpd_bought / RNEST_CAP_FASE1_BPD,
                        "multiplo_vs_planejada_230k": bpd_bought / RNEST_CAP_PLANEJADA_BPD,
                    }
                )

        # Comparativo direto: RNEST (cenário A, cap fase1) vs parque a US$18k
        base_rnest = next(
            r
            for r in rnest_rows
            if r["capex_scenario"] == "A_orcamento_atualizado_TCU_US19bi"
            and r["capacidade_ref"] == "fase1_115kbpd"
        )
        park_18 = next(
            r for r in park_rows if r["benchmark_capex_por_bpd"] == "media_mundial_greenfield"
        )
        park_9 = next(
            r for r in park_rows if r["benchmark_capex_por_bpd"] == "Jamnagar_Reliance"
        )

        results["windows"][wname] = {
            "y0": y0,
            "y1": y1,
            "years_effective": years_eff,
            "rnest_prod_cum_barris": r_cum,
            "brasil_prod_cum_barris": b_cum,
            "rnest_avg_bpd": r_avg,
            "brasil_avg_bpd": b_avg,
            "rnest_util_fase1": utilization(r_avg, RNEST_CAP_FASE1_BPD),
            "rnest_util_planejada": utilization(r_avg, RNEST_CAP_PLANEJADA_BPD),
            "brasil_util": utilization(b_avg, BRASIL_CAP_BPD),
            "rnest_metrics": rnest_rows,
            "parque_reposicao_metrics": park_rows,
            "counterfactual_capacidade_com_mesmo_capex": cf,
            "headline_compare": {
                "rnest_capex_por_barril_usd": base_rnest[
                    "custo_implicito_capex_por_barril_produzido_usd"
                ],
                "rnest_carga_capital_por_barril_usd": base_rnest[
                    "carga_capital_por_barril_usd"
                ],
                "parque_18k_capex_por_barril_usd": park_18[
                    "custo_implicito_capex_por_barril_produzido_usd"
                ],
                "parque_18k_carga_capital_por_barril_usd": park_18[
                    "carga_capital_por_barril_usd"
                ],
                "parque_9k_capex_por_barril_usd": park_9[
                    "custo_implicito_capex_por_barril_produzido_usd"
                ],
                "multiplo_rnest_vs_parque18k_capex_por_barril": (
                    base_rnest["custo_implicito_capex_por_barril_produzido_usd"]
                    / park_18["custo_implicito_capex_por_barril_produzido_usd"]
                ),
                "multiplo_rnest_vs_parque18k_carga_anualizada": (
                    base_rnest["carga_capital_por_barril_usd"]
                    / park_18["carga_capital_por_barril_usd"]
                ),
            },
        }

    return results


def write_xlsx(results: dict, path: Path) -> None:
    wb = Workbook()

    # Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = results["title"]
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Janela preferencial: 2015–2025 (anos cheios). 2026 = YTD parcial."

    w = results["windows"]["2015_2025"]
    h = w["headline_compare"]
    rows = [
        ("Produção acumulada RNEST 2015–2025 (barris)", w["rnest_prod_cum_barris"]),
        ("Produção média RNEST (bpd)", w["rnest_avg_bpd"]),
        ("Utilização vs fase 1 115 kbpd", w["rnest_util_fase1"]),
        ("Utilização vs planejada 230 kbpd", w["rnest_util_planejada"]),
        ("Produção acumulada Brasil 2015–2025 (barris)", w["brasil_prod_cum_barris"]),
        ("Produção média Brasil (bpd)", w["brasil_avg_bpd"]),
        ("Utilização parque nacional (proxy)", w["brasil_util"]),
        ("", ""),
        (
            "RNEST Capex/barril produzido (cenário A US$19bi)",
            h["rnest_capex_por_barril_usd"],
        ),
        (
            "RNEST carga capital anualizada/barril (A, WACC 10%, 30a)",
            h["rnest_carga_capital_por_barril_usd"],
        ),
        (
            "Parque nacional Capex/barril se reposição a US$18k/bpd",
            h["parque_18k_capex_por_barril_usd"],
        ),
        (
            "Parque nacional carga capital/barril (US$18k/bpd)",
            h["parque_18k_carga_capital_por_barril_usd"],
        ),
        (
            "Múltiplo RNEST ÷ parque@18k (Capex/barril)",
            h["multiplo_rnest_vs_parque18k_capex_por_barril"],
        ),
        (
            "Múltiplo RNEST ÷ parque@18k (carga anualizada/barril)",
            h["multiplo_rnest_vs_parque18k_carga_anualizada"],
        ),
    ]
    ws["A4"] = "Indicador"
    ws["B4"] = "Valor"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)
    for i, (lab, val) in enumerate(rows, start=5):
        ws[f"A{i}"] = lab
        ws[f"B{i}"] = val
        if isinstance(val, float):
            if "Utilização" in lab or "Múltiplo" in lab:
                ws[f"B{i}"].number_format = "0.00%"
                if "Múltiplo" in lab:
                    ws[f"B{i}"].number_format = "0.00"
            elif "bpd" in lab and "US$" not in lab and "Capex" not in lab:
                ws[f"B{i}"].number_format = "#,##0.0"
            elif "barris" in lab:
                ws[f"B{i}"].number_format = "#,##0"
            else:
                ws[f"B{i}"].number_format = '#,##0.00'

    # Produção anual
    ws2 = wb.create_sheet("Producao_anual")
    ws2.append(["ano", "RNEST_barris", "BRASIL_barris", "RNEST_share"])
    for y in range(2015, 2027):
        r = results["production_annual"]["RNEST"].get(y, 0.0)
        b = results["production_annual"]["BRASIL"].get(y, 0.0)
        ws2.append([y, r, b, (r / b if b else None)])
        ws2.cell(ws2.max_row, 2).number_format = "#,##0.00"
        ws2.cell(ws2.max_row, 3).number_format = "#,##0.00"
        ws2.cell(ws2.max_row, 4).number_format = "0.00%"

    # Métricas RNEST
    ws3 = wb.create_sheet("RNEST_metricas")
    headers = [
        "janela",
        "capex_scenario",
        "capacidade_ref",
        "capex_usd",
        "producao_acumulada_barris",
        "anos_efetivos",
        "producao_media_bpd",
        "capacidade_bpd",
        "fator_utilizacao",
        "capex_por_bpd_capacidade_usd",
        "custo_implicito_capex_por_barril_usd",
        "carga_capital_anualizada_usd",
        "carga_capital_por_barril_usd",
    ]
    ws3.append(headers)
    for wname, wdata in results["windows"].items():
        for r in wdata["rnest_metrics"]:
            ws3.append(
                [
                    wname,
                    r["capex_scenario"],
                    r["capacidade_ref"],
                    r["capex_usd"],
                    r["producao_acumulada_barris"],
                    r["anos"],
                    r["producao_media_bpd"],
                    r["capacidade_bpd"],
                    r["fator_utilizacao_sobre_capacidade"],
                    r["capex_por_bpd_capacidade_usd"],
                    r["custo_implicito_capex_por_barril_produzido_usd"],
                    r["carga_capital_anualizada_usd"],
                    r["carga_capital_por_barril_usd"],
                ]
            )

    # Parque
    ws4 = wb.create_sheet("Parque_reposicao")
    ws4.append(
        [
            "janela",
            "benchmark",
            "capex_por_bpd_usd",
            "capex_usd",
            "producao_acumulada_barris",
            "producao_media_bpd",
            "fator_utilizacao",
            "custo_implicito_capex_por_barril_usd",
            "carga_capital_por_barril_usd",
        ]
    )
    for wname, wdata in results["windows"].items():
        for r in wdata["parque_reposicao_metrics"]:
            ws4.append(
                [
                    wname,
                    r["benchmark_capex_por_bpd"],
                    r["capex_por_bpd_usd"],
                    r["capex_usd"],
                    r["producao_acumulada_barris"],
                    r["producao_media_bpd"],
                    r["fator_utilizacao_sobre_capacidade"],
                    r["custo_implicito_capex_por_barril_produzido_usd"],
                    r["carga_capital_por_barril_usd"],
                ]
            )

    # Contrafactual
    ws5 = wb.create_sheet("Contrafactual_capacidade")
    ws5.append(
        [
            "janela",
            "capex_scenario",
            "benchmark",
            "capex_usd",
            "capacidade_compravel_bpd",
            "multiplo_vs_fase1_115k",
            "multiplo_vs_planejada_230k",
        ]
    )
    for wname, wdata in results["windows"].items():
        for r in wdata["counterfactual_capacidade_com_mesmo_capex"]:
            ws5.append(
                [
                    wname,
                    r["capex_scenario"],
                    r["benchmark"],
                    r["capex_usd"],
                    r["capacidade_compravel_bpd"],
                    r["multiplo_vs_fase1_115k"],
                    r["multiplo_vs_planejada_230k"],
                ]
            )

    ws6 = wb.create_sheet("Notas")
    for i, n in enumerate(results["notes"], start=1):
        ws6[f"A{i}"] = n
        ws6[f"A{i}"].alignment = Alignment(wrap_text=True)
    ws6.column_dimensions["A"].width = 110
    for sheet in wb.worksheets:
        if sheet.title != "Notas":
            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_md(results: dict, path: Path) -> None:
    w = results["windows"]["2015_2025"]
    h = w["headline_compare"]
    lines = [
        "# Custo implícito de capital — RNEST vs parque nacional",
        "",
        "Janela principal: **2015–2025** (anos-calendário cheios). Produção em **barris de derivados** (ANP).",
        "",
        "## Produção e utilização",
        "",
        f"- RNEST acumulado: **{w['rnest_prod_cum_barris']:,.0f} barris** "
        f"(média **{w['rnest_avg_bpd']:,.0f} bpd**)",
        f"- Utilização proxy vs fase 1 (115 kbpd): **{w['rnest_util_fase1']:.1%}**",
        f"- Utilização proxy vs capacidade planejada (230 kbpd): **{w['rnest_util_planejada']:.1%}**",
        f"- Brasil acumulado: **{w['brasil_prod_cum_barris']:,.0f} barris** "
        f"(média **{w['brasil_avg_bpd']:,.0f} bpd**; util. proxy vs 2,43 mi bpd: **{w['brasil_util']:.1%}**)",
        "",
        "## Custo implícito (Capex ÷ produção acumulada)",
        "",
        "Cenário-base RNEST: **Capex US$ 19 bi** (orçamento atualizado no TCU) e capacidade fase 1 **115 kbpd**.",
        "",
        "| Referência | Capex implícito por barril produzido | Carga de capital anualizada por barril (WACC 10%, 30a) |",
        "|---|---:|---:|",
        f"| RNEST (US$ 19 bi) | **US$ {h['rnest_capex_por_barril_usd']:,.2f}** | **US$ {h['rnest_carga_capital_por_barril_usd']:,.2f}** |",
        f"| Parque nacional se reposição a US$ 18k/bpd | US$ {h['parque_18k_capex_por_barril_usd']:,.2f} | US$ {h['parque_18k_carga_capital_por_barril_usd']:,.2f} |",
        f"| Parque nacional se reposição a US$ 9k/bpd (Jamnagar) | US$ {h['parque_9k_capex_por_barril_usd']:,.2f} | — |",
        "",
        f"**Múltiplo RNEST ÷ parque@US$18k/bpd (Capex/barril): {h['multiplo_rnest_vs_parque18k_capex_por_barril']:.1f}×**",
        "",
        f"**Múltiplo na carga anualizada/barril: {h['multiplo_rnest_vs_parque18k_carga_anualizada']:.1f}×**",
        "",
        "## Cenários de Capex RNEST (2015–2025)",
        "",
        "| Cenário Capex | Capex | Capex ÷ prod. acum. (US$/b) | Carga cap./barril (US$/b) | Capex/bpd @115k | Capex/bpd @230k |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    # pick fase1 rows for each scenario
    by_scen = {}
    for r in w["rnest_metrics"]:
        if r["capacidade_ref"] == "fase1_115kbpd":
            by_scen[r["capex_scenario"]] = r
    planejada = {
        r["capex_scenario"]: r
        for r in w["rnest_metrics"]
        if r["capacidade_ref"] == "planejada_230kbpd"
    }
    labels = {
        "A_orcamento_atualizado_TCU_US19bi": "A — US$ 19 bi (TCU)",
        "B_revisao_publica_US20_1bi": "B — US$ 20,1 bi",
        "C_cenario_base_2009_US13bi": "C — US$ 13 bi (base 2009)",
        "D_programa_expansao_atribuido_100kbpd_US30bi": "D — US$ 30 bi (programa→100 kbpd)",
    }
    for key, lab in labels.items():
        r = by_scen[key]
        p = planejada[key]
        lines.append(
            f"| {lab} | {r['capex_usd']/1e9:.1f} bi | "
            f"{r['custo_implicito_capex_por_barril_produzido_usd']:,.2f} | "
            f"{r['carga_capital_por_barril_usd']:,.2f} | "
            f"{r['capex_por_bpd_capacidade_usd']:,.0f} | "
            f"{p['capex_por_bpd_capacidade_usd']:,.0f} |"
        )

    lines += [
        "",
        "## Capacidade que o mesmo Capex compraria (benchmarks)",
        "",
        "| Capex RNEST | Benchmark | Capacidade comprável | × fase 1 (115k) | × planejada (230k) |",
        "|---|---|---:|---:|---:|",
    ]
    for r in w["counterfactual_capacidade_com_mesmo_capex"]:
        if r["capex_scenario"] not in (
            "A_orcamento_atualizado_TCU_US19bi",
            "D_programa_expansao_atribuido_100kbpd_US30bi",
        ):
            continue
        lines.append(
            f"| {r['capex_usd']/1e9:.1f} bi | {r['benchmark']} | "
            f"{r['capacidade_compravel_bpd']:,.0f} bpd | "
            f"{r['multiplo_vs_fase1_115k']:.1f}× | "
            f"{r['multiplo_vs_planejada_230k']:.1f}× |"
        )

    lines += [
        "",
        "## Notas",
        "",
    ]
    for n in results["notes"]:
        lines.append(f"- {n}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    results = build()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUT_DIR / "rnest_custo_capital_implicito.json"
    xlsx_path = OUT_DIR / "rnest_custo_capital_implicito.xlsx"
    md_path = OUT_DIR / "rnest_custo_capital_implicito.md"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    write_xlsx(results, xlsx_path)
    write_md(results, md_path)

    w = results["windows"]["2015_2025"]
    h = w["headline_compare"]
    print("=== HEADLINE 2015-2025 ===")
    print(f"RNEST prod cum: {w['rnest_prod_cum_barris']:,.0f} bbl")
    print(f"RNEST avg: {w['rnest_avg_bpd']:,.0f} bpd | util fase1={w['rnest_util_fase1']:.1%} | util plan={w['rnest_util_planejada']:.1%}")
    print(f"Brasil avg: {w['brasil_avg_bpd']:,.0f} bpd | util={w['brasil_util']:.1%}")
    print(f"RNEST Capex/bbl (A/19bi): US$ {h['rnest_capex_por_barril_usd']:,.2f}")
    print(f"RNEST capital charge/bbl: US$ {h['rnest_carga_capital_por_barril_usd']:,.2f}")
    print(f"Parque@18k Capex/bbl: US$ {h['parque_18k_capex_por_barril_usd']:,.2f}")
    print(f"Parque@18k charge/bbl: US$ {h['parque_18k_carga_capital_por_barril_usd']:,.2f}")
    print(f"Múltiplo Capex/bbl: {h['multiplo_rnest_vs_parque18k_capex_por_barril']:.2f}x")
    print(f"Wrote {json_path}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    main()
