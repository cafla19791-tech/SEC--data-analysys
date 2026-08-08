#!/usr/bin/env python3
"""DBGG counterfactual: no selected tax expenditures + SELIC 4% + zero 2020 net emissions."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, numbers

ROOT = Path(__file__).resolve().parent
SRC_DBGG = ROOT / "Dbggindexp.xlsx"
SRC_OSU = ROOT.parent / "osu_2025" / "osu_2025-anexos-publicacao.xlsx"
OUT_XLSX = ROOT / "dbgg_gt_selic4_sem_emissoes_2020.xlsx"
OUT_JSON = ROOT / "dbgg_gt_selic4_sem_emissoes_2020.json"

SELIC_AA = 0.04
MONTHS_PT = {
    "Jan": 1,
    "Fev": 2,
    "Mar": 3,
    "Abr": 4,
    "Mai": 5,
    "Jun": 6,
    "Jul": 7,
    "Ago": 8,
    "Set": 9,
    "Out": 10,
    "Nov": 11,
    "Dez": 12,
}

# OSU Tab_1 benefit names (exact)
GT_NAMES = [
    "Desenvolvimento Regional",
    "Pesquisas Científicas e Inovação Tecnológica",
    "Informática e Automação",
]


def load_dbgg_series(ws, start_scan: int = 1):
    rows = []
    year = None
    for r in range(start_scan, ws.max_row + 1):
        y = ws.cell(r, 1).value
        m = ws.cell(r, 2).value
        if y is not None:
            try:
                year = int(y)
            except (TypeError, ValueError):
                pass
        if m is None or year is None or m not in MONTHS_PT:
            continue
        rows.append(
            {
                "year": year,
                "month": MONTHS_PT[m],
                "selic": float(ws.cell(r, 10).value or 0),
                "total": float(ws.cell(r, 15).value or 0),
            }
        )
    return rows


def load_osu_gt(path: Path) -> dict[int, dict[str, float]]:
    """Return {year: {name: R$ milhões}} from OSU Tab_1 (source in R$ mil)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tab_1"]
    years = []
    for c in range(4, ws.max_column + 1):
        y = ws.cell(3, c).value
        if y is not None:
            years.append((c, int(y)))
    out: dict[int, dict[str, float]] = {y: {} for _, y in years}
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name not in GT_NAMES:
            continue
        for c, y in years:
            v = ws.cell(r, c).value
            # OSU: R$ mil → BCB: R$ milhões
            out[y][name] = (float(v) / 1000.0) if isinstance(v, (int, float)) else 0.0
    return out


def compound_preperiod_gt(gt: dict[int, dict[str, float]], rate: float) -> float:
    """Compound 2003–2006 annual GT to Dec/2006 at constant annual rate (year-end)."""
    total = 0.0
    for y in range(2003, 2007):
        annual = sum(gt.get(y, {}).get(n, 0.0) for n in GT_NAMES)
        years_ahead = 2006 - y
        total += annual * ((1.0 + rate) ** years_ahead)
    return total


def simulate():
    wb = openpyxl.load_workbook(SRC_DBGG, data_only=True)
    div = load_dbgg_series(wb["DividaR$"])
    prim = load_dbgg_series(wb["PrimarioR$"])
    juros = load_dbgg_series(wb["JurosR$"])

    D = {(x["year"], x["month"]): x for x in div}
    P = {(x["year"], x["month"]): x for x in prim}
    J = {(x["year"], x["month"]): x for x in juros}

    gt = load_osu_gt(SRC_OSU)
    pre_gt = compound_preperiod_gt(gt, SELIC_AA)
    r_m = (1.0 + SELIC_AA) ** (1.0 / 12.0) - 1.0

    start = D[(2006, 12)]
    selic0_actual = start["selic"]
    selic0_cf = selic0_actual - pre_gt
    nonselic_cum_emis_removed = 0.0

    months = []
    stock_selic = selic0_cf
    annual = {}

    keys = sorted(k for k in D if k >= (2007, 1) and k in P and k in J)
    for y, m in keys:
        prev_y, prev_m = (y, m - 1) if m > 1 else (y - 1, 12)
        prev_s = D[(prev_y, prev_m)]["selic"]
        cur_s = D[(y, m)]["selic"]
        cur_t = D[(y, m)]["total"]
        emis_s = P[(y, m)]["selic"]
        emis_t = P[(y, m)]["total"]
        juros_s = J[(y, m)]["selic"]
        residual = cur_s - prev_s - emis_s - juros_s

        gt_year = sum(gt.get(y, {}).get(n, 0.0) for n in GT_NAMES) if y in gt else 0.0
        gt_month = gt_year / 12.0

        emis_s_base = 0.0 if y == 2020 else emis_s
        emis_s_cf = emis_s_base - gt_month

        emis_ns = emis_t - emis_s
        emis_ns_removed = emis_ns if y == 2020 else 0.0
        nonselic_cum_emis_removed += emis_ns_removed

        j_cf = stock_selic * r_m
        stock_selic = stock_selic + j_cf + emis_s_cf + residual

        nonselic_actual = cur_t - cur_s
        nonselic_cf = nonselic_actual - nonselic_cum_emis_removed
        dbgg_cf = nonselic_cf + stock_selic
        dbgg_actual = cur_t

        row = {
            "year": y,
            "month": m,
            "period": f"{y}-{m:02d}",
            "selic_stock_actual": cur_s,
            "selic_stock_cf": stock_selic,
            "emis_selic_actual": emis_s,
            "emis_selic_cf": emis_s_cf,
            "emis_selic_removed_2020": emis_s if y == 2020 else 0.0,
            "emis_nonselic_removed_2020": emis_ns_removed,
            "gt_month": gt_month,
            "juros_selic_actual": juros_s,
            "juros_selic_cf": j_cf,
            "juros_saved": juros_s - j_cf,
            "residual": residual,
            "nonselic_actual": nonselic_actual,
            "nonselic_cf": nonselic_cf,
            "dbgg_total_actual": dbgg_actual,
            "dbgg_total_cf": dbgg_cf,
            "dbgg_diff": dbgg_actual - dbgg_cf,
        }
        months.append(row)

        a = annual.setdefault(
            y,
            {
                "year": y,
                "gt_desenv_regional": gt.get(y, {}).get(GT_NAMES[0], 0.0),
                "gt_pesquisa_inovacao": gt.get(y, {}).get(GT_NAMES[1], 0.0),
                "gt_informatica": gt.get(y, {}).get(GT_NAMES[2], 0.0),
                "gt_total": gt_year,
                "emis_selic_actual": 0.0,
                "emis_selic_cf": 0.0,
                "emis_selic_removed_2020": 0.0,
                "emis_nonselic_removed_2020": 0.0,
                "juros_selic_actual": 0.0,
                "juros_selic_cf": 0.0,
                "juros_saved": 0.0,
            },
        )
        a["emis_selic_actual"] += emis_s
        a["emis_selic_cf"] += emis_s_cf
        a["emis_selic_removed_2020"] += row["emis_selic_removed_2020"]
        a["emis_nonselic_removed_2020"] += emis_ns_removed
        a["juros_selic_actual"] += juros_s
        a["juros_selic_cf"] += j_cf
        a["juros_saved"] += row["juros_saved"]
        a["selic_stock_actual_end"] = cur_s
        a["selic_stock_cf_end"] = stock_selic
        a["dbgg_actual_end"] = dbgg_actual
        a["dbgg_cf_end"] = dbgg_cf
        a["dbgg_diff_end"] = dbgg_actual - dbgg_cf

    end = months[-1]
    summary = {
        "source_dbgg": "https://www.bcb.gov.br/content/estatisticas/Documents/Tabelas_especiais/Dbggindexp.xlsx",
        "source_osu": (
            "https://www.gov.br/planejamento/pt-br/assuntos/avaliacao-de-politicas-publicas/"
            "arquivos/orcamento-de-subsidios-da-uniao/osu_2025-anexos-publicacao.xlsx"
        ),
        "selic_counterfactual_aa": SELIC_AA,
        "gt_categories": GT_NAMES,
        "assumption": [
            "SELIC counterfactual = 4% a.a. every month from Jan/2007 to end of sample",
            "Monthly rate = (1.04)^(1/12)-1 on beginning-of-month Selic CF stock",
            "Remove OSU tax expenditures: Desenvolvimento Regional; Pesquisas Científicas e Inovação Tecnológica; Informática e Automação",
            "Annual GT (OSU R$ mil) converted to R$ milhões and spread evenly over 12 months as extra primary surplus (negative Selic net emission)",
            "GT 2003–2006 compounded to Dec/2006 at 4% a.a. (year-end) and subtracted from initial Selic stock",
            "GT applied through 2024 (last OSU year); 2025–2026 months get GT=0",
            "Net emissions in 2020 set to zero for all indexators (Selic emis_cf base=0; non-Selic cumulative emissions removed from non-Selic stock)",
            "Outside 2020, Selic residual (ΔS-E-J) kept as actual; non-Selic path = actual non-Selic minus cumulative 2020 non-Selic emissions removed",
            "DBGG_cf = nonSelic_cf + Selic_cf",
            "Units: R$ milhões",
        ],
        "initial_r$_mi": {
            "selic_stock_actual_dez2006": selic0_actual,
            "gt_2003_2006_compounded_4pct": pre_gt,
            "selic_stock_cf_dez2006": selic0_cf,
        },
        "totals_r$_mi": {
            "selic_stock_actual_end": end["selic_stock_actual"],
            "selic_stock_cf_end": end["selic_stock_cf"],
            "selic_stock_reduction": end["selic_stock_actual"] - end["selic_stock_cf"],
            "juros_selic_actual_cum": sum(x["juros_selic_actual"] for x in months),
            "juros_selic_cf_cum": sum(x["juros_selic_cf"] for x in months),
            "juros_saved_cum": sum(x["juros_saved"] for x in months),
            "gt_applied_2007_2024_undiscounted": sum(
                sum(gt.get(y, {}).get(n, 0.0) for n in GT_NAMES) for y in range(2007, 2025)
            ),
            "gt_pre_2003_2006_compounded": pre_gt,
            "emis_selic_2020_removed": sum(x["emis_selic_removed_2020"] for x in months),
            "emis_nonselic_2020_removed": sum(x["emis_nonselic_removed_2020"] for x in months),
            "emis_total_2020_removed": sum(
                x["emis_selic_removed_2020"] + x["emis_nonselic_removed_2020"] for x in months
            ),
            "dbgg_actual_end": end["dbgg_total_actual"],
            "dbgg_cf_end": end["dbgg_total_cf"],
            "dbgg_reduction": end["dbgg_diff"],
        },
        "annual": [annual[y] for y in sorted(annual)],
        "months": months,
    }
    return summary


def write_xlsx(summary: dict) -> None:
    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)

    # Resumo
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "DBGG contrafactual combinado"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Premissas"
    ws["A3"].font = header_font
    for i, a in enumerate(summary["assumption"], start=4):
        ws[f"A{i}"] = a
    r0 = 4 + len(summary["assumption"]) + 1
    ws[f"A{r0}"] = "Totais (R$ milhões)"
    ws[f"A{r0}"].font = header_font
    r = r0 + 1
    for k, v in summary["initial_r$_mi"].items():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1
    for k, v in summary["totals_r$_mi"].items():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22

    # Anual
    ws = wb.create_sheet("Anual")
    headers = [
        "year",
        "gt_desenv_regional",
        "gt_pesquisa_inovacao",
        "gt_informatica",
        "gt_total",
        "emis_selic_actual",
        "emis_selic_cf",
        "emis_selic_removed_2020",
        "emis_nonselic_removed_2020",
        "juros_selic_actual",
        "juros_selic_cf",
        "juros_saved",
        "selic_stock_actual_end",
        "selic_stock_cf_end",
        "dbgg_actual_end",
        "dbgg_cf_end",
        "dbgg_diff_end",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(summary["annual"], start=2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(i, c, row.get(h))
            cell.border = thin
            if c > 1:
                cell.number_format = "#,##0.00"

    # Mensal
    ws = wb.create_sheet("Mensal")
    m_headers = [
        "period",
        "year",
        "month",
        "selic_stock_actual",
        "selic_stock_cf",
        "emis_selic_actual",
        "emis_selic_cf",
        "gt_month",
        "emis_selic_removed_2020",
        "emis_nonselic_removed_2020",
        "juros_selic_actual",
        "juros_selic_cf",
        "juros_saved",
        "residual",
        "nonselic_actual",
        "nonselic_cf",
        "dbgg_total_actual",
        "dbgg_total_cf",
        "dbgg_diff",
    ]
    for c, h in enumerate(m_headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(summary["months"], start=2):
        for c, h in enumerate(m_headers, 1):
            cell = ws.cell(i, c, row.get(h))
            cell.border = thin
            if h not in ("period", "year", "month"):
                cell.number_format = "#,##0.00"

    # GT OSU
    ws = wb.create_sheet("GT_OSU")
    ws["A1"] = "Gastos tributários removidos (R$ milhões, de OSU R$ mil / 1000)"
    ws["A1"].font = header_font
    ws["A2"] = "year"
    for c, n in enumerate(GT_NAMES, 2):
        ws.cell(2, c, n)
    ws.cell(2, 5, "total")
    # rebuild from annual
    for i, row in enumerate(summary["annual"], start=3):
        ws.cell(i, 1, row["year"])
        ws.cell(i, 2, row["gt_desenv_regional"])
        ws.cell(i, 3, row["gt_pesquisa_inovacao"])
        ws.cell(i, 4, row["gt_informatica"])
        ws.cell(i, 5, row["gt_total"])

    wb.save(OUT_XLSX)


def main():
    summary = simulate()
    OUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(summary)
    t = summary["totals_r$_mi"]
    print("End period:", summary["months"][-1]["period"])
    print("DBGG actual (R$ bi):", t["dbgg_actual_end"] / 1000)
    print("DBGG CF (R$ bi):", t["dbgg_cf_end"] / 1000)
    print("Reduction (R$ bi):", t["dbgg_reduction"] / 1000)
    print("Selic reduction (R$ bi):", t["selic_stock_reduction"] / 1000)
    print("GT 2007-24 (R$ bi):", t["gt_applied_2007_2024_undiscounted"] / 1000)
    print("GT pre compounded (R$ bi):", t["gt_pre_2003_2006_compounded"] / 1000)
    print("Emis 2020 removed (R$ bi):", t["emis_total_2020_removed"] / 1000)
    print("Wrote", OUT_XLSX)
    print("Wrote", OUT_JSON)


if __name__ == "__main__":
    main()
