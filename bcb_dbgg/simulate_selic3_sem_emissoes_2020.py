#!/usr/bin/env python3
"""DBGG counterfactual: SELIC 3% a.a. + zero net emissions in 2020."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Border, Side

ROOT = Path(__file__).resolve().parent
SRC_DBGG = ROOT / "Dbggindexp.xlsx"
OUT_XLSX = ROOT / "dbgg_selic_3pct_sem_emissoes_2020.xlsx"
OUT_JSON = ROOT / "dbgg_selic_3pct_sem_emissoes_2020.json"

SELIC_AA = 0.03
MONTHS_PT = {
    "Jan": 1,
    "Fev": 2,
    "Mar": 3,
    "Abr": 4,
    "Mai": 5,
    "Jun": 6,
    "Jul": 7,
    "Ago": 8,
    "Set": 9,
    "Out": 10,
    "Nov": 11,
    "Dez": 12,
}


def load_dbgg_series(ws, start_scan: int = 1):
    rows = []
    year = None
    for r in range(start_scan, ws.max_row + 1):
        y = ws.cell(r, 1).value
        m = ws.cell(r, 2).value
        if y is not None:
            try:
                year = int(y)
            except (TypeError, ValueError):
                pass
        if m is None or year is None or m not in MONTHS_PT:
            continue
        rows.append(
            {
                "year": year,
                "month": MONTHS_PT[m],
                "selic": float(ws.cell(r, 10).value or 0),
                "total": float(ws.cell(r, 15).value or 0),
            }
        )
    return rows


def simulate():
    wb = openpyxl.load_workbook(SRC_DBGG, data_only=True)
    div = load_dbgg_series(wb["DividaR$"])
    prim = load_dbgg_series(wb["PrimarioR$"])
    juros = load_dbgg_series(wb["JurosR$"])

    D = {(x["year"], x["month"]): x for x in div}
    P = {(x["year"], x["month"]): x for x in prim}
    J = {(x["year"], x["month"]): x for x in juros}

    r_m = (1.0 + SELIC_AA) ** (1.0 / 12.0) - 1.0
    selic0 = D[(2006, 12)]["selic"]
    stock_selic = selic0
    nonselic_cum_emis_removed = 0.0

    months = []
    annual = {}
    keys = sorted(k for k in D if k >= (2007, 1) and k in P and k in J)

    for y, m in keys:
        prev_y, prev_m = (y, m - 1) if m > 1 else (y - 1, 12)
        prev_s = D[(prev_y, prev_m)]["selic"]
        cur_s = D[(y, m)]["selic"]
        cur_t = D[(y, m)]["total"]
        emis_s = P[(y, m)]["selic"]
        emis_t = P[(y, m)]["total"]
        juros_s = J[(y, m)]["selic"]
        residual = cur_s - prev_s - emis_s - juros_s

        emis_s_cf = 0.0 if y == 2020 else emis_s
        emis_ns = emis_t - emis_s
        emis_ns_removed = emis_ns if y == 2020 else 0.0
        nonselic_cum_emis_removed += emis_ns_removed

        j_cf = stock_selic * r_m
        stock_selic = stock_selic + j_cf + emis_s_cf + residual

        nonselic_actual = cur_t - cur_s
        nonselic_cf = nonselic_actual - nonselic_cum_emis_removed
        dbgg_cf = nonselic_cf + stock_selic

        row = {
            "year": y,
            "month": m,
            "period": f"{y}-{m:02d}",
            "selic_stock_actual": cur_s,
            "selic_stock_cf": stock_selic,
            "emis_selic_actual": emis_s,
            "emis_selic_cf": emis_s_cf,
            "emis_selic_removed_2020": emis_s if y == 2020 else 0.0,
            "emis_nonselic_removed_2020": emis_ns_removed,
            "juros_selic_actual": juros_s,
            "juros_selic_cf": j_cf,
            "juros_saved": juros_s - j_cf,
            "residual": residual,
            "nonselic_actual": nonselic_actual,
            "nonselic_cf": nonselic_cf,
            "dbgg_total_actual": cur_t,
            "dbgg_total_cf": dbgg_cf,
            "dbgg_diff": cur_t - dbgg_cf,
        }
        months.append(row)

        a = annual.setdefault(
            y,
            {
                "year": y,
                "emis_selic_actual": 0.0,
                "emis_selic_cf": 0.0,
                "emis_selic_removed_2020": 0.0,
                "emis_nonselic_removed_2020": 0.0,
                "juros_selic_actual": 0.0,
                "juros_selic_cf": 0.0,
                "juros_saved": 0.0,
            },
        )
        a["emis_selic_actual"] += emis_s
        a["emis_selic_cf"] += emis_s_cf
        a["emis_selic_removed_2020"] += row["emis_selic_removed_2020"]
        a["emis_nonselic_removed_2020"] += emis_ns_removed
        a["juros_selic_actual"] += juros_s
        a["juros_selic_cf"] += j_cf
        a["juros_saved"] += row["juros_saved"]
        a["selic_stock_actual_end"] = cur_s
        a["selic_stock_cf_end"] = stock_selic
        a["dbgg_actual_end"] = cur_t
        a["dbgg_cf_end"] = dbgg_cf
        a["dbgg_diff_end"] = cur_t - dbgg_cf

    end = months[-1]
    return {
        "source": "https://www.bcb.gov.br/content/estatisticas/Documents/Tabelas_especiais/Dbggindexp.xlsx",
        "selic_counterfactual_aa": SELIC_AA,
        "assumption": [
            "SELIC counterfactual = 3% a.a. every month from Jan/2007 to end of sample",
            "Monthly rate = (1.03)^(1/12)-1 on beginning-of-month Selic CF stock",
            "Net emissions in 2020 set to zero for all indexators",
            "Outside 2020, Selic net emissions (PrimarioR$) kept as actual",
            "Selic residual (ΔS - E - J) kept as actual",
            "Non-Selic path = actual non-Selic minus cumulative 2020 non-Selic emissions removed",
            "DBGG_cf = nonSelic_cf + Selic_cf",
            "Units: R$ milhões",
        ],
        "initial_r$_mi": {"selic_stock_dez2006": selic0},
        "totals_r$_mi": {
            "selic_stock_actual_end": end["selic_stock_actual"],
            "selic_stock_cf_end": end["selic_stock_cf"],
            "selic_stock_reduction": end["selic_stock_actual"] - end["selic_stock_cf"],
            "juros_selic_actual_cum": sum(x["juros_selic_actual"] for x in months),
            "juros_selic_cf_cum": sum(x["juros_selic_cf"] for x in months),
            "juros_saved_cum": sum(x["juros_saved"] for x in months),
            "emis_selic_2020_removed": sum(x["emis_selic_removed_2020"] for x in months),
            "emis_nonselic_2020_removed": sum(x["emis_nonselic_removed_2020"] for x in months),
            "emis_total_2020_removed": sum(
                x["emis_selic_removed_2020"] + x["emis_nonselic_removed_2020"] for x in months
            ),
            "dbgg_actual_end": end["dbgg_total_actual"],
            "dbgg_cf_end": end["dbgg_total_cf"],
            "dbgg_reduction": end["dbgg_diff"],
        },
        "annual": [annual[y] for y in sorted(annual)],
        "months": months,
    }


def write_xlsx(summary: dict) -> None:
    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)

    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "DBGG contrafactual: SELIC 3% + sem emissões 2020"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Premissas"
    ws["A3"].font = header_font
    for i, a in enumerate(summary["assumption"], start=4):
        ws[f"A{i}"] = a
    r = 4 + len(summary["assumption"]) + 1
    ws[f"A{r}"] = "Totais (R$ milhões)"
    ws[f"A{r}"].font = header_font
    r += 1
    for k, v in {**summary["initial_r$_mi"], **summary["totals_r$_mi"]}.items():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    # Decomposicao simples
    ws = wb.create_sheet("Decomposicao")
    ws["A1"] = "Referência (R$ bilhões)"
    ws["A1"].font = header_font
    t = summary["totals_r$_mi"]
    rows = [
        ("DBGG real (fim)", t["dbgg_actual_end"] / 1000),
        ("DBGG contrafactual (fim)", t["dbgg_cf_end"] / 1000),
        ("Redução total DBGG", t["dbgg_reduction"] / 1000),
        ("Redução estoque Selic", t["selic_stock_reduction"] / 1000),
        ("Emissões 2020 removidas (total)", t["emis_total_2020_removed"] / 1000),
        ("  — Selic", t["emis_selic_2020_removed"] / 1000),
        ("  — Não-Selic", t["emis_nonselic_2020_removed"] / 1000),
        ("Juros Selic economizados (acumulado)", t["juros_saved_cum"] / 1000),
    ]
    ws["A3"] = "Métrica"
    ws["B3"] = "R$ bi"
    ws["A3"].font = header_font
    ws["B3"].font = header_font
    for i, (name, val) in enumerate(rows, start=4):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = val
    ws.column_dimensions["A"].width = 44

    ws = wb.create_sheet("Anual")
    headers = [
        "year",
        "emis_selic_actual",
        "emis_selic_cf",
        "emis_selic_removed_2020",
        "emis_nonselic_removed_2020",
        "juros_selic_actual",
        "juros_selic_cf",
        "juros_saved",
        "selic_stock_actual_end",
        "selic_stock_cf_end",
        "dbgg_actual_end",
        "dbgg_cf_end",
        "dbgg_diff_end",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(summary["annual"], start=2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(i, c, row.get(h))
            cell.border = thin
            if c > 1:
                cell.number_format = "#,##0.00"

    ws = wb.create_sheet("Mensal")
    m_headers = [
        "period",
        "year",
        "month",
        "selic_stock_actual",
        "selic_stock_cf",
        "emis_selic_actual",
        "emis_selic_cf",
        "emis_selic_removed_2020",
        "emis_nonselic_removed_2020",
        "juros_selic_actual",
        "juros_selic_cf",
        "juros_saved",
        "residual",
        "nonselic_actual",
        "nonselic_cf",
        "dbgg_total_actual",
        "dbgg_total_cf",
        "dbgg_diff",
    ]
    for c, h in enumerate(m_headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.border = thin
    for i, row in enumerate(summary["months"], start=2):
        for c, h in enumerate(m_headers, 1):
            cell = ws.cell(i, c, row.get(h))
            cell.border = thin
            if h not in ("period", "year", "month"):
                cell.number_format = "#,##0.00"

    wb.save(OUT_XLSX)


def main():
    summary = simulate()
    OUT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(summary)
    t = summary["totals_r$_mi"]
    print("End:", summary["months"][-1]["period"])
    print(f"DBGG actual R$ bi: {t['dbgg_actual_end']/1000:.2f}")
    print(f"DBGG CF     R$ bi: {t['dbgg_cf_end']/1000:.2f}")
    print(f"Reduction   R$ bi: {t['dbgg_reduction']/1000:.2f}")
    print(f"Selic red.  R$ bi: {t['selic_stock_reduction']/1000:.2f}")
    print(f"Emis 2020   R$ bi: {t['emis_total_2020_removed']/1000:.2f}")
    print(f"Juros saved R$ bi: {t['juros_saved_cum']/1000:.2f}")
    print("Wrote", OUT_XLSX)


if __name__ == "__main__":
    main()
