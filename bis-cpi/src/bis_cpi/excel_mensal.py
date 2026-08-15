"""Excel: 1 aba/país com série mensal de CPI (índice + YoY + acumulada)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import area_names, loaders

_INVALID_SHEET = re.compile(r"[\\/*?:\[\]]")


def sheet_name(code: str, name: str) -> str:
    base = f"{code} - {name}"
    base = _INVALID_SHEET.sub("-", base).strip() or code
    return base[:31]


def _autosize(ws, max_col: int = 8, max_width: int = 40) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        width = 12
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _print_layout(ws, header_left: str | None = None) -> None:
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.75 if header_left else 0.5, bottom=0.5)
    if header_left:
        ws.oddHeader.left.text = header_left
        ws.evenHeader.left.text = header_left


def _build_country_rows(
    index_pts: list[tuple[str, float]],
    yoy_pts: list[tuple[str, float]],
) -> list[tuple[str, float | None, float | None, float | None]]:
    yoy_map = dict(yoy_pts)
    if not index_pts:
        return []
    base = index_pts[0][1]
    rows: list[tuple[str, float | None, float | None, float | None]] = []
    for period, idx in index_pts:
        acum = ((idx / base) - 1.0) * 100.0 if base else None
        rows.append((period, idx, yoy_map.get(period), acum))
    return rows


def gerar_excel_mensal(
    csv_path: Path,
    out_path: Path,
    *,
    areas: set[str] | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Path:
    idx, names_en = loaders.load_series(
        csv_path, freq="M", unit=loaders.UNIT_INDEX, areas=areas,
        date_from=date_from, date_to=date_to,
    )
    yoy, _ = loaders.load_series(
        csv_path, freq="M", unit=loaders.UNIT_YOY, areas=areas,
        date_from=date_from, date_to=date_to,
    )

    wb = Workbook()
    # Legenda
    ws0 = wb.active
    ws0.title = "00_Legenda"
    legend = [
        ["BIS WS_LONG_CPI — séries mensais por país"],
        ["Fonte", "https://data.bis.org/static/bulk/WS_LONG_CPI_csv_col.zip"],
        [],
        ["Colunas"],
        ["Mês", "YYYY-MM"],
        ["Índice (2010=100)", "UNIT_MEASURE 628"],
        ["Variação YoY (%)", "UNIT_MEASURE 771 — variação acumulada em 12 meses"],
        [
            "Inflação acumulada (%)",
            "(Índice_t / Índice_primeiro_mês_da_série − 1) × 100",
        ],
        [],
        ["Países", str(len(idx))],
    ]
    for row in legend:
        ws0.append(row)
    ws0["A1"].font = Font(bold=True, size=14)
    _autosize(ws0, 2)

    # Índice
    ws_i = wb.create_sheet("01_Indice")
    ws_i.append(["Código", "País", "Nome BIS", "Início", "Fim", "N obs."])
    for code in sorted(idx):
        pts = idx[code]
        nome = area_names.display_name(code, names_en.get(code, code))
        ws_i.append([
            code, nome, names_en.get(code, ""),
            pts[0][0], pts[-1][0], len(pts),
        ])
    _autosize(ws_i, 6)
    _print_layout(ws_i)
    for row in ws_i.iter_rows(min_row=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for code in sorted(idx):
        nome = area_names.display_name(code, names_en.get(code, code))
        ws = wb.create_sheet(sheet_name(code, nome))
        ws.append([
            "Mês",
            "Índice (2010=100)",
            "Variação YoY (%)",
            "Inflação acumulada (%)",
        ])
        for period, ix, yy, acum in _build_country_rows(idx[code], yoy.get(code, [])):
            ws.append([period, ix, yy, acum])
        ws.freeze_panes = "A2"
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for row in ws.iter_rows(min_row=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        _print_layout(ws, header_left=nome)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
