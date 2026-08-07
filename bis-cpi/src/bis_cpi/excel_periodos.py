"""Rankings de inflação acumulada por períodos (via índice CPI)."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import area_names, loaders


@dataclass(frozen=True)
class Periodo:
    sheet: str
    titulo: str
    inicio: str  # YYYY-MM
    fim: str  # YYYY-MM


# Mesmos recortes usados no CBPOL (ajustados a competência mensal).
PERIODOS: list[Periodo] = [
    Periodo("01_1995_a_2002", "Inflação acumulada 01/1995 a 12/2002", "1995-01", "2002-12"),
    Periodo("02_2003_a_2016-04", "Inflação acumulada 01/2003 a 04/2016", "2003-01", "2016-04"),
    Periodo("03_2016-05_a_2018", "Inflação acumulada 05/2016 a 12/2018", "2016-05", "2018-12"),
    Periodo("04_2019_a_2022", "Inflação acumulada 01/2019 a 12/2022", "2019-01", "2022-12"),
    Periodo("05_2023_a_2026-06", "Inflação acumulada 01/2023 a 06/2026", "2023-01", "2026-06"),
    Periodo("06_2003_a_2026-06", "Inflação acumulada 01/2003 a 06/2026", "2003-01", "2026-06"),
]


def _nearest_on_or_after(pts: list[tuple[str, float]], target: str) -> tuple[str, float] | None:
    for p, v in pts:
        if p >= target:
            return p, v
    return None


def _nearest_on_or_before(pts: list[tuple[str, float]], target: str) -> tuple[str, float] | None:
    hit = None
    for p, v in pts:
        if p <= target:
            hit = (p, v)
        else:
            break
    return hit


def inflacao_periodo(
    pts: list[tuple[str, float]],
    inicio: str,
    fim: str,
) -> dict | None:
    """Inflação = (I_fim / I_inicio − 1) × 100 usando extremos disponíveis no intervalo."""
    a = _nearest_on_or_after(pts, inicio)
    b = _nearest_on_or_before(pts, fim)
    if not a or not b:
        return None
    if a[0] > fim or b[0] < inicio or a[0] > b[0]:
        return None
    if a[1] == 0:
        return None
    infl = (b[1] / a[1] - 1.0) * 100.0
    return {
        "inicio_obs": a[0],
        "fim_obs": b[0],
        "indice_inicio": a[1],
        "indice_fim": b[1],
        "inflacao_pct": infl,
    }


def gerar_excel_periodos(
    csv_path: Path,
    out_path: Path,
    *,
    areas: set[str] | None = None,
    periodos: list[Periodo] | None = None,
) -> Path:
    idx, names_en = loaders.load_series(
        csv_path, freq="M", unit=loaders.UNIT_INDEX, areas=areas,
    )
    periodos = periodos or PERIODOS

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "00_Legenda"
    ws0.append(["BIS WS_LONG_CPI — inflação acumulada por período (ranking)"])
    ws0.append(["Método", "(Índice_fim / Índice_início − 1) × 100 no intervalo"])
    ws0.append(["Fonte", "https://data.bis.org/static/bulk/WS_LONG_CPI_csv_col.zip"])
    ws0.append([])
    ws0.append(["Período", "Início", "Fim"])
    for p in periodos:
        ws0.append([p.titulo, p.inicio, p.fim])
    ws0["A1"].font = Font(bold=True, size=14)
    for col in range(1, 4):
        ws0.column_dimensions[get_column_letter(col)].width = 48

    for per in periodos:
        ws = wb.create_sheet(per.sheet[:31])
        ws.append([per.titulo])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])
        ws.append([
            "Rank",
            "Código",
            "País",
            "Inflação acumulada (%)",
            "Índice início",
            "Índice fim",
            "Mês início obs.",
            "Mês fim obs.",
        ])
        rows = []
        for code, pts in idx.items():
            info = inflacao_periodo(pts, per.inicio, per.fim)
            if not info:
                continue
            rows.append((code, info))
        rows.sort(key=lambda x: x[1]["inflacao_pct"], reverse=True)
        for i, (code, info) in enumerate(rows, 1):
            nome = area_names.display_name(code, names_en.get(code, code))
            ws.append([
                i,
                code,
                nome,
                info["inflacao_pct"],
                info["indice_inicio"],
                info["indice_fim"],
                info["inicio_obs"],
                info["fim_obs"],
            ])
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 24
        from openpyxl.styles import Alignment
        from openpyxl.worksheet.page import PageMargins

        for row in ws.iter_rows(min_row=3, max_col=8):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
