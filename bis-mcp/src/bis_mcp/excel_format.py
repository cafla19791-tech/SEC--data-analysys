"""Helpers de formatacao Excel (largura de colunas e alinhamento)."""

from __future__ import annotations

from typing import Any, Iterable, Sequence


def _display_width(value: Any) -> int:
    if value is None:
        return 0
    try:
        import math

        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return 0
    except Exception:
        pass
    s = str(value)
    # Excel width ~ ASCII chars; accented letters still count as 1 visually enough.
    return len(s)


def column_widths_for_frame(
    columns: Sequence[Any],
    rows: Iterable[Sequence[Any]] | None = None,
    *,
    min_width: float = 10,
    max_width: float = 60,
    padding: float = 3,
) -> list[float]:
    """Calcula larguras para caber rotulo + conteudo (aprox. chars Excel)."""
    widths = [max(min_width, _display_width(c) + padding) for c in columns]
    if rows is not None:
        for row in rows:
            for i, cell in enumerate(row):
                if i >= len(widths):
                    break
                widths[i] = max(widths[i], min(max_width, _display_width(cell) + padding))
    return [min(max_width, w) for w in widths]


def apply_column_widths(
    worksheet: Any,
    widths: Sequence[float],
    *,
    engine: str,
    formats: Sequence[Any] | None = None,
) -> None:
    """Aplica larguras no worksheet (xlsxwriter ou openpyxl)."""
    if engine == "xlsxwriter":
        for i, width in enumerate(widths):
            fmt = formats[i] if formats and i < len(formats) else None
            if fmt is not None:
                worksheet.set_column(i, i, width, fmt)
            else:
                worksheet.set_column(i, i, width)
        return

    # openpyxl
    from openpyxl.utils import get_column_letter

    for i, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = width


def make_center_formats(
    workbook: Any,
    num_formats: Sequence[str | None],
) -> list[Any]:
    """Cria formatos xlsxwriter centrados (opcionalmente com num_format)."""
    out: list[Any] = []
    for nf in num_formats:
        opts: dict[str, Any] = {"align": "center", "valign": "vcenter"}
        if nf:
            opts["num_format"] = nf
        out.append(workbook.add_format(opts))
    return out


def center_align_dataframe_sheet(
    writer: Any,
    sheet_name: str,
    df: Any,
    *,
    engine: str,
    header_row: int = 0,
    bold_header: bool = True,
    wrap_header: bool = True,
) -> None:
    """Centraliza cabecalhos (e dados no openpyxl) na aba.

    No xlsxwriter, os dados herdam alinhamento via set_column / col_formats
    (ver make_center_formats); aqui so reescrevemos a linha de cabecalho.
    """
    ws = writer.sheets[sheet_name]
    n_cols = len(df.columns)
    if n_cols == 0:
        return

    if engine == "xlsxwriter":
        header_fmt = writer.book.add_format(
            {
                "align": "center",
                "valign": "vcenter",
                "bold": bold_header,
                "text_wrap": wrap_header,
            }
        )
        for c, name in enumerate(df.columns):
            ws.write(header_row, c, name, header_fmt)
        return

    # openpyxl
    from openpyxl.styles import Alignment, Font

    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=wrap_header
    )
    cell_align = Alignment(horizontal="center", vertical="center")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row + 1, column=c)
        cell.alignment = header_align
        if bold_header:
            cell.font = Font(bold=True)

    n_rows = len(df)
    if n_rows == 0:
        return
    data_start = header_row + 2
    data_end = header_row + 1 + n_rows
    for row in ws.iter_rows(
        min_row=data_start, max_row=data_end, min_col=1, max_col=n_cols
    ):
        for cell in row:
            cell.alignment = cell_align


def autosize_dataframe_sheet(
    writer: Any,
    sheet_name: str,
    df: Any,
    *,
    engine: str,
    col_formats: Sequence[Any] | None = None,
    sample_rows: int = 200,
    min_width: float = 10,
    max_width: float = 60,
    padding: float = 3,
    extra_title_width: float | None = None,
    center: bool = False,
    header_row: int = 0,
) -> None:
    """Ajusta colunas de uma aba ja escrita com pandas.to_excel."""
    ws = writer.sheets[sheet_name]
    cols = list(df.columns)
    # Amostra para nao varrer milhoes de linhas
    sample = df.head(sample_rows)
    rows = sample.itertuples(index=False, name=None)
    widths = column_widths_for_frame(
        cols,
        rows,
        min_width=min_width,
        max_width=max_width,
        padding=padding,
    )
    if extra_title_width is not None and widths:
        widths[0] = max(widths[0], extra_title_width)

    if center and engine == "xlsxwriter" and col_formats is None:
        col_formats = make_center_formats(writer.book, [None] * len(cols))

    apply_column_widths(ws, widths, engine=engine, formats=col_formats)

    if center:
        center_align_dataframe_sheet(
            writer,
            sheet_name,
            df,
            engine=engine,
            header_row=header_row,
        )
