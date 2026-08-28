#!/usr/bin/env python3
"""Build Eletrobras (AXIA Energia) gross debt and interest paid tables 2002–2025."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "eletrobras"
FACTS = ROOT / "data/raw/eletrobras_companyfacts.json"

# Dívida bruta aproximada 2005–2015 (R$ bi) — pesquisa prévia em 20-F/releases
# (empréstimos + financiamentos + debêntures consolidados). Quebras IFRS 11 em 2013.
DEBT_EARLY_BI = {
    2005: 37.3,
    2006: 34.3,
    2007: 24.5,
    2008: 32.6,
    2009: 29.5,
    2010: 33.1,
    2011: 42.5,  # ~42–43 antes IFRS 11
    2012: 50.0,  # ~50 antes IFRS 11
    2013: 32.5,  # ~32–33 após IFRS 11
    2014: 45.0,  # aproximado (entre 2013 pós-IFRS11 e 2015)
    2015: 47.0,
}

# Juros pagos DFC 2009–2014 (R$ mi) — pesquisa prévia 20-F
INTEREST_EARLY_MI = {
    2009: 1104.0,
    2010: 1453.0,
    2011: 1368.0,
    2012: 1813.0,
    2013: 1306.0,
    2014: 1306.0,
}

# Estimativas 2005–2008 (sem linha DFC comparável)
INTEREST_EST_MI = {
    2005: 2536.0,
    2006: 2343.0,
    2007: 1722.0,
    2008: 2086.0,
}


def load_sec_series() -> tuple[dict[int, dict], dict[int, dict]]:
    facts = json.loads(FACTS.read_text())["facts"]["ifrs-full"]

    debt: dict[int, dict] = {}
    for r in facts["Borrowings"]["units"]["BRL"]:
        if r.get("form") not in ("20-F", "20-F/A"):
            continue
        end = r.get("end") or ""
        if not end.endswith("-12-31"):
            continue
        y = int(end[:4])
        val = float(r["val"]) / 1e6
        filed = r.get("filed") or ""
        prev = debt.get(y)
        # prefer later filing (captures restatements)
        if prev is None or filed > prev["filed"]:
            debt[y] = {
                "gross_debt_brl_mi": round(val, 1),
                "source": f"SEC 20-F Borrowings (filed {filed})",
                "filed": filed,
                "quality": "official",
            }

    # 2025 from 20-F Total loans line (not yet in companyfacts Borrowings end-2025)
    debt[2025] = {
        "gross_debt_brl_mi": 74295.8,
        "source": "SEC 20-F 2025 — Total loans, financing and debentures",
        "filed": "2026-04-09",
        "quality": "official",
    }

    interest: dict[int, dict] = {}
    for r in facts["InterestPaidClassifiedAsOperatingActivities"]["units"]["BRL"]:
        if r.get("form") not in ("20-F", "20-F/A"):
            continue
        start = r.get("start") or ""
        end = r.get("end") or ""
        if not (end.endswith("-12-31") and start[:4] == end[:4]):
            continue
        y = int(end[:4])
        val = abs(float(r["val"])) / 1e6
        filed = r.get("filed") or ""
        prev = interest.get(y)
        if prev is None or filed > prev["filed"]:
            interest[y] = {
                "interest_paid_brl_mi": round(val, 1),
                "source": f"SEC 20-F InterestPaid (operating) filed {filed}",
                "filed": filed,
                "quality": "official",
            }

    # 2025 Payment of interests
    interest[2025] = {
        "interest_paid_brl_mi": 5831.6,
        "source": "SEC 20-F 2025 — Payment of interests (DFC)",
        "filed": "2026-04-09",
        "quality": "official",
    }
    return debt, interest


def build_rows() -> list[dict]:
    sec_debt, sec_int = load_sec_series()
    rows = []
    for y in range(2002, 2026):
        if y in sec_debt:
            d = sec_debt[y]
            debt_mi = d["gross_debt_brl_mi"]
            debt_src = d["source"]
            debt_q = d["quality"]
        elif y in DEBT_EARLY_BI:
            debt_mi = DEBT_EARLY_BI[y] * 1000.0
            debt_src = "20-F/releases (série histórica consolidada; aproximado)"
            debt_q = "approx"
            if y in (2011, 2012):
                debt_src += " — antes do IFRS 11"
            if y == 2013:
                debt_src += " — após IFRS 11"
            if y == 2014:
                debt_src += " — interpolado/aproximado (lacuna na série homogênea)"
        else:
            debt_mi = None
            debt_src = "n/d — sem série consolidada homogênea de dívida bruta"
            debt_q = "missing"

        if y in sec_int:
            i = sec_int[y]
            int_mi = i["interest_paid_brl_mi"]
            int_src = i["source"]
            int_q = i["quality"]
        elif y in INTEREST_EARLY_MI:
            int_mi = INTEREST_EARLY_MI[y]
            int_src = "20-F DFC — Payment of interests / encargos financeiros"
            int_q = "official_research"
        elif y in INTEREST_EST_MI:
            int_mi = INTEREST_EST_MI[y]
            int_src = "estimativa (taxa média × dívida; sem linha DFC comparável)"
            int_q = "estimated"
        else:
            int_mi = None
            int_src = "n/d"
            int_q = "missing"

        rows.append(
            {
                "year": y,
                "gross_debt_brl_mi": debt_mi,
                "gross_debt_brl_bi": round(debt_mi / 1000.0, 2) if debt_mi is not None else None,
                "gross_debt_source": debt_src,
                "gross_debt_quality": debt_q,
                "interest_paid_brl_mi": int_mi,
                "interest_paid_source": int_src,
                "interest_paid_quality": int_q,
            }
        )
    return rows


def write_outputs(rows: list[dict]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    payload = {
        "company": "Eletrobras / AXIA Energia S.A.",
        "cik": "0001439124",
        "period": "2002-2025",
        "currency": "BRL",
        "notes": [
            "Dívida bruta = empréstimos + financiamentos + debêntures consolidados (Total loans / Borrowings).",
            "2016–2025: SEC companyfacts / 20-F (valores reapresentados mais recentes quando houver).",
            "2005–2015: série histórica de 20-F/releases (aproximada; quebra IFRS 11 em 2013).",
            "2002–2004: sem série consolidada homogênea de dívida bruta.",
            "Juros pagos = caixa (InterestPaid / Payment of interests no DFC).",
            "2015–2025: SEC XBRL; 2009–2014: DFC 20-F; 2005–2008: estimativa; 2002–2004: n/d.",
            "2020 juros: valor reapresentado no 20-F 2022 (R$ 1.370,7 mi); o 20-F 2020/2021 reportava R$ 1.701,1 mi.",
            "2023 dívida: valor reapresentado no 20-F 2024 (R$ 59.460 mi); o 20-F 2023 reportava R$ 60.780 mi.",
        ],
        "rows": rows,
    }
    (OUT / "eletrobras_divida_juros_2002_2025.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin", color="9AA7B5"),
        right=Side(style="thin", color="9AA7B5"),
        top=Side(style="thin", color="9AA7B5"),
        bottom=Side(style="thin", color="9AA7B5"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    zebra = PatternFill("solid", fgColor="F7F9FC")
    est_fill = PatternFill("solid", fgColor="FFF2CC")
    miss_fill = PatternFill("solid", fgColor="FCE4D6")

    def style_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

    def row_fill(quality: str, i: int):
        if quality in ("estimated", "approx"):
            return est_fill
        if quality == "missing":
            return miss_fill
        return zebra if i % 2 else None

    # Debt
    ws = wb.active
    ws.title = "Divida_Bruta"
    ws["A1"] = "Eletrobras (AXIA Energia) — Evolução da Dívida Bruta (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Empréstimos + financiamentos + debêntures consolidados (R$). Amarelo = aproximado; laranja = n/d."
    )
    headers = ["Ano", "Dívida bruta (R$ mi)", "Dívida bruta (R$ bi)", "Δ YoY %", "Qualidade", "Fonte"]
    style_header(ws, 4, headers)
    prev = None
    for i, r in enumerate(rows):
        yoy = None
        if prev and r["gross_debt_brl_bi"] is not None:
            yoy = (r["gross_debt_brl_bi"] / prev - 1) * 100
        vals = [
            r["year"],
            r["gross_debt_brl_mi"],
            r["gross_debt_brl_bi"],
            round(yoy, 1) if yoy is not None else None,
            r["gross_debt_quality"],
            r["gross_debt_source"],
        ]
        fill = row_fill(r["gross_debt_quality"], i)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(5 + i, c, v)
            cell.border = thin
            if fill:
                cell.fill = fill
            if c in (2,) and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"
            if c in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = "0.00" if c == 3 else "0.0"
        if r["gross_debt_brl_bi"] is not None:
            prev = r["gross_debt_brl_bi"]
    for col, w in enumerate([8, 22, 20, 12, 14, 70], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Interest
    ws = wb.create_sheet("Juros_Pagos")
    ws["A1"] = "Eletrobras (AXIA Energia) — Evolução dos Juros Pagos (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Juros/encargos financeiros pagos em caixa (DFC). Amarelo = estimativa; laranja = n/d."
    headers = ["Ano", "Juros pagos (R$ mi)", "Juros pagos (R$ bi)", "Δ YoY %", "Qualidade", "Fonte"]
    style_header(ws, 4, headers)
    prev = None
    for i, r in enumerate(rows):
        ip = r["interest_paid_brl_mi"]
        yoy = None
        if prev and ip is not None:
            yoy = (ip / prev - 1) * 100
        vals = [
            r["year"],
            ip,
            round(ip / 1000.0, 3) if ip is not None else None,
            round(yoy, 1) if yoy is not None else None,
            r["interest_paid_quality"],
            r["interest_paid_source"],
        ]
        fill = row_fill(r["interest_paid_quality"], i)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(5 + i, c, v)
            cell.border = thin
            if fill:
                cell.fill = fill
            if c == 2 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"
            if c in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = "0.000" if c == 3 else "0.0"
        if ip is not None:
            prev = ip
    for col, w in enumerate([8, 22, 20, 12, 16, 70], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Combined
    ws = wb.create_sheet("Consolidado")
    ws["A1"] = "Eletrobras — Dívida bruta e juros pagos (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Ano", "Dívida bruta (R$ bi)", "Juros pagos (R$ mi)"]
    style_header(ws, 3, headers)
    for i, r in enumerate(rows):
        for c, v in enumerate(
            [r["year"], r["gross_debt_brl_bi"], r["interest_paid_brl_mi"]], 1
        ):
            cell = ws.cell(4 + i, c, v)
            cell.border = thin
            if i % 2:
                cell.fill = zebra
            if c == 2 and isinstance(v, (int, float)):
                cell.number_format = "0.00"
            if c == 3 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"

    ws = wb.create_sheet("Notas")
    for i, n in enumerate(payload["notes"], 1):
        ws[f"A{i}"] = n
    ws.column_dimensions["A"].width = 120

    xlsx = OUT / "eletrobras_divida_juros_2002_2025.xlsx"
    wb.save(xlsx)

    md = [
        "# Eletrobras (AXIA Energia) — Dívida bruta e juros pagos (2002–2025)",
        "",
        "## Dívida bruta (R$ bi)",
        "",
        "| Ano | Dívida bruta | Qualidade |",
        "|---:|---:|---|",
    ]
    for r in rows:
        v = f"{r['gross_debt_brl_bi']:.2f}" if r["gross_debt_brl_bi"] is not None else "n/d"
        md.append(f"| {r['year']} | {v} | {r['gross_debt_quality']} |")
    md += ["", "## Juros pagos (R$ mi)", "", "| Ano | Juros pagos | Qualidade |", "|---:|---:|---|"]
    for r in rows:
        v = f"{r['interest_paid_brl_mi']:.1f}" if r["interest_paid_brl_mi"] is not None else "n/d"
        md.append(f"| {r['year']} | {v} | {r['interest_paid_quality']} |")
    (OUT / "eletrobras_divida_juros_2002_2025.md").write_text("\n".join(md) + "\n", encoding="utf-8")
    print("Wrote", xlsx)


def main():
    rows = build_rows()
    write_outputs(rows)
    for r in rows:
        print(
            f"{r['year']}: debt={r['gross_debt_brl_bi']} bi | interest={r['interest_paid_brl_mi']} ({r['interest_paid_quality']})"
        )


if __name__ == "__main__":
    main()
