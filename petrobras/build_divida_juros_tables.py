#!/usr/bin/env python3
"""Build Petrobras gross debt and interest paid tables (2002–2025)."""

from __future__ import annotations

import json
import re
import time
import urllib.request
from html import unescape
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "petrobras"
CACHE = Path("/tmp/pbr_20f")
INDEX = Path("/tmp/pbr_20f_index.json")
UA = {
    "User-Agent": "SEC-data-analysys research bot contact@example.com",
    "Accept-Encoding": "identity",
}

# Continuity series for Petrobras gross debt (US$ bi). Official overrides below.
WSN_GROSS_DEBT_USD_BI = {
    2002: 8.70,
    2003: 14.45,
    2004: 19.28,
    2005: 18.42,
    2006: 20.36,
    2007: 21.16,
    2008: 26.80,
    2009: 57.96,
    2010: 68.98,
    2011: 82.51,
    2012: 94.86,
    2013: 112.30,
    2014: 130.52,
    2015: 122.94,
    2016: 116.91,
    2017: 108.98,
    2018: 84.34,
    2019: 87.29,
    2020: 75.57,
    2021: 58.85,
    2022: 53.17,
    2023: 62.39,
    2024: 60.45,
    2025: 70.08,
}

OFFICIAL_GROSS_DEBT_USD_MI = {
    2019: 87121.0,
    2024: 60311.0,
    2025: 69793.0,
}


def fetch_filing(year: int, meta: tuple) -> Path:
    CACHE.mkdir(parents=True, exist_ok=True)
    _rd, _form, acc, doc, _fd = meta
    dest = CACHE / f"{year}_{doc}"
    if dest.exists() and dest.stat().st_size > 10_000:
        return dest
    url = f"https://www.sec.gov/Archives/edgar/data/1119639/{acc.replace('-', '')}/{doc}"
    req = urllib.request.Request(url, headers=UA)
    dest.write_bytes(urllib.request.urlopen(req, timeout=180).read())
    time.sleep(0.35)
    return dest


def parse_num(token: str) -> float | None:
    s = token.strip().replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    elif s.startswith("(") and not s.endswith(")"):
        # broken token; reject
        return None
    s = s.replace(",", "")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", s):
        return None
    v = float(s)
    return -v if neg else v


def numbers_after(html: str, start: int, limit: int = 1200) -> list[float]:
    window = re.sub(r"<[^>]+>", " ", html[start : start + limit])
    window = unescape(window)
    # normalize "(7,308 )" -> "(7,308)"
    window = re.sub(r"\(\s*([\d,]+(?:\.\d+)?)\s*\)", r"(\1)", window)
    vals = []
    for n in re.findall(
        r"\(\d{1,3}(?:,\d{3})+(?:\.\d+)?\)|\(\d{3,5}(?:\.\d+)?\)|\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d{3,5}(?:\.\d+)?",
        window,
    ):
        v = parse_num(n)
        if v is None:
            continue
        av = abs(v)
        if 50 <= av <= 30000:
            vals.append(av)
        if len(vals) >= 3:
            break
    return vals


def extract_interest_candidates(html: str, filing_year: int) -> list[tuple[int, float, str]]:
    """Return (year, value_usd_mi, source) candidates from one 20-F."""
    out: list[tuple[int, float, str]] = []
    text = unescape(re.sub(r"<[^>]+>", " ", html))
    text = re.sub(r"\s+", " ", text)

    # A) Modern finance-debt interest line (preferred from 2019+)
    for pat, src in [
        (
            r"(?i)Repayment of finance debt\s*[-–]\s*interest",
            "20-F CF: Repayment of finance debt - interest",
        ),
        (
            r"(?i)Repayment of interest\s*[-–]\s*finance debt",
            "20-F CF: Repayment of interest - finance debt",
        ),
        (
            r"(?i)(?<!lease )interest\s*[-–]\s*finance debt",
            "20-F CF: interest - finance debt",
        ),
    ]:
        for m in re.finditer(pat, html):
            vals = numbers_after(html, m.start())
            if not vals:
                continue
            for i, v in enumerate(vals[:3]):
                out.append((filing_year - i, v, src))
            return out  # strongest modern label

    # B) XBRL InterestPaidClassifiedAsFinancingActivities
    for m in re.finditer(
        r'<ix:nonFraction[^>]*name="ifrs-full:InterestPaidClassifiedAsFinancingActivities"[^>]*'
        r'contextRef="([^"]*)"[^>]*>([^<]+)</ix:nonFraction>',
        html,
        re.I,
    ):
        ctx, raw = m.group(1), m.group(2)
        ym = re.search(r"(20\d{2})-01-01.*?(20\d{2})-12-31", ctx)
        if not ym:
            # From2024-01-01to2024-12-31
            ym = re.search(r"(20\d{2})-01-01to(20\d{2})-12-31", ctx)
        if ym and ym.group(1) == ym.group(2):
            v = parse_num(raw)
            if v is not None:
                out.append(
                    (
                        int(ym.group(1)),
                        abs(v),
                        "20-F ix:InterestPaidClassifiedAsFinancingActivities",
                    )
                )
    if out:
        return out

    # C) Pre-IFRS16 / mid years: Repayment of interest (avoid tiny lease-only rows)
    for m in re.finditer(r"(?i)Repayment of interest(?!\s+on capital)", html):
        vals = numbers_after(html, m.start())
        # discard obvious lease-only tiny first values when all < 600 and filing>=2019
        if vals and not (filing_year >= 2019 and vals[0] < 600):
            for i, v in enumerate(vals[:3]):
                out.append((filing_year - i, v, "20-F CF: Repayment of interest"))
            return out

    # D) Early US-GAAP supplemental: Interest, net of amount capitalized (US$ mi)
    m = re.search(
        r"(?i)Cash paid during the (?:year|period) for Interest, net of amount capitalized\s+"
        r"([\d,\.\(\) ]{5,80})",
        text,
    )
    if m:
        vals = []
        for n in re.findall(r"\d{1,3}(?:,\d{3})*|\d{3,5}", m.group(1)):
            v = parse_num(n)
            if v is not None and 50 <= v <= 20000:
                vals.append(v)
        for i, v in enumerate(vals[:3]):
            out.append(
                (
                    filing_year - i,
                    v,
                    "20-F supplemental: Interest, net of amount capitalized",
                )
            )
        return out

    # E) Very early: Cash paid during the period for Interest <small millions>
    m = re.search(
        r"(?i)Cash paid during the (?:year|period) for Interest(?!,)\s+"
        r"([\d,\.\(\) ]{5,60})",
        text,
    )
    if m:
        vals = []
        for n in re.findall(r"\d{1,3}(?:,\d{3})*|\d{2,5}", m.group(1)):
            v = parse_num(n)
            if v is not None and 50 <= v <= 20000:
                vals.append(v)
            elif v is not None and v < 50:
                # 2002-era small values still valid
                if 20 <= v <= 20000:
                    vals.append(v)
        # avoid the R$-thousand supplemental tables (values like 1,517,259)
        vals = [v for v in vals if v < 100000]
        for i, v in enumerate(vals[:3]):
            out.append(
                (filing_year - i, v, "20-F supplemental: Cash paid for Interest")
            )
    return out


def load_sec_borrowings() -> dict[int, float]:
    path = ROOT / "data/raw/petrobras_companyfacts.json"
    if not path.exists():
        return {}
    rows = (
        json.loads(path.read_text())["facts"]["ifrs-full"]
        .get("Borrowings", {})
        .get("units", {})
        .get("USD", [])
    )
    out: dict[int, float] = {}
    for r in rows:
        if r.get("form") not in ("20-F", "20-F/A") or r.get("fp") != "FY":
            continue
        end = r.get("end") or ""
        fy = r.get("fy")
        if fy and end.endswith("-12-31"):
            out[int(fy)] = float(r["val"]) / 1_000_000.0
    return out


def build_interest_map(by_year: dict) -> dict[int, dict]:
    """Aggregate interest paid; prefer own-year filing, then later comparatives."""
    # rank: own filing first column > later comparative > older
    best: dict[int, tuple[int, int, float, str]] = {}
    # key year -> (priority, filing_year, value, source)
    for y in range(2002, 2026):
        meta = by_year[str(y)]
        path = fetch_filing(y, meta)
        html = path.read_text("utf-8", errors="ignore")
        cands = extract_interest_candidates(html, y)
        for year, val, src in cands:
            if year < 2002 or year > 2025:
                continue
            own = 1 if year == y else 0
            # higher own preferred; among same, prefer more recent filing
            prio = (own, y)
            prev = best.get(year)
            if prev is None or prio > (prev[0], prev[1]):
                best[year] = (own, y, val, src)
        print(f"filing {y}: extracted {cands[:3]}")
    return {
        y: {"interest_paid_usd_mi": round(v, 1), "interest_paid_source": src, "from_filing": fy}
        for y, (own, fy, v, src) in sorted(best.items())
    }


def build_rows(by_year: dict) -> list[dict]:
    interest = build_interest_map(by_year)
    borrowings = load_sec_borrowings()
    # 2025 finance debt from official table if missing
    if 2025 not in borrowings:
        borrowings[2025] = 26441.0  # FY2025 earnings: Dívida Financeira

    rows = []
    for y in range(2002, 2026):
        meta = by_year[str(y)]
        if y in OFFICIAL_GROSS_DEBT_USD_MI:
            gross_mi = OFFICIAL_GROSS_DEBT_USD_MI[y]
            gross_src = "Petrobras official / 20-F endividamento"
        else:
            gross_mi = WSN_GROSS_DEBT_USD_BI[y] * 1000.0
            gross_src = "Historical total debt series (Petrobras gross-debt concept)"

        ip = interest.get(y, {})
        rows.append(
            {
                "year": y,
                "gross_debt_usd_mi": round(gross_mi, 1),
                "gross_debt_usd_bi": round(gross_mi / 1000.0, 2),
                "gross_debt_source": gross_src,
                "finance_debt_usd_mi": round(borrowings[y], 1) if y in borrowings else None,
                "interest_paid_usd_mi": ip.get("interest_paid_usd_mi"),
                "interest_paid_source": ip.get("interest_paid_source"),
                "filing_url": (
                    f"https://www.sec.gov/Archives/edgar/data/1119639/"
                    f"{meta[2].replace('-', '')}/{meta[3]}"
                ),
            }
        )
    return rows


def write_outputs(rows: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "company": "Petrobras (Petróleo Brasileiro S.A.)",
        "cik": "0001119639",
        "period": "2002-2025",
        "units": "US$ millions / billions",
        "notes": [
            "Dívida bruta = conceito Petrobras (dívida financeira + arrendamentos a partir de IFRS 16/2019).",
            "Juros pagos = caixa de juros da dívida financeira no 20-F (não inclui accrual puro nem juros de arrendamento, salvo quando a linha histórica não segregava).",
            "2002–2010: supplemental 'Cash paid for Interest, net of amount capitalized' (US$ mi).",
            "2011–2018: 'Repayment of interest' no DFC.",
            "2019–2025: 'Repayment of interest - finance debt' / InterestPaidClassifiedAsFinancingActivities.",
        ],
        "rows": rows,
    }
    (OUT_DIR / "petrobras_divida_juros_2002_2025.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin", color="9AA7B5"),
        right=Side(style="thin", color="9AA7B5"),
        top=Side(style="thin", color="9AA7B5"),
        bottom=Side(style="thin", color="9AA7B5"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    zebra = PatternFill("solid", fgColor="F7F9FC")

    def style_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Debt
    ws = wb.active
    ws.title = "Divida_Bruta"
    ws["A1"] = "Petrobras — Evolução da Dívida Bruta (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Conceito Petrobras de dívida bruta (dívida financeira + arrendamentos a partir de 2019/IFRS 16). US$."
    )
    headers = [
        "Ano",
        "Dívida bruta (US$ mi)",
        "Dívida bruta (US$ bi)",
        "Δ YoY %",
        "Dívida financeira (US$ mi)",
        "Fonte",
    ]
    style_header(ws, 4, headers)
    prev = None
    for i, r in enumerate(rows):
        yoy = None
        if prev:
            yoy = (r["gross_debt_usd_bi"] / prev - 1) * 100
        vals = [
            r["year"],
            r["gross_debt_usd_mi"],
            r["gross_debt_usd_bi"],
            round(yoy, 1) if yoy is not None else None,
            r["finance_debt_usd_mi"],
            r["gross_debt_source"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(5 + i, c, v)
            cell.border = thin
            if i % 2:
                cell.fill = zebra
            if c in (2, 5) and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"
            if c == 3 and isinstance(v, (int, float)):
                cell.number_format = "0.00"
            if c == 4 and isinstance(v, (int, float)):
                cell.number_format = "0.0"
        prev = r["gross_debt_usd_bi"]
    for col, w in enumerate([8, 22, 20, 12, 26, 55], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Interest
    ws = wb.create_sheet("Juros_Pagos")
    ws["A1"] = "Petrobras — Evolução dos Juros Pagos (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Juros pagos em caixa sobre dívida financeira (US$ milhões), extraídos dos 20-F."
    headers = ["Ano", "Juros pagos (US$ mi)", "Juros pagos (US$ bi)", "Δ YoY %", "Fonte"]
    style_header(ws, 4, headers)
    prev = None
    for i, r in enumerate(rows):
        ip = r["interest_paid_usd_mi"]
        yoy = None
        if prev and ip is not None:
            yoy = (ip / prev - 1) * 100
        vals = [
            r["year"],
            ip,
            round(ip / 1000.0, 3) if ip is not None else None,
            round(yoy, 1) if yoy is not None else None,
            r["interest_paid_source"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(5 + i, c, v)
            cell.border = thin
            if i % 2:
                cell.fill = zebra
            if c == 2 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"
            if c in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = "0.000" if c == 3 else "0.0"
        if ip is not None:
            prev = ip
    for col, w in enumerate([8, 22, 20, 12, 70], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Combined pretty table
    ws = wb.create_sheet("Consolidado")
    ws["A1"] = "Petrobras — Dívida bruta e juros pagos (2002–2025)"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Ano", "Dívida bruta (US$ bi)", "Juros pagos (US$ mi)"]
    style_header(ws, 3, headers)
    for i, r in enumerate(rows):
        for c, v in enumerate(
            [r["year"], r["gross_debt_usd_bi"], r["interest_paid_usd_mi"]], 1
        ):
            cell = ws.cell(4 + i, c, v)
            cell.border = thin
            if i % 2:
                cell.fill = zebra
            if c == 2 and isinstance(v, (int, float)):
                cell.number_format = "0.00"
            if c == 3 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.0"

    ws = wb.create_sheet("Notas")
    notes = [
        "1. Dívida bruta: conceito de endividamento da Petrobras (financeira + arrendamentos pós-IFRS 16/2019).",
        "2. Série de dívida bruta em US$; overrides oficiais em 2019 (87.121), 2024 (60.311) e 2025 (69.793).",
        "3. Dívida financeira: tag IFRS Borrowings (SEC), quando disponível; 2025 da tabela de endividamento do release.",
        "4. Juros pagos: caixa de juros da dívida financeira nos Form 20-F (SEC).",
        "5. 2002–2010 usam disclosure suplementar 'Interest, net of amount capitalized' / 'Cash paid for Interest'.",
        "6. 2011–2018: linha DFC 'Repayment of interest'.",
        "7. 2019–2025: linha 'Repayment of interest/finance debt' (exclui amortização de juros de arrendamento).",
    ]
    for i, n in enumerate(notes, 1):
        ws[f"A{i}"] = n
    ws.column_dimensions["A"].width = 130

    xlsx = OUT_DIR / "petrobras_divida_juros_2002_2025.xlsx"
    wb.save(xlsx)
    print("Wrote", xlsx)

    # Markdown tables for quick view
    md = ["# Petrobras — Dívida bruta e juros pagos (2002–2025)", "", "## Dívida bruta (US$ bi)", "", "| Ano | Dívida bruta |", "|---:|---:|"]
    for r in rows:
        md.append(f"| {r['year']} | {r['gross_debt_usd_bi']:.2f} |")
    md += ["", "## Juros pagos (US$ mi)", "", "| Ano | Juros pagos |", "|---:|---:|"]
    for r in rows:
        ip = r["interest_paid_usd_mi"]
        md.append(f"| {r['year']} | {ip if ip is not None else 'n/d'} |")
    (OUT_DIR / "petrobras_divida_juros_2002_2025.md").write_text("\n".join(md) + "\n", encoding="utf-8")


def main():
    by_year = json.loads(INDEX.read_text())
    rows = build_rows(by_year)
    write_outputs(rows)
    print("\nFinal series:")
    for r in rows:
        print(
            f"{r['year']}: debt={r['gross_debt_usd_bi']:.2f} bi | interest={r['interest_paid_usd_mi']}"
        )
    missing = [r["year"] for r in rows if r["interest_paid_usd_mi"] is None]
    print("Missing interest:", missing)


if __name__ == "__main__":
    main()
