"""Alíquotas oficiais de recolhimento compulsório (2002–2026).

Fonte primária até jul/2019:
  https://www.bcb.gov.br/ftp/notaecon/compulsorios.xls
  (série histórica Deban / Nota para Imprensa)

Complemento 2020–2026: Circulares e Resoluções BCB + quadro resumo vigente
  https://www.bcb.gov.br/conteudo/dadosabertos/BCBDeban/Resumo%20das%20normas%20dos%20compulsórios.csv

Uso:
  python3 scripts/aliquotas_compulsorio.py
  python3 scripts/aliquotas_compulsorio.py --sem-download
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.taxas_credito_bcb import _escrever_aba  # noqa: E402

DATA_DIR = ROOT / "data" / "compulsorio"
OUTPUT_DIR = ROOT / "output"
XLS_URL = "https://www.bcb.gov.br/ftp/notaecon/compulsorios.xls"
QUADRO_URL = (
    "https://www.bcb.gov.br/conteudo/dadosabertos/BCBDeban/"
    "Resumo%20das%20normas%20dos%20compuls%C3%B3rios.csv"
)
ANO_INI = 2002
ANO_FIM = 2026

MESES = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}

COLUNAS = [
    "a_vista",
    "a_prazo",
    "poupanca_habitacional",
    "poupanca_rural",
    "adic_a_vista",
    "adic_a_prazo",
    "adic_poupanca",
]

# Alterações posteriores ao XLS oficial (última linha: jul/2019).
POS_XLS = [
    {
        "data": "2020-03-16",
        "a_prazo": 25.0,
        "norma": "Circular BCB 3.987, de 20/02/2020",
        "nota": "Redução de 31% para 25% (efeito em 16/03/2020).",
    },
    {
        "data": "2020-03-30",
        "a_prazo": 17.0,
        "norma": "Circular BCB 3.993, de 23/03/2020",
        "nota": "Redução provisória de 25% para 17% (Covid-19).",
    },
    {
        "data": "2021-11-29",
        "a_prazo": 20.0,
        "norma": "Resolução BCB 78/2021 e Resolução BCB 145/2021",
        "nota": "Alíquota a prazo fixada em 20% (período de cálculo a partir de 29/11/2021).",
    },
]


def eh_aspas(valor: object) -> bool:
    txt = str(valor).strip()
    return txt in {'"', "“", "”", "''"}


def eh_traco(valor: object) -> bool:
    txt = str(valor).strip()
    return txt in {"-", "–", "—"}


def para_pct(valor: object) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if eh_aspas(valor) or eh_traco(valor):
        return None
    try:
        n = float(valor)
    except (TypeError, ValueError):
        return None
    if n <= 1.5:
        n *= 100.0
    return float(n)


def preencher(anterior: float | None, valor: object) -> float | None:
    if eh_aspas(valor):
        return anterior
    if eh_traco(valor):
        return None
    lido = para_pct(valor)
    return anterior if lido is None else lido


def ler_xls_oficial(path: Path) -> pd.DataFrame:
    bruto = pd.read_excel(path, sheet_name="Português", header=None)
    estado = {c: None for c in COLUNAS}
    ano = None
    linhas = []
    for rec in bruto.itertuples(index=False):
        c0, c1, c2, _c3, c4, _c5, c6, c7, _c8, _c9, _c10, _c11, c12, c13, _c14, c15 = rec[:16]
        if isinstance(c0, (int, float)) and not pd.isna(c0) and 1900 < float(c0) < 2100:
            ano = int(c0)
        mes = MESES.get(str(c1).strip().lower()[:3]) if not pd.isna(c1) else None
        if ano is None or mes is None:
            continue
        estado["a_vista"] = preencher(estado["a_vista"], c2)
        estado["a_prazo"] = preencher(estado["a_prazo"], c4)
        estado["poupanca_habitacional"] = preencher(estado["poupanca_habitacional"], c6)
        estado["poupanca_rural"] = preencher(estado["poupanca_rural"], c7)
        estado["adic_a_vista"] = preencher(estado["adic_a_vista"], c12)
        estado["adic_a_prazo"] = preencher(estado["adic_a_prazo"], c13)
        estado["adic_poupanca"] = preencher(estado["adic_poupanca"], c15)
        linhas.append(
            {
                "data": pd.Timestamp(year=ano, month=mes, day=1),
                **estado,
                "norma": "Série histórica BCB (compulsorios.xls)",
                "nota": "",
            }
        )
    return pd.DataFrame(linhas)


def aplicar_pos_xls(base: pd.DataFrame) -> pd.DataFrame:
    if base.empty:
        atual = {c: None for c in COLUNAS}
    else:
        atual = {c: base.iloc[-1][c] for c in COLUNAS}
    extra = []
    for item in POS_XLS:
        atual = {**atual, **{k: v for k, v in item.items() if k in COLUNAS}}
        extra.append(
            {
                "data": pd.Timestamp(item["data"]),
                **atual,
                "norma": item["norma"],
                "nota": item["nota"],
            }
        )
    out = pd.concat([base, pd.DataFrame(extra)], ignore_index=True)
    return out.sort_values("data").drop_duplicates("data", keep="last").reset_index(drop=True)


def _chave(valor: object) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    return float(valor)


def snapshot_anual(alteracoes: pd.DataFrame, ano_ini: int, ano_fim: int) -> pd.DataFrame:
    linhas = []
    for ano in range(ano_ini, ano_fim + 1):
        corte = pd.Timestamp(year=ano, month=12, day=31)
        ate = alteracoes[alteracoes["data"] <= corte]
        if ate.empty:
            rec = {c: None for c in COLUNAS}
            rec.update({"ano": ano, "desde": None, "norma": ""})
        else:
            ult = ate.iloc[-1]
            rec = {c: ult[c] for c in COLUNAS}
            alvo = tuple(_chave(ult[c]) for c in COLUNAS)
            desde = ult
            for i in range(len(ate) - 1, -1, -1):
                row = ate.iloc[i]
                if tuple(_chave(row[c]) for c in COLUNAS) != alvo:
                    break
                desde = row
            rec.update(
                {
                    "ano": ano,
                    "desde": desde["data"].strftime("%d/%m/%Y"),
                    "norma": ult["norma"],
                }
            )
        linhas.append(rec)
    return pd.DataFrame(linhas)


def baixar_xls(cache_dir: Path, baixar: bool) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    dest = cache_dir / "compulsorios.xls"
    if dest.exists() and dest.stat().st_size > 0:
        return dest
    if not baixar:
        raise FileNotFoundError(dest)
    resp = requests.get(XLS_URL, timeout=90)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return dest


def baixar_quadro(cache_dir: Path, baixar: bool) -> pd.DataFrame:
    dest = cache_dir / "quadro_resumo.csv"
    if dest.exists():
        return pd.read_csv(dest, sep=";")
    if not baixar:
        raise FileNotFoundError(dest)
    resp = requests.get(QUADRO_URL, timeout=90)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    return pd.read_csv(dest, sep=";", encoding="latin-1")


def gerar_planilha(
    anual: pd.DataFrame,
    alteracoes: pd.DataFrame,
    quadro: pd.DataFrame,
    path: Path,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Notas"
    ws0["A1"] = "Alíquotas de recolhimento compulsório no Brasil — 2002 a 2026"
    ws0["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    notas = [
        "Percentuais nominais estabelecidos pelo Banco Central (alíquota da exigibilidade "
        "sobre a base de cálculo). Não é a alíquota efetiva (que desconta deduções de PR, "
        "caixa até 2018 e direcionamentos).",
        "Até julho de 2019 a fonte é a série histórica oficial "
        "https://www.bcb.gov.br/ftp/notaecon/compulsorios.xls (Deban / Nota para Imprensa). "
        "Aspas na planilha original significam manutenção da alíquota anterior.",
        "De 2020 em diante o XLS não foi republicado. As mudanças vêm das Circulares "
        "3.987/2020 e 3.993/2020 e das Resoluções BCB 78/2021 e 145/2021 (recursos a prazo). "
        "À vista permanece 21% desde nov/2018 (Circular 3.917/2018; Resolução BCB 189/2022). "
        "Poupança permanece 20% desde abr/2018 (Circular 3.890/2018; Resolução BCB 188/2022).",
        "Aba Anual: alíquota vigente em 31 de dezembro de cada ano (2026 = regra em vigor "
        "na data da extração). Aba Alteracoes: cada mudança. Aba Vigente: quadro resumo "
        "atual do Deban (à vista 21%, a prazo 20%, poupança 20%).",
        "Há ainda direcionamento obrigatório (poupança SBPE 65%, microfinanças 2%, crédito "
        "rural sobre à vista), que não é recolhimento compulsório — aparece só no quadro vigente.",
        "A Resolução BCB 551/2026 criou dedução da exigibilidade (antecipação de contribuições "
        "ao FGC). Isso reduz o valor a recolher, mas não altera as alíquotas nominais de 21% "
        "(à vista) e 20% (a prazo e poupança).",
    ]
    for i, txt in enumerate(notas, start=3):
        ws0[f"A{i}"] = txt
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
        ws0[f"A{i}"].alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[i].height = 48
    ws0.column_dimensions["A"].width = 28

    cabs_a = [
        "Ano",
        "À vista %",
        "A prazo %",
        "Poupança habitacional/livre %",
        "Poupança rural %",
        "Adicional à vista %",
        "Adicional a prazo %",
        "Adicional poupança %",
        "Vigente desde",
        "Fonte da última mudança",
    ]
    linhas_a = []
    for rec in anual.to_dict("records"):
        linhas_a.append(
            [
                int(rec["ano"]),
                rec["a_vista"],
                rec["a_prazo"],
                rec["poupanca_habitacional"],
                rec["poupanca_rural"],
                rec["adic_a_vista"],
                rec["adic_a_prazo"],
                rec["adic_poupanca"],
                rec["desde"],
                rec["norma"],
            ]
        )
    _escrever_aba(wb.create_sheet("Anual"), cabs_a, linhas_a, [8, 12, 12, 26, 16, 16, 16, 18, 14, 46])

    alt = alteracoes[(alteracoes["data"].dt.year >= ANO_INI) & (alteracoes["data"].dt.year <= ANO_FIM)].copy()
    cabs_c = [
        "Data",
        "À vista %",
        "A prazo %",
        "Poupança hab./livre %",
        "Poupança rural %",
        "Adicional à vista %",
        "Adicional a prazo %",
        "Adicional poupança %",
        "Norma",
        "Nota",
    ]
    linhas_c = []
    for rec in alt.to_dict("records"):
        linhas_c.append(
            [
                rec["data"].strftime("%d/%m/%Y"),
                rec["a_vista"],
                rec["a_prazo"],
                rec["poupanca_habitacional"],
                rec["poupanca_rural"],
                rec["adic_a_vista"],
                rec["adic_a_prazo"],
                rec["adic_poupanca"],
                rec["norma"],
                rec["nota"],
            ]
        )
    _escrever_aba(wb.create_sheet("Alteracoes"), cabs_c, linhas_c, [12, 12, 12, 18, 16, 16, 16, 18, 46, 52])

    if not quadro.empty:
        cabs_v = [str(c) for c in quadro.columns]
        linhas_v = []
        for rec in quadro.fillna("").to_dict("records"):
            linhas_v.append([rec.get(c, "") for c in quadro.columns])
        _escrever_aba(
            wb.create_sheet("Vigente"),
            cabs_v,
            linhas_v,
            [28] + [36] * (len(cabs_v) - 1),
        )
    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR / "aliquotas_compulsorio_2002_2026.xlsx")
    parser.add_argument("--sem-download", action="store_true")
    args = parser.parse_args(argv)

    xls = baixar_xls(args.cache_dir, baixar=not args.sem_download)
    hist = ler_xls_oficial(xls)
    alt = aplicar_pos_xls(hist)
    anual = snapshot_anual(alt, ANO_INI, ANO_FIM)
    try:
        quadro = baixar_quadro(args.cache_dir, baixar=not args.sem_download)
    except Exception as exc:  # noqa: BLE001
        print(f"Quadro vigente indisponível: {exc}", flush=True)
        quadro = pd.DataFrame()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    anual.to_csv(args.output.parent / "aliquotas_compulsorio_anual_2002_2026.csv", index=False)
    alt.assign(data=alt["data"].dt.strftime("%Y-%m-%d")).to_csv(
        args.output.parent / "aliquotas_compulsorio_alteracoes.csv", index=False
    )
    path = gerar_planilha(anual, alt, quadro, args.output)
    print(f"Planilha: {path}")
    cols = ["ano", "a_vista", "a_prazo", "poupanca_habitacional", "poupanca_rural"]
    print(anual[cols].to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
