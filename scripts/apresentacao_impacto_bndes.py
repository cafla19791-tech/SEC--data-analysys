#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolida Excel de apresentação a partir dos resumos do agregador ContAgil.

Lê (na pasta ``saida/`` ou ``--pasta``):

  - impacto_fiscal_por_ano.csv|.xlsx
  - resumo_por_agente.csv|.xlsx
  - (opcional) fluxos_por_ano_contrato/RESUMO.csv
  - (opcional) resumo_impacto_bndes.xlsx aba Totais

Gera::

  saida/APRESENTACAO_IMPACTO_BNDES_INDIRETAS.xlsx

Abas: Capa | Sumario | Por_Ano | Top_20_Agentes | Por_Agente | Notas

Uso (ContAgil)::

  python sec_scripts\\apresentacao_impacto_bndes.py --pasta saida
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MARKER = "apresentacao-impacto-bndes-20260816a"

# Paleta sóbria (compatível com demais workbooks do repo)
AZUL = "1F4E79"
AZUL_CLARO = "D6E3F0"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
PRETO = "1A1A1A"

FILL_H = PatternFill("solid", fgColor=AZUL)
FILL_ALT = PatternFill("solid", fgColor=CINZA)
FILL_KPI = PatternFill("solid", fgColor=AZUL_CLARO)
FONT_H = Font(color=BRANCO, bold=True, name="Calibri", size=11)
FONT_TITULO = Font(color=AZUL, bold=True, name="Calibri", size=18)
FONT_SUB = Font(color=PRETO, name="Calibri", size=11)
FONT_KPI_LBL = Font(color=AZUL, bold=True, name="Calibri", size=10)
FONT_KPI_VAL = Font(color=PRETO, bold=True, name="Calibri", size=14)
FONT_CELL = Font(name="Calibri", size=10, color=PRETO)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

FMT_RS = 'R$#,##0.00'
FMT_RS_MI = 'R$#,##0.00" mi"'
FMT_INT = "#,##0"
FMT_PCT = "0.0%"

COL_AGENTE_ALIASES = ("Instituição Financeira", "Instituicao Financeira", "Agente", "agente")
COL_IMPACTO = "Impacto Fiscal 2026 (R$)"
COL_SUBSIDIO_ANO = "Soma Subsídio Nominal (R$)"
COL_SUBSIDIO_AG = "Total Subsídio (R$)"
COL_PARCELAS = "Quantidade de Parcelas"
COL_QTD_PARCELAS_AG = "Qtd Parcelas"
COL_CONTRATOS = "Qtd Contratos"
COL_ANO = "Ano"


def _ler_tabela(pasta: Path, stem: str) -> pd.DataFrame:
    """Lê CSV (preferido) ou XLSX com o mesmo stem."""
    csv_p = pasta / f"{stem}.csv"
    xlsx_p = pasta / f"{stem}.xlsx"
    if csv_p.is_file():
        return pd.read_csv(csv_p)
    if xlsx_p.is_file():
        return pd.read_excel(xlsx_p)
    raise FileNotFoundError(
        f"Não encontrado {stem}.csv nem {stem}.xlsx em {pasta}. "
        "Rode agregar_impacto_saida.ps1 primeiro."
    )


def _col_agente(df: pd.DataFrame) -> str:
    for c in COL_AGENTE_ALIASES:
        if c in df.columns:
            return c
    raise ValueError(
        f"Coluna de agente não encontrada. Colunas: {list(df.columns)}"
    )


def _norm_por_ano(df: pd.DataFrame) -> pd.DataFrame:
    need = {COL_ANO, COL_SUBSIDIO_ANO, COL_IMPACTO}
    missing = need - set(df.columns)
    if missing:
        raise ValueError(f"impacto_fiscal_por_ano sem colunas: {sorted(missing)}")
    out = df.copy()
    out[COL_ANO] = pd.to_numeric(out[COL_ANO], errors="coerce").astype("Int64")
    out[COL_SUBSIDIO_ANO] = pd.to_numeric(out[COL_SUBSIDIO_ANO], errors="coerce").fillna(0.0)
    out[COL_IMPACTO] = pd.to_numeric(out[COL_IMPACTO], errors="coerce").fillna(0.0)
    if COL_PARCELAS in out.columns:
        out[COL_PARCELAS] = pd.to_numeric(out[COL_PARCELAS], errors="coerce").fillna(0).astype(int)
    out = out.dropna(subset=[COL_ANO]).sort_values(COL_ANO).reset_index(drop=True)
    tot_imp = float(out[COL_IMPACTO].sum()) or 1.0
    out["Participação Impacto"] = out[COL_IMPACTO] / tot_imp
    out["Impacto (R$ bi)"] = out[COL_IMPACTO] / 1e9
    out["Subsídio (R$ bi)"] = out[COL_SUBSIDIO_ANO] / 1e9
    return out


def _norm_por_agente(df: pd.DataFrame) -> pd.DataFrame:
    ag = _col_agente(df)
    if COL_IMPACTO not in df.columns:
        raise ValueError(f"resumo_por_agente sem {COL_IMPACTO}")
    out = df.copy()
    out = out.rename(columns={ag: "Instituição Financeira"})
    out["Instituição Financeira"] = (
        out["Instituição Financeira"].fillna("NÃO INFORMADO").astype(str).str.strip()
    )
    out[COL_IMPACTO] = pd.to_numeric(out[COL_IMPACTO], errors="coerce").fillna(0.0)
    sub_col = COL_SUBSIDIO_AG if COL_SUBSIDIO_AG in out.columns else None
    if sub_col:
        out[sub_col] = pd.to_numeric(out[sub_col], errors="coerce").fillna(0.0)
    for c in (COL_CONTRATOS, COL_QTD_PARCELAS_AG):
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    out = out.sort_values(COL_IMPACTO, ascending=False).reset_index(drop=True)
    out.insert(0, "Ranking", range(1, len(out) + 1))
    tot_imp = float(out[COL_IMPACTO].sum()) or 1.0
    out["Participação Impacto"] = out[COL_IMPACTO] / tot_imp
    out["Impacto acumulado %"] = out["Participação Impacto"].cumsum()
    out["Impacto (R$ bi)"] = out[COL_IMPACTO] / 1e9
    if sub_col:
        out["Subsídio (R$ bi)"] = out[sub_col] / 1e9
    return out


def _ler_resumo_contratos(pasta: Path) -> pd.DataFrame | None:
    p = pasta / "fluxos_por_ano_contrato" / "RESUMO.csv"
    if not p.is_file():
        return None
    df = pd.read_csv(p)
    return df


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = FILL_H
        cell.font = FONT_H
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _autosize(ws, min_w: int = 10, max_w: int = 42) -> None:
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for row in range(1, min((ws.max_row or 1) + 1, 80)):
            val = ws.cell(row, col).value
            if val is None:
                continue
            maxlen = max(maxlen, len(str(val)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, maxlen + 2))


def _escrever_df(
    ws,
    df: pd.DataFrame,
    *,
    money_cols: set[str] | None = None,
    pct_cols: set[str] | None = None,
    int_cols: set[str] | None = None,
    bi_cols: set[str] | None = None,
) -> None:
    money_cols = money_cols or set()
    pct_cols = pct_cols or set()
    int_cols = int_cols or set()
    bi_cols = bi_cols or set()
    cols = list(df.columns)
    ws.append(cols)
    _style_header(ws, len(cols))
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, val if not (isinstance(val, float) and pd.isna(val)) else None)
            cell.font = FONT_CELL
            cell.border = THIN
            if i % 2 == 0:
                cell.fill = FILL_ALT
            name = cols[j - 1]
            if name in money_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = FMT_RS
            elif name in bi_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = "0.000"
            elif name in pct_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = FMT_PCT
            elif name in int_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.number_format = FMT_INT
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32
    _autosize(ws)


def _aba_capa(
    wb: Workbook,
    *,
    tot_impacto: float,
    tot_subsidio: float,
    n_anos: int,
    n_agentes: int,
    n_parcelas: int,
    top_agente: str,
    top_impacto: float,
    gerado_em: str,
) -> None:
    ws = wb.active
    ws.title = "Capa"
    ws["A1"] = "Impacto Fiscal — BNDES Indiretas"
    ws["A1"].font = FONT_TITULO
    ws["A2"] = "Fluxos por ano de contrato (NUMERADOS) · capitalização à data de referência"
    ws["A2"].font = FONT_SUB
    ws["A3"] = f"Gerado em {gerado_em}  |  {MARKER}"
    ws["A3"].font = Font(name="Calibri", size=9, color="666666", italic=True)

    kpis = [
        ("Total impacto fiscal 2026", tot_impacto, FMT_RS),
        ("Total subsídio nominal", tot_subsidio, FMT_RS),
        ("Impacto (R$ bilhões)", tot_impacto / 1e9, "0.00"),
        ("Anos com fluxo", n_anos, FMT_INT),
        ("Agentes financeiros", n_agentes, FMT_INT),
        ("Parcelas agregadas", n_parcelas, FMT_INT),
        ("Maior agente (impacto)", top_agente, None),
        ("Impacto do líder", top_impacto, FMT_RS),
    ]
    ws["A5"] = "Indicadores"
    ws["A5"].font = Font(color=AZUL, bold=True, size=13, name="Calibri")
    row = 6
    for label, val, fmt in kpis:
        ws.cell(row, 1, label).font = FONT_KPI_LBL
        ws.cell(row, 1).fill = FILL_KPI
        ws.cell(row, 1).border = THIN
        c = ws.cell(row, 2, val)
        c.font = FONT_KPI_VAL
        c.fill = FILL_KPI
        c.border = THIN
        if fmt and isinstance(val, (int, float)):
            c.number_format = fmt
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 48
    ws["A15"] = (
        "Fonte: saida/fluxos_por_ano_contrato/YYYY.csv → agregar_impacto_fluxos.py (modo coluna). "
        "Abrir abas Sumario, Por_Ano e Por_Agente para detalhe."
    )
    ws["A15"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A15:B17")


def _aba_sumario(
    wb: Workbook,
    por_ano: pd.DataFrame,
    por_agente: pd.DataFrame,
) -> None:
    ws = wb.create_sheet("Sumario")
    ws["A1"] = "Sumário executivo"
    ws["A1"].font = FONT_TITULO

    tot_imp = float(por_ano[COL_IMPACTO].sum())
    tot_sub = float(por_ano[COL_SUBSIDIO_ANO].sum())
    ws["A3"] = "Totais"
    ws["A3"].font = Font(color=AZUL, bold=True, size=12)
    ws["A4"] = "Subsídio nominal (R$)"
    ws["B4"] = tot_sub
    ws["B4"].number_format = FMT_RS
    ws["A5"] = "Impacto fiscal 2026 (R$)"
    ws["B5"] = tot_imp
    ws["B5"].number_format = FMT_RS
    ws["A6"] = "Impacto (R$ bi)"
    ws["B6"] = tot_imp / 1e9
    ws["B6"].number_format = "0.00"
    for r in range(4, 7):
        ws.cell(r, 1).font = FONT_KPI_LBL
        ws.cell(r, 2).font = FONT_KPI_VAL

    ws["A8"] = "Top 10 instituições (impacto fiscal 2026)"
    ws["A8"].font = Font(color=AZUL, bold=True, size=12)

    top = por_agente.head(10).copy()
    cols_show = [
        "Ranking",
        "Instituição Financeira",
        COL_IMPACTO,
        "Impacto (R$ bi)",
        "Participação Impacto",
        "Impacto acumulado %",
    ]
    if COL_CONTRATOS in top.columns:
        cols_show.insert(2, COL_CONTRATOS)
    if COL_SUBSIDIO_AG in top.columns:
        cols_show.insert(-3, COL_SUBSIDIO_AG)

    # escrever tabela a partir da linha 9
    start = 9
    subset = top[[c for c in cols_show if c in top.columns]]
    for j, name in enumerate(subset.columns, start=1):
        cell = ws.cell(start, j, name)
        cell.fill = FILL_H
        cell.font = FONT_H
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    for i, row in enumerate(subset.itertuples(index=False), start=start + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, None if (isinstance(val, float) and pd.isna(val)) else val)
            cell.font = FONT_CELL
            cell.border = THIN
            if i % 2 == 0:
                cell.fill = FILL_ALT
            name = subset.columns[j - 1]
            if name in {COL_IMPACTO, COL_SUBSIDIO_AG} and isinstance(val, (int, float)):
                cell.number_format = FMT_RS
            elif name in {"Participação Impacto", "Impacto acumulado %"} and isinstance(
                val, (int, float)
            ):
                cell.number_format = FMT_PCT
            elif name == "Impacto (R$ bi)" and isinstance(val, (int, float)):
                cell.number_format = "0.000"
            elif name in {COL_CONTRATOS, "Ranking"} and isinstance(val, (int, float)):
                cell.number_format = FMT_INT
    ws.freeze_panes = "A10"
    _autosize(ws)
    ws.column_dimensions["B"].width = 48


def _aba_notas(wb: Workbook) -> None:
    ws = wb.create_sheet("Notas")
    ws["A1"] = "Notas metodológicas"
    ws["A1"].font = FONT_TITULO
    linhas = [
        "",
        "1. Universo: operações BNDES Indiretas numeradas (N-AAAA) em BNDES_INDIRETAS_NUMERADOS.xlsx.",
        "2. Fluxos: gerados por ano de contrato (saida/fluxos_por_ano_contrato/YYYY.csv).",
        "3. Subsídio: diferença entre taxa SELIC e taxa do contrato em cada parcela.",
        "4. Impacto fiscal 2026: valor do subsídio capitalizado até a data de referência (30/06/2026),",
        "   usando o fator acumulado SELIC/TJLP/TLP da geração ContAgil (coluna impacto_fiscal do CSV).",
        "5. Agregação: streaming (agregar_impacto_fluxos.py, modo coluna) — soma por ano do fluxo e por agente.",
        "6. Excel consolidado de fluxos (FLUXOS_...POR_ANO_CONTRATO.xlsx) é amostra; a fonte da verdade é o CSV.",
        "7. Este workbook é de apresentação; valores em R$ e R$ bilhões (bi = 10^9).",
        "",
        f"Marcador: {MARKER}",
    ]
    for i, txt in enumerate(linhas, start=2):
        ws.cell(i, 1, txt).font = FONT_SUB
    ws.column_dimensions["A"].width = 110


def construir_apresentacao(
    pasta: Path,
    saida: Path | None = None,
) -> Path:
    """Monta o workbook de apresentação e grava em disco."""
    pasta = Path(pasta)
    por_ano = _norm_por_ano(_ler_tabela(pasta, "impacto_fiscal_por_ano"))
    por_agente = _norm_por_agente(_ler_tabela(pasta, "resumo_por_agente"))
    resumo_ct = _ler_resumo_contratos(pasta)

    tot_impacto = float(por_ano[COL_IMPACTO].sum())
    tot_subsidio = float(por_ano[COL_SUBSIDIO_ANO].sum())
    n_parcelas = (
        int(por_ano[COL_PARCELAS].sum()) if COL_PARCELAS in por_ano.columns else 0
    )
    top_row = por_agente.iloc[0]
    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()
    _aba_capa(
        wb,
        tot_impacto=tot_impacto,
        tot_subsidio=tot_subsidio,
        n_anos=len(por_ano),
        n_agentes=len(por_agente),
        n_parcelas=n_parcelas,
        top_agente=str(top_row["Instituição Financeira"]),
        top_impacto=float(top_row[COL_IMPACTO]),
        gerado_em=gerado_em,
    )
    _aba_sumario(wb, por_ano, por_agente)

    ws_ano = wb.create_sheet("Por_Ano")
    cols_ano = [
        COL_ANO,
        COL_SUBSIDIO_ANO,
        "Subsídio (R$ bi)",
        COL_IMPACTO,
        "Impacto (R$ bi)",
        "Participação Impacto",
    ]
    if COL_PARCELAS in por_ano.columns:
        cols_ano.append(COL_PARCELAS)
    _escrever_df(
        ws_ano,
        por_ano[cols_ano],
        money_cols={COL_SUBSIDIO_ANO, COL_IMPACTO},
        pct_cols={"Participação Impacto"},
        int_cols={COL_ANO, COL_PARCELAS},
        bi_cols={"Subsídio (R$ bi)", "Impacto (R$ bi)"},
    )

    ws_top = wb.create_sheet("Top_20_Agentes")
    top20 = por_agente.head(20)
    cols_ag = [
        "Ranking",
        "Instituição Financeira",
        COL_IMPACTO,
        "Impacto (R$ bi)",
        "Participação Impacto",
        "Impacto acumulado %",
    ]
    money_ag = {COL_IMPACTO}
    if COL_SUBSIDIO_AG in top20.columns:
        cols_ag.insert(2, COL_SUBSIDIO_AG)
        cols_ag.insert(3, "Subsídio (R$ bi)")
        money_ag.add(COL_SUBSIDIO_AG)
    if COL_CONTRATOS in top20.columns:
        cols_ag.insert(2, COL_CONTRATOS)
    if COL_QTD_PARCELAS_AG in top20.columns:
        cols_ag.insert(3 if COL_CONTRATOS in top20.columns else 2, COL_QTD_PARCELAS_AG)
    _escrever_df(
        ws_top,
        top20[[c for c in cols_ag if c in top20.columns]],
        money_cols=money_ag,
        pct_cols={"Participação Impacto", "Impacto acumulado %"},
        int_cols={"Ranking", COL_CONTRATOS, COL_QTD_PARCELAS_AG},
        bi_cols={"Impacto (R$ bi)", "Subsídio (R$ bi)"},
    )
    ws_top.column_dimensions["B"].width = 48

    ws_ag = wb.create_sheet("Por_Agente")
    _escrever_df(
        ws_ag,
        por_agente[[c for c in cols_ag if c in por_agente.columns]],
        money_cols=money_ag,
        pct_cols={"Participação Impacto", "Impacto acumulado %"},
        int_cols={"Ranking", COL_CONTRATOS, COL_QTD_PARCELAS_AG},
        bi_cols={"Impacto (R$ bi)", "Subsídio (R$ bi)"},
    )
    ws_ag.column_dimensions["B"].width = 48

    if resumo_ct is not None and not resumo_ct.empty:
        ws_r = wb.create_sheet("Resumo_Geracao")
        _escrever_df(ws_r, resumo_ct)

    _aba_notas(wb)

    out = Path(saida) if saida is not None else pasta / "APRESENTACAO_IMPACTO_BNDES_INDIRETAS.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _resolver_pasta(explicit: Path | None) -> Path:
    if explicit is not None:
        return Path(explicit)
    contagil = Path(
        r"C:\Arquivos de Programas RFB\ContAgilAppBeta64\python_jep\winpython\saida"
    )
    if contagil.exists():
        return contagil
    cwd_saida = Path.cwd() / "saida"
    if cwd_saida.exists():
        return cwd_saida
    return Path.cwd() / "output"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pasta", type=Path, default=None, help="Pasta com os CSVs do agregador")
    p.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Caminho do xlsx (default: pasta/APRESENTACAO_IMPACTO_BNDES_INDIRETAS.xlsx)",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    pasta = _resolver_pasta(args.pasta)
    print(f"[{MARKER}]")
    print(f"Pasta : {pasta}")
    try:
        out = construir_apresentacao(pasta, args.saida)
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 1
    print(f"OK → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
