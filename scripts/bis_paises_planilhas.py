"""Planilhas do BIS (política monetária e crédito): uma aba por país.

Fonte: https://data.bis.org/bulkdownload  (CSV flat)

Uso:
  python3 scripts/bis_paises_planilhas.py
  python3 scripts/bis_paises_planilhas.py --sem-download
  python3 scripts/bis_paises_planilhas.py --apenas WS_DSR,WS_CREDIT_GAP
"""

from __future__ import annotations

import argparse
import sys
import zipfile
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATA_DIR = ROOT / "data" / "bis"
OUTPUT_DIR = ROOT / "output" / "bis_paises"
BIS_BULK = "https://data.bis.org/static/bulk/{stem}_csv_flat.zip"

INDICADORES = [
    ("WS_CBPOL", "Taxas de política dos bancos centrais"),
    ("WS_CBTA", "Ativos totais dos bancos centrais"),
    ("WS_TC", "Crédito ao setor não financeiro"),
    ("WS_CREDIT_GAP", "Hiato crédito/PIB"),
    ("WS_DSR", "Índice de serviço da dívida"),
    ("WS_GLI", "Indicadores de liquidez global"),
]

COLUNAS_PAIS = (
    "REF_AREA:Reference area",
    "BORROWERS_CTY:Borrowers' country",
)

DESCARTAR = {
    "STRUCTURE",
    "STRUCTURE_ID",
    "ACTION",
    "OBS_CONF:Observation confidentiality",
    "OBS_PRE_BREAK:Pre-Break Observation",
    "OBS_PRE_BREAK:Observation pre-break value",
    "CONF_STATUS:Confidentiality - status",
}

PAISES_PT = {
    "AR": "Argentina",
    "AT": "Áustria",
    "AU": "Austrália",
    "BE": "Bélgica",
    "BR": "Brasil",
    "CA": "Canadá",
    "CH": "Suíça",
    "CL": "Chile",
    "CN": "China",
    "CO": "Colômbia",
    "CZ": "Tchéquia",
    "DE": "Alemanha",
    "DK": "Dinamarca",
    "ES": "Espanha",
    "FI": "Finlândia",
    "FR": "França",
    "GB": "Reino Unido",
    "GR": "Grécia",
    "HK": "Hong Kong",
    "HR": "Croácia",
    "HU": "Hungria",
    "ID": "Indonésia",
    "IE": "Irlanda",
    "IL": "Israel",
    "IN": "Índia",
    "IS": "Islândia",
    "IT": "Itália",
    "JP": "Japão",
    "KR": "Coreia do Sul",
    "KW": "Kuwait",
    "MA": "Marrocos",
    "MK": "Macedônia do Norte",
    "MX": "México",
    "MY": "Malásia",
    "NL": "Países Baixos",
    "NO": "Noruega",
    "NZ": "Nova Zelândia",
    "PE": "Peru",
    "PH": "Filipinas",
    "PL": "Polônia",
    "PT": "Portugal",
    "RO": "Romênia",
    "RS": "Sérvia",
    "RU": "Rússia",
    "SA": "Arábia Saudita",
    "SE": "Suécia",
    "SK": "Eslováquia",
    "TH": "Tailândia",
    "TR": "Turquia",
    "US": "Estados Unidos",
    "XM": "Zona do euro",
    "5C": "Área do euro",
    "ZA": "África do Sul",
    "G2": "G20",
    "4U": "AL e Caribe emergentes",
    "4Y": "Ásia e Pacífico emergentes",
}


def _borda() -> Border:
    lado = Side(style="thin", color="1A1A1A")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def partir_codigo_nome(valor: object) -> tuple[str, str]:
    """Separa 'BR: Brazil' em código e rótulo em inglês."""
    txt = "" if valor is None or (isinstance(valor, float) and pd.isna(valor)) else str(valor).strip()
    if not txt:
        return "", ""
    if ":" in txt:
        cod, nome = txt.split(":", 1)
        return cod.strip(), nome.strip()
    return txt, txt


def nome_pais(codigo: str, nome_en: str) -> str:
    return PAISES_PT.get(codigo, nome_en or codigo)


def nome_aba(codigo: str, nome: str, usados: set[str]) -> str:
    base = f"{codigo} {nome}".strip() or codigo or "pais"
    for ch in r"\/*?:[]":
        base = base.replace(ch, "-")
    base = " ".join(base.split())[:31] or codigo[:31] or "pais"
    cand = base
    n = 2
    while cand in usados:
        suf = f" {n}"
        cand = (base[: 31 - len(suf)] + suf)
        n += 1
    usados.add(cand)
    return cand


def coluna_pais(colunas: list[str]) -> str:
    for cand in COLUNAS_PAIS:
        if cand in colunas:
            return cand
    for col in colunas:
        up = col.upper()
        if up.startswith("REF_AREA") or up.startswith("BORROWERS_CTY"):
            return col
    raise ValueError(f"Sem coluna de país em {colunas}")


def colunas_uteis(df: pd.DataFrame) -> pd.DataFrame:
    out = df.drop(columns=[c for c in DESCARTAR if c in df.columns], errors="ignore")
    vazias = [c for c in out.columns if out[c].isna().all() or (out[c].astype(str).str.strip() == "").all()]
    return out.drop(columns=vazias, errors="ignore")


def baixar_zip(stem: str, cache_dir: Path, baixar: bool) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    zip_path = cache_dir / f"{stem}_csv_flat.zip"
    if zip_path.exists() and zip_path.stat().st_size > 0:
        return zip_path
    if not baixar:
        raise FileNotFoundError(zip_path)
    url = BIS_BULK.format(stem=stem)
    print(f"  baixando {url}", flush=True)
    resp = requests.get(url, timeout=300)
    resp.raise_for_status()
    zip_path.write_bytes(resp.content)
    return zip_path


def ler_csv_flat(zip_path: Path) -> pd.DataFrame:
    with zipfile.ZipFile(zip_path) as zf:
        nome = next(n for n in zf.namelist() if n.lower().endswith(".csv"))
        with zf.open(nome) as fh:
            return pd.read_csv(fh, dtype=str, low_memory=False)


def preparar(df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    col = coluna_pais(list(df.columns))
    out = colunas_uteis(df)
    partes = out[col].map(partir_codigo_nome)
    out = out.copy()
    out["_codigo"] = [p[0] for p in partes]
    out["_nome_en"] = [p[1] for p in partes]
    out["_pais"] = [nome_pais(c, n) for c, n in partes]
    if "OBS_VALUE:Observation Value" in out.columns:
        out["OBS_VALUE:Observation Value"] = pd.to_numeric(out["OBS_VALUE:Observation Value"], errors="coerce")
    tempo = "TIME_PERIOD:Time period or range"
    if tempo in out.columns:
        chaves = [tempo] + [c for c in out.columns if c not in {tempo, "_codigo", "_nome_en", "_pais"}]
        out = out.sort_values(chaves, kind="mergesort")
    return out, col


def _escrever_cabecalho(ws, cabecalhos: list[str], borda: Border) -> None:
    fill = PatternFill("solid", fgColor="E8E8E8")
    fonte = Font(name="Calibri", size=10, bold=True)
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(1, col, cab)
        cell.font = fonte
        cell.fill = fill
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _rotulo_coluna(nome: str) -> str:
    return nome.split(":", 1)[-1] if ":" in nome and not nome.startswith("_") else nome


def _estilizar_aba(ws, n_linhas: int, n_cols: int, bordar_dados: bool) -> None:
    borda = _borda()
    fill = PatternFill("solid", fgColor="E8E8E8")
    fonte_cab = Font(name="Calibri", size=10, bold=True)
    fonte = Font(name="Calibri", size=9)
    for col in range(1, n_cols + 1):
        cell = ws.cell(1, col)
        cell.font = fonte_cab
        cell.fill = fill
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 22
    if bordar_dados:
        fill_alt = PatternFill("solid", fgColor="F4F4F4")
        for row in ws.iter_rows(min_row=2, max_row=n_linhas + 1, max_col=n_cols):
            par = row[0].row % 2 == 0
            for cell in row:
                cell.font = fonte
                cell.border = borda
                if par:
                    cell.fill = fill_alt
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.0000"
    ws.freeze_panes = "A2"
    if n_linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_linhas + 1}"
    ws.row_dimensions[1].height = 30


def gerar_planilha(df: pd.DataFrame, titulo: str, fonte: str, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    grupos = []
    for (codigo, pais), g in df.groupby(["_codigo", "_pais"], sort=False):
        grupos.append((str(codigo), str(pais), g))
    grupos.sort(key=lambda x: (x[0] != "BR", x[1].casefold(), x[0]))

    usados: set[str] = {"Notas", "Indice"}
    indice_linhas = []
    tempo = "TIME_PERIOD:Time period or range"
    drop_interno = {"_codigo", "_nome_en", "_pais"}
    abas: list[tuple[str, pd.DataFrame]] = []

    for codigo, pais, g in grupos:
        aba = nome_aba(codigo, pais, usados)
        dados = g.drop(columns=[c for c in drop_interno if c in g.columns]).copy()
        dados.columns = [_rotulo_coluna(c) for c in dados.columns]
        abas.append((aba, dados))
        periodos = g[tempo] if tempo in g.columns else pd.Series(dtype=str)
        indice_linhas.append(
            {
                "Código": codigo,
                "País": pais,
                "Aba": aba,
                "Observações": int(len(dados)),
                "Início": "" if periodos.empty else str(periodos.min()),
                "Fim": "" if periodos.empty else str(periodos.max()),
            }
        )

    notas = pd.DataFrame(
        {
            "Nota": [
                titulo,
                "Uma aba por país/economia com as séries e dimensões do BIS para esse referente.",
                f"Fonte: {fonte}",
                "Catálogo: https://data.bis.org/bulkdownload",
                "Arquivo CSV flat oficial (SDMX). Agregados (G20, área do euro, regiões) "
                "entram como abas quando o BIS os publica.",
                "Locational/consolidated banking e títulos de dívida do catálogo não cabem "
                "em Excel (dezenas a centenas de MB) e ficam fora deste recorte.",
            ]
        }
    )
    indice = pd.DataFrame(indice_linhas)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        notas.to_excel(writer, sheet_name="Notas", index=False, header=False)
        indice.to_excel(writer, sheet_name="Indice", index=False)
        for aba, dados in abas:
            dados.to_excel(writer, sheet_name=aba, index=False)

    wb = load_workbook(path)
    ws0 = wb["Notas"]
    ws0["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws0.column_dimensions["A"].width = 120
    for row in ws0.iter_rows(min_row=2, max_row=ws0.max_row, max_col=1):
        row[0].alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[row[0].row].height = 32
    _estilizar_aba(wb["Indice"], len(indice), indice.shape[1], bordar_dados=True)
    for aba, dados in abas:
        _estilizar_aba(wb[aba], len(dados), dados.shape[1], bordar_dados=len(dados) <= 2500)
    wb.save(path)
    return path


def processar(stem: str, titulo: str, cache_dir: Path, output_dir: Path, baixar: bool) -> Path:
    print(f"{stem} — {titulo}", flush=True)
    zip_path = baixar_zip(stem, cache_dir, baixar=baixar)
    bruto = ler_csv_flat(zip_path)
    prep, _col = preparar(bruto)
    dest = output_dir / f"bis_{stem}_paises.xlsx"
    fonte = BIS_BULK.format(stem=stem)
    path = gerar_planilha(prep, f"{titulo} — BIS {stem}", fonte, dest)
    n_paises = prep["_codigo"].nunique()
    print(f"  {len(prep)} linhas, {n_paises} abas → {path}", flush=True)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--sem-download", action="store_true")
    parser.add_argument("--apenas", default="", help="Lista de códigos WS_* separados por vírgula")
    args = parser.parse_args(argv)

    escolhidos = {x.strip().upper() for x in args.apenas.split(",") if x.strip()}
    alvos = [p for p in INDICADORES if not escolhidos or p[0] in escolhidos]
    if not alvos:
        raise SystemExit("Nenhum indicador selecionado.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    gerados = []
    for stem, titulo in alvos:
        gerados.append((stem, titulo, processar(stem, titulo, args.cache_dir, args.output_dir, not args.sem_download)))

    cat = args.output_dir / "bis_indice_indicadores.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"
    borda = _borda()
    ws["A1"] = "Indicadores BIS de política monetária e crédito — uma planilha por tema, uma aba por país"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
    ws.merge_cells("A1:D1")
    cabs = ["Código", "Indicador", "Arquivo", "Fonte"]
    for col, cab in enumerate(cabs, start=1):
        cell = ws.cell(3, col, cab)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")
        cell.border = borda
    for i, (stem, titulo, path) in enumerate(gerados):
        vals = [stem, titulo, path.name, BIS_BULK.format(stem=stem)]
        for col, valor in enumerate(vals, start=1):
            cell = ws.cell(i + 4, col, valor)
            cell.font = Font(name="Calibri", size=9)
            cell.border = borda
    for col, w in enumerate([16, 44, 36, 70], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(cat)
    print(f"Índice: {cat}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
