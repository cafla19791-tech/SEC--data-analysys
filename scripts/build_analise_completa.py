#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build consolidated Excel answering the 10 fiscal/monetary questions."""

from __future__ import annotations

import csv
import json
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path("/workspace")
RAW = ROOT / "data" / "raw"
BNDES = ROOT / "data" / "bndes"
CVM = ROOT / "data" / "cvm"
OUT = ROOT / "output"
OUT.mkdir(parents=True, exist_ok=True)

MONTH_MAP = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def load_bcb_json(path: Path) -> pd.DataFrame:
    data = json.loads(path.read_text())
    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["data"], dayfirst=True)
    df["value"] = pd.to_numeric(df["valor"], errors="coerce")
    return df.sort_values("date").reset_index(drop=True)


def parse_dbgg_sheet(sheet: str, value_col: int = 14) -> pd.DataFrame:
    df = pd.read_excel(RAW / "Dbggindexp.xlsx", sheet_name=sheet, header=None)
    rows = []
    year = None
    for _, row in df.iterrows():
        y = pd.to_numeric(row[0], errors="coerce")
        if pd.notna(y):
            year = int(y)
        mraw = str(row[1]).strip().lower() if pd.notna(row[1]) else ""
        if year is None or mraw not in MONTH_MAP:
            continue
        val = pd.to_numeric(row[value_col], errors="coerce")
        if pd.isna(val):
            continue
        rows.append(
            {
                "year": year,
                "month": MONTH_MAP[mraw],
                "date": pd.Timestamp(year, MONTH_MAP[mraw], 1),
                "value": float(val),
            }
        )
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def q1_dbgg_fatores():
    stock = parse_dbgg_sheet("DividaR$")
    juros = parse_dbgg_sheet("JurosR$")
    emis = parse_dbgg_sheet("PrimarioR$")
    dbgg_pib = load_bcb_json(RAW / "bcb_series" / "13762_dbgg_pib.json")
    dbgg_rm = load_bcb_json(RAW / "bcb_series" / "13761_dbgg_rm.json")
    dbgg_pib_old = load_bcb_json(RAW / "bcb_series" / "4537_dbgg_pib_ate2007.json")

    dec = stock[stock["month"] == 12].copy()
    jun2026 = stock[(stock["year"] == 2026) & (stock["month"] == 6)]
    stocks = pd.concat([dec, jun2026], ignore_index=True)

    juros_a = juros.groupby("year", as_index=False)["value"].sum().rename(
        columns={"value": "juros_nominais_rs_mi"}
    )
    emis_a = emis.groupby("year", as_index=False)["value"].sum().rename(
        columns={"value": "emissoes_liquidas_rs_mi"}
    )

    sgs = dbgg_rm.copy()
    sgs["year"] = sgs["date"].dt.year
    sgs["month"] = sgs["date"].dt.month
    sgs_dec = sgs[sgs["month"] == 12][["year", "value"]].rename(
        columns={"value": "dbgg_sgs_rs_mi"}
    )
    sgs_jun26 = sgs[(sgs["year"] == 2026) & (sgs["month"] == 6)][
        ["year", "value"]
    ].rename(columns={"value": "dbgg_sgs_rs_mi"})
    sgs_ye = pd.concat([sgs_dec, sgs_jun26], ignore_index=True).drop_duplicates(
        "year", keep="last"
    )

    pib = dbgg_pib.copy()
    pib["year"] = pib["date"].dt.year
    pib["month"] = pib["date"].dt.month
    pib_dec = pib[pib["month"] == 12][["year", "value"]].rename(
        columns={"value": "dbgg_pct_pib"}
    )
    pib_jun26 = pib[(pib["year"] == 2026) & (pib["month"] == 6)][
        ["year", "value"]
    ].rename(columns={"value": "dbgg_pct_pib"})
    pib_ye = pd.concat([pib_dec, pib_jun26], ignore_index=True).drop_duplicates(
        "year", keep="last"
    )
    # Pre-2007 methodology (% PIB only)
    old = dbgg_pib_old.copy()
    old["year"] = old["date"].dt.year
    old["month"] = old["date"].dt.month
    old_dec = old[old["month"] == 12][["year", "value"]].rename(columns={"value": "dbgg_pct_pib_old"})
    pib_ye = pib_ye.merge(old_dec, on="year", how="outer")
    pib_ye["dbgg_pct_pib"] = pib_ye["dbgg_pct_pib"].fillna(pib_ye["dbgg_pct_pib_old"])
    pib_ye = pib_ye.drop(columns=["dbgg_pct_pib_old"])

    base = pd.DataFrame({"year": list(range(2002, 2027))})
    base = base.merge(sgs_ye, on="year", how="left")
    base = base.merge(pib_ye, on="year", how="left")
    base = base.merge(juros_a, on="year", how="left")
    base = base.merge(emis_a, on="year", how="left")

    stocks_map = stocks.drop_duplicates("year", keep="last").set_index("year")["value"]
    base["dbgg_rs_mi"] = base["year"].map(stocks_map)
    base["dbgg_rs_mi"] = base["dbgg_rs_mi"].fillna(base["dbgg_sgs_rs_mi"])
    base["dbgg_ant_rs_mi"] = base["dbgg_rs_mi"].shift(1)
    base["variacao_rs_mi"] = base["dbgg_rs_mi"] - base["dbgg_ant_rs_mi"]
    base["demais_fatores_rs_mi"] = (
        base["variacao_rs_mi"]
        - base["juros_nominais_rs_mi"].fillna(0)
        - base["emissoes_liquidas_rs_mi"].fillna(0)
    )
    mask_flow = base["juros_nominais_rs_mi"].notna() | base["emissoes_liquidas_rs_mi"].notna()
    base.loc[~mask_flow, "demais_fatores_rs_mi"] = np.nan

    base["dbgg_pct_pib_ant"] = base["dbgg_pct_pib"].shift(1)
    base["variacao_pp_pib"] = base["dbgg_pct_pib"] - base["dbgg_pct_pib_ant"]
    base["pib_impl_rs_mi"] = base["dbgg_rs_mi"] / (base["dbgg_pct_pib"] / 100.0)
    base["efeito_pib_pp"] = (
        -base["dbgg_pct_pib_ant"]
        * (base["pib_impl_rs_mi"] - base["pib_impl_rs_mi"].shift(1))
        / base["pib_impl_rs_mi"]
    )
    base["juros_pp_pib"] = 100.0 * base["juros_nominais_rs_mi"] / base["pib_impl_rs_mi"]
    base["emissoes_pp_pib"] = 100.0 * base["emissoes_liquidas_rs_mi"] / base["pib_impl_rs_mi"]
    base["demais_pp_pib"] = (
        base["variacao_pp_pib"]
        - base["juros_pp_pib"].fillna(0)
        - base["emissoes_pp_pib"].fillna(0)
        - base["efeito_pib_pp"].fillna(0)
    )
    base.loc[~mask_flow, "demais_pp_pib"] = np.nan

    detail = base[
        [
            "year", "dbgg_rs_mi", "dbgg_pct_pib", "variacao_rs_mi", "variacao_pp_pib",
            "juros_nominais_rs_mi", "emissoes_liquidas_rs_mi", "demais_fatores_rs_mi",
            "juros_pp_pib", "emissoes_pp_pib", "efeito_pib_pp", "demais_pp_pib",
        ]
    ].copy()
    detail.columns = [
        "Ano", "DBGG (R$ milhoes)", "DBGG (% PIB)", "Variacao (R$ milhoes)",
        "Variacao (p.p. PIB)", "Juros nominais (R$ milhoes)",
        "Emissoes liquidas (R$ milhoes)",
        "Demais fatores R$ (cambio/indices/reconhecimento/outros)",
        "Juros nominais (p.p. PIB)", "Emissoes liquidas (p.p. PIB)",
        "Efeito PIB nominal (p.p. PIB)", "Demais fatores (p.p. PIB)",
    ]
    notes = (
        "Fontes: BCB Tabelas Especiais Dbggindexp.xlsx; BCB SGS 13761/13762. "
        "Identidade em R$: dDBGG = juros + emissoes liquidas + demais. "
        "Em p.p. do PIB: juros e emissoes / PIB implicito; efeito PIB = "
        "-DBGG/PIB(t-1)*dPIB/PIB(t); demais fecha a variacao DBGG/PIB. "
        "Fluxos detalhados a partir de 2007 (Dbggindexp desde dez/2006). DBGG % PIB 2002-2006: serie SGS 4537 (metodologia ate 2007); a partir de 2006-12: SGS 13762 (metodologia a partir de 2008). "
        "2026 usa o ultimo mes disponivel (nao e fechamento de ano)."
    )
    monthly = stock.merge(
        dbgg_pib.rename(columns={"value": "dbgg_pct_pib", "date": "date_pib"})[
            ["date_pib", "dbgg_pct_pib"]
        ],
        left_on="date", right_on="date_pib", how="left",
    )
    monthly = monthly[["date", "value", "dbgg_pct_pib"]].rename(
        columns={"date": "Data", "value": "DBGG (R$ milhoes)", "dbgg_pct_pib": "DBGG (% PIB)"}
    )
    return detail, monthly, notes


def compound_daily_selic(path: Path):
    rows = json.loads(path.read_text())
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["data"], dayfirst=True)
    df["daily_pct"] = pd.to_numeric(df["valor"], errors="coerce")
    df = df[(df["date"] >= "2003-01-01") & (df["date"] <= "2026-06-30")].copy()
    df["factor"] = 1.0 + df["daily_pct"] / 100.0
    acc = float(df["factor"].prod() - 1.0)
    df["acumulado"] = df["factor"].cumprod() - 1.0
    return acc, df


def q2_q4_policy_rates():
    selic_acc, selic_df = compound_daily_selic(RAW / "selic_diaria.json")
    selic_summary = pd.DataFrame(
        [{
            "Indicador": "SELIC acumulada (BCB serie 11, taxa diaria)",
            "Inicio": "02/01/2003",
            "Fim": "30/06/2026",
            "Fator acumulado": 1.0 + selic_acc,
            "Taxa acumulada (%)": 100.0 * selic_acc,
            "Observacoes": "Produto dos fatores diarios (1 + i_t/100) - 1",
        }]
    )

    path = RAW / "bis_cbpol" / "WS_CBPOL_csv_flat.csv"
    series: dict[str, list[tuple[str, float]]] = {}
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if not row["FREQ:Frequency"].startswith("M"):
                continue
            tp = row["TIME_PERIOD:Time period or range"]
            if tp < "2003-01" or tp > "2026-06":
                continue
            val = row["OBS_VALUE:Observation Value"]
            if val is None or val == "":
                continue
            try:
                fval = float(val)
            except ValueError:
                continue
            if not np.isfinite(fval):
                continue
            area = row["REF_AREA:Reference area"]
            series.setdefault(area, []).append((tp, fval))

    expected = set(pd.period_range("2003-01", "2026-06", freq="M").strftime("%Y-%m"))
    records = []
    for area, pts in series.items():
        pts = sorted(pts)
        months = {p[0] for p in pts}
        coverage = len(months & expected) / len(expected)
        factor = 1.0
        for _, r in pts:
            factor *= (1.0 + r / 100.0) ** (1.0 / 12.0)
        if not np.isfinite(factor):
            continue
        code, name = area.split(": ", 1) if ": " in area else (area, area)
        records.append({
            "Codigo": code,
            "Pais_Area": name,
            "Meses com dado": len(months),
            "Cobertura (%)": round(100 * coverage, 1),
            "Taxa acumulada (%)": 100.0 * (factor - 1.0),
            "Fator acumulado": factor,
            "Primeiro mes": pts[0][0],
            "Ultimo mes": pts[-1][0],
        })
    ranking = pd.DataFrame(records).sort_values("Taxa acumulada (%)").reset_index(drop=True)
    ranking.insert(0, "Rank", ranking.index + 1)
    others = ranking[ranking["Codigo"] != "BR"].copy()
    notes = (
        "Q2: SELIC diaria BCB SGS 11, acumulada por produto dos fatores diarios "
        "entre 02/01/2003 e 30/06/2026. "
        "Q3/Q4: BIS WS_CBPOL csv flat, frequencia mensal fim de periodo; "
        "acumulacao product((1+r/100)^(1/12))-1 de jan/2003 a jun/2026. "
        f"Jurisdicoes disponiveis: {len(ranking)} (arquivo BIS atual; nao 49). "
        "Brasil no ranking usa metodologia BIS para comparabilidade; "
        "a taxa oficial da Q2 e a SELIC diaria do BCB."
    )
    return selic_summary, others, ranking, notes


def extract_cvm_company(cd_cvm: str, years: range) -> pd.DataFrame:
    rows = []
    for y in years:
        zpath = CVM / f"dfp_{y}.zip"
        if not zpath.exists() or zpath.stat().st_size < 1000:
            continue
        with zipfile.ZipFile(zpath) as zf:
            bpp_name = f"dfp_cia_aberta_BPP_con_{y}.csv"
            dre_name = f"dfp_cia_aberta_DRE_con_{y}.csv"
            if bpp_name not in zf.namelist():
                continue
            bpp = pd.read_csv(zf.open(bpp_name), sep=";", dtype=str, encoding="latin-1")
            dre = pd.read_csv(zf.open(dre_name), sep=";", dtype=str, encoding="latin-1")
        bpp = bpp[
            (bpp["CD_CVM"].astype(str).str.lstrip("0") == str(int(cd_cvm)))
            & (bpp["ORDEM_EXERC"].str.upper().str.startswith("ULT"))
            & (bpp["DT_REFER"].str.startswith(str(y)))
        ]
        dre = dre[
            (dre["CD_CVM"].astype(str).str.lstrip("0") == str(int(cd_cvm)))
            & (dre["ORDEM_EXERC"].str.upper().str.startswith("ULT"))
            & (dre["DT_REFER"].str.startswith(str(y)))
        ]

        def acc(df, codes):
            sub = df[df["CD_CONTA"].isin(codes)]
            for c in codes:
                s = sub[sub["CD_CONTA"] == c]
                if not s.empty:
                    return pd.to_numeric(s.iloc[0]["VL_CONTA"], errors="coerce")
            return np.nan

        st = acc(bpp, ["2.01.04"])
        lt = acc(bpp, ["2.02.01"])
        debt = np.nan if pd.isna(st) and pd.isna(lt) else (0 if pd.isna(st) else st) + (0 if pd.isna(lt) else lt)
        rows.append({
            "Ano": y,
            "Divida bruta (R$ mil)": debt,
            "Resultado financeiro (R$ mil)": acc(dre, ["3.06"]),
            "Lucro liquido (R$ mil)": acc(dre, ["3.11", "3.09"]),
        })
    return pd.DataFrame(rows)


def q5_q7_estatais():
    petro = extract_cvm_company("9512", range(2010, 2026))
    eletro = extract_cvm_company("2437", range(2010, 2026))
    petro.insert(1, "Empresa", "Petrobras")
    eletro.insert(1, "Empresa", "Eletrobras/Axia")

    dlsp = load_bcb_json(RAW / "bcb_series" / "4474_dlsp_estatais_rm.json")
    juros = load_bcb_json(RAW / "bcb_series" / "4612_juros_estatais.json")
    res = load_bcb_json(RAW / "bcb_series" / "4579_resultado_nominal_estatais.json")

    def annualize_stock(df, value_name):
        d = df.copy()
        d["Ano"] = d["date"].dt.year
        d["month"] = d["date"].dt.month
        dec = d[d["month"] == 12][["Ano", "value"]].rename(columns={"value": value_name})
        jun = d[(d["Ano"] == 2026) & (d["month"] == 6)][["Ano", "value"]].rename(columns={"value": value_name})
        return pd.concat([dec, jun], ignore_index=True).drop_duplicates("Ano", keep="last")

    def annualize_flow(df, value_name):
        d = df.copy()
        d["Ano"] = d["date"].dt.year
        return d.groupby("Ano", as_index=False)["value"].sum().rename(columns={"value": value_name})

    dem_debt = annualize_stock(dlsp, "Valor")
    dem_j = annualize_flow(juros, "Valor")
    dem_r = annualize_flow(res, "Valor")

    def pack(company_df, value_col, conceito, unidade, empresa):
        out = company_df[["Ano", value_col]].rename(columns={value_col: "Valor"}).copy()
        out.insert(1, "Empresa", empresa)
        out["Conceito"] = conceito
        out["Unidade"] = unidade
        return out[["Ano", "Empresa", "Conceito", "Unidade", "Valor"]]

    debt = pd.concat([
        pack(petro, "Divida bruta (R$ mil)", "Divida bruta contabil CVM 2.01.04+2.02.01", "R$ mil", "Petrobras"),
        pack(eletro, "Divida bruta (R$ mil)", "Divida bruta contabil CVM 2.01.04+2.02.01", "R$ mil", "Eletrobras/Axia"),
        pack(dem_debt.assign(Empresa="x"), "Valor", "Divida liquida DLSP estatais BCB (nao bruta)", "R$ milhoes",
             "Demais estatais (BCB; exc. Petrobras/Eletrobras/bancos)").assign(
                 **{"Empresa": "Demais estatais (BCB; exc. Petrobras/Eletrobras/bancos)"}
             ) if False else dem_debt.assign(
                 Empresa="Demais estatais (BCB; exc. Petrobras/Eletrobras/bancos)",
                 Conceito="Divida liquida DLSP estatais BCB (nao bruta)",
                 Unidade="R$ milhoes",
             )[["Ano", "Empresa", "Conceito", "Unidade", "Valor"]],
    ], ignore_index=True)

    juros_df = pd.concat([
        pack(petro, "Resultado financeiro (R$ mil)", "Resultado financeiro DRE conta 3.06", "R$ mil", "Petrobras"),
        pack(eletro, "Resultado financeiro (R$ mil)", "Resultado financeiro DRE conta 3.06", "R$ mil", "Eletrobras/Axia"),
        dem_j.assign(
            Empresa="Demais estatais (BCB; exc. Petrobras/Eletrobras/bancos)",
            Conceito="Juros nominais NFSP SGS 4612",
            Unidade="R$ milhoes",
        )[["Ano", "Empresa", "Conceito", "Unidade", "Valor"]],
    ], ignore_index=True)

    ll_df = pd.concat([
        pack(petro, "Lucro liquido (R$ mil)", "Lucro/prejuizo liquido consolidado conta 3.11", "R$ mil", "Petrobras"),
        pack(eletro, "Lucro liquido (R$ mil)", "Lucro/prejuizo liquido consolidado conta 3.11", "R$ mil", "Eletrobras/Axia"),
        dem_r.assign(
            Empresa="Demais estatais (BCB; exc. Petrobras/Eletrobras/bancos)",
            Conceito="Resultado nominal NFSP SGS 4579 (conceito fiscal)",
            Unidade="R$ milhoes",
        )[["Ano", "Empresa", "Conceito", "Unidade", "Valor"]],
    ], ignore_index=True)

    notes = (
        "Petrobras e Eletrobras/Axia: CVM DFP consolidado 2010-2025. "
        "Divida bruta = 2.01.04 + 2.02.01. Valores CVM em R$ mil. "
        "Demais estatais: BCB ja exclui Petrobras, Eletrobras e bancos. "
        "Estoque = DLSP (liquida) SGS 4474; juros SGS 4612; resultado nominal SGS 4579. "
        "Sem DFP CVM 2002-2009 no layout atual."
    )
    return debt, juros_df, ll_df, notes


def q8_bndes_ipca():
    frames = []
    for path in sorted(BNDES.glob("bndes_*.xlsx")):
        try:
            df = pd.read_excel(path, sheet_name="DESEMBOLSOS_BASE DE DADOS", header=2)
        except Exception:
            continue
        cols = {str(c).strip().replace("\n", " "): c for c in df.columns}
        ano_col = next((cols[k] for k in cols if k.upper().startswith("ANO")), None)
        val_col = next((cols[k] for k in cols if "DESEMBOLSO" in k.upper() and "R$" in k.upper()), None)
        if ano_col is None or val_col is None:
            continue
        tmp = df[[ano_col, val_col]].copy()
        tmp.columns = ["Ano", "Desembolso_R$"]
        tmp["Ano"] = pd.to_numeric(tmp["Ano"], errors="coerce")
        tmp["Desembolso_R$"] = pd.to_numeric(tmp["Desembolso_R$"], errors="coerce")
        tmp = tmp.dropna(subset=["Ano", "Desembolso_R$"])
        tmp["Ano"] = tmp["Ano"].astype(int)
        frames.append(tmp)
    all_d = pd.concat(frames, ignore_index=True)
    annual = all_d.groupby("Ano", as_index=False)["Desembolso_R$"].sum()
    annual = annual[(annual["Ano"] >= 2003) & (annual["Ano"] <= 2026)]

    ipca = load_bcb_json(RAW / "bcb_series" / "433_ipca.json").sort_values("date")
    ipca["factor"] = 1.0 + ipca["value"] / 100.0
    ipca["index"] = 100.0 * ipca["factor"].cumprod() / ipca["factor"].iloc[0]
    target = ipca[(ipca["date"].dt.year == 2026) & (ipca["date"].dt.month == 6)]
    if target.empty:
        target_idx = float(ipca["index"].iloc[-1])
        target_label = ipca["date"].iloc[-1].strftime("%m/%Y")
    else:
        target_idx = float(target["index"].iloc[0])
        target_label = "06/2026"
    ipca["Ano"] = ipca["date"].dt.year
    idx_ano = ipca.groupby("Ano", as_index=False)["index"].mean()
    annual = annual.merge(idx_ano, on="Ano", how="left")
    annual["Desembolso atualizado IPCA (R$)"] = annual["Desembolso_R$"] * target_idx / annual["index"]
    annual = annual.rename(columns={
        "Desembolso_R$": "Desembolso corrente Sistema BNDES (R$)",
        "index": "IPCA medio do ano (indice)",
    })
    annual["IPCA referencia"] = target_label
    notes = (
        f"Desembolsos do Sistema BNDES (bases oficiais informadas). "
        f"Atualizacao IPCA BCB SGS 433 ate {target_label}, via IPCA medio do ano. "
        "As bases nao discriminam BNB nem BASA; cobrem o Sistema BNDES."
    )
    return annual, notes


def q9_q10_gdp_ppp():
    csv_path = next((RAW / "wb_gdp_ppp").glob("API_NY.GDP.PCAP.PP.KD*.csv"))
    df = pd.read_csv(csv_path, skiprows=4)
    df = df.loc[:, ~df.columns.str.match(r"^Unnamed")]
    meta = next((RAW / "wb_gdp_ppp").glob("Metadata_Country*.csv"))
    meta_df = pd.read_csv(meta)
    countries = meta_df[meta_df["Region"].notna() & (meta_df["Region"] != "")]["Country Code"].tolist()
    df = df[df["Country Code"].isin(countries)].copy()
    years = [str(y) for y in range(2002, 2026)]
    wide = df[["Country Name", "Country Code"] + [y for y in years if y in df.columns]].copy()

    rankings = {}
    for y in range(2002, 2026):
        ys = str(y)
        if ys not in df.columns:
            continue
        sub = df[["Country Name", "Country Code", ys]].copy()
        sub[ys] = pd.to_numeric(sub[ys], errors="coerce")
        sub = sub.dropna(subset=[ys]).sort_values(ys, ascending=False)
        sub.insert(0, "Rank", range(1, len(sub) + 1))
        sub = sub.rename(columns={
            ys: "PIB per capita PPP (USD int. const. 2021)",
            "Country Name": "Pais", "Country Code": "Codigo",
        })
        rankings[y] = sub.reset_index(drop=True)

    v = df[["Country Name", "Country Code", "2002", "2016"]].copy()
    v["2002"] = pd.to_numeric(v["2002"], errors="coerce")
    v["2016"] = pd.to_numeric(v["2016"], errors="coerce")
    v = v.dropna(subset=["2002", "2016"])
    v = v[v["2002"] > 0]
    v["Variacao % 2002-2016"] = 100.0 * (v["2016"] / v["2002"] - 1.0)
    v = v.sort_values("Variacao % 2002-2016", ascending=False).reset_index(drop=True)
    v.insert(0, "Rank", v.index + 1)
    v = v.rename(columns={
        "Country Name": "Pais", "Country Code": "Codigo",
        "2002": "PIB pc PPP 2002", "2016": "PIB pc PPP 2016",
    })
    notes = (
        "Fonte: WDI NY.GDP.PCAP.PP.KD (PPP constant 2021 international $), "
        "atualizacao 2026-07-13. Exclui agregados (Region vazia)."
    )
    return wide, rankings, v, notes


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:200]:
            if cell.value is not None:
                length = max(length, min(60, len(str(cell.value))))
        ws.column_dimensions[letter].width = max(12, length + 2)


def main():
    print("Q1...")
    q1, q1_m, n1 = q1_dbgg_fatores()
    print("Q2-Q4...")
    q2, q3, q4, n234 = q2_q4_policy_rates()
    print("Q5-Q7...")
    q5, q6, q7, n567 = q5_q7_estatais()
    print("Q8...")
    q8, n8 = q8_bndes_ipca()
    print("Q9-Q10...")
    wide, rankings, var, n910 = q9_q10_gdp_ppp()

    out_path = OUT / "analise_divida_juros_estatais_pib.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({
            "Pergunta": [
                "1 DBGG fatores", "2 SELIC acumulada", "3 Taxas BIS demais paises",
                "4 Ranking taxas", "5 Divida estatais", "6 Juros estatais",
                "7 Resultados", "8 BNDES IPCA", "9 PIB PPP rankings", "10 Var PPP 2002-2016",
            ],
            "Aba": [
                "Q1_DBGG_fatores", "Q2_SELIC_acumulada", "Q3_BIS_taxas_paises",
                "Q4_Ranking_taxas", "Q5_Divida_estatais", "Q6_Juros_estatais",
                "Q7_Resultados_estatais", "Q8_BNDES_IPCA", "Q9_PIB_PPP_niveis / Q9_rank_YYYY",
                "Q10_Var_PPP_2002_2016",
            ],
            "Notas": [n1, n234, n234, n234, n567, n567, n567, n8, n910, n910],
        }).to_excel(writer, sheet_name="Indice", index=False)
        q1.to_excel(writer, sheet_name="Q1_DBGG_fatores", index=False)
        q1_m.to_excel(writer, sheet_name="Q1_DBGG_mensal", index=False)
        q2.to_excel(writer, sheet_name="Q2_SELIC_acumulada", index=False)
        q3.to_excel(writer, sheet_name="Q3_BIS_taxas_paises", index=False)
        q4.to_excel(writer, sheet_name="Q4_Ranking_taxas", index=False)
        q5.to_excel(writer, sheet_name="Q5_Divida_estatais", index=False)
        q6.to_excel(writer, sheet_name="Q6_Juros_estatais", index=False)
        q7.to_excel(writer, sheet_name="Q7_Resultados_estatais", index=False)
        q8.to_excel(writer, sheet_name="Q8_BNDES_IPCA", index=False)
        wide.to_excel(writer, sheet_name="Q9_PIB_PPP_niveis", index=False)
        for y, rdf in rankings.items():
            rdf.to_excel(writer, sheet_name=f"Q9_rank_{y}"[:31], index=False)
        var.to_excel(writer, sheet_name="Q10_Var_PPP_2002_2016", index=False)
        pd.DataFrame({"Notas": [n1, n234, n567, n8, n910]}).to_excel(
            writer, sheet_name="Notas_metodologicas", index=False
        )
        for ws in writer.book.worksheets:
            ws["A1"].font = Font(bold=True)
            autosize(ws)

    br = q4.loc[q4["Codigo"] == "BR"].iloc[0]
    md = (
        f"# Respostas consolidadas\n\n"
        f"## 2) SELIC acumulada (02/01/2003-30/06/2026)\n"
        f"**{q2.iloc[0]['Taxa acumulada (%)']:.4f}%** (fator {q2.iloc[0]['Fator acumulado']:.6f})\n\n"
        f"## 4) Ranking BIS - Brasil\n"
        f"Rank **{int(br['Rank'])}** de {len(q4)} - taxa acumulada **{br['Taxa acumulada (%)']:.4f}%**\n\n"
        f"Arquivo: `{out_path.name}`\n"
    )
    (OUT / "respostas_resumo.md").write_text(md, encoding="utf-8")
    print(md)
    print("Wrote", out_path)


if __name__ == "__main__":
    main()
