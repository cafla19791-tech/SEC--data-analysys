#!/usr/bin/env python3
"""Build OSU_2025_Anexos.xlsx from the official source spreadsheet.

Reads data/osu_2025/OSU_2025_anexos_fonte.xlsx (Tab_1..Tab_8) and writes a
cleaned workbook with index and summary sheets.
"""

from __future__ import annotations

import re
from collections import OrderedDict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "osu_2025" / "OSU_2025_anexos_fonte.xlsx"
OUT = ROOT / "OSU_2025_Anexos.xlsx"

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
NORMAL = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="D6E3F0")
GROUP_FILL = PatternFill("solid", fgColor="E2EFDA")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
NUM = "#,##0.00"
PCT = "0.00%"
NOTE = Font(name="Calibri", italic=True, size=9, color="666666")

GROUP_LABELS = {
    "Benefícios Financeiros",
    "Benefícios Creditícios",
    "Benefícios Tributários",
    "TOTAL",
}

SHEET_META = [
    ("Tab_1", "T1_Nominais", "serie"),
    ("Tab_2", "T2_Constantes", "serie"),
    ("Tab_3", "T3_Comparativo", "comp"),
    ("Tab_4", "T4_Pct_PIB", "serie_pct"),
    ("Tab_5", "T5_Pct_Despesa", "serie_pct"),
    ("Tab_6", "T6_Regiao", "regiao"),
    ("Tab_7", "T7_Tributarios_Receita", "receita"),
    ("Tab_8", "T8_Variaveis", "vars"),
]

REGIONAL_RE = re.compile(
    "|".join(
        [
            r"\bSUDAM\b",
            r"\bSUDENE\b",
            r"Desenvolvimento Regional",
            r"Zona Franca",
            r"\bFDNE\b",
            r"\bFDA\b",
            r"\bFDCO\b",
            r"Fundos Constitucionais",
            r"\bFNE\b",
            r"\bFNO\b",
            r"\bFCO\b",
            r"Amazônia",
            r"\bADENE\b",
            r"\bADA\b",
            r"\bFINOR\b",
            r"\bFINAM\b",
        ]
    ),
    re.IGNORECASE,
)


def trim_row(row):
    r = list(row)
    while r and (r[-1] is None or r[-1] == ""):
        r.pop()
    return r


def is_blank(row):
    return not any(v is not None and v != "" for v in row)


def style_header_row(ws, row_idx, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, min_width=10, max_width=48, first_col_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            ws.column_dimensions[letter].width = first_col_width
            continue
        maxlen = 0
        for cell in col[:80]:
            if cell.value is not None:
                maxlen = max(maxlen, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, maxlen + 2))


def read_rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def write_serie(out_ws, rows, value_kind="number"):
    header_idx = next(i for i, row in enumerate(rows) if row and len(row) > 1 and row[1] == "Discriminação")
    title = None
    for row in rows[:header_idx]:
        for v in row:
            if isinstance(v, str) and v.startswith("Tabela"):
                title = v
                break
    out_ws["A1"] = title or out_ws.title
    out_ws["A1"].font = TITLE_FONT

    cols = trim_row(rows[header_idx])[1:]
    years = [c for c in cols[2:] if isinstance(c, (int, float))]
    for j, h in enumerate(cols, 1):
        out_ws.cell(3, j, h)
    style_header_row(out_ws, 3, 1, len(cols))
    out_ws.row_dimensions[3].height = 30

    out_row = 4
    fonte = None
    for row in rows[header_idx + 1 :]:
        if is_blank(row):
            continue
        label = row[1] if len(row) > 1 else None
        if isinstance(label, str) and label.startswith("Fontes"):
            fonte = label
            continue
        tip = row[2] if len(row) > 2 else None
        vals = row[3 : 3 + len(years)] if len(row) > 3 else []
        out_ws.cell(out_row, 1, label).font = BOLD if label in GROUP_LABELS else NORMAL
        out_ws.cell(out_row, 2, tip).font = NORMAL
        out_ws.cell(out_row, 2).alignment = Alignment(horizontal="center")
        for j, v in enumerate(vals, 3):
            cell = out_ws.cell(out_row, j, v if v != "" else None)
            cell.font = BOLD if label in GROUP_LABELS else NORMAL
            cell.border = THIN
            if isinstance(v, (int, float)):
                cell.number_format = PCT if value_kind == "pct" else NUM
        out_ws.cell(out_row, 1).border = THIN
        out_ws.cell(out_row, 2).border = THIN
        if label in GROUP_LABELS:
            fill = TOTAL_FILL if label == "TOTAL" else GROUP_FILL
            for c in range(1, 3 + len(years)):
                out_ws.cell(out_row, c).fill = fill
        elif out_row % 2 == 0:
            for c in range(1, 3 + len(years)):
                out_ws.cell(out_row, c).fill = ALT_FILL
        out_row += 1

    if fonte:
        out_ws.cell(out_row + 1, 1, fonte).font = NOTE
    autosize(out_ws, first_col_width=62)
    out_ws.freeze_panes = "C4"
    out_ws.auto_filter.ref = f"A3:{get_column_letter(2 + len(years))}{out_row - 1}"
    return years


def write_comp(out_ws, rows):
    header_idx = next(i for i, row in enumerate(rows) if row and len(row) > 1 and row[1] == "Discriminação")
    title = None
    for row in rows[:header_idx]:
        for v in row:
            if isinstance(v, str) and v.startswith("Tabela"):
                title = v
                break
    out_ws["A1"] = title or "Comparativo"
    out_ws["A1"].font = TITLE_FONT
    header = trim_row(rows[header_idx])[1:]
    for j, h in enumerate(header, 1):
        out_ws.cell(3, j, h)
    style_header_row(out_ws, 3, 1, len(header))

    out_row = 4
    fonte = None
    for row in rows[header_idx + 1 :]:
        if is_blank(row):
            continue
        label = row[1]
        if isinstance(label, str) and label.startswith(("Fontes", "Elaboração", "Nota")):
            fonte = label
            continue
        vals = row[2 : 2 + len(header) - 1]
        out_ws.cell(out_row, 1, label).font = BOLD if label in GROUP_LABELS else NORMAL
        out_ws.cell(out_row, 1).border = THIN
        for j, v in enumerate(vals, 2):
            cell = out_ws.cell(out_row, j, None if v in ("", "-") else v)
            cell.border = THIN
            cell.font = BOLD if label in GROUP_LABELS else NORMAL
            if isinstance(v, (int, float)):
                cell.number_format = PCT if j == len(header) and abs(v) < 50 else NUM
        if label in GROUP_LABELS:
            fill = TOTAL_FILL if label == "TOTAL" else GROUP_FILL
            for c in range(1, len(header) + 1):
                out_ws.cell(out_row, c).fill = fill
        out_row += 1
    if fonte:
        out_ws.cell(out_row + 1, 1, fonte).font = NOTE
    autosize(out_ws, first_col_width=62)
    out_ws.freeze_panes = "B4"
    out_ws.auto_filter.ref = f"A3:{get_column_letter(len(header))}{out_row - 1}"


def write_regiao(out_ws, rows):
    title = rows[0][1] if rows and rows[0] and len(rows[0]) > 1 else "Tabela 6"
    out_ws["A1"] = title
    out_ws["A1"].font = TITLE_FONT
    year_row, region_row = rows[2], rows[3]

    year_map = []
    current_year = None
    for col_i, val in enumerate(year_row):
        if col_i <= 1:
            continue
        if isinstance(val, (int, float)):
            current_year = int(val)
        region = region_row[col_i] if col_i < len(region_row) else None
        if region:
            year_map.append((current_year, region, col_i))

    year_cols: OrderedDict = OrderedDict()
    for idx, (year, region, src_col) in enumerate(year_map):
        year_cols.setdefault(year, []).append((idx + 2, region, src_col))

    out_ws.cell(3, 1, "Discriminação")
    for year, items in year_cols.items():
        start, end = items[0][0], items[-1][0]
        out_ws.cell(3, start, year)
        if end > start:
            out_ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        for excel_col, region, _ in items:
            out_ws.cell(4, excel_col, region)
    style_header_row(out_ws, 3, 1, 1 + len(year_map))
    style_header_row(out_ws, 4, 1, 1 + len(year_map))
    out_ws.merge_cells("A3:A4")

    out_row = 5
    flat_items = [i for items in year_cols.values() for i in items]
    for row in rows[4:]:
        if is_blank(row):
            continue
        label = row[1]
        if isinstance(label, str) and label.startswith(("Fontes", "Elaboração", "Nota")):
            out_ws.cell(out_row + 1, 1, label).font = NOTE
            continue
        out_ws.cell(out_row, 1, label).font = BOLD if label in GROUP_LABELS else NORMAL
        out_ws.cell(out_row, 1).border = THIN
        for excel_col, _region, src_col in flat_items:
            v = row[src_col] if src_col < len(row) else None
            cell = out_ws.cell(out_row, excel_col, None if v == "" else v)
            cell.border = THIN
            cell.font = BOLD if label in GROUP_LABELS else NORMAL
            if isinstance(v, (int, float)):
                cell.number_format = NUM
        if label in GROUP_LABELS:
            fill = TOTAL_FILL if label == "TOTAL" else GROUP_FILL
            for c in range(1, 2 + len(year_map)):
                out_ws.cell(out_row, c).fill = fill
        out_row += 1
    autosize(out_ws, first_col_width=55, max_width=14, min_width=11)
    out_ws.freeze_panes = "B5"


def write_receita(out_ws, rows):
    header_idx = next(i for i, row in enumerate(rows) if row and len(row) > 1 and row[1] == "Discriminação")
    title = rows[0][1] if rows[0] and len(rows[0]) > 1 else "Tabela 7"
    out_ws["A1"] = title
    out_ws["A1"].font = TITLE_FONT
    header = trim_row(rows[header_idx])[1:]
    for j, h in enumerate(header, 1):
        out_ws.cell(3, j, h)
    style_header_row(out_ws, 3, 1, len(header))
    out_row = 4
    for row in rows[header_idx + 1 :]:
        if is_blank(row):
            continue
        label = row[1]
        if isinstance(label, str) and label.startswith("Fontes"):
            out_ws.cell(out_row + 1, 1, label).font = NOTE
            continue
        out_ws.cell(out_row, 1, label).font = BOLD if label == "TOTAL" else NORMAL
        out_ws.cell(out_row, 1).border = THIN
        for j, v in enumerate(row[2 : 2 + len(header) - 1], 2):
            cell = out_ws.cell(out_row, j, None if v == "" else v)
            cell.border = THIN
            if isinstance(v, (int, float)):
                cell.number_format = NUM
            if label == "TOTAL":
                cell.fill = TOTAL_FILL
                cell.font = BOLD
        if label == "TOTAL":
            out_ws.cell(out_row, 1).fill = TOTAL_FILL
        out_row += 1
    autosize(out_ws, first_col_width=70)
    out_ws.freeze_panes = "B4"


def write_vars(out_ws, rows):
    header_idx = next(i for i, row in enumerate(rows) if row and len(row) > 1 and row[1] == "Discriminação")
    title = rows[0][1] if rows[0] and len(rows[0]) > 1 else "Tabela 8"
    out_ws["A1"] = title
    out_ws["A1"].font = TITLE_FONT
    header = trim_row(rows[header_idx])[1:]
    for j, h in enumerate(header, 1):
        out_ws.cell(3, j, h)
    style_header_row(out_ws, 3, 1, len(header))
    out_row = 4
    for row in rows[header_idx + 1 :]:
        if is_blank(row):
            continue
        label = row[1]
        if isinstance(label, str) and label.startswith("Elaboração"):
            out_ws.cell(out_row + 1, 1, label).font = NOTE
            continue
        for j, v in enumerate(row[1 : 1 + len(header)], 1):
            cell = out_ws.cell(out_row, j, None if v == "" else v)
            cell.border = THIN
            cell.font = NORMAL
            if j >= 4 and isinstance(v, (int, float)):
                cell.number_format = "0.00" if label and "IPCA" in str(label) else NUM
        out_row += 1
    autosize(out_ws, first_col_width=55, max_width=16)
    out_ws.freeze_panes = "D4"


def parse_tab1(rows):
    header_idx = next(i for i, r in enumerate(rows) if r and len(r) > 1 and r[1] == "Discriminação")
    hdr = trim_row(rows[header_idx])
    years = [int(y) for y in hdr[3:] if isinstance(y, (int, float))]
    year_to_col = {y: 3 + i for i, y in enumerate(years)}
    groups, details = {}, []
    for row in rows[header_idx + 1 :]:
        if is_blank(row):
            continue
        label = row[1]
        if isinstance(label, str) and label.startswith("Fontes"):
            continue
        tip = row[2] if len(row) > 2 else None
        series = {
            y: (row[year_to_col[y]] if year_to_col[y] < len(row) and row[year_to_col[y]] != "" else None)
            for y in years
        }
        rec = {"label": label, "tipologia": tip, **series}
        if label in GROUP_LABELS:
            groups[label] = rec
        else:
            details.append(rec)
    return years, groups, details


def build() -> Path:
    if not SRC.exists():
        raise SystemExit(f"Missing source file: {SRC}\nRun: python data/osu_2025/download_osu_2025.py")

    src = openpyxl.load_workbook(SRC, data_only=True)
    wb = Workbook()

    idx = wb.active
    idx.title = "Índice"
    idx["B2"] = "ORÇAMENTO DE SUBSÍDIOS DA UNIÃO (OSU) 2025 — ANEXOS"
    idx["B2"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    idx["B3"] = "Fonte oficial: Ministério do Planejamento e Orçamento (MPO) / SMA"
    idx["B3"].font = Font(name="Calibri", size=10, color="666666")
    idx["B4"] = "Arquivo-fonte: data/osu_2025/OSU_2025_anexos_fonte.xlsx"
    idx["B4"].font = Font(name="Calibri", size=10, color="666666")
    idx["B5"] = (
        "URL: https://www.gov.br/planejamento/pt-br/assuntos/avaliacao-de-politicas-publicas/"
        "arquivos/orcamento-de-subsidios-da-uniao/osu_2025-anexos-publicacao.xlsx"
    )
    idx["B5"].font = Font(name="Calibri", size=9, color="2F5496")
    idx["B6"] = f"Processado em: {date.today().isoformat()}"
    idx["B6"].font = Font(name="Calibri", size=10, color="666666")

    idx["B8"] = "Nº"
    idx["C8"] = "Planilha"
    idx["D8"] = "Descrição"
    style_header_row(idx, 8, 2, 4)
    descriptions = [
        (1, "T1_Nominais", "Subsídios por tipo — valores nominais (R$ mil correntes), 2003–2024"),
        (2, "T2_Constantes", "Subsídios por tipo — valores constantes (R$ mil de 2023), 2003–2024"),
        (3, "T3_Comparativo", "Comparativo 2023 × 2024 (R$ mil e variação %)"),
        (4, "T4_Pct_PIB", "Subsídios por tipo — % do PIB, 2003–2024"),
        (5, "T5_Pct_Despesa", "Subsídios por tipo — % da Despesa Primária, 2003–2024"),
        (6, "T6_Regiao", "Subsídios por tipo e região, 2011–2024"),
        (7, "T7_Tributarios_Receita", "Subsídios tributários por tipo de receita, 2003–2024"),
        (8, "T8_Variaveis", "Variáveis auxiliares: Despesa Primária, PIB e IPCA médio"),
        (9, "Resumo_Totais", "Totais por grande grupo (Financeiros, Creditícios, Tributários) — 2015–2024"),
        (10, "Resumo_Top_2024", "Maiores itens de subsídio em 2024 e variação vs 2023"),
        (11, "Sudam_Sudene_OSU", "Itens do OSU relacionados a SUDAM/SUDENE e desenvolvimento regional"),
    ]
    for i, (n, name, desc) in enumerate(descriptions):
        r = 9 + i
        idx.cell(r, 2, n).font = NORMAL
        idx.cell(r, 3, name).font = BOLD
        idx.cell(r, 4, desc).font = NORMAL
        for c in range(2, 5):
            idx.cell(r, c).border = THIN
            if i % 2 == 0:
                idx.cell(r, c).fill = ALT_FILL
    idx["B21"] = "Notas"
    idx["B21"].font = SUBTITLE_FONT
    idx["B22"] = "• Valores monetários das Tabelas 1–3, 6 e 7 estão em R$ mil, conforme publicação oficial."
    idx["B23"] = (
        "• As abas Gráfico*/Tabela* do arquivo-fonte (série 2003–2016) ficam só na fonte; "
        "esta planilha prioriza Tab_1–Tab_8 do OSU 2025."
    )
    idx["B24"] = "• Totais de grupo e TOTAL geral vêm da própria publicação."
    idx.column_dimensions["B"].width = 6
    idx.column_dimensions["C"].width = 26
    idx.column_dimensions["D"].width = 95

    tab_data = {}
    for src_name, out_name, kind in SHEET_META:
        rows = read_rows(src[src_name])
        tab_data[src_name] = rows
        ws = wb.create_sheet(out_name)
        if kind == "serie":
            write_serie(ws, rows, "number")
        elif kind == "serie_pct":
            write_serie(ws, rows, "pct")
        elif kind == "comp":
            write_comp(ws, rows)
        elif kind == "regiao":
            write_regiao(ws, rows)
        elif kind == "receita":
            write_receita(ws, rows)
        elif kind == "vars":
            write_vars(ws, rows)

    years, groups, details = parse_tab1(tab_data["Tab_1"])
    focus_years = [y for y in years if y >= 2015]
    y2023, y2024 = 2023, 2024

    resumo = wb.create_sheet("Resumo_Totais")
    resumo["A1"] = "Resumo — Totais por grande grupo (R$ mil correntes)"
    resumo["A1"].font = TITLE_FONT
    resumo["A2"] = "Valores nominais da Tabela 1. Também em R$ bilhões para leitura rápida."
    resumo["A2"].font = Font(name="Calibri", size=10, color="666666")
    headers = ["Discriminação", "Tipologia"] + focus_years
    for j, h in enumerate(headers, 1):
        resumo.cell(4, j, h)
    style_header_row(resumo, 4, 1, len(headers))
    order = ["Benefícios Financeiros", "Benefícios Creditícios", "Benefícios Tributários", "TOTAL"]
    for i, name in enumerate(order):
        r = 5 + i
        rec = groups[name]
        resumo.cell(r, 1, name).font = BOLD
        resumo.cell(r, 2, rec.get("tipologia") or ("—" if name == "TOTAL" else "")).font = NORMAL
        for j, y in enumerate(focus_years, 3):
            cell = resumo.cell(r, j, rec.get(y))
            cell.number_format = NUM
            cell.font = BOLD
            cell.border = THIN
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for c in range(1, 3 + len(focus_years)):
            resumo.cell(r, c).border = THIN
            resumo.cell(r, c).fill = fill

    resumo["A11"] = "Mesmos totais em R$ bilhões correntes"
    resumo["A11"].font = SUBTITLE_FONT
    for j, h in enumerate(headers, 1):
        resumo.cell(12, j, h)
    style_header_row(resumo, 12, 1, len(headers))
    for i, name in enumerate(order):
        r = 13 + i
        rec = groups[name]
        resumo.cell(r, 1, name).font = BOLD
        resumo.cell(r, 2, rec.get("tipologia") or ("—" if name == "TOTAL" else "")).font = NORMAL
        for j, y in enumerate(focus_years, 3):
            v = rec.get(y)
            cell = resumo.cell(r, j, (v / 1_000_000) if isinstance(v, (int, float)) else None)
            cell.number_format = "#,##0.00"
            cell.border = THIN
            cell.font = BOLD
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for c in range(1, 3 + len(focus_years)):
            resumo.cell(r, c).border = THIN
            resumo.cell(r, c).fill = fill

    resumo["A19"] = "Participação no TOTAL (%)"
    resumo["A19"].font = SUBTITLE_FONT
    for j, h in enumerate(["Discriminação"] + focus_years, 1):
        resumo.cell(20, j, h)
    style_header_row(resumo, 20, 1, 1 + len(focus_years))
    for i, name in enumerate(["Benefícios Financeiros", "Benefícios Creditícios", "Benefícios Tributários"]):
        r = 21 + i
        resumo.cell(r, 1, name).font = NORMAL
        resumo.cell(r, 1).border = THIN
        for j, y in enumerate(focus_years, 2):
            tot = groups["TOTAL"].get(y)
            v = groups[name].get(y)
            share = (v / tot) if isinstance(v, (int, float)) and isinstance(tot, (int, float)) and tot else None
            cell = resumo.cell(r, j, share)
            cell.number_format = PCT
            cell.border = THIN
    autosize(resumo, first_col_width=32, max_width=14)
    resumo.freeze_panes = "C5"

    items = []
    for rec in details:
        v24, v23 = rec.get(y2024), rec.get(y2023)
        if not isinstance(v24, (int, float)):
            continue
        var_abs = (v24 - v23) if isinstance(v23, (int, float)) else None
        var_pct = (var_abs / abs(v23)) if isinstance(var_abs, (int, float)) and v23 not in (0, None) else None
        items.append(
            {
                "label": rec["label"],
                "tipologia": rec["tipologia"],
                "v2023": v23,
                "v2024": v24,
                "var_abs": var_abs,
                "var_pct": var_pct,
            }
        )
    items.sort(key=lambda x: x["v2024"], reverse=True)
    tot2024 = groups["TOTAL"][y2024]

    top = wb.create_sheet("Resumo_Top_2024")
    top["A1"] = "Maiores itens de subsídio em 2024 (valores nominais da Tabela 1)"
    top["A1"].font = TITLE_FONT
    top["A2"] = "Exclui linhas de totalização de grupo. Valores em R$ mil e R$ bi."
    top["A2"].font = Font(name="Calibri", size=10, color="666666")
    headers = ["#", "Discriminação", "Tipologia", "2023 (R$ mil)", "2024 (R$ mil)", "2024 (R$ bi)", "Δ R$ mil", "Δ %", "% do TOTAL 2024"]
    for j, h in enumerate(headers, 1):
        top.cell(4, j, h)
    style_header_row(top, 4, 1, len(headers))
    for i, rec in enumerate(items[:40], 1):
        r = 4 + i
        top.cell(r, 1, i).border = THIN
        top.cell(r, 2, rec["label"]).border = THIN
        top.cell(r, 3, rec["tipologia"]).border = THIN
        for col, key, fmt in [(4, "v2023", NUM), (5, "v2024", NUM), (7, "var_abs", NUM), (8, "var_pct", PCT)]:
            cell = top.cell(r, col, rec[key])
            cell.number_format = fmt
            cell.border = THIN
        cell = top.cell(r, 6, rec["v2024"] / 1_000_000)
        cell.number_format = "#,##0.00"
        cell.border = THIN
        cell = top.cell(r, 9, rec["v2024"] / tot2024 if tot2024 else None)
        cell.number_format = PCT
        cell.border = THIN
        if i % 2 == 0:
            for c in range(1, 10):
                top.cell(r, c).fill = ALT_FILL

    headers2 = ["#", "Discriminação", "Tipologia", "Δ R$ mil", "Δ %", "2024 (R$ bi)"]
    top["A47"] = "Maiores altas absolutas 2023→2024"
    top["A47"].font = SUBTITLE_FONT
    for j, h in enumerate(headers2, 1):
        top.cell(48, j, h)
    style_header_row(top, 48, 1, len(headers2))
    up = sorted([x for x in items if isinstance(x["var_abs"], (int, float))], key=lambda x: x["var_abs"], reverse=True)[:15]
    for i, rec in enumerate(up, 1):
        r = 48 + i
        top.cell(r, 1, i).border = THIN
        top.cell(r, 2, rec["label"]).border = THIN
        top.cell(r, 3, rec["tipologia"]).border = THIN
        for col, key, fmt in [(4, "var_abs", NUM), (5, "var_pct", PCT)]:
            cell = top.cell(r, col, rec[key])
            cell.number_format = fmt
            cell.border = THIN
        cell = top.cell(r, 6, rec["v2024"] / 1_000_000)
        cell.number_format = "#,##0.00"
        cell.border = THIN

    top["A66"] = "Maiores quedas absolutas 2023→2024"
    top["A66"].font = SUBTITLE_FONT
    for j, h in enumerate(headers2, 1):
        top.cell(67, j, h)
    style_header_row(top, 67, 1, len(headers2))
    down = sorted([x for x in items if isinstance(x["var_abs"], (int, float))], key=lambda x: x["var_abs"])[:15]
    for i, rec in enumerate(down, 1):
        r = 67 + i
        top.cell(r, 1, i).border = THIN
        top.cell(r, 2, rec["label"]).border = THIN
        top.cell(r, 3, rec["tipologia"]).border = THIN
        for col, key, fmt in [(4, "var_abs", NUM), (5, "var_pct", PCT)]:
            cell = top.cell(r, col, rec[key])
            cell.number_format = fmt
            cell.border = THIN
        cell = top.cell(r, 6, rec["v2024"] / 1_000_000)
        cell.number_format = "#,##0.00"
        cell.border = THIN
    autosize(top, first_col_width=6)
    top.column_dimensions["B"].width = 70
    top.column_dimensions["C"].width = 12
    top.freeze_panes = "D5"

    matched = [rec for rec in details if REGIONAL_RE.search(rec["label"] or "")]
    matched.sort(key=lambda x: x.get(y2024) or 0, reverse=True)
    sud = wb.create_sheet("Sudam_Sudene_OSU")
    sud["A1"] = "Itens do OSU 2025 relacionados a SUDAM/SUDENE e desenvolvimento regional"
    sud["A1"].font = TITLE_FONT
    sud["A2"] = (
        "Filtro sobre a Tabela 1 (nominais): SUDAM, SUDENE, ZFM, FDNE/FDA/FDCO, "
        "Fundos Constitucionais (FNE/FNO/FCO), Desenvolvimento Regional."
    )
    sud["A2"].font = Font(name="Calibri", size=10, color="666666")
    headers = ["Discriminação", "Tipologia"] + focus_years + ["2024 (R$ bi)"]
    for j, h in enumerate(headers, 1):
        sud.cell(4, j, h)
    style_header_row(sud, 4, 1, len(headers))
    for i, rec in enumerate(matched):
        r = 5 + i
        sud.cell(r, 1, rec["label"]).font = NORMAL
        sud.cell(r, 1).border = THIN
        sud.cell(r, 2, rec["tipologia"]).border = THIN
        for j, y in enumerate(focus_years, 3):
            cell = sud.cell(r, j, rec.get(y) if isinstance(rec.get(y), (int, float)) else None)
            cell.number_format = NUM
            cell.border = THIN
        v = rec.get(y2024)
        cell = sud.cell(r, 3 + len(focus_years), (v / 1_000_000) if isinstance(v, (int, float)) else None)
        cell.number_format = "#,##0.00"
        cell.border = THIN
        if i % 2 == 1:
            for c in range(1, 4 + len(focus_years)):
                sud.cell(r, c).fill = ALT_FILL
    sud.cell(
        5 + len(matched) + 2,
        1,
        "Nota: “Subvenção nas Operações de Crédito Rural na área da SUDENE/SUDAM” é benefício "
        "financeiro explícito; “Desenvolvimento Regional” e “Zona Franca de Manaus…” são gastos tributários.",
    ).font = NOTE
    autosize(sud, first_col_width=70, max_width=14)
    sud.freeze_panes = "C5"

    wb.save(OUT)
    src.close()
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Saved {path} ({path.stat().st_size:,} bytes)")
