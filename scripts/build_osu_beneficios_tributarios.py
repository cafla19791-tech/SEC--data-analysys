#!/usr/bin/env python3
"""Discriminativo de benefícios tributários selecionados — 2003 a 2025.

Fonte 2003–2024:
  OSU 2025 – Anexos (MPO), aba Tab_1 (R$ mil correntes)
  https://www.gov.br/planejamento/.../osu_2025-anexos-publicacao.xlsx

Fonte 2025 (projeção):
  PLDO 2025 – Anexo IV, Quadro X (Principais Gastos Tributários)
  https://www.gov.br/planejamento/.../pldo2025-anexoiv-...-ano2025.pdf

Atualização: IPCA BCB SGS 433, índice médio do ano → jun/2026.
"""

from __future__ import annotations

import json
import urllib.request
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OSU_URL = (
    "https://www.gov.br/planejamento/pt-br/assuntos/avaliacao-de-politicas-publicas/"
    "arquivos/orcamento-de-subsidios-da-uniao/osu_2025-anexos-publicacao.xlsx/"
    "@@download/file"
)
PLDO_URL = (
    "https://www.gov.br/planejamento/pt-br/assuntos/orcamento/orcamentos-anuais/2025/"
    "pldo/4-14-1-pldo2025-anexoiv-14-renunciareceitaquadrosi_a_xxv-ano2025.pdf/"
    "@@download/file"
)
OSU_PATH = ROOT / "data" / "osu" / "osu_2025_anexos.xlsx"
OUT_XLSX = ROOT / "output" / "osu_beneficios_tributarios_2003_2025_ipca.xlsx"
OUT_MD = ROOT / "output" / "osu_beneficios_tributarios_2003_2025_ipca.md"

# Ordem e rótulos pedidos pelo usuário → nome exato no OSU/PLDO
BENEFICIOS: list[tuple[str, str]] = [
    ("Desenvolvimento Regional", "Desenvolvimento Regional"),
    (
        "Entidades Sem Fins Lucrativos - Imunes e Isentas",
        "Entidades Sem Fins Lucrativos - Imunes / Isentas",
    ),
    (
        "Pesquisas Científicas e Tecnológicas",
        "Pesquisas Científicas e Inovação Tecnológica",
    ),
    ("Informática e Automação", "Informática e Automação"),
    ("Zona Franca de Manaus", "Zona Franca de Manaus e Áreas de Livre Comércio"),
    ("Cultura e Audiovisual", "Cultura e Audiovisual"),
    ("Regime Automotivo", "Setor Automotivo"),
    ("Água Mineral", "Água Mineral"),
    ("Fundos Constitucionais", "Fundos Constitucionais"),
]

# Projeções LDO 2025 – Quadro X (R$ 1,00)
VALORES_2025: dict[str, float] = {
    "Entidades Sem Fins Lucrativos - Imunes / Isentas": 47_307_022_164.0,
    "Zona Franca de Manaus e Áreas de Livre Comércio": 30_654_355_755.0,
    "Desenvolvimento Regional": 28_302_137_256.0,
    "Setor Automotivo": 11_394_412_832.0,
    "Pesquisas Científicas e Inovação Tecnológica": 8_761_970_462.0,
    "Informática e Automação": 8_080_098_606.0,
    "Cultura e Audiovisual": 2_903_522_043.0,
    "Fundos Constitucionais": 1_808_747_247.0,
    "Água Mineral": 360_934_631.0,
}


def download_osu() -> Path:
    OSU_PATH.parent.mkdir(parents=True, exist_ok=True)
    if OSU_PATH.exists() and OSU_PATH.stat().st_size > 10_000:
        return OSU_PATH
    req = urllib.request.Request(OSU_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        OSU_PATH.write_bytes(resp.read())
    return OSU_PATH


def load_ipca() -> tuple[pd.Series, float]:
    path = ROOT / "data" / "raw" / "bcb_series" / "433_ipca.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        rows = json.loads(path.read_text())
    else:
        url = (
            "https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados"
            "?formato=json&dataInicial=01/01/1999&dataFinal=01/06/2026"
        )
        with urllib.request.urlopen(url, timeout=90) as resp:
            rows = json.loads(resp.read().decode())
        path.write_text(json.dumps(rows), encoding="utf-8")
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["data"], dayfirst=True)
    df["var"] = df["valor"].astype(float)
    df = df.sort_values("date").reset_index(drop=True)
    df["factor"] = 1.0 + df["var"] / 100.0
    df["index"] = 100.0 * df["factor"].cumprod() / df["factor"].iloc[0]
    df["year"] = df["date"].dt.year
    ann = df.groupby("year")["index"].mean()
    target = df[(df["date"].dt.year == 2026) & (df["date"].dt.month == 6)]
    if target.empty:
        raise RuntimeError("IPCA jun/2026 não encontrado")
    return ann, float(target["index"].iloc[0])


def extract_osu_series() -> dict[str, pd.Series]:
    """Retorna séries em R$ (não mil) para 2003–2024."""
    download_osu()
    wb = load_workbook(OSU_PATH, read_only=True, data_only=True)
    rows = list(wb["Tab_1"].iter_rows(values_only=True))
    wb.close()
    header = list(rows[2])
    ycols = {
        int(h): i
        for i, h in enumerate(header)
        if isinstance(h, (int, float)) and 2000 <= h <= 2035
    }
    out: dict[str, pd.Series] = {}
    for display, osu_name in BENEFICIOS:
        found = None
        for r in rows[3:]:
            if not r[1]:
                continue
            name = str(r[1]).strip()
            if osu_name == "Fundos Constitucionais":
                if name == "Fundos Constitucionais":
                    found = r
                    break
            elif name == osu_name:
                found = r
                break
        if found is None:
            raise KeyError(f"Não encontrado no OSU Tab_1: {osu_name}")
        data = {
            y: (float(found[i]) * 1_000.0 if found[i] is not None else 0.0)
            for y, i in ycols.items()
        }
        out[display] = pd.Series(data).sort_index()
    return out


def build() -> dict[str, pd.DataFrame]:
    series = extract_osu_series()
    ipca_ann, ipca_jun = load_ipca()

    records = []
    for display, osu_name in BENEFICIOS:
        s = series[display]
        for ano, val in s.items():
            records.append(
                {
                    "Beneficio": display,
                    "Nome_fonte": osu_name,
                    "Ano": int(ano),
                    "Valor_corrente_R$": float(val),
                    "Fonte": "OSU 2025 Anexos – Tab_1 (DGT Bases Efetivas)",
                    "Natureza": "gasto tributário (benefício tributário implícito)",
                }
            )
        # 2025 projeção PLDO
        v2025 = VALORES_2025[osu_name]
        records.append(
            {
                "Beneficio": display,
                "Nome_fonte": osu_name,
                "Ano": 2025,
                "Valor_corrente_R$": v2025,
                "Fonte": "PLDO 2025 Anexo IV – Quadro X (projeção)",
                "Natureza": "gasto tributário (projeção LDO 2025)",
            }
        )

    long = pd.DataFrame(records).sort_values(["Beneficio", "Ano"]).reset_index(drop=True)
    years = long["Ano"].astype(int)
    missing = sorted({y for y in years if y not in ipca_ann.index})
    if missing:
        raise RuntimeError(f"IPCA anual ausente: {missing}")
    long["Fator_IPCA_media_ano_para_jun2026"] = (
        ipca_jun / ipca_ann.loc[years].values
    )
    long["Atualizado_IPCA_30_06_2026_R$"] = (
        long["Valor_corrente_R$"] * long["Fator_IPCA_media_ano_para_jun2026"]
    )
    long["Valor_corrente_R$_bi"] = long["Valor_corrente_R$"] / 1e9
    long["Atualizado_IPCA_30_06_2026_R$_bi"] = (
        long["Atualizado_IPCA_30_06_2026_R$"] / 1e9
    )

    # Matrizes largas
    corrente = long.pivot(index="Ano", columns="Beneficio", values="Valor_corrente_R$_bi")
    corrente = corrente[[b for b, _ in BENEFICIOS]]
    corrente["TOTAL_9_beneficios"] = corrente.sum(axis=1)
    corrente = corrente.reset_index()

    ipca_w = long.pivot(
        index="Ano", columns="Beneficio", values="Atualizado_IPCA_30_06_2026_R$_bi"
    )
    ipca_w = ipca_w[[b for b, _ in BENEFICIOS]]
    ipca_w["TOTAL_9_beneficios"] = ipca_w.sum(axis=1)
    ipca_w = ipca_w.reset_index()

    # Totais por benefício
    tot = (
        long.groupby(["Beneficio", "Nome_fonte"], as_index=False)
        .agg(
            Soma_corrente_2003_2025=("Valor_corrente_R$", "sum"),
            Soma_IPCA_30_06_2026=("Atualizado_IPCA_30_06_2026_R$", "sum"),
        )
        .sort_values("Soma_corrente_2003_2025", ascending=False)
    )
    tot["Soma_corrente_bi"] = tot["Soma_corrente_2003_2025"] / 1e9
    tot["Soma_IPCA_bi"] = tot["Soma_IPCA_30_06_2026"] / 1e9

    # Aba só 2024 vs 2025
    yoy = []
    for display, osu_name in BENEFICIOS:
        v24 = float(long[(long.Beneficio == display) & (long.Ano == 2024)]["Valor_corrente_R$"].iloc[0])
        v25 = float(long[(long.Beneficio == display) & (long.Ano == 2025)]["Valor_corrente_R$"].iloc[0])
        yoy.append(
            {
                "Beneficio": display,
                "2024_corrente_R$_bi": v24 / 1e9,
                "2025_proj_R$_bi": v25 / 1e9,
                "Var_%": (v25 / v24 - 1.0) if v24 else None,
                "2024_fonte": "OSU 2025",
                "2025_fonte": "PLDO 2025 Quadro X",
            }
        )
    yoy_df = pd.DataFrame(yoy)

    metodologia = pd.DataFrame(
        [
            {"Item": "Pedido", "Valor": "Discriminativo de 9 benefícios tributários, 2003–2025"},
            {"Item": "Fonte 2003–2024", "Valor": OSU_URL},
            {"Item": "Fonte 2025", "Valor": PLDO_URL},
            {
                "Item": "Unidade OSU",
                "Valor": "Tab_1 em R$ mil correntes → convertidos para R$",
            },
            {
                "Item": "Unidade PLDO 2025",
                "Valor": "Quadro X em R$ 1,00 (projeção LDO 2025)",
            },
            {
                "Item": "Mapeamento de nomes",
                "Valor": (
                    "Regime Automotivo → Setor Automotivo; "
                    "Pesquisas Científicas e Tecnológicas → "
                    "Pesquisas Científicas e Inovação Tecnológica; "
                    "Zona Franca de Manaus → Zona Franca de Manaus e Áreas de Livre Comércio; "
                    "Entidades… → Imunes / Isentas"
                ),
            },
            {
                "Item": "Fundos Constitucionais",
                "Valor": (
                    "Linha de GASTO TRIBUTÁRIO (IOF etc.), distinta do benefício "
                    "creditício implícito FNE/FNO/FCO"
                ),
            },
            {
                "Item": "Atualização IPCA",
                "Valor": "BCB SGS 433; fator = índice jun/2026 / média anual do ano",
            },
            {
                "Item": "Soma correntes 9 benefícios 2003–2025",
                "Valor": float(long["Valor_corrente_R$"].sum()),
            },
            {
                "Item": "Soma IPCA 30/06/2026 9 benefícios 2003–2025",
                "Valor": float(long["Atualizado_IPCA_30_06_2026_R$"].sum()),
            },
            {
                "Item": "Ressalva 2025",
                "Valor": (
                    "2025 é projeção do PLDO; 2003–2024 vêm do OSU (bases efetivas). "
                    "Séries são comparáveis em conceito, mas de publicações distintas."
                ),
            },
        ]
    )

    return {
        "long": long,
        "corrente": corrente,
        "ipca": ipca_w,
        "totais": tot,
        "yoy": yoy_df,
        "metodologia": metodologia,
    }


def write_outputs(tables: dict[str, pd.DataFrame]) -> None:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    long = tables["long"]
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        long.to_excel(writer, sheet_name="Serie_longa", index=False)
        tables["corrente"].to_excel(writer, sheet_name="Corrente_R$_bi", index=False)
        tables["ipca"].to_excel(writer, sheet_name="IPCA_30_06_2026_R$_bi", index=False)
        tables["totais"].to_excel(writer, sheet_name="Totais_por_beneficio", index=False)
        tables["yoy"].to_excel(writer, sheet_name="Comparativo_2024_2025", index=False)
        tables["metodologia"].to_excel(writer, sheet_name="Metodologia", index=False)

        # Uma aba por benefício (facilita leitura)
        for display, _ in BENEFICIOS:
            sub = long[long["Beneficio"] == display][
                [
                    "Ano",
                    "Valor_corrente_R$",
                    "Valor_corrente_R$_bi",
                    "Fator_IPCA_media_ano_para_jun2026",
                    "Atualizado_IPCA_30_06_2026_R$",
                    "Atualizado_IPCA_30_06_2026_R$_bi",
                    "Fonte",
                ]
            ].copy()
            # Excel sheet name max 31 chars
            sheet = display[:31]
            sub.to_excel(writer, sheet_name=sheet, index=False)

    # Markdown resumo
    corr = tables["corrente"]
    ipca = tables["ipca"]
    lines = [
        "# Discriminativo — Benefícios tributários selecionados (2003–2025)",
        "",
        "- **2003–2024:** OSU 2025 Anexos (MPO), Tab_1",
        "- **2025:** PLDO 2025 Anexo IV, Quadro X (projeção)",
        "- **Atualização:** IPCA até 30/06/2026",
        "",
        f"Arquivo: `{OUT_XLSX.relative_to(ROOT)}`",
        "",
        "## Totais por benefício (soma 2003–2025)",
        "",
        "| Benefício | Corrente (R$ bi) | IPCA 30/06/2026 (R$ bi) |",
        "|:----------|-----------------:|------------------------:|",
    ]
    for _, r in tables["totais"].iterrows():
        lines.append(
            f"| {r['Beneficio']} | {r['Soma_corrente_bi']:.2f} | {r['Soma_IPCA_bi']:.2f} |"
        )
    lines += [
        "",
        f"| **TOTAL** | **{tables['totais']['Soma_corrente_bi'].sum():.2f}** | "
        f"**{tables['totais']['Soma_IPCA_bi'].sum():.2f}** |",
        "",
        "## Série anual — valores correntes (R$ bi)",
        "",
    ]
    # Compact table: Ano + TOTAL + each benefit abbreviated
    headers = ["Ano"] + [b for b, _ in BENEFICIOS] + ["TOTAL"]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|----:|" + "|".join(["---:"] * (len(headers) - 1)) + "|")
    for _, r in corr.iterrows():
        cells = [str(int(r["Ano"]))]
        for b, _ in BENEFICIOS:
            cells.append(f"{r[b]:.2f}")
        cells.append(f"{r['TOTAL_9_beneficios']:.2f}")
        lines.append("| " + " | ".join(cells) + " |")

    lines += [
        "",
        "## Série anual — atualizado IPCA 30/06/2026 (R$ bi)",
        "",
    ]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|----:|" + "|".join(["---:"] * (len(headers) - 1)) + "|")
    for _, r in ipca.iterrows():
        cells = [str(int(r["Ano"]))]
        for b, _ in BENEFICIOS:
            cells.append(f"{r[b]:.2f}")
        cells.append(f"{r['TOTAL_9_beneficios']:.2f}")
        lines.append("| " + " | ".join(cells) + " |")

    lines += [
        "",
        "## Comparativo 2024 (OSU) × 2025 (projeção PLDO)",
        "",
        "| Benefício | 2024 (R$ bi) | 2025 proj. (R$ bi) | Var. % |",
        "|:----------|-------------:|-------------------:|-------:|",
    ]
    for _, r in tables["yoy"].iterrows():
        var = "" if r["Var_%"] is None else f"{100*r['Var_%']:.1f}%"
        lines.append(
            f"| {r['Beneficio']} | {r['2024_corrente_R$_bi']:.2f} | "
            f"{r['2025_proj_R$_bi']:.2f} | {var} |"
        )
    lines += [
        "",
        "```bash",
        "python3 scripts/build_osu_beneficios_tributarios.py",
        "```",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")


if __name__ == "__main__":
    write_outputs(build())
