#!/usr/bin/env python3
"""Build OSU 2025 discriminativo with values updated by IPCA to latest available month.

Methodology (aligned with Renúncia Fiscal SUDAM-SUDENE in this repo):
- Annual flow of calendar year Y treated as of 31/12/Y
- Factor(Y) = IPCA(target_month) / IPCA(dez/Y)
- Target = latest published IPCA month (currently jun/2026)
- Source index: IBGE SIDRA table 1737, variable 2266 (número-índice)
"""

from __future__ import annotations

import json
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "osu_2025" / "OSU_2025_anexos_fonte.xlsx"
OUT = ROOT / "OSU_2025_Discriminativo_IPCA.xlsx"
IPCA_CACHE = ROOT / "data" / "osu_2025" / "ipca_sidra_1737.json"

SIDRA_URL = (
    "https://apisidra.ibge.gov.br/values/t/1737/n1/1/v/2266/"
    "p/200212,200312,200412,200512,200612,200712,200812,200912,201012,"
    "201112,201212,201312,201412,201512,201612,201712,201812,201912,"
    "202012,202112,202212,202312,202412,202512,"
    "202601,202602,202603,202604,202605,202606"
    "?formato=json"
)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUB = Font(name="Calibri", bold=True, size=11, color="2F5496")
NORMAL = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)
NOTE = Font(name="Calibri", italic=True, size=9, color="666666")
TOTAL_FILL = PatternFill("solid", fgColor="D6E3F0")
GROUP_FILL = PatternFill("solid", fgColor="E2EFDA")
ALT = PatternFill("solid", fgColor="F2F2F2")
NUM = "#,##0.00"
PCT = "0.0000%"
FACTOR = "0.000000"

GROUP_LABELS = {
    "Benefícios Financeiros",
    "Benefícios Creditícios",
    "Benefícios Tributários",
    "TOTAL",
}


def style_header(ws, row, start, end):
    for c in range(start, end + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, first=55, max_w=14, min_w=10):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        if col[0].column == 1:
            ws.column_dimensions[letter].width = first
            continue
        maxlen = 0
        for cell in col[:60]:
            if cell.value is not None:
                maxlen = max(maxlen, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, maxlen + 2))


def fetch_ipca() -> dict[str, float]:
    req = urllib.request.Request(SIDRA_URL, headers={"User-Agent": "SEC-data-analysys/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    idx = {}
    for row in data[1:]:
        idx[row["D3C"]] = float(row["V"])
    IPCA_CACHE.write_text(json.dumps({"fetched": date.today().isoformat(), "indices": idx}, indent=2))
    return idx


def load_or_fetch_ipca() -> dict[str, float]:
    try:
        return fetch_ipca()
    except Exception as exc:  # noqa: BLE001
        if IPCA_CACHE.exists():
            cached = json.loads(IPCA_CACHE.read_text())
            print(f"SIDRA unavailable ({exc}); using cache {IPCA_CACHE}")
            return {k: float(v) for k, v in cached["indices"].items()}
        raise


def latest_month(indices: dict[str, float]) -> tuple[str, str, float, str]:
    import calendar

    code = max(indices)
    year, month = int(code[:4]), int(code[4:])
    names = [
        "",
        "jan",
        "fev",
        "mar",
        "abr",
        "mai",
        "jun",
        "jul",
        "ago",
        "set",
        "out",
        "nov",
        "dez",
    ]
    label = f"{names[month]}/{year}"
    last = calendar.monthrange(year, month)[1]
    end_label = f"{last:02d}/{month:02d}/{year}"
    return code, label, indices[code], end_label


def parse_tab1(src_path: Path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    rows = list(wb["Tab_1"].iter_rows(values_only=True))
    wb.close()
    header_idx = next(i for i, r in enumerate(rows) if r and len(r) > 1 and r[1] == "Discriminação")
    hdr = list(rows[header_idx])
    while hdr and hdr[-1] is None:
        hdr.pop()
    years = [int(y) for y in hdr[3:] if isinstance(y, (int, float))]
    year_to_col = {y: 3 + i for i, y in enumerate(years)}
    records = []
    fonte = None
    for row in rows[header_idx + 1 :]:
        label = row[1] if len(row) > 1 else None
        if not label:
            continue
        if isinstance(label, str) and label.startswith("Fontes"):
            fonte = label
            continue
        tip = row[2] if len(row) > 2 else None
        series = {}
        for y in years:
            v = row[year_to_col[y]] if year_to_col[y] < len(row) else None
            if v == "":
                v = None
            series[y] = v
        records.append({"label": label, "tipologia": tip, "series": series})
    return years, records, fonte


def build() -> Path:
    if not SRC.exists():
        raise SystemExit(f"Missing {SRC}")

    indices = load_or_fetch_ipca()
    target_code, target_label, target_idx, target_end = latest_month(indices)

    years, records, fonte = parse_tab1(SRC)

    factors = {}
    for y in years:
        dec_code = f"{y}12"
        if dec_code not in indices:
            raise SystemExit(f"Missing IPCA dez/{y} in SIDRA extract")
        factors[y] = target_idx / indices[dec_code]

    wb = Workbook()

    # --- Índice ---
    idx = wb.active
    idx.title = "Índice"
    idx["B2"] = "OSU 2025 — Discriminativo atualizado pelo IPCA"
    idx["B2"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    idx["B3"] = f"Valores em R$ mil de {target_label} (IPCA até {target_end})"
    idx["B3"].font = Font(name="Calibri", size=11, color="2F5496")
    idx["B4"] = "Fonte dos fluxos: Orçamento de Subsídios da União 2025 — Anexos (MPO/SMA), Tabela 1 (nominais)"
    idx["B4"].font = Font(name="Calibri", size=10, color="666666")
    idx["B5"] = "Fonte do índice: IBGE SIDRA tabela 1737, variável 2266 (número-índice do IPCA)"
    idx["B5"].font = Font(name="Calibri", size=10, color="666666")
    idx["B6"] = (
        f"Metodologia: fator(ano Y) = IPCA({target_label}) / IPCA(dez/Y). "
        "Fluxo do ano-calendário tratado como posição em 31/12/Y "
        "(mesmo critério da planilha Renúncia Fiscal SUDAM-SUDENE)."
    )
    idx["B6"].font = Font(name="Calibri", size=10, color="666666")
    idx["B7"] = f"Processado em: {date.today().isoformat()} | IPCA alvo: {target_idx:.2f} ({target_code})"
    idx["B7"].font = Font(name="Calibri", size=10, color="666666")

    idx["B9"] = "Nº"
    idx["C9"] = "Planilha"
    idx["D9"] = "Descrição"
    style_header(idx, 9, 2, 4)
    sheets_desc = [
        (1, "IPCA_Fatores", f"Número-índice de dez/Y e fator de atualização até {target_label}"),
        (2, "Discriminativo_IPCA", f"Todos os itens × anos 2003–2024 em R$ mil de {target_label}"),
        (3, "Discriminativo_Nominal", "Mesmos itens em valores nominais originais (R$ mil correntes)"),
        (4, "Resumo_Grupos", "Totais Financeiros / Creditícios / Tributários — nominal e IPCA"),
        (5, "Top_2024_IPCA", f"Ranking 2024 em valores de {target_label} e variação vs 2023"),
        (6, "Acumulado_IPCA", f"Soma 2003–2024 de cada item em R$ mil / R$ bi de {target_label}"),
    ]
    for i, (n, name, desc) in enumerate(sheets_desc):
        r = 10 + i
        idx.cell(r, 2, n).border = THIN
        idx.cell(r, 3, name).font = BOLD
        idx.cell(r, 3).border = THIN
        idx.cell(r, 4, desc).border = THIN
        if i % 2 == 0:
            for c in range(2, 5):
                idx.cell(r, c).fill = ALT
    idx["B18"] = "Notas"
    idx["B18"].font = SUB
    idx["B19"] = (
        f"• O IPCA de jul/2026 ainda não havia sido divulgado na data do processamento "
        f"(divulgação prevista em 11/08/2026); usa-se {target_label}."
    )
    idx["B20"] = "• Valores monetários estão em R$ mil (milhares de reais), como na publicação oficial."
    idx["B21"] = "• Células vazias indicam ausência do programa/benefício naquele ano na Tabela 1."
    idx.column_dimensions["B"].width = 6
    idx.column_dimensions["C"].width = 26
    idx.column_dimensions["D"].width = 90

    # --- IPCA_Fatores ---
    fat = wb.create_sheet("IPCA_Fatores")
    fat["A1"] = f"Fatores IPCA até {target_end}"
    fat["A1"].font = TITLE
    fat["A2"] = f"IPCA alvo ({target_label}): {target_idx:.2f} | Fator = IPCA_alvo / IPCA_dezembro_do_ano"
    fat["A2"].font = Font(name="Calibri", size=10, color="666666")
    headers = [
        "Ano-Calendário",
        "Data-base",
        f"IPCA dez/ano",
        f"IPCA {target_label}",
        f"Fator até {target_end}",
        "Fonte/observação",
    ]
    for j, h in enumerate(headers, 1):
        fat.cell(4, j, h)
    style_header(fat, 4, 1, len(headers))
    for i, y in enumerate(years):
        r = 5 + i
        dec = indices[f"{y}12"]
        fat.cell(r, 1, y).border = THIN
        fat.cell(r, 2, f"31/12/{y}").border = THIN
        c = fat.cell(r, 3, dec)
        c.number_format = "0.00"
        c.border = THIN
        c = fat.cell(r, 4, target_idx)
        c.number_format = "0.00"
        c.border = THIN
        c = fat.cell(r, 5, factors[y])
        c.number_format = FACTOR
        c.border = THIN
        fat.cell(
            r,
            6,
            f"Atualização de dez/{y} até {target_label} (SIDRA 1737)",
        ).border = THIN
        if i % 2 == 1:
            for c in range(1, 7):
                fat.cell(r, c).fill = ALT
    fat.cell(5 + len(years) + 1, 1, "Série mensal recente (2026):").font = SUB
    r0 = 5 + len(years) + 2
    for j, h in enumerate(["Código", "Mês", "Número-índice"], 1):
        fat.cell(r0, j, h)
    style_header(fat, r0, 1, 3)
    recent = sorted(k for k in indices if k.startswith("2026"))
    for i, code in enumerate(recent):
        r = r0 + 1 + i
        fat.cell(r, 1, code).border = THIN
        y, m = int(code[:4]), int(code[4:])
        names = ["", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
        fat.cell(r, 2, f"{names[m]}/{y}").border = THIN
        c = fat.cell(r, 3, indices[code])
        c.number_format = "0.00"
        c.border = THIN
    autosize(fat, first=16, max_w=55, min_w=12)
    fat.column_dimensions["F"].width = 55

    def write_matrix(ws, title, subtitle, value_fn, money=True):
        ws["A1"] = title
        ws["A1"].font = TITLE
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", size=10, color="666666")
        headers = ["Discriminação", "Tipologia"] + years + ["Acumulado 2003-2024"]
        for j, h in enumerate(headers, 1):
            ws.cell(4, j, h)
        style_header(ws, 4, 1, len(headers))
        ws.row_dimensions[4].height = 30
        out_row = 5
        for rec in records:
            label = rec["label"]
            tip = rec["tipologia"]
            ws.cell(out_row, 1, label).font = BOLD if label in GROUP_LABELS else NORMAL
            ws.cell(out_row, 1).border = THIN
            ws.cell(out_row, 2, tip).font = NORMAL
            ws.cell(out_row, 2).border = THIN
            ws.cell(out_row, 2).alignment = Alignment(horizontal="center")
            acc = 0.0
            has_acc = False
            for j, y in enumerate(years, 3):
                raw = rec["series"].get(y)
                val = value_fn(raw, y) if isinstance(raw, (int, float)) else None
                cell = ws.cell(out_row, j, val)
                cell.border = THIN
                cell.font = BOLD if label in GROUP_LABELS else NORMAL
                if money and isinstance(val, (int, float)):
                    cell.number_format = NUM
                    acc += val
                    has_acc = True
            cell = ws.cell(out_row, 3 + len(years), acc if has_acc else None)
            cell.border = THIN
            cell.font = BOLD
            if money:
                cell.number_format = NUM
            if label in GROUP_LABELS:
                fill = TOTAL_FILL if label == "TOTAL" else GROUP_FILL
                for c in range(1, 4 + len(years)):
                    ws.cell(out_row, c).fill = fill
            elif out_row % 2 == 0:
                for c in range(1, 4 + len(years)):
                    ws.cell(out_row, c).fill = ALT
            out_row += 1
        if fonte:
            ws.cell(out_row + 1, 1, fonte).font = NOTE
        autosize(ws, first=62, max_w=13)
        ws.freeze_panes = "C5"
        ws.auto_filter.ref = f"A4:{get_column_letter(3 + len(years))}{out_row - 1}"

    write_matrix(
        wb.create_sheet("Discriminativo_IPCA"),
        f"Discriminativo — valores atualizados pelo IPCA até {target_end}",
        f"R$ mil de {target_label}. Fator(Y) = IPCA({target_label}) / IPCA(dez/Y).",
        lambda raw, y: raw * factors[y],
    )
    write_matrix(
        wb.create_sheet("Discriminativo_Nominal"),
        "Discriminativo — valores nominais (Tabela 1 OSU 2025)",
        "R$ mil correntes, sem correção. Incluído para confronto com a aba IPCA.",
        lambda raw, y: raw,
    )

    # --- Resumo_Grupos ---
    res = wb.create_sheet("Resumo_Grupos")
    res["A1"] = f"Resumo por grande grupo — nominal × IPCA ({target_label})"
    res["A1"].font = TITLE
    groups = {r["label"]: r for r in records if r["label"] in GROUP_LABELS}
    order = ["Benefícios Financeiros", "Benefícios Creditícios", "Benefícios Tributários", "TOTAL"]

    res["A3"] = "Valores nominais (R$ bilhões correntes)"
    res["A3"].font = SUB
    headers = ["Discriminação"] + years
    for j, h in enumerate(headers, 1):
        res.cell(4, j, h)
    style_header(res, 4, 1, len(headers))
    for i, name in enumerate(order):
        r = 5 + i
        res.cell(r, 1, name).font = BOLD
        res.cell(r, 1).border = THIN
        for j, y in enumerate(years, 2):
            v = groups[name]["series"].get(y)
            cell = res.cell(r, j, (v / 1_000_000) if isinstance(v, (int, float)) else None)
            cell.number_format = "#,##0.00"
            cell.border = THIN
            cell.font = BOLD
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for c in range(1, 2 + len(years)):
            res.cell(r, c).fill = fill

    res["A11"] = f"Valores atualizados pelo IPCA (R$ bilhões de {target_label})"
    res["A11"].font = SUB
    for j, h in enumerate(headers, 1):
        res.cell(12, j, h)
    style_header(res, 12, 1, len(headers))
    for i, name in enumerate(order):
        r = 13 + i
        res.cell(r, 1, name).font = BOLD
        res.cell(r, 1).border = THIN
        for j, y in enumerate(years, 2):
            v = groups[name]["series"].get(y)
            cell = res.cell(
                r,
                j,
                (v * factors[y] / 1_000_000) if isinstance(v, (int, float)) else None,
            )
            cell.number_format = "#,##0.00"
            cell.border = THIN
            cell.font = BOLD
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for c in range(1, 2 + len(years)):
            res.cell(r, c).fill = fill

    # Accumulated totals
    res["A19"] = f"Acumulado 2003–2024 (R$ bilhões de {target_label})"
    res["A19"].font = SUB
    for j, h in enumerate(["Discriminação", "Acumulado IPCA", "% do TOTAL"], 1):
        res.cell(20, j, h)
    style_header(res, 20, 1, 3)
    tot_acc = sum(
        (groups["TOTAL"]["series"][y] * factors[y] / 1_000_000)
        for y in years
        if isinstance(groups["TOTAL"]["series"].get(y), (int, float))
    )
    for i, name in enumerate(order):
        r = 21 + i
        acc = sum(
            (groups[name]["series"][y] * factors[y] / 1_000_000)
            for y in years
            if isinstance(groups[name]["series"].get(y), (int, float))
        )
        res.cell(r, 1, name).font = BOLD
        res.cell(r, 1).border = THIN
        c = res.cell(r, 2, acc)
        c.number_format = "#,##0.00"
        c.border = THIN
        c.font = BOLD
        c = res.cell(r, 3, acc / tot_acc if tot_acc else None)
        c.number_format = PCT
        c.border = THIN
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for col in range(1, 4):
            res.cell(r, col).fill = fill

    # Latest year highlight
    res["A27"] = f"2024 em valores de {target_label}"
    res["A27"].font = SUB
    for j, h in enumerate(
        ["Discriminação", "2024 nominal (R$ bi)", f"2024 IPCA (R$ bi)", "Fator dez/2024"],
        1,
    ):
        res.cell(28, j, h)
    style_header(res, 28, 1, 4)
    for i, name in enumerate(order):
        r = 29 + i
        v = groups[name]["series"].get(2024)
        res.cell(r, 1, name).font = BOLD
        res.cell(r, 1).border = THIN
        c = res.cell(r, 2, v / 1_000_000 if isinstance(v, (int, float)) else None)
        c.number_format = "#,##0.00"
        c.border = THIN
        c = res.cell(r, 3, v * factors[2024] / 1_000_000 if isinstance(v, (int, float)) else None)
        c.number_format = "#,##0.00"
        c.border = THIN
        c = res.cell(r, 4, factors[2024])
        c.number_format = FACTOR
        c.border = THIN
    autosize(res, first=32, max_w=12)
    res.freeze_panes = "B5"

    # --- Top_2024_IPCA ---
    top = wb.create_sheet("Top_2024_IPCA")
    top["A1"] = f"Maiores itens em 2024 — valores de {target_label}"
    top["A1"].font = TITLE
    top["A2"] = "Exclui linhas de totalização de grupo. Compara também o valor nominal original."
    top["A2"].font = Font(name="Calibri", size=10, color="666666")
    details = [r for r in records if r["label"] not in GROUP_LABELS]
    items = []
    for rec in details:
        v24 = rec["series"].get(2024)
        v23 = rec["series"].get(2023)
        if not isinstance(v24, (int, float)):
            continue
        ipca24 = v24 * factors[2024]
        ipca23 = v23 * factors[2023] if isinstance(v23, (int, float)) else None
        var = (ipca24 - ipca23) if ipca23 is not None else None
        var_pct = (var / abs(ipca23)) if isinstance(var, (int, float)) and ipca23 else None
        items.append(
            {
                "label": rec["label"],
                "tip": rec["tipologia"],
                "nom24": v24,
                "ipca24": ipca24,
                "ipca23": ipca23,
                "var": var,
                "var_pct": var_pct,
            }
        )
    items.sort(key=lambda x: x["ipca24"], reverse=True)
    tot_ipca24 = groups["TOTAL"]["series"][2024] * factors[2024]
    headers = [
        "#",
        "Discriminação",
        "Tipologia",
        "2024 nominal (R$ mil)",
        f"2024 IPCA (R$ mil)",
        f"2024 IPCA (R$ bi)",
        f"2023 IPCA (R$ mil)",
        "Δ IPCA (R$ mil)",
        "Δ %",
        "% do TOTAL IPCA 2024",
    ]
    for j, h in enumerate(headers, 1):
        top.cell(4, j, h)
    style_header(top, 4, 1, len(headers))
    for i, rec in enumerate(items[:50], 1):
        r = 4 + i
        top.cell(r, 1, i).border = THIN
        top.cell(r, 2, rec["label"]).border = THIN
        top.cell(r, 3, rec["tip"]).border = THIN
        for col, key in [(4, "nom24"), (5, "ipca24"), (7, "ipca23"), (8, "var")]:
            cell = top.cell(r, col, rec[key])
            cell.number_format = NUM
            cell.border = THIN
        cell = top.cell(r, 6, rec["ipca24"] / 1_000_000)
        cell.number_format = "#,##0.00"
        cell.border = THIN
        cell = top.cell(r, 9, rec["var_pct"])
        cell.number_format = "0.00%"
        cell.border = THIN
        cell = top.cell(r, 10, rec["ipca24"] / tot_ipca24 if tot_ipca24 else None)
        cell.number_format = "0.00%"
        cell.border = THIN
        if i % 2 == 0:
            for c in range(1, 11):
                top.cell(r, c).fill = ALT
    autosize(top, first=6, max_w=14)
    top.column_dimensions["B"].width = 70
    top.column_dimensions["C"].width = 12
    top.freeze_panes = "D5"

    # --- Acumulado_IPCA ---
    acc_ws = wb.create_sheet("Acumulado_IPCA")
    acc_ws["A1"] = f"Acumulado 2003–2024 em valores de {target_label}"
    acc_ws["A1"].font = TITLE
    acc_ws["A2"] = (
        "Soma dos fluxos anuais de cada item, cada ano atualizado pelo fator IPCA "
        f"correspondente até {target_end}."
    )
    acc_ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    headers = [
        "#",
        "Discriminação",
        "Tipologia",
        f"Acumulado (R$ mil de {target_label})",
        f"Acumulado (R$ bi)",
        "% do TOTAL acumulado",
    ]
    for j, h in enumerate(headers, 1):
        acc_ws.cell(4, j, h)
    style_header(acc_ws, 4, 1, len(headers))

    acc_items = []
    for rec in details:
        total = 0.0
        anyv = False
        for y in years:
            v = rec["series"].get(y)
            if isinstance(v, (int, float)):
                total += v * factors[y]
                anyv = True
        if anyv:
            acc_items.append({"label": rec["label"], "tip": rec["tipologia"], "acc": total})
    acc_items.sort(key=lambda x: x["acc"], reverse=True)
    tot_acc_detail = sum(
        groups["TOTAL"]["series"][y] * factors[y]
        for y in years
        if isinstance(groups["TOTAL"]["series"].get(y), (int, float))
    )
    for i, rec in enumerate(acc_items, 1):
        r = 4 + i
        acc_ws.cell(r, 1, i).border = THIN
        acc_ws.cell(r, 2, rec["label"]).border = THIN
        acc_ws.cell(r, 3, rec["tip"]).border = THIN
        c = acc_ws.cell(r, 4, rec["acc"])
        c.number_format = NUM
        c.border = THIN
        c = acc_ws.cell(r, 5, rec["acc"] / 1_000_000)
        c.number_format = "#,##0.00"
        c.border = THIN
        c = acc_ws.cell(r, 6, rec["acc"] / tot_acc_detail if tot_acc_detail else None)
        c.number_format = "0.00%"
        c.border = THIN
        if i % 2 == 0:
            for c in range(1, 7):
                acc_ws.cell(r, c).fill = ALT

    # group totals at bottom
    r = 4 + len(acc_items) + 2
    acc_ws.cell(r, 1, "Totais de grupo").font = SUB
    for j, h in enumerate(["Discriminação", f"Acumulado (R$ bi de {target_label})", "%"], 1):
        acc_ws.cell(r + 1, j, h)
    style_header(acc_ws, r + 1, 1, 3)
    for i, name in enumerate(order):
        rr = r + 2 + i
        acc = sum(
            groups[name]["series"][y] * factors[y] / 1_000_000
            for y in years
            if isinstance(groups[name]["series"].get(y), (int, float))
        )
        acc_ws.cell(rr, 1, name).font = BOLD
        acc_ws.cell(rr, 1).border = THIN
        c = acc_ws.cell(rr, 2, acc)
        c.number_format = "#,##0.00"
        c.border = THIN
        c = acc_ws.cell(rr, 3, acc / (tot_acc_detail / 1_000_000) if tot_acc_detail else None)
        c.number_format = "0.00%"
        c.border = THIN
        fill = TOTAL_FILL if name == "TOTAL" else GROUP_FILL
        for col in range(1, 4):
            acc_ws.cell(rr, col).fill = fill

    autosize(acc_ws, first=6, max_w=16)
    acc_ws.column_dimensions["B"].width = 70
    acc_ws.column_dimensions["C"].width = 12
    acc_ws.freeze_panes = "D5"

    wb.save(OUT)

    # Print validation
    t24_nom = groups["TOTAL"]["series"][2024] / 1e6
    t24_ipca = groups["TOTAL"]["series"][2024] * factors[2024] / 1e6
    t_acc = tot_acc_detail / 1e6
    print(f"Target IPCA: {target_label} = {target_idx:.2f}")
    print(f"Fator 2024: {factors[2024]:.6f}")
    print(f"Fator 2015: {factors[2015]:.6f}")
    print(f"TOTAL 2024 nominal R$ bi: {t24_nom:.2f}")
    print(f"TOTAL 2024 IPCA R$ bi: {t24_ipca:.2f}")
    print(f"TOTAL acumulado 2003-2024 IPCA R$ bi: {t_acc:.2f}")
    print(f"Top 3 2024 IPCA:")
    for rec in items[:3]:
        print(f"  {rec['label']}: {rec['ipca24']/1e6:.2f} bi")
    print(f"Saved {OUT} ({OUT.stat().st_size:,} bytes)")
    return OUT


if __name__ == "__main__":
    # Fix latest_month return annotation usage
    build()
