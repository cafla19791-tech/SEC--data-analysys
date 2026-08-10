#!/usr/bin/env python3
"""Extrai do OSU 2025 (anexos) a série de subsídios dos Fundos Constitucionais.

Fonte:
  https://www.gov.br/planejamento/pt-br/assuntos/avaliacao-de-politicas-publicas/
  arquivos/orcamento-de-subsidios-da-uniao/osu_2025-anexos-publicacao.xlsx

Abas usadas:
  - Tab_1: valores nominais (R$ mil correntes), 2003–2024
  - Tab_2: valores constantes (R$ mil de 2023)
  - Tab_3: comparativo 2023–2024
  - Tab_6: rateio regional (Norte/Nordeste/Centro-Oeste)

Importante: a linha é agregada **FNE + FNO + FCO** (benefício creditício implícito),
não o volume contratado. Em Tab_6 o MPO aplica chaves fixas 20%/60%/20%
(Norte/Nordeste/Centro-Oeste); a coluna Norte é proxy imputado do FNO no
*custo de subsídio*, não a contratação do BASA.
"""

from __future__ import annotations

import json
import urllib.request
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC_URL = (
    "https://www.gov.br/planejamento/pt-br/assuntos/avaliacao-de-politicas-publicas/"
    "arquivos/orcamento-de-subsidios-da-uniao/osu_2025-anexos-publicacao.xlsx/"
    "@@download/file"
)
SRC_PATH = ROOT / "data" / "osu" / "osu_2025_anexos.xlsx"
OUT_XLSX = ROOT / "output" / "osu_2025_fundos_constitucionais.xlsx"
OUT_MD = ROOT / "output" / "osu_2025_fundos_constitucionais.md"

LABEL_FC = "Fundos Constitucionais de Financiamento - FNE, FNO e FCO"
LABEL_GT = "Fundos Constitucionais"  # gasto tributário (outra linha)


def download_source() -> Path:
    SRC_PATH.parent.mkdir(parents=True, exist_ok=True)
    if SRC_PATH.exists() and SRC_PATH.stat().st_size > 10_000:
        return SRC_PATH
    req = urllib.request.Request(SRC_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        SRC_PATH.write_bytes(resp.read())
    return SRC_PATH


def _year_cols(header: list) -> dict[int, int]:
    return {
        int(h): i
        for i, h in enumerate(header)
        if isinstance(h, (int, float)) and 2000 <= h <= 2035
    }


def _find_row(rows: list, label: str):
    for r in rows:
        if r[1] and str(r[1]).strip() == label:
            return r
    for r in rows:
        if r[1] and label in str(r[1]):
            return r
    raise KeyError(label)


def load_tab_series(wb, sheet: str, label: str) -> pd.Series:
    rows = list(wb[sheet].iter_rows(values_only=True))
    header = list(rows[2])
    ycols = _year_cols(header)
    row = _find_row(rows, label)
    data = {y: float(row[i]) if row[i] is not None else None for y, i in ycols.items()}
    s = pd.Series(data, name=label).sort_index()
    s.index.name = "Ano"
    return s


def load_tab6_regional(wb, label: str) -> pd.DataFrame:
    rows = list(wb["Tab_6"].iter_rows(values_only=True))
    year_row, reg_row = rows[2], rows[3]
    cols: dict[tuple[int, str], int] = {}
    current_year: int | None = None
    for i, (y, reg) in enumerate(zip(year_row, reg_row)):
        if isinstance(y, (int, float)) and 2000 <= y <= 2035:
            current_year = int(y)
        if current_year is not None and reg not in (None, ""):
            cols[(current_year, str(reg))] = i
    row = _find_row(rows, label)
    records = []
    years = sorted({y for y, _ in cols})
    for y in years:
        rec = {"Ano": y}
        for reg in ("Norte", "Nordeste", "Centro-Oeste", "Sudeste", "Sul", "Total"):
            rec[reg] = float(row[cols[(y, reg)]])
        tot = rec["Total"]
        rec["Share_Norte"] = rec["Norte"] / tot if tot else None
        rec["Share_Nordeste"] = rec["Nordeste"] / tot if tot else None
        rec["Share_CentroOeste"] = rec["Centro-Oeste"] / tot if tot else None
        records.append(rec)
    return pd.DataFrame(records)


def load_ipca_jun2026() -> tuple[pd.Series, float]:
    path = ROOT / "data" / "raw" / "bcb_series" / "433_ipca.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        rows = json.loads(path.read_text())
    else:
        url = (
            "https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados"
            "?formato=json&dataInicial=01/01/1999&dataFinal=01/06/2026"
        )
        with urllib.request.urlopen(url, timeout=90) as resp:
            rows = json.loads(resp.read().decode())
        path.write_text(json.dumps(rows), encoding="utf-8")
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["data"], dayfirst=True)
    df["var"] = df["valor"].astype(float)
    df = df.sort_values("date").reset_index(drop=True)
    df["factor"] = 1.0 + df["var"] / 100.0
    df["index"] = 100.0 * df["factor"].cumprod() / df["factor"].iloc[0]
    df["year"] = df["date"].dt.year
    ann = df.groupby("year")["index"].mean()
    target = df[(df["date"].dt.year == 2026) & (df["date"].dt.month == 6)]
    if target.empty:
        raise RuntimeError("IPCA jun/2026 não encontrado")
    return ann, float(target["index"].iloc[0])


def build() -> dict[str, pd.DataFrame]:
    download_source()
    wb = load_workbook(SRC_PATH, read_only=True, data_only=True)

    corr_mil = load_tab_series(wb, "Tab_1", LABEL_FC)  # R$ mil
    const_mil = load_tab_series(wb, "Tab_2", LABEL_FC)
    gt_mil = load_tab_series(wb, "Tab_1", LABEL_GT)
    reg = load_tab6_regional(wb, LABEL_FC)

    # Tab_3 comparativo
    rows = list(wb["Tab_3"].iter_rows(values_only=True))
    r3 = _find_row(rows, LABEL_FC)
    tab3 = pd.DataFrame(
        [
            {
                "Discriminação": LABEL_FC,
                "2023_R$_mil": float(r3[2]),
                "2024_R$_mil": float(r3[3]),
                "Variacao_R$_mil": float(r3[4]),
                "Variacao_%": float(r3[5]),
            }
        ]
    )

    tip = None
    for r in wb["Tab_1"].iter_rows(values_only=True):
        if r[1] and LABEL_FC in str(r[1]):
            tip = r[2]
            break
    wb.close()

    ipca_ann, ipca_jun = load_ipca_jun2026()
    years = corr_mil.index.astype(int)
    fator = ipca_jun / ipca_ann.loc[years]

    serie = pd.DataFrame(
        {
            "Ano": years,
            "Subsídio_FNE_FNO_FCO_corrente_R$": corr_mil.values * 1_000.0,
            "Subsídio_FNE_FNO_FCO_corrente_R$_mil": corr_mil.values,
            "Subsídio_constante_OSU_2023_R$": const_mil.reindex(years).values * 1_000.0,
            "Subsídio_constante_OSU_2023_R$_mil": const_mil.reindex(years).values,
            "Fator_IPCA_media_ano_para_jun2026": fator.values,
            "Atualizado_IPCA_30_06_2026_R$": corr_mil.values * 1_000.0 * fator.values,
            "Tipologia_OSU": tip,
            "Fonte": "OSU 2025 – Anexos (MPO), Tab_1/Tab_2",
        }
    )

    bi = pd.DataFrame(
        {
            "Ano": serie["Ano"],
            "Corrente (R$ bi)": serie["Subsídio_FNE_FNO_FCO_corrente_R$"] / 1e9,
            "Constante OSU 2023 (R$ bi)": serie["Subsídio_constante_OSU_2023_R$"] / 1e9,
            "IPCA 30/06/2026 (R$ bi)": serie["Atualizado_IPCA_30_06_2026_R$"] / 1e9,
        }
    )

    # Proxy FNO = share Norte (20% fixo no OSU)
    norte = reg.set_index("Ano")["Norte"] * 1_000.0
    proxy = pd.DataFrame(
        {
            "Ano": reg["Ano"],
            "Proxy_FNO_subsídio_Norte_R$": norte.values,
            "Proxy_FNO_subsídio_Norte_R$_bi": norte.values / 1e9,
            "Share_Norte_no_total": reg["Share_Norte"].values,
            "Total_FNE_FNO_FCO_R$_bi": (reg["Total"] * 1_000.0 / 1e9).values,
            "Observacao": (
                "Rateio OSU Tab_6 com chave fixa 20% Norte / 60% Nordeste / "
                "20% Centro-Oeste — não é volume contratado do FNO."
            ),
        }
    )
    proxy["Atualizado_IPCA_30_06_2026_R$"] = (
        proxy["Proxy_FNO_subsídio_Norte_R$"]
        * (ipca_jun / ipca_ann.loc[proxy["Ano"].astype(int)].values)
    )

    gt = pd.DataFrame(
        {
            "Ano": gt_mil.index.astype(int),
            "Gasto_tributario_Fundos_Constitucionais_R$_mil": gt_mil.values,
            "Gasto_tributario_Fundos_Constitucionais_R$": gt_mil.values * 1_000.0,
            "Nota": (
                "Linha distinta (gasto tributário), não confundir com o "
                "benefício creditício implícito FNE/FNO/FCO."
            ),
        }
    )

    reg_out = reg.copy()
    for c in ("Norte", "Nordeste", "Centro-Oeste", "Sudeste", "Sul", "Total"):
        reg_out[f"{c}_R$"] = reg_out[c] * 1_000.0
        reg_out[f"{c}_R$_bi"] = reg_out[c] / 1e6

    metodologia = pd.DataFrame(
        [
            {"Item": "Documento", "Valor": SRC_URL},
            {
                "Item": "Conceito da linha principal",
                "Valor": (
                    "Benefício creditício implícito dos Fundos Constitucionais "
                    "de Financiamento (FNE + FNO + FCO), tipificado como Implícito "
                    "no OSU — custo de subsídio, não contratação."
                ),
            },
            {
                "Item": "Unidade original",
                "Valor": "R$ mil correntes (Tab_1); R$ mil de 2023 (Tab_2)",
            },
            {
                "Item": "Proxy FNO",
                "Valor": (
                    "Coluna Norte da Tab_6 (= 20% do total em todos os anos "
                    "2011–2024). Usar só como rateio do subsídio agregado."
                ),
            },
            {
                "Item": "Atualização adicional",
                "Valor": "IPCA BCB SGS 433, média anual → jun/2026 (opcional).",
            },
            {
                "Item": "Soma correntes 2003–2024 (FNE+FNO+FCO)",
                "Valor": float(serie["Subsídio_FNE_FNO_FCO_corrente_R$"].sum()),
            },
            {
                "Item": "Soma IPCA 30/06/2026 2003–2024",
                "Valor": float(serie["Atualizado_IPCA_30_06_2026_R$"].sum()),
            },
            {
                "Item": "2023 e 2024 (correntes)",
                "Valor": (
                    f"2023 = R$ {serie.loc[serie.Ano==2023, 'Subsídio_FNE_FNO_FCO_corrente_R$'].iloc[0]/1e9:.3f} bi; "
                    f"2024 = R$ {serie.loc[serie.Ano==2024, 'Subsídio_FNE_FNO_FCO_corrente_R$'].iloc[0]/1e9:.3f} bi"
                ),
            },
        ]
    )

    return {
        "serie": serie,
        "bi": bi,
        "proxy_fno": proxy,
        "regional": reg_out,
        "gasto_tributario": gt,
        "tab3": tab3,
        "metodologia": metodologia,
    }


def write_outputs(tables: dict[str, pd.DataFrame]) -> None:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        tables["serie"].to_excel(writer, sheet_name="FNE_FNO_FCO_subsidio", index=False)
        tables["bi"].to_excel(writer, sheet_name="Serie_R$_bi", index=False)
        tables["proxy_fno"].to_excel(writer, sheet_name="Proxy_FNO_Norte_20pct", index=False)
        tables["regional"].to_excel(writer, sheet_name="Tab6_regional", index=False)
        tables["gasto_tributario"].to_excel(
            writer, sheet_name="GT_Fundos_Constitucionais", index=False
        )
        tables["tab3"].to_excel(writer, sheet_name="Comparativo_2023_2024", index=False)
        tables["metodologia"].to_excel(writer, sheet_name="Metodologia", index=False)

    bi = tables["bi"]
    lines = [
        "# OSU 2025 — Subsídios dos Fundos Constitucionais (FNE + FNO + FCO)",
        "",
        f"Fonte: [Anexos OSU 2025 (MPO)]({SRC_URL})",
        "",
        f"Arquivo: `{OUT_XLSX.relative_to(ROOT)}`",
        "",
        "**Conceito:** benefício creditício implícito (custo de subsídio), "
        "não volume contratado.",
        "",
        "## Série agregada FNE+FNO+FCO",
        "",
        "| Ano | Corrente (R$ bi) | Constante OSU 2023 (R$ bi) | IPCA 30/06/2026 (R$ bi) |",
        "|----:|-----------------:|--------------------------:|------------------------:|",
    ]
    for _, r in bi.iterrows():
        lines.append(
            f"| {int(r['Ano'])} | {r['Corrente (R$ bi)']:.3f} | "
            f"{r['Constante OSU 2023 (R$ bi)']:.3f} | "
            f"{r['IPCA 30/06/2026 (R$ bi)']:.3f} |"
        )
    proxy = tables["proxy_fno"]
    lines += [
        "",
        f"Soma 2003–2024 (correntes): R$ {bi['Corrente (R$ bi)'].sum():.2f} bi",
        f"Soma 2003–2024 (IPCA 30/06/2026): R$ {bi['IPCA 30/06/2026 (R$ bi)'].sum():.2f} bi",
        "",
        "## Proxy FNO (Norte = 20% fixo na Tab_6 do OSU)",
        "",
        "| Ano | Proxy FNO subsídio (R$ bi) | IPCA 30/06/2026 (R$ bi) |",
        "|----:|---------------------------:|------------------------:|",
    ]
    for _, r in proxy.iterrows():
        lines.append(
            f"| {int(r['Ano'])} | {r['Proxy_FNO_subsídio_Norte_R$_bi']:.3f} | "
            f"{r['Atualizado_IPCA_30_06_2026_R$'] / 1e9:.3f} |"
        )
    lines += [
        "",
        "Regenerar:",
        "",
        "```bash",
        "python3 scripts/build_osu_fundos_constitucionais.py",
        "```",
    ]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")


if __name__ == "__main__":
    write_outputs(build())
