#!/usr/bin/env python3
"""
OPERACOES DIRETAS — colunas K, L, M, N

Para cada contrato (uma linha):
  K — Valor desembolsado atualizado pelo IPCA (até a data de referência)
  L — SELIC total capitalizada mensalmente no prazo de amortização
  M — Taxa de juros total do contrato capitalizada mensalmente no prazo
  N — Diferença L − M

Fórmulas:
  fator_ipca(t) = Π (1 + ipca_m/100) até o mês t
  K = valor_desembolsado × fator_ipca(ref) / fator_ipca(mês da contratação)

  L = Π_{i=0}^{n-1} (1 + selic_{t0+i}/100)     n = prazo_amortização (meses)
  M = (1 + taxa_contrato_mensal)^n
  N = L − M

  taxa_contrato_mensal:
    - TJLP/TLP: (1+0,06)^(1/12)×(1+juros/100)^(1/12) − 1
    - demais:   (1+juros/100)^(1/12) − 1

Uso (WinPython ContAgil):
  python scripts\\calcular_diretas_ipca_selic.py ^
    --excel "C:\\Arquivos de Programas RFB\\ContAgilAppBeta64\\python_jep\\winpython\\OPERACOES DIRETAS - 2002 a 2018.xlsx"

  # Saída padrão: mesmo nome com sufixo _calculado.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

from scripts.gerar_fluxos import (
    CONTAGIL_WINPYTHON,
    DATA_DIR,
    _mapear_colunas_contratos,
    limpar_valor,
    parse_datas,
    taxa_contrato_efetiva,
)

BCB_SGS = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
IPCA_COD = 433
SELIC_MENSAL_COD = 4390  # % a.m. (Selic acumulada no mês anualizada ≈ % a.m. ContAgil)
DATA_REF_DEFAULT = datetime(2026, 6, 30)

HEADERS_KLMN = {
    "K": "Valor_Desembolsado_IPCA",
    "L": "SELIC_Total_Cap_Mensal",
    "M": "Juros_Contrato_Total_Cap_Mensal",
    "N": "Diferenca_L_menos_M",
}


def _baixar_sgs(cod: int, inicio: str = "01/01/2000", fim: str | None = None) -> pd.DataFrame:
    """Baixa série SGS Bacen (json) → DataFrame data (mês) + valor."""
    if fim is None:
        fim = datetime.now().strftime("%d/%m/%Y")
    url = BCB_SGS.format(cod=cod)
    # Bacen limita janelas longas — baixa em blocos de ~10 anos
    inicio_ts = pd.Timestamp(day=1, month=1, year=int(inicio[-4:]))
    fim_ts = pd.Timestamp(datetime.strptime(fim, "%d/%m/%Y"))
    partes: list[pd.DataFrame] = []
    cursor = inicio_ts
    while cursor <= fim_ts:
        bloco_fim = min(cursor + pd.DateOffset(years=9, months=11), fim_ts)
        params = {
            "formato": "json",
            "dataInicial": cursor.strftime("%d/%m/%Y"),
            "dataFinal": bloco_fim.strftime("%d/%m/%Y"),
        }
        resp = requests.get(url, params=params, timeout=120)
        resp.raise_for_status()
        dados = resp.json()
        if dados:
            df = pd.DataFrame(dados)
            df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
            partes.append(df.dropna())
        cursor = bloco_fim + pd.DateOffset(days=1)
    if not partes:
        raise RuntimeError(f"Série SGS {cod} vazia")
    out = (
        pd.concat(partes, ignore_index=True)
        .drop_duplicates(subset=["data"])
        .sort_values("data")
        .reset_index(drop=True)
    )
    out["mes"] = out["data"].dt.to_period("M").dt.to_timestamp()
    return out[["mes", "valor"]]


def carregar_ipca(path: Path | None = None) -> pd.DataFrame:
    """IPCA mensal % a.m. + fator acumulado."""
    if path is not None and path.exists():
        raw = pd.read_excel(path)
        cols = {str(c).strip().lower(): c for c in raw.columns}
        data_col = cols.get("data", raw.columns[0])
        taxa_col = next(
            (cols[k] for k in cols if "ipca" in k or "taxa" in k or "%" in k),
            raw.columns[1],
        )
        df = pd.DataFrame(
            {
                "mes": pd.to_datetime(raw[data_col], dayfirst=True, errors="coerce")
                .dt.to_period("M")
                .dt.to_timestamp(),
                "valor": pd.to_numeric(raw[taxa_col], errors="coerce"),
            }
        ).dropna()
    else:
        print("Baixando IPCA (Bacen SGS 433)...")
        df = _baixar_sgs(IPCA_COD)
    df = df.sort_values("mes").drop_duplicates("mes")
    # valor em % a.m.
    med = float(df["valor"].median())
    if med > 50:  # índice (ex. 100+) — não esperado na 433
        raise ValueError("IPCA não parece variação % a.m.")
    df["fator"] = (1.0 + df["valor"] / 100.0).cumprod()
    return df.reset_index(drop=True)


def carregar_selic_mensal(path: Path | None = None) -> pd.DataFrame:
    """SELIC mensal % a.m. (ContAgil selic_mensal.xlsx ou Bacen 4390)."""
    candidatos = []
    if path is not None:
        candidatos.append(path)
    candidatos.extend(
        [
            Path.cwd() / "selic_mensal.xlsx",
            CONTAGIL_WINPYTHON / "selic_mensal.xlsx",
            DATA_DIR / "selic_mensal.xlsx",
        ]
    )
    for cand in candidatos:
        if cand is not None and cand.exists():
            print(f"SELIC mensal: {cand}")
            raw = pd.read_excel(cand)
            data_col = raw.columns[0]
            taxa_col = raw.columns[1]
            df = pd.DataFrame(
                {
                    "mes": pd.to_datetime(raw[data_col], dayfirst=True, errors="coerce")
                    .dt.to_period("M")
                    .dt.to_timestamp(),
                    "valor": pd.to_numeric(raw[taxa_col], errors="coerce"),
                }
            ).dropna()
            return df.sort_values("mes").drop_duplicates("mes").reset_index(drop=True)

    print("Baixando SELIC mensal (Bacen SGS 4390)...")
    df = _baixar_sgs(SELIC_MENSAL_COD)
    return df.sort_values("mes").drop_duplicates("mes").reset_index(drop=True)


def _idx_mes(serie: pd.DataFrame, mes: pd.Timestamp) -> int:
    """Índice do mês mais próximo ≤ mes (senão o primeiro)."""
    mes = pd.Timestamp(mes).to_period("M").to_timestamp()
    idxs = serie.index[serie["mes"] <= mes]
    if len(idxs):
        return int(idxs[-1])
    return 0


def fator_ipca_entre(
    ipca: pd.DataFrame, mes_ini: pd.Timestamp, mes_fim: pd.Timestamp
) -> float:
    """fator_ipca(fim) / fator_ipca(ini)."""
    i0 = _idx_mes(ipca, mes_ini)
    i1 = _idx_mes(ipca, mes_fim)
    f0 = float(ipca.loc[i0, "fator"])
    f1 = float(ipca.loc[i1, "fator"])
    if f0 <= 0:
        return 1.0
    return f1 / f0


def selic_cap_mensal(selic: pd.DataFrame, mes_ini: pd.Timestamp, n: int) -> float:
    """Π (1+selic_m/100) por n meses a partir de mes_ini (extrapolando último se faltar)."""
    if n <= 0:
        return 1.0
    i0 = _idx_mes(selic, mes_ini)
    taxas = selic["valor"].to_numpy(dtype=float)
    fator = 1.0
    last = float(taxas[i0]) if len(taxas) else 0.0
    for k in range(int(n)):
        idx = i0 + k
        if idx < len(taxas):
            last = float(taxas[idx])
        fator *= 1.0 + last / 100.0
    return fator


def juros_contrato_cap(custo: str | None, juros_pct: float, n: int) -> float:
    """(1 + taxa_contrato_mensal)^n."""
    if n <= 0:
        return 1.0
    taxa_m = taxa_contrato_efetiva(custo, juros_pct)
    return float((1.0 + taxa_m) ** int(n))


def _resolver_excel(path: Path | None) -> Path:
    if path is not None and path.exists():
        return path
    candidatos = [
        path,
        Path.cwd() / "OPERACOES DIRETAS - 2002 a 2018.xlsx",
        CONTAGIL_WINPYTHON / "OPERACOES DIRETAS - 2002 a 2018.xlsx",
        Path.cwd() / "OPERACOES DIRETAS.xlsx",
        CONTAGIL_WINPYTHON / "OPERACOES DIRETAS.xlsx",
    ]
    for c in candidatos:
        if c is not None and c.exists():
            return c
    raise FileNotFoundError(
        "Excel de operações diretas não encontrado.\n"
        "Informe --excel com o caminho completo, ex.:\n"
        r'  --excel "C:\Arquivos de Programas RFB\ContAgilAppBeta64\python_jep\winpython\OPERACOES DIRETAS - 2002 a 2018.xlsx"'
    )


def preparar_contratos(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Mapeia colunas BNDES/ContAgil e devolve campos numéricos necessários."""
    mapped, rename = _mapear_colunas_contratos(df_raw)
    if rename:
        print(f"Colunas mapeadas: {rename}")

    if "valor_desembolsado" not in mapped.columns and "valor_contratado" in mapped.columns:
        mapped["valor_desembolsado"] = mapped["valor_contratado"]
    elif "valor_desembolsado" in mapped.columns and "valor_contratado" in mapped.columns:
        vd = limpar_valor(mapped["valor_desembolsado"])
        vc = limpar_valor(mapped["valor_contratado"])
        mapped["valor_desembolsado"] = vd.fillna(vc)

    required = ["data_contratacao", "valor_desembolsado", "juros", "prazo_amortizacao"]
    missing = [c for c in required if c not in mapped.columns]
    if missing:
        raise ValueError(
            f"Colunas ausentes: {missing}. Disponíveis: {list(df_raw.columns)}"
        )

    out = pd.DataFrame(
        {
            "data_contratacao": parse_datas(mapped["data_contratacao"]),
            "valor_desembolsado": limpar_valor(mapped["valor_desembolsado"]),
            "juros": limpar_valor(mapped["juros"]).fillna(0.0),
            "prazo_amortizacao": limpar_valor(mapped["prazo_amortizacao"]).fillna(0).astype(int),
            "custo_financeiro": (
                mapped["custo_financeiro"].astype(str)
                if "custo_financeiro" in mapped.columns
                else ""
            ),
        }
    )
    return out


def calcular_klmn(
    contratos: pd.DataFrame,
    ipca: pd.DataFrame,
    selic: pd.DataFrame,
    data_ref: datetime = DATA_REF_DEFAULT,
) -> pd.DataFrame:
    """Calcula colunas K, L, M, N para cada linha."""
    ref = pd.Timestamp(data_ref).to_period("M").to_timestamp()
    ks, ls, ms, ns = [], [], [], []
    for _, row in contratos.iterrows():
        data = row["data_contratacao"]
        if pd.isna(data) or pd.isna(row["valor_desembolsado"]):
            ks.append(np.nan)
            ls.append(np.nan)
            ms.append(np.nan)
            ns.append(np.nan)
            continue
        mes0 = pd.Timestamp(data).to_period("M").to_timestamp()
        n = int(row["prazo_amortizacao"])
        valor = float(row["valor_desembolsado"])
        k = valor * fator_ipca_entre(ipca, mes0, ref) if valor > 0 else np.nan
        l = selic_cap_mensal(selic, mes0, n)
        m = juros_contrato_cap(row.get("custo_financeiro"), float(row["juros"]), n)
        ks.append(round(k, 2) if pd.notna(k) else np.nan)
        ls.append(round(l, 8))
        ms.append(round(m, 8))
        ns.append(round(l - m, 8))
    return pd.DataFrame(
        {
            HEADERS_KLMN["K"]: ks,
            HEADERS_KLMN["L"]: ls,
            HEADERS_KLMN["M"]: ms,
            HEADERS_KLMN["N"]: ns,
        }
    )


def gravar_colunas_klmn(
    excel_path: Path,
    calc: pd.DataFrame,
    saida: Path,
    header_row: int = 1,
) -> Path:
    """Escreve K–N no Excel (openpyxl), preservando demais colunas."""
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # Cabeçalhos
    for col_letter, titulo in HEADERS_KLMN.items():
        ws[f"{col_letter}{header_row}"] = titulo

    # Dados: linhas Excel = header_row+1 ...
    for i, row in calc.iterrows():
        excel_row = header_row + 1 + int(i)
        ws[f"K{excel_row}"] = row[HEADERS_KLMN["K"]]
        ws[f"L{excel_row}"] = row[HEADERS_KLMN["L"]]
        ws[f"M{excel_row}"] = row[HEADERS_KLMN["M"]]
        ws[f"N{excel_row}"] = row[HEADERS_KLMN["N"]]

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


def detectar_header_row(path: Path) -> int:
    """Retorna 1-based header row ( ContAgil = 1; portal às vezes 6)."""
    from scripts.gerar_fluxos import _excel_tem_colunas_contratos

    for h0 in (0, 5, 1, 2, 3, 4):
        try:
            df = pd.read_excel(path, header=h0, nrows=5)
        except Exception:
            continue
        if _excel_tem_colunas_contratos(df):
            return h0 + 1  # openpyxl 1-based
    return 1


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="OPERACOES DIRETAS - 2002 a 2018.xlsx (ContAgil).",
    )
    p.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Excel de saída (default: *_calculado.xlsx).",
    )
    p.add_argument("--ipca", type=Path, default=None, help="Excel IPCA opcional.")
    p.add_argument("--selic", type=Path, default=None, help="selic_mensal.xlsx ContAgil.")
    p.add_argument(
        "--data-ref",
        type=str,
        default="2026-06-30",
        help="Data de referência do IPCA (YYYY-MM-DD).",
    )
    p.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="Linha do cabeçalho (1-based). Default: auto.",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    excel = _resolver_excel(args.excel)
    print(f"Excel: {excel}")

    header_row = args.header_row or detectar_header_row(excel)
    header0 = header_row - 1
    print(f"Header Excel: linha {header_row}")

    df_raw = pd.read_excel(excel, header=header0)
    print(f"Linhas: {len(df_raw):,}")
    contratos = preparar_contratos(df_raw)
    print(
        f"Contratos com data/valor: "
        f"{contratos['data_contratacao'].notna().sum():,} | "
        f"prazo>0: {(contratos['prazo_amortizacao'] > 0).sum():,}"
    )

    ipca = carregar_ipca(args.ipca)
    selic = carregar_selic_mensal(args.selic)
    print(f"IPCA: {len(ipca)} meses | SELIC: {len(selic)} meses")

    data_ref = datetime.strptime(args.data_ref, "%Y-%m-%d")
    calc = calcular_klmn(contratos, ipca, selic, data_ref=data_ref)

    saida = args.saida
    if saida is None:
        saida = excel.with_name(excel.stem + "_calculado.xlsx")

    gravar_colunas_klmn(excel, calc, saida, header_row=header_row)
    print(f"✅ Salvo: {saida}")
    print(calc.describe(include="all").to_string())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
