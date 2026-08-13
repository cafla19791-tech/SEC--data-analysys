#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Calculo de fluxos e impactos - BNDES Indiretos (capitalizacao mensal).

Este arquivo e PYTHON. Nao cole aqui o conteudo de contagil_fluxos_bndes.bat.

Uso (uma linha):
  python scripts/contagil_fluxos.py --massa-dados dados --pasta-saida saida --arquivo-fatores fator_acumulado_SELIC_TJLP_TLP.xlsx

Aliases: --arquivo-fatores | --fatores
"""

from __future__ import annotations

import argparse
import glob
import importlib.util
import os
import sys
import types
from datetime import datetime
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).resolve().parent
ROOT = _SCRIPTS_DIR.parent


def _load_sibling(mod_name: str):
    full = f"scripts.{mod_name}"
    if full in sys.modules:
        return sys.modules[full]
    path = _SCRIPTS_DIR / f"{mod_name}.py"
    if not path.is_file():
        print(f"ERRO: falta o arquivo {path}")
        raise SystemExit(2)
    if "scripts" not in sys.modules:
        pkg = types.ModuleType("scripts")
        pkg.__path__ = [str(_SCRIPTS_DIR)]
        pkg.__package__ = "scripts"
        sys.modules["scripts"] = pkg
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    spec = importlib.util.spec_from_file_location(full, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Nao foi possivel carregar {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[full] = mod
    sys.modules[mod_name] = mod
    spec.loader.exec_module(mod)
    return mod


_gf = _load_sibling("gerar_fluxos")

import numpy as np
import pandas as pd

CONTAGIL_PASTA_DADOS = _gf.CONTAGIL_PASTA_DADOS
CONTAGIL_PASTA_SAIDA = _gf.CONTAGIL_PASTA_SAIDA
CONTAGIL_WINPYTHON = _gf.CONTAGIL_WINPYTHON
DATA_DIR = _gf.DATA_DIR
DATA_IMPACTO = _gf.DATA_IMPACTO
OUTPUT_DIR = _gf.OUTPUT_DIR
SelicSerie = _gf.SelicSerie
_excel_tem_colunas_contratos = _gf._excel_tem_colunas_contratos
_mapear_colunas_contratos = _gf._mapear_colunas_contratos
gerar_fluxos = _gf.gerar_fluxos
gerar_e_gravar_fluxos = _gf.gerar_e_gravar_fluxos
load_from_excel = _gf.load_from_excel
normalizar_colunas = _gf.normalizar_colunas

# Acima deste nº de contratos, grava em lotes (CSV streaming) para não estourar RAM.
LIMITE_MEMORIA_CONTRATOS = 5_000
LOTE_FLUXOS = 2_000

FATORES_DEFAULT_NOME = "fator_acumulado_SELIC_TJLP_TLP.xlsx"
DATA_REF_DEFAULT = datetime(2026, 6, 1)


def _banner() -> None:
    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    print("=" * 70)
    print("CALCULO DE FLUXOS E IMPACTOS - BNDES INDIRETOS")
    print(f"Início: {agora}")
    print("=" * 70)


def _parece_caminho_contagil(path: Path | None) -> bool:
    if path is None:
        return False
    texto = str(path).replace("/", "\\").upper()
    return "CONTAGIL" in texto or "WINPYTHON" in texto or texto.startswith("C:\\ARQUIVOS")


def resolver_pasta_dados(arg: Path | None) -> Path:
    if arg is not None:
        if arg.exists():
            return arg
        if _parece_caminho_contagil(arg):
            local = DATA_DIR / "contagil_winpython" / "dados"
            local.mkdir(parents=True, exist_ok=True)
            print(f"⚠️ Massa ContAgil ausente: {arg}")
            print(f"   Usando espelho local: {local}")
            return local
        return arg
    if CONTAGIL_PASTA_DADOS.exists():
        return CONTAGIL_PASTA_DADOS
    # Script na pasta winpython: ./dados
    cwd_dados = Path.cwd() / "dados"
    if cwd_dados.exists():
        return cwd_dados
    local = DATA_DIR / "contagil_winpython" / "dados"
    local.mkdir(parents=True, exist_ok=True)
    return local


def resolver_pasta_saida(arg: Path | None) -> Path:
    if arg is not None:
        if arg.exists() or not _parece_caminho_contagil(arg):
            arg.mkdir(parents=True, exist_ok=True)
            return arg
        local = DATA_DIR / "contagil_winpython" / "saida"
        local.mkdir(parents=True, exist_ok=True)
        print(f"⚠️ Saída ContAgil ausente: {arg}")
        print(f"   Usando espelho local: {local}")
        return local
    if CONTAGIL_PASTA_SAIDA.exists():
        return CONTAGIL_PASTA_SAIDA
    cwd_saida = Path.cwd() / "saida"
    if cwd_saida.exists() or (Path.cwd() / "dados").exists():
        cwd_saida.mkdir(parents=True, exist_ok=True)
        return cwd_saida
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


def resolver_fatores(arg: Path | None) -> Path:
    candidatos: list[Path] = []
    if arg is not None:
        candidatos.append(arg)
    candidatos.extend(
        [
            Path.cwd() / FATORES_DEFAULT_NOME,
            CONTAGIL_WINPYTHON / FATORES_DEFAULT_NOME,
            DATA_DIR / FATORES_DEFAULT_NOME,
            DATA_DIR / "selic_mensal.xlsx",
            DATA_DIR / "selic_taxas_contagil.xlsx",
        ]
    )
    for cand in candidatos:
        if cand is not None and cand.exists():
            return cand
    raise FileNotFoundError(
        f"Arquivo de fatores não encontrado. Procurado: {FATORES_DEFAULT_NOME}\n"
        "Coloque fator_acumulado_SELIC_TJLP_TLP.xlsx na pasta winpython "
        "ou informe --arquivo-fatores / --fatores."
    )


def _contar_serie_mensal(df: pd.DataFrame) -> int:
    """Conta linhas válidas (data + valor numérico > 0) em uma série mensal."""
    if df is None or df.empty:
        return 0
    cols = list(df.columns)
    if not cols:
        return 0
    data_col = cols[0]
    for c in cols:
        if str(c).strip().lower() in {"data", "date", "mes", "mês"}:
            data_col = c
            break
    valor_col = cols[-1]
    for c in cols:
        low = str(c).strip().lower()
        if "fator" in low or "taxa" in low or "selic" in low or "tjlp" in low or low == "tlp":
            valor_col = c
            break
    datas = pd.to_datetime(df[data_col], dayfirst=True, errors="coerce")
    vals = pd.to_numeric(df[valor_col], errors="coerce")
    return int((datas.notna() & vals.notna()).sum())


def _contar_meses_auxiliares(path: Path) -> tuple[int, int]:
    """Conta meses TJLP/TLP em abas do Excel ou arquivos irmãos na mesma pasta."""
    n_tjlp = 0
    n_tlp = 0
    try:
        with pd.ExcelFile(path) as xl:
            for sheet in xl.sheet_names:
                low = str(sheet).strip().lower()
                try:
                    part = pd.read_excel(xl, sheet_name=sheet)
                except Exception:  # noqa: BLE001
                    continue
                if "tjlp" in low:
                    n_tjlp = max(n_tjlp, _contar_serie_mensal(part))
                elif low.strip() == "tlp" or (low.startswith("tlp") and "tjlp" not in low):
                    n_tlp = max(n_tlp, _contar_serie_mensal(part))
                else:
                    # Colunas nomeadas na aba principal
                    for c in part.columns:
                        cl = str(c).strip().lower()
                        if "tjlp" in cl and (
                            "fator" in cl or "acumul" in cl or "taxa" in cl
                        ):
                            n_tjlp = max(
                                n_tjlp,
                                int(pd.to_numeric(part[c], errors="coerce").notna().sum()),
                            )
                        elif "tlp" in cl and "tjlp" not in cl and (
                            "fator" in cl or "acumul" in cl or "taxa" in cl or cl == "tlp"
                        ):
                            n_tlp = max(
                                n_tlp,
                                int(pd.to_numeric(part[c], errors="coerce").notna().sum()),
                            )
    except Exception:  # noqa: BLE001
        pass

    pasta = path.parent
    if n_tjlp == 0:
        for nome in ("tjlp_mensal.xlsx", "TJLP_mensal.xlsx"):
            cand = pasta / nome
            if cand.exists():
                try:
                    n_tjlp = _contar_serie_mensal(pd.read_excel(cand))
                except Exception:  # noqa: BLE001
                    pass
                break
    if n_tlp == 0:
        for nome in ("tlp_mensal.xlsx", "TLP_mensal.xlsx"):
            cand = pasta / nome
            if cand.exists():
                try:
                    n_tlp = _contar_serie_mensal(pd.read_excel(cand))
                except Exception:  # noqa: BLE001
                    pass
                break
    return n_tjlp, n_tlp


def carregar_fatores_mensais(path: Path) -> SelicSerie:
    """Carrega fator_acumulado_SELIC_TJLP_TLP.xlsx (Data + Fator_Acumulado)."""
    print(f"Carregando fatores combinados: {path}")
    raw = pd.read_excel(path)
    print(f"Colunas encontradas: {list(raw.columns)}")

    cols_norm = {str(c).strip().lower(): c for c in raw.columns}
    data_col = None
    for cand in ("data", "date", "mes", "mês"):
        if cand in cols_norm:
            data_col = cols_norm[cand]
            break
    if data_col is None:
        data_col = raw.columns[0]

    fator_col = None
    for cand in ("fator_acumulado", "fator", "fator acumulado", "fator_selic"):
        if cand in cols_norm:
            fator_col = cols_norm[cand]
            break
    if fator_col is None:
        # Preferência: coluna com 'selic' + fator; senão qualquer 'fator'
        for c in raw.columns:
            low = str(c).lower()
            if "fator" in low and "selic" in low:
                fator_col = c
                break
        if fator_col is None:
            for c in raw.columns:
                if "fator" in str(c).lower():
                    fator_col = c
                    break
    if fator_col is None:
        fator_col = raw.columns[-1]

    datas = pd.to_datetime(raw[data_col], dayfirst=True, errors="coerce")
    fatores = pd.to_numeric(raw[fator_col], errors="coerce")
    mask = datas.notna() & fatores.notna() & (fatores > 0)
    datas_arr = datas[mask].to_numpy(dtype="datetime64[ns]")
    fatores_arr = fatores[mask].to_numpy(dtype=float)

    # Referência: 01/06/2026 (ou último fator ≤ data de impacto)
    data_ref = np.datetime64(DATA_REF_DEFAULT)
    ate_ref = fatores_arr[datas_arr <= data_ref]
    if len(ate_ref):
        fator_ref = float(ate_ref[-1])
        data_ref_usada = pd.Timestamp(
            datas_arr[datas_arr <= data_ref][-1]
        ).strftime("%Y-%m-%d")
    else:
        fator_ref = float(fatores_arr[-1])
        data_ref_usada = pd.Timestamp(datas_arr[-1]).strftime("%Y-%m-%d")

    n_selic = len(fatores_arr)
    n_tjlp, n_tlp = _contar_meses_auxiliares(path)

    print(f"Referência de atualização: {data_ref_usada} | Fator SELIC = {fator_ref:.8f}")
    print(f"SELIC: {n_selic} meses | TJLP: {n_tjlp} meses | TLP: {n_tlp} meses")

    return SelicSerie(
        datas_arr,
        fatores_arr,
        origem=f"mensal:{path.name}",
        fator_referencia=fator_ref,
    )


def listar_contratos(pasta: Path) -> list[Path]:
    arquivos = sorted(Path(p) for p in glob.glob(os.path.join(str(pasta), "*.xlsx")))
    saida: list[Path] = []
    for arq in arquivos:
        nome = arq.name.upper()
        if nome.startswith("STP") or "SELIC" in nome or "FATOR" in nome or "TJLP" in nome:
            continue
        if nome.startswith("~$"):
            continue
        saida.append(arq)
    return saida


def diagnosticar_colunas(path: Path) -> None:
    """Imprime colunas brutas e mapeamento — ajuda quando o arquivo é pulado."""
    for h in (0, 5, 1, 2, 3, 4):
        try:
            df = pd.read_excel(path, sheet_name=0, header=h, nrows=5)
        except Exception as exc:  # noqa: BLE001
            print(f"    header={h}: erro ao ler ({exc})")
            continue
        mapped, rename = _mapear_colunas_contratos(df)
        ok = _excel_tem_colunas_contratos(df)
        print(f"    header={h}: colunas={list(df.columns)[:12]}")
        print(f"             mapeadas={rename or '{}'} | ok={ok}")
        if ok:
            missing = [
                c
                for c in (
                    "data_contratacao",
                    "valor_desembolsado",
                    "juros",
                    "prazo_carencia",
                    "prazo_amortizacao",
                )
                if c not in mapped.columns
            ]
            if missing:
                print(f"             ainda faltam: {missing}")
            break


def _contar_linhas_excel(path: Path) -> int | None:
    """Conta linhas de dados sem carregar a planilha inteira em memória."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # max_row inclui header; reporta linhas brutas como o log ContAgil
        n = ws.max_row or 0
        wb.close()
        return int(n)
    except Exception:  # noqa: BLE001
        return None


def _ler_bruto_com_header(path: Path) -> pd.DataFrame:
    """Lê o Excel bruto na linha de header correta (antes de normalizar_colunas)."""
    for h in (0, 5, 1, 2, 3, 4, 6, 7, 8):
        try:
            candidato = pd.read_excel(path, sheet_name=0, header=h)
        except Exception:  # noqa: BLE001
            continue
        if _excel_tem_colunas_contratos(candidato):
            if h != 0:
                print(f"    Header Excel detectado na linha {h}")
            return candidato
    # Fallback: tenta load_from_excel (já mapeia) — devolve bruto da linha 0
    return pd.read_excel(path, sheet_name=0, header=0)


def processar_arquivo(
    arquivo: Path,
    pasta_saida: Path,
    serie: SelicSerie,
    max_contratos: int | None = None,
) -> Path | None:
    print(f"\n>>> {arquivo.name} ...")
    n_linhas = _contar_linhas_excel(arquivo)
    if n_linhas is not None:
        print(f"    Linhas: {n_linhas:,}")

    try:
        # Fluxo ContAgil: read → normalizar_colunas → gerar_fluxos
        # (a definição de normalizar_colunas estava ausente no script WinPython)
        try:
            bruto = _ler_bruto_com_header(arquivo)
            df = normalizar_colunas(bruto)
        except ValueError:
            # Header/aba atípicos: tenta o loader completo
            df = load_from_excel(arquivo)

        if n_linhas is None:
            print(f"    Linhas: {len(df):,} (após limpeza)")

        if max_contratos is not None:
            df = df.head(int(max_contratos)).copy()
            df["contrato"] = df.index

        if df.empty:
            print("    [AVISO] Nenhum contrato válido após limpeza. Pulando.")
            return None

        pasta_saida.mkdir(parents=True, exist_ok=True)
        saida = pasta_saida / f"fluxos_{arquivo.stem}.xlsx"
        # Massa grande (ex.: BNDES INDIRETAS ~700k+): streaming em lotes.
        # Sem isso o processo fica minutos/horas sem log e estoura a memória.
        if len(df) >= LIMITE_MEMORIA_CONTRATOS:
            gerar_e_gravar_fluxos(
                df,
                serie,
                saida_xlsx=saida,
                lote=LOTE_FLUXOS,
            )
            return saida

        fluxos = gerar_fluxos(df, serie)
        # Excel ~1M linhas: se passar, grava CSV completo + amostra xlsx
        if len(fluxos) > 1_000_000:
            csv_path = saida.with_suffix(".csv")
            fluxos.to_csv(csv_path, index=False)
            fluxos.head(1_000_000).to_excel(saida, index=False)
            print(f"    → CSV completo: {csv_path} ({len(fluxos):,} parcelas)")
            print(f"    → Excel (amostra 1M): {saida}")
        else:
            fluxos.to_excel(saida, index=False)
            print(f"    → Salvo: {saida} ({len(fluxos):,} parcelas)")
        return saida
    except Exception as exc:  # noqa: BLE001 — espelha log ContAgil "ERRO: ..."
        print(f"    ERRO: {exc}")
        diagnosticar_colunas(arquivo)
        return None


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--massa-dados",
        "--pasta-dados",
        dest="massa_dados",
        type=Path,
        default=None,
        help="Pasta com BNDES INDIRETAS *.xlsx (default: winpython/dados).",
    )
    p.add_argument(
        "--pasta-saida",
        type=Path,
        default=None,
        help="Pasta de saída (default: winpython/saida).",
    )
    p.add_argument(
        "--arquivo-fatores",
        "--fatores",
        dest="fatores",
        type=Path,
        default=None,
        metavar="FILE",
        help=f"Excel de fatores mensais (default: {FATORES_DEFAULT_NOME}).",
    )
    p.add_argument(
        "--max-contratos",
        type=int,
        default=None,
        help="Limita contratos por arquivo (teste rápido).",
    )
    p.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Processa um único Excel/CSV em vez da massa.",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    _banner()

    pasta_dados = resolver_pasta_dados(args.massa_dados)
    pasta_saida = resolver_pasta_saida(args.pasta_saida)
    print(f"Massa de dados : {pasta_dados}")
    print(f"Pasta de saída : {pasta_saida}")
    print()

    fatores_path = resolver_fatores(args.fatores)
    serie = carregar_fatores_mensais(fatores_path)
    print()

    if args.input is not None:
        arquivos = [Path(args.input)]
    else:
        if not pasta_dados.exists():
            print(f"[ERRO] Massa de dados não encontrada: {pasta_dados}")
            return 1
        arquivos = listar_contratos(pasta_dados)

    print(f"Arquivos de contratos ({len(arquivos)}):")
    for a in arquivos:
        print(f"  - {a.name}")
    print()

    if not arquivos:
        print("Nenhum arquivo de contrato encontrado.")
        return 1

    ok: list[Path] = []
    for arq in arquivos:
        out = processar_arquivo(
            arq, pasta_saida, serie, max_contratos=args.max_contratos
        )
        if out is not None:
            ok.append(out)

    print()
    if not ok:
        print("Nenhum contrato válido processado.")
        return 2

    print(f"Concluído: {len(ok)} arquivo(s) processado(s).")
    print(f"Referência de impacto: {DATA_IMPACTO.date()} (capitalização mensal).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
