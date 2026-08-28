#!/usr/bin/env python3
"""Cotejamento IPCA (Brasil) × CPI All Items (EUA), 1990–2025.

Critério comum: variação dezembro/dezembro.
  IPCA — IBGE / Bacen SGS 13522 (acumulado em 12 meses, dezembro)
  CPI All Items — BLS CPI-U NSA / FRED CPIAUCNS (Dez_t / Dez_{t-1} − 1)
"""

from __future__ import annotations

import argparse
import io
import sys
import time
from pathlib import Path

import pandas as pd
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

ANO_INICIO_DEFAULT = 1990
ANO_FIM_DEFAULT = 2025
SGS_IPCA_12M = 13522
FRED_CPI = "CPIAUCNS"
SGS_API = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
FRED_CSV = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={sid}"

COLUNAS_ANUAL = [
    "ano",
    "ipca_pct",
    "cpi_all_items_pct",
    "diferenca_pp",
    "maior_no_ano",
    "indice_ipca",
    "indice_cpi",
    "razao_indices",
]


def anos_periodo(ano_inicio: int, ano_fim: int) -> list[int]:
    if ano_fim < ano_inicio:
        raise ValueError("ano_fim deve ser >= ano_inicio")
    return list(range(ano_inicio, ano_fim + 1))


def _get_json(http, url: str, params: dict, tentativas: int = 4):
    ultimo_erro: Exception | None = None
    for tentativa in range(tentativas):
        resp = http.get(url, params=params, timeout=120)
        if resp.status_code == 404:
            return []
        corpo = getattr(resp, "content", b"x") or b""
        if resp.status_code in {429, 500, 502, 503} or not corpo.strip():
            ultimo_erro = RuntimeError(f"HTTP {resp.status_code} vazio")
            time.sleep(1.5 * (tentativa + 1))
            continue
        resp.raise_for_status()
        try:
            return resp.json() or []
        except ValueError as exc:
            ultimo_erro = exc
            time.sleep(1.5 * (tentativa + 1))
    raise RuntimeError(f"Falha ao ler {url}: {ultimo_erro}") from ultimo_erro


def baixar_sgs(codigo: int, inicio: str, fim: str, session=None) -> pd.DataFrame:
    http = session or requests
    dados = _get_json(
        http,
        SGS_API.format(cod=codigo),
        {"formato": "json", "dataInicial": inicio, "dataFinal": fim},
    )
    if not dados:
        return pd.DataFrame(columns=["data", "valor"])
    df = pd.DataFrame(dados)
    df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    return (
        df.dropna(subset=["data", "valor"])
        .drop_duplicates("data")
        .sort_values("data")
        .reset_index(drop=True)[["data", "valor"]]
    )


def baixar_fred(series_id: str, session=None) -> pd.DataFrame:
    http = session or requests
    ultimo_erro: Exception | None = None
    for tentativa in range(4):
        resp = http.get(FRED_CSV.format(sid=series_id), timeout=120)
        if resp.status_code in {429, 500, 502, 503} or not (resp.content or b"").strip():
            ultimo_erro = RuntimeError(f"FRED HTTP {resp.status_code}")
            time.sleep(1.5 * (tentativa + 1))
            continue
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))
        df = df.rename(columns={df.columns[0]: "data", df.columns[1]: "valor"})
        df["data"] = pd.to_datetime(df["data"], errors="coerce")
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        return (
            df.dropna(subset=["data", "valor"])
            .drop_duplicates("data")
            .sort_values("data")
            .reset_index(drop=True)[["data", "valor"]]
        )
    raise RuntimeError(f"Falha ao ler FRED {series_id}: {ultimo_erro}")


def ipca_dez_dez(df: pd.DataFrame, ano: int) -> float | None:
    dez = df.loc[(df["data"].dt.year == ano) & (df["data"].dt.month == 12), "valor"]
    if dez.empty:
        return None
    return float(dez.iloc[0]) / 100.0


def cpi_dez_dez(df: pd.DataFrame, ano: int) -> float | None:
    dez = df.loc[(df["data"].dt.year == ano) & (df["data"].dt.month == 12), "valor"]
    ant = df.loc[(df["data"].dt.year == ano - 1) & (df["data"].dt.month == 12), "valor"]
    if dez.empty or ant.empty or float(ant.iloc[0]) == 0:
        return None
    return float(dez.iloc[0]) / float(ant.iloc[0]) - 1.0


def acumulado(taxas: list[float]) -> float:
    fator = 1.0
    for taxa in taxas:
        fator *= 1.0 + taxa
    return fator - 1.0


def media_geometrica(taxas: list[float]) -> float:
    if not taxas:
        return float("nan")
    return (1.0 + acumulado(taxas)) ** (1.0 / len(taxas)) - 1.0


def montar_cotejamento(
    ipca: pd.DataFrame,
    cpi: pd.DataFrame,
    anos: list[int],
) -> pd.DataFrame:
    linhas = []
    idx_ipca = 100.0
    idx_cpi = 100.0
    for ano in anos:
        p_ipca = ipca_dez_dez(ipca, ano)
        p_cpi = cpi_dez_dez(cpi, ano)
        if p_ipca is None or p_cpi is None:
            raise RuntimeError(f"Série incompleta em {ano}")
        idx_ipca *= 1.0 + p_ipca
        idx_cpi *= 1.0 + p_cpi
        if p_ipca > p_cpi:
            maior = "IPCA"
        elif p_cpi > p_ipca:
            maior = "CPI All Items"
        else:
            maior = "Empate"
        linhas.append(
            {
                "ano": ano,
                "ipca_pct": p_ipca * 100.0,
                "cpi_all_items_pct": p_cpi * 100.0,
                "diferenca_pp": (p_ipca - p_cpi) * 100.0,
                "maior_no_ano": maior,
                "indice_ipca": idx_ipca,
                "indice_cpi": idx_cpi,
                "razao_indices": idx_ipca / idx_cpi,
            }
        )
    return pd.DataFrame(linhas, columns=COLUNAS_ANUAL)


def resumo_periodo(df: pd.DataFrame, ano_ini: int, ano_fim: int, rotulo: str) -> dict:
    recorte = df.loc[(df["ano"] >= ano_ini) & (df["ano"] <= ano_fim)].copy()
    ipca = (recorte["ipca_pct"] / 100.0).tolist()
    cpi = (recorte["cpi_all_items_pct"] / 100.0).tolist()
    n_ipca = int((recorte["diferenca_pp"] > 0).sum())
    n_cpi = int((recorte["diferenca_pp"] < 0).sum())
    return {
        "periodo": rotulo,
        "anos": len(recorte),
        "ipca_acumulado_pct": acumulado(ipca) * 100.0,
        "cpi_acumulado_pct": acumulado(cpi) * 100.0,
        "ipca_media_aritmetica_pct": recorte["ipca_pct"].mean(),
        "cpi_media_aritmetica_pct": recorte["cpi_all_items_pct"].mean(),
        "ipca_media_geometrica_pct": media_geometrica(ipca) * 100.0,
        "cpi_media_geometrica_pct": media_geometrica(cpi) * 100.0,
        "diferenca_media_pp": recorte["diferenca_pp"].mean(),
        "anos_ipca_maior": n_ipca,
        "anos_cpi_maior": n_cpi,
        "maior_diferenca_pp": recorte["diferenca_pp"].max(),
        "menor_diferenca_pp": recorte["diferenca_pp"].min(),
    }


def formatar_pct(valor: float, casas: int = 2) -> str:
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".") + "%"


def formatar_num(valor: float, casas: int = 2) -> str:
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def markdown_cotejamento(anual: pd.DataFrame, resumos: pd.DataFrame) -> str:
    ini = int(anual["ano"].min())
    fim = int(anual["ano"].max())
    linhas = [
        f"# Cotejamento IPCA × CPI All Items ({ini}–{fim})",
        "",
        "Variação no ano: **dezembro/dezembro**.",
        "IPCA: IBGE / Bacen SGS 13522. CPI All Items: BLS CPI-U NSA (FRED `CPIAUCNS`).",
        "Índices: dezembro de 1989 = 100 (base imediata anterior ao primeiro ano).",
        "",
        "| Ano | IPCA | CPI All Items | Diferença (p.p.) | Maior no ano | Índice IPCA | Índice CPI | Razão IPCA/CPI |",
        "|----:|-----:|--------------:|-----------------:|:-------------|------------:|-----------:|---------------:|",
    ]
    for _, r in anual.iterrows():
        linhas.append(
            f"| {int(r['ano'])} | "
            f"{formatar_pct(r['ipca_pct'])} | "
            f"{formatar_pct(r['cpi_all_items_pct'])} | "
            f"{formatar_num(r['diferenca_pp'])} | "
            f"{r['maior_no_ano']} | "
            f"{formatar_num(r['indice_ipca'])} | "
            f"{formatar_num(r['indice_cpi'])} | "
            f"{formatar_num(r['razao_indices'], 3)} |"
        )
    linhas.extend(["", "## Síntese", ""])
    linhas.append(
        "| Período | IPCA acum. | CPI acum. | Média geom. IPCA | Média geom. CPI | "
        "Anos IPCA > CPI | Anos CPI > IPCA |"
    )
    linhas.append(
        "|---------|-----------:|----------:|-----------------:|----------------:|----------------:|----------------:|"
    )
    for _, r in resumos.iterrows():
        linhas.append(
            f"| {r['periodo']} | "
            f"{formatar_pct(r['ipca_acumulado_pct'])} | "
            f"{formatar_pct(r['cpi_acumulado_pct'])} | "
            f"{formatar_pct(r['ipca_media_geometrica_pct'])} | "
            f"{formatar_pct(r['cpi_media_geometrica_pct'])} | "
            f"{int(r['anos_ipca_maior'])} | "
            f"{int(r['anos_cpi_maior'])} |"
        )
    linhas.extend(
        [
            "",
            "A diferença é **IPCA − CPI All Items**, em pontos percentuais.",
            "1990–1994 no Brasil cobrem a hiperinflação e o Plano Real (1º/jul/1994); "
            "o recorte 1995–2025 isola o período em real.",
            "",
        ]
    )
    return "\n".join(linhas)


def _borda(ws):
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin


def _cabecalho(ws, n_cols: int):
    fill = PatternFill("solid", fgColor="1B4F72")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"


def gravar_xlsx(
    anual: pd.DataFrame,
    resumos: pd.DataFrame,
    caminho: Path,
) -> Path:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    titulos = [
        "Ano",
        "IPCA (%)",
        "CPI All Items (%)",
        "Diferença IPCA−CPI (p.p.)",
        "Maior no ano",
        "Índice IPCA (dez/1989=100)",
        "Índice CPI (dez/1989=100)",
        "Razão índices (IPCA/CPI)",
    ]
    ws = wb.active
    ws.title = "Anual_1990_2025"
    for r_idx, row in enumerate(dataframe_to_rows(anual, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    for col, titulo in enumerate(titulos, start=1):
        ws.cell(1, col, titulo)
    fill_ipca = PatternFill("solid", fgColor="F5B7B1")
    fill_cpi = PatternFill("solid", fgColor="AED6F1")
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        for col in (2, 3, 4):
            ws.cell(row, col).number_format = "0.00"
        ws.cell(row, 6).number_format = "#,##0.00"
        ws.cell(row, 7).number_format = "#,##0.00"
        ws.cell(row, 8).number_format = "0.000"
        maior = ws.cell(row, 5).value
        if maior == "IPCA":
            ws.cell(row, 2).fill = fill_ipca
            ws.cell(row, 5).fill = fill_ipca
        elif maior == "CPI All Items":
            ws.cell(row, 3).fill = fill_cpi
            ws.cell(row, 5).fill = fill_cpi
        if (ws.cell(row, 4).value or 0) < 0:
            ws.cell(row, 4).font = Font(color="1B4F72")
        else:
            ws.cell(row, 4).font = Font(color="922B21")
    _cabecalho(ws, 8)
    _borda(ws)
    for i, w in enumerate([8, 12, 18, 24, 16, 26, 24, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Sintese")
    titulos2 = [
        "Período",
        "Anos",
        "IPCA acumulado (%)",
        "CPI All Items acumulado (%)",
        "Média aritmética IPCA (%)",
        "Média aritmética CPI (%)",
        "Média geométrica IPCA (%)",
        "Média geométrica CPI (%)",
        "Diferença média (p.p.)",
        "Anos IPCA > CPI",
        "Anos CPI > IPCA",
        "Maior diferença (p.p.)",
        "Menor diferença (p.p.)",
    ]
    for r_idx, row in enumerate(dataframe_to_rows(resumos, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(r_idx, c_idx, value)
    for col, titulo in enumerate(titulos2, start=1):
        ws2.cell(1, col, titulo)
    for row in range(2, ws2.max_row + 1):
        for col in range(3, 10):
            ws2.cell(row, col).number_format = "#,##0.00"
        for col in (12, 13):
            ws2.cell(row, col).number_format = "0.00"
    _cabecalho(ws2, 13)
    _borda(ws2)
    for i in range(1, 14):
        ws2.column_dimensions[get_column_letter(i)].width = 22
    ws2.column_dimensions["A"].width = 28

    ws3 = wb.create_sheet("Fonte")
    ws3["A1"] = "Fonte e conceito"
    ws3["A1"].font = Font(bold=True, size=14)
    notas = [
        "Cotejamento da inflação oficial ao consumidor: Brasil (IPCA) e Estados Unidos (CPI All Items).",
        "Critério: variação dezembro/dezembro — o mesmo da “inflação no ano” do IBGE e da variação de 12 meses do BLS.",
        "IPCA: IBGE, índice oficial de inflação ao consumidor e da meta do Banco Central. Série Bacen SGS 13522 (acumulado em 12 meses, valor de dezembro).",
        "CPI All Items: BLS, Consumer Price Index for All Urban Consumers: All Items, sem ajuste sazonal (série CUUR0000SA0 / FRED CPIAUCNS).",
        "Diferença em pontos percentuais = IPCA − CPI All Items.",
        "Índices encadeados com base dezembro/1989 = 100, imediatamente anterior a 1990.",
        "O recorte 1995–2025 exclui a hiperinflação e o ano do Plano Real (1º de julho de 1994).",
    ]
    for i, texto in enumerate(notas, start=3):
        ws3[f"A{i}"] = texto
    ws3.column_dimensions["A"].width = 140

    wb.save(caminho)
    return caminho


def gravar(
    anual: pd.DataFrame,
    resumos: pd.DataFrame,
    pasta: Path,
    stem: str = "cotejamento_ipca_cpi_all_items_1990_2025",
) -> dict[str, Path]:
    pasta.mkdir(parents=True, exist_ok=True)
    csv = pasta / f"{stem}.csv"
    md = pasta / f"{stem}.md"
    xlsx = pasta / f"{stem}.xlsx"
    anual.to_csv(csv, index=False, float_format="%.6f")
    md.write_text(markdown_cotejamento(anual, resumos), encoding="utf-8")
    gravar_xlsx(anual, resumos, xlsx)
    return {"csv": csv, "md": md, "xlsx": xlsx}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Cotejamento IPCA × CPI All Items.")
    parser.add_argument("--ano-inicio", type=int, default=ANO_INICIO_DEFAULT)
    parser.add_argument("--ano-fim", type=int, default=ANO_FIM_DEFAULT)
    parser.add_argument("--output-dir", type=Path, default=ROOT / "output")
    args = parser.parse_args(argv)

    anos = anos_periodo(args.ano_inicio, args.ano_fim)
    ipca = baixar_sgs(SGS_IPCA_12M, "01/12/1989", "31/12/2025")
    cpi = baixar_fred(FRED_CPI)
    anual = montar_cotejamento(ipca, cpi, anos)
    resumos = pd.DataFrame(
        [
            resumo_periodo(anual, args.ano_inicio, args.ano_fim, f"{args.ano_inicio}–{args.ano_fim}"),
            resumo_periodo(anual, 1995, args.ano_fim, f"1995–{args.ano_fim} (pós-Real)"),
        ]
    )
    caminhos = gravar(anual, resumos, args.output_dir)
    print(markdown_cotejamento(anual, resumos))
    for nome, path in caminhos.items():
        print(f"[OK] {nome}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
