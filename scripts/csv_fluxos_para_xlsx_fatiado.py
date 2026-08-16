#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Converte ``fluxos_por_ano_contrato/YYYY.csv`` em ``YYYY.xlsx`` fatiado.

O Excel limita ~1.048.576 linhas por aba. Anos com milhões de parcelas
são gravados em várias abas: ``2002_p01``, ``2002_p02``, …

Lê o CSV em streaming (não carrega o ano inteiro na memória) e grava com
openpyxl ``write_only``.

Uso (ContAgil)::

  python sec_scripts\\csv_fluxos_para_xlsx_fatiado.py --pasta saida\\fluxos_por_ano_contrato

  # só um ano / retomar os que já têm xlsx:
  python sec_scripts\\csv_fluxos_para_xlsx_fatiado.py --pasta saida\\fluxos_por_ano_contrato --ano 2011
  python sec_scripts\\csv_fluxos_para_xlsx_fatiado.py --pasta saida\\fluxos_por_ano_contrato --retomar
"""

from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

from openpyxl import Workbook

MARKER = "csv-para-xlsx-fatiado-20260816a"

# Cabeçalho conta como 1 linha → margem abaixo do limite do Excel (1_048_576)
LINHAS_POR_ABA_DEFAULT = 1_000_000
EXCEL_HARD_MAX = 1_048_575  # dados + 1 header

CONTAGIL_CSV = Path(
    r"C:\Arquivos de Programas RFB\ContAgilAppBeta64\python_jep\winpython"
    r"\saida\fluxos_por_ano_contrato"
)


def listar_csvs_ano(pasta: Path) -> list[Path]:
    pasta = Path(pasta)
    if not pasta.is_dir():
        raise FileNotFoundError(f"Pasta não encontrada: {pasta}")
    out = sorted(
        p
        for p in pasta.glob("*.csv")
        if p.stem.isdigit() and len(p.stem) == 4
    )
    if not out:
        raise FileNotFoundError(
            f"Nenhum YYYY.csv em {pasta}. "
            "Rode fluxos_por_ano_contrato_numerados.bat primeiro."
        )
    return out


def _nome_aba(ano: str, parte: int) -> str:
    # Excel: máx. 31 caracteres
    return f"{ano}_p{parte:02d}"[:31]


def converter_csv_para_xlsx(
    csv_path: Path,
    xlsx_path: Path | None = None,
    *,
    linhas_por_aba: int = LINHAS_POR_ABA_DEFAULT,
) -> dict:
    """Converte um CSV de fluxos em XLSX com várias abas.

    Retorna dict com metadados (abas, linhas, segundos, caminho).
    """
    csv_path = Path(csv_path)
    if not csv_path.is_file():
        raise FileNotFoundError(csv_path)

    linhas_por_aba = int(linhas_por_aba)
    if linhas_por_aba < 1:
        raise ValueError("linhas_por_aba deve ser >= 1")
    if linhas_por_aba > EXCEL_HARD_MAX:
        linhas_por_aba = EXCEL_HARD_MAX

    xlsx_path = Path(xlsx_path) if xlsx_path else csv_path.with_suffix(".xlsx")
    ano = csv_path.stem
    t0 = time.time()

    # write_only: não há aba ativa por padrão; create_sheet obrigatório
    wb = Workbook(write_only=True)
    ws = None
    parte = 0
    na_aba = 0
    total = 0
    header: list[str] | None = None
    abas: list[str] = []

    # utf-8-sig cobre CSV com BOM do Excel Windows
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"CSV vazio: {csv_path}") from None

        for row in reader:
            if ws is None or na_aba >= linhas_por_aba:
                if ws is not None:
                    # fecha aba anterior implicitamente ao criar a próxima
                    pass
                parte += 1
                nome = _nome_aba(ano, parte)
                ws = wb.create_sheet(title=nome)
                ws.append(header)
                abas.append(nome)
                na_aba = 0

            ws.append(row)
            na_aba += 1
            total += 1
            if total % 500_000 == 0:
                elapsed = max(time.time() - t0, 1e-6)
                print(
                    f"  [{ano}] {total:,} linhas | "
                    f"aba {parte} ({na_aba:,}) | "
                    f"{total / elapsed:,.0f} lin/s",
                    flush=True,
                )

    if total == 0:
        # só cabeçalho — ainda assim gera uma aba vazia de dados
        parte = 1
        nome = _nome_aba(ano, parte)
        ws = wb.create_sheet(title=nome)
        ws.append(header or [])
        abas.append(nome)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    # grava em arquivo temporário e renomeia (evita xlsx pela metade se falhar)
    tmp = xlsx_path.with_suffix(".xlsx.partial")
    if tmp.exists():
        tmp.unlink()
    wb.save(tmp)
    tmp.replace(xlsx_path)

    secs = round(time.time() - t0, 1)
    mb = xlsx_path.stat().st_size / (1024 * 1024)
    return {
        "ano": ano,
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        "linhas": total,
        "abas": abas,
        "n_abas": len(abas),
        "segundos": secs,
        "mb": round(mb, 1),
    }


def processar_pasta(
    pasta: Path,
    *,
    ano: int | None = None,
    retomar: bool = False,
    linhas_por_aba: int = LINHAS_POR_ABA_DEFAULT,
) -> list[dict]:
    csvs = listar_csvs_ano(pasta)
    if ano is not None:
        csvs = [p for p in csvs if p.stem == str(int(ano))]
        if not csvs:
            raise FileNotFoundError(f"Não há {ano}.csv em {pasta}")

    resultados: list[dict] = []
    for i, csv_p in enumerate(csvs, start=1):
        xlsx_p = csv_p.with_suffix(".xlsx")
        if retomar and xlsx_p.is_file() and xlsx_p.stat().st_size > 0:
            print(f"[{i}/{len(csvs)}] {csv_p.name}: XLSX já existe — pulando (--retomar)")
            continue
        print(f"[{i}/{len(csvs)}] {csv_p.name} → {xlsx_p.name} ...", flush=True)
        try:
            mb_csv = csv_p.stat().st_size / (1024 * 1024)
            print(f"  CSV: {mb_csv:,.1f} MB | até {linhas_por_aba:,} linhas/aba")
        except OSError:
            pass
        info = converter_csv_para_xlsx(
            csv_p, xlsx_p, linhas_por_aba=linhas_por_aba
        )
        print(
            f"  OK {info['ano']}: {info['linhas']:,} linhas | "
            f"{info['n_abas']} aba(s) | {info['mb']:,.1f} MB | {info['segundos']}s"
        )
        print(f"     abas: {', '.join(info['abas'])}")
        resultados.append(info)
    return resultados


def _resolver_pasta(explicit: Path | None) -> Path:
    if explicit is not None:
        return Path(explicit)
    if CONTAGIL_CSV.exists():
        return CONTAGIL_CSV
    cand = Path.cwd() / "saida" / "fluxos_por_ano_contrato"
    if cand.exists():
        return cand
    return Path.cwd() / "fluxos_por_ano_contrato"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--pasta",
        type=Path,
        default=None,
        help="Pasta com YYYY.csv (default: ContAgil saida/fluxos_por_ano_contrato)",
    )
    p.add_argument("--ano", type=int, default=None, help="Converter só este ano")
    p.add_argument(
        "--retomar",
        action="store_true",
        help="Pula anos que já têm .xlsx não vazio",
    )
    p.add_argument(
        "--linhas-por-aba",
        type=int,
        default=LINHAS_POR_ABA_DEFAULT,
        help=f"Linhas de dados por aba (default {LINHAS_POR_ABA_DEFAULT:,}; máx {EXCEL_HARD_MAX:,})",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    pasta = _resolver_pasta(args.pasta)
    print(f"[{MARKER}]")
    print("=" * 70)
    print("CSV → XLSX FATIADO (fluxos por ano de contrato)")
    print(f"Pasta : {pasta}")
    print(f"Linhas/aba: {args.linhas_por_aba:,}")
    print("=" * 70)
    print(
        "AVISO: arquivos .xlsx ficam muito grandes e a conversão pode levar "
        "horas. O CSV continua sendo a fonte completa."
    )
    print()
    try:
        resultados = processar_pasta(
            pasta,
            ano=args.ano,
            retomar=args.retomar,
            linhas_por_aba=args.linhas_por_aba,
        )
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 1

    print()
    print(f"[OK] Convertidos: {len(resultados)} arquivo(s)")
    for r in resultados:
        print(f"  → {r['xlsx']} ({r['n_abas']} abas, {r['linhas']:,} linhas)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
