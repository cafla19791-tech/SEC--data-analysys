"""Custos de captação dos recursos de crédito (BCB, 2001–2026).

Dois recortes oficiais:

1. Indicadores de mercado (Selic, CDI, CDB/RDB, poupança, TBF), disponíveis
   no SGS para o intervalo 2001–2026.
2. Custo de captação referencial por modalidade das estatísticas de crédito:
   taxa média das novas operações minus o spread médio publicado pelo BCB
   (definição oficial da Nota para Imprensa). Série a partir de março de 2011.

Uso:
  python3 scripts/custo_captacao_bcb.py
  python3 scripts/custo_captacao_bcb.py --sem-download
"""

from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.taxas_credito_bcb import (  # noqa: E402
    ANO_FIM,
    ANO_INI,
    DATA_DIR,
    OUTPUT_DIR,
    SERIES,
    SGS_REST,
    _escrever_aba,
    baixar_sgs,
    carregar_series_sgs,
    media_anual,
    taxas_medias_anuais,
)

import requests  # noqa: E402

# Indicadores de mercado. ``unidade``: aa = já em % a.a.; am = % no mês → anualiza.
SERIES_MERCADO = [
    (4189, "Selic over", "aa", "Taxa Selic mensal anualizada (média das taxas diárias)"),
    (4391, "CDI", "am", "CDI acumulado no mês, anualizado por capitalização"),
    (28663, "CDB/RDB pós-fixado", "aa", "Taxa média mensal dos depósitos a prazo pós-fixados"),
    (25, "Caderneta de poupança", "am", "Rentabilidade mensal da poupança (regra vigente)"),
    (256, "TBF", "aa", "Taxa Básica Financeira (referência histórica de CDB prefixado / TR)"),
]


def am_para_aa(taxa_am: float) -> float:
    """Converte percentual ao mês em percentual ao ano, com capitalização."""
    return ((1.0 + float(taxa_am) / 100.0) ** 12 - 1.0) * 100.0


def custo_referencial(taxa_aa: float, spread_pp: float) -> float:
    """Custo de captação referencial do BCB: taxa média minus spread médio."""
    return float(taxa_aa) - float(spread_pp)


def codigo_spread(codigo_taxa: int) -> int | None:
    """Nas estatísticas de crédito, o spread da modalidade 207xx é o código + 69."""
    if 20714 <= int(codigo_taxa) <= 20782:
        return int(codigo_taxa) + 69
    return None


def baixar_sgs_retry(cod: int, inicio: str, fim: str, tentativas: int = 5) -> pd.DataFrame:
    ultimo: Exception | None = None
    for i in range(tentativas):
        try:
            return baixar_sgs(cod, inicio, fim)
        except Exception as exc:  # noqa: BLE001
            ultimo = exc
            time.sleep(1.5 * (i + 1))
            try:
                resp = requests.get(
                    SGS_REST.format(cod=cod),
                    params={"formato": "json", "dataInicial": inicio, "dataFinal": fim},
                    timeout=90,
                )
                resp.raise_for_status()
                dados = resp.json()
                if not dados:
                    return pd.DataFrame(columns=["data", "taxa"])
                out = pd.DataFrame(dados)
                out["data"] = pd.to_datetime(out["data"], dayfirst=True)
                out["taxa"] = pd.to_numeric(out["valor"], errors="coerce")
                return out[["data", "taxa"]].dropna()
            except Exception as exc2:  # noqa: BLE001
                ultimo = exc2
                time.sleep(2.0 * (i + 1))
    raise RuntimeError(f"Falha ao baixar SGS {cod}: {ultimo}")


def _para_mensal_aa(df: pd.DataFrame, unidade: str) -> pd.Series:
    s = df.set_index("data")["taxa"].sort_index()
    if unidade == "am":
        s = s.map(am_para_aa)
    # Séries diárias (poupança): média das observações de cada mês.
    if len(s) > 400:
        s = s.groupby(s.index.to_period("M")).mean()
        s.index = s.index.to_timestamp()
    return s


def carregar_mercado(cache_dir: Path, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / "sgs_captacao_mercado.csv"
    if cache.exists():
        return pd.read_csv(cache, parse_dates=["data"])
    if not baixar:
        raise FileNotFoundError(cache)
    cache_dir.mkdir(parents=True, exist_ok=True)
    partes = []
    for cod, nome, unidade, nota in SERIES_MERCADO:
        print(f"  mercado SGS {cod} {nome}", flush=True)
        df = baixar_sgs_retry(cod, "01/01/2001", "01/12/2026")
        if df.empty:
            print(f"    vazia {cod}", flush=True)
            continue
        df = df.copy()
        df["codigo"] = cod
        df["indicador"] = nome
        df["unidade_origem"] = unidade
        df["nota"] = nota
        partes.append(df)
        time.sleep(0.08)
    if not partes:
        raise RuntimeError("Nenhuma série de mercado de captação foi baixada.")
    out = pd.concat(partes, ignore_index=True)
    out.to_csv(cache, index=False)
    return out


def medias_anuais_mercado(mensal: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for (cod, nome, unidade, nota), g in mensal.groupby(
        ["codigo", "indicador", "unidade_origem", "nota"], sort=False
    ):
        s = _para_mensal_aa(g[["data", "taxa"]], str(unidade))
        # CDI/Selic acumulados no mês corrente ainda incompleto distorcem a anualização.
        if int(cod) in {4390, 4391} and not s.empty:
            hoje = pd.Timestamp.today().normalize()
            if s.index.max().to_period("M") == hoje.to_period("M") and hoje.day < 28:
                s = s.iloc[:-1]
        anual = media_anual(s)
        rec = {
            "codigo": int(cod),
            "indicador": nome,
            "unidade": "% a.a.",
            "inicio": s.index.min().strftime("%m/%Y"),
            "fim": s.index.max().strftime("%m/%Y"),
            "nota": nota,
        }
        for ano in range(ANO_INI, ANO_FIM + 1):
            rec[str(ano)] = float(anual[ano]) if ano in anual.index else float("nan")
        linhas.append(rec)
    return pd.DataFrame(linhas)


def carregar_spreads(cache_dir: Path, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / "sgs_spreads_modalidade.csv"
    if cache.exists():
        return pd.read_csv(cache, parse_dates=["data"])
    if not baixar:
        raise FileNotFoundError(cache)
    cache_dir.mkdir(parents=True, exist_ok=True)
    partes = []
    vistos: set[int] = set()
    for cod, nome, segmento, origem in SERIES:
        sc = codigo_spread(int(cod))
        if sc is None or sc in vistos:
            continue
        vistos.add(sc)
        print(f"  spread SGS {sc} ← taxa {cod} {nome}", flush=True)
        try:
            df = baixar_sgs_retry(sc, "01/03/2011", "01/12/2026")
        except Exception as exc:  # noqa: BLE001
            print(f"    falha {sc}: {exc}", flush=True)
            continue
        if df.empty:
            print(f"    vazia {sc}", flush=True)
            continue
        df = df.copy()
        df["codigo_spread"] = sc
        df["codigo_taxa"] = int(cod)
        df["modalidade"] = nome
        df["segmento"] = segmento
        df["origem"] = origem
        df = df.rename(columns={"taxa": "spread"})
        partes.append(df)
        time.sleep(0.08)
    if not partes:
        raise RuntimeError("Nenhuma série de spread foi baixada.")
    out = pd.concat(partes, ignore_index=True)
    out.to_csv(cache, index=False)
    return out


def custos_mensais(taxas: pd.DataFrame, spreads: pd.DataFrame) -> pd.DataFrame:
    t = taxas.rename(columns={"codigo": "codigo_taxa"})
    cols = ["data", "codigo_taxa", "spread"]
    if "codigo_spread" in spreads.columns:
        cols.append("codigo_spread")
    s = spreads[cols].copy()
    m = t.merge(s, on=["data", "codigo_taxa"], how="inner")
    m["custo"] = m["taxa"] - m["spread"]
    return m


def custos_anuais(mensal: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    chaves = ["codigo_taxa", "codigo_spread", "modalidade", "segmento", "origem"]
    tem_spread_cod = "codigo_spread" in mensal.columns
    if not tem_spread_cod:
        chaves = ["codigo_taxa", "modalidade", "segmento", "origem"]
    for chave, g in mensal.groupby(chaves, sort=False):
        if not isinstance(chave, tuple):
            chave = (chave,)
        rec_keys = dict(zip(chaves, chave))
        s = g.set_index("data")["custo"].sort_index()
        anual = media_anual(s)
        rec = {
            "codigo_taxa": int(rec_keys["codigo_taxa"]),
            "codigo_spread": int(rec_keys["codigo_spread"]) if tem_spread_cod else None,
            "modalidade": rec_keys["modalidade"],
            "segmento": rec_keys["segmento"],
            "origem": rec_keys["origem"],
            "inicio": s.index.min().strftime("%m/%Y"),
            "fim": s.index.max().strftime("%m/%Y"),
        }
        for ano in range(ANO_INI, ANO_FIM + 1):
            rec[str(ano)] = float(anual[ano]) if ano in anual.index else float("nan")
        linhas.append(rec)
    return pd.DataFrame(linhas)


def comparativo_anual(taxas_a: pd.DataFrame, spreads_m: pd.DataFrame, custos_a: pd.DataFrame) -> pd.DataFrame:
    """Empilha taxa, spread e custo nas mesmas modalidades/anos."""
    sp = spreads_m.copy()
    if "codigo" in sp.columns and "codigo_spread" in sp.columns:
        sp = sp.drop(columns=["codigo"])
    spread_a = taxas_medias_anuais(
        sp.rename(columns={"spread": "taxa", "codigo_spread": "codigo"})[
            ["data", "taxa", "codigo", "modalidade", "segmento", "origem"]
        ]
    )
    anos = [str(a) for a in range(ANO_INI, ANO_FIM + 1)]
    partes = []
    for rotulo, df, col_cod in (
        ("Taxa de crédito", taxas_a, "codigo"),
        ("Spread", spread_a, "codigo"),
        ("Custo de captação", custos_a, "codigo_taxa"),
    ):
        tmp = df.copy()
        tmp["indicador"] = rotulo
        tmp = tmp.rename(columns={col_cod: "codigo"})
        partes.append(tmp[["codigo", "modalidade", "segmento", "origem", "indicador"] + anos])
    return pd.concat(partes, ignore_index=True)


def gerar_planilha(
    mercado: pd.DataFrame,
    referencial: pd.DataFrame,
    comparativo: pd.DataFrame,
    path: Path,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Notas"
    ws0["A1"] = "Custo de captação dos recursos de crédito no Brasil — 2001–2026"
    ws0["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    notas = [
        "O Banco Central não publica, por parcela, o custo de captação de cada banco. "
        "O que existe no SGS é (a) o custo de mercado dos instrumentos de funding e "
        "(b) o custo de captação referencial das estatísticas de crédito, definido como "
        "a taxa média das novas operações minus o spread médio da mesma modalidade.",
        "Aba Mercado: média aritmética, em cada ano-calendário, das taxas oficiais de "
        "Selic (SGS 4189, % a.a.), CDI (SGS 4391, acumulada no mês e anualizada), "
        "CDB/RDB pós-fixado (SGS 28663, % a.a.), caderneta de poupança (SGS 25, "
        "rentabilidade mensal anualizada) e TBF (SGS 256, % a.a.).",
        "A série de CDB/RDB 28663 termina em janeiro de 2024 — o BCB não republicou "
        "essa taxa média depois dessa data. A partir de 2024 o CDI (e a Selic) é o "
        "referencial de mercado dos depósitos pós-fixados.",
        "Aba Custo_referencial: para cada modalidade da Nota para Imprensa, "
        "custo = taxa SGS 207xx − spread SGS (mesmo código + 69). A metodologia "
        "atual começa em março de 2011. O custo dos recursos livres acompanha a "
        "Selic/CDI; o dos recursos direcionados acompanha a poupança e os fundos públicos.",
        "Não há série oficial de spread (logo, nem de custo referencial) para as "
        "modalidades de cartão 22021–22024. O funding desses produtos entra no "
        "agregado de recursos livres. 2026 usa os meses já publicados (crédito até "
        "junho; Selic/CDI até o último mês completo).",
        "Fontes: https://www.bcb.gov.br/estatisticas/estatisticasmonetariascredito  |  "
        "https://www3.bcb.gov.br/sgspub/",
    ]
    for i, txt in enumerate(notas, start=3):
        ws0[f"A{i}"] = txt
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
        ws0[f"A{i}"].alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[i].height = 42
    ws0.column_dimensions["A"].width = 28

    anos = [str(a) for a in range(ANO_INI, ANO_FIM + 1)]
    cabs_m = ["Código SGS", "Indicador", "Unidade", "Início", "Fim", "Nota"] + anos
    linhas_m = []
    for rec in mercado.to_dict("records"):
        lin = [int(rec["codigo"]), rec["indicador"], rec["unidade"], rec["inicio"], rec["fim"], rec["nota"]]
        for a in anos:
            v = rec.get(a)
            lin.append(None if v is None or pd.isna(v) else float(v))
        linhas_m.append(lin)
    _escrever_aba(
        wb.create_sheet("Mercado"),
        cabs_m,
        linhas_m,
        [12, 24, 10, 10, 10, 62] + [9] * len(anos),
    )

    cabs_r = [
        "Código taxa",
        "Código spread",
        "Modalidade",
        "Segmento",
        "Origem",
        "Início",
        "Fim",
    ] + anos
    linhas_r = []
    for rec in referencial.to_dict("records"):
        lin = [
            int(rec["codigo_taxa"]),
            None if rec.get("codigo_spread") is None or pd.isna(rec.get("codigo_spread")) else int(rec["codigo_spread"]),
            rec["modalidade"],
            rec["segmento"],
            rec["origem"],
            rec["inicio"],
            rec["fim"],
        ]
        for a in anos:
            v = rec.get(a)
            lin.append(None if v is None or pd.isna(v) else float(v))
        linhas_r.append(lin)
    _escrever_aba(
        wb.create_sheet("Custo_referencial"),
        cabs_r,
        linhas_r,
        [12, 14, 46, 10, 18, 10, 10] + [9] * len(anos),
    )

    cabs_c = ["Código", "Modalidade", "Segmento", "Origem", "Indicador"] + anos
    linhas_c = []
    for rec in comparativo.to_dict("records"):
        lin = [int(rec["codigo"]), rec["modalidade"], rec["segmento"], rec["origem"], rec["indicador"]]
        for a in anos:
            v = rec.get(a)
            lin.append(None if v is None or pd.isna(v) else float(v))
        linhas_c.append(lin)
    _escrever_aba(
        wb.create_sheet("Taxa_spread_custo"),
        cabs_c,
        linhas_c,
        [12, 46, 10, 18, 22] + [9] * len(anos),
    )

    # Destaque da linha de cabeçalho das notas
    fill = PatternFill("solid", fgColor="D6EAF8")
    ws0["A1"].fill = fill
    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR / "custo_captacao_bcb_2001_2026.xlsx")
    parser.add_argument("--sem-download", action="store_true")
    args = parser.parse_args(argv)

    print("Indicadores de mercado de captação...", flush=True)
    mercado_m = carregar_mercado(args.cache_dir, baixar=not args.sem_download)
    mercado_a = medias_anuais_mercado(mercado_m)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    mercado_a.to_csv(args.output.parent / "custo_captacao_mercado_2001_2026.csv", index=False)

    print("Taxas e spreads oficiais por modalidade...", flush=True)
    taxas_m = carregar_series_sgs(args.cache_dir, baixar=not args.sem_download)
    taxas_m = taxas_m.rename(columns={"codigo": "codigo_taxa"}) if "codigo_taxa" not in taxas_m.columns else taxas_m
    if "codigo" in taxas_m.columns and "codigo_taxa" not in taxas_m.columns:
        taxas_m = taxas_m.rename(columns={"codigo": "codigo_taxa"})
    spreads_m = carregar_spreads(args.cache_dir, baixar=not args.sem_download)
    mensal = custos_mensais(taxas_m, spreads_m)
    if mensal.empty:
        raise RuntimeError("Não foi possível cruzar taxa e spread.")
    ref_a = custos_anuais(mensal)
    ref_a.to_csv(args.output.parent / "custo_captacao_referencial_2001_2026.csv", index=False)

    taxas_a = taxas_medias_anuais(taxas_m.rename(columns={"codigo_taxa": "codigo"}))
    comp = comparativo_anual(taxas_a, spreads_m, ref_a)
    comp.to_csv(args.output.parent / "custo_captacao_taxa_spread_2001_2026.csv", index=False)

    path = gerar_planilha(mercado_a, ref_a, comp, args.output)
    print(f"Planilha: {path}")
    print(mercado_a[["indicador", "inicio", "fim", "2024", "2025", "2026"]].to_string(index=False))
    cols = [c for c in ("modalidade", "inicio", "fim", "2011", "2024", "2025", "2026") if c in ref_a.columns]
    print(ref_a[cols].to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
