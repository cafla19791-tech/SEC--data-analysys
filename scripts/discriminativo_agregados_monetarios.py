#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Discriminativo dos agregados monetários M1, M2, M3 e M4.

Séries SGS «Novo» (revisão metodológica de 2018, histórico desde dez/2001):

* 27791 M1 — papel-moeda em poder do público + depósitos à vista
* 27810 M2 — M1 + poupança + títulos privados de instituições depositárias
* 27813 M3 — M2 + quotas de fundos + compromissadas
* 27815 M4 — M3 + títulos públicos federais em poder do público

Pedido: janeiro/2001 a julho/2026. A metodologia oficial só publica o
quarteto completo a partir de dez/2001 (M1 em geral desde jan/2002).

Uso::

  python scripts/discriminativo_agregados_monetarios.py
"""

from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

MARKER = "agregados-m1-m4-20260830"
BCB_SGS = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
ESCALA_MILHOES = 1_000.0
MENOS = "\u2212"
FMT_NUM = f'#,##0.0;"{MENOS}"#,##0.0;"—"'
FMT_PCT = f'0.00%;"{MENOS}"0.00%;"—"'
VERMELHO_TEXTO = "9B1B1B"
VERMELHO_FUNDO = "FFC7CE"


@dataclass(frozen=True)
class Serie:
    codigo: int
    nome: str
    papel: str  # agregado | componente


SERIES: tuple[Serie, ...] = (
    Serie(27791, "M1", "agregado"),
    Serie(27810, "M2", "agregado"),
    Serie(27813, "M3", "agregado"),
    Serie(27815, "M4", "agregado"),
    Serie(27789, "Papel-moeda em poder do público", "componente"),
    Serie(27790, "Depósitos à vista", "componente"),
)

AGREGADOS = ("M1", "M2", "M3", "M4")
CODIGO_AGREGADO = {s.nome: s.codigo for s in SERIES if s.papel == "agregado"}


def _http_get(url: str, params: dict, tentativas: int = 5) -> list:
    ultimo: Exception | None = None
    headers = {"User-Agent": f"SEC-data-analysys/{MARKER}"}
    for i in range(tentativas):
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=120)
            if resp.status_code == 404:
                return []
            if resp.status_code != 200 or not resp.text.strip():
                raise RuntimeError(f"HTTP {resp.status_code} vazio")
            if resp.text.lstrip().startswith("<") or resp.text.lstrip().startswith("<?"):
                raise RuntimeError("resposta XML/HTML")
            dados = resp.json()
            if not isinstance(dados, list):
                raise RuntimeError(f"JSON inesperado: {type(dados)}")
            return dados
        except Exception as exc:  # noqa: BLE001
            ultimo = exc
            time.sleep(1.5 * (2**i))
    raise RuntimeError(f"falha ao baixar {url} {params}: {ultimo}") from ultimo


def baixar_sgs(
    codigo: int,
    inicio: date,
    fim: date,
    *,
    cache: Path | None = None,
    usar_cache: bool = True,
) -> pd.DataFrame:
    if cache is not None and usar_cache and cache.exists() and cache.stat().st_size > 20:
        df = pd.read_csv(cache, parse_dates=["mes"])
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        return df.dropna(subset=["mes"]).sort_values("mes").reset_index(drop=True)

    url = BCB_SGS.format(cod=codigo)
    partes: list[pd.DataFrame] = []
    cursor = pd.Timestamp(inicio)
    fim_ts = pd.Timestamp(fim)
    while cursor <= fim_ts:
        bloco_fim = min(cursor + pd.DateOffset(years=8, months=11), fim_ts)
        dados = _http_get(
            url,
            {
                "formato": "json",
                "dataInicial": cursor.strftime("%d/%m/%Y"),
                "dataFinal": bloco_fim.strftime("%d/%m/%Y"),
            },
        )
        if dados:
            df = pd.DataFrame(dados)
            df["mes"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
            partes.append(df[["mes", "valor"]].dropna())
        cursor = bloco_fim + pd.DateOffset(days=1)
        time.sleep(0.2)

    if not partes:
        out = pd.DataFrame(columns=["mes", "valor"])
    else:
        out = (
            pd.concat(partes, ignore_index=True)
            .drop_duplicates("mes")
            .sort_values("mes")
            .reset_index(drop=True)
        )
    if cache is not None:
        cache.parent.mkdir(parents=True, exist_ok=True)
        out.to_csv(cache, index=False)
    return out


def carregar_painel(
    pasta_cache: Path,
    *,
    usar_cache: bool = True,
    arquivos: dict[int, Path] | None = None,
    inicio: date = date(2001, 1, 1),
    fim: date = date(2026, 7, 31),
    series: Iterable[Serie] = SERIES,
) -> pd.DataFrame:
    pasta_cache.mkdir(parents=True, exist_ok=True)
    cols: dict[str, pd.Series] = {}
    for s in series:
        if arquivos is not None:
            if s.codigo not in arquivos:
                continue
            df = pd.read_csv(arquivos[s.codigo], parse_dates=["mes"])
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        else:
            print(f"[SGS] {s.codigo} {s.nome}")
            df = baixar_sgs(
                s.codigo,
                inicio,
                fim,
                cache=pasta_cache / f"sgs_{s.codigo}.csv",
                usar_cache=usar_cache,
            )
        if df.empty:
            continue
        ser = df.set_index("mes")["valor"]
        ser.index = pd.to_datetime(ser.index)
        cols[s.nome] = ser
    if not cols:
        raise RuntimeError("nenhuma série SGS carregada")
    painel = pd.DataFrame(cols).sort_index()
    painel.index.name = "mes"
    return painel


def tabela_mensal(painel: pd.DataFrame) -> pd.DataFrame:
    """Estoque mensal + camadas (M2−M1 …) + variação % 12 meses."""
    out = pd.DataFrame(index=painel.index)
    out["Mês"] = painel.index
    for nome in AGREGADOS:
        if nome in painel.columns:
            out[nome] = painel[nome]
    if {"M1", "M2"}.issubset(out.columns):
        out["M2 − M1"] = out["M2"] - out["M1"]
    if {"M2", "M3"}.issubset(out.columns):
        out["M3 − M2"] = out["M3"] - out["M2"]
    if {"M3", "M4"}.issubset(out.columns):
        out["M4 − M3"] = out["M4"] - out["M3"]
    for nome in AGREGADOS:
        if nome in out.columns:
            out[f"{nome} Δ% 12m"] = out[nome].pct_change(12)
            out[f"{nome} Δ% mês"] = out[nome].pct_change(1)
    out = out.reset_index(drop=True)
    return out


def tabela_anual(mensal: pd.DataFrame) -> pd.DataFrame:
    """Último mês de cada ano (dezembro, ou julho em 2026)."""
    tmp = mensal.copy()
    tmp["ano"] = pd.to_datetime(tmp["Mês"]).dt.year
    tmp["mes_num"] = pd.to_datetime(tmp["Mês"]).dt.month
    idx = tmp.groupby("ano")["mes_num"].idxmax()
    return tmp.loc[idx].reset_index(drop=True)


def tabela_composicao(painel: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in ("Papel-moeda em poder do público", "Depósitos à vista", "M1") if c in painel.columns]
    if not cols:
        return pd.DataFrame()
    out = painel[cols].copy()
    if {"Papel-moeda em poder do público", "Depósitos à vista"}.issubset(out.columns):
        out["PMPP + depósitos à vista"] = (
            out["Papel-moeda em poder do público"] + out["Depósitos à vista"]
        )
        if "M1" in out.columns:
            out["Resíduo M1"] = out["M1"] - out["PMPP + depósitos à vista"]
    out = out.reset_index()
    return out


# --- Excel -----------------------------------------------------------------

AZUL = "1F4E79"
AZUL_CLARO = "D6E3F0"
DOURADO = "FFF2CC"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
COLS_PCT_SUFIXOS = ("Δ% 12m", "Δ% mês")


def _cab(ws: Worksheet, headers: list[str], linha: int) -> None:
    fill = PatternFill("solid", fgColor=AZUL)
    font = Font(color=BRANCO, bold=True, name="Calibri", size=9)
    for col, nome in enumerate(headers, start=1):
        cell = ws.cell(linha, col, nome)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _fmt_milhoes(valor: float) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    return float(valor) / ESCALA_MILHOES


def _pintar(cell, valor, *, pct: bool = False) -> None:
    cell.border = THIN
    cell.alignment = Alignment(horizontal="center")
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        cell.value = None
        return
    v = float(valor)
    if pct:
        cell.value = v
        cell.number_format = FMT_PCT
    else:
        cell.value = _fmt_milhoes(v)
        cell.number_format = FMT_NUM
        v = cell.value if cell.value is not None else 0.0
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return
    if float(v) < -1e-9:
        cell.font = Font(name="Calibri", size=8, bold=True, color=VERMELHO_TEXTO)
        cell.fill = PatternFill("solid", fgColor=VERMELHO_FUNDO)
    else:
        cell.font = Font(name="Calibri", size=8)


def _aba_metodologia(wb: Workbook, gerado_em: datetime, primeiro: pd.Timestamp, ultimo: pd.Timestamp) -> None:
    ws = wb.active
    ws.title = "Metodologia"
    ws["A1"] = "Discriminativo — agregados monetários M1, M2, M3 e M4"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=AZUL)
    ws.merge_cells("A1:B1")
    blocos = [
        ("Gerado em", gerado_em.strftime("%d/%m/%Y %H:%M")),
        ("Marker", MARKER),
        ("Fonte", "Banco Central do Brasil — SGS, séries «Novo» (revisão de ago/2018)"),
        ("Unidade", "R$ milhões (SGS em R$ mil ÷ 1.000)"),
        ("Pedido", "Janeiro/2001 a julho/2026"),
        ("Cobertura efetiva", f"{primeiro.strftime('%m/%Y')} a {ultimo.strftime('%m/%Y')}"),
        (
            "Início em 2001",
            "A revisão metodológica (Nota Técnica nº 48/2018) republicou o "
            "histórico desde dezembro/2001. O M1 (27791) entra em geral em "
            "janeiro/2002; M2/M3/M4 têm ponto em dezembro/2001. "
            "Janeiro–novembro/2001 não existem na metodologia vigente — "
            "não se misturam séries antigas, que não são comparáveis.",
        ),
        (
            "M1 (27791)",
            "Papel-moeda em poder do público + depósitos à vista "
            "(inclui cooperativas de crédito).",
        ),
        (
            "M2 (27810)",
            "M1 + depósitos de poupança + títulos privados emitidos por "
            "instituições depositárias.",
        ),
        (
            "M3 (27813)",
            "M2 + quotas de fundos de investimento depositários + "
            "operações compromissadas com títulos públicos e privados.",
        ),
        (
            "M4 (27815)",
            "M3 + títulos públicos federais em poder do público (Selic).",
        ),
        (
            "Camadas",
            "M2 − M1, M3 − M2 e M4 − M3 isolam o que cada agregado acrescenta. "
            "Identidade: M1 + (M2−M1) = M2, e analogamente até M4.",
        ),
        (
            "Variações",
            "Δ% mês = M_t / M_{t−1} − 1; Δ% 12m = M_t / M_{t−12} − 1. "
            "Negativos: sinal − vermelho negrito com fundo vermelho.",
        ),
        (
            "Abas",
            "Discriminativo — painel mensal. Anual — último mês de cada ano. "
            "Composição_M1 — PMPP + depósitos à vista. Grafico — estoques.",
        ),
    ]
    _cab(ws, ["Campo", "Descrição"], 3)
    for i, (campo, desc) in enumerate(blocos, start=4):
        ws.cell(i, 1, campo).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(i, 2, desc).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 48 if len(desc) > 80 else 20
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 112
    ws.freeze_panes = "A4"


def _escrever_painel(
    ws: Worksheet,
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str,
    *,
    linha0: int = 4,
) -> None:
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    n_cols = df.shape[1]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Calibri", size=9, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 32

    vis = df.copy()
    if "Mês" in vis.columns:
        vis["Mês"] = pd.to_datetime(vis["Mês"]).dt.strftime("%Y-%m")
    if "ano" in vis.columns:
        vis = vis.drop(columns=["ano", "mes_num"], errors="ignore")
    headers = list(vis.columns)
    _cab(ws, headers, linha0)
    ouro = {f"M{i}" for i in (1, 2, 3, 4)}
    for i, row in enumerate(vis.itertuples(index=False), start=linha0 + 1):
        for col, (nome, valor) in enumerate(zip(headers, row), start=1):
            cell = ws.cell(i, col)
            if nome == "Mês":
                cell.value = valor
                cell.border = THIN
                cell.font = Font(name="Calibri", size=8, bold=True)
                cell.alignment = Alignment(horizontal="center")
                continue
            pct = any(nome.endswith(s) for s in COLS_PCT_SUFIXOS)
            _pintar(cell, valor, pct=pct)
            if nome in ouro:
                cell.font = Font(name="Calibri", size=8, bold=True)
                if not (isinstance(valor, float) and pd.notna(valor) and float(valor) < 0):
                    cell.fill = PatternFill("solid", fgColor=AZUL_CLARO if nome == "M4" else CINZA)
    ws.freeze_panes = f"B{linha0 + 1}"
    ws.auto_filter.ref = f"A{linha0}:{get_column_letter(n_cols)}{linha0 + len(vis)}"
    ws.row_dimensions[linha0].height = 30
    ws.column_dimensions["A"].width = 12
    for j in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _aba_grafico(wb: Workbook, mensal: pd.DataFrame) -> None:
    ws = wb.create_sheet("Grafico")
    ws["A1"] = "Estoque dos agregados M1–M4 (R$ milhões)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    tmp = mensal[["Mês"] + [c for c in AGREGADOS if c in mensal.columns]].copy()
    tmp["Mês"] = pd.to_datetime(tmp["Mês"])
    # amostra anual (dezembro) para o gráfico não ficar ilegível
    anual = tmp[tmp["Mês"].dt.month == 12].copy()
    if tmp["Mês"].max().month != 12:
        anual = pd.concat([anual, tmp.tail(1)], ignore_index=True)
    headers = ["Mês"] + [c for c in AGREGADOS if c in anual.columns]
    _cab(ws, headers, 3)
    for i, row in enumerate(anual.itertuples(index=False), start=4):
        ws.cell(i, 1, row[0].strftime("%Y-%m")).border = THIN
        for j in range(2, len(headers) + 1):
            cell = ws.cell(i, j, _fmt_milhoes(row[j - 1]))
            cell.number_format = FMT_NUM
            cell.border = THIN
    last = 3 + len(anual)
    chart = LineChart()
    chart.title = "M1, M2, M3 e M4 — estoque em dezembro (R$ milhões)"
    chart.style = 10
    chart.y_axis.title = "R$ milhões"
    chart.height = 12
    chart.width = 22
    data = Reference(ws, min_col=2, max_col=len(headers), min_row=3, max_row=last)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A20")
    ws.column_dimensions["A"].width = 12
    for j in range(2, 6):
        ws.column_dimensions[get_column_letter(j)].width = 16


def escrever_planilha(
    *,
    mensal: pd.DataFrame,
    anual: pd.DataFrame,
    composicao: pd.DataFrame,
    saida: Path,
    primeiro: pd.Timestamp,
    ultimo: pd.Timestamp,
) -> Path:
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _aba_metodologia(wb, datetime.now(), primeiro, ultimo)
    ws_d = wb.create_sheet("Discriminativo", 1)
    _escrever_painel(
        ws_d,
        mensal,
        "Discriminativo — agregados monetários M1, M2, M3 e M4 (saldo em final de período)",
        f"Mensal {primeiro.strftime('%m/%Y')} a {ultimo.strftime('%m/%Y')}. "
        "R$ milhões. Camadas M2−M1, M3−M2 e M4−M3. "
        "Δ% 12m e Δ% mês. Negativos em − vermelho negrito.",
    )
    ws_a = wb.create_sheet("Anual")
    _escrever_painel(
        ws_a,
        anual,
        "Agregados em final de cada ano (dezembro; 2026 = último mês)",
        "Mesmo conteúdo do discriminativo, uma linha por ano. R$ milhões.",
    )
    if not composicao.empty:
        ws_c = wb.create_sheet("Composicao_M1")
        comp = composicao.copy()
        if "mes" in comp.columns:
            comp = comp.rename(columns={"mes": "Mês"})
        _escrever_painel(
            ws_c,
            comp,
            "Composição do M1 — papel-moeda em poder do público + depósitos à vista",
            "Identidade: PMPP + depósitos à vista = M1 (resíduo de arredondamento do SGS).",
        )
    _aba_grafico(wb, mensal)
    wb.save(saida)
    print(f"[OK] Planilha: {saida} ({saida.stat().st_size / 1024:.1f} KB)")
    return saida


def processar(
    *,
    pasta_cache: Path,
    saida: Path,
    usar_cache: bool = True,
    arquivos: dict[int, Path] | None = None,
) -> Path:
    print(f"[{MARKER}]")
    painel = carregar_painel(pasta_cache, usar_cache=usar_cache, arquivos=arquivos)
    mensal = tabela_mensal(painel)
    # só linhas com pelo menos um agregado
    tem = mensal[[c for c in AGREGADOS if c in mensal.columns]].notna().any(axis=1)
    mensal = mensal.loc[tem].reset_index(drop=True)
    anual = tabela_anual(mensal)
    composicao = tabela_composicao(painel)
    primeiro = pd.to_datetime(mensal["Mês"].min())
    ultimo = pd.to_datetime(mensal["Mês"].max())
    print(f"[INFO] {primeiro.date()} → {ultimo.date()} | {len(mensal)} meses")
    path = escrever_planilha(
        mensal=mensal,
        anual=anual,
        composicao=composicao,
        saida=saida,
        primeiro=primeiro,
        ultimo=ultimo,
    )
    print("\n=== Estoque (R$ milhões) — recortes ===")
    for ts in (mensal["Mês"].min(), pd.Timestamp("2010-12-01"), pd.Timestamp("2020-12-01"), mensal["Mês"].max()):
        fatia = mensal[pd.to_datetime(mensal["Mês"]) == pd.Timestamp(ts)]
        if fatia.empty:
            continue
        r = fatia.iloc[0]
        print(
            f"  {pd.Timestamp(r['Mês']).strftime('%m/%Y')}  "
            + "  ".join(f"{n}={_fmt_milhoes(r[n]):,.1f}" for n in AGREGADOS if n in r.index and pd.notna(r[n]))
        )
    return path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pasta-cache", type=Path, default=ROOT / "data" / "sgs")
    p.add_argument(
        "--saida",
        type=Path,
        default=ROOT / "output" / "discriminativo_agregados_monetarios_m1_m4.xlsx",
    )
    p.add_argument("--sem-cache", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        processar(
            pasta_cache=args.pasta_cache,
            saida=args.saida,
            usar_cache=not args.sem_cache,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
