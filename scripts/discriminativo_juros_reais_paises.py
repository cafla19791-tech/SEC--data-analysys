#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Discriminativo de taxas básicas de juros reais acumuladas e CAGR.

Compara as taxas básicas (policy rates) do BIS, deflacionadas pelo IPC
mensal do próprio BIS, em cinco recortes solicitados. Cada período vira
uma aba, com ranking em ordem decrescente da taxa básica real acumulada.

Uso::

  python scripts/discriminativo_juros_reais_paises.py
  python discriminativo_juros_reais_paises.py --saida output/discriminativo.xlsx
"""

from __future__ import annotations

import argparse
import calendar
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from io import StringIO
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

MARKER = "discriminativo-juros-reais-paises-20260830"
DIAS_ANO = 365.25
COBERTURA_MINIMA = 0.80

# Membros da zona do euro presentes no BIS CBPOL. Após o fim da série
# nacional, a taxa do BCE (XM) é usada com o IPC do próprio país.
EURO_MEMBROS = frozenset(
    {"AT", "BE", "DE", "ES", "FR", "GR", "HR", "IT", "NL", "PT"}
)

BIS_CBPOL = (
    "https://stats.bis.org/api/v2/data/dataflow/BIS/WS_CBPOL/1.0/M"
    "?format=csvfile&startPeriod=1994-01"
)
BIS_CPI = (
    "https://stats.bis.org/api/v2/data/dataflow/BIS/WS_LONG_CPI/1.0/M"
    "?format=csvfile&startPeriod=1993-01"
)

# 628 = índice de preços ao consumidor; 771 = variação em 12 meses (%).
CPI_UNIDADE_INDICE = 628
CPI_UNIDADE_YOY = 771

PAISES: dict[str, str] = {
    "AR": "Argentina",
    "AT": "Áustria",
    "AU": "Austrália",
    "BE": "Bélgica",
    "BR": "Brasil",
    "CA": "Canadá",
    "CH": "Suíça",
    "CL": "Chile",
    "CN": "China",
    "CO": "Colômbia",
    "CZ": "Tchéquia",
    "DE": "Alemanha",
    "DK": "Dinamarca",
    "ES": "Espanha",
    "FR": "França",
    "GB": "Reino Unido",
    "GR": "Grécia",
    "HK": "Hong Kong",
    "HR": "Croácia",
    "HU": "Hungria",
    "ID": "Indonésia",
    "IL": "Israel",
    "IN": "Índia",
    "IS": "Islândia",
    "IT": "Itália",
    "JP": "Japão",
    "KR": "Coreia do Sul",
    "KW": "Kuwait",
    "MA": "Marrocos",
    "MK": "Macedônia do Norte",
    "MX": "México",
    "MY": "Malásia",
    "NL": "Países Baixos",
    "NO": "Noruega",
    "NZ": "Nova Zelândia",
    "PE": "Peru",
    "PH": "Filipinas",
    "PL": "Polônia",
    "PT": "Portugal",
    "RO": "Romênia",
    "RS": "Sérvia",
    "RU": "Rússia",
    "SA": "Arábia Saudita",
    "SE": "Suécia",
    "TH": "Tailândia",
    "TR": "Turquia",
    "US": "Estados Unidos",
    "XM": "Zona do Euro",
    "ZA": "África do Sul",
}


@dataclass(frozen=True)
class Periodo:
    id: int
    inicio: date
    fim: date
    aba: str
    titulo: str


PERIODOS: tuple[Periodo, ...] = (
    Periodo(
        1,
        date(1995, 1, 1),
        date(2002, 12, 31),
        "1_1995-01_2002-12",
        "01/01/1995 a 31/12/2002",
    ),
    Periodo(
        2,
        date(2003, 1, 1),
        date(2016, 5, 11),
        "2_2003-01_2016-05-11",
        "01/01/2003 a 11/05/2016",
    ),
    Periodo(
        3,
        date(2016, 5, 12),
        date(2018, 12, 31),
        "3_2016-05-12_2018-12",
        "12/05/2016 a 31/12/2018",
    ),
    Periodo(
        4,
        date(2019, 1, 1),
        date(2022, 12, 31),
        "4_2019-01_2022-12",
        "01/01/2019 a 31/12/2022",
    ),
    Periodo(
        5,
        date(2023, 1, 1),
        date(2026, 8, 28),
        "5_2023-01_2026-08-28",
        "01/01/2023 a 28/08/2026",
    ),
)


def nome_pais(codigo: str) -> str:
    return PAISES.get(codigo, codigo)


def ultimo_dia_mes(ano: int, mes: int) -> date:
    return date(ano, mes, calendar.monthrange(ano, mes)[1])


def mes_ref(d: date) -> date:
    return date(d.year, d.month, 1)


def mes_anterior(d: date) -> date:
    if d.month == 1:
        return date(d.year - 1, 12, 1)
    return date(d.year, d.month - 1, 1)


def meses_sobrepostos(inicio: date, fim: date) -> list[date]:
    """Primeiro dia de cada mês que intersecta [inicio, fim]."""
    atual = mes_ref(inicio)
    ultimo = mes_ref(fim)
    out: list[date] = []
    while atual <= ultimo:
        out.append(atual)
        if atual.month == 12:
            atual = date(atual.year + 1, 1, 1)
        else:
            atual = date(atual.year, atual.month + 1, 1)
    return out


def intersecao_mes(mes: date, inicio: date, fim: date) -> tuple[date, date] | None:
    m0 = mes_ref(mes)
    m1 = ultimo_dia_mes(m0.year, m0.month)
    a = max(inicio, m0)
    b = min(fim, m1)
    if a > b:
        return None
    return a, b


def dias_inclusive(a: date, b: date) -> int:
    return (b - a).days + 1


def taxa_real_fisher(nominal_aa: float, inflacao: float) -> float:
    """Fisher: (1+i)/(1+π) − 1. Entradas em fração (0,10 = 10%)."""
    if inflacao <= -0.999999:
        raise ValueError("inflação inválida para Fisher")
    return (1.0 + nominal_aa) / (1.0 + inflacao) - 1.0


def fator_proporcional(taxa_aa: float, dias: int, base: float = DIAS_ANO) -> float:
    """Capitaliza taxa anual efetiva por ``dias``/``base``."""
    return (1.0 + taxa_aa) ** (dias / base)


def cagr(fator: float, dias: int, base: float = DIAS_ANO) -> float:
    if dias <= 0 or fator <= 0:
        return float("nan")
    return fator ** (base / dias) - 1.0


def anos_fracao_meses(inicio: date, fim: date) -> float:
    """Duração em anos com pró-rata mensal (mês cheio = 1/12)."""
    anos = 0.0
    for mes in meses_sobrepostos(inicio, fim):
        trecho = intersecao_mes(mes, inicio, fim)
        if trecho is None:
            continue
        dias = dias_inclusive(*trecho)
        dias_mes = dias_inclusive(mes_ref(mes), ultimo_dia_mes(mes.year, mes.month))
        anos += (dias / dias_mes) / 12.0
    return anos


def cagr_meses(fator: float, anos: float) -> float:
    if anos <= 0 or fator <= 0:
        return float("nan")
    return fator ** (1.0 / anos) - 1.0


def parse_ano_mes(valor: object) -> date | None:
    texto = str(valor).strip()
    if len(texto) < 7:
        return None
    try:
        ano = int(texto[:4])
        mes = int(texto[5:7])
        return date(ano, mes, 1)
    except ValueError:
        return None


def _http_get(url: str, tentativas: int = 4) -> str:
    ultimo: Exception | None = None
    headers = {"User-Agent": f"SEC-data-analysys/{MARKER}"}
    for i in range(tentativas):
        try:
            resp = requests.get(url, headers=headers, timeout=120)
            resp.raise_for_status()
            if not resp.text or resp.text.lstrip().startswith("<?xml"):
                raise RuntimeError(f"resposta inválida de {url[:80]}")
            return resp.text
        except Exception as exc:  # noqa: BLE001 — retry de rede
            ultimo = exc
            time.sleep(2 ** i)
    raise RuntimeError(f"falha ao baixar {url}: {ultimo}") from ultimo


def _ler_csv_bis(texto: str) -> pd.DataFrame:
    return pd.read_csv(StringIO(texto))


def baixar_ou_cache(url: str, destino: Path, *, usar_cache: bool = True) -> pd.DataFrame:
    destino.parent.mkdir(parents=True, exist_ok=True)
    if usar_cache and destino.exists() and destino.stat().st_size > 1000:
        print(f"[CACHE] {destino.name} ({destino.stat().st_size:,} bytes)")
        return pd.read_csv(destino)
    print(f"[BAIXAR] {url}")
    texto = _http_get(url)
    destino.write_text(texto, encoding="utf-8")
    print(f"[OK] {destino.name} ({destino.stat().st_size:,} bytes)")
    return _ler_csv_bis(texto)


def preparar_policy(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["mes"] = out["TIME_PERIOD"].map(parse_ano_mes)
    out["taxa_aa"] = pd.to_numeric(out["OBS_VALUE"], errors="coerce") / 100.0
    out = out.dropna(subset=["mes", "taxa_aa", "REF_AREA"])
    return (
        out[["REF_AREA", "mes", "taxa_aa"]]
        .drop_duplicates(["REF_AREA", "mes"], keep="last")
        .sort_values(["REF_AREA", "mes"])
        .reset_index(drop=True)
    )


def completar_policy_euro(policy: pd.DataFrame) -> pd.DataFrame:
    """Depois do fim da série nacional, usa a taxa do BCE (XM)."""
    xm = policy.loc[policy["REF_AREA"] == "XM", ["mes", "taxa_aa"]].set_index("mes")[
        "taxa_aa"
    ]
    if xm.empty:
        return policy
    extra: list[pd.DataFrame] = []
    for codigo in sorted(EURO_MEMBROS):
        nativo = policy.loc[policy["REF_AREA"] == codigo, "mes"]
        if nativo.empty:
            meses = xm.index
        else:
            meses = xm.index[xm.index > nativo.max()]
        if len(meses) == 0:
            continue
        extra.append(
            pd.DataFrame(
                {
                    "REF_AREA": codigo,
                    "mes": list(meses),
                    "taxa_aa": xm.loc[meses].to_numpy(),
                }
            )
        )
    if not extra:
        return policy
    out = pd.concat([policy, *extra], ignore_index=True)
    return (
        out.drop_duplicates(["REF_AREA", "mes"], keep="first")
        .sort_values(["REF_AREA", "mes"])
        .reset_index(drop=True)
    )


def preparar_cpi(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()
    out["mes"] = out["TIME_PERIOD"].map(parse_ano_mes)
    out["valor"] = pd.to_numeric(out["OBS_VALUE"], errors="coerce")
    out = out.dropna(subset=["mes", "valor", "REF_AREA", "UNIT_MEASURE"])
    indice = out.loc[out["UNIT_MEASURE"] == CPI_UNIDADE_INDICE, ["REF_AREA", "mes", "valor"]]
    yoy = out.loc[out["UNIT_MEASURE"] == CPI_UNIDADE_YOY, ["REF_AREA", "mes", "valor"]]
    indice = (
        indice.rename(columns={"valor": "ipc"})
        .drop_duplicates(["REF_AREA", "mes"], keep="last")
        .sort_values(["REF_AREA", "mes"])
        .reset_index(drop=True)
    )
    yoy = (
        yoy.rename(columns={"valor": "inflacao_12m"})
        .assign(inflacao_12m=lambda d: d["inflacao_12m"] / 100.0)
        .drop_duplicates(["REF_AREA", "mes"], keep="last")
        .sort_values(["REF_AREA", "mes"])
        .reset_index(drop=True)
    )
    return indice, yoy


@dataclass
class ResultadoPais:
    codigo: str
    pais: str
    ranking: int
    taxa_real_acumulada: float
    cagr_real: float
    taxa_nominal_acumulada: float
    cagr_nominal: float
    inflacao_acumulada: float
    cagr_inflacao: float
    taxa_real_media_fisher: float
    taxa_nominal_media: float
    inflacao_12m_media: float
    dias_periodo: int
    dias_com_dados: int
    cobertura: float
    primeiro_mes: date | None
    ultimo_mes: date | None
    no_ranking: bool
    observacao: str


def _series_pais(
    policy: pd.DataFrame,
    ipc: pd.DataFrame,
    yoy: pd.DataFrame,
    codigo: str,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    pol = policy.loc[policy["REF_AREA"] == codigo].set_index("mes")["taxa_aa"]
    idx = ipc.loc[ipc["REF_AREA"] == codigo].set_index("mes")["ipc"]
    inf = yoy.loc[yoy["REF_AREA"] == codigo].set_index("mes")["inflacao_12m"]
    return pol, idx, inf


def fator_inflacao_ipc(
    ipc: pd.Series,
    inicio: date,
    fim: date,
) -> tuple[float, int, date | None, date | None]:
    """Fator de preços no intervalo, com pró-rata nos meses parciais."""
    fator = 1.0
    dias_usados = 0
    primeiro: date | None = None
    ultimo: date | None = None
    for mes in meses_sobrepostos(inicio, fim):
        trecho = intersecao_mes(mes, inicio, fim)
        if trecho is None:
            continue
        a, b = trecho
        prev = mes_anterior(mes)
        if mes not in ipc.index or prev not in ipc.index:
            continue
        c0 = float(ipc.loc[prev])
        c1 = float(ipc.loc[mes])
        if c0 <= 0 or c1 <= 0:
            continue
        mom = c1 / c0
        dias_mes = dias_inclusive(mes_ref(mes), ultimo_dia_mes(mes.year, mes.month))
        dias = dias_inclusive(a, b)
        fator *= mom ** (dias / dias_mes)
        dias_usados += dias
        if primeiro is None:
            primeiro = mes
        ultimo = mes
    return fator, dias_usados, primeiro, ultimo


def fator_nominal_policy(
    policy: pd.Series,
    inicio: date,
    fim: date,
) -> tuple[float, int, date | None, date | None]:
    """Capitaliza a taxa básica anual pelos dias de cada mês no recorte.

    Converte % a.a. em fator mensal ``(1+i)^{1/12}`` e aplica pró-rata
    ``dias / dias_do_mês``, a mesma convenção do IPC. Assim, 12 meses
    cheios a uma taxa constante reproduzem exatamente ``(1+i)``.
    """
    fator = 1.0
    dias_usados = 0
    primeiro: date | None = None
    ultimo: date | None = None
    for mes in meses_sobrepostos(inicio, fim):
        trecho = intersecao_mes(mes, inicio, fim)
        if trecho is None:
            continue
        if mes not in policy.index:
            continue
        taxa = float(policy.loc[mes])
        if pd.isna(taxa):
            continue
        a, b = trecho
        dias = dias_inclusive(a, b)
        dias_mes = dias_inclusive(mes_ref(mes), ultimo_dia_mes(mes.year, mes.month))
        fator *= (1.0 + taxa) ** (dias / dias_mes / 12.0)
        dias_usados += dias
        if primeiro is None:
            primeiro = mes
        ultimo = mes
    return fator, dias_usados, primeiro, ultimo


def media_ponderada_dias(serie: pd.Series, inicio: date, fim: date) -> tuple[float, int]:
    num = 0.0
    den = 0
    for mes in meses_sobrepostos(inicio, fim):
        trecho = intersecao_mes(mes, inicio, fim)
        if trecho is None or mes not in serie.index:
            continue
        valor = float(serie.loc[mes])
        if pd.isna(valor):
            continue
        dias = dias_inclusive(*trecho)
        num += valor * dias
        den += dias
    if den == 0:
        return float("nan"), 0
    return num / den, den


def calcular_pais(
    codigo: str,
    policy: pd.Series,
    ipc: pd.Series,
    yoy: pd.Series,
    periodo: Periodo,
    *,
    cobertura_minima: float = COBERTURA_MINIMA,
) -> ResultadoPais:
    dias_periodo = dias_inclusive(periodo.inicio, periodo.fim)
    fat_nom, dias_nom, p0, p1 = fator_nominal_policy(policy, periodo.inicio, periodo.fim)
    fat_ipc, dias_ipc, i0, i1 = fator_inflacao_ipc(ipc, periodo.inicio, periodo.fim)
    dias_com = min(dias_nom, dias_ipc)
    cobertura = dias_com / dias_periodo if dias_periodo else 0.0

    if dias_nom == 0 or dias_ipc == 0 or fat_nom <= 0 or fat_ipc <= 0:
        return ResultadoPais(
            codigo=codigo,
            pais=nome_pais(codigo),
            ranking=0,
            taxa_real_acumulada=float("nan"),
            cagr_real=float("nan"),
            taxa_nominal_acumulada=float("nan"),
            cagr_nominal=float("nan"),
            inflacao_acumulada=float("nan"),
            cagr_inflacao=float("nan"),
            taxa_real_media_fisher=float("nan"),
            taxa_nominal_media=float("nan"),
            inflacao_12m_media=float("nan"),
            dias_periodo=dias_periodo,
            dias_com_dados=0,
            cobertura=0.0,
            primeiro_mes=None,
            ultimo_mes=None,
            no_ranking=False,
            observacao="sem série sobreposta de taxa básica e IPC",
        )

    # Alinha o recorte efetivo ao menor intervalo com as duas séries.
    inicio_efetivo = periodo.inicio
    fim_efetivo = periodo.fim
    if p0 and i0:
        inicio_efetivo = max(periodo.inicio, max(p0, i0))
    if p1 and i1:
        fim_mes = ultimo_dia_mes(min(p1, i1).year, min(p1, i1).month)
        fim_efetivo = min(periodo.fim, fim_mes)
    if inicio_efetivo > periodo.inicio or fim_efetivo < periodo.fim:
        fat_nom, dias_nom, p0, p1 = fator_nominal_policy(
            policy, inicio_efetivo, fim_efetivo
        )
        fat_ipc, dias_ipc, i0, i1 = fator_inflacao_ipc(ipc, inicio_efetivo, fim_efetivo)
        dias_com = min(dias_nom, dias_ipc)

    real_acc = fat_nom / fat_ipc - 1.0
    anos = anos_fracao_meses(inicio_efetivo, fim_efetivo)
    nom_media, _ = media_ponderada_dias(policy, inicio_efetivo, fim_efetivo)
    yoy_media, _ = media_ponderada_dias(yoy, inicio_efetivo, fim_efetivo)
    fisher_media = (
        taxa_real_fisher(nom_media, yoy_media)
        if pd.notna(nom_media) and pd.notna(yoy_media)
        else float("nan")
    )

    obs: list[str] = []
    if inicio_efetivo != periodo.inicio or fim_efetivo != periodo.fim:
        obs.append(
            f"recorte efetivo {inicio_efetivo.isoformat()} a {fim_efetivo.isoformat()}"
        )
    if cobertura < cobertura_minima:
        obs.append(f"cobertura {cobertura:.1%} abaixo de {cobertura_minima:.0%}")

    primeiro = max(x for x in (p0, i0) if x is not None) if p0 and i0 else p0 or i0
    ultimo = min(x for x in (p1, i1) if x is not None) if p1 and i1 else p1 or i1

    return ResultadoPais(
        codigo=codigo,
        pais=nome_pais(codigo),
        ranking=0,
        taxa_real_acumulada=real_acc,
        cagr_real=cagr_meses(1.0 + real_acc, anos),
        taxa_nominal_acumulada=fat_nom - 1.0,
        cagr_nominal=cagr_meses(fat_nom, anos),
        inflacao_acumulada=fat_ipc - 1.0,
        cagr_inflacao=cagr_meses(fat_ipc, anos),
        taxa_real_media_fisher=fisher_media,
        taxa_nominal_media=nom_media,
        inflacao_12m_media=yoy_media,
        dias_periodo=dias_periodo,
        dias_com_dados=dias_com,
        cobertura=cobertura,
        primeiro_mes=primeiro,
        ultimo_mes=ultimo,
        no_ranking=cobertura >= cobertura_minima and pd.notna(real_acc),
        observacao="; ".join(obs),
    )


def ranquear(resultados: list[ResultadoPais]) -> list[ResultadoPais]:
    validos = [r for r in resultados if r.no_ranking]
    invalidos = [r for r in resultados if not r.no_ranking]
    validos.sort(key=lambda r: r.taxa_real_acumulada, reverse=True)
    invalidos.sort(key=lambda r: (r.cobertura, r.pais), reverse=True)
    out: list[ResultadoPais] = []
    for i, r in enumerate(validos, start=1):
        out.append(
            ResultadoPais(
                **{**r.__dict__, "ranking": i},
            )
        )
    for r in invalidos:
        out.append(ResultadoPais(**{**r.__dict__, "ranking": 0}))
    return out


def resultados_para_df(resultados: list[ResultadoPais]) -> pd.DataFrame:
    linhas = []
    for r in resultados:
        linhas.append(
            {
                "Ranking": r.ranking if r.ranking else "—",
                "País": r.pais,
                "Código": r.codigo,
                "Taxa básica real acumulada": r.taxa_real_acumulada,
                "CAGR real": r.cagr_real,
                "Taxa básica nominal acumulada": r.taxa_nominal_acumulada,
                "CAGR nominal": r.cagr_nominal,
                "Inflação acumulada (IPC)": r.inflacao_acumulada,
                "CAGR inflação": r.cagr_inflacao,
                "Taxa real média Fisher (i e π 12m)": r.taxa_real_media_fisher,
                "Taxa básica nominal média": r.taxa_nominal_media,
                "Inflação 12 meses média": r.inflacao_12m_media,
                "Dias do período": r.dias_periodo,
                "Dias com dados": r.dias_com_dados,
                "Cobertura": r.cobertura,
                "Primeiro mês": r.primeiro_mes.isoformat() if r.primeiro_mes else "",
                "Último mês": r.ultimo_mes.isoformat() if r.ultimo_mes else "",
                "No ranking": "sim" if r.no_ranking else "não",
                "Observação": r.observacao,
            }
        )
    return pd.DataFrame(linhas)


def calcular_periodos(
    policy: pd.DataFrame,
    ipc: pd.DataFrame,
    yoy: pd.DataFrame,
    periodos: Iterable[Periodo] = PERIODOS,
    *,
    cobertura_minima: float = COBERTURA_MINIMA,
) -> dict[int, list[ResultadoPais]]:
    codigos = sorted(set(policy["REF_AREA"]).intersection(ipc["REF_AREA"]))
    saida: dict[int, list[ResultadoPais]] = {}
    for periodo in periodos:
        print(f"[CALC] Período {periodo.id}: {periodo.titulo} ({len(codigos)} países)")
        brutos: list[ResultadoPais] = []
        for codigo in codigos:
            pol, idx, inf = _series_pais(policy, ipc, yoy, codigo)
            brutos.append(
                calcular_pais(
                    codigo,
                    pol,
                    idx,
                    inf,
                    periodo,
                    cobertura_minima=cobertura_minima,
                )
            )
        saida[periodo.id] = ranquear(brutos)
    return saida


# --- Excel -----------------------------------------------------------------

AZUL = "1F4E79"
AZUL_CLARO = "D6E3F0"
DOURADO = "FFF2CC"
VERDE_BR = "C6EFCE"
CINZA = "F2F2F2"
VERMELHO_SUAVE = "FCE4D6"
BRANCO = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

COLS_PCT = {
    "Taxa básica real acumulada",
    "CAGR real",
    "Taxa básica nominal acumulada",
    "CAGR nominal",
    "Inflação acumulada (IPC)",
    "CAGR inflação",
    "Taxa real média Fisher (i e π 12m)",
    "Taxa básica nominal média",
    "Inflação 12 meses média",
    "Cobertura",
}


def _preencher_cabecalho(ws: Worksheet, headers: list[str], linha: int) -> None:
    fill = PatternFill("solid", fgColor=AZUL)
    font = Font(color=BRANCO, bold=True, name="Calibri", size=10)
    for col, nome in enumerate(headers, start=1):
        cell = ws.cell(linha, col, nome)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _escrever_df_ranking(ws: Worksheet, df: pd.DataFrame, linha0: int) -> None:
    headers = list(df.columns)
    _preencher_cabecalho(ws, headers, linha0)
    for i, row in enumerate(df.itertuples(index=False), start=1):
        excel_row = linha0 + i
        is_br = str(row[1]) == "Brasil"  # coluna País
        is_top = row[0] == 1
        for col, valor in enumerate(row, start=1):
            cell = ws.cell(excel_row, col, None if (isinstance(valor, float) and pd.isna(valor)) else valor)
            nome = headers[col - 1]
            cell.border = THIN
            cell.font = Font(name="Calibri", size=10, bold=is_br)
            cell.alignment = Alignment(vertical="center")
            if nome in COLS_PCT and isinstance(valor, float) and pd.notna(valor):
                cell.number_format = "0.00%"
            elif nome in {"Dias do período", "Dias com dados"} and isinstance(valor, (int, float)):
                cell.number_format = "#,##0"
            if is_br:
                cell.fill = PatternFill("solid", fgColor=VERDE_BR)
            elif is_top:
                cell.fill = PatternFill("solid", fgColor=DOURADO)
            elif excel_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=CINZA)
            if nome == "No ranking" and valor == "não":
                cell.fill = PatternFill("solid", fgColor=VERMELHO_SUAVE)
    ultima = linha0 + len(df)
    col_real = headers.index("Taxa básica real acumulada") + 1
    col_cagr = headers.index("CAGR real") + 1
    regra = ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    )
    if ultima >= linha0 + 1:
        ws.conditional_formatting.add(
            f"{get_column_letter(col_real)}{linha0+1}:{get_column_letter(col_real)}{ultima}",
            regra,
        )
        ws.conditional_formatting.add(
            f"{get_column_letter(col_cagr)}{linha0+1}:{get_column_letter(col_cagr)}{ultima}",
            regra,
        )


def _larguras(ws: Worksheet, headers: list[str]) -> None:
    larguras = {
        "Ranking": 10,
        "País": 22,
        "Código": 10,
        "Observação": 48,
        "Primeiro mês": 14,
        "Último mês": 14,
        "No ranking": 12,
    }
    for i, nome in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(nome, 18)


def _aba_metodologia(wb: Workbook, gerado_em: datetime, n_paises: int) -> None:
    ws = wb.active
    ws.title = "Metodologia"
    ws["A1"] = "Discriminativo — taxas básicas de juros reais acumuladas e CAGR"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=AZUL)
    ws.merge_cells("A1:B1")

    blocos = [
        ("Gerado em", gerado_em.strftime("%d/%m/%Y %H:%M")),
        ("Marker", MARKER),
        ("Países com taxa básica + IPC", str(n_paises)),
        (
            "Fonte taxa básica",
            "BIS WS_CBPOL — Central bank policy rates, mensal, fim do período (% a.a.)",
        ),
        (
            "Fonte inflação",
            "BIS WS_LONG_CPI — índice de preços ao consumidor (unidade 628) e variação 12 meses (unidade 771)",
        ),
        (
            "Taxa real mensal",
            "A taxa básica anual do mês é capitalizada pelos dias corridos do recorte. "
            "A inflação do mês é a variação do índice IPC frente ao mês anterior, "
            "pró-rata pelos dias do mês que caem no período. "
            "Fator real = fator nominal / fator de preços.",
        ),
        (
            "Taxa básica real acumulada",
            "Π (1+i_t)^{d_t/(12·D_t)}  ÷  Π (IPC_t/IPC_{t-1})^{d_t/D_t}  −  1, "
            "com i_t = taxa básica anual do mês, d_t = dias do mês t no recorte "
            "e D_t = dias do mês t.",
        ),
        (
            "CAGR real",
            "(1 + taxa real acumulada)^{1 / anos} − 1, "
            "em que anos = Σ d_t/(12·D_t) (mês cheio = 1/12).",
        ),
        (
            "Taxa real média Fisher",
            "Média ponderada por dias de i e da inflação em 12 meses, depois (1+i)/(1+π)−1. "
            "É o nível médio da taxa real ex-post usualmente citada na imprensa; "
            "não é usada no ranking.",
        ),
        (
            "Ranking",
            "Ordem decrescente da taxa básica real acumulada. "
            f"Entram países com cobertura ≥ {COBERTURA_MINIMA:.0%} dos dias do período. "
            "Brasil destacado em verde; 1º lugar em dourado.",
        ),
        (
            "Períodos",
            " ; ".join(f"{p.id}) {p.titulo}" for p in PERIODOS),
        ),
        (
            "Período 5 / defasagem",
            "O pedido vai até 28/08/2026. As séries mensais do BIS em geral fecham "
            "no último mês publicado (tipicamente o mês anterior). "
            "A coluna Observação registra o recorte efetivo quando a série termina antes.",
        ),
        (
            "Zona do Euro",
            "A série XM é a taxa do BCE. Depois que o BIS encerra a série nacional "
            "(adoção do euro), AT/BE/DE/ES/FR/GR/HR/IT/NL/PT passam a usar a taxa "
            "do BCE com o IPC nacional, para permanecerem no ranking.",
        ),
        (
            "O que este número mede",
            "Retorno real (poder de compra) de rolar a taxa básica do banco central "
            "ao longo do período, descontado o IPC do próprio país. "
            "Não é taxa de juros de mercado nem taxa de empréstimo bancário.",
        ),
    ]
    ws["A3"] = "Campo"
    ws["B3"] = "Descrição"
    _preencher_cabecalho(ws, ["Campo", "Descrição"], 3)
    for i, (campo, desc) in enumerate(blocos, start=4):
        ws.cell(i, 1, campo).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(i, 2, desc).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 48 if len(desc) > 80 else 20
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 24


def _aba_periodo(wb: Workbook, periodo: Periodo, resultados: list[ResultadoPais]) -> None:
    ws = wb.create_sheet(periodo.aba)
    ws["A1"] = f"Período {periodo.id} — {periodo.titulo}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    ws.merge_cells("A1:S1")
    n_rank = sum(1 for r in resultados if r.no_ranking)
    br = next((r for r in resultados if r.codigo == "BR"), None)
    resumo = (
        f"Ranking em ordem decrescente da taxa básica real acumulada. "
        f"{n_rank} países no ranking (cobertura ≥ {COBERTURA_MINIMA:.0%})."
    )
    if br and br.no_ranking:
        resumo += (
            f" Brasil: {br.ranking}º lugar | "
            f"acumulada {br.taxa_real_acumulada:.2%} | "
            f"CAGR {br.cagr_real:.2%} a.a."
        )
    ws["A2"] = resumo
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells("A2:S2")
    df = resultados_para_df(resultados)
    _escrever_df_ranking(ws, df, 4)
    _larguras(ws, list(df.columns))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(df.shape[1])}{4 + len(df)}"
    ws.row_dimensions[4].height = 32
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.left.text = f"Período {periodo.id} — {periodo.titulo}"
    ws.print_title_rows = "4:4"


def _aba_comparativo(
    wb: Workbook, por_periodo: dict[int, list[ResultadoPais]]
) -> None:
    ws = wb.create_sheet("Comparativo")
    ws["A1"] = "Comparativo entre períodos — ranking da taxa básica real acumulada"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    ws.merge_cells("A1:Q1")

    paises = sorted(
        {r.codigo for lista in por_periodo.values() for r in lista},
        key=lambda c: nome_pais(c),
    )
    headers = ["País", "Código"]
    for p in PERIODOS:
        headers.extend(
            [
                f"P{p.id} ranking",
                f"P{p.id} real acum.",
                f"P{p.id} CAGR real",
            ]
        )
    _preencher_cabecalho(ws, headers, 3)

    lookup = {
        pid: {r.codigo: r for r in lista} for pid, lista in por_periodo.items()
    }
    linha = 4
    for codigo in paises:
        valores: list[object] = [nome_pais(codigo), codigo]
        for p in PERIODOS:
            r = lookup.get(p.id, {}).get(codigo)
            if r is None:
                valores.extend(["—", None, None])
            else:
                valores.extend(
                    [
                        r.ranking if r.ranking else "—",
                        r.taxa_real_acumulada,
                        r.cagr_real,
                    ]
                )
        is_br = codigo == "BR"
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(linha, col, None if (isinstance(valor, float) and pd.isna(valor)) else valor)
            cell.border = THIN
            cell.font = Font(name="Calibri", size=10, bold=is_br)
            nome = headers[col - 1]
            if "real" in nome.lower() or "CAGR" in nome:
                if isinstance(valor, float) and pd.notna(valor):
                    cell.number_format = "0.00%"
            if is_br:
                cell.fill = PatternFill("solid", fgColor=VERDE_BR)
            elif linha % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=CINZA)
        linha += 1

    # bloco Brasil
    bloco = linha + 2
    ws.cell(bloco, 1, "Destaque — Brasil").font = Font(
        name="Calibri", size=12, bold=True, color=AZUL
    )
    br_headers = [
        "Período",
        "Datas",
        "Ranking",
        "N no ranking",
        "Taxa real acumulada",
        "CAGR real",
        "CAGR nominal",
        "CAGR inflação",
    ]
    _preencher_cabecalho(ws, br_headers, bloco + 1)
    for i, p in enumerate(PERIODOS):
        r = lookup.get(p.id, {}).get("BR")
        n_rank = sum(1 for x in por_periodo.get(p.id, []) if x.no_ranking)
        row = bloco + 2 + i
        dados = [
            p.id,
            p.titulo,
            r.ranking if r and r.ranking else "—",
            n_rank,
            r.taxa_real_acumulada if r else None,
            r.cagr_real if r else None,
            r.cagr_nominal if r else None,
            r.cagr_inflacao if r else None,
        ]
        for col, valor in enumerate(dados, start=1):
            cell = ws.cell(row, col, valor)
            cell.border = THIN
            cell.fill = PatternFill("solid", fgColor=VERDE_BR)
            cell.font = Font(name="Calibri", size=10, bold=True)
            if col >= 5 and isinstance(valor, float):
                cell.number_format = "0.00%"

    _larguras(ws, headers)
    for i, nome in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 16 if i > 2 else 22
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{linha - 1}"


def escrever_planilha(
    por_periodo: dict[int, list[ResultadoPais]],
    saida: Path,
    *,
    n_paises: int,
) -> Path:
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _aba_metodologia(wb, datetime.now(), n_paises)
    por_id = {p.id: p for p in PERIODOS}
    for pid in sorted(por_periodo):
        _aba_periodo(wb, por_id[pid], por_periodo[pid])
    _aba_comparativo(wb, por_periodo)
    wb.save(saida)
    print(f"[OK] Planilha: {saida} ({saida.stat().st_size / 1024:.1f} KB)")
    return saida


def carregar_fontes(
    pasta_cache: Path,
    *,
    usar_cache: bool = True,
    policy_csv: Path | None = None,
    cpi_csv: Path | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if policy_csv is not None:
        policy_bruto = pd.read_csv(policy_csv)
    else:
        policy_bruto = baixar_ou_cache(
            BIS_CBPOL, pasta_cache / "bis_cbpol_m.csv", usar_cache=usar_cache
        )
    if cpi_csv is not None:
        cpi_bruto = pd.read_csv(cpi_csv)
    else:
        cpi_bruto = baixar_ou_cache(
            BIS_CPI, pasta_cache / "bis_long_cpi_m.csv", usar_cache=usar_cache
        )
    policy = completar_policy_euro(preparar_policy(policy_bruto))
    ipc, yoy = preparar_cpi(cpi_bruto)
    print(
        f"[INFO] policy={policy['REF_AREA'].nunique()} países, "
        f"{policy['mes'].min()} → {policy['mes'].max()}"
    )
    print(
        f"[INFO] ipc={ipc['REF_AREA'].nunique()} países, "
        f"{ipc['mes'].min()} → {ipc['mes'].max()}"
    )
    return policy, ipc, yoy


def processar(
    *,
    pasta_cache: Path,
    saida: Path,
    usar_cache: bool = True,
    policy_csv: Path | None = None,
    cpi_csv: Path | None = None,
    cobertura_minima: float = COBERTURA_MINIMA,
) -> Path:
    print(f"[{MARKER}]")
    policy, ipc, yoy = carregar_fontes(
        pasta_cache,
        usar_cache=usar_cache,
        policy_csv=policy_csv,
        cpi_csv=cpi_csv,
    )
    n_paises = len(set(policy["REF_AREA"]).intersection(ipc["REF_AREA"]))
    por_periodo = calcular_periodos(
        policy, ipc, yoy, cobertura_minima=cobertura_minima
    )
    path = escrever_planilha(por_periodo, saida, n_paises=n_paises)
    for p in PERIODOS:
        lista = por_periodo[p.id]
        print(f"\n=== Período {p.id}: {p.titulo} ===")
        for r in lista[:8]:
            if not r.no_ranking:
                continue
            print(
                f"  {r.ranking:2d}. {r.pais:<22} "
                f"acum={r.taxa_real_acumulada:8.2%}  "
                f"CAGR={r.cagr_real:7.2%}  "
                f"cob={r.cobertura:5.1%}"
            )
        br = next((x for x in lista if x.codigo == "BR"), None)
        if br:
            print(
                f"  → Brasil: {br.ranking}º | "
                f"acum={br.taxa_real_acumulada:.2%} | "
                f"CAGR={br.cagr_real:.2%}"
            )
    return path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--pasta-cache",
        type=Path,
        default=ROOT / "data" / "bis",
        help="Cache CSV do BIS (default: data/bis)",
    )
    p.add_argument(
        "--saida",
        type=Path,
        default=ROOT / "output" / "discriminativo_juros_reais_paises.xlsx",
    )
    p.add_argument("--policy-csv", type=Path, default=None, help="CSV BIS CBPOL local")
    p.add_argument("--cpi-csv", type=Path, default=None, help="CSV BIS LONG_CPI local")
    p.add_argument("--sem-cache", action="store_true", help="Força download do BIS")
    p.add_argument(
        "--cobertura-minima",
        type=float,
        default=COBERTURA_MINIMA,
        help="Fração mínima de dias com dados para entrar no ranking (default 0,80)",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        processar(
            pasta_cache=args.pasta_cache,
            saida=args.saida,
            usar_cache=not args.sem_cache,
            policy_csv=args.policy_csv,
            cpi_csv=args.cpi_csv,
            cobertura_minima=args.cobertura_minima,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
