#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta discriminativos Excel para PDF (marcadores) e HTML (abas).

O PDF **não** tem abas nativas como o Excel. O que dá para preservar:

  - uma seção por aba, com o mesmo quadro
  - sumário/marcadores (painel esquerdo do leitor) para saltar entre abas
  - página Índice com links internos
  - rodapé «Índice | aba anterior | aba seguinte»

O HTML replica a faixa de abas: clicar troca o conteúdo, como no Excel.

Uso::

  python scripts/discriminativo_para_pdf.py \\
      --entrada output/discriminativo_ranking_juros_reais.xlsx
  python scripts/discriminativo_para_pdf.py \\
      --entrada output/discriminativo_juros_reais_paises.xlsx
"""

from __future__ import annotations

import argparse
import html
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    NextPageTemplate,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.flowables import Flowable

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

MARKER = "discriminativo-pdf-20260829"


def _slug(nome: str) -> str:
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", str(nome).strip())
    texto = texto.strip("_") or "aba"
    if texto[0].isdigit():
        texto = "aba_" + texto
    return texto


def _hex_fill(cell) -> str | None:
    fill = getattr(cell, "fill", None)
    if fill is None or fill.fill_type not in ("solid", "pattern"):
        if fill is None or not fill.fgColor:
            return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    rgb = getattr(fg, "rgb", None)
    if not rgb or rgb in ("00000000", "0"):
        theme = getattr(fg, "theme", None)
        if theme is None:
            return None
        return None
    rgb = str(rgb)
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6 or rgb.upper() == "000000":
        return None
    return f"#{rgb}"


def _hex_font(cell) -> str | None:
    font = getattr(cell, "font", None)
    if font is None or font.color is None:
        return None
    rgb = getattr(font.color, "rgb", None)
    if not rgb:
        return None
    rgb = str(rgb)
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    return f"#{rgb}"


def formatar_valor(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Sim" if v else "Não"
    fmt = str(cell.number_format or "")
    if isinstance(v, (int, float)) and "%" in fmt:
        casas = 4 if "0.0000" in fmt else 2
        return f"{v * 100:,.{casas}f}%".replace(",", "X").replace(".", ",").replace("X", ".")
    if isinstance(v, float) and any(tok in fmt for tok in ("0.00", "#,##0")):
        return f"{v:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def _alinhamento(cell) -> str:
    h = getattr(getattr(cell, "alignment", None), "horizontal", None)
    if h == "center":
        return "CENTER"
    if h == "right":
        return "RIGHT"
    if h == "left":
        return "LEFT"
    if isinstance(cell.value, (int, float)):
        return "RIGHT"
    return "LEFT"


def ler_aba(ws) -> dict:
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    # recorta colunas/linhas vazias no fim
    ultima_r, ultima_c = 0, 0
    grid = []
    for r in range(1, max_r + 1):
        linha = []
        vazia = True
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                linha.append({"texto": "", "fill": None, "font": None, "bold": False, "align": "LEFT"})
                continue
            texto = formatar_valor(cell)
            if texto:
                vazia = False
                ultima_c = max(ultima_c, c)
            linha.append(
                {
                    "texto": texto,
                    "fill": _hex_fill(cell),
                    "font": _hex_font(cell),
                    "bold": bool(getattr(cell.font, "bold", False)),
                    "align": _alinhamento(cell),
                }
            )
        if not vazia:
            ultima_r = r
        grid.append(linha)
    grid = [ln[:ultima_c] for ln in grid[:ultima_r]]
    larguras = []
    for c in range(1, ultima_c + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        larguras.append(max(float(w) if w else 12.0, 10.0))
    # títulos longos na 1ª coluna (célula mesclada) não podem ficar com 10–12
    if grid and grid[0] and len(grid[0][0]["texto"]) > 40:
        larguras[0] = max(larguras[0], 36.0)
    merges = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row > ultima_r or rng.min_col > ultima_c:
            continue
        merges.append(
            (
                rng.min_row,
                rng.min_col,
                min(rng.max_row, ultima_r),
                min(rng.max_col, ultima_c),
            )
        )
    return {"nome": ws.title, "grid": grid, "larguras": larguras, "merges": merges}


def ler_workbook(path: Path, abas: list[str] | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    nomes = list(wb.sheetnames)
    if abas:
        pedidas = set(abas)
        nomes = [n for n in nomes if n in pedidas]
        if not nomes:
            raise ValueError(f"Nenhuma aba pedida encontrada em {path.name}")
    return [ler_aba(wb[n]) for n in nomes]


class Destino(Flowable):
    """Cria destino nomeado + entrada no sumário (marcadores) do PDF."""

    def __init__(self, key: str, titulo: str, nivel: int = 0):
        super().__init__()
        self.key = key
        self.titulo = titulo
        self.nivel = nivel
        self.width = 0
        self.height = 0

    def draw(self):
        self.canv.bookmarkPage(self.key)
        self.canv.addOutlineEntry(self.titulo, self.key, level=self.nivel, closed=0)


def _rgb(hex_cor: str | None, default="#FFFFFF"):
    h = (hex_cor or default).lstrip("#")
    if len(h) != 6:
        h = default.lstrip("#")
    return colors.HexColor("#" + h)


class _CabecalhoRodape:
    def __init__(self, titulo: str, nomes: list[str], slugs: list[str]):
        self.titulo = titulo
        self.nomes = nomes
        self.slugs = slugs
        self.atual = slugs[0] if slugs else "indice"

    def __call__(self, canv, doc):
        canv.saveState()
        w, h = landscape(A4)
        canv.setFillColor(colors.HexColor("#1B4F72"))
        canv.rect(0, h - 16 * mm, w, 16 * mm, fill=1, stroke=0)
        canv.setFillColor(colors.white)
        canv.setFont("Helvetica-Bold", 9)
        canv.drawString(12 * mm, h - 10 * mm, self.titulo[:90])
        canv.setFont("Helvetica", 8)
        canv.drawRightString(w - 12 * mm, h - 10 * mm, f"Aba: {getattr(doc, 'aba_atual', '')}")
        canv.setFillColor(colors.HexColor("#34495E"))
        canv.setFont("Helvetica", 8)
        y = 8 * mm
        canv.drawString(12 * mm, y, "PDF com marcadores: abra o painel de sumário do leitor para trocar de aba.")
        canv.drawRightString(w - 12 * mm, y, f"{doc.page}")
        # links Índice / ant / prox
        slugs = self.slugs
        atual = getattr(doc, "aba_slug", slugs[0] if slugs else "indice")
        try:
            i = slugs.index(atual)
        except ValueError:
            i = 0
        canv.setFillColor(colors.HexColor("#1B4F72"))
        canv.setFont("Helvetica-Bold", 8)
        canv.drawString(12 * mm, 13 * mm, "Índice")
        canv.linkAbsolute("Índice", "indice", (12 * mm, 11 * mm, 28 * mm, 16 * mm))
        if i > 0:
            canv.drawString(36 * mm, 13 * mm, "< anterior")
            canv.linkAbsolute("ant", slugs[i - 1], (36 * mm, 11 * mm, 62 * mm, 16 * mm))
        if i < len(slugs) - 1:
            canv.drawString(66 * mm, 13 * mm, "proxima >")
            canv.linkAbsolute("prox", slugs[i + 1], (66 * mm, 11 * mm, 92 * mm, 16 * mm))
        canv.restoreState()


def _estilos():
    base = getSampleStyleSheet()
    return {
        "titulo": ParagraphStyle(
            "tit",
            parent=base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=14,
            textColor=colors.HexColor("#1B4F72"),
            spaceAfter=6,
        ),
        "corpo": ParagraphStyle(
            "cor",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
        ),
        "link": ParagraphStyle(
            "lnk",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.HexColor("#1B4F72"),
            leading=12,
        ),
        "cel": ParagraphStyle(
            "cel",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=8,
        ),
        "celb": ParagraphStyle(
            "celb",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=6.5,
            leading=8,
        ),
    }


def _par_cel(cel: dict, sty: dict) -> Paragraph:
    estilo = sty["celb"] if cel["bold"] else sty["cel"]
    cor = cel["font"] or ("#FFFFFF" if (cel["fill"] or "").upper() == "#1B4F72" else "#1B2A3A")
    align = {"LEFT": TA_LEFT, "RIGHT": TA_RIGHT, "CENTER": TA_CENTER}[cel["align"]]
    local = ParagraphStyle(
        f"c{id(cel)}",
        parent=estilo,
        textColor=colors.HexColor(cor),
        alignment=align,
    )
    texto = html.escape(cel["texto"]).replace("\n", "<br/>") or "&nbsp;"
    return Paragraph(texto, local)


def _tabela_aba(aba: dict, sty: dict, largura_util: float) -> Table:
    grid = aba["grid"]
    if not grid:
        return Table([[Paragraph("(aba vazia)", sty["corpo"])]])
    n_col = max(len(r) for r in grid)
    pesos = list(aba["larguras"]) + [14.0] * (n_col - len(aba["larguras"]))
    pesos = pesos[:n_col]
    # cabeçalho longo: garante largura mínima pela maior linha de texto da coluna
    for j in range(n_col):
        maior = 0
        for linha in grid[:8]:
            if j < len(linha):
                maior = max(maior, max((len(p) for p in linha[j]["texto"].split("\n")), default=0))
        pesos[j] = max(pesos[j], min(maior * 0.85, 36.0))
    s = sum(pesos) or 1.0
    col_w = [largura_util * p / s for p in pesos]
    data = []
    for linha in grid:
        linha = linha + [{"texto": "", "fill": None, "font": None, "bold": False, "align": "LEFT"}] * (
            n_col - len(linha)
        )
        data.append([_par_cel(c, sty) for c in linha[:n_col]])
    tab = Table(data, colWidths=col_w, repeatRows=0)
    cmds = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#B0BEC5")),
    ]
    for r1, c1, r2, c2 in aba.get("merges") or []:
        cmds.append(("SPAN", (c1 - 1, r1 - 1), (c2 - 1, r2 - 1)))
    for i, linha in enumerate(grid):
        for j, cel in enumerate(linha[:n_col]):
            fill = cel["fill"]
            if fill:
                cmds.append(("BACKGROUND", (j, i), (j, i), _rgb(fill)))
    tab.setStyle(TableStyle(cmds))
    return tab


def exportar_pdf(abas: list[dict], saida: Path, titulo: str) -> Path:
    saida.parent.mkdir(parents=True, exist_ok=True)
    slugs = [_slug(a["nome"]) for a in abas]
    # desambiguar
    vistos: dict[str, int] = {}
    for i, s in enumerate(slugs):
        n = vistos.get(s, 0) + 1
        vistos[s] = n
        if n > 1:
            slugs[i] = f"{s}_{n}"
    sty = _estilos()
    page = landscape(A4)
    margem = 12 * mm
    frame = Frame(
        margem,
        18 * mm,
        page[0] - 2 * margem,
        page[1] - 36 * mm,
        id="corpo",
    )
    cab = _CabecalhoRodape(titulo, [a["nome"] for a in abas], slugs)
    doc = BaseDocTemplate(
        str(saida),
        pagesize=page,
        title=titulo,
        author="SEC--data-analysys",
        subject="Discriminativo com uma seção por aba e marcadores PDF",
    )
    doc.addPageTemplates([PageTemplate(id="paisagem", frames=[frame], onPage=cab)])
    doc.aba_atual = "Índice"
    doc.aba_slug = "indice"

    story: list = [Destino("indice", "Índice de abas", 0)]
    story.append(Paragraph(html.escape(titulo), sty["titulo"]))
    story.append(
        Paragraph(
            "O PDF não possui abas como o Excel. Use o <b>sumário/marcadores</b> do leitor "
            "(ícone de lista à esquerda) ou os links abaixo — cada item abre a aba correspondente. "
            "No rodapé: Índice, aba anterior e aba seguinte.",
            sty["corpo"],
        )
    )
    story.append(Spacer(1, 6))
    links = []
    for nome, slug in zip((a["nome"] for a in abas), slugs):
        links.append(Paragraph(f'<link href="#{slug}" color="#1B4F72"><u>{html.escape(nome)}</u></link>', sty["link"]))
    n = 4
    while len(links) % n:
        links.append(Paragraph("", sty["link"]))
    grade = [links[i : i + n] for i in range(0, len(links), n)]
    tab_idx = Table(grade, colWidths=[(page[0] - 2 * margem) / n] * n)
    tab_idx.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tab_idx)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<font size='7' color='#5D6D7E'>{MARKER}</font>", sty["corpo"]))

    largura_util = page[0] - 2 * margem
    for aba, slug in zip(abas, slugs):
        story.append(PageBreak())
        bloco = [
            Destino(slug, aba["nome"], 0),
            Paragraph(html.escape(aba["nome"]), sty["titulo"]),
            _tabela_aba(aba, sty, largura_util),
        ]
        story.append(KeepTogether(bloco[:2]))
        # KeepTogether só no título; tabela pode quebrar páginas
        story.append(bloco[2])

    # injeta nome da aba no doc a cada destino via afterFlowable
    def after(flowable):
        if isinstance(flowable, Destino):
            if flowable.key == "indice":
                doc.aba_atual = "Índice"
                doc.aba_slug = "indice"
            else:
                doc.aba_atual = flowable.titulo
                doc.aba_slug = flowable.key

    doc.afterFlowable = after
    doc.build(story)
    print(f"[OK] PDF: {saida} ({saida.stat().st_size / 1e6:.2f} MB)")
    return saida


def _spans_html(aba: dict) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    origem: dict[tuple[int, int], tuple[int, int]] = {}
    pular: set[tuple[int, int]] = set()
    n_r = len(aba["grid"])
    n_c = max((len(ln) for ln in aba["grid"]), default=0)
    for r1, c1, r2, c2 in aba.get("merges") or []:
        origem[(r1 - 1, c1 - 1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1 - 1, min(r2, n_r)):
            for c in range(c1 - 1, min(c2, n_c)):
                if (r, c) != (r1 - 1, c1 - 1):
                    pular.add((r, c))
    return origem, pular


def exportar_html(abas: list[dict], saida: Path, titulo: str) -> Path:
    """HTML com faixa de abas — navegação equivalente à do Excel."""
    saida.parent.mkdir(parents=True, exist_ok=True)
    slugs = []
    vistos: dict[str, int] = {}
    for a in abas:
        s = _slug(a["nome"])
        vistos[s] = vistos.get(s, 0) + 1
        slugs.append(s if vistos[s] == 1 else f"{s}_{vistos[s]}")

    partes = [
        "<!DOCTYPE html><html lang='pt-BR'><head><meta charset='utf-8'>",
        f"<title>{html.escape(titulo)}</title>",
        "<style>",
        "body{margin:0;font-family:Calibri,Arial,sans-serif;background:#f4f6f7;color:#1b2a3a}",
        ".bar{background:#1B4F72;color:#fff;padding:10px 16px;font-weight:bold}",
        ".tabs{display:flex;flex-wrap:wrap;gap:4px;background:#d6dbdf;padding:6px 8px;",
        "border-bottom:1px solid #85929e;position:sticky;top:0;z-index:2}",
        ".tabs label{background:#eaeded;border:1px solid #aab7b8;border-bottom:none;",
        "padding:4px 10px;border-radius:6px 6px 0 0;cursor:pointer;font-size:13px}",
        ".tabs label:hover{background:#fff}",
        "input[type=radio]{display:none}",
        ".sheet{display:none;padding:12px 16px 32px;overflow:auto;background:#fff}",
        "table{border-collapse:collapse;font-size:12px}",
        "td{border:1px solid #b0bec5;padding:6px 8px;white-space:normal;vertical-align:middle}",
        "tr:first-child td, tr:nth-child(4) td{min-width:7em}",
        f"{''.join(f'#{s}:checked~#{s}_box{{display:block}}#{s}:checked~div.tabs label[for={s}]{{background:#fff;font-weight:bold;color:#1B4F72}}' for s in slugs)}",
        "</style></head><body>",
        f"<div class='bar'>{html.escape(titulo)} — clique nas abas (equivalente às folhas do Excel)</div>",
    ]
    for i, s in enumerate(slugs):
        chk = " checked" if i == 0 else ""
        partes.append(f"<input type='radio' name='aba' id='{s}'{chk}>")
    partes.append("<div class='tabs'>")
    for a, s in zip(abas, slugs):
        partes.append(f"<label for='{s}'>{html.escape(a['nome'])}</label>")
    partes.append("</div>")
    for a, s in zip(abas, slugs):
        origem, pular = _spans_html(a)
        partes.append(f"<div class='sheet' id='{s}_box'><table>")
        for i, linha in enumerate(a["grid"]):
            partes.append("<tr>")
            for j, cel in enumerate(linha):
                if (i, j) in pular:
                    continue
                rs, cs = origem.get((i, j), (1, 1))
                span = ""
                if rs > 1:
                    span += f" rowspan='{rs}'"
                if cs > 1:
                    span += f" colspan='{cs}'"
                fill = f"background:{cel['fill']};" if cel["fill"] else ""
                cor = f"color:{cel['font']};" if cel["font"] else ""
                if (cel["fill"] or "").upper() == "#1B4F72" and not cel["font"]:
                    cor = "color:#fff;"
                peso = "font-weight:bold;" if cel["bold"] else ""
                al = {"LEFT": "left", "RIGHT": "right", "CENTER": "center"}[cel["align"]]
                txt = html.escape(cel["texto"]).replace("\n", "<br>")
                partes.append(
                    f"<td{span} style='{fill}{cor}{peso}text-align:{al};white-space:normal;"
                    f"min-width:6em'>{txt}</td>"
                )
            partes.append("</tr>")
        partes.append("</table></div>")
    partes.append(f"<p style='margin:8px 16px;color:#5d6d7e;font-size:12px'>{MARKER}</p>")
    partes.append("</body></html>")
    saida.write_text("".join(partes), encoding="utf-8")
    print(f"[OK] HTML (abas): {saida} ({saida.stat().st_size / 1e6:.2f} MB)")
    return saida


def processar(
    entrada: Path,
    saida_pdf: Path | None = None,
    saida_html: Path | None = None,
    *,
    abas: list[str] | None = None,
    titulo: str | None = None,
) -> tuple[Path, Path]:
    entrada = Path(entrada)
    if not entrada.exists():
        raise FileNotFoundError(entrada)
    if saida_pdf is None:
        saida_pdf = entrada.with_suffix(".pdf")
    if saida_html is None:
        saida_html = entrada.with_suffix(".html")
    print(f"[{MARKER}] {entrada}")
    folhas = ler_workbook(entrada, abas=abas)
    titulo = titulo or entrada.stem.replace("_", " ")
    exportar_pdf(folhas, Path(saida_pdf), titulo)
    exportar_html(folhas, Path(saida_html), titulo)
    return Path(saida_pdf), Path(saida_html)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Excel do discriminativo → PDF (marcadores) + HTML (abas).")
    p.add_argument("--entrada", type=Path, required=True)
    p.add_argument("--pdf", type=Path, default=None)
    p.add_argument("--html", type=Path, default=None)
    p.add_argument("--abas", type=str, default="", help="Nomes de abas separados por vírgula")
    p.add_argument("--titulo", type=str, default=None)
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    abas = [x.strip() for x in args.abas.split(",") if x.strip()] or None
    processar(args.entrada, args.pdf, args.html, abas=abas, titulo=args.titulo)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
