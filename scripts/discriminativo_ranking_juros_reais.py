#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Discriminativo de ranking das taxas básicas de juros reais anuais.

Uma aba por ano (1995–2026), países ordenados da maior para a menor
taxa real acumulada no ano (Fisher composto, BIS CBPOL + CPI).

Anos completos (12 meses com dezembro): ranking oficial.
2026 (e qualquer ano sem dezembro): ranking parcial (acumulado até o
último mês disponível).

Uso::

  python scripts/discriminativo_ranking_juros_reais.py
  python scripts/discriminativo_ranking_juros_reais.py --ano-inicio 1995 --ano-fim 2026
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.discriminativo_juros_reais_paises import (  # noqa: E402
    BIS_BULK,
    ROOT as REPO_ROOT,
    _aplicar_larguras,
    _estilo_base,
    carregar_series_paises,
    estatisticas_anuais,
)

MARKER = "ranking-juros-reais-1995-2026-20260829"
ANO_INICIO_DEFAULT = 1995
ANO_FIM_DEFAULT = 2026

COL_POS = "Posição"
COL_PAIS = "País"
COL_COD = "Código"
COL_REAL = "Taxa básica real acumulada no ano (%)"
COL_INFL = "Inflação acumulada no ano (%)"
COL_NOM = "Taxa básica nominal composta no ano (%)"
COL_NOM_FIM = "Taxa básica nominal no fim do período (% a.a.)"
COL_MESES = "Meses"
COL_COB = "Cobertura"
COL_VAR = "Variação da real vs ano anterior (p.p.)"

COLS_RANK = (
    COL_POS,
    COL_PAIS,
    COL_COD,
    COL_REAL,
    COL_INFL,
    COL_NOM,
    COL_NOM_FIM,
    COL_MESES,
    COL_COB,
    COL_VAR,
)

# Cabeçalhos com quebra de linha para caber sem reticências.
HEADERS_RANK = (
    "Posição",
    "País",
    "Código",
    "Taxa básica real\nacumulada no ano (%)",
    "Inflação acumulada\nno ano (%)",
    "Taxa básica nominal\ncomposta no ano (%)",
    "Taxa básica nominal\nno fim do período (% a.a.)",
    "Meses",
    "Cobertura",
    "Variação da real vs\nano anterior (p.p.)",
)

# Larguras em caracteres Excel — cabem o título completo + valores %.
LARGURAS_ABA_ANO = {
    1: 12,
    2: 28,
    3: 12,
    4: 28,
    5: 24,
    6: 28,
    7: 32,
    8: 12,
    9: 30,
    10: 26,
}


def consolidar_anuais(por_serie: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames = [estatisticas_anuais(s) for s in por_serie.values() if not s.empty]
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def ranking_do_ano(
    anuais: pd.DataFrame,
    ano: int,
    *,
    anuais_ant: pd.DataFrame | None = None,
    exigir_completo: bool | None = None,
) -> pd.DataFrame:
    """Ranking do ano: oficiais (completos) primeiro, incompletos sem posição."""
    bloco = anuais[anuais["ano"] == int(ano)].copy()
    if bloco.empty:
        return pd.DataFrame(columns=list(COLS_RANK) + ["oficial"])
    if exigir_completo is None:
        exigir_completo = bool(bloco["completo"].any())

    if anuais_ant is not None and not anuais_ant.empty:
        prev = anuais_ant.set_index("codigo")["real_aa"]
        bloco["var_pp"] = bloco["codigo"].map(prev)
        bloco["var_pp"] = (bloco["real_aa"] - bloco["var_pp"]) * 100.0
    else:
        bloco["var_pp"] = pd.NA

    oficiais = bloco[bloco["completo"]] if exigir_completo else bloco
    extras = bloco[~bloco["completo"]] if exigir_completo else bloco.iloc[0:0]

    def _montar(parte: pd.DataFrame, ranquear: bool) -> pd.DataFrame:
        if parte.empty:
            return pd.DataFrame(columns=list(COLS_RANK) + ["oficial"])
        parte = parte.sort_values(
            ["real_aa", "pais"], ascending=[False, True], kind="mergesort"
        ).reset_index(drop=True)
        rows = []
        for i, rec in enumerate(parte.itertuples(index=False), start=1):
            cob = "ano completo (12 meses)" if rec.completo else f"parcial ({int(rec.n_meses)} meses)"
            rows.append(
                {
                    COL_POS: i if ranquear else pd.NA,
                    COL_PAIS: rec.pais,
                    COL_COD: rec.codigo,
                    COL_REAL: rec.real_aa,
                    COL_INFL: rec.inflacao_aa,
                    COL_NOM: rec.nominal_composta,
                    COL_NOM_FIM: rec.nominal_fim,
                    COL_MESES: int(rec.n_meses),
                    COL_COB: cob,
                    COL_VAR: rec.var_pp,
                    "oficial": ranquear,
                }
            )
        return pd.DataFrame(rows)

    out = pd.concat(
        [_montar(oficiais, True), _montar(extras, False)],
        ignore_index=True,
    )
    return out


def resumo_rankings(por_ano: dict[int, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for ano in sorted(por_ano):
        df = por_ano[ano]
        oficiais = df[df["oficial"] == True]  # noqa: E712
        if oficiais.empty:
            oficiais = df
        if oficiais.empty:
            continue
        br = oficiais[oficiais[COL_COD] == "BR"]
        top = oficiais.head(3)

        def _pais(i: int) -> str | None:
            return None if len(top) <= i else str(top.iloc[i][COL_PAIS])

        def _taxa(i: int) -> float | None:
            if len(top) <= i:
                return None
            v = top.iloc[i][COL_REAL]
            return float(v) if pd.notna(v) else None

        cobertura = (
            "parcial"
            if oficiais.empty or oficiais[COL_COB].astype(str).str.contains("parcial").all()
            else "ano completo"
        )
        rows.append(
            {
                "ano": ano,
                "n_paises": int(len(oficiais)),
                "cobertura": cobertura,
                "1_pais": _pais(0),
                "1_taxa": _taxa(0),
                "2_pais": _pais(1),
                "2_taxa": _taxa(1),
                "3_pais": _pais(2),
                "3_taxa": _taxa(2),
                "brasil_pos": int(br.iloc[0][COL_POS]) if not br.empty and pd.notna(br.iloc[0][COL_POS]) else None,
                "brasil_taxa": float(br.iloc[0][COL_REAL]) if not br.empty and pd.notna(br.iloc[0][COL_REAL]) else None,
                "mediana": float(oficiais[COL_REAL].median()) if oficiais[COL_REAL].notna().any() else None,
            }
        )
    return pd.DataFrame(rows)


def evolucao_brasil(por_ano: dict[int, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for ano in sorted(por_ano):
        df = por_ano[ano]
        br = df[df[COL_COD] == "BR"]
        if br.empty:
            continue
        rec = br.iloc[0]
        rows.append(
            {
                "ano": ano,
                COL_POS: rec[COL_POS],
                COL_REAL: rec[COL_REAL],
                COL_INFL: rec[COL_INFL],
                COL_NOM: rec[COL_NOM],
                COL_NOM_FIM: rec[COL_NOM_FIM],
                COL_MESES: rec[COL_MESES],
                COL_COB: rec[COL_COB],
                COL_VAR: rec[COL_VAR],
                "n_paises": int((df["oficial"] == True).sum()) or int(len(df)),  # noqa: E712
            }
        )
    return pd.DataFrame(rows)


def escrever_capa(wb: Workbook, resumo: pd.DataFrame, sty: dict, gerado: datetime, anos: list[int]) -> None:
    ws = wb.active
    ws.title = "Capa"
    ws["A1"] = "Discriminativo — ranking das taxas básicas de juros reais anuais"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1B4F72")
    ws.merge_cells("A1:B1")
    periodo = f"{anos[0]}–{anos[-1]}" if anos else ""
    linhas = [
        ("Período", periodo),
        ("Fonte", f"BIS Data Portal — {BIS_BULK}"),
        ("Taxa básica nominal", "WS_CBPOL, mensal, fim de período"),
        ("Índice de inflação", "WS_LONG_CPI, índice 2010=100"),
        ("Real anual", "R = Π(1+r_m) − 1  (Fisher composto no ano-calendário)"),
        ("Ranking oficial", "Somente países com 12 meses e dezembro no ano"),
        ("Ranking parcial", "Ano sem dezembro (ex.: 2026): acumulado até o último mês"),
        ("Ordem", "Maior taxa real acumulada → menor"),
        ("Anos com ranking", str(len(resumo))),
        ("Gerado em", gerado.strftime("%d/%m/%Y %H:%M")),
        ("Marker", MARKER),
    ]
    ws["A3"] = "Campo"
    ws["B3"] = "Valor"
    ws["A3"].font = sty["header"]
    ws["B3"].font = sty["header"]
    ws["A3"].fill = sty["header_fill"]
    ws["B3"].fill = sty["header_fill"]
    for i, (k, v) in enumerate(linhas, start=4):
        ws.cell(i, 1, k).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(i, 2, v).font = Font(name="Calibri", size=10)
    _aplicar_larguras(ws, {1: 28, 2: 100})


def escrever_resumo(wb: Workbook, resumo: pd.DataFrame, sty: dict) -> None:
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "Pódio e posição do Brasil — um ranking por ano"
    ws["A1"].font = sty["title"]
    ws.merge_cells("A1:L1")
    headers = [
        "Ano",
        "Cobertura",
        "N países no ranking",
        "1º país",
        "1º taxa real",
        "2º país",
        "2º taxa real",
        "3º país",
        "3º taxa real",
        "Posição do Brasil",
        "Taxa real do Brasil",
        "Mediana do ranking",
    ]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(3, j, h)
        cell.font = sty["header"]
        cell.fill = sty["header_fill"]
        cell.alignment = sty["center"]
        cell.border = sty["thin"]
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"
    if resumo.empty:
        return
    ws.auto_filter.ref = f"A3:L{3 + len(resumo)}"
    for i, rec in enumerate(resumo.to_dict("records"), start=4):
        vals = [
            rec["ano"],
            rec["cobertura"],
            rec["n_paises"],
            rec["1_pais"],
            rec["1_taxa"],
            rec["2_pais"],
            rec["2_taxa"],
            rec["3_pais"],
            rec["3_taxa"],
            rec["brasil_pos"],
            rec["brasil_taxa"],
            rec["mediana"],
        ]
        fill = sty["alt"] if i % 2 == 0 else None
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(i, j, None if val is None or (isinstance(val, float) and pd.isna(val)) else val)
            cell.border = sty["thin"]
            cell.font = sty["mes"]
            if fill:
                cell.fill = fill
            if j in (5, 7, 9, 11, 12) and val is not None and pd.notna(val):
                cell.number_format = sty["pct"]
    _aplicar_larguras(
        ws,
        {1: 8, 2: 16, 3: 22, 4: 22, 5: 16, 6: 22, 7: 16, 8: 22, 9: 16, 10: 20, 11: 22, 12: 20},
    )


def escrever_brasil(wb: Workbook, evo: pd.DataFrame, sty: dict) -> None:
    ws = wb.create_sheet("Brasil")
    ws["A1"] = "Brasil no ranking anual de juros reais"
    ws["A1"].font = sty["title"]
    ws.merge_cells("A1:I1")
    headers = [
        "Ano",
        COL_POS,
        COL_REAL,
        COL_INFL,
        COL_NOM,
        COL_NOM_FIM,
        COL_MESES,
        COL_COB,
        COL_VAR,
        "Países no ranking",
    ]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(3, j, h)
        cell.font = sty["header"]
        cell.fill = sty["header_fill"]
        cell.alignment = sty["center"]
        cell.border = sty["thin"]
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"
    if evo.empty:
        return
    ws.auto_filter.ref = f"A3:J{3 + len(evo)}"
    for i, rec in enumerate(evo.to_dict("records"), start=4):
        vals = [
            rec["ano"],
            rec[COL_POS],
            rec[COL_REAL],
            rec[COL_INFL],
            rec[COL_NOM],
            rec[COL_NOM_FIM],
            rec[COL_MESES],
            rec[COL_COB],
            rec[COL_VAR],
            rec["n_paises"],
        ]
        fill = PatternFill("solid", fgColor="FDEBD0")
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(i, j, None if val is None or (isinstance(val, float) and pd.isna(val)) else val)
            cell.border = sty["thin"]
            cell.font = sty["mes"]
            cell.fill = fill
            if j in (3, 4, 5, 6) and val is not None and pd.notna(val):
                cell.number_format = sty["pct"]
    _aplicar_larguras(ws, {1: 10, 2: 12, 3: 36, 4: 28, 5: 36, 6: 42, 7: 12, 8: 30, 9: 36, 10: 20})


def escrever_aba_ano(wb: Workbook, ano: int, ranking: pd.DataFrame, sty: dict) -> None:
    ws = wb.create_sheet(str(ano))
    parcial = ranking.empty or not bool(ranking["oficial"].fillna(False).any())
    if not ranking.empty and ranking["oficial"].any():
        parcial = not ranking.loc[ranking["oficial"] == True, COL_COB].astype(str).str.contains("completo").any()  # noqa: E712
    titulo = f"Ranking {ano} — taxa básica de juros real acumulada no ano"
    if parcial:
        titulo += " (parcial)"
    ws["A1"] = titulo
    ws["A1"].font = sty["title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 22
    ws["A2"] = (
        "Ordem: maior juro real → menor. "
        "Oficial: 12 meses com dezembro. "
        "Série incompleta (sem posição) listada ao final. "
        f"Fonte: BIS CBPOL + CPI. {MARKER}"
    )
    ws["A2"].font = sty["subtitle"]
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 28

    wrap_cab = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, h in enumerate(HEADERS_RANK, start=1):
        cell = ws.cell(4, j, h)
        cell.font = sty["header"]
        cell.fill = sty["header_fill"]
        cell.alignment = wrap_cab
        cell.border = sty["thin"]
    ws.row_dimensions[4].height = 48
    ws.freeze_panes = "A5"
    if ranking.empty:
        ws["A5"] = "Sem países com taxa real neste ano."
        return
    ws.auto_filter.ref = f"A4:J{4 + len(ranking)}"

    ouro = PatternFill("solid", fgColor="F7DC6F")
    prata = PatternFill("solid", fgColor="D5D8DC")
    bronze = PatternFill("solid", fgColor="EDBB99")
    brasil = PatternFill("solid", fgColor="F5B041")
    incompleto = sty["parcial_fill"]

    r = 5
    for rec in ranking.itertuples(index=False):
        pos, pais, cod, real, infl, nom, nom_fim, meses, cob, var, oficial = rec
        fill = None
        font = sty["mes"]
        if not oficial:
            fill = incompleto
        elif pais == "Brasil":
            fill = brasil
            font = sty["acum"]
        elif pos == 1:
            fill = ouro
            font = sty["acum"]
        elif pos == 2:
            fill = prata
            font = sty["acum"]
        elif pos == 3:
            fill = bronze
            font = sty["acum"]
        elif r % 2 == 0:
            fill = sty["alt"]
        valores = (pos, pais, cod, real, infl, nom, nom_fim, meses, cob, var)
        for j, val in enumerate(valores, start=1):
            cell = ws.cell(r, j)
            cell.font = font
            cell.border = sty["thin"]
            if fill is not None:
                cell.fill = fill
            if val is None or (not isinstance(val, str) and pd.isna(val)):
                cell.value = None
            else:
                cell.value = val
            if j in (4, 5, 6, 7) and val is not None and not isinstance(val, str) and pd.notna(val):
                cell.number_format = sty["pct"]
            if j == 1:
                cell.alignment = sty["center"]
        r += 1

    _aplicar_larguras(ws, LARGURAS_ABA_ANO)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.horizontalCentered = True
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is not None:
        ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.view = "normal"


def escrever_planilha(
    por_ano: dict[int, pd.DataFrame],
    saida: Path,
    *,
    anos_pedido: list[int],
) -> Path:
    saida.parent.mkdir(parents=True, exist_ok=True)
    sty = _estilo_base()
    wb = Workbook()
    resumo = resumo_rankings(por_ano)
    evo = evolucao_brasil(por_ano)
    anos_com_dados = sorted(por_ano)
    escrever_capa(wb, resumo, sty, datetime.now(), anos_pedido)
    escrever_resumo(wb, resumo, sty)
    escrever_brasil(wb, evo, sty)
    for ano in anos_pedido:
        ranking = por_ano.get(ano, pd.DataFrame(columns=list(COLS_RANK) + ["oficial"]))
        print(f"[ABA] {ano}: {len(ranking)} países")
        escrever_aba_ano(wb, ano, ranking, sty)
    wb.save(saida)
    print(f"[OK] Planilha: {saida} ({saida.stat().st_size / 1e6:.2f} MB) | anos {anos_com_dados}")
    return saida


def montar_rankings(
    por_serie: dict[str, pd.DataFrame],
    anos: list[int],
) -> dict[int, pd.DataFrame]:
    anuais = consolidar_anuais(por_serie)
    out: dict[int, pd.DataFrame] = {}
    for ano in anos:
        ant = anuais[anuais["ano"] == ano - 1] if not anuais.empty else None
        out[ano] = ranking_do_ano(anuais, ano, anuais_ant=ant)
    return out


def processar(
    pasta_cache: Path,
    saida: Path,
    *,
    baixar: bool = True,
    ano_inicio: int = ANO_INICIO_DEFAULT,
    ano_fim: int = ANO_FIM_DEFAULT,
    paises: set[str] | None = None,
    cbpol_zip: Path | None = None,
    cpi_zip: Path | None = None,
) -> Path:
    print(f"[{MARKER}] ranking {ano_inicio}–{ano_fim}")
    por_serie, _ = carregar_series_paises(
        pasta_cache,
        baixar=baixar,
        ano_inicio=ano_inicio,
        ano_fim=ano_fim,
        paises=paises,
        cbpol_zip=cbpol_zip,
        cpi_zip=cpi_zip,
    )
    anos = list(range(int(ano_inicio), int(ano_fim) + 1))
    por_ano = montar_rankings(por_serie, anos)
    return escrever_planilha(por_ano, saida, anos_pedido=anos)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Ranking anual das taxas básicas de juros reais (uma aba por ano)."
    )
    p.add_argument(
        "--pasta-cache",
        type=Path,
        default=REPO_ROOT / "data" / "raw" / "bis",
    )
    p.add_argument(
        "--saida",
        type=Path,
        default=REPO_ROOT / "output" / "discriminativo_ranking_juros_reais.xlsx",
    )
    p.add_argument("--ano-inicio", type=int, default=ANO_INICIO_DEFAULT)
    p.add_argument("--ano-fim", type=int, default=ANO_FIM_DEFAULT)
    p.add_argument("--paises", type=str, default="")
    p.add_argument("--cbpol", type=Path, default=None)
    p.add_argument("--cpi", type=Path, default=None)
    p.add_argument("--sem-download", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    paises = {x.strip().upper() for x in args.paises.split(",") if x.strip()} or None
    processar(
        args.pasta_cache,
        args.saida,
        baixar=not args.sem_download,
        ano_inicio=args.ano_inicio,
        ano_fim=args.ano_fim,
        paises=paises,
        cbpol_zip=args.cbpol,
        cpi_zip=args.cpi,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
