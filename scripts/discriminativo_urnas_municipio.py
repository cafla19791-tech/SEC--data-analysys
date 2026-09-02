#!/usr/bin/env python3
"""Discriminativo municipal: UE2020 vs urnas anteriores a 2020.

Lê urnas_2t_presidente.csv (2º turno 2022) e, por município, compara
votos e vitórias nas urnas de modelo 2020 com as de modelos anteriores
(UE2009–UE2015).

Saída (ContAgil):
  saida/tse2022/discriminativo_municipio_ue2020.xlsx
  saida/tse2022/discriminativo_municipio_ue2020.csv

Uso:
  python discriminativo_urnas_municipio.py
  python discriminativo_urnas_municipio.py --entrada saida/tse2022/urnas_2t_presidente.csv
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

GERACAO_UE2020 = "UE2020"
GERACAO_ANTERIOR = "ANTERIOR_2020"
GERACAO_SEM_FAIXA = "SEM_FAIXA"

CHAVES_MUN = ("SG_UF", "CD_MUNICIPIO", "NM_MUNICIPIO")

COLUNAS_MUN = [
    "SG_UF",
    "CD_MUNICIPIO",
    "NM_MUNICIPIO",
    "COMPARAVEL",
    "QT_URNAS_PRE2020",
    "QT_URNAS_UE2020",
    "QT_VOTOS_LULA_PRE2020",
    "QT_VOTOS_BOLSONARO_PRE2020",
    "QT_VOTOS_VALIDOS_PRE2020",
    "PCT_LULA_PRE2020",
    "PCT_BOLSONARO_PRE2020",
    "VENCEDOR_VOTOS_PRE2020",
    "QT_VITORIAS_LULA_PRE2020",
    "QT_VITORIAS_BOLSONARO_PRE2020",
    "PCT_VITORIAS_LULA_PRE2020",
    "PCT_VITORIAS_BOLSONARO_PRE2020",
    "QT_VOTOS_LULA_UE2020",
    "QT_VOTOS_BOLSONARO_UE2020",
    "QT_VOTOS_VALIDOS_UE2020",
    "PCT_LULA_UE2020",
    "PCT_BOLSONARO_UE2020",
    "VENCEDOR_VOTOS_UE2020",
    "QT_VITORIAS_LULA_UE2020",
    "QT_VITORIAS_BOLSONARO_UE2020",
    "PCT_VITORIAS_LULA_UE2020",
    "PCT_VITORIAS_BOLSONARO_UE2020",
    "DIF_PCT_LULA",
    "DIF_PCT_BOLSONARO",
    "DIF_PCT_VITORIAS_LULA",
    "DIF_PCT_VITORIAS_BOLSONARO",
    "INVERTEU_VENCEDOR_VOTOS",
]


def classificar_geracao(nr_modelo: object) -> str:
    if nr_modelo is None or pd.isna(nr_modelo):
        return GERACAO_SEM_FAIXA
    try:
        ano = int(nr_modelo)
    except (TypeError, ValueError):
        return GERACAO_SEM_FAIXA
    if ano >= 2020:
        return GERACAO_UE2020
    if ano > 0:
        return GERACAO_ANTERIOR
    return GERACAO_SEM_FAIXA


def vencedor_votos(lula: float, bolso: float) -> str:
    if lula > bolso:
        return "Lula"
    if bolso > lula:
        return "Bolsonaro"
    return "Empate"


def _pct(parte: float, total: float) -> float | None:
    if total is None or pd.isna(total) or total <= 0:
        return None
    return round(100.0 * float(parte) / float(total), 2)


def preparar_urnas(df: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por urna, com geração do modelo e vencedor da urna."""
    base = df.copy()
    if "NR_MODELO" not in base.columns:
        raise ValueError("Falta NR_MODELO. Use urnas_2t_presidente.csv gerado pelo script TSE.")
    for col in ("QT_VOTOS_LULA", "QT_VOTOS_BOLSONARO", "QT_VOTOS_VALIDOS", "NR_MODELO"):
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce")
    if "QT_VOTOS_VALIDOS" not in base.columns:
        base["QT_VOTOS_VALIDOS"] = base["QT_VOTOS_LULA"].fillna(0) + base[
            "QT_VOTOS_BOLSONARO"
        ].fillna(0)
    base["GERACAO"] = base["NR_MODELO"].map(classificar_geracao)
    base["VENCEDOR_URNA"] = [
        vencedor_votos(l, b)
        for l, b in zip(
            base["QT_VOTOS_LULA"].fillna(0),
            base["QT_VOTOS_BOLSONARO"].fillna(0),
        )
    ]
    return base


def _bloco(grupo: pd.DataFrame, sufixo: str) -> dict[str, object]:
    urnas = int(len(grupo))
    lula = float(grupo["QT_VOTOS_LULA"].fillna(0).sum())
    bolso = float(grupo["QT_VOTOS_BOLSONARO"].fillna(0).sum())
    validos = float(grupo["QT_VOTOS_VALIDOS"].fillna(0).sum())
    vit_lula = int((grupo["VENCEDOR_URNA"] == "Lula").sum())
    vit_bolso = int((grupo["VENCEDOR_URNA"] == "Bolsonaro").sum())
    return {
        f"QT_URNAS_{sufixo}": urnas,
        f"QT_VOTOS_LULA_{sufixo}": int(lula),
        f"QT_VOTOS_BOLSONARO_{sufixo}": int(bolso),
        f"QT_VOTOS_VALIDOS_{sufixo}": int(validos),
        f"PCT_LULA_{sufixo}": _pct(lula, validos),
        f"PCT_BOLSONARO_{sufixo}": _pct(bolso, validos),
        f"VENCEDOR_VOTOS_{sufixo}": vencedor_votos(lula, bolso) if validos > 0 else "",
        f"QT_VITORIAS_LULA_{sufixo}": vit_lula,
        f"QT_VITORIAS_BOLSONARO_{sufixo}": vit_bolso,
        f"PCT_VITORIAS_LULA_{sufixo}": _pct(vit_lula, urnas),
        f"PCT_VITORIAS_BOLSONARO_{sufixo}": _pct(vit_bolso, urnas),
    }


def _vazio(sufixo: str) -> dict[str, object]:
    return {
        f"QT_URNAS_{sufixo}": 0,
        f"QT_VOTOS_LULA_{sufixo}": 0,
        f"QT_VOTOS_BOLSONARO_{sufixo}": 0,
        f"QT_VOTOS_VALIDOS_{sufixo}": 0,
        f"PCT_LULA_{sufixo}": None,
        f"PCT_BOLSONARO_{sufixo}": None,
        f"VENCEDOR_VOTOS_{sufixo}": "",
        f"QT_VITORIAS_LULA_{sufixo}": 0,
        f"QT_VITORIAS_BOLSONARO_{sufixo}": 0,
        f"PCT_VITORIAS_LULA_{sufixo}": None,
        f"PCT_VITORIAS_BOLSONARO_{sufixo}": None,
    }


def _linha_comparacao(pre: dict, ue: dict) -> dict[str, object]:
    comparavel = (
        int(pre["QT_VOTOS_VALIDOS_PRE2020"]) > 0
        and int(ue["QT_VOTOS_VALIDOS_UE2020"]) > 0
    )
    dif_lula = None
    dif_bolso = None
    dif_vit_lula = None
    dif_vit_bolso = None
    if comparavel:
        dif_lula = round(ue["PCT_LULA_UE2020"] - pre["PCT_LULA_PRE2020"], 2)
        dif_bolso = round(ue["PCT_BOLSONARO_UE2020"] - pre["PCT_BOLSONARO_PRE2020"], 2)
        if pre["PCT_VITORIAS_LULA_PRE2020"] is not None and ue[
            "PCT_VITORIAS_LULA_UE2020"
        ] is not None:
            dif_vit_lula = round(
                ue["PCT_VITORIAS_LULA_UE2020"] - pre["PCT_VITORIAS_LULA_PRE2020"], 2
            )
            dif_vit_bolso = round(
                ue["PCT_VITORIAS_BOLSONARO_UE2020"]
                - pre["PCT_VITORIAS_BOLSONARO_PRE2020"],
                2,
            )
    v_pre = pre["VENCEDOR_VOTOS_PRE2020"]
    v_ue = ue["VENCEDOR_VOTOS_UE2020"]
    inverteu = "S" if comparavel and v_pre and v_ue and v_pre != v_ue else "N"
    return {
        "COMPARAVEL": "S" if comparavel else "N",
        **pre,
        **ue,
        "DIF_PCT_LULA": dif_lula,
        "DIF_PCT_BOLSONARO": dif_bolso,
        "DIF_PCT_VITORIAS_LULA": dif_vit_lula,
        "DIF_PCT_VITORIAS_BOLSONARO": dif_vit_bolso,
        "INVERTEU_VENCEDOR_VOTOS": inverteu,
    }


def discriminar_municipios(df: pd.DataFrame) -> pd.DataFrame:
    """Uma linha por município: votos e vitórias pré-2020 vs UE2020."""
    urnas = preparar_urnas(df)
    linhas: list[dict] = []
    chaves = [c for c in CHAVES_MUN if c in urnas.columns]
    for chaves_mun, bloco in urnas.groupby(chaves, dropna=False):
        if not isinstance(chaves_mun, tuple):
            chaves_mun = (chaves_mun,)
        meta = dict(zip(chaves, chaves_mun))
        pre = bloco.loc[bloco["GERACAO"] == GERACAO_ANTERIOR]
        ue = bloco.loc[bloco["GERACAO"] == GERACAO_UE2020]
        pre_b = _bloco(pre, "PRE2020") if len(pre) else _vazio("PRE2020")
        ue_b = _bloco(ue, "UE2020") if len(ue) else _vazio("UE2020")
        linhas.append({**meta, **_linha_comparacao(pre_b, ue_b)})
    out = pd.DataFrame(linhas)
    for col in COLUNAS_MUN:
        if col not in out.columns:
            out[col] = None
    out = out[COLUNAS_MUN]
    return out.sort_values(["SG_UF", "NM_MUNICIPIO", "CD_MUNICIPIO"]).reset_index(
        drop=True
    )


def discriminar_ufs(mun: pd.DataFrame) -> pd.DataFrame:
    """Reagrega o discriminativo municipal por UF (só municípios comparáveis)."""
    base = mun.loc[mun["COMPARAVEL"] == "S"].copy()
    if base.empty:
        return pd.DataFrame()
    g = (
        base.groupby("SG_UF", dropna=False)
        .agg(
            QT_MUNICIPIOS=("CD_MUNICIPIO", "nunique"),
            QT_INVERTERAM=("INVERTEU_VENCEDOR_VOTOS", lambda s: int((s == "S").sum())),
            QT_URNAS_PRE2020=("QT_URNAS_PRE2020", "sum"),
            QT_URNAS_UE2020=("QT_URNAS_UE2020", "sum"),
            QT_VOTOS_LULA_PRE2020=("QT_VOTOS_LULA_PRE2020", "sum"),
            QT_VOTOS_BOLSONARO_PRE2020=("QT_VOTOS_BOLSONARO_PRE2020", "sum"),
            QT_VOTOS_VALIDOS_PRE2020=("QT_VOTOS_VALIDOS_PRE2020", "sum"),
            QT_VOTOS_LULA_UE2020=("QT_VOTOS_LULA_UE2020", "sum"),
            QT_VOTOS_BOLSONARO_UE2020=("QT_VOTOS_BOLSONARO_UE2020", "sum"),
            QT_VOTOS_VALIDOS_UE2020=("QT_VOTOS_VALIDOS_UE2020", "sum"),
            QT_VITORIAS_LULA_PRE2020=("QT_VITORIAS_LULA_PRE2020", "sum"),
            QT_VITORIAS_BOLSONARO_PRE2020=("QT_VITORIAS_BOLSONARO_PRE2020", "sum"),
            QT_VITORIAS_LULA_UE2020=("QT_VITORIAS_LULA_UE2020", "sum"),
            QT_VITORIAS_BOLSONARO_UE2020=("QT_VITORIAS_BOLSONARO_UE2020", "sum"),
        )
        .reset_index()
    )
    g["PCT_LULA_PRE2020"] = [
        _pct(a, b) for a, b in zip(g["QT_VOTOS_LULA_PRE2020"], g["QT_VOTOS_VALIDOS_PRE2020"])
    ]
    g["PCT_LULA_UE2020"] = [
        _pct(a, b) for a, b in zip(g["QT_VOTOS_LULA_UE2020"], g["QT_VOTOS_VALIDOS_UE2020"])
    ]
    g["DIF_PCT_LULA"] = [
        round(u - p, 2) if p is not None and u is not None else None
        for p, u in zip(g["PCT_LULA_PRE2020"], g["PCT_LULA_UE2020"])
    ]
    g["PCT_MUN_INVERTERAM"] = [
        _pct(a, b) for a, b in zip(g["QT_INVERTERAM"], g["QT_MUNICIPIOS"])
    ]
    return g.sort_values("SG_UF").reset_index(drop=True)


def _linha_brasil(recorte: str, pre: pd.DataFrame, ue: pd.DataFrame, *, mun_comp: pd.DataFrame | None = None) -> dict:
    lula_pre = float(pre["QT_VOTOS_LULA"].fillna(0).sum()) if "QT_VOTOS_LULA" in pre.columns else float(pre.get("QT_VOTOS_LULA_PRE2020", pd.Series(dtype=float)).fillna(0).sum())
    # Usado só no recorte municipal agregado (colunas já sufixadas).
    if "QT_VOTOS_LULA_PRE2020" in (mun_comp.columns if mun_comp is not None else []):
        bloco = mun_comp
        lula_pre = int(bloco["QT_VOTOS_LULA_PRE2020"].sum())
        bolso_pre = int(bloco["QT_VOTOS_BOLSONARO_PRE2020"].sum())
        val_pre = int(bloco["QT_VOTOS_VALIDOS_PRE2020"].sum())
        lula_ue = int(bloco["QT_VOTOS_LULA_UE2020"].sum())
        bolso_ue = int(bloco["QT_VOTOS_BOLSONARO_UE2020"].sum())
        val_ue = int(bloco["QT_VOTOS_VALIDOS_UE2020"].sum())
        n_mun = int(len(bloco))
        n_inv = int((bloco["INVERTEU_VENCEDOR_VOTOS"] == "S").sum())
        urnas_pre = int(bloco["QT_URNAS_PRE2020"].sum())
        urnas_ue = int(bloco["QT_URNAS_UE2020"].sum())
    else:
        bolso_pre = float(pre["QT_VOTOS_BOLSONARO"].fillna(0).sum())
        val_pre = float(pre["QT_VOTOS_VALIDOS"].fillna(0).sum())
        lula_ue = float(ue["QT_VOTOS_LULA"].fillna(0).sum())
        bolso_ue = float(ue["QT_VOTOS_BOLSONARO"].fillna(0).sum())
        val_ue = float(ue["QT_VOTOS_VALIDOS"].fillna(0).sum())
        n_mun = None
        n_inv = None
        urnas_pre = int(len(pre))
        urnas_ue = int(len(ue))
    pct_pre = _pct(lula_pre, val_pre)
    pct_ue = _pct(lula_ue, val_ue)
    return {
        "RECORTE": recorte,
        "QT_MUNICIPIOS": n_mun,
        "QT_INVERTERAM": n_inv,
        "PCT_MUN_INVERTERAM": _pct(n_inv, n_mun) if n_mun else None,
        "QT_URNAS_PRE2020": urnas_pre,
        "QT_URNAS_UE2020": urnas_ue,
        "PCT_LULA_PRE2020": pct_pre,
        "PCT_LULA_UE2020": pct_ue,
        "DIF_PCT_LULA": round(pct_ue - pct_pre, 2)
        if pct_pre is not None and pct_ue is not None
        else None,
        "PCT_BOLSONARO_PRE2020": _pct(bolso_pre, val_pre),
        "PCT_BOLSONARO_UE2020": _pct(bolso_ue, val_ue),
        "VENCEDOR_VOTOS_PRE2020": vencedor_votos(lula_pre, bolso_pre),
        "VENCEDOR_VOTOS_UE2020": vencedor_votos(lula_ue, bolso_ue),
    }


def discriminar_brasil(mun: pd.DataFrame, urnas: pd.DataFrame | None = None) -> pd.DataFrame:
    comparavel = mun.loc[mun["COMPARAVEL"] == "S"]
    if comparavel.empty:
        comparavel = mun
    linhas = [
        _linha_brasil(
            "Municípios com os dois tipos de urna",
            pd.DataFrame(),
            pd.DataFrame(),
            mun_comp=comparavel,
        )
    ]
    if urnas is not None:
        prep = preparar_urnas(urnas)
        linhas.append(
            _linha_brasil(
                "Todas as urnas do país",
                prep.loc[prep["GERACAO"] == GERACAO_ANTERIOR],
                prep.loc[prep["GERACAO"] == GERACAO_UE2020],
            )
        )
    return pd.DataFrame(linhas)


def ler_urnas(caminho: Path) -> pd.DataFrame:
    if not caminho.exists():
        gz = caminho.with_suffix(caminho.suffix + ".gz")
        if caminho.suffix == ".csv" and gz.exists():
            caminho = gz
        elif caminho.with_name("urnas_2t_presidente.csv.gz").exists():
            caminho = caminho.with_name("urnas_2t_presidente.csv.gz")
        else:
            raise FileNotFoundError(
                f"Não achei {caminho}. No ContAgil rode antes:\n"
                "  python baixar_boletins_urna_2022.py --somente-resultado-github"
            )
    if str(caminho).endswith(".gz"):
        return pd.read_csv(caminho, compression="gzip")
    return pd.read_csv(caminho)


def _estilo_aba(ws, linhas: int, colunas: int) -> None:
    header = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = header
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(colunas)}{max(1, linhas)}"
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 32
    for col in range(1, colunas + 1):
        letra = get_column_letter(col)
        tamanho = 14
        for row in range(1, min(linhas, 40) + 1):
            val = ws.cell(row, col).value
            if val is not None:
                tamanho = max(tamanho, min(28, len(str(val)) + 2))
        ws.column_dimensions[letra].width = tamanho
        for row in range(1, linhas + 1):
            ws.cell(row, col).border = thin


def escrever_planilha(
    mun: pd.DataFrame,
    ufs: pd.DataFrame,
    brasil: pd.DataFrame,
    destino: Path,
) -> Path:
    destino.parent.mkdir(parents=True, exist_ok=True)
    comparaveis = mun.loc[mun["COMPARAVEL"] == "S"].copy()
    inverteram = comparaveis.loc[comparaveis["INVERTEU_VENCEDOR_VOTOS"] == "S"].copy()
    leia = pd.DataFrame(
        [
            {
                "campo": "Fonte",
                "valor": "Boletins de Urna oficiais TSE, 2º turno 2022 (Presidente)",
            },
            {
                "campo": "UE2020",
                "valor": "Urnas modelo 2020 (NR_MODELO >= 2020)",
            },
            {
                "campo": "PRE2020",
                "valor": "Urnas UE2009, UE2010, UE2011, UE2013 e UE2015",
            },
            {
                "campo": "COMPARAVEL",
                "valor": "S = município tem votos válidos nos dois grupos de urna",
            },
            {
                "campo": "PCT_LULA_*",
                "valor": "% dos votos válidos (Lula / (Lula+Bolsonaro)) naquele grupo",
            },
            {
                "campo": "PCT_VITORIAS_*",
                "valor": "% das urnas do município em que aquele candidato teve mais votos",
            },
            {
                "campo": "DIF_PCT_LULA",
                "valor": "PCT_LULA_UE2020 − PCT_LULA_PRE2020 (pontos percentuais). Positivo = Lula foi melhor nas UE2020 daquele município",
            },
            {
                "campo": "INVERTEU_VENCEDOR_VOTOS",
                "valor": "S = o vencedor pelos votos válidos muda entre pré-2020 e UE2020 no mesmo município",
            },
        ]
    )
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        brasil.to_excel(writer, sheet_name="Brasil", index=False)
        ufs.to_excel(writer, sheet_name="Por_UF", index=False)
        comparaveis.to_excel(writer, sheet_name="Municipios_comparaveis", index=False)
        inverteram.to_excel(writer, sheet_name="Municipios_inverteram", index=False)
        mun.to_excel(writer, sheet_name="Todos_municipios", index=False)
        leia.to_excel(writer, sheet_name="Leia-me", index=False)
        for nome, df in (
            ("Brasil", brasil),
            ("Por_UF", ufs),
            ("Municipios_comparaveis", comparaveis),
            ("Municipios_inverteram", inverteram),
            ("Todos_municipios", mun),
            ("Leia-me", leia),
        ):
            ws = writer.sheets[nome]
            _estilo_aba(ws, len(df) + 1, max(1, len(df.columns)))
    return destino


def descobrir_entrada(argv_entrada: Path | None) -> Path:
    if argv_entrada is not None:
        return Path(argv_entrada)
    cwd = Path.cwd()
    candidatos = [
        cwd / "saida" / "tse2022" / "urnas_2t_presidente.csv",
        cwd / "saida" / "tse2022" / "urnas_2t_presidente.csv.gz",
        Path("output/tse2022/urnas_2t_presidente.csv"),
        Path("output/tse2022/urnas_2t_presidente.csv.gz"),
    ]
    for cand in candidatos:
        if cand.exists():
            return cand
    return candidatos[0]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--entrada", type=Path, default=None, help="CSV/CSV.GZ de urnas")
    p.add_argument(
        "--pasta-saida",
        type=Path,
        default=None,
        help="Pasta saida (CSVs em saida/tse2022) ou pasta tse2022.",
    )
    return p.parse_args(argv)


def pasta_saida(args: argparse.Namespace, entrada: Path) -> Path:
    if args.pasta_saida is not None:
        pasta = Path(args.pasta_saida)
        return pasta if pasta.name.lower() == "tse2022" else pasta / "tse2022"
    return entrada.parent


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    entrada = descobrir_entrada(args.entrada)
    saida = pasta_saida(args, entrada)
    saida.mkdir(parents=True, exist_ok=True)
    print(f"Entrada: {entrada}", flush=True)
    print(f"Saida  : {saida}", flush=True)
    urnas = ler_urnas(entrada)
    mun = discriminar_municipios(urnas)
    ufs = discriminar_ufs(mun)
    brasil = discriminar_brasil(mun, urnas)
    csv_path = saida / "discriminativo_municipio_ue2020.csv"
    xlsx_path = saida / "discriminativo_municipio_ue2020.xlsx"
    mun.to_csv(csv_path, index=False, encoding="utf-8")
    escrever_planilha(mun, ufs, brasil, xlsx_path)
    n_comp = int((mun["COMPARAVEL"] == "S").sum())
    n_inv = int((mun["INVERTEU_VENCEDOR_VOTOS"] == "S").sum())
    print(f"Municípios: {len(mun)}")
    print(f"Comparáveis (pré-2020 e UE2020): {n_comp}")
    print(f"Inverteram o vencedor: {n_inv}")
    if not brasil.empty:
        row = brasil.iloc[0]
        print(
            f"Brasil (comparáveis): Lula pré-2020 {row['PCT_LULA_PRE2020']}% "
            f"vs UE2020 {row['PCT_LULA_UE2020']}% "
            f"(dif {row['DIF_PCT_LULA']} p.p.)"
        )
    print(f"CSV : {csv_path}")
    print(f"XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
