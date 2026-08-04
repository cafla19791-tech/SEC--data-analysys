#!/usr/bin/env python3
"""Discriminativos de OPERAÇÕES DIRETAS com valor atualizado pelo IPCA.

Lê a planilha ContAgil (ex.: OPERAÇÕES DIRETAS2002 A 302026.xlsx), acrescenta
a coluna:

  VALOR DESEMBOLSADO ATUALIZADO - IPCA-30 DE JUNHO DE 2026

e gera Excel com 4 abas por lapso temporal:

  1) 2002
  2) 2003-2018
  3) 2019-2022
  4) 2023-atual

Em cada aba: ordenação CLIENTE + DATA DO CONTRATO (crescente);
a cada mudança de cliente, linha de subtotal (Valor Desembolsado + IPCA)
e quebra de página.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

ROOT = Path(__file__).resolve().parents[1]
_SCRIPTS = Path(__file__).resolve().parent
for _p in (str(ROOT), str(_SCRIPTS)):
    if _p not in sys.path:
        sys.path.insert(0, _p)


def _load_sibling(mod_name: str, filename: str):
    """Importa modulo irmao: arquivo local primeiro (ContAgil sec_scripts)."""
    import importlib.util

    path = _SCRIPTS / filename
    if path.exists():
        spec = importlib.util.spec_from_file_location(f"sec_{mod_name}", path)
        if spec is not None and spec.loader is not None:
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    try:
        return __import__(f"scripts.{mod_name}", fromlist=["*"])
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(f"Nao achou {path} nem scripts.{mod_name}") from exc


_calc = _load_sibling("calcular_diretas_ipca_selic", "calcular_diretas_ipca_selic.py")
_flux = _load_sibling("gerar_fluxos", "gerar_fluxos.py")
DATA_REF_DEFAULT = _calc.DATA_REF_DEFAULT
IPCA_COD = _calc.IPCA_COD
_baixar_sgs = _calc._baixar_sgs
carregar_ipca = _calc.carregar_ipca
detectar_header_row = _calc.detectar_header_row
CONTAGIL_WINPYTHON = _flux.CONTAGIL_WINPYTHON
_mapear_colunas_contratos = _flux._mapear_colunas_contratos
limpar_valor = _flux.limpar_valor
parse_datas = _flux.parse_datas

COL_IPCA = "VALOR DESEMBOLSADO ATUALIZADO - IPCA-30 DE JUNHO DE 2026"
COL_VALOR = "Valor Desembolsado"
COL_CLIENTE = "Cliente"
COL_DATA = "Data do Contrato"

PERIODOS = [
    ("2002", 2002, 2002),
    ("2003-2018", 2003, 2018),
    ("2019-2022", 2019, 2022),
    ("2023-atual", 2023, 9999),
]

CANDIDATOS_EXCEL = [
    "OPERAÇÕES DIRETAS2002 A 302026.xlsx",
    "OPERACOES DIRETAS2002 A 302026.xlsx",
    "OPERAÇÕES DIRETAS 2002 A 302026.xlsx",
    "OPERACOES DIRETAS 2002 A 302026.xlsx",
    "OPERACOES DIRETAS - 2002 a 2018.xlsx",
    "OPERAÇÕES DIRETAS - 2002 a 2018.xlsx",
    "OPERACOES DIRETAS.xlsx",
]


def resolver_excel(path: Optional[Path]) -> Path:
    if path is not None and path.exists():
        return path
    bases = [Path.cwd(), CONTAGIL_WINPYTHON, ROOT]
    nomes = []
    if path is not None:
        nomes.append(path.name)
    nomes.extend(CANDIDATOS_EXCEL)
    for base in bases:
        for nome in nomes:
            cand = base / nome
            if cand.exists():
                return cand
        # glob frouxo
        for pat in ("*DIRETA*2002*", "*DIRETAS*2026*", "*OPERA*DIRETA*.xlsx"):
            hits = sorted(base.glob(pat))
            if hits:
                return hits[0]
    raise FileNotFoundError(
        "Planilha de OPERAÇÕES DIRETAS não encontrada.\n"
        "Informe --excel com o caminho completo, ex.:\n"
        r'  --excel "C:\Arquivos de Programas RFB\ContAgilAppBeta64\python_jep\winpython\OPERAÇÕES DIRETAS2002 A 302026.xlsx"'
    )


def carregar_ipca_cobrindo(path: Optional[Path], ano_min: int = 2000) -> pd.DataFrame:
    if path is not None and Path(path).exists():
        ipca = carregar_ipca(Path(path))
        if ipca["mes"].min().year <= ano_min:
            return ipca
        print("[AVISO] IPCA local incompleto; baixando Bacen SGS 433...")
    print(f"[INFO] Baixando IPCA (Bacen SGS 433) desde 01/01/{ano_min}...")
    raw = _baixar_sgs(IPCA_COD, inicio=f"01/01/{ano_min}")
    df = raw.sort_values("mes").drop_duplicates("mes").copy()
    df["fator"] = (1.0 + df["valor"] / 100.0).cumprod()
    return df.reset_index(drop=True)


def _fator_mapa(ipca: pd.DataFrame) -> pd.Series:
    s = ipca.copy()
    s["periodo"] = s["mes"].dt.to_period("M")
    return s.drop_duplicates("periodo").set_index("periodo")["fator"].astype(float).sort_index()


def atualizar_valores_ipca(
    datas: pd.Series,
    valores: pd.Series,
    ipca: pd.DataFrame,
    data_ref: datetime | pd.Timestamp,
) -> pd.Series:
    """valor × fator(ref) / fator(mês contratação) — vetorizado."""
    fator_por_mes = _fator_mapa(ipca)
    per_ref = pd.Period(pd.Timestamp(data_ref), freq="M")
    if per_ref not in fator_por_mes.index:
        ant = fator_por_mes.index[fator_por_mes.index <= per_ref]
        if len(ant) == 0:
            raise ValueError(f"IPCA sem dados até {data_ref}")
        per_ref = ant.max()
    fator_ref = float(fator_por_mes.loc[per_ref])

    periodos = pd.to_datetime(datas, errors="coerce").dt.to_period("M")
    full = pd.period_range(
        fator_por_mes.index.min(),
        max(periodos.dropna().max() if periodos.notna().any() else per_ref, per_ref),
        freq="M",
    )
    ffill = fator_por_mes.reindex(full).ffill()
    fator_ini = periodos.map(ffill)
    out = pd.to_numeric(valores, errors="coerce") * (fator_ref / fator_ini)
    return out.astype(float)


def preparar_base(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Mantém colunas originais + campos canônicos necessários."""
    mapped, rename = _mapear_colunas_contratos(df_raw)
    if rename:
        print(f"[INFO] Colunas mapeadas: {rename}")

    # cliente: se não mapeado, tenta coluna original
    if "cliente" not in mapped.columns:
        for c in df_raw.columns:
            if str(c).strip().lower() in {"cliente", "client", "nome do cliente"}:
                mapped["cliente"] = df_raw[c]
                break
        else:
            mapped["cliente"] = "(sem cliente)"

    if "valor_desembolsado" not in mapped.columns and "valor_contratado" in mapped.columns:
        mapped["valor_desembolsado"] = mapped["valor_contratado"]

    if "data_contratacao" not in mapped.columns or "valor_desembolsado" not in mapped.columns:
        raise ValueError(
            "É preciso ter Data da contratação e Valor desembolsado. "
            f"Colunas: {list(df_raw.columns)}"
        )

    out = mapped.copy()
    out["_cliente"] = out["cliente"].astype(str).str.strip().replace({"": "(sem cliente)", "nan": "(sem cliente)"})
    out["_data"] = parse_datas(out["data_contratacao"])
    out["_valor"] = limpar_valor(out["valor_desembolsado"]).fillna(0.0)
    out["_ano"] = out["_data"].dt.year
    return out


def filtrar_periodo(df: pd.DataFrame, ano_ini: int, ano_fim: int) -> pd.DataFrame:
    m = df["_ano"].between(ano_ini, ano_fim) & df["_data"].notna()
    parte = df.loc[m].copy()
    return parte.sort_values(["_cliente", "_data", "_valor"], kind="mergesort").reset_index(drop=True)


def colunas_exibicao(df: pd.DataFrame) -> tuple[list[str], dict[str, str]]:
    """Colunas amigáveis para o discriminativo (ordem ContAgil + IPCA ao final)."""
    preferidas = [
        ("cliente", COL_CLIENTE),
        ("cnpj", "CNPJ"),
        ("uf", "UF"),
        ("numero_contrato", "Número do contrato"),
        ("data_contratacao", COL_DATA),
        ("valor_desembolsado", COL_VALOR),
        ("custo_financeiro", "Custo financeiro"),
        ("juros", "Juros"),
        ("prazo_carencia", "Prazo carência (meses)"),
        ("prazo_amortizacao", "Prazo amortização (meses)"),
        ("forma_de_apoio", "Forma de apoio"),
    ]
    cols: list[str] = []
    rename: dict[str, str] = {}
    for canon, titulo in preferidas:
        if canon in df.columns or canon in {"cliente", "data_contratacao", "valor_desembolsado"}:
            rename[canon] = titulo
            if titulo not in cols:
                cols.append(titulo)
    for obrig in (COL_CLIENTE, COL_DATA, COL_VALOR):
        if obrig not in cols:
            cols.append(obrig)
    return cols, rename


def _celula_linha(row: pd.Series, titulo: str, rename: dict[str, str]):
    if titulo == COL_CLIENTE:
        return row["_cliente"]
    if titulo == COL_DATA:
        d = row["_data"]
        return d.strftime("%d/%m/%Y") if pd.notna(d) else ""
    if titulo == COL_VALOR:
        return float(row["_valor"])
    canon = next((c for c, t in rename.items() if t == titulo), None)
    if canon and canon in row.index:
        val = row[canon]
        if canon == "data_contratacao":
            return pd.Timestamp(val).strftime("%d/%m/%Y") if pd.notna(val) else ""
        return val if pd.notna(val) else ""
    return ""


def montar_linhas_aba(df: pd.DataFrame) -> tuple[list[str], list[list], list[int]]:
    """Retorna cabeçalho, linhas (dados+subtotais) e nº das linhas com quebra (1-based Excel)."""
    cols_titulo, rename = colunas_exibicao(df)
    header = cols_titulo + [COL_IPCA]
    idx_cliente = header.index(COL_CLIENTE)
    idx_valor = header.index(COL_VALOR)
    idx_ipca = header.index(COL_IPCA)

    linhas: list[list] = []
    quebras: list[int] = []

    if df.empty:
        return header, linhas, quebras

    grupos = list(df.groupby("_cliente", sort=False))
    for gi, (cliente, g) in enumerate(grupos):
        for _, row in g.iterrows():
            rec = [_celula_linha(row, h, rename) for h in cols_titulo]
            rec.append(float(row["_valor_ipca"]) if pd.notna(row["_valor_ipca"]) else None)
            linhas.append(rec)

        sub = [""] * len(header)
        sub[idx_cliente] = f"SUBTOTAL — {cliente}"
        sub[idx_valor] = float(g["_valor"].sum())
        sub[idx_ipca] = float(g["_valor_ipca"].sum())
        linhas.append(sub)

        # quebra após o subtotal (exceto último cliente)
        excel_row = 1 + len(linhas)  # header na linha 1
        if gi < len(grupos) - 1:
            quebras.append(excel_row)

    tot = [""] * len(header)
    tot[idx_cliente] = "TOTAL DA ABA"
    tot[idx_valor] = float(df["_valor"].sum())
    tot[idx_ipca] = float(df["_valor_ipca"].sum())
    linhas.append(tot)

    return header, linhas, quebras


def _estilo_planilha(ws, n_header_cols: int, n_data_rows: int, quebras: Iterable[int]) -> None:
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fill_header = PatternFill("solid", fgColor="1F4E79")
    font_header = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    fill_sub = PatternFill("solid", fgColor="D9E2F3")
    font_sub = Font(bold=True, name="Calibri", size=11)
    fill_tot = PatternFill("solid", fgColor="FFF2CC")
    font_tot = Font(bold=True, name="Calibri", size=11)

    for col in range(1, n_header_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for r in range(2, n_data_rows + 2):
        first = str(ws.cell(r, 1).value or "")
        is_sub = first.startswith("SUBTOTAL")
        is_tot = first.startswith("TOTAL DA ABA")
        for c in range(1, n_header_cols + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            if is_sub:
                cell.fill = fill_sub
                cell.font = font_sub
            elif is_tot:
                cell.fill = fill_tot
                cell.font = font_tot
            # valores monetários nas duas últimas colunas típicas
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    for row_id in quebras:
        ws.row_breaks.append(Break(id=row_id))

    # larguras
    for col in range(1, n_header_cols + 1):
        letter = get_column_letter(col)
        header = str(ws.cell(1, col).value or "")
        width = 18
        if "CLIENTE" in header.upper() or header.upper().startswith("SUBTOTAL"):
            width = 36
        elif "IPCA" in header.upper():
            width = 28
        elif "DATA" in header.upper():
            width = 14
        elif "DESEMBOLSADO" in header.upper() or "VALOR" in header.upper():
            width = 18
        ws.column_dimensions[letter].width = width


def escrever_workbook(
    periodos_dados: list[tuple[str, pd.DataFrame]],
    saida: Path,
) -> Path:
    wb = Workbook()
    # remove sheet default depois de criar a primeira
    default = wb.active
    first = True
    for nome, df in periodos_dados:
        header, linhas, quebras = montar_linhas_aba(df)
        if first:
            ws = default
            ws.title = nome[:31]
            first = False
        else:
            ws = wb.create_sheet(nome[:31])

        ws.append(header)
        for row in linhas:
            ws.append(row)
        _estilo_planilha(ws, len(header), len(linhas), quebras)

        # cabeçalho de impressão
        ws.oddHeader.center.text = f"Discriminativo OPERAÇÕES DIRETAS — {nome}"
        ws.oddFooter.center.text = "Página &P de &N | IPCA ref. 30/06/2026"

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    return saida


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--excel", type=Path, default=None, help="Planilha OPERAÇÕES DIRETAS")
    p.add_argument("--ipca", type=Path, default=None, help="IPCA_MENSAL.xlsx (opcional)")
    p.add_argument("--data-ref", default="2026-06-30", help="Data referência IPCA")
    p.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Excel de saída (default: saida/DISCRIMINATIVOS_DIRETAS_IPCA.xlsx)",
    )
    p.add_argument("--header-row", type=int, default=None, help="Linha cabeçalho 1-based")
    args = p.parse_args(argv)

    excel = resolver_excel(args.excel)
    print(f"[INFO] Excel: {excel}")

    header_row = args.header_row or detectar_header_row(excel)
    df_raw = pd.read_excel(excel, header=header_row - 1)
    print(f"[INFO] Linhas brutas: {len(df_raw):,}")

    base = preparar_base(df_raw)
    data_ref = datetime.strptime(args.data_ref, "%Y-%m-%d")

    ipca_path = args.ipca
    if ipca_path is None:
        for cand in (
            Path.cwd() / "IPCA_MENSAL.xlsx",
            CONTAGIL_WINPYTHON / "IPCA_MENSAL.xlsx",
        ):
            if cand.exists():
                ipca_path = cand
                break
    ipca = carregar_ipca_cobrindo(ipca_path, ano_min=2000)
    print(f"[INFO] IPCA: {ipca['mes'].min().date()} → {ipca['mes'].max().date()}")

    base["_valor_ipca"] = atualizar_valores_ipca(
        base["_data"], base["_valor"], ipca, data_ref
    )
    n_ok = int(base["_valor_ipca"].notna().sum())
    print(f"[INFO] Valores atualizados IPCA: {n_ok:,}")

    periodos_dados = []
    for nome, a0, a1 in PERIODOS:
        parte = filtrar_periodo(base, a0, a1)
        print(
            f"[INFO] Aba {nome}: {len(parte):,} contratos | "
            f"corrente={parte['_valor'].sum():,.2f} | IPCA={parte['_valor_ipca'].sum():,.2f}"
        )
        periodos_dados.append((nome, parte))

    if args.saida is None:
        saida_dir = Path.cwd() / "saida"
        if not saida_dir.exists() and (CONTAGIL_WINPYTHON / "saida").exists():
            saida_dir = CONTAGIL_WINPYTHON / "saida"
        elif not saida_dir.exists():
            saida_dir = ROOT / "output"
        saida = saida_dir / "DISCRIMINATIVOS_DIRETAS_IPCA.xlsx"
    else:
        saida = args.saida

    escrever_workbook(periodos_dados, saida)
    print(f"[OK] Discriminativos salvos em: {saida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
