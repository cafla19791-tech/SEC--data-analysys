#!/usr/bin/env python3
"""Discriminativos de fluxos BNDES INDIRETAS por ano do contrato.

Regra de negócio
----------------
O impacto fiscal de cada parcela continua calculado na ``data_fluxo``
(capitalização SELIC até a data de referência). O que muda é só a
**discriminação**: todas as parcelas de um contrato entram na planilha
do **ano da contratação**.

Exemplo: contrato em 12/12/2022 com 180 parcelas → as 180 linhas vão
para a aba/arquivo ``2022``, mesmo que as parcelas ocorram em 2023–2037.

Entrada típica (após ``contagil_fluxos.py``)::

  saida/fluxos_*.csv   (com colunas ano_contrato ou data_contratacao)

Saída::

  saida/discriminativos_ano_contrato/
      fluxos_ano_contrato_2002.csv
      fluxos_ano_contrato_2003.csv
      ...
      RESUMO_POR_ANO_CONTRATO.xlsx

Uso ContAgil::

  python scripts/discriminativos_indiretas_ano_contrato.py --pasta saida
"""

from __future__ import annotations

import argparse
import csv
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.gerar_fluxos import (  # noqa: E402
    CONTAGIL_PASTA_SAIDA,
    CONTAGIL_WINPYTHON,
    OUTPUT_DIR,
)

CHUNK_DEFAULT = 500_000
MARKER = "discriminativos-ano-contrato-20260813"


def listar_csvs_fluxos(pasta: Path) -> list[Path]:
    pasta = Path(pasta)
    if not pasta.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {pasta}")
    csvs = sorted(
        p
        for p in pasta.glob("fluxos_*.csv")
        if "diario" not in p.stem.lower()
        and "ano_contrato" not in p.stem.lower()
    )
    if not csvs:
        raise FileNotFoundError(
            f"Nenhum fluxos_*.csv em {pasta}. "
            "Gere com contagil_fluxos.py / contagil_fluxos_seguro.py primeiro."
        )
    return csvs


def _resolver_pasta(pasta: Optional[Path]) -> Path:
    if pasta is not None:
        return Path(pasta)
    if CONTAGIL_PASTA_SAIDA.exists():
        return CONTAGIL_PASTA_SAIDA
    if (CONTAGIL_WINPYTHON / "saida").exists():
        return CONTAGIL_WINPYTHON / "saida"
    return OUTPUT_DIR


def _ano_contrato_series(df: pd.DataFrame) -> pd.Series:
    """Obtém ano do contrato a partir das colunas disponíveis."""
    if "ano_contrato" in df.columns:
        return pd.to_numeric(df["ano_contrato"], errors="coerce")
    if "data_contratacao" in df.columns:
        return pd.to_datetime(df["data_contratacao"], errors="coerce").dt.year
    raise ValueError(
        "Fluxos sem 'ano_contrato' nem 'data_contratacao'. "
        "Regenere com contagil_fluxos.py (versão que grava ano_contrato) "
        f"[{MARKER}]."
    )


def repartir_streaming(
    arquivos: list[Path],
    pasta_saida: Path,
    *,
    chunksize: int = CHUNK_DEFAULT,
    ano_min: Optional[int] = None,
    ano_max: Optional[int] = None,
) -> dict[int, dict]:
    """Lê fluxos_*.csv em chunks e grava um CSV por ano_contrato.

    Retorna estatísticas por ano: parcelas, contratos, subsidio, impacto.
    """
    pasta_saida = Path(pasta_saida)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    writers: dict[int, csv.DictWriter] = {}
    handles: dict[int, object] = {}
    fieldnames: list[str] | None = None
    stats: dict[int, dict] = defaultdict(
        lambda: {
            "parcelas": 0,
            "contratos": set(),
            "subsidio": 0.0,
            "impacto_fiscal": 0.0,
        }
    )

    t0 = time.time()
    total_linhas = 0

    try:
        for arq in arquivos:
            print(f"[INFO] Lendo {arq.name} ...")
            for chunk in pd.read_csv(arq, chunksize=chunksize, low_memory=False):
                anos = _ano_contrato_series(chunk)
                chunk = chunk.copy()
                chunk["_ano_contrato"] = anos
                chunk = chunk.dropna(subset=["_ano_contrato"])
                chunk["_ano_contrato"] = chunk["_ano_contrato"].astype(int)
                if ano_min is not None:
                    chunk = chunk[chunk["_ano_contrato"] >= int(ano_min)]
                if ano_max is not None:
                    chunk = chunk[chunk["_ano_contrato"] <= int(ano_max)]
                if chunk.empty:
                    continue

                if "ano_contrato" not in chunk.columns:
                    chunk["ano_contrato"] = chunk["_ano_contrato"]

                if fieldnames is None:
                    fieldnames = [
                        c for c in chunk.columns if c != "_ano_contrato"
                    ]

                for ano, g in chunk.groupby("_ano_contrato", sort=False):
                    ano_i = int(ano)
                    if ano_i not in writers:
                        path = pasta_saida / f"fluxos_ano_contrato_{ano_i}.csv"
                        fh = path.open("w", newline="", encoding="utf-8")
                        handles[ano_i] = fh
                        writers[ano_i] = csv.DictWriter(
                            fh, fieldnames=fieldnames, extrasaction="ignore"
                        )
                        writers[ano_i].writeheader()

                    rows = g.drop(columns=["_ano_contrato"]).to_dict(
                        orient="records"
                    )
                    writers[ano_i].writerows(rows)

                    st = stats[ano_i]
                    st["parcelas"] += len(g)
                    if "contrato" in g.columns:
                        st["contratos"].update(
                            g["contrato"].dropna().unique().tolist()
                        )
                    if "subsidio" in g.columns:
                        st["subsidio"] += float(
                            pd.to_numeric(g["subsidio"], errors="coerce")
                            .fillna(0)
                            .sum()
                        )
                    if "impacto_fiscal" in g.columns:
                        st["impacto_fiscal"] += float(
                            pd.to_numeric(g["impacto_fiscal"], errors="coerce")
                            .fillna(0)
                            .sum()
                        )
                    elif "impacto" in g.columns:
                        st["impacto_fiscal"] += float(
                            pd.to_numeric(g["impacto"], errors="coerce")
                            .fillna(0)
                            .sum()
                        )

                total_linhas += len(chunk)
                if total_linhas % (chunksize * 4) < chunksize:
                    elapsed = max(time.time() - t0, 1e-6)
                    print(
                        f"  … {total_linhas:,} parcelas "
                        f"({total_linhas / elapsed:,.0f} linhas/s) | "
                        f"anos={len(stats)}"
                    )
    finally:
        for fh in handles.values():
            fh.close()

    # serializa sets
    out_stats: dict[int, dict] = {}
    for ano, st in stats.items():
        out_stats[ano] = {
            "ano_contrato": ano,
            "qtd_contratos": len(st["contratos"]),
            "qtd_parcelas": st["parcelas"],
            "subsidio": st["subsidio"],
            "impacto_fiscal": st["impacto_fiscal"],
            "arquivo": str(pasta_saida / f"fluxos_ano_contrato_{ano}.csv"),
        }
    return out_stats


def escrever_resumo_excel(stats: dict[int, dict], saida: Path) -> Path:
    saida = Path(saida)
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Por_Ano_Contrato"
    headers = [
        "ano_contrato",
        "qtd_contratos",
        "qtd_parcelas",
        "subsidio",
        "impacto_fiscal",
        "arquivo",
    ]
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for ano in sorted(stats):
        st = stats[ano]
        ws.append([st[h] for h in headers])

    # nota metodológica
    ws2 = wb.create_sheet("Nota")
    ws2["A1"] = (
        "Discriminativo por ANO DO CONTRATO. "
        "Todas as parcelas de um contrato (ex.: 180) entram na planilha "
        "do ano em que o contrato foi celebrado. "
        "O impacto fiscal de cada parcela continua capitalizado na data_fluxo "
        "(mesma metodologia ContAgil / SELIC)."
    )
    ws2.column_dimensions["A"].width = 100
    ws2.row_dimensions[1].height = 60

    for col, width in enumerate([14, 14, 14, 18, 18, 60], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    wb.save(saida)
    return saida


def main(argv: Optional[Iterable[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--pasta",
        type=Path,
        default=None,
        help="Pasta com fluxos_*.csv (default: ContAgil saida/ ou output/)",
    )
    p.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Pasta dos discriminativos (default: <pasta>/discriminativos_ano_contrato)",
    )
    p.add_argument("--ano-min", type=int, default=None)
    p.add_argument("--ano-max", type=int, default=None)
    p.add_argument("--chunksize", type=int, default=CHUNK_DEFAULT)
    args = p.parse_args(list(argv) if argv is not None else None)

    pasta = _resolver_pasta(args.pasta)
    arquivos = listar_csvs_fluxos(pasta)
    saida = (
        Path(args.saida)
        if args.saida is not None
        else pasta / "discriminativos_ano_contrato"
    )

    print(f"[{MARKER}] pasta={pasta}")
    print(f"[{MARKER}] arquivos={len(arquivos)} → {saida}")

    stats = repartir_streaming(
        arquivos,
        saida,
        chunksize=args.chunksize,
        ano_min=args.ano_min,
        ano_max=args.ano_max,
    )
    if not stats:
        print("[ERRO] Nenhuma parcela classificada por ano_contrato.")
        return 1

    resumo = escrever_resumo_excel(
        stats, saida / "RESUMO_POR_ANO_CONTRATO.xlsx"
    )
    print(f"[OK] Resumo: {resumo}")
    for ano in sorted(stats):
        st = stats[ano]
        print(
            f"  {ano}: {st['qtd_contratos']:,} contratos | "
            f"{st['qtd_parcelas']:,} parcelas | "
            f"impacto={st['impacto_fiscal']:,.2f}"
        )
    print(
        "\nNota: impacto fiscal por parcela = mesma fórmula de sempre "
        "(data_fluxo). Só a pasta/aba do discriminativo usa ano_contrato."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
