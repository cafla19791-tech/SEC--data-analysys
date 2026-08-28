#!/usr/bin/env python3
"""Resultado oficial do 2º turno presidencial de 2022 por UF e DF.

Fonte: TSE, eleição 545, arquivos de divulgação oficial
https://resultados.tse.jus.br/oficial/ele2022/545/
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

TSE_URL = (
    "https://resultados.tse.jus.br/oficial/ele2022/545/"
    "dados-simplificados/{uf}/{uf}-c0001-e000545-r.json"
)

UFS = [
    "ac", "al", "ap", "am", "ba", "ce", "df", "es", "go", "ma", "mt", "ms",
    "mg", "pa", "pb", "pr", "pe", "pi", "rj", "rn", "rs", "ro", "rr", "sc",
    "sp", "se", "to", "zz",
]

NOMES = {
    "AC": "Acre",
    "AL": "Alagoas",
    "AP": "Amapá",
    "AM": "Amazonas",
    "BA": "Bahia",
    "CE": "Ceará",
    "DF": "Distrito Federal",
    "ES": "Espírito Santo",
    "GO": "Goiás",
    "MA": "Maranhão",
    "MT": "Mato Grosso",
    "MS": "Mato Grosso do Sul",
    "MG": "Minas Gerais",
    "PA": "Pará",
    "PB": "Paraíba",
    "PR": "Paraná",
    "PE": "Pernambuco",
    "PI": "Piauí",
    "RJ": "Rio de Janeiro",
    "RN": "Rio Grande do Norte",
    "RS": "Rio Grande do Sul",
    "RO": "Rondônia",
    "RR": "Roraima",
    "SC": "Santa Catarina",
    "SP": "São Paulo",
    "SE": "Sergipe",
    "TO": "Tocantins",
    "ZZ": "Exterior",
}

REGIAO = {
    "AC": "Norte",
    "AL": "Nordeste",
    "AP": "Norte",
    "AM": "Norte",
    "BA": "Nordeste",
    "CE": "Nordeste",
    "DF": "Centro-Oeste",
    "ES": "Sudeste",
    "GO": "Centro-Oeste",
    "MA": "Nordeste",
    "MT": "Centro-Oeste",
    "MS": "Centro-Oeste",
    "MG": "Sudeste",
    "PA": "Norte",
    "PB": "Nordeste",
    "PR": "Sul",
    "PE": "Nordeste",
    "PI": "Nordeste",
    "RJ": "Sudeste",
    "RN": "Nordeste",
    "RS": "Sul",
    "RO": "Norte",
    "RR": "Norte",
    "SC": "Sul",
    "SP": "Sudeste",
    "SE": "Nordeste",
    "TO": "Norte",
    "ZZ": "Exterior",
}

COLUNAS = [
    "uf",
    "unidade",
    "regiao",
    "lula_votos",
    "lula_pct_validos",
    "bolsonaro_votos",
    "bolsonaro_pct_validos",
    "votos_validos",
    "votos_brancos",
    "votos_nulos",
    "comparecimento",
    "abstencao",
    "eleitores_aptos",
    "vencedor",
    "diferenca_votos",
    "diferenca_pp",
]


def pct_oficial(texto: str) -> float:
    return float(texto.replace(".", "").replace(",", "."))


def parse_uf(payload: dict) -> dict:
    cands = {c["n"]: c for c in payload["cand"]}
    lula = int(cands["13"]["vap"])
    bolso = int(cands["22"]["vap"])
    uf = payload["cdabr"].upper()
    return {
        "uf": uf,
        "unidade": NOMES[uf],
        "regiao": REGIAO[uf],
        "lula_votos": lula,
        "lula_pct_validos": pct_oficial(cands["13"]["pvap"]),
        "bolsonaro_votos": bolso,
        "bolsonaro_pct_validos": pct_oficial(cands["22"]["pvap"]),
        "votos_validos": int(payload["vv"]),
        "votos_brancos": int(payload["vb"]),
        "votos_nulos": int(payload["vn"]),
        "comparecimento": int(payload["c"]),
        "abstencao": int(payload["a"]),
        "eleitores_aptos": int(payload["e"]),
        "vencedor": "Lula" if lula > bolso else "Bolsonaro",
        "diferenca_votos": abs(lula - bolso),
        "diferenca_pp": round(
            abs(pct_oficial(cands["13"]["pvap"]) - pct_oficial(cands["22"]["pvap"])),
            2,
        ),
        "secoes_totalizadas_pct": pct_oficial(payload["pst"]),
    }


def baixar_tse(session: requests.Session | None = None) -> list[dict]:
    http = session or requests
    linhas = []
    for uf in UFS:
        url = TSE_URL.format(uf=uf)
        resp = http.get(url, timeout=60)
        resp.raise_for_status()
        linhas.append(parse_uf(resp.json()))
    return linhas


def montar_tabelas(linhas: list[dict]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    todos = pd.DataFrame(linhas)
    ufs = todos.loc[todos["uf"] != "ZZ", COLUNAS].copy()
    exterior = todos.loc[todos["uf"] == "ZZ", COLUNAS].copy()
    lula_br = int(todos["lula_votos"].sum())
    bolso_br = int(todos["bolsonaro_votos"].sum())
    validos_br = int(todos["votos_validos"].sum())
    brasil = pd.DataFrame(
        [
            {
                "abrangencia": "Brasil (TSE, 100% das seções)",
                "lula_votos": lula_br,
                "lula_pct_validos": round(100.0 * lula_br / validos_br, 2),
                "bolsonaro_votos": bolso_br,
                "bolsonaro_pct_validos": round(100.0 * bolso_br / validos_br, 2),
                "votos_validos": validos_br,
                "votos_brancos": int(todos["votos_brancos"].sum()),
                "votos_nulos": int(todos["votos_nulos"].sum()),
                "comparecimento": int(todos["comparecimento"].sum()),
                "abstencao": int(todos["abstencao"].sum()),
                "eleitores_aptos": int(todos["eleitores_aptos"].sum()),
                "vencedor": "Lula" if lula_br > bolso_br else "Bolsonaro",
                "diferenca_votos": abs(lula_br - bolso_br),
                "diferenca_pp": round(abs(100.0 * lula_br / validos_br - 100.0 * bolso_br / validos_br), 2),
            }
        ]
    )
    return ufs.reset_index(drop=True), exterior.reset_index(drop=True), brasil


def _estilo_cabecalho(ws, colunas: int):
    fill = PatternFill("solid", fgColor="1B4F72")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, colunas + 1):
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(colunas)}{ws.max_row}"


def _borda_fina(ws):
    thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin


def _pintar_vencedor(ws, col_vencedor: int, col_lula: int, col_bolso: int):
    fill_lula = PatternFill("solid", fgColor="F5B7B1")
    fill_bolso = PatternFill("solid", fgColor="AED6F1")
    for row in range(2, ws.max_row + 1):
        vencedor = ws.cell(row, col_vencedor).value
        if vencedor == "Lula":
            ws.cell(row, col_lula).fill = fill_lula
            ws.cell(row, col_vencedor).fill = fill_lula
        elif vencedor == "Bolsonaro":
            ws.cell(row, col_bolso).fill = fill_bolso
            ws.cell(row, col_vencedor).fill = fill_bolso


def _escrever_aba(ws, df: pd.DataFrame, titulos: list[str]):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    for col, titulo in enumerate(titulos, start=1):
        ws.cell(1, col, titulo)


def gravar_xlsx(
    ufs: pd.DataFrame,
    exterior: pd.DataFrame,
    brasil: pd.DataFrame,
    caminho: Path,
) -> Path:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    titulos_uf = [
        "UF",
        "Unidade federativa",
        "Região",
        "Lula (votos)",
        "Lula (% válidos)",
        "Bolsonaro (votos)",
        "Bolsonaro (% válidos)",
        "Votos válidos",
        "Votos brancos",
        "Votos nulos",
        "Comparecimento",
        "Abstenção",
        "Eleitores aptos",
        "Vencedor",
        "Diferença (votos)",
        "Diferença (p.p.)",
    ]
    ws = wb.active
    ws.title = "Por_UF"
    _escrever_aba(ws, ufs, titulos_uf)
    inteiros = {4, 6, 8, 9, 10, 11, 12, 13, 15}
    percentuais = {5, 7, 16}
    for row in range(2, ws.max_row + 1):
        for col in inteiros:
            ws.cell(row, col).number_format = "#,##0"
        for col in percentuais:
            ws.cell(row, col).number_format = "0.00"
        for col in range(1, 17):
            ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="left")
    _pintar_vencedor(ws, 14, 4, 6)
    _estilo_cabecalho(ws, 16)
    _borda_fina(ws)
    larguras = [6, 22, 14, 14, 16, 16, 18, 14, 14, 13, 16, 13, 16, 12, 16, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Brasil_e_exterior")
    bloco = pd.concat(
        [
            brasil.assign(uf="BR", unidade="Brasil", regiao="Brasil")[COLUNAS],
            exterior,
        ],
        ignore_index=True,
    )
    _escrever_aba(ws2, bloco, titulos_uf)
    for row in range(2, ws2.max_row + 1):
        for col in inteiros:
            ws2.cell(row, col).number_format = "#,##0"
        for col in percentuais:
            ws2.cell(row, col).number_format = "0.00"
        for col in range(1, 17):
            ws2.cell(row, col).alignment = Alignment(horizontal="center")
        ws2.cell(row, 2).alignment = Alignment(horizontal="left")
    _pintar_vencedor(ws2, 14, 4, 6)
    _estilo_cabecalho(ws2, 16)
    _borda_fina(ws2)
    for i, w in enumerate(larguras, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Fonte")
    ws3["A1"] = "Fonte e conceito"
    ws3["A1"].font = Font(bold=True, size=14)
    notas = [
        "Tribunal Superior Eleitoral (TSE).",
        "Eleição 545 — 2º turno das Eleições Gerais 2022, cargo de presidente.",
        "Data do pleito: 30/10/2022. Totalização nacional: 31/10/2022, 00h18, 100% das seções.",
        "Percentuais oficiais sobre votos válidos (excluídos brancos e nulos).",
        "Arquivos: https://resultados.tse.jus.br/oficial/ele2022/545/dados-simplificados/{uf}/{uf}-c0001-e000545-r.json",
        "Aba Por_UF: 26 estados + Distrito Federal.",
        "Aba Brasil_e_exterior: total nacional (soma das UFs + voto no exterior) e linha do exterior (ZZ).",
        "Brasil oficial: Lula 60.345.999 (50,90%); Bolsonaro 58.206.354 (49,10%); válidos 118.552.353.",
    ]
    for i, texto in enumerate(notas, start=3):
        ws3[f"A{i}"] = texto
    ws3.column_dimensions["A"].width = 140

    wb.save(caminho)
    return caminho


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Planilha TSE do 2º turno presidencial de 2022 por UF."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT / "output" / "eleicoes_2022_presidente_2t_por_uf.xlsx",
    )
    args = parser.parse_args(argv)
    linhas = baixar_tse()
    ufs, exterior, brasil = montar_tabelas(linhas)
    caminho = gravar_xlsx(ufs, exterior, brasil, args.output)
    print(f"[OK] {caminho}")
    print(ufs.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
