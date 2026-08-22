"""Evolução da balança comercial e das reservas internacionais (1995–2025).

Fonte: Banco Central do Brasil — SGS (BPM6 / liquidez internacional).

Séries:
  22707  Balança comercial (bens) — saldo mensal (US$ milhões) — fluxo
  22708  Exportação de bens — mensal (US$ milhões) — fluxo
  22709  Importação de bens — mensal (US$ milhões) — fluxo
  3546   Reservas internacionais — total, conceito de liquidez (US$ milhões) — estoque

Uso:
  python3 scripts/evolucao_balanca_reservas.py
  python3 scripts/evolucao_balanca_reservas.py --cache-dir data --output-dir output
"""

from __future__ import annotations

import argparse
import re
import time
from datetime import datetime
from html import unescape
from pathlib import Path

import matplotlib.image as mpimg
import matplotlib.pyplot as plt
import pandas as pd
import requests
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ESTILO_TABELA = (
    "border-collapse:collapse;border-spacing:0;"
    "border:1.5px solid #1a1a1a;width:100%;"
)
ESTILO_CELULA = (
    "border:1px solid #1a1a1a;padding:6px 8px;"
    "font-family:Calibri,Arial,sans-serif;font-size:13px;"
)
ESTILO_TH = ESTILO_CELULA + "background:#1f4e79;color:#fff;font-weight:700;"
BORDA_CONTINUA = Border(
    left=Side(style="thin", color="1A1A1A"),
    right=Side(style="thin", color="1A1A1A"),
    top=Side(style="thin", color="1A1A1A"),
    bottom=Side(style="thin", color="1A1A1A"),
)

BCB_SGS = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
BCB_SGS_SOAP = "https://www3.bcb.gov.br/wssgs/services/FachadaWSSGS"


def _parse_sgs_data(texto: str) -> datetime | None:
    texto = (texto or "").strip()
    for fmt in ("%d/%m/%Y", "%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    partes = texto.split("/")
    if len(partes) == 2:
        try:
            return datetime(int(partes[1]), int(partes[0]), 1)
        except ValueError:
            return None
    return None


def _sgs_para_mensal(df: pd.DataFrame) -> pd.DataFrame:
    """Estoque de fim de mês (última observação do mês, diária ou mensal)."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["mes", "valor"])
    out = df.dropna(subset=["data", "valor"]).sort_values("data")
    out["mes"] = out["data"].dt.to_period("M").dt.to_timestamp()
    return (
        out.drop_duplicates(subset=["mes"], keep="last")[["mes", "valor"]]
        .reset_index(drop=True)
    )


def parse_sgs_soap_xml(texto: str) -> pd.DataFrame:
    """Extrai DATA/VALOR do XML interno de getValoresSeriesXML."""
    interno = unescape(texto or "")
    itens = re.findall(
        r"<ITEM>\s*<DATA>([^<]+)</DATA>\s*<VALOR>([^<]*)</VALOR>",
        interno,
        flags=re.I,
    )
    if not itens:
        return pd.DataFrame(columns=["data", "valor"])
    linhas = []
    for data_txt, valor_txt in itens:
        dt = _parse_sgs_data(data_txt)
        if dt is None:
            continue
        try:
            valor = float(valor_txt.replace(",", "."))
        except ValueError:
            continue
        linhas.append({"data": pd.Timestamp(dt), "valor": valor})
    if not linhas:
        return pd.DataFrame(columns=["data", "valor"])
    return pd.DataFrame(linhas).drop_duplicates(subset=["data"]).sort_values("data")


def _baixar_sgs_soap_intervalo(cod: int, inicio: str, fim: str) -> pd.DataFrame:
    """Uma chamada SOAP. Levanta HTTPError em 4xx/5xx ou Fault."""
    corpo = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                  xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <soapenv:Body>
    <ns1:getValoresSeriesXML soapenv:encodingStyle="http://schemas.xmlsoap.org/soap/encoding/"
        xmlns:ns1="https://www3.bcb.gov.br/wssgs/services/FachadaWSSGS">
      <in0 xmlns:soapenc="http://schemas.xmlsoap.org/soap/encoding/" soapenc:arrayType="xsd:long[1]">
        <item xsi:type="xsd:long">{int(cod)}</item>
      </in0>
      <in1 xsi:type="xsd:string">{inicio}</in1>
      <in2 xsi:type="xsd:string">{fim}</in2>
    </ns1:getValoresSeriesXML>
  </soapenv:Body>
</soapenv:Envelope>"""
    resp = requests.post(
        BCB_SGS_SOAP,
        data=corpo.encode("utf-8"),
        headers={"Content-Type": "text/xml; charset=utf-8", "SOAPAction": ""},
        timeout=120,
    )
    resp.raise_for_status()
    if re.search(r"<soap(?:env)?:Fault\b", resp.text or "", flags=re.I):
        raise requests.HTTPError("SOAP fault", response=resp)
    return parse_sgs_soap_xml(resp.text)


def _baixar_sgs_soap(cod: int, inicio: str, fim: str) -> pd.DataFrame:
    """Fallback quando a API JSON do SGS é bloqueada (WAF).

    Intervalos longos (sobretudo séries diárias) devolvem HTTP 500; nesse
    caso a série é baixada em blocos anuais.
    """
    try:
        bruto = _baixar_sgs_soap_intervalo(cod, inicio, fim)
        return _sgs_para_mensal(bruto)
    except (requests.HTTPError, requests.Timeout, requests.ConnectionError):
        pass
    inicio_ts = pd.Timestamp(datetime.strptime(inicio, "%d/%m/%Y"))
    fim_ts = pd.Timestamp(datetime.strptime(fim, "%d/%m/%Y"))
    partes: list[pd.DataFrame] = []
    cursor = inicio_ts
    while cursor <= fim_ts:
        bloco_fim = min(cursor + pd.DateOffset(years=1) - pd.DateOffset(days=1), fim_ts)
        try:
            parte = _baixar_sgs_soap_intervalo(
                cod,
                cursor.strftime("%d/%m/%Y"),
                bloco_fim.strftime("%d/%m/%Y"),
            )
        except (requests.HTTPError, requests.Timeout, requests.ConnectionError):
            parte = pd.DataFrame(columns=["data", "valor"])
        if not parte.empty:
            partes.append(parte)
        cursor = bloco_fim + pd.DateOffset(days=1)
        time.sleep(0.08)
    if not partes:
        return pd.DataFrame(columns=["mes", "valor"])
    return _sgs_para_mensal(pd.concat(partes, ignore_index=True))

SERIES = {
    22707: "saldo_comercial",
    22708: "exportacoes",
    22709: "importacoes",
    3546: "reservas",
}

SERIES_FLUXO = {"saldo_comercial", "exportacoes", "importacoes"}
ANO_INICIO = 1995
ANO_FIM = 2025

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
DATA_DIR = ROOT / "data"


def _json_sgs_payload(resp: requests.Response):
    """Lista JSON da API, ou None se a resposta for WAF/erro."""
    if resp.status_code == 404:
        return []
    texto = (resp.text or "").lstrip()
    if not texto.startswith("[") and not texto.startswith("{"):
        return None
    try:
        dados = resp.json()
    except (requests.JSONDecodeError, ValueError):
        return None
    if isinstance(dados, dict):
        return None
    if isinstance(dados, list):
        return dados
    return None


def baixar_sgs(
    cod: int,
    inicio: str = "01/01/1995",
    fim: str = "31/12/2025",
) -> pd.DataFrame:
    """Baixa série SGS Bacen (json em blocos; SOAP se a API REST estiver bloqueada)."""
    url = BCB_SGS.format(cod=cod)
    inicio_ts = pd.Timestamp(datetime.strptime(inicio, "%d/%m/%Y"))
    fim_ts = pd.Timestamp(datetime.strptime(fim, "%d/%m/%Y"))
    partes: list[pd.DataFrame] = []
    usar_soap = False
    cursor = inicio_ts
    while cursor <= fim_ts and not usar_soap:
        bloco_fim = min(cursor + pd.DateOffset(years=9, months=11), fim_ts)
        params = {
            "formato": "json",
            "dataInicial": cursor.strftime("%d/%m/%Y"),
            "dataFinal": bloco_fim.strftime("%d/%m/%Y"),
        }
        dados = None
        for tentativa in range(2):
            resp = requests.get(url, params=params, timeout=120)
            dados = _json_sgs_payload(resp)
            if dados is not None:
                break
            if resp.status_code in {429, 502, 503, 504}:
                time.sleep(2 ** tentativa)
                continue
            usar_soap = True
            break
        if usar_soap:
            break
        if dados:
            df = pd.DataFrame(dados)
            df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
            partes.append(df.dropna(subset=["data", "valor"]))
        cursor = bloco_fim + pd.DateOffset(days=1)
    if usar_soap or not partes:
        soap = _baixar_sgs_soap(cod, inicio, fim)
        if not soap.empty:
            return soap
        if partes:
            return _sgs_para_mensal(pd.concat(partes, ignore_index=True))
        raise RuntimeError(f"Série SGS {cod} vazia")
    return _sgs_para_mensal(pd.concat(partes, ignore_index=True))


def carregar_series(
    cache_dir: Path | None = None,
    baixar: bool = True,
) -> dict[str, pd.DataFrame]:
    """Carrega as quatro séries (cache CSV ou Bacen)."""
    out: dict[str, pd.DataFrame] = {}
    for cod, nome in SERIES.items():
        cache = None if cache_dir is None else cache_dir / f"sgs_{cod}_{nome}.csv"
        if cache is not None and cache.exists():
            df = pd.read_csv(cache, parse_dates=["mes"])
            out[nome] = df
            continue
        if not baixar:
            raise FileNotFoundError(f"Cache ausente para {nome}: {cache}")
        print(f"Baixando SGS {cod} ({nome}) 01/01/{ANO_INICIO}..31/12/{ANO_FIM}...")
        df = baixar_sgs(cod, f"01/01/{ANO_INICIO}", f"31/12/{ANO_FIM}")
        if cache is not None:
            cache.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(cache, index=False)
        out[nome] = df
    return out


def _volume_importacoes(serie: pd.Series) -> pd.Series:
    """Importações no BPM6 podem vir com sinal negativo (débito)."""
    if serie.dropna().empty:
        return serie
    if float(serie.mean()) < 0:
        return -serie
    return serie


def agregar_anual(series: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Fluxos somados no ano; reservas = estoque de dezembro (ou último mês)."""
    frames: list[pd.DataFrame] = []
    for nome, df in series.items():
        work = df.copy()
        work["ano"] = pd.to_datetime(work["mes"]).dt.year
        work = work[(work["ano"] >= ANO_INICIO) & (work["ano"] <= ANO_FIM)]
        if nome in SERIES_FLUXO:
            if nome == "importacoes":
                work = work.copy()
                work["valor"] = _volume_importacoes(work["valor"])
            anual = work.groupby("ano", as_index=False)["valor"].sum()
        else:
            ultimo = work.sort_values("mes").groupby("ano", as_index=False).last()
            anual = ultimo[["ano", "valor"]]
        anual = anual.rename(columns={"valor": nome})
        frames.append(anual)
    out = frames[0]
    for extra in frames[1:]:
        out = out.merge(extra, on="ano", how="outer")
    out = out.sort_values("ano").reset_index(drop=True)
    if {"exportacoes", "importacoes"}.issubset(out.columns):
        out["saldo_reconstruido"] = out["exportacoes"] - out["importacoes"]
    out["var_reservas"] = out["reservas"].diff()
    return out


def _fmt_numero(valor: float, casas: int = 1) -> str:
    return f"{valor:.{casas}f}".replace(".", ",")


def _fmt_bi(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    return _fmt_numero(valor / 1000.0, casas)


def _fmt_bi_signed(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    corpo = _fmt_numero(abs(valor) / 1000.0, casas)
    if valor > 0:
        return f"+{corpo}"
    if valor < 0:
        return f"-{corpo}"
    return corpo


def fases_historicas(anual: pd.DataFrame) -> list[dict]:
    """Recortes usuais da política externa brasileira no período."""
    recortes = [
        (1995, 1998, "Plano Real / âncora cambial"),
        (1999, 2002, "Flutuação e ajuste pós-desvalorização"),
        (2003, 2007, "Boom de commodities e superávits"),
        (2008, 2013, "Crise global e acúmulo de reservas"),
        (2014, 2016, "Recessão e retração das importações"),
        (2017, 2019, "Recuperação moderada"),
        (2020, 2021, "Pandemia"),
        (2022, 2025, "Superávits recordes e nova platô de reservas"),
    ]
    linhas = []
    for ini, fim, rotulo in recortes:
        bloco = anual[(anual["ano"] >= ini) & (anual["ano"] <= fim)]
        if bloco.empty:
            continue
        linhas.append(
            {
                "periodo": f"{ini}–{fim}",
                "rotulo": rotulo,
                "saldo_medio": float(bloco["saldo_comercial"].mean()),
                "saldo_acumulado": float(bloco["saldo_comercial"].sum()),
                "reservas_inicio": float(bloco["reservas"].iloc[0]),
                "reservas_fim": float(bloco["reservas"].iloc[-1]),
                "var_reservas": float(bloco["reservas"].iloc[-1] - bloco["reservas"].iloc[0]),
            }
        )
    return linhas


def gerar_graficos(anual: pd.DataFrame, output_dir: Path) -> list[Path]:
    """Gera PNGs estáticos da evolução 1995–2025."""
    output_dir.mkdir(parents=True, exist_ok=True)
    anos = anual["ano"].tolist()
    saldo = anual["saldo_comercial"] / 1000.0
    reservas = anual["reservas"] / 1000.0
    exportacoes = anual["exportacoes"] / 1000.0
    importacoes = anual["importacoes"] / 1000.0
    caminhos: list[Path] = []

    fig, ax = plt.subplots(figsize=(12, 5.2))
    cores = ["#1b7f4a" if v >= 0 else "#b42318" for v in saldo]
    ax.bar(anos, saldo, color=cores, width=0.8)
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_title("Balança comercial (bens) — saldo anual, BPM6")
    ax.set_ylabel("US$ bilhões")
    ax.set_xlabel("Ano")
    ax.set_xticks(anos[::2])
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p1 = output_dir / "grafico_saldo_comercial_1995_2025.png"
    fig.savefig(p1, dpi=140)
    plt.close(fig)
    caminhos.append(p1)

    fig, ax = plt.subplots(figsize=(12, 5.2))
    ax.plot(anos, reservas, color="#0b4f8a", linewidth=2.2, marker="o", markersize=3.5)
    ax.fill_between(anos, reservas, color="#0b4f8a", alpha=0.12)
    ax.set_title("Reservas internacionais — estoque em dezembro (liquidez)")
    ax.set_ylabel("US$ bilhões")
    ax.set_xlabel("Ano")
    ax.set_xticks(anos[::2])
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p2 = output_dir / "grafico_reservas_1995_2025.png"
    fig.savefig(p2, dpi=140)
    plt.close(fig)
    caminhos.append(p2)

    fig, ax1 = plt.subplots(figsize=(12, 5.4))
    cores = ["#1b7f4a" if v >= 0 else "#b42318" for v in saldo]
    ax1.bar(anos, saldo, color=cores, width=0.75, alpha=0.85, label="Saldo comercial")
    ax1.axhline(0, color="#333", linewidth=0.8)
    ax1.set_ylabel("Saldo comercial (US$ bi)", color="#1b7f4a")
    ax2 = ax1.twinx()
    ax2.plot(anos, reservas, color="#0b4f8a", linewidth=2.2, marker="o", markersize=3.2, label="Reservas")
    ax2.set_ylabel("Reservas (US$ bi)", color="#0b4f8a")
    ax1.set_title("Saldo comercial anual e estoque de reservas — 1995 a 2025")
    ax1.set_xticks(anos[::2])
    ax1.grid(axis="y", linestyle=":", alpha=0.4)
    fig.tight_layout()
    p3 = output_dir / "grafico_saldo_e_reservas_1995_2025.png"
    fig.savefig(p3, dpi=140)
    plt.close(fig)
    caminhos.append(p3)

    fig, ax = plt.subplots(figsize=(12, 5.2))
    ax.plot(anos, exportacoes, color="#0b4f8a", linewidth=2, label="Exportações")
    ax.plot(anos, importacoes, color="#b54708", linewidth=2, label="Importações")
    ax.fill_between(
        anos,
        exportacoes,
        importacoes,
        where=exportacoes >= importacoes,
        color="#1b7f4a",
        alpha=0.15,
        interpolate=True,
    )
    ax.fill_between(
        anos,
        exportacoes,
        importacoes,
        where=exportacoes < importacoes,
        color="#b42318",
        alpha=0.15,
        interpolate=True,
    )
    ax.set_title("Exportações e importações de bens (BPM6)")
    ax.set_ylabel("US$ bilhões")
    ax.set_xlabel("Ano")
    ax.set_xticks(anos[::2])
    ax.legend(frameon=False)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p4 = output_dir / "grafico_exportacoes_importacoes_1995_2025.png"
    fig.savefig(p4, dpi=140)
    plt.close(fig)
    caminhos.append(p4)
    return caminhos


def cabecalhos_anual() -> list[str]:
    return [
        "Ano",
        "Exportações",
        "Importações",
        "Saldo comercial",
        "Reservas (dez.)",
        "Δ reservas",
    ]


def cabecalhos_fases() -> list[str]:
    return [
        "Período",
        "Contexto",
        "Saldo médio",
        "Saldo acumulado",
        "Reservas início",
        "Reservas fim",
        "Δ reservas",
    ]


def linhas_tabela_anual(anual: pd.DataFrame) -> list[list[str]]:
    linhas = []
    for row in anual.itertuples(index=False):
        linhas.append(
            [
                str(int(row.ano)),
                _fmt_bi(row.exportacoes),
                _fmt_bi(row.importacoes),
                _fmt_bi_signed(row.saldo_comercial),
                _fmt_bi(row.reservas),
                _fmt_bi_signed(row.var_reservas),
            ]
        )
    return linhas


def linhas_tabela_fases(fases: list[dict]) -> list[list[str]]:
    linhas = []
    for f in fases:
        linhas.append(
            [
                f["periodo"],
                f["rotulo"],
                _fmt_bi_signed(f["saldo_medio"]),
                _fmt_bi_signed(f["saldo_acumulado"]),
                _fmt_bi(f["reservas_inicio"]),
                _fmt_bi(f["reservas_fim"]),
                _fmt_bi_signed(f["var_reservas"]),
            ]
        )
    return linhas


def tabela_html(
    cabecalhos: list[str],
    linhas: list[list[str]],
    aligns: list[str] | None = None,
) -> str:
    """Tabela HTML com grade contínua (border-collapse + 1px solid em todas as células)."""
    aligns = aligns or (["center"] + ["right"] * (len(cabecalhos) - 1))
    if len(aligns) != len(cabecalhos):
        aligns = ["center"] * len(cabecalhos)
    ths = "".join(
        f'<th style="{ESTILO_TH}text-align:{aligns[i]};">{h}</th>'
        for i, h in enumerate(cabecalhos)
    )
    corpo = []
    for r, linha in enumerate(linhas):
        fundo = "#ffffff" if r % 2 == 0 else "#eef3f8"
        tds = "".join(
            f'<td style="{ESTILO_CELULA}background:{fundo};text-align:{aligns[i]};">{val}</td>'
            for i, val in enumerate(linha)
        )
        corpo.append(f"<tr>{tds}</tr>")
    return (
        f'<table style="{ESTILO_TABELA}">'
        f"<thead><tr>{ths}</tr></thead>"
        f"<tbody>{''.join(corpo)}</tbody>"
        "</table>"
    )


def _tabela_anual_html(anual: pd.DataFrame) -> str:
    return tabela_html(cabecalhos_anual(), linhas_tabela_anual(anual))


def _tabela_fases_html(fases: list[dict]) -> str:
    aligns = ["center", "left"] + ["right"] * 5
    return tabela_html(cabecalhos_fases(), linhas_tabela_fases(fases), aligns)


def desenhar_tabela_png(
    cabecalhos: list[str],
    linhas: list[list[str]],
    path: Path,
    titulo: str,
    larguras: list[float] | None = None,
) -> Path:
    """PNG com grade contínua (todas as arestas das células em traço sólido)."""
    n_lin = len(linhas) + 1
    n_col = len(cabecalhos)
    fig_w = max(11.0, 1.7 * n_col)
    fig_h = 0.36 * n_lin + 0.85
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_axis_off()
    ax.set_title(titulo, fontsize=12, pad=10, loc="left")
    tab = ax.table(
        cellText=linhas,
        colLabels=cabecalhos,
        loc="center",
        cellLoc="center",
        colWidths=larguras,
    )
    tab.auto_set_font_size(False)
    tab.set_fontsize(8)
    tab.scale(1, 1.25)
    for (r, _c), cell in tab.get_celld().items():
        cell.set_edgecolor("#1a1a1a")
        cell.set_linewidth(0.9)
        cell.visible_edges = "BTRL"
        if r == 0:
            cell.set_facecolor("#1f4e79")
            cell.get_text().set_color("white")
            cell.get_text().set_fontweight("bold")
        elif r % 2 == 0:
            cell.set_facecolor("#eef3f8")
        else:
            cell.set_facecolor("white")
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


A4_PAISAGEM = (11.69, 8.27)


def _pdf_txt(texto) -> str:
    """Escapa $ para o matplotlib não tratar R$ como mathtext."""
    return str(texto).replace("$", r"\$")


def contar_paginas_pdf(path: Path) -> int:
    """Conta objetos /Type /Page (exclui /Pages)."""
    bruto = Path(path).read_bytes()
    return len(re.findall(rb"/Type\s*/Page(?!s)\b", bruto))


def _figura_capa_pdf(titulo: str, notas: list[str]):
    fig = plt.figure(figsize=A4_PAISAGEM, facecolor="white")
    fig.text(0.07, 0.84, _pdf_txt(titulo), fontsize=18, fontweight="bold", color="#1f4e79", va="top")
    y = 0.72
    for nota in notas:
        fig.text(0.07, y, _pdf_txt(nota), fontsize=10, color="#222", va="top", wrap=True)
        y -= 0.055 + 0.012 * nota.count("\n")
    fig.text(
        0.07,
        0.08,
        "Fonte: Banco Central do Brasil — SGS. Tabelas com grade contínua.",
        fontsize=8,
        color="#555",
    )
    return fig


def _figura_tabela_pdf(
    cabecalhos: list[str],
    linhas: list[list[str]],
    titulo: str,
    larguras: list[float] | None = None,
):
    fig = plt.figure(figsize=A4_PAISAGEM, facecolor="white")
    ax = fig.add_axes([0.03, 0.04, 0.94, 0.86])
    ax.set_axis_off()
    fig.suptitle(_pdf_txt(titulo), fontsize=12, x=0.04, ha="left", y=0.96, color="#1f4e79")
    n_col = max(len(cabecalhos), 1)
    fonte = 8 if n_col <= 8 else 6.5
    tab = ax.table(
        cellText=[[_pdf_txt(c) for c in row] for row in linhas],
        colLabels=[_pdf_txt(h) for h in cabecalhos],
        loc="center",
        cellLoc="center",
        colWidths=larguras,
    )
    tab.auto_set_font_size(False)
    tab.set_fontsize(fonte)
    tab.scale(1, 1.18)
    for (r, _c), cell in tab.get_celld().items():
        cell.set_edgecolor("#1a1a1a")
        cell.set_linewidth(0.7)
        cell.visible_edges = "BTRL"
        if r == 0:
            cell.set_facecolor("#1f4e79")
            cell.get_text().set_color("white")
            cell.get_text().set_fontweight("bold")
        elif r % 2 == 0:
            cell.set_facecolor("#eef3f8")
        else:
            cell.set_facecolor("white")
    return fig


def _figura_imagem_pdf(imagem: Path, titulo: str | None = None):
    fig = plt.figure(figsize=A4_PAISAGEM, facecolor="white")
    if titulo:
        fig.suptitle(_pdf_txt(titulo), fontsize=11, x=0.04, ha="left", y=0.97, color="#1f4e79")
        ax = fig.add_axes([0.04, 0.05, 0.92, 0.86])
    else:
        ax = fig.add_axes([0.03, 0.03, 0.94, 0.94])
    ax.imshow(mpimg.imread(imagem))
    ax.set_axis_off()
    return fig


def exportar_pdf_relatorio(
    path: Path,
    titulo: str,
    notas: list[str],
    tabelas: list[tuple[str, list[str], list[list[str]], list[float] | None]] | None = None,
    imagens: list[Path] | None = None,
) -> Path:
    """PDF em A4 paisagem: capa, tabelas com grade contínua e gráficos."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with PdfPages(path) as pdf:
        fig = _figura_capa_pdf(titulo, notas)
        pdf.savefig(fig)
        plt.close(fig)
        for item in tabelas or []:
            tit, cabs, lins, larg = item
            fig = _figura_tabela_pdf(cabs, lins, tit, larg)
            pdf.savefig(fig)
            plt.close(fig)
        for imagem in imagens or []:
            img = Path(imagem)
            if not img.exists():
                continue
            fig = _figura_imagem_pdf(img, img.stem.replace("_", " "))
            pdf.savefig(fig)
            plt.close(fig)
    return path


def _escrever_aba_excel(ws, cabecalhos: list[str], linhas: list[list[str]]) -> None:
    preench_cab = PatternFill("solid", fgColor="1F4E79")
    preench_alt = PatternFill("solid", fgColor="EEF3F8")
    fonte_cab = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fonte_cel = Font(name="Calibri", size=11)
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(1, col, cab)
        cell.border = BORDA_CONTINUA
        cell.fill = preench_cab
        cell.font = fonte_cab
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, linha in enumerate(linhas, start=2):
        for col, val in enumerate(linha, start=1):
            cell = ws.cell(i, col, val)
            cell.border = BORDA_CONTINUA
            cell.font = fonte_cel
            cell.alignment = Alignment(
                horizontal="left" if col == 2 and ws.title == "Fases" else "center",
                vertical="center",
            )
            if i % 2 == 0:
                cell.fill = preench_alt
    for col in range(1, len(cabecalhos) + 1):
        letras = [str(ws.cell(r, col).value or "") for r in range(1, len(linhas) + 2)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(max(letras, key=len)) + 3, 12), 42)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridlines = True


def exportar_excel_grade(anual: pd.DataFrame, output_dir: Path) -> Path:
    """Planilha com borda contínua (thin) em todas as células."""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "balanca_reservas_tabelas_1995_2025.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Serie_anual"
    _escrever_aba_excel(ws1, cabecalhos_anual(), linhas_tabela_anual(anual))
    ws2 = wb.create_sheet("Fases")
    _escrever_aba_excel(ws2, cabecalhos_fases(), linhas_tabela_fases(fases_historicas(anual)))
    wb.save(path)
    return path


def gerar_tabelas_png(anual: pd.DataFrame, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    p1 = desenhar_tabela_png(
        cabecalhos_anual(),
        linhas_tabela_anual(anual),
        output_dir / "tabela_anual_1995_2025.png",
        "Balança comercial e reservas — série anual (US$ bilhões)",
        larguras=[0.10, 0.16, 0.16, 0.20, 0.20, 0.18],
    )
    p2 = desenhar_tabela_png(
        cabecalhos_fases(),
        linhas_tabela_fases(fases_historicas(anual)),
        output_dir / "tabela_fases_1995_2025.png",
        "Fases históricas — saldo médio e reservas (US$ bilhões)",
        larguras=[0.10, 0.28, 0.12, 0.14, 0.12, 0.12, 0.12],
    )
    return [p1, p2]


def _destaques(anual: pd.DataFrame) -> dict[str, float | int]:
    primeiro = anual.iloc[0]
    ultimo = anual.iloc[-1]
    max_saldo = anual.loc[anual["saldo_comercial"].idxmax()]
    min_saldo = anual.loc[anual["saldo_comercial"].idxmin()]
    max_res = anual.loc[anual["reservas"].idxmax()]
    min_res = anual.loc[anual["reservas"].idxmin()]
    deficits = anual[anual["saldo_comercial"] < 0]
    return {
        "ano_ini": int(primeiro.ano),
        "ano_fim": int(ultimo.ano),
        "saldo_ini": float(primeiro.saldo_comercial),
        "saldo_fim": float(ultimo.saldo_comercial),
        "reservas_ini": float(primeiro.reservas),
        "reservas_fim": float(ultimo.reservas),
        "mult_reservas": float(ultimo.reservas / primeiro.reservas),
        "saldo_acumulado": float(anual["saldo_comercial"].sum()),
        "n_deficit": int((anual["saldo_comercial"] < 0).sum()),
        "n_superavit": int((anual["saldo_comercial"] > 0).sum()),
        "anos_deficit": ", ".join(str(int(a)) for a in deficits["ano"]) if not deficits.empty else "nenhum",
        "max_saldo": float(max_saldo.saldo_comercial),
        "ano_max_saldo": int(max_saldo.ano),
        "min_saldo": float(min_saldo.saldo_comercial),
        "ano_min_saldo": int(min_saldo.ano),
        "max_reservas": float(max_res.reservas),
        "ano_max_reservas": int(max_res.ano),
        "min_reservas": float(min_res.reservas),
        "ano_min_reservas": int(min_res.ano),
        "x_ini": float(primeiro.exportacoes),
        "x_fim": float(ultimo.exportacoes),
        "m_ini": float(primeiro.importacoes),
        "m_fim": float(ultimo.importacoes),
    }


def gerar_relatorio(anual: pd.DataFrame, output_dir: Path) -> Path:
    """Escreve o relatório markdown com tabelas e leitura histórica."""
    d = _destaques(anual)
    fases = fases_historicas(anual)
    gerado = datetime.now().strftime("%Y-%m-%d")
    texto = f"""# Evolução da balança comercial e das reservas internacionais (1995–2025)

**Fonte:** Banco Central do Brasil, SGS. Balança comercial no conceito do
balanço de pagamentos (BPM6): séries 22707 (saldo), 22708 (exportações) e
22709 (importações), em US$ milhões; fluxos mensais somados no ano-calendário.
Reservas internacionais: série 3546 (conceito de liquidez internacional),
estoque de dezembro, em US$ milhões. Valores nas tabelas em **US$ bilhões**.
**Consulta:** {gerado}.
Tabelas com **grade contínua** (borda sólida em todas as células).

Os números do Banco Central **não coincidem** com a balança comercial da
Secex/MDIC. O BP registra transferência de propriedade entre residentes e
não residentes (inclusive operações fictas e ajustes de cobertura); a Secex
registra a movimentação física pelo Siscomex.

## Síntese

Entre {d['ano_ini']} e {d['ano_fim']} o Brasil saiu de um regime de **déficit
comercial com reservas baixas** para um de **superávits persistentes e
reservas de centenas de bilhões de dólares**.

- Reservas em dezembro de {d['ano_ini']}: **US$ {_fmt_bi(d['reservas_ini'])} bi**.
- Reservas em dezembro de {d['ano_fim']}: **US$ {_fmt_bi(d['reservas_fim'])} bi**
  ({_fmt_numero(d['mult_reservas'])} vezes o estoque de {d['ano_ini']}).
- Saldo comercial: **US$ {_fmt_bi_signed(d['saldo_ini'])} bi** em {d['ano_ini']}
  e **US$ {_fmt_bi_signed(d['saldo_fim'])} bi** em {d['ano_fim']}.
- Superávit comercial acumulado no período: **US$ {_fmt_bi_signed(d['saldo_acumulado'])} bi**.
- Anos de déficit: {d['n_deficit']} ({d['anos_deficit']}). Anos de superávit: {d['n_superavit']}.
- Maior superávit anual: **US$ {_fmt_bi_signed(d['max_saldo'])} bi** ({d['ano_max_saldo']}).
- Maior déficit anual: **US$ {_fmt_bi_signed(d['min_saldo'])} bi** ({d['ano_min_saldo']}).
- Pico de reservas: **US$ {_fmt_bi(d['max_reservas'])} bi** ({d['ano_max_reservas']}).
- Vale de reservas: **US$ {_fmt_bi(d['min_reservas'])} bi** ({d['ano_min_reservas']}).
- Exportações: US$ {_fmt_bi(d['x_ini'])} bi ({d['ano_ini']}) → US$ {_fmt_bi(d['x_fim'])} bi ({d['ano_fim']}).
- Importações: US$ {_fmt_bi(d['m_ini'])} bi ({d['ano_ini']}) → US$ {_fmt_bi(d['m_fim'])} bi ({d['ano_fim']}).

Reservas são **estoque**; o saldo comercial é **fluxo**. Superávit comercial
ajuda a gerar divisas, mas a variação das reservas também reflete conta
financeira, intervenção cambial, valuation de ativos e operações do Banco
Central. Por isso as duas séries sobem juntas no longo prazo, sem se mover
na mesma proporção ano a ano.

## Fases

{_tabela_fases_html(fases)}

### 1995–1998 — âncora cambial e vulnerabilidade externa

No primeiro governo do Real a âncora cambial barateou importações e o saldo
comercial ficou **negativo** em todos os anos (média de cerca de US$ 6,6 bi
de déficit). As reservas foram de US$ 51,8 bi (dez/1995) a um pico de
US$ 60,1 bi (1996) e recuaram para US$ 44,6 bi no fim de 1998, na esteira
das crises da Ásia e da Rússia, que antecederam a maxidesvalorização de
janeiro de 1999.

### 1999–2002 — flutuação e ajuste

A mudança para câmbio flutuante e o aperto externo inverteram o saldo: o
déficit encolheu em 1999–2000 e virou superávit a partir de 2001
(US$ 12,0 bi em 2002). As reservas, porém, tocaram o **mínimo da série**
em dezembro de 2000 (US$ 33,0 bi) e ainda estavam em US$ 37,8 bi no fim
de 2002, com risco-país elevado no ciclo eleitoral.

### 2003–2007 — boom de commodities

A alta dos preços de soja, minério e petróleo, mais a demanda chinesa,
ampliou as exportações (de US$ 72,9 bi em 2003 para US$ 160,5 bi em 2007).
O saldo médio do quinquênio foi de US$ 36,7 bi. O acúmulo de reservas
acelerou em 2006–2007: o estoque saltou de US$ 53,8 bi (dez/2005) para
US$ 180,3 bi (dez/2007), com ingresso de capitais além do superávit
comercial.

### 2008–2013 — crise e platô elevado

A crise de 2008 cortou o comércio (exportações de US$ 198,2 bi em 2008
para US$ 153,5 bi em 2009), mas o Brasil já tinha colchão externo. Na
retomada as reservas atravessaram US$ 350 bi em 2011 e chegaram a
US$ 373,1 bi em 2012. Em 2013 o saldo comercial quase zerou (US$ 0,4 bi)
e as reservas recuaram para US$ 358,8 bi.

### 2014–2016 — recessão e superávit por retração

2014 foi o **único déficit depois de 2000** (US$ 6,7 bi). Em seguida a
queda da atividade e do investimento reduziu importações mais que
exportações (de US$ 230,7 bi em 2014 para US$ 139,7 bi em 2016). O saldo
voltou a US$ 44,5 bi em 2016 — ajuste sobretudo via quantidade, com o
fim do superciclo de preços.

### 2017–2021 — recuperação, pandemia e novo impulso

O saldo ficou elevado (US$ 57,4 bi em 2017; US$ 29,6 bi em 2019). Em
2018 as reservas atingiram o **pico da série** (US$ 374,7 bi). A pandemia
cortou volumes em 2020, mas o superávit se manteve (US$ 35,7 bi) e
subiu a US$ 42,3 bi em 2021 com a recuperação dos preços. As reservas
permaneceram no intervalo de US$ 350–375 bi.

### 2022–2025 — superávits recordes

O saldo comercial atingiu o **máximo da série em 2023** (US$ 92,3 bi),
com agro, petróleo e minério; 2024 e 2025 seguiram altos (US$ 65,8 bi e
US$ 59,7 bi). As reservas oscilaram: queda a US$ 324,7 bi em 2022
(valuation e intervenção), recuperação a US$ 355,0 bi em 2023, novo
recuo a US$ 329,7 bi em 2024 e fechamento de 2025 em **US$ 358,2 bi**.
O colchão externo deixou de ser a restrição dominante da macroeconomia
brasileira — o platô de 2012–2025 substituiu o acúmulo acelerado de
2006–2011.

## Série anual (US$ bilhões)

{_tabela_anual_html(anual)}

## Arquivos gerados

- `balanca_reservas_anual_1995_2025.csv` — série anual
- `balanca_reservas_fases_1995_2025.csv` — recortes históricos
- `balanca_reservas_tabelas_1995_2025.xlsx` — mesmas tabelas com borda contínua em todas as células
- `tabela_anual_1995_2025.png` / `tabela_fases_1995_2025.png` — grade contínua
- `grafico_saldo_comercial_1995_2025.png`
- `grafico_reservas_1995_2025.png`
- `grafico_saldo_e_reservas_1995_2025.png`
- `grafico_exportacoes_importacoes_1995_2025.png`

API: `https://api.bcb.gov.br/dados/serie/bcdata.sgs.{{cod}}/dados`.
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "evolucao_balanca_reservas_1995_2025.md"
    path.write_text(texto, encoding="utf-8")
    return path


def exportar_tabelas(anual: pd.DataFrame, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_anual = output_dir / "balanca_reservas_anual_1995_2025.csv"
    anual_bi = anual.copy()
    for col in ["saldo_comercial", "exportacoes", "importacoes", "reservas", "var_reservas", "saldo_reconstruido"]:
        if col in anual_bi.columns:
            anual_bi[col] = anual_bi[col] / 1000.0
    anual_bi.to_csv(csv_anual, index=False, float_format="%.3f")
    fases = pd.DataFrame(fases_historicas(anual))
    csv_fases = output_dir / "balanca_reservas_fases_1995_2025.csv"
    for col in ["saldo_medio", "saldo_acumulado", "reservas_inicio", "reservas_fim", "var_reservas"]:
        fases[col] = fases[col] / 1000.0
    fases.to_csv(csv_fases, index=False, float_format="%.3f")
    xlsx = exportar_excel_grade(anual, output_dir)
    return [csv_anual, csv_fases, xlsx]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--sem-download", action="store_true")
    parser.add_argument("--sem-graficos", action="store_true")
    args = parser.parse_args(argv)

    series = carregar_series(cache_dir=args.cache_dir, baixar=not args.sem_download)
    anual = agregar_anual(series)
    caminhos = exportar_tabelas(anual, args.output_dir)
    relatorio = gerar_relatorio(anual, args.output_dir)
    caminhos.append(relatorio)
    if not args.sem_graficos:
        caminhos.extend(gerar_graficos(anual, args.output_dir))
        caminhos.extend(gerar_tabelas_png(anual, args.output_dir))
    print(f"Anos: {int(anual['ano'].min())}–{int(anual['ano'].max())} ({len(anual)} linhas)")
    print(
        "Reservas dez/{ini}: US$ {r0:.1f} bi → dez/{fim}: US$ {r1:.1f} bi".format(
            ini=int(anual.iloc[0].ano),
            fim=int(anual.iloc[-1].ano),
            r0=anual.iloc[0].reservas / 1000,
            r1=anual.iloc[-1].reservas / 1000,
        )
    )
    print(
        "Saldo {ini}: US$ {s0:+.1f} bi → {fim}: US$ {s1:+.1f} bi".format(
            ini=int(anual.iloc[0].ano),
            fim=int(anual.iloc[-1].ano),
            s0=anual.iloc[0].saldo_comercial / 1000,
            s1=anual.iloc[-1].saldo_comercial / 1000,
        )
    )
    for p in caminhos:
        print(f"  {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
