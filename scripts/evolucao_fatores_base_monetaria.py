"""Fatores condicionantes da base monetária (1995–hoje).

Fonte: Banco Central do Brasil, SGS, saldo em final de período
(milhares de R$). Tabelas em **R$ bilhões**.

A variação anual de cada fator é a diferença do estoque de dezembro
(no ano corrente, último mês publicado). Os estoques SGS são contas
de operações, não uma partição da base: a soma não reproduz o nível
nem, necessariamente, a variação da base (há quebras e contas fora
desta lista). O resíduo = Δbase − Σ Δfatores.

  1788   Base monetária restrita
  1810   Tesouro Nacional — conta única
  1809   Operações com títulos públicos federais (total)
  1811   Operações com o setor externo
  1815   Depósitos de instituições financeiras
  1818   Autoridade monetária — outras operações
 12484   Redesconto do Banco Central
 12487   Operações com derivativos (ajustes)
 28724   Linhas temporárias especiais de liquidez
 29004   Títulos — mercado primário
 29006   Títulos — mercado secundário

Uso:
  python3 scripts/evolucao_fatores_base_monetaria.py
  python3 scripts/evolucao_fatores_base_monetaria.py --sem-download
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook

from scripts.evolucao_balanca_reservas import (
    _escrever_aba_excel,
    _fmt_numero,
    baixar_sgs,
    desenhar_tabela_png,
    exportar_pdf_relatorio,
    tabela_html,
)

SERIES = {
    1788: "base",
    1810: "tesouro",
    1809: "titulos",
    1811: "externo",
    1815: "depositos_if",
    1818: "outras",
    12484: "redesconto",
    12487: "derivativos",
    28724: "linhas_temp",
    29004: "titulos_primario",
    29006: "titulos_secundario",
}

FATORES = (
    "tesouro",
    "titulos",
    "externo",
    "depositos_if",
    "outras",
    "redesconto",
    "derivativos",
    "linhas_temp",
)

ANO_INICIO = 1995
ANO_FIM = 2026
MILHARES_PARA_BI = 1_000_000.0

DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"


def _fmt_bi(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    return _fmt_numero(float(valor), casas)


def _fmt_bi_signed(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    n = float(valor)
    txt = _fmt_numero(abs(n), casas)
    if n > 0:
        return f"+{txt}"
    if n < 0:
        return f"−{txt}"
    return txt


def carregar_series(cache_dir: Path | None = None, baixar: bool = True) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    fim = datetime.now().strftime("%d/%m/%Y")
    for cod, nome in SERIES.items():
        cache = None if cache_dir is None else cache_dir / f"sgs_{cod}_{nome}.csv"
        if cache is not None and cache.exists():
            out[nome] = pd.read_csv(cache, parse_dates=["mes"])
            continue
        if not baixar:
            raise FileNotFoundError(f"Cache ausente para {nome}: {cache}")
        print(f"Baixando SGS {cod} ({nome}) 01/01/{ANO_INICIO}..{fim}...", flush=True)
        try:
            df = baixar_sgs(cod, f"01/01/{ANO_INICIO}", fim)
        except RuntimeError:
            df = pd.DataFrame(columns=["mes", "valor"])
        if cache is not None:
            cache.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(cache, index=False)
        out[nome] = df
    return out


def _estoque_ano(df: pd.DataFrame, ano: int) -> float | None:
    if df is None or df.empty:
        return None
    bloco = df[pd.to_datetime(df["mes"]).dt.year == ano]
    if bloco.empty:
        return None
    v = bloco.sort_values("mes").iloc[-1]["valor"]
    if pd.isna(v):
        return None
    return float(v) / MILHARES_PARA_BI


def mes_referencia(series: dict[str, pd.DataFrame], ano: int) -> str | None:
    df = series.get("base")
    if df is None or df.empty:
        return None
    bloco = df[pd.to_datetime(df["mes"]).dt.year == ano]
    if bloco.empty:
        return None
    return pd.to_datetime(bloco["mes"].max()).strftime("%b/%Y")


def agregar_anual(series: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Estoque de dezembro (ou último mês) e variação anual, em R$ bilhões."""
    linhas = []
    for ano in range(ANO_INICIO, ANO_FIM + 1):
        row: dict = {"ano": ano}
        for nome in SERIES.values():
            row[nome] = _estoque_ano(series[nome], ano)
        if row["base"] is None:
            continue
        linhas.append(row)
    out = pd.DataFrame(linhas)
    if out.empty:
        return out
    out["soma_fatores"] = out[list(FATORES)].sum(axis=1, min_count=1)
    out["residuo_estoque"] = out["base"] - out["soma_fatores"]
    for nome in ("base",) + FATORES:
        prev = out[nome].shift(1)
        out[f"d_{nome}"] = out[nome] - prev
        primeira = out[nome].notna() & prev.isna() & (out["ano"] > ANO_INICIO)
        out.loc[primeira, f"d_{nome}"] = out.loc[primeira, nome]
    out["d_soma"] = out[[f"d_{n}" for n in FATORES]].sum(axis=1, min_count=1)
    out["residuo_var"] = out["d_base"] - out["d_soma"]
    return out


def fases_historicas(anual: pd.DataFrame) -> list[dict]:
    recortes = [
        (1995, 1998, "Pós-Real e âncora cambial"),
        (1999, 2002, "Metas de inflação e flutuação"),
        (2003, 2008, "Acúmulo de reservas e crédito"),
        (2009, 2010, "Crise internacional"),
        (2011, 2016, "Desaceleração"),
        (2017, 2019, "Recomposição"),
        (2020, 2021, "Pandemia e linhas temporárias"),
        (2022, 2026, "Aperto e enxugamento"),
    ]
    linhas = []
    for ini, fim, rotulo in recortes:
        bloco = anual[(anual["ano"] >= ini) & (anual["ano"] <= fim)]
        if bloco.empty:
            continue
        linhas.append(
            {
                "periodo": f"{ini}–{fim}",
                "rotulo": rotulo,
                "base_ini": float(bloco["base"].dropna().iloc[0]),
                "base_fim": float(bloco["base"].dropna().iloc[-1]),
                "d_tesouro": float(bloco["d_tesouro"].sum(min_count=1) or 0.0),
                "d_titulos": float(bloco["d_titulos"].sum(min_count=1) or 0.0),
                "d_externo": float(bloco["d_externo"].sum(min_count=1) or 0.0),
                "d_base": float(bloco["base"].dropna().iloc[-1] - bloco["base"].dropna().iloc[0]),
            }
        )
    return linhas


def _rotulo_ano(ano: int, ultimo_ano: int, mes_ultimo: str | None) -> str:
    if ano != ultimo_ano or not mes_ultimo:
        return str(int(ano))
    prefixo = mes_ultimo.split("/")[0].strip().lower()[:3]
    if prefixo in {"dec", "dez"}:
        return str(int(ano))
    return f"{ano}*"


def cabecalhos_estoque() -> list[str]:
    return [
        "Ano",
        "Base",
        "Tesouro",
        "Títulos",
        "Externo",
        "Dep. IF",
        "Outras",
        "Redesc.",
        "Deriv.",
        "Linhas",
    ]


def cabecalhos_var() -> list[str]:
    return [
        "Ano",
        "Δ Base",
        "Δ Tesouro",
        "Δ Títulos",
        "Δ Externo",
        "Δ Dep. IF",
        "Δ Outras",
        "Δ Redesc.",
        "Δ Deriv.",
        "Δ Linhas",
        "Resíduo",
    ]


def cabecalhos_fases() -> list[str]:
    return [
        "Período",
        "Contexto",
        "Base início",
        "Base fim",
        "Δ Base",
        "Δ Tesouro",
        "Δ Títulos",
        "Δ Externo",
    ]


def linhas_estoque(anual: pd.DataFrame, mes_ultimo: str | None) -> list[list[str]]:
    ultimo = int(anual["ano"].max())
    linhas = []
    for rec in anual.to_dict("records"):
        linhas.append(
            [
                _rotulo_ano(int(rec["ano"]), ultimo, mes_ultimo),
                _fmt_bi(rec.get("base")),
                _fmt_bi(rec.get("tesouro")),
                _fmt_bi(rec.get("titulos")),
                _fmt_bi(rec.get("externo")),
                _fmt_bi(rec.get("depositos_if")),
                _fmt_bi(rec.get("outras")),
                _fmt_bi(rec.get("redesconto")),
                _fmt_bi(rec.get("derivativos")),
                _fmt_bi(rec.get("linhas_temp")),
            ]
        )
    return linhas


def linhas_var(anual: pd.DataFrame, mes_ultimo: str | None) -> list[list[str]]:
    ultimo = int(anual["ano"].max())
    linhas = []
    for rec in anual.to_dict("records"):
        linhas.append(
            [
                _rotulo_ano(int(rec["ano"]), ultimo, mes_ultimo),
                _fmt_bi_signed(rec.get("d_base")),
                _fmt_bi_signed(rec.get("d_tesouro")),
                _fmt_bi_signed(rec.get("d_titulos")),
                _fmt_bi_signed(rec.get("d_externo")),
                _fmt_bi_signed(rec.get("d_depositos_if")),
                _fmt_bi_signed(rec.get("d_outras")),
                _fmt_bi_signed(rec.get("d_redesconto")),
                _fmt_bi_signed(rec.get("d_derivativos")),
                _fmt_bi_signed(rec.get("d_linhas_temp")),
                _fmt_bi_signed(rec.get("residuo_var")),
            ]
        )
    return linhas


def linhas_fases(fases: list[dict]) -> list[list[str]]:
    return [
        [
            f["periodo"],
            f["rotulo"],
            _fmt_bi(f["base_ini"]),
            _fmt_bi(f["base_fim"]),
            _fmt_bi_signed(f["d_base"]),
            _fmt_bi_signed(f["d_tesouro"]),
            _fmt_bi_signed(f["d_titulos"]),
            _fmt_bi_signed(f["d_externo"]),
        ]
        for f in fases
    ]


def gerar_graficos(anual: pd.DataFrame, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    anos = anual["ano"]
    caminhos: list[Path] = []

    fig, ax = plt.subplots(figsize=(12, 5.4))
    ax.plot(anos, anual["base"], color="#111111", linewidth=2.2, label="Base")
    ax.plot(anos, anual["tesouro"], color="#0b4f8a", linewidth=1.8, label="Tesouro")
    ax.plot(anos, anual["titulos"], color="#b54708", linewidth=1.8, label="Títulos")
    ax.plot(anos, anual["externo"], color="#1b7f4a", linewidth=1.8, label="Externo")
    ax.axhline(0, color="#888", linewidth=0.7)
    ax.set_title("Base monetária e estoques dos fatores (fim de período)")
    ax.set_ylabel("R$ bilhões")
    ax.set_xticks(list(anos[::2]))
    ax.legend(frameon=False, ncol=4)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p1 = output_dir / "grafico_fatores_base_estoque_1995_2026.png"
    fig.savefig(p1, dpi=140)
    plt.close(fig)
    caminhos.append(p1)

    fig, ax = plt.subplots(figsize=(12, 5.6))
    cores = ["#0b4f8a", "#b54708", "#1b7f4a", "#6b21a8", "#9a3412", "#0f766e", "#a16207", "#334155"]
    bottom_pos = pd.Series(0.0, index=anual.index)
    bottom_neg = pd.Series(0.0, index=anual.index)
    for nome, cor in zip(FATORES, cores):
        s = anual[f"d_{nome}"].fillna(0.0)
        pos = s.clip(lower=0)
        neg = s.clip(upper=0)
        ax.bar(anos, pos, bottom=bottom_pos, color=cor, width=0.8, label=nome.replace("_", " "))
        ax.bar(anos, neg, bottom=bottom_neg, color=cor, width=0.8)
        bottom_pos = bottom_pos + pos
        bottom_neg = bottom_neg + neg
    ax.plot(anos, anual["d_base"], color="#111", linewidth=1.6, marker="o", markersize=3, label="Δ Base")
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_title("Contribuição anual dos fatores para a variação da base")
    ax.set_ylabel("R$ bilhões")
    ax.set_xticks(list(anos[::2]))
    ax.legend(frameon=False, ncol=3, fontsize=8)
    ax.grid(axis="y", linestyle=":", alpha=0.4)
    fig.tight_layout()
    p2 = output_dir / "grafico_fatores_base_variacao_1995_2026.png"
    fig.savefig(p2, dpi=140)
    plt.close(fig)
    caminhos.append(p2)
    return caminhos


def _destaques(anual: pd.DataFrame) -> dict:
    primeiro = anual.iloc[0]
    ultimo = anual.iloc[-1]
    return {
        "ano_ini": int(primeiro.ano),
        "ano_fim": int(ultimo.ano),
        "base_ini": float(primeiro.base),
        "base_fim": float(ultimo.base),
        "mult": float(ultimo.base / primeiro.base) if primeiro.base else float("nan"),
        "tesouro_fim": float(ultimo.tesouro) if pd.notna(ultimo.tesouro) else None,
        "titulos_fim": float(ultimo.titulos) if pd.notna(ultimo.titulos) else None,
        "externo_fim": float(ultimo.externo) if pd.notna(ultimo.externo) else None,
    }


def gerar_relatorio(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> Path:
    d = _destaques(anual)
    fases = fases_historicas(anual)
    gerado = datetime.now().strftime("%Y-%m-%d")
    nota = f" *{d['ano_fim']} = estoque de {mes_ultimo}." if mes_ultimo else ""
    html_est = tabela_html(cabecalhos_estoque(), linhas_estoque(anual, mes_ultimo))
    html_var = tabela_html(cabecalhos_var(), linhas_var(anual, mes_ultimo))
    html_fases = tabela_html(
        cabecalhos_fases(),
        linhas_fases(fases),
        ["center", "left"] + ["right"] * 6,
    )
    texto = f"""# Fatores condicionantes da base monetária (1995–2026)

**Fonte:** Banco Central do Brasil, SGS, saldo em final de período
(milhares de R$ → **R$ bilhões**). **Consulta:** {gerado}.{nota}

Os fatores oficiais medem as operações do Banco Central que expandem
ou contraem a base (papel-moeda em circulação + reservas bancárias).
Estoque positivo de um fator = posição acumulada expansionista; a
**variação anual** mostra o sentido da operação naquele ano. Os
estoques **não** somam a base (resíduo de nível cresce com o tempo).

- **Tesouro (1810):** conta única. Saque do Tesouro expande a base;
  recolhimento contrai.
- **Títulos (1809):** operações com títulos públicos (primário 29004 +
  secundário 29006). Venda/compromissada do Bacen enxuga.
- **Setor externo (1811):** compra de divisas expande.
- **Depósitos de IF (1815):** compulsório e demais depósitos.
- **Redesconto (12484), derivativos (12487), linhas temporárias (28724)
  e outras (1818).**

Tabelas com **grade contínua**. Sinal **+** = expansão da base.

## Síntese

Base restrita: R$ {_fmt_bi(d['base_ini'])} bi em dez/{d['ano_ini']} →
R$ {_fmt_bi(d['base_fim'])} bi em {d['ano_fim']}
({_fmt_numero(d['mult'])} vezes).

Estoque mais recente: Tesouro R$ {_fmt_bi(d['tesouro_fim'])} bi;
títulos R$ {_fmt_bi(d['titulos_fim'])} bi; externo
R$ {_fmt_bi(d['externo_fim'])} bi.

## Fases — variação acumulada dos fatores

{html_fases}

## Variação anual dos fatores (R$ bilhões)

{html_var}

## Estoque em fim de período (R$ bilhões)

{html_est}

## Arquivos

- `fatores_base_monetaria_anual_1995_2026.csv`
- `fatores_base_monetaria_tabelas_1995_2026.xlsx`
- `tabela_fatores_base_estoque_1995_2026.png` / `tabela_fatores_base_variacao_1995_2026.png`
- `grafico_fatores_base_estoque_1995_2026.png` / `grafico_fatores_base_variacao_1995_2026.png`
- `evolucao_fatores_base_monetaria_1995_2026.pdf`
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "evolucao_fatores_base_monetaria_1995_2026.md"
    path.write_text(texto, encoding="utf-8")
    return path


def exportar_tabelas(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_anual = output_dir / "fatores_base_monetaria_anual_1995_2026.csv"
    anual.to_csv(csv_anual, index=False, float_format="%.3f")
    fases = pd.DataFrame(fases_historicas(anual))
    csv_fases = output_dir / "fatores_base_monetaria_fases_1995_2026.csv"
    fases.to_csv(csv_fases, index=False, float_format="%.3f")
    xlsx = output_dir / "fatores_base_monetaria_tabelas_1995_2026.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Variacao_anual"
    _escrever_aba_excel(ws1, cabecalhos_var(), linhas_var(anual, mes_ultimo))
    ws2 = wb.create_sheet("Estoque")
    _escrever_aba_excel(ws2, cabecalhos_estoque(), linhas_estoque(anual, mes_ultimo))
    ws3 = wb.create_sheet("Fases")
    _escrever_aba_excel(ws3, cabecalhos_fases(), linhas_fases(fases_historicas(anual)))
    wb.save(xlsx)
    return [csv_anual, csv_fases, xlsx]


def gerar_tabelas_png(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> list[Path]:
    p1 = desenhar_tabela_png(
        cabecalhos_var(),
        linhas_var(anual, mes_ultimo),
        output_dir / "tabela_fatores_base_variacao_1995_2026.png",
        "Fatores da base monetária — variação anual (R$ bilhões)",
    )
    p2 = desenhar_tabela_png(
        cabecalhos_estoque(),
        linhas_estoque(anual, mes_ultimo),
        output_dir / "tabela_fatores_base_estoque_1995_2026.png",
        "Fatores da base monetária — estoque de fim de período (R$ bilhões)",
    )
    p3 = desenhar_tabela_png(
        cabecalhos_fases(),
        linhas_fases(fases_historicas(anual)),
        output_dir / "tabela_fatores_base_fases_1995_2026.png",
        "Fases — fatores da base monetária (R$ bilhões)",
        larguras=[0.10, 0.28, 0.10, 0.10, 0.10, 0.11, 0.11, 0.10],
    )
    return [p1, p2, p3]


def gerar_pdf(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> Path:
    return exportar_pdf_relatorio(
        output_dir / "evolucao_fatores_base_monetaria_1995_2026.pdf",
        "Fatores condicionantes da base monetária (1995–2026)",
        [
            f"SGS 1788 (base), 1810 Tesouro, 1809 títulos, 1811 externo, 1815 depósitos de IF, "
            f"1818 outras, 12484 redesconto, 12487 derivativos, 28724 linhas temporárias. "
            f"Valores em R$ bilhões. *último estoque: {mes_ultimo or 'n/d'}.",
            "Estoque positivo = posição expansionista. Os estoques não particionam a base; "
            "resíduo = Δbase − Σ Δfatores. Primeira publicação de uma série conta como variação a partir de zero.",
            "Sinal + = expansão da base. Fonte: Banco Central do Brasil.",
        ],
        tabelas=[
            (
                "Fases — variação acumulada (R$ bilhões)",
                cabecalhos_fases(),
                linhas_fases(fases_historicas(anual)),
                [0.10, 0.26, 0.10, 0.10, 0.11, 0.11, 0.11, 0.11],
            ),
            ("Variação anual dos fatores (R$ bilhões)", cabecalhos_var(), linhas_var(anual, mes_ultimo), None),
            ("Estoque em fim de período (R$ bilhões)", cabecalhos_estoque(), linhas_estoque(anual, mes_ultimo), None),
        ],
        imagens=[
            output_dir / "grafico_fatores_base_estoque_1995_2026.png",
            output_dir / "grafico_fatores_base_variacao_1995_2026.png",
        ],
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--sem-download", action="store_true")
    parser.add_argument("--sem-graficos", action="store_true")
    args = parser.parse_args(argv)

    series = carregar_series(cache_dir=args.cache_dir, baixar=not args.sem_download)
    anual = agregar_anual(series)
    ultimo = int(anual["ano"].max())
    mes_ultimo = mes_referencia(series, ultimo)
    caminhos = exportar_tabelas(anual, args.output_dir, mes_ultimo)
    caminhos.append(gerar_relatorio(anual, args.output_dir, mes_ultimo))
    if not args.sem_graficos:
        caminhos.extend(gerar_graficos(anual, args.output_dir))
        caminhos.extend(gerar_tabelas_png(anual, args.output_dir, mes_ultimo))
        caminhos.append(gerar_pdf(anual, args.output_dir, mes_ultimo))
    print(f"Anos: {int(anual['ano'].min())}–{int(anual['ano'].max())} ({len(anual)} linhas)")
    if mes_ultimo:
        print(f"Último estoque: {mes_ultimo}")
    print(anual[["ano", "base", "d_base", "d_tesouro", "d_titulos", "d_externo"]].to_string(index=False))
    for p in caminhos:
        print(f"  {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
