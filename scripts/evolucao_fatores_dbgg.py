"""Fatores que influenciam a DBGG (1995–hoje).

Fonte: Banco Central do Brasil, SGS.

A DBGG em % do PIB começa em dez/2001 (SGS 4537). O saldo em R$
(SGS 4502) existe desde jan/1998. De 1995 a 1997 mostramos só os
**drivers**: NFSP e primário do Governo Federal + Bacen, Selic e
câmbio. Ajustes patrimoniais/cambiais entram em 2001.

Séries (dezembro; no ano corrente, último mês):

  4502   DBGG — saldo (R$ milhões)
  4537   DBGG / PIB (%)
  4501   DLSP consolidada — saldo (R$ milhões)
  4513   DLSP consolidada / PIB (%)
  5793   Primário consolidado 12 meses / PIB (%)  [desde 2002]
  5783   Primário GF+Bacen 12 meses / PIB (%)     [desde 1991]
  5760   Juros nominais consolidados 12 meses / PIB (%)
  5727   NFSP consolidada 12 meses / PIB (%)
  5717   NFSP GF+Bacen 12 meses / PIB (%)         [desde 1991]
 10820   Ajuste de privatização — saldo (R$ milhões)
 10821   Ajuste patrimonial — saldo (R$ milhões)
 10822   Ajuste metodológico dívida externa — saldo (R$ milhões)
 10824   Ajuste metodológico dívida interna — saldo (R$ milhões)
  4189   Meta Selic (% a.a.)
     1   Taxa de câmbio — US$ comercial (média)
 28199   Dívida externa do governo geral (R$ milhões, desde 2013)

Uso:
  python3 scripts/evolucao_fatores_dbgg.py
  python3 scripts/evolucao_fatores_dbgg.py --sem-download
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook

from scripts.evolucao_balanca_reservas import (
    _escrever_aba_excel,
    _fmt_numero,
    baixar_sgs,
    desenhar_tabela_png,
    tabela_html,
)

SERIES = {
    4502: "dbgg_rs",
    4537: "dbgg_pib",
    4501: "dlsp_rs",
    4513: "dlsp_pib",
    5793: "primario_cons",
    5783: "primario_gfbc",
    5760: "juros_cons",
    5727: "nfsp_cons",
    5717: "nfsp_gfbc",
    10820: "ajuste_priv",
    10821: "ajuste_patrim",
    10822: "ajuste_camb_ext",
    10824: "ajuste_met_int",
    4189: "selic",
    1: "usd",
    28199: "ext_gg",
}

ANO_INICIO = 1995
ANO_FIM = 2026
MI_PARA_BI = 1_000.0

DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"


def _fmt(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    return _fmt_numero(float(valor), casas)


def _fmt_signed(valor: float | None, casas: int = 1) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    n = float(valor)
    txt = _fmt_numero(abs(n), casas)
    if n > 0:
        return f"+{txt}"
    if n < 0:
        return f"−{txt}"
    return txt


def carregar_series(cache_dir: Path | None = None, baixar: bool = True) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    fim = datetime.now().strftime("%d/%m/%Y")
    for cod, nome in SERIES.items():
        cache = None if cache_dir is None else cache_dir / f"sgs_{cod}_{nome}.csv"
        if cache is not None and cache.exists():
            out[nome] = pd.read_csv(cache, parse_dates=["mes"])
            continue
        if not baixar:
            raise FileNotFoundError(f"Cache ausente para {nome}: {cache}")
        print(f"Baixando SGS {cod} ({nome}) 01/01/{ANO_INICIO}..{fim}...", flush=True)
        try:
            df = baixar_sgs(cod, f"01/01/{ANO_INICIO}", fim)
        except RuntimeError:
            df = pd.DataFrame(columns=["mes", "valor"])
        if cache is not None:
            cache.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(cache, index=False)
        out[nome] = df
    return out


def _ultimo_ano(df: pd.DataFrame, ano: int, escala: float = 1.0) -> float | None:
    if df is None or df.empty:
        return None
    bloco = df[pd.to_datetime(df["mes"]).dt.year == ano]
    if bloco.empty:
        return None
    v = bloco.sort_values("mes").iloc[-1]["valor"]
    if pd.isna(v):
        return None
    return float(v) / escala


def mes_referencia(series: dict[str, pd.DataFrame], ano: int) -> str | None:
    meses = []
    for nome in ("dbgg_rs", "dbgg_pib", "nfsp_gfbc", "selic", "usd"):
        df = series.get(nome)
        if df is None or df.empty:
            continue
        bloco = df[pd.to_datetime(df["mes"]).dt.year == ano]
        if bloco.empty:
            continue
        meses.append(pd.to_datetime(bloco["mes"].max()))
    if not meses:
        return None
    return max(meses).strftime("%b/%Y")


def agregar_anual(series: dict[str, pd.DataFrame]) -> pd.DataFrame:
    linhas = []
    for ano in range(ANO_INICIO, ANO_FIM + 1):
        row = {
            "ano": ano,
            "dbgg_rs": _ultimo_ano(series["dbgg_rs"], ano, MI_PARA_BI),
            "dbgg_pib": _ultimo_ano(series["dbgg_pib"], ano),
            "dlsp_rs": _ultimo_ano(series["dlsp_rs"], ano, MI_PARA_BI),
            "dlsp_pib": _ultimo_ano(series["dlsp_pib"], ano),
            "primario_cons": _ultimo_ano(series["primario_cons"], ano),
            "primario_gfbc": _ultimo_ano(series["primario_gfbc"], ano),
            "juros_cons": _ultimo_ano(series["juros_cons"], ano),
            "nfsp_cons": _ultimo_ano(series["nfsp_cons"], ano),
            "nfsp_gfbc": _ultimo_ano(series["nfsp_gfbc"], ano),
            "ajuste_priv": _ultimo_ano(series["ajuste_priv"], ano, MI_PARA_BI),
            "ajuste_patrim": _ultimo_ano(series["ajuste_patrim"], ano, MI_PARA_BI),
            "ajuste_camb_ext": _ultimo_ano(series["ajuste_camb_ext"], ano, MI_PARA_BI),
            "ajuste_met_int": _ultimo_ano(series["ajuste_met_int"], ano, MI_PARA_BI),
            "selic": _ultimo_ano(series["selic"], ano),
            "usd": _ultimo_ano(series["usd"], ano),
            "ext_gg": _ultimo_ano(series["ext_gg"], ano, MI_PARA_BI),
        }
        if all(row[k] is None for k in row if k != "ano"):
            continue
        linhas.append(row)
    out = pd.DataFrame(linhas)
    if out.empty:
        return out
    out["primario"] = out["primario_cons"].combine_first(out["primario_gfbc"])
    out["nfsp"] = out["nfsp_cons"].combine_first(out["nfsp_gfbc"])
    out["juros"] = out["juros_cons"]
    mask = out["juros"].isna() & out["nfsp"].notna() & out["primario"].notna()
    out.loc[mask, "juros"] = out.loc[mask, "nfsp"] - out.loc[mask, "primario"]
    out["d_dbgg_pib"] = out["dbgg_pib"].diff()
    out["d_dbgg_rs"] = out["dbgg_rs"].diff()
    out["d_priv"] = out["ajuste_priv"].diff()
    out["d_patrim"] = out["ajuste_patrim"].diff()
    out["d_camb"] = out["ajuste_camb_ext"].diff()
    out["d_met_int"] = out["ajuste_met_int"].diff()
    out["pib_imp"] = out["dbgg_rs"] / (out["dbgg_pib"] / 100.0)
    out["g_pib"] = out["pib_imp"].pct_change() * 100.0
    # Efeito do crescimento do PIB sobre a razão: −d_{t−1} × g/(1+g)
    g = out["pib_imp"].pct_change()
    out["efeito_pib"] = -out["dbgg_pib"].shift(1) * g / (1.0 + g)
    out["share_ext"] = 100.0 * out["ext_gg"] / out["dbgg_rs"]
    return out


def fases_historicas(anual: pd.DataFrame) -> list[dict]:
    recortes = [
        (1995, 1998, "Pós-Real e âncora"),
        (1999, 2002, "Desvalorização e metas"),
        (2003, 2008, "Ajuste e queda da dívida"),
        (2009, 2010, "Crise internacional"),
        (2011, 2016, "Recessão e alta da DBGG"),
        (2017, 2019, "Teto e recomposição"),
        (2020, 2021, "Pandemia"),
        (2022, 2026, "Juros altos e recomposição"),
    ]
    linhas = []
    for ini, fim, rotulo in recortes:
        bloco = anual[(anual["ano"] >= ini) & (anual["ano"] <= fim)]
        if bloco.empty:
            continue
        dbgg = bloco["dbgg_pib"].dropna()
        linhas.append(
            {
                "periodo": f"{ini}–{fim}",
                "rotulo": rotulo,
                "dbgg_ini": float(dbgg.iloc[0]) if not dbgg.empty else None,
                "dbgg_fim": float(dbgg.iloc[-1]) if not dbgg.empty else None,
                "primario_med": float(bloco["primario"].mean()) if bloco["primario"].notna().any() else None,
                "juros_med": float(bloco["juros"].mean()) if bloco["juros"].notna().any() else None,
                "selic_fim": float(bloco["selic"].dropna().iloc[-1]) if bloco["selic"].notna().any() else None,
                "usd_fim": float(bloco["usd"].dropna().iloc[-1]) if bloco["usd"].notna().any() else None,
            }
        )
    return linhas


def _rotulo_ano(ano: int, ultimo_ano: int, mes_ultimo: str | None) -> str:
    if ano != ultimo_ano or not mes_ultimo:
        return str(int(ano))
    prefixo = mes_ultimo.split("/")[0].strip().lower()[:3]
    if prefixo in {"dec", "dez"}:
        return str(int(ano))
    return f"{ano}*"


def cabecalhos_estoque() -> list[str]:
    return [
        "Ano",
        "DBGG R$ bi",
        "DBGG % PIB",
        "Δ p.p.",
        "DLSP % PIB",
        "Externa R$ bi",
        "Ext/DBGG %",
    ]


def cabecalhos_fatores() -> list[str]:
    return [
        "Ano",
        "Primário",
        "Juros",
        "NFSP",
        "Efeito PIB",
        "Δ Privatiz.",
        "Δ Patrimonial",
        "Δ Câmbio ext.",
        "Δ Mét. int.",
        "Selic",
        "US$",
    ]


def cabecalhos_fases() -> list[str]:
    return [
        "Período",
        "Contexto",
        "DBGG início",
        "DBGG fim",
        "Primário méd.",
        "Juros méd.",
        "Selic fim",
        "US$ fim",
    ]


def linhas_estoque(anual: pd.DataFrame, mes_ultimo: str | None) -> list[list[str]]:
    ultimo = int(anual["ano"].max())
    return [
        [
            _rotulo_ano(int(r["ano"]), ultimo, mes_ultimo),
            _fmt(r.get("dbgg_rs")),
            _fmt(r.get("dbgg_pib")),
            _fmt_signed(r.get("d_dbgg_pib")),
            _fmt(r.get("dlsp_pib")),
            _fmt(r.get("ext_gg")),
            _fmt(r.get("share_ext")),
        ]
        for r in anual.to_dict("records")
    ]


def linhas_fatores(anual: pd.DataFrame, mes_ultimo: str | None) -> list[list[str]]:
    ultimo = int(anual["ano"].max())
    return [
        [
            _rotulo_ano(int(r["ano"]), ultimo, mes_ultimo),
            _fmt(r.get("primario")),
            _fmt(r.get("juros")),
            _fmt(r.get("nfsp")),
            _fmt_signed(r.get("efeito_pib")),
            _fmt_signed(r.get("d_priv")),
            _fmt_signed(r.get("d_patrim")),
            _fmt_signed(r.get("d_camb")),
            _fmt_signed(r.get("d_met_int")),
            _fmt(r.get("selic")),
            _fmt(r.get("usd"), 2),
        ]
        for r in anual.to_dict("records")
    ]


def linhas_fases(fases: list[dict]) -> list[list[str]]:
    return [
        [
            f["periodo"],
            f["rotulo"],
            _fmt(f["dbgg_ini"]),
            _fmt(f["dbgg_fim"]),
            _fmt(f["primario_med"]),
            _fmt(f["juros_med"]),
            _fmt(f["selic_fim"]),
            _fmt(f["usd_fim"], 2),
        ]
        for f in fases
    ]


def gerar_graficos(anual: pd.DataFrame, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    anos = anual["ano"]
    caminhos: list[Path] = []

    fig, ax = plt.subplots(figsize=(12, 5.4))
    ax.plot(anos, anual["dbgg_pib"], color="#111", linewidth=2.2, label="DBGG / PIB")
    ax.plot(anos, anual["dlsp_pib"], color="#0b4f8a", linewidth=2.0, label="DLSP / PIB")
    ax.set_title("DBGG e DLSP em % do PIB")
    ax.set_ylabel("% do PIB")
    ax.set_xticks(list(anos[::2]))
    ax.legend(frameon=False)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p1 = output_dir / "grafico_dbgg_dlsp_pib_1995_2026.png"
    fig.savefig(p1, dpi=140)
    plt.close(fig)
    caminhos.append(p1)

    fig, ax = plt.subplots(figsize=(12, 5.4))
    ax.bar(anos, anual["primario"].fillna(0), color="#1b7f4a", width=0.7, label="Primário 12m")
    ax.plot(anos, anual["juros"], color="#b54708", linewidth=2.0, label="Juros 12m")
    ax.plot(anos, anual["nfsp"], color="#111", linewidth=1.6, linestyle="--", label="NFSP 12m")
    ax.axhline(0, color="#333", linewidth=0.7)
    ax.set_title("Primário, juros e NFSP (% PIB, 12 meses)")
    ax.set_ylabel("% do PIB")
    ax.set_xticks(list(anos[::2]))
    ax.legend(frameon=False)
    ax.grid(axis="y", linestyle=":", alpha=0.4)
    fig.tight_layout()
    p2 = output_dir / "grafico_dbgg_primario_juros_1995_2026.png"
    fig.savefig(p2, dpi=140)
    plt.close(fig)
    caminhos.append(p2)

    fig, ax1 = plt.subplots(figsize=(12, 5.4))
    ax1.plot(anos, anual["selic"], color="#0b4f8a", linewidth=2.0, label="Selic")
    ax1.set_ylabel("Selic % a.a.", color="#0b4f8a")
    ax2 = ax1.twinx()
    ax2.plot(anos, anual["usd"], color="#b54708", linewidth=2.0, label="US$")
    ax2.set_ylabel("R$ / US$", color="#b54708")
    ax1.set_title("Selic e câmbio (fim de período)")
    ax1.set_xticks(list(anos[::2]))
    ax1.grid(axis="y", linestyle=":", alpha=0.4)
    fig.tight_layout()
    p3 = output_dir / "grafico_dbgg_selic_cambio_1995_2026.png"
    fig.savefig(p3, dpi=140)
    plt.close(fig)
    caminhos.append(p3)
    return caminhos


def _destaques(anual: pd.DataFrame) -> dict:
    com_dbgg = anual[anual["dbgg_pib"].notna()]
    primeiro = com_dbgg.iloc[0] if not com_dbgg.empty else anual.iloc[0]
    ultimo = anual.iloc[-1]
    return {
        "ano_ini": int(anual.iloc[0].ano),
        "ano_dbgg": int(primeiro.ano),
        "ano_fim": int(ultimo.ano),
        "dbgg_ini": float(primeiro.dbgg_pib) if pd.notna(primeiro.dbgg_pib) else None,
        "dbgg_fim": float(ultimo.dbgg_pib) if pd.notna(ultimo.dbgg_pib) else None,
        "dbgg_rs_fim": float(ultimo.dbgg_rs) if pd.notna(ultimo.dbgg_rs) else None,
        "dlsp_fim": float(ultimo.dlsp_pib) if pd.notna(ultimo.dlsp_pib) else None,
        "primario_fim": float(ultimo.primario) if pd.notna(ultimo.primario) else None,
        "juros_fim": float(ultimo.juros) if pd.notna(ultimo.juros) else None,
        "selic_fim": float(ultimo.selic) if pd.notna(ultimo.selic) else None,
        "usd_fim": float(ultimo.usd) if pd.notna(ultimo.usd) else None,
    }


def gerar_relatorio(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> Path:
    d = _destaques(anual)
    fases = fases_historicas(anual)
    gerado = datetime.now().strftime("%Y-%m-%d")
    nota = (
        f" *{d['ano_fim']}: último ponto disponível ({mes_ultimo});"
        " dívida/NFSP em geral atrasam Selic/câmbio."
        if mes_ultimo
        else ""
    )
    html_est = tabela_html(cabecalhos_estoque(), linhas_estoque(anual, mes_ultimo))
    html_fat = tabela_html(cabecalhos_fatores(), linhas_fatores(anual, mes_ultimo))
    html_fases = tabela_html(
        cabecalhos_fases(),
        linhas_fases(fases),
        ["center", "left"] + ["right"] * 6,
    )
    texto = f"""# Fatores que influenciam a DBGG (1995–2026)

**Fonte:** Banco Central do Brasil, SGS. **Consulta:** {gerado}.{nota}

A **DBGG** (dívida bruta do governo geral) soma os débitos da União,
estados e municípios junto ao setor privado, ao setor público
financeiro e ao resto do mundo, **incluindo as compromissadas do
Bacen**. Não inclui estatais. % do PIB (SGS 4537) desde dez/2001;
saldo em R$ (4502) desde 1998. A revisão de 2008 (SGS 13762) altera
o tratamento das compromissadas; a série longa 4537 continua
publicada. Selic e US$ no ano corrente usam o último dia útil
disponível (podem ser mais recentes que o estoque da dívida).

Fatores oficiais da dinâmica (tabelas especiais / SGS):

- **Resultado primário** (5793 consolidado; 5783 GF+Bacen em 1995–2001):
  déficit aumenta a dívida. Sinal do Bacen: NFSP positiva = déficit.
- **Juros nominais** (5760): principal motor da alta nominal.
- **NFSP** (5727 / 5717) = primário + juros, sem desvalorização cambial.
- **Crescimento do PIB** (efeito calculado): dilui a razão DBGG/PIB.
- **Câmbio / ajuste metodológico externo** (10822): varia o estoque
  em reais da dívida externa.
- **Ajuste patrimonial** (10821): reconhecimento de dívidas
  (“esqueletos”).
- **Privatizações** (10820): venda de ativos reduz a dívida líquida.
- **Ajuste metodológico interno** (10824): dívida interna indexada
  ao câmbio.
- **Selic (4189)** e **US$ (1)**: preços que alimentam juros e câmbio.

Δ privatização / patrimonial / câmbio estão em **R$ bilhões**
(variação do estoque acumulado). Primário, juros, NFSP e efeito PIB
estão em **% do PIB**. Tabelas com **grade contínua**.

## Síntese

DBGG: {_fmt(d['dbgg_ini'])}% do PIB em {d['ano_dbgg']} →
{_fmt(d['dbgg_fim'])}% em {d['ano_fim']}
(R$ {_fmt(d['dbgg_rs_fim'])} bi). DLSP: {_fmt(d['dlsp_fim'])}% do PIB.

Último fluxo 12 meses: primário {_fmt(d['primario_fim'])}% do PIB;
juros {_fmt(d['juros_fim'])}%. Selic {_fmt(d['selic_fim'])}%;
US$ {_fmt(d['usd_fim'], 2)}.

## Fases

{html_fases}

## Fatores anuais

Primário / juros / NFSP / efeito PIB em % do PIB. Ajustes em R$ bilhões.
Selic em % a.a.; US$ em R$/US$.

{html_fat}

## Estoque da DBGG

{html_est}

## Arquivos

- `fatores_dbgg_anual_1995_2026.csv`
- `fatores_dbgg_tabelas_1995_2026.xlsx`
- `tabela_fatores_dbgg_1995_2026.png` / `tabela_dbgg_estoque_1995_2026.png`
- `grafico_dbgg_dlsp_pib_1995_2026.png`
- `grafico_dbgg_primario_juros_1995_2026.png`
- `grafico_dbgg_selic_cambio_1995_2026.png`
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "evolucao_fatores_dbgg_1995_2026.md"
    path.write_text(texto, encoding="utf-8")
    return path


def exportar_tabelas(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_anual = output_dir / "fatores_dbgg_anual_1995_2026.csv"
    anual.to_csv(csv_anual, index=False, float_format="%.3f")
    csv_fases = output_dir / "fatores_dbgg_fases_1995_2026.csv"
    pd.DataFrame(fases_historicas(anual)).to_csv(csv_fases, index=False, float_format="%.3f")
    xlsx = output_dir / "fatores_dbgg_tabelas_1995_2026.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Fatores"
    _escrever_aba_excel(ws1, cabecalhos_fatores(), linhas_fatores(anual, mes_ultimo))
    ws2 = wb.create_sheet("Estoque")
    _escrever_aba_excel(ws2, cabecalhos_estoque(), linhas_estoque(anual, mes_ultimo))
    ws3 = wb.create_sheet("Fases")
    _escrever_aba_excel(ws3, cabecalhos_fases(), linhas_fases(fases_historicas(anual)))
    wb.save(xlsx)
    return [csv_anual, csv_fases, xlsx]


def gerar_tabelas_png(anual: pd.DataFrame, output_dir: Path, mes_ultimo: str | None) -> list[Path]:
    p1 = desenhar_tabela_png(
        cabecalhos_fatores(),
        linhas_fatores(anual, mes_ultimo),
        output_dir / "tabela_fatores_dbgg_1995_2026.png",
        "Fatores da DBGG (primário/juros/NFSP/efeito PIB em % PIB; ajustes em R$ bi)",
    )
    p2 = desenhar_tabela_png(
        cabecalhos_estoque(),
        linhas_estoque(anual, mes_ultimo),
        output_dir / "tabela_dbgg_estoque_1995_2026.png",
        "Estoque da DBGG e da DLSP",
    )
    p3 = desenhar_tabela_png(
        cabecalhos_fases(),
        linhas_fases(fases_historicas(anual)),
        output_dir / "tabela_fatores_dbgg_fases_1995_2026.png",
        "Fases — DBGG e drivers",
        larguras=[0.10, 0.24, 0.11, 0.10, 0.12, 0.11, 0.11, 0.11],
    )
    return [p1, p2, p3]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--sem-download", action="store_true")
    parser.add_argument("--sem-graficos", action="store_true")
    args = parser.parse_args(argv)

    series = carregar_series(cache_dir=args.cache_dir, baixar=not args.sem_download)
    anual = agregar_anual(series)
    ultimo = int(anual["ano"].max())
    mes_ultimo = mes_referencia(series, ultimo)
    caminhos = exportar_tabelas(anual, args.output_dir, mes_ultimo)
    caminhos.append(gerar_relatorio(anual, args.output_dir, mes_ultimo))
    if not args.sem_graficos:
        caminhos.extend(gerar_graficos(anual, args.output_dir))
        caminhos.extend(gerar_tabelas_png(anual, args.output_dir, mes_ultimo))
    print(f"Anos: {int(anual['ano'].min())}–{int(anual['ano'].max())} ({len(anual)} linhas)")
    if mes_ultimo:
        print(f"Último ponto: {mes_ultimo}")
    cols = [c for c in ("ano", "dbgg_pib", "primario", "juros", "nfsp", "selic") if c in anual.columns]
    print(anual[cols].to_string(index=False))
    for p in caminhos:
        print(f"  {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
