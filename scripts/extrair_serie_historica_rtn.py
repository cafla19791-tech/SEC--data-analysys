"""Extrai series anuais do boletim RTN serie_historica_*.xlsx (Tesouro Nacional).

Uso (ContAgil WinPython)::

    python extrair_serie_historica_rtn.py "serie_historica_mai26 (2).xlsx" --constantes-ipca --out rtn_ipca_mai26.csv

Abas:
  1.1   — valores correntes
  1.1-A — valores constantes IPCA (referencia no titulo, ex.: Mai/2026)
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

ROW_NEEDLES_1_1 = {
    "receita_total": "1. RECEITA TOTAL",
    "despesa_total": "4. DESPESA TOTAL",
    "resultado_primario": "5. RESULTADO PRIMÁRIO GOVERNO CENTRAL - ACIMA DA LINHA",
    "juros_nominais": "9. JUROS NOMINAIS",
    "resultado_nominal": "10. RESULTADO NOMINAL DO GOVERNO CENTRAL",
}


def _load_workbook(path: str | Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Pacote openpyxl necessario. Instale com: pip install openpyxl"
        ) from exc
    return load_workbook(path, read_only=True, data_only=True)


def _match_row(label: str, needle: str) -> bool:
    return label.strip().startswith(needle) or needle in label


def _year_of(header: Any) -> int | None:
    if isinstance(header, datetime):
        return header.year
    if isinstance(header, int) and 1900 < header < 2100:
        return header
    return None


def annual_sum_from_monthly_sheet(
    path: str | Path,
    sheet: str,
    needles: dict[str, str] | None = None,
    *,
    year_from: int = 2001,
    year_to: int = 2025,
    min_months: int = 12,
) -> dict[str, Any]:
    """Soma colunas mensais em totais anuais (R$ milhoes)."""
    needles = needles or ROW_NEEDLES_1_1
    wb = _load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet}' nao encontrada. Disponiveis: {wb.sheetnames}")
    ws = wb[sheet]

    headers: tuple[Any, ...] | None = None
    matched: dict[str, tuple[Any, ...]] = {}
    unit = None
    title = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 2:
            title = row[0]
        if i == 3 and isinstance(row[0], str):
            unit = row[0]
        if i == 5:
            headers = row
            continue
        lab = row[0]
        if not isinstance(lab, str):
            continue
        for key, needle in needles.items():
            if key not in matched and _match_row(lab, needle):
                matched[key] = row

    if headers is None:
        raise ValueError(f"Cabecalho (linha 5) nao encontrado na aba {sheet}")

    totals: dict[str, dict[int, float]] = {k: defaultdict(float) for k in needles}
    counts: dict[str, dict[int, int]] = {k: defaultdict(int) for k in needles}

    for col, h in enumerate(headers):
        if col == 0:
            continue
        y = _year_of(h)
        if y is None or y < year_from or y > year_to:
            continue
        for key, row in matched.items():
            val = row[col] if col < len(row) else None
            if val is None:
                continue
            try:
                totals[key][y] += float(val)
                counts[key][y] += 1
            except (TypeError, ValueError):
                continue

    series: dict[str, dict[int, float]] = {}
    for key in needles:
        series[key] = {
            y: totals[key][y]
            for y in sorted(totals[key])
            if counts[key][y] >= min_months
        }

    return {
        "path": str(path),
        "sheet": sheet,
        "title": title,
        "unit": unit,
        "labels": {k: matched[k][0] for k in matched},
        "missing": [k for k in needles if k not in matched],
        "series": series,
        "month_counts": {k: dict(counts[k]) for k in counts},
    }


def extract_annual_rtn(
    path: str | Path,
    *,
    year_from: int = 2001,
    year_to: int = 2025,
    constantes_ipca: bool = False,
) -> dict[str, Any]:
    """Tabela anual a partir de serie_historica_*.xlsx.

    constantes_ipca=True → aba 1.1-A (IPCA da referencia da planilha, ex. Mai/2026).
    """
    sheet = "1.1-A" if constantes_ipca else "1.1"
    core = annual_sum_from_monthly_sheet(
        path,
        sheet,
        ROW_NEEDLES_1_1,
        year_from=year_from,
        year_to=year_to,
    )
    years = sorted(set().union(*[set(v.keys()) for v in core["series"].values()]))
    rows = []
    for y in years:
        row: dict[str, Any] = {"ano": y}
        for key, series in core["series"].items():
            mi = series.get(y)
            row[f"{key}_R$mi"] = None if mi is None else round(mi, 2)
            row[f"{key}_R$bi"] = None if mi is None else round(mi / 1000.0, 2)
        rows.append(row)

    return {
        "path": str(path),
        "sheet": sheet,
        "constantes_ipca": constantes_ipca,
        "unit": core.get("unit"),
        "year_from": year_from,
        "year_to": year_to,
        "count": len(rows),
        "labels": core.get("labels"),
        "missing": core.get("missing"),
        "rows": rows,
        "provider": "extrair_serie_historica_rtn",
    }


def write_csv(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        out_path.write_text("", encoding="utf-8")
        return
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extrai series anuais do RTN serie_historica_*.xlsx"
    )
    p.add_argument("path", help="Caminho do XLSX")
    p.add_argument("--from", dest="year_from", type=int, default=2001)
    p.add_argument("--to", dest="year_to", type=int, default=2025)
    p.add_argument(
        "--constantes-ipca",
        action="store_true",
        help="Usa aba 1.1-A (IPCA) em vez de 1.1 (correntes)",
    )
    p.add_argument("--out", default="", help="CSV de saida")
    p.add_argument("--list-sheets", action="store_true", help="Lista abas e sai")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    path = Path(args.path)
    if not path.exists():
        print(json.dumps({"error": f"arquivo nao encontrado: {path}"}, ensure_ascii=False))
        return 1
    try:
        if args.list_sheets:
            wb = _load_workbook(path)
            print(
                json.dumps(
                    {
                        "path": str(path),
                        "sheets": list(wb.sheetnames),
                        "note": "Abas *-A = valores constantes IPCA",
                    },
                    indent=2,
                    ensure_ascii=False,
                )
            )
            return 0
        table = extract_annual_rtn(
            path,
            year_from=args.year_from,
            year_to=args.year_to,
            constantes_ipca=args.constantes_ipca,
        )
        if args.out:
            out_path = Path(args.out)
            write_csv(table["rows"], out_path)
            preview = {
                **{k: v for k, v in table.items() if k != "rows"},
                "out": str(out_path.resolve()),
                "rows_preview": table["rows"][:3],
                "rows_tail": table["rows"][-2:],
            }
            print(json.dumps(preview, indent=2, ensure_ascii=False, default=str))
        else:
            print(json.dumps(table, indent=2, ensure_ascii=False, default=str))
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
