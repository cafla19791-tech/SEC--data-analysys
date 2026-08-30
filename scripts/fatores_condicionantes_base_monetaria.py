#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fatores condicionantes da base monetária (fim de período), 2000–2026.

Baixa as séries SGS do Banco Central e monta a tabela oficial:

* os fatores (Tesouro, setor externo, títulos, depósitos, redesconto,
  derivativos, linhas de liquidez, outras) são o **fluxo do mês**,
  apesar do rótulo «saldo em final de período»;
* a base monetária restrita (papel-moeda + reservas) é o **estoque**
  no último dia do mês.

Identidade: Σ fatores do mês = base_t − base_{t−1} (diferença de R$ 1 mil
por arredondamento). No ano, a soma dos fluxos mensais reconstitui a
variação do estoque entre os fins de período.

Uso::

  python scripts/fatores_condicionantes_base_monetaria.py
  python fatores_condicionantes_base_monetaria.py --saida output/fatores.xlsx
"""

from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

MARKER = "fatores-base-monetaria-20260830"
BCB_SGS = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
ANO_INICIO = 2000
ANO_FIM = 2026
# Unidade SGS = milhares de R$. A Nota para a Imprensa publica em R$ milhões.
ESCALA_MILHOES = 1_000.0
TOLERANCIA_IDENTIDADE = 500.0  # R$ mil no nível SGS (R$ 0,5 milhão; arredondamento)
COL_VARIACAO = "Variação da base monetária"


@dataclass(frozen=True)
class Serie:
    codigo: int
    nome: str
    papel: str  # fator | detalhe | estoque
    inclui_soma: bool = False
    recuo: int = 0


SERIES: tuple[Serie, ...] = (
    Serie(1810, "Tesouro Nacional — Conta única", "fator", True),
    Serie(1811, "Operações com o setor externo", "fator", True),
    Serie(1809, "Operações com títulos públicos federais — Total", "fator", True),
    Serie(29004, "Títulos — mercado primário", "detalhe", False, 1),
    Serie(29006, "Títulos — mercado secundário", "detalhe", False, 1),
    Serie(1815, "Depósitos de instituições financeiras", "fator", True),
    Serie(12484, "Redesconto do Banco Central", "fator", True),
    Serie(12487, "Operações com derivativos — ajustes", "fator", True),
    Serie(
        28724,
        "Linhas temporárias especiais de liquidez",
        "fator",
        True,
    ),
    Serie(1818, "Autoridade Monetária — Outras operações", "fator", True),
    Serie(1788, "Base monetária restrita", "estoque"),
    Serie(1786, "Papel-moeda em circulação", "estoque", recuo=1),
    Serie(1787, "Reservas bancárias", "estoque", recuo=1),
)

FATORES_SOMA = tuple(s for s in SERIES if s.inclui_soma)
DETALHES = tuple(s for s in SERIES if s.papel == "detalhe")
ESTOQUES = tuple(s for s in SERIES if s.papel == "estoque")

# Série mensal começa depois da diária correspondente: usa o último
# dia do mês só onde o mensal ainda não existe (evita furo na identidade).
PREENCHIMENTO_DIARIO: dict[int, int] = {
    12487: 12485,  # derivativos-ajustes (diário desde mai/2002; mensal desde jun/2002)
    28724: 28723,  # LTEL (diário desde abr/2020; mensal desde mai/2020)
}


def _http_get(url: str, params: dict, tentativas: int = 5) -> list:
    ultimo: Exception | None = None
    headers = {"User-Agent": f"SEC-data-analysys/{MARKER}"}
    for i in range(tentativas):
        try:
            resp = requests.get(url, params=params, headers=headers, timeout=120)
            # Janela anterior ao início da série (ex.: 29004 só existe desde 2015).
            if resp.status_code == 404:
                return []
            if resp.status_code != 200 or not resp.text.strip():
                raise RuntimeError(f"HTTP {resp.status_code} vazio")
            if resp.text.lstrip().startswith("<") or resp.text.lstrip().startswith("<?xml"):
                raise RuntimeError("resposta XML/HTML em vez de JSON")
            dados = resp.json()
            if not isinstance(dados, list):
                raise RuntimeError(f"JSON inesperado: {type(dados)}")
            return dados
        except Exception as exc:  # noqa: BLE001 — retry de rede/SGS
            ultimo = exc
            time.sleep(1.5 * (2**i))
    raise RuntimeError(f"falha ao baixar {url} {params}: {ultimo}") from ultimo


def baixar_sgs(
    codigo: int,
    inicio: date,
    fim: date,
    *,
    cache: Path | None = None,
    usar_cache: bool = True,
) -> pd.DataFrame:
    """Série mensal SGS → colunas mes (Timestamp) e valor (R$ mil)."""
    if cache is not None and usar_cache and cache.exists() and cache.stat().st_size > 20:
        df = pd.read_csv(cache, parse_dates=["mes"])
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        return df.dropna(subset=["mes"]).sort_values("mes").reset_index(drop=True)

    url = BCB_SGS.format(cod=codigo)
    partes: list[pd.DataFrame] = []
    cursor = pd.Timestamp(inicio)
    fim_ts = pd.Timestamp(fim)
    while cursor <= fim_ts:
        bloco_fim = min(cursor + pd.DateOffset(years=8, months=11), fim_ts)
        dados = _http_get(
            url,
            {
                "formato": "json",
                "dataInicial": cursor.strftime("%d/%m/%Y"),
                "dataFinal": bloco_fim.strftime("%d/%m/%Y"),
            },
        )
        if dados:
            df = pd.DataFrame(dados)
            df["mes"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
            partes.append(df[["mes", "valor"]].dropna())
        cursor = bloco_fim + pd.DateOffset(days=1)
        time.sleep(0.25)

    if not partes:
        out = pd.DataFrame(columns=["mes", "valor"])
    else:
        out = (
            pd.concat(partes, ignore_index=True)
            .drop_duplicates("mes")
            .sort_values("mes")
            .reset_index(drop=True)
        )
    if cache is not None:
        cache.parent.mkdir(parents=True, exist_ok=True)
        out.to_csv(cache, index=False)
    return out


def carregar_painel(
    pasta_cache: Path,
    *,
    usar_cache: bool = True,
    series: Iterable[Serie] = SERIES,
    inicio: date = date(1999, 12, 1),
    fim: date = date(2026, 8, 31),
    arquivos: dict[int, Path] | None = None,
) -> pd.DataFrame:
    """Painel mensal: índice = mês, colunas = código SGS."""
    pasta_cache.mkdir(parents=True, exist_ok=True)
    cols: dict[int, pd.Series] = {}
    for s in series:
        if arquivos and s.codigo in arquivos:
            df = pd.read_csv(arquivos[s.codigo], parse_dates=["mes"])
            df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        else:
            print(f"[SGS] {s.codigo} {s.nome}")
            df = baixar_sgs(
                s.codigo,
                inicio,
                fim,
                cache=pasta_cache / f"sgs_{s.codigo}.csv",
                usar_cache=usar_cache,
            )
        if df.empty:
            continue
        ser = df.set_index("mes")["valor"]
        ser.index = pd.to_datetime(ser.index)
        cols[s.codigo] = ser
    if not cols:
        raise RuntimeError("nenhuma série SGS carregada")
    painel = pd.DataFrame(cols).sort_index()
    painel.index.name = "mes"
    return preencher_com_diario(
        painel,
        pasta_cache=pasta_cache,
        usar_cache=usar_cache,
        arquivos=arquivos,
        inicio=inicio,
        fim=fim,
    )


def ultimo_dia_do_mes(diario: pd.DataFrame) -> pd.Series:
    """Última observação de cada mês civil."""
    tmp = diario.dropna(subset=["mes", "valor"]).copy()
    tmp["ref"] = pd.to_datetime(tmp["mes"]).dt.to_period("M").dt.to_timestamp()
    return tmp.sort_values("mes").groupby("ref")["valor"].last()


def preencher_com_diario(
    painel: pd.DataFrame,
    *,
    pasta_cache: Path,
    usar_cache: bool,
    arquivos: dict[int, Path] | None,
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """Completa furos do mensal com o fechamento da série diária irmã."""
    out = painel.copy()
    for mensal, diario_cod in PREENCHIMENTO_DIARIO.items():
        if arquivos is not None:
            if diario_cod not in arquivos:
                continue
            bruto = pd.read_csv(arquivos[diario_cod], parse_dates=["mes"])
            bruto["valor"] = pd.to_numeric(bruto["valor"], errors="coerce")
        else:
            print(f"[SGS] {diario_cod} (diário → preenche {mensal})")
            bruto = baixar_sgs(
                diario_cod,
                inicio,
                fim,
                cache=pasta_cache / f"sgs_{diario_cod}.csv",
                usar_cache=usar_cache,
            )
        if bruto.empty:
            continue
        fechamento = ultimo_dia_do_mes(bruto)
        if mensal not in out.columns:
            out[mensal] = fechamento
        else:
            out[mensal] = out[mensal].combine_first(fechamento)
    return out.sort_index()


def ultimo_mes_ano(painel: pd.DataFrame, ano: int) -> pd.Timestamp | None:
    idx = painel.index[(painel.index.year == ano)]
    if len(idx) == 0:
        return None
    return idx.max()


def estoque_fim(painel: pd.DataFrame, codigo: int, ano: int) -> float:
    mes = ultimo_mes_ano(painel, ano)
    if mes is None or codigo not in painel.columns:
        return float("nan")
    return float(painel.loc[mes, codigo])


def fluxo_ano(painel: pd.DataFrame, codigo: int, ano: int) -> float:
    if codigo not in painel.columns:
        return 0.0
    fatia = painel.loc[painel.index.year == ano, codigo]
    if fatia.empty:
        return 0.0
    return float(fatia.fillna(0.0).sum())


def fluxo_dezembro(painel: pd.DataFrame, codigo: int, ano: int) -> float:
    if codigo not in painel.columns:
        return float("nan")
    alvo = [i for i in painel.index if i.year == ano and i.month == 12]
    if not alvo:
        # 2026 (e anos incompletos): último mês disponível
        mes = ultimo_mes_ano(painel, ano)
        if mes is None:
            return float("nan")
        return float(painel.loc[mes, codigo])
    return float(painel.loc[alvo[0], codigo])


def anos_com_dados(painel: pd.DataFrame, ano_ini: int = ANO_INICIO, ano_fim: int = ANO_FIM) -> list[int]:
    presentes = sorted({int(y) for y in painel.index.year.unique()})
    return [a for a in presentes if ano_ini <= a <= ano_fim]


def soma_fatores_mes(painel: pd.DataFrame, mes: pd.Timestamp) -> float:
    total = 0.0
    for s in FATORES_SOMA:
        if s.codigo not in painel.columns:
            continue
        v = painel.loc[mes, s.codigo]
        if pd.notna(v):
            total += float(v)
    return total


def identidade_mensal(painel: pd.DataFrame) -> pd.DataFrame:
    """Compara Σ fatores do mês com a variação do estoque da base."""
    if 1788 not in painel.columns:
        raise KeyError("série 1788 (base restrita) ausente")
    base = painel[1788]
    linhas = []
    meses = [m for m in painel.index if m >= pd.Timestamp(2000, 1, 1)]
    for mes in meses:
        prevs = base.index[base.index < mes]
        if len(prevs) == 0 or pd.isna(base.loc[mes]):
            continue
        ant = prevs.max()
        if pd.isna(base.loc[ant]):
            continue
        delta = float(base.loc[mes] - base.loc[ant])
        soma = soma_fatores_mes(painel, mes)
        linhas.append(
            {
                "mes": mes,
                "base": float(base.loc[mes]),
                "delta_base": delta,
                "soma_fatores": soma,
                "residuo": soma - delta,
            }
        )
    return pd.DataFrame(linhas)


def tabela_anual(painel: pd.DataFrame, anos: Iterable[int] | None = None) -> pd.DataFrame:
    """Uma coluna por ano; linhas = fatores (fluxo no ano) + estoques de fim."""
    if anos is None:
        anos = anos_com_dados(painel)
    anos = list(anos)
    rotulos: list[str] = []
    valores: list[list[float]] = []
    papeis: list[str] = []
    recuos: list[int] = []
    codigos: list[int | str] = []

    def _add(nome: str, vals: list[float], papel: str, recuo: int = 0, codigo: int | str = ""):
        rotulos.append(nome)
        valores.append(vals)
        papeis.append(papel)
        recuos.append(recuo)
        codigos.append(codigo)

    for s in SERIES:
        if s.papel == "estoque":
            continue
        if s.papel == "fator":
            vals = [fluxo_ano(painel, s.codigo, a) for a in anos]
        else:
            vals = [fluxo_ano(painel, s.codigo, a) for a in anos]
        _add(s.nome, vals, s.papel, s.recuo, s.codigo)

    soma = [
        sum(fluxo_ano(painel, s.codigo, a) for s in FATORES_SOMA) for a in anos
    ]
    _add("Variação da base (= soma dos fatores)", soma, "total", 0, "Σ")

    delta_estoque = []
    for a in anos:
        atual = estoque_fim(painel, 1788, a)
        anterior = estoque_fim(painel, 1788, a - 1)
        if pd.isna(atual) or pd.isna(anterior):
            delta_estoque.append(float("nan"))
        else:
            delta_estoque.append(atual - anterior)
    _add("Memória: Δ estoque da base (fim a fim)", delta_estoque, "memoria", 0, "Δ")
    residuo = [
        (s - d) if pd.notna(s) and pd.notna(d) else float("nan")
        for s, d in zip(soma, delta_estoque)
    ]
    _add("Memória: resíduo (soma − Δ estoque)", residuo, "memoria", 0, "ε")

    for s in ESTOQUES:
        vals = [estoque_fim(painel, s.codigo, a) for a in anos]
        _add(s.nome, vals, "estoque", s.recuo, s.codigo)

    out = pd.DataFrame(valores, columns=anos)
    out.insert(0, "Item", rotulos)
    out.insert(1, "SGS", [str(c) for c in codigos])
    out.insert(2, "Papel", papeis)
    out.insert(3, "Recuo", recuos)
    return out


def tabela_discriminativo(
    painel: pd.DataFrame, anos: Iterable[int] | None = None
) -> pd.DataFrame:
    """Um ano por linha; fatores em colunas; última coluna-chave = soma algébrica."""
    if anos is None:
        anos = anos_com_dados(painel)
    anos = list(anos)
    linhas = []
    for a in anos:
        row: dict[str, object] = {"Ano": a}
        soma = 0.0
        for s in FATORES_SOMA:
            v = fluxo_ano(painel, s.codigo, a)
            row[s.nome] = v
            soma += v
        row[COL_VARIACAO] = soma
        for s in DETALHES:
            row[s.nome] = fluxo_ano(painel, s.codigo, a)
        for s in ESTOQUES:
            row[s.nome] = estoque_fim(painel, s.codigo, a)
        linhas.append(row)
    return pd.DataFrame(linhas)


def tabela_dezembro(painel: pd.DataFrame, anos: Iterable[int] | None = None) -> pd.DataFrame:
    """Fluxo do mês de dezembro (ou último mês, se o ano ainda não fechou)."""
    if anos is None:
        anos = anos_com_dados(painel)
    anos = list(anos)
    rows = []
    for s in SERIES:
        if s.papel == "estoque":
            vals = [estoque_fim(painel, s.codigo, a) for a in anos]
        else:
            vals = [fluxo_dezembro(painel, s.codigo, a) for a in anos]
        rows.append(
            {
                "Item": s.nome,
                "SGS": s.codigo,
                "Papel": s.papel,
                "Recuo": s.recuo,
                **{a: v for a, v in zip(anos, vals)},
            }
        )
    return pd.DataFrame(rows)


def tabela_mensal(painel: pd.DataFrame) -> pd.DataFrame:
    out = painel.copy()
    out.index.name = "mes"
    out = out.reset_index()
    out.insert(1, "ano", out["mes"].dt.year)
    out.insert(2, "mes_num", out["mes"].dt.month)
    rename = {s.codigo: f"{s.codigo} {s.nome}" for s in SERIES if s.codigo in painel.columns}
    out = out.rename(columns=rename)
    if 1788 in painel.columns:
        base = painel[1788]
        delta = base.diff()
        out["Δ base restrita"] = out["mes"].map(delta)
        out["Σ fatores"] = out["mes"].map(lambda m: soma_fatores_mes(painel, m))
        out["Resíduo"] = out["Σ fatores"] - out["Δ base restrita"]
    return out


# --- Excel -----------------------------------------------------------------

AZUL = "1F4E79"
AZUL_CLARO = "D6E3F0"
DOURADO = "FFF2CC"
VERDE = "C6EFCE"
VERMELHO = "FCE4D6"
CINZA = "F2F2F2"
BRANCO = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _cab(ws: Worksheet, headers: list[str], linha: int) -> None:
    fill = PatternFill("solid", fgColor=AZUL)
    font = Font(color=BRANCO, bold=True, name="Calibri", size=10)
    for col, nome in enumerate(headers, start=1):
        cell = ws.cell(linha, col, nome)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _fmt_milhoes(valor: float) -> float:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return float("nan")
    return float(valor) / ESCALA_MILHOES


def _pintar_numero(cell, valor: float, papel: str) -> None:
    cell.number_format = '#,##0.0;(#,##0.0);"—"'
    if pd.isna(valor):
        cell.value = None
        return
    cell.value = _fmt_milhoes(valor)
    if papel in {"fator", "detalhe", "total"} and abs(valor) > 0.5:
        if valor > 0:
            cell.font = Font(name="Calibri", size=9, color="006100")
        else:
            cell.font = Font(name="Calibri", size=9, color="9C0006")


def _aba_metodologia(
    wb: Workbook,
    gerado_em: datetime,
    ultimo: pd.Timestamp,
    n_meses: int,
) -> None:
    ws = wb.active
    ws.title = "Metodologia"
    ws["A1"] = "Fatores condicionantes da base monetária — saldo em final de período"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=AZUL)
    ws.merge_cells("A1:B1")

    blocos = [
        ("Gerado em", gerado_em.strftime("%d/%m/%Y %H:%M")),
        ("Marker", MARKER),
        ("Fonte", "Banco Central do Brasil — SGS (Departamento de Estatísticas)"),
        ("Unidade publicada", "R$ milhões (séries SGS em R$ mil ÷ 1.000)"),
        ("Cobertura", f"{ANO_INICIO}–{ANO_FIM}; último mês disponível: {ultimo.strftime('%m/%Y')}"),
        ("Meses na amostra", str(n_meses)),
        (
            "O que é um fator condicionante",
            "Relação entre o Banco Central e o resto da economia (Tesouro, "
            "setor externo, sistema bancário, títulos) que injeta ou recolhe "
            "moeda nacional. A soma dos fatores no período é a variação da "
            "base monetária restrita.",
        ),
        (
            "Saldo em final de período",
            "Nas séries dos fatores, o ponto mensal do SGS é o fluxo daquele "
            "mês (contribuição para a variação da base), não o estoque da "
            "conta. Na base monetária (1788/1786/1787) o ponto é o estoque "
            "no último dia do mês. A aba Anual soma os fluxos de janeiro a "
            "dezembro (ou até o último mês, em 2026) e mostra o estoque de "
            "fim de ano.",
        ),
        (
            "Identidade",
            "Σ fatores_t = Base_t − Base_{t−1}. Primário (29004) + secundário "
            "(29006) = total de títulos (1809), a partir de 2015. Primário e "
            "secundário não entram de novo na soma. "
            "O mensal de derivativos (12487) e o das linhas LTEL (28724) "
            "nascem um mês depois das séries diárias 12485 e 28723; o "
            "fechamento diário preenche esse primeiro mês (mai/2002 e abr/2020).",
        ),
        (
            "Sinal",
            "Valor positivo = expansão da base; negativo = contração. "
            "Ex.: recolhimento para a Conta Única ou venda líquida de títulos "
            "no open market entram com sinal negativo.",
        ),
        (
            "Abas",
            "Discriminativo — um ano por linha, um fator por coluna, "
            f"e a coluna «{COL_VARIACAO}» = soma algébrica dos oito fatores "
            "(fórmula Excel SOMA, sem primário/secundário). "
            "Anual — a mesma informação transposta (fatores nas linhas). "
            "Dezembro — fluxo só do mês de dezembro. "
            "Mensal — painel completo. Identidade — resíduo mês a mês.",
        ),
        (
            "Séries SGS",
            " ; ".join(f"{s.codigo} {s.nome}" for s in SERIES),
        ),
        (
            "2026",
            "Ano incompleto: fatores somados até o último mês publicado; "
            "estoque da base nesse mesmo mês. Coluna marcada com *. "
            "Os meses mais recentes ainda podem ser revisados pelo Bacen "
            "(resíduo de cerca de R$ 100 milhões em 2026).",
        ),
        (
            "Resíduos conhecidos",
            "A identidade anual fecha (resíduo < R$ 1 milhão) em 24 dos 27 anos. "
            "Exceções: 2000 (R$ 31 milhões, meses de fev e jun) e 2026 "
            "(revisão das estatísticas mais recentes). 2002 e 2020 fecham "
            "após o preenchimento do primeiro mês das séries diárias.",
        ),
    ]
    _cab(ws, ["Campo", "Descrição"], 3)
    for i, (campo, desc) in enumerate(blocos, start=4):
        ws.cell(i, 1, campo).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(i, 2, desc).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 52 if len(desc) > 90 else 22
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 112
    ws.freeze_panes = "A4"


def _escrever_matriz(
    ws: Worksheet,
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str,
    anos: list[int],
    ultimo: pd.Timestamp,
) -> None:
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(anos))
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + len(anos))

    headers = ["Item", "SGS"] + [
        f"{a}*" if a == ultimo.year and ultimo.month < 12 else str(a) for a in anos
    ]
    _cab(ws, headers, 4)

    # df tem Item, SGS, Papel, Recuo, anos...
    vis = df.drop(columns=["Papel", "Recuo"])
    linha = 5
    for row in vis.itertuples(index=False):
        item, sgs, *nums = row
        papel = df.loc[df["Item"] == item, "Papel"].iloc[0]
        recuo = int(df.loc[df["Item"] == item, "Recuo"].iloc[0])
        nome = ("    " * recuo) + item
        c0 = ws.cell(linha, 1, nome)
        c1 = ws.cell(linha, 2, sgs)
        c0.border = THIN
        c1.border = THIN
        c0.alignment = Alignment(vertical="center")
        c1.alignment = Alignment(horizontal="center")
        bold = papel in {"total", "estoque"} and recuo == 0
        c0.font = Font(name="Calibri", size=9, bold=bold)
        c1.font = Font(name="Calibri", size=8, color="666666")
        fill = None
        if papel == "total":
            fill = PatternFill("solid", fgColor=DOURADO)
        elif papel == "estoque" and recuo == 0:
            fill = PatternFill("solid", fgColor=AZUL_CLARO)
        elif papel == "memoria":
            fill = PatternFill("solid", fgColor=CINZA)
        elif papel == "detalhe":
            fill = PatternFill("solid", fgColor="EEF2F7")
        if fill:
            c0.fill = fill
            c1.fill = fill
        for col, valor in enumerate(nums, start=3):
            cell = ws.cell(linha, col)
            cell.border = THIN
            cell.fill = fill or PatternFill("solid", fgColor=BRANCO)
            _pintar_numero(cell, float(valor) if pd.notna(valor) else float("nan"), papel)
            if fill and papel in {"total", "estoque", "memoria"}:
                cell.font = Font(name="Calibri", size=9, bold=bold)
        linha += 1

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 10
    for i in range(3, 3 + len(anos)):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(2 + len(anos))}{linha - 1}"
    ws.row_dimensions[4].height = 28
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    nota = (
        f"Valores em R$ milhões. Verde = expansão da base; vermelho = contração. "
        f"Coluna {ultimo.year}* = até {ultimo.strftime('%m/%Y')}."
    )
    ws.cell(linha + 1, 1, nota).font = Font(name="Calibri", size=8, italic=True, color="666666")


def _aba_mensal(wb: Workbook, mensal: pd.DataFrame) -> None:
    ws = wb.create_sheet("Mensal")
    ws["A1"] = "Painel mensal — fatores (fluxo) e base (estoque)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    vis = mensal.copy()
    vis["mes"] = vis["mes"].dt.strftime("%Y-%m")
    for col in vis.columns:
        if col in {"mes", "ano", "mes_num"}:
            continue
        vis[col] = vis[col].map(lambda v: _fmt_milhoes(v) if pd.notna(v) else None)
    headers = list(vis.columns)
    _cab(ws, headers, 3)
    for i, row in enumerate(dataframe_to_rows(vis, index=False, header=False), start=4):
        for col, valor in enumerate(row, start=1):
            cell = ws.cell(i, col, valor)
            cell.border = THIN
            cell.font = Font(name="Calibri", size=8)
            if headers[col - 1] not in {"mes", "ano", "mes_num"} and isinstance(valor, float):
                cell.number_format = '#,##0.0;(#,##0.0);"—"'
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=CINZA)
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(vis)}"
    for i, nome in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i > 3 else (12 if i > 1 else 12)
    ws.column_dimensions["A"].width = 12


def _aba_identidade(wb: Workbook, ident: pd.DataFrame) -> None:
    ws = wb.create_sheet("Identidade")
    ws["A1"] = "Identidade mensal — Σ fatores = Δ base restrita"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    n_ok = int((ident["residuo"].abs() <= TOLERANCIA_IDENTIDADE).sum()) if not ident.empty else 0
    ws["A2"] = (
        f"{n_ok} de {len(ident)} meses com |resíduo| ≤ R$ {TOLERANCIA_IDENTIDADE:.0f} mil "
        f"(arredondamento do SGS). Valores da tabela em R$ milhões."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    headers = ["Mês", "Base (estoque)", "Δ base", "Σ fatores", "Resíduo"]
    _cab(ws, headers, 4)
    for i, r in enumerate(ident.itertuples(index=False), start=5):
        vals = [
            r.mes.strftime("%Y-%m"),
            _fmt_milhoes(r.base),
            _fmt_milhoes(r.delta_base),
            _fmt_milhoes(r.soma_fatores),
            _fmt_milhoes(r.residuo),
        ]
        ok = abs(r.residuo) <= TOLERANCIA_IDENTIDADE
        for col, valor in enumerate(vals, start=1):
            cell = ws.cell(i, col, valor)
            cell.border = THIN
            cell.font = Font(name="Calibri", size=8)
            if col > 1:
                cell.number_format = '#,##0.0;(#,##0.0);"—"'
            cell.fill = PatternFill("solid", fgColor=VERDE if ok else VERMELHO)
    ws.freeze_panes = "A5"
    for i, w in enumerate([12, 16, 14, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _aba_grafico(wb: Workbook, anual: pd.DataFrame, anos: list[int]) -> None:
    ws = wb.create_sheet("Grafico")
    ws["A1"] = "Contribuição anual dos fatores (R$ milhões)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    fatores = anual[anual["Papel"] == "fator"].copy()
    headers = ["Fator"] + [str(a) for a in anos]
    _cab(ws, headers, 3)
    for i, (_, row) in enumerate(fatores.iterrows(), start=4):
        ws.cell(i, 1, row["Item"]).border = THIN
        ws.cell(i, 1).font = Font(name="Calibri", size=8)
        for j, a in enumerate(anos, start=2):
            cell = ws.cell(i, j, _fmt_milhoes(row[a]))
            cell.number_format = '#,##0.0;(#,##0.0);"—"'
            cell.border = THIN
    # Bloco transposto (anos nas linhas) para o gráfico empilhado
    start = 5 + len(fatores)
    ws.cell(start, 1, "Ano").font = Font(bold=True, color=BRANCO)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=AZUL)
    for j, (_, row) in enumerate(fatores.iterrows(), start=2):
        cell = ws.cell(start, j, row["Item"])
        cell.font = Font(bold=True, color=BRANCO, name="Calibri", size=8)
        cell.fill = PatternFill("solid", fgColor=AZUL)
    for i, a in enumerate(anos, start=1):
        ws.cell(start + i, 1, a)
        for j, (_, row) in enumerate(fatores.iterrows(), start=2):
            cell = ws.cell(start + i, j, _fmt_milhoes(row[a]))
            cell.number_format = "#,##0"
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Fatores condicionantes da base — fluxo anual"
    chart.y_axis.title = "R$ milhões"
    chart.style = 10
    chart.height = 12
    chart.width = 22
    data = Reference(
        ws,
        min_col=2,
        max_col=1 + len(fatores),
        min_row=start,
        max_row=start + len(anos),
    )
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(anos))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A20")
    ws.column_dimensions["A"].width = 44
    for i in range(2, 4 + len(anos)):
        ws.column_dimensions[get_column_letter(i)].width = 11


def _aba_discriminativo(
    wb: Workbook,
    disc: pd.DataFrame,
    ultimo: pd.Timestamp,
) -> None:
    """Anos nas linhas, fatores nas colunas, variação = SOMA algébrica."""
    ws = wb.create_sheet("Discriminativo", 1)
    n_fatores = len(FATORES_SOMA)
    col_var = 2 + n_fatores  # 1=Ano, 2..9=fatores, 10=variação
    ultima_col = 1 + len(disc.columns)
    ws["A1"] = (
        "Discriminativo — fatores que influenciam a base monetária (2000–2026)"
    )
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=AZUL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ultima_col)
    ws["A2"] = (
        "Cada linha é um ano. As colunas dos fatores são o fluxo acumulado no ano "
        f"(R$ milhões). A coluna «{COL_VARIACAO}» é a soma algébrica dos oito "
        "fatores (fórmula SOMA; primário e secundário são detalhe do total de "
        "títulos e não entram de novo). 2026* = até o último mês publicado."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ultima_col)
    ws.row_dimensions[2].height = 36

    nomes_fatores = [s.nome for s in FATORES_SOMA]
    nomes_detalhe = [s.nome for s in DETALHES]
    nomes_estoque = [s.nome for s in ESTOQUES]
    headers = (
        ["Ano"]
        + nomes_fatores
        + [COL_VARIACAO]
        + nomes_detalhe
        + nomes_estoque
    )
    _cab(ws, headers, 4)

    fill_var = PatternFill("solid", fgColor=DOURADO)
    fill_est = PatternFill("solid", fgColor=AZUL_CLARO)
    letra_ini = get_column_letter(2)
    letra_fim = get_column_letter(1 + n_fatores)

    for i, row in enumerate(disc.itertuples(index=False), start=5):
        ano = int(row.Ano)
        rotulo = f"{ano}*" if ano == ultimo.year and ultimo.month < 12 else str(ano)
        c_ano = ws.cell(i, 1, rotulo)
        c_ano.border = THIN
        c_ano.font = Font(name="Calibri", size=10, bold=True)
        c_ano.alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            c_ano.fill = PatternFill("solid", fgColor=CINZA)

        valores = {disc.columns[j]: row[j] for j in range(len(disc.columns))}

        for j, nome in enumerate(nomes_fatores, start=2):
            cell = ws.cell(i, j)
            cell.border = THIN
            _pintar_numero(cell, float(valores[nome]), "fator")
            if i % 2 == 0 and cell.fill.fgColor is None:
                cell.fill = PatternFill("solid", fgColor=CINZA)

        c_var = ws.cell(i, col_var)
        c_var.value = f"=SUM({letra_ini}{i}:{letra_fim}{i})"
        c_var.number_format = '#,##0.0;(#,##0.0);"—"'
        c_var.font = Font(name="Calibri", size=9, bold=True)
        c_var.border = THIN
        c_var.fill = fill_var

        col = col_var + 1
        for nome in nomes_detalhe:
            cell = ws.cell(i, col)
            cell.border = THIN
            _pintar_numero(cell, float(valores[nome]) if pd.notna(valores[nome]) else float("nan"), "detalhe")
            col += 1
        for nome in nomes_estoque:
            cell = ws.cell(i, col)
            cell.border = THIN
            cell.fill = fill_est
            _pintar_numero(cell, float(valores[nome]) if pd.notna(valores[nome]) else float("nan"), "estoque")
            cell.font = Font(name="Calibri", size=9, bold=True)
            col += 1

    # linha de totais 2000–2026 (soma dos fluxos; estoque = último ano)
    tot = 5 + len(disc)
    ws.cell(tot, 1, "Total fluxos").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(tot, 1).fill = fill_var
    ws.cell(tot, 1).border = THIN
    primeira = 5
    ultima = 4 + len(disc)
    for j in range(2, col_var + 1):
        letra = get_column_letter(j)
        cell = ws.cell(tot, j, f"=SUM({letra}{primeira}:{letra}{ultima})")
        cell.number_format = '#,##0.0;(#,##0.0);"—"'
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.fill = fill_var
        cell.border = THIN
    for j in range(col_var + 1, ultima_col + 1):
        ws.cell(tot, j).border = THIN
        ws.cell(tot, j).fill = fill_var

    ws.column_dimensions["A"].width = 12
    for j, nome in enumerate(headers, start=1):
        if j == 1:
            continue
        ws.column_dimensions[get_column_letter(j)].width = 16 if j != col_var else 18
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ultima_col)}{ultima}"
    ws.row_dimensions[4].height = 42
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.cell(
        tot + 2,
        1,
        "Valores em R$ milhões. Verde = expansão; vermelho = contração. "
        f"«{COL_VARIACAO}» = SOMA algébrica das oito colunas de fatores da mesma linha.",
    ).font = Font(name="Calibri", size=8, italic=True, color="666666")


def escrever_planilha(
    *,
    anual: pd.DataFrame,
    discriminativo: pd.DataFrame,
    dezembro: pd.DataFrame,
    mensal: pd.DataFrame,
    ident: pd.DataFrame,
    saida: Path,
    ultimo: pd.Timestamp,
) -> Path:
    saida.parent.mkdir(parents=True, exist_ok=True)
    anos = [c for c in anual.columns if isinstance(c, int)]
    wb = Workbook()
    _aba_metodologia(wb, datetime.now(), ultimo, len(mensal))
    _aba_discriminativo(wb, discriminativo, ultimo)
    ws_a = wb.create_sheet("Anual")
    _escrever_matriz(
        ws_a,
        anual,
        "Fatores condicionantes da base monetária — fluxo no ano e estoque de fim de período",
        "Fluxo acumulado de janeiro a dezembro (2026 até o último mês publicado). "
        "Estoque da base = último dia do último mês do ano. Unidade: R$ milhões.",
        anos,
        ultimo,
    )
    ws_d = wb.create_sheet("Dezembro")
    _escrever_matriz(
        ws_d,
        dezembro,
        "Fatores no mês de dezembro (definição literal do SGS «final de período»)",
        "Fluxo somente do mês de dezembro de cada ano (em 2026, último mês disponível). "
        "Estoque da base no mesmo mês. Unidade: R$ milhões.",
        anos,
        ultimo,
    )
    _aba_mensal(wb, mensal)
    _aba_identidade(wb, ident)
    _aba_grafico(wb, anual, anos)
    wb.save(saida)
    print(f"[OK] Planilha: {saida} ({saida.stat().st_size / 1024:.1f} KB)")
    return saida


def processar(
    *,
    pasta_cache: Path,
    saida: Path,
    usar_cache: bool = True,
    arquivos: dict[int, Path] | None = None,
) -> Path:
    print(f"[{MARKER}]")
    painel = carregar_painel(pasta_cache, usar_cache=usar_cache, arquivos=arquivos)
    anos = anos_com_dados(painel)
    ultimo = painel.index.max()
    print(f"[INFO] painel {painel.index.min().date()} → {ultimo.date()} | anos={anos[0]}–{anos[-1]}")
    anual = tabela_anual(painel, anos)
    disc = tabela_discriminativo(painel, anos)
    dezembro = tabela_dezembro(painel, anos)
    mensal = tabela_mensal(painel)
    ident = identidade_mensal(painel)
    if not ident.empty:
        ok = (ident["residuo"].abs() <= TOLERANCIA_IDENTIDADE).mean()
        print(f"[INFO] identidade mensal ok em {ok:.1%} dos meses")
    path = escrever_planilha(
        anual=anual,
        discriminativo=disc,
        dezembro=dezembro,
        mensal=mensal,
        ident=ident,
        saida=saida,
        ultimo=ultimo,
    )
    # resumo no stdout
    print("\n=== Fluxo anual dos fatores e estoque da base (R$ milhões) ===")
    fatores = anual[anual["Papel"].isin(["fator", "total", "estoque"])]
    for a in anos:
        if a not in (2000, 2010, 2016, 2020, 2023, 2024, 2025, anos[-1]):
            continue
        print(f"\n-- {a} --")
        for _, r in fatores.iterrows():
            if r["Papel"] == "estoque" and r["Recuo"] == 1:
                continue
            print(f"  {r['Item']:<52} {_fmt_milhoes(r[a]):>12,.1f}")
    return path


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pasta-cache", type=Path, default=ROOT / "data" / "sgs")
    p.add_argument(
        "--saida",
        type=Path,
        default=ROOT / "output" / "fatores_condicionantes_base_monetaria.xlsx",
    )
    p.add_argument("--sem-cache", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        processar(
            pasta_cache=args.pasta_cache,
            saida=args.saida,
            usar_cache=not args.sem_cache,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
