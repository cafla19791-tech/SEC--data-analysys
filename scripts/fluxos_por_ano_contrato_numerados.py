#!/usr/bin/env python3
"""Gera fluxos a partir de BNDES_INDIRETAS_NUMERADOS.xlsx (uma aba por ano).

Para cada aba de contratos (2002, 2003, … com números ``N-AAAA``), gera as
parcelas correspondentes e grava:

  - CSV completo: ``saida/fluxos_por_ano_contrato/YYYY.csv``  ← fonte da verdade
  - Excel consolidado (amostra por ano + RESUMO):
    ``saida/FLUXOS_BNDES_INDIRETAS_POR_ANO_CONTRATO.xlsx``

Regra: todas as parcelas de um contrato ficam na aba/arquivo do **ano do
contrato** (ano da aba de origem). Impacto fiscal continua capitalizado na
``data_fluxo``.

Por padrão só grava CSV por ano (Excel por ano trava o ContAgil em massas
grandes). Rode de novo para **retomar** anos ainda sem CSV.

Uso ContAgil::

  python scripts/fluxos_por_ano_contrato_numerados.py
  python scripts/fluxos_por_ano_contrato_numerados.py ^
    --numerados saida\\BNDES_INDIRETAS_NUMERADOS.xlsx ^
    --fatores fator_acumulado_SELIC_TJLP_TLP.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

_SCRIPTS = Path(__file__).resolve().parent
ROOT = _SCRIPTS.parent
for _p in (str(ROOT), str(_SCRIPTS)):
    if _p not in sys.path:
        sys.path.insert(0, _p)


def _configure_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconf = getattr(stream, "reconfigure", None)
        if reconf is None:
            continue
        try:
            reconf(encoding="utf-8", errors="replace")
        except Exception:
            try:
                reconf(errors="replace")
            except Exception:
                pass


_configure_stdio()


def _load_mod(name: str):
    """Carrega sibling em scripts.* (mesmo nome que contagil_fluxos_seguro).

    Importante: usar o mesmo ``sys.modules`` key evita duas classes SelicSerie
    (isinstance falhava e virava ``float(SelicSerie)`` → zero parcelas).
    """
    import importlib.util
    import types

    full = f"scripts.{name}"
    if full in sys.modules:
        return sys.modules[full]

    path = _SCRIPTS / f"{name}.py"
    if path.exists():
        if "scripts" not in sys.modules:
            pkg = types.ModuleType("scripts")
            pkg.__path__ = [str(_SCRIPTS)]
            pkg.__package__ = "scripts"
            sys.modules["scripts"] = pkg
        if str(ROOT) not in sys.path:
            sys.path.insert(0, str(ROOT))
        spec = importlib.util.spec_from_file_location(full, path)
        if spec is not None and spec.loader is not None:
            mod = importlib.util.module_from_spec(spec)
            sys.modules[full] = mod
            sys.modules[name] = mod
            spec.loader.exec_module(mod)
            return mod
    return __import__(f"scripts.{name}", fromlist=["*"])


_gf = _load_mod("gerar_fluxos")
try:
    _seg = _load_mod("contagil_fluxos_seguro")
except Exception:  # noqa: BLE001
    _seg = None

CONTAGIL_WINPYTHON = _gf.CONTAGIL_WINPYTHON
CONTAGIL_PASTA_SAIDA = _gf.CONTAGIL_PASTA_SAIDA
normalizar_colunas = _gf.normalizar_colunas
gerar_fluxos = _gf.gerar_fluxos
gerar_e_gravar_fluxos = _gf.gerar_e_gravar_fluxos

EXCEL_MAX = 1_000_000
# Amostra por aba no Excel consolidado (CSV completo fica em fluxos_por_ano_contrato/)
EXCEL_AMOSTRA_ABA = 50_000
_ANO_SHEET = re.compile(r"^(19|20)\d{2}$")
MARKER = "fluxos-por-ano-contrato-numerados-20260815a"


def resolver_numerados(path: Optional[Path]) -> Path:
    candidatos = []
    if path is not None:
        candidatos.append(Path(path))
    candidatos.extend(
        [
            Path.cwd() / "saida" / "BNDES_INDIRETAS_NUMERADOS.xlsx",
            CONTAGIL_PASTA_SAIDA / "BNDES_INDIRETAS_NUMERADOS.xlsx",
            CONTAGIL_WINPYTHON / "saida" / "BNDES_INDIRETAS_NUMERADOS.xlsx",
            ROOT / "output" / "BNDES_INDIRETAS_NUMERADOS_DEMO.xlsx",
            ROOT / "saida" / "BNDES_INDIRETAS_NUMERADOS.xlsx",
        ]
    )
    for c in candidatos:
        if c.exists():
            return c
    raise FileNotFoundError(
        "BNDES_INDIRETAS_NUMERADOS.xlsx não encontrado. "
        "Rode numerar_contratos_indiretas.bat antes."
    )


def resolver_fatores(path: Optional[Path]):
    if _seg is None:
        print("[AVISO] contagil_fluxos_seguro indisponivel - SELIC 14,5% a.a.")
        return 0.145
    candidatos: list[Path] = []
    if path is not None:
        candidatos.append(Path(path))
    candidatos.extend(
        [
            Path.cwd() / "fator_acumulado_SELIC_TJLP_TLP.xlsx",
            CONTAGIL_WINPYTHON / "fator_acumulado_SELIC_TJLP_TLP.xlsx",
            ROOT / "data" / "selic_taxas_contagil.xlsx",
        ]
    )
    for c in candidatos:
        if not c.exists():
            continue
        try:
            print(f"[INFO] Fatores: {c}")
            return _seg.carregar_fatores_mensais(c)
        except Exception as exc:  # noqa: BLE001 — cai para SELIC constante
            print(
                f"[AVISO] Falha ao carregar fatores {c}: "
                f"{type(exc).__name__}: {exc}"
            )
            continue
    print("[AVISO] Sem arquivo de fatores valido - usando SELIC 14,5% a.a. composta.")
    return 0.145


def listar_abas_ano(path: Path) -> list[str]:
    xl = pd.ExcelFile(path)
    anos = [s for s in xl.sheet_names if _ANO_SHEET.match(str(s).strip())]
    if not anos:
        # fallback: qualquer aba cujo nome contenha ano
        anos = [s for s in xl.sheet_names if re.search(r"(19|20)\d{2}", str(s))]
    if not anos:
        raise ValueError(f"Nenhuma aba de ano em {path}: {xl.sheet_names}")
    return sorted(anos, key=lambda s: int(re.search(r"(19|20)\d{2}", str(s)).group(0)))


def _escrever_aba_excel(wb: Workbook, nome: str, df: pd.DataFrame, *, first: bool) -> None:
    if first:
        ws = wb.active
        ws.title = str(nome)[:31]
    else:
        ws = wb.create_sheet(str(nome)[:31])
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(list(row))
        if i == 0:
            for col in range(1, len(row) + 1):
                cell = ws.cell(1, col)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"


def _contar_linhas_csv(path: Path) -> int:
    """Linhas de dados (sem cabeçalho)."""
    if not path.exists() or path.stat().st_size == 0:
        return 0
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        n = sum(1 for _ in f)
    return max(0, n - 1)


def _salvar_consolidado(
    pasta_saida: Path,
    pasta_csv: Path,
    resumo_rows: list[dict],
    *,
    amostra_por_aba: int,
) -> Path:
    xlsx_out = pasta_saida / "FLUXOS_BNDES_INDIRETAS_POR_ANO_CONTRATO.xlsx"
    wb = Workbook()
    first_sheet = True

    for row in resumo_rows:
        if row.get("status") not in ("ok", "retomado"):
            continue
        csv_ano = Path(row["csv"])
        if not csv_ano.exists():
            continue
        n_parc = int(row.get("qtd_parcelas") or 0)
        df_aba = pd.read_csv(csv_ano, nrows=amostra_por_aba)
        if n_parc > amostra_por_aba:
            print(
                f"  [INFO] Aba Excel {row['ano_contrato']}: amostra "
                f"{len(df_aba):,}/{n_parc:,}; CSV completo em {csv_ano.name}"
            )
        _escrever_aba_excel(wb, str(row["ano_contrato"]), df_aba, first=first_sheet)
        first_sheet = False

    if first_sheet:
        wb.active.title = "vazio"
    else:
        ws = wb.create_sheet("RESUMO", 0)
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        headers = list(resumo_rows[0].keys()) if resumo_rows else []
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.fill = fill
            cell.font = font
        for row in resumo_rows:
            ws.append([row.get(h) for h in headers])
        ws2 = wb.create_sheet("NOTA")
        ws2["A1"] = (
            "Cada aba YYYY é amostra dos fluxos dos contratos da aba YYYY de "
            "BNDES_INDIRETAS_NUMERADOS.xlsx (numeração N-AAAA). "
            "CSVs COMPLETOS: pasta fluxos_por_ano_contrato\\YYYY.csv. "
            "Todas as parcelas de um contrato ficam no ano do contrato; "
            "o impacto fiscal continua capitalizado na data_fluxo. "
            f"Pasta: {pasta_csv}"
        )
        ws2.column_dimensions["A"].width = 110

    pasta_saida.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_out)
    return xlsx_out


def processar(
    numerados: Path,
    pasta_saida: Path,
    *,
    fatores,
    ano_min: Optional[int] = None,
    ano_max: Optional[int] = None,
    lote: int = 2_000,
    excel_max: int = EXCEL_MAX,
    amostra_excel: int = EXCEL_AMOSTRA_ABA,
    retomar: bool = True,
    gravar_excel_ano: bool = False,
) -> Path:
    """Gera um CSV por ano de contrato; Excel consolidado com amostra.

    Por padrão NÃO grava Excel por ano (só CSV) — gravar ~1M linhas em .xlsx
    após 2002 costumava travar o ContAgil e parecer que "só gerou 2002".
    Com ``retomar=True``, anos que já têm CSV com linhas são pulados.
    """
    del excel_max  # mantido na assinatura por compatibilidade
    pasta_saida = Path(pasta_saida)
    pasta_csv = pasta_saida / "fluxos_por_ano_contrato"
    pasta_csv.mkdir(parents=True, exist_ok=True)

    abas = listar_abas_ano(numerados)
    resumo_rows: list[dict] = []

    print(f"[{MARKER}] numerados={numerados}")
    print(f"[{MARKER}] abas encontradas ({len(abas)}): {abas}")
    print(f"[{MARKER}] saida CSV: {pasta_csv}")
    print(f"[{MARKER}] retomar={retomar} | excel_por_ano={gravar_excel_ano}")
    sys.stdout.flush()

    for aba in abas:
        m = re.search(r"(19|20)\d{2}", str(aba))
        ano = int(m.group(0)) if m else None
        if ano is None:
            continue
        if ano_min is not None and ano < ano_min:
            continue
        if ano_max is not None and ano > ano_max:
            continue

        csv_ano = pasta_csv / f"{ano}.csv"
        print(f"\n=== Ano {ano} (aba '{aba}') ===")
        sys.stdout.flush()

        if retomar:
            n_exist = _contar_linhas_csv(csv_ano)
            if n_exist > 0:
                print(f"  [RETOMAR] ja existe {csv_ano.name} com {n_exist:,} parcelas - pulando")
                resumo_rows.append(
                    {
                        "ano_contrato": ano,
                        "qtd_contratos": "",
                        "qtd_parcelas": n_exist,
                        "status": "retomado",
                        "csv": str(csv_ano),
                        "erro": "",
                    }
                )
                continue

        try:
            bruto = pd.read_excel(numerados, sheet_name=aba)
            if bruto.empty:
                print("  [AVISO] aba vazia - pulando")
                resumo_rows.append(
                    {
                        "ano_contrato": ano,
                        "qtd_contratos": 0,
                        "qtd_parcelas": 0,
                        "status": "vazio",
                        "csv": str(csv_ano),
                        "erro": "aba vazia",
                    }
                )
                continue

            contratos = normalizar_colunas(bruto)
            if contratos.empty:
                print("  [AVISO] nenhum contrato valido - pulando")
                resumo_rows.append(
                    {
                        "ano_contrato": ano,
                        "qtd_contratos": 0,
                        "qtd_parcelas": 0,
                        "status": "vazio",
                        "csv": str(csv_ano),
                        "erro": "sem contratos válidos",
                    }
                )
                continue

            t0 = time.time()
            stats = gerar_e_gravar_fluxos(
                contratos,
                fatores,
                saida_csv=csv_ano,
                saida_xlsx=(pasta_csv / f"{ano}.xlsx") if gravar_excel_ano else None,
                lote=lote,
                gravar_excel=gravar_excel_ano,
            )
            n_parc = int(stats["parcelas"])
            print(
                f"  OK ano {ano}: {stats['contratos']:,} contratos | "
                f"{n_parc:,} parcelas | {time.time() - t0:.1f}s"
            )
            resumo_rows.append(
                {
                    "ano_contrato": ano,
                    "qtd_contratos": stats["contratos"],
                    "qtd_parcelas": n_parc,
                    "status": "ok",
                    "csv": str(csv_ano),
                    "erro": "",
                }
            )
        except Exception as exc:  # noqa: BLE001 — um ano não derruba os demais
            print(f"  [ERRO] ano {ano}: {type(exc).__name__}: {exc}")
            sys.stdout.flush()
            resumo_rows.append(
                {
                    "ano_contrato": ano,
                    "qtd_contratos": "",
                    "qtd_parcelas": 0,
                    "status": "erro",
                    "csv": str(csv_ano),
                    "erro": f"{type(exc).__name__}: {exc}",
                }
            )
            continue

    # RESUMO parcial em CSV (visível mesmo se Excel falhar)
    resumo_csv = pasta_csv / "RESUMO.csv"
    pd.DataFrame(resumo_rows).to_csv(resumo_csv, index=False)
    print(f"\n[OK] Resumo: {resumo_csv}")

    xlsx_out = _salvar_consolidado(
        pasta_saida,
        pasta_csv,
        resumo_rows,
        amostra_por_aba=amostra_excel,
    )
    feitos = [r for r in resumo_rows if r.get("status") in ("ok", "retomado")]
    erros = [r for r in resumo_rows if r.get("status") == "erro"]
    print(f"[OK] Excel consolidado: {xlsx_out}")
    print(f"[OK] CSVs por ano: {pasta_csv}")
    print(f"[OK] Anos com fluxo: {len(feitos)} | erros: {len(erros)}")
    if feitos:
        print("     -> " + ", ".join(str(r["ano_contrato"]) for r in feitos))
    if erros:
        print("     -> falhas: " + ", ".join(str(r["ano_contrato"]) for r in erros))
    return xlsx_out


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--numerados",
        type=Path,
        default=None,
        help="BNDES_INDIRETAS_NUMERADOS.xlsx (default: saida/ ContAgil)",
    )
    p.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Pasta de saída (default: saida ContAgil ou ./saida)",
    )
    p.add_argument("--fatores", type=Path, default=None)
    p.add_argument("--ano-min", type=int, default=None)
    p.add_argument("--ano-max", type=int, default=None)
    p.add_argument("--lote", type=int, default=2_000)
    p.add_argument(
        "--sem-retomar",
        action="store_true",
        help="Refaz todos os anos mesmo se o CSV já existir",
    )
    p.add_argument(
        "--excel-por-ano",
        action="store_true",
        help="Também grava YYYY.xlsx por ano (lento; padrão é só CSV)",
    )
    p.add_argument(
        "--amostra-excel",
        type=int,
        default=EXCEL_AMOSTRA_ABA,
        help="Linhas por aba no Excel consolidado (default 50000)",
    )
    args = p.parse_args(argv)

    numerados = resolver_numerados(args.numerados)
    if args.saida is not None:
        pasta_saida = Path(args.saida)
    elif CONTAGIL_PASTA_SAIDA.exists():
        pasta_saida = CONTAGIL_PASTA_SAIDA
    else:
        pasta_saida = Path.cwd() / "saida"
        pasta_saida.mkdir(exist_ok=True)

    fatores = resolver_fatores(args.fatores)
    processar(
        numerados,
        pasta_saida,
        fatores=fatores,
        ano_min=args.ano_min,
        ano_max=args.ano_max,
        lote=args.lote,
        retomar=not args.sem_retomar,
        gravar_excel_ano=args.excel_por_ano,
        amostra_excel=args.amostra_excel,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
