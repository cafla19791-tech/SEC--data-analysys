"""Taxa acumulada das taxas básicas do BIS em cada lapso Selic (105–161).

Fonte das taxas: BIS WS_CBPOL
  https://data.bis.org/static/bulk/WS_CBPOL_csv_col.zip

Uso:
  python3 scripts/lapsos_bis_juros.py
  python3 scripts/lapsos_bis_juros.py --sem-download
"""

from __future__ import annotations

import argparse
import sys
import zipfile
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
BIS_URL = "https://data.bis.org/static/bulk/WS_CBPOL_csv_col.zip"
BIS_DIR = DATA_DIR / "bis"
LAPSOS_CSV = OUTPUT_DIR / "copom_selic_lapsos_pregao_1999_2026.csv"
PREGAO_CSV = DATA_DIR / "sgs_11_selic_pregao.csv"
LAPSO_INI = 105
LAPSO_FIM = 161
BASE_DIAS = 252

PAISES_PT = {
    "AR": "Argentina",
    "AT": "Áustria",
    "AU": "Austrália",
    "BE": "Bélgica",
    "BR": "Brasil",
    "CA": "Canadá",
    "CH": "Suíça",
    "CL": "Chile",
    "CN": "China",
    "CO": "Colômbia",
    "CZ": "Tchéquia",
    "DE": "Alemanha",
    "DK": "Dinamarca",
    "ES": "Espanha",
    "FR": "França",
    "GB": "Reino Unido",
    "GR": "Grécia",
    "HK": "Hong Kong",
    "HR": "Croácia",
    "HU": "Hungria",
    "ID": "Indonésia",
    "IL": "Israel",
    "IN": "Índia",
    "IS": "Islândia",
    "IT": "Itália",
    "JP": "Japão",
    "KR": "Coreia do Sul",
    "KW": "Kuwait",
    "MA": "Marrocos",
    "MK": "Macedônia do Norte",
    "MX": "México",
    "MY": "Malásia",
    "NL": "Países Baixos",
    "NO": "Noruega",
    "NZ": "Nova Zelândia",
    "PE": "Peru",
    "PH": "Filipinas",
    "PL": "Polônia",
    "PT": "Portugal",
    "RO": "Romênia",
    "RS": "Sérvia",
    "RU": "Rússia",
    "SA": "Arábia Saudita",
    "SE": "Suécia",
    "TH": "Tailândia",
    "TR": "Turquia",
    "US": "Estados Unidos",
    "XM": "Zona do euro",
    "ZA": "África do Sul",
}


def _fmt(valor: float, casas: int = 4) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    txt = f"{float(valor):,.{casas}f}"
    return txt.replace(",", "X").replace(".", ",").replace("X", ".")


def baixar_bis(cache_dir: Path, baixar: bool = True) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    csv_path = cache_dir / "WS_CBPOL_csv_col.csv"
    if csv_path.exists():
        return csv_path
    zip_path = cache_dir / "WS_CBPOL_csv_col.zip"
    if not zip_path.exists():
        if not baixar:
            raise FileNotFoundError(zip_path)
        print(f"Baixando {BIS_URL} ...", flush=True)
        resp = requests.get(BIS_URL, timeout=180)
        resp.raise_for_status()
        zip_path.write_bytes(resp.content)
    with zipfile.ZipFile(zip_path) as zf:
        nome = next(n for n in zf.namelist() if n.lower().endswith(".csv"))
        csv_path.write_bytes(zf.read(nome))
    return csv_path


def carregar_bis_diario(csv_path: Path, inicio: str = "2014-01-01") -> pd.DataFrame:
    """Datas × países (código ISO), série diária do BIS, valores em % a.a."""

    def _usar(nome: str) -> bool:
        if nome in {"FREQ", "REF_AREA", "Reference area"}:
            return True
        return len(nome) == 10 and nome[4] == "-" and nome >= inicio

    bruto = pd.read_csv(csv_path, usecols=_usar)
    diario = bruto.loc[bruto["FREQ"] == "D"].copy()
    datas = [c for c in diario.columns if len(str(c)) == 10 and str(c)[4] == "-"]
    largo = diario.set_index("REF_AREA")[datas]
    largo.columns = pd.to_datetime(largo.columns)
    # datas × países
    return largo.T.sort_index().apply(pd.to_numeric, errors="coerce")


def taxa_acumulada(taxas_aa: pd.Series, base: int = BASE_DIAS) -> float:
    """Produto diário (1 + i/base) − 1, em %; i é a taxa anual em %."""
    s = pd.to_numeric(taxas_aa, errors="coerce").dropna()
    if s.empty:
        return float("nan")
    return float((1.0 + s.to_numpy() / 100.0 / float(base)).prod() - 1.0) * 100.0


def taxas_no_pregao(serie: pd.Series, pregao: pd.DatetimeIndex) -> pd.Series:
    """Taxa vigente em cada pregão (última observação do BIS até aquele dia)."""
    serie = pd.to_numeric(serie, errors="coerce")
    serie.index = pd.to_datetime(serie.index)
    idx = serie.index.union(pregao).sort_values()
    return serie.reindex(idx).ffill().reindex(pregao)


def carregar_lapsos(path: Path, ordem_ini: int, ordem_fim: int) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["inicio", "fim"])
    return df[(df["ordem"] >= ordem_ini) & (df["ordem"] <= ordem_fim)].sort_values("ordem")


def carregar_pregao(path: Path) -> pd.DatetimeIndex:
    df = pd.read_csv(path, parse_dates=["data"])
    return pd.DatetimeIndex(pd.to_datetime(df["data"])).sort_values()


def pregao_do_lapso(pregao: pd.DatetimeIndex, inicio: pd.Timestamp, fim: pd.Timestamp) -> pd.DatetimeIndex:
    return pregao[(pregao >= pd.Timestamp(inicio)) & (pregao <= pd.Timestamp(fim))]


def acumulado_paises(
    bis: pd.DataFrame,
    pregao: pd.DatetimeIndex,
    nomes: dict[str, str] | None = None,
) -> pd.DataFrame:
    nomes = nomes or PAISES_PT
    linhas = []
    for codigo in bis.columns:
        vals = taxas_no_pregao(bis[codigo], pregao)
        usados = vals.dropna()
        if usados.empty:
            continue
        linhas.append(
            {
                "codigo": codigo,
                "pais": nomes.get(codigo, codigo),
                "taxa_ini": float(usados.iloc[0]),
                "taxa_fim": float(usados.iloc[-1]),
                "taxa_media": float(usados.mean()),
                "taxa_acumulada": taxa_acumulada(usados),
                "n_pregao": int(len(usados)),
            }
        )
    out = pd.DataFrame(linhas)
    if out.empty:
        return out
    out["brasil"] = out["codigo"] == "BR"
    return out.sort_values(["brasil", "taxa_acumulada"], ascending=[False, False]).drop(columns=["brasil"])


def _borda() -> Border:
    lado = Side(style="thin", color="1A1A1A")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _escrever_cabecalho(ws, lapso: pd.Series, n_pregao: int, ultima_bis: pd.Timestamp) -> int:
    fonte_t = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    fonte = Font(name="Calibri", size=10)
    ws["A1"] = f"Lapso {int(lapso.ordem)} — taxa básica acumulada (BIS)"
    ws["A1"].font = fonte_t
    ws.merge_cells("A1:F1")
    linhas = [
        f"Selic do Copom no lapso: {_fmt(lapso.selic, 2)}% a.a.",
        f"Termo inicial: {pd.Timestamp(lapso.inicio).strftime('%d/%m/%Y')}    "
        f"Termo final: {pd.Timestamp(lapso.fim).strftime('%d/%m/%Y')}    "
        f"Dias com pregão (Brasil): {n_pregao}",
        "Fonte: BIS WS_CBPOL — https://data.bis.org/static/bulk/WS_CBPOL_csv_col.zip",
        f"Acumulação: produto nos pregões (1 + i/{BASE_DIAS}) − 1; i = taxa básica "
        f"do banco central (% a.a.). Série diária do BIS até "
        f"{ultima_bis.strftime('%d/%m/%Y')} (última observação vigente é repetida).",
    ]
    for i, txt in enumerate(linhas, start=2):
        ws[f"A{i}"] = txt
        ws[f"A{i}"].font = fonte
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    return 7


def _escrever_tabela(ws, linha0: int, df: pd.DataFrame) -> None:
    cabs = [
        "País",
        "Código",
        "Taxa inicial (% a.a.)",
        "Taxa final (% a.a.)",
        "Taxa média (% a.a.)",
        "Taxa acumulada (%)",
    ]
    campos = ["pais", "codigo", "taxa_ini", "taxa_fim", "taxa_media", "taxa_acumulada"]
    borda = _borda()
    fill_cab = PatternFill("solid", fgColor="E8E8E8")
    fill_br = PatternFill("solid", fgColor="D6EAF8")
    fill_alt = PatternFill("solid", fgColor="F4F4F4")
    fonte_cab = Font(name="Calibri", size=10, bold=True)
    fonte = Font(name="Calibri", size=10)
    for col, cab in enumerate(cabs, start=1):
        cell = ws.cell(linha0, col, cab)
        cell.font = fonte_cab
        cell.fill = fill_cab
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, rec in enumerate(df.to_dict("records")):
        row = linha0 + 1 + i
        br = rec["codigo"] == "BR"
        for col, campo in enumerate(campos, start=1):
            valor = rec[campo]
            if col >= 3:
                cell = ws.cell(row, col, None if pd.isna(valor) else float(valor))
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell = ws.cell(row, col, valor)
                cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
            cell.font = Font(name="Calibri", size=10, bold=br)
            cell.border = borda
            if br:
                cell.fill = fill_br
            elif i % 2 == 1:
                cell.fill = fill_alt
    larguras = [28, 10, 22, 22, 22, 22]
    for col, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A{linha0}:F{linha0 + len(df)}"
    ws.freeze_panes = f"A{linha0 + 1}"


def gerar_planilha(
    lapsos: pd.DataFrame,
    bis: pd.DataFrame,
    pregao: pd.DatetimeIndex,
    path: Path,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    ultima_bis = pd.Timestamp(bis.index.max())
    wb = Workbook()

    ws_idx = wb.active
    ws_idx.title = "Indice"
    ws_idx["A1"] = "Lapsos 105–161 — taxas básicas do BIS acumuladas nos pregões"
    ws_idx["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws_idx.merge_cells("A1:G1")
    ws_idx["A2"] = (
        "Uma aba por lapso da meta Selic do Copom. Em cada aba, a taxa acumulada "
        f"a juros de cada país, com (1 + i/{BASE_DIAS}) em cada pregão brasileiro."
    )
    ws_idx.merge_cells("A2:G2")

    cabs_idx = [
        "Lapso",
        "Selic (% a.a.)",
        "Termo inicial",
        "Termo final",
        "Pregões",
        "Acumulada Brasil (%)",
        "Países",
    ]
    borda = _borda()
    fill_cab = PatternFill("solid", fgColor="E8E8E8")
    for col, cab in enumerate(cabs_idx, start=1):
        cell = ws_idx.cell(4, col, cab)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.fill = fill_cab
        cell.border = borda
        cell.alignment = Alignment(horizontal="center")

    for i, lapso in enumerate(lapsos.itertuples(index=False)):
        dias = pregao_do_lapso(pregao, lapso.inicio, lapso.fim)
        tab = acumulado_paises(bis, dias)
        nome = f"Lapso {int(lapso.ordem)}"
        ws = wb.create_sheet(nome)
        linha0 = _escrever_cabecalho(ws, lapso, int(len(dias)), ultima_bis)
        _escrever_tabela(ws, linha0, tab)
        br = tab.loc[tab["codigo"] == "BR", "taxa_acumulada"]
        acum_br = float(br.iloc[0]) if len(br) else float("nan")
        vals_idx = [
            int(lapso.ordem),
            float(lapso.selic),
            pd.Timestamp(lapso.inicio).strftime("%d/%m/%Y"),
            pd.Timestamp(lapso.fim).strftime("%d/%m/%Y"),
            int(len(dias)),
            None if pd.isna(acum_br) else acum_br,
            int(len(tab)),
        ]
        for col, valor in enumerate(vals_idx, start=1):
            cell = ws_idx.cell(5 + i, col, valor)
            cell.font = Font(name="Calibri", size=10)
            cell.border = borda
            cell.alignment = Alignment(horizontal="center" if col != 6 else "right")
            if col in (2, 6) and isinstance(valor, float):
                cell.number_format = "#,##0.0000" if col == 6 else "#,##0.00"
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F4F4F4")

    for col, w in enumerate([10, 16, 16, 16, 12, 24, 10], start=1):
        ws_idx.column_dimensions[get_column_letter(col)].width = w
    ws_idx.freeze_panes = "A5"
    ws_idx.auto_filter.ref = f"A4:G{4 + len(lapsos)}"

    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=BIS_DIR)
    parser.add_argument("--lapsos", type=Path, default=LAPSOS_CSV)
    parser.add_argument("--pregao", type=Path, default=PREGAO_CSV)
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR / "lapsos_bis_juros_105_161.xlsx")
    parser.add_argument("--sem-download", action="store_true")
    parser.add_argument("--lapso-ini", type=int, default=LAPSO_INI)
    parser.add_argument("--lapso-fim", type=int, default=LAPSO_FIM)
    args = parser.parse_args(argv)

    csv_bis = baixar_bis(args.cache_dir, baixar=not args.sem_download)
    print(f"BIS: {csv_bis}", flush=True)
    bis = carregar_bis_diario(csv_bis)
    print(f"Série diária: {bis.shape[1]} países, {bis.index.min().date()}–{bis.index.max().date()}")
    lapsos = carregar_lapsos(args.lapsos, args.lapso_ini, args.lapso_fim)
    pregao = carregar_pregao(args.pregao)
    print(f"Lapsos {args.lapso_ini}–{args.lapso_fim}: {len(lapsos)}")
    path = gerar_planilha(lapsos, bis, pregao, args.output)
    print(f"Planilha: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
