#!/usr/bin/env python3
"""Gera planilha (xlsx) e PDF com lucros líquidos anuais de Itaú, Unibanco,
Itaú Unibanco, Bradesco, Santander Brasil e BTG Pactual (2002–2026).

Valores em R$ bilhões. Células vazias = dado não comparável ou inexistente
naquele perímetro societário. 2026 é parcial (1º semestre, quando houver).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Ano: (itau_contabil, itau_recorrente, unibanco, bradesco_contabil,
#       bradesco_recorrente, santander_ifrs, btg)
# None = não aplicável / não localizado com critério comparável
ROWS: list[tuple] = [
    (2002, 2.83, None, 1.01, 2.14, None, None, None),
    (2003, 3.27, None, 1.05, 2.29, None, 1.62, None),
    (2004, 4.63, None, 1.28, 3.33, None, None, None),
    (2005, 5.45, None, 1.84, 6.31, None, None, None),
    (2006, 5.88, None, 2.21, 6.46, None, 0.80, None),
    (2007, 7.66, None, 3.45, 7.91, None, 1.90, None),
    (2008, 4.85, 10.57, 2.85, 7.02, None, 2.38, 2.65),
    (2009, 14.09, None, None, 9.22, None, 5.51, None),
    (2010, 11.71, None, None, 9.94, None, 7.38, None),
    (2011, 13.84, None, None, 10.96, None, 7.74, None),
    (2012, 12.63, None, None, 11.29, None, 5.48, 3.26),
    (2013, 16.42, None, None, 12.40, None, 5.72, 2.78),
    (2014, 21.55, None, None, 15.31, None, 5.63, 3.41),
    (2015, 25.74, None, None, 18.13, 17.87, 9.78, 4.62),
    (2016, 21.63, None, None, 17.89, None, 7.33, None),
    (2017, 23.19, None, None, 17.09, None, 8.92, 2.60),
    (2018, 24.91, None, None, 16.58, 21.56, 12.58, 2.36),
    (2019, 27.11, 28.36, None, 21.02, None, 16.41, 3.83),
    (2020, 18.90, 18.50, None, 15.84, None, 13.42, 4.05),
    (2021, 26.76, 26.88, None, 23.17, None, 15.53, 6.34),
    (2022, 29.21, 30.79, None, 21.22, 20.70, 14.29, 7.84),
    (2023, 33.10, 35.62, None, 14.25, 16.30, 9.45, 9.93),
    (2024, 41.09, 41.40, None, 17.25, 19.55, 13.37, 11.79),
    (2025, 44.86, 46.83, None, 23.21, 24.65, 12.77, 15.95),
    (2026, None, 24.68, None, None, None, None, None),
]

HEADERS = [
    "Ano",
    "Itaú / Itaú Unibanco\n(contábil)",
    "Itaú Unibanco\n(recorrente)",
    "Unibanco",
    "Bradesco\n(contábil)",
    "Bradesco\n(recorrente)",
    "Santander Brasil\n(IFRS)",
    "BTG Pactual",
]

NAVY = "1B4F72"
GOLD = "B7950B"
YELLOW = "FCF3CF"
ORANGE = "FDEBD0"
ALT = "EBF5FB"
WHITE = "FFFFFF"
GREEN = "D5F5E3"
THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def fmt(v) -> str:
    if v is None:
        return "—"
    return f"{v:.2f}".replace(".", ",")


def write_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lucros 2002-2026"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.freeze_panes = "A6"

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Lucros líquidos anuais — Itaú, Unibanco, Itaú Unibanco, Bradesco, Santander Brasil e BTG Pactual"
    c.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center")
    for col in range(1, 9):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        "Valores em R$ bilhões, ano civil. 2026 = 1º semestre (Itaú Unibanco, recorrente). "
        "Itaú e Unibanco fundiram-se em nov/2008. BTG e BTG Pactual são a mesma franquia após 2009. "
        "Célula “—” = não aplicável ou sem dado comparável."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="2C3E50")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32
    for col in range(1, 9):
        ws.cell(2, col).fill = PatternFill("solid", fgColor="D4E6F1")

    ws.merge_cells("A3:H3")
    ws["A3"].value = (
        "2008 Itaú Unibanco recorrente = pro forma conjunto (Itaú 7,72 + Unibanco 2,85). "
        "2008 BTG Pactual = UBS Pactual. 2017 BTG ≈ R$ 2,6 bi. "
        "Santander 2025 gerencial = R$ 15,6 bi; BTG 2025 ajustado = R$ 16,68 bi. "
        "Fontes: RI dos bancos, IFRS/BRGAAP, VCP Scanner (ITUB, BBD, BSBR), Folha/Valor/G1."
    )
    ws["A3"].font = Font(name="Calibri", size=9, color="5D6D7E")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 36

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(5, col, h.replace("\n", " "))
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[5].height = 28

    for i, row in enumerate(ROWS):
        r = 6 + i
        year = row[0]
        vals = row[1:]
        ws.cell(r, 1, year).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
        for c, v in enumerate(vals, 2):
            cell = ws.cell(r, c, v)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Calibri", size=10)
        fill = YELLOW if year == 2026 else (ORANGE if year == 2008 else (ALT if i % 2 else WHITE))
        if year == 2025:
            fill = GREEN
        for c in range(1, 9):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, c).border = THIN

    widths = [10, 22, 20, 14, 16, 18, 20, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    notes = wb.create_sheet("Notas e fontes")
    notes["A1"] = "Notas metodológicas"
    notes["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    notes["A1"].fill = PatternFill("solid", fgColor=NAVY)
    notes.merge_cells("A1:B1")
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 110
    items = [
        ("Unidade", "R$ bilhões, preços correntes (não deflacionados)."),
        ("Período", "Ano civil 2002–2025. 2026 = apenas 1º semestre do Itaú Unibanco (R$ 24,68 bi recorrentes)."),
        ("Itaú vs Itaú Unibanco", "Até 2007 a coluna é o Itaú Holding. Em 2008 a fusão com o Unibanco foi em novembro; o contábil 4,85 não soma o ano inteiro dos dois bancos. A partir de 2009 a série é Itaú Unibanco."),
        ("Unibanco", "Série própria até 2008 (BRGAAP / releases). Depois foi incorporado."),
        ("BTG vs BTG Pactual", "Não são duas instituições paralelas após 2009. Pactual foi ao UBS em 2006; o BTG comprou o Pactual em 2009; IPO em 2012. 2008 = UBS Pactual."),
        ("Santander Brasil", "Perímetro muda em 2008–2009 com a incorporação do Banco Real. Série IFRS da holding listada. Em 2025 o lucro gerencial foi R$ 15,6 bi (contábil IFRS 12,77)."),
        ("Contábil vs recorrente", "O recorrente/gerencial exclui itens extraordinários e é o número que o mercado usa. Não some as duas colunas."),
        ("BTG 2016–2017", "Queda após a crise de 2015 (prisão de André Esteves). 2017 ≈ R$ 2,6 bi; 2016 sem número combinado comparável na mesma base."),
        ("Fontes", "Releases de RI (Itaú, Bradesco, Santander, BTG Pactual); demonstrações IFRS/BRGAAP; VCP Scanner (ITUB, BBD, BSBR); Folha, Valor e G1 para Unibanco e o pro forma de 2008."),
        ("Elaboração", "Série compilada em 16/08/2026. Sujeita a revisão das demonstrações originais."),
    ]
    for i, (k, v) in enumerate(items, start=3):
        notes.cell(i, 1, k).font = Font(name="Calibri", bold=True, color=WHITE)
        notes.cell(i, 1).fill = PatternFill("solid", fgColor=NAVY)
        notes.cell(i, 2, v).alignment = Alignment(wrap_text=True, vertical="center")
        notes.row_dimensions[i].height = 36
        notes.cell(i, 1).border = notes.cell(i, 2).border = THIN

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_pdf(path: Path) -> None:
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A3),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Lucros líquidos dos bancos 2002–2026",
        author="Série compilada a partir de RI / IFRS",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "T",
        parent=styles["Title"],
        fontName="Times-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1B4F72"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "S",
        parent=styles["Normal"],
        fontName="Times-Italic",
        fontSize=9.5,
        leading=12,
        textColor=colors.HexColor("#2C3E50"),
        alignment=TA_JUSTIFY,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "B",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=9,
        leading=12,
        alignment=TA_JUSTIFY,
        spaceAfter=4,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontName="Times-Bold",
        fontSize=12,
        textColor=colors.HexColor("#1B4F72"),
        spaceBefore=8,
        spaceAfter=4,
    )
    cell = ParagraphStyle(
        "C",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
    )
    cell_b = ParagraphStyle("CB", parent=cell, fontName="Times-Bold")
    head = ParagraphStyle(
        "HD",
        parent=cell,
        fontName="Times-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.white,
    )

    story = []
    story.append(Paragraph(
        "Lucros líquidos anuais dos bancos Itaú, Unibanco, Itaú Unibanco, "
        "Bradesco, Santander Brasil e BTG Pactual — 2002 a 2026",
        title,
    ))
    story.append(Paragraph(
        "Valores em <b>R$ bilhões</b>, ano civil, preços correntes. "
        "A coluna Itaú / Itaú Unibanco é o Itaú Holding até 2007 e o Itaú Unibanco a partir de 2008. "
        "Unibanco existe como série própria só até 2008 (fusão em novembro). "
        "BTG e BTG Pactual não são duas franquias paralelas depois de 2009: o BTG comprou o Pactual "
        "do UBS naquele ano. 2026 ainda não fechou — figura só o 1º semestre do Itaú Unibanco.",
        sub,
    ))

    header_row = [Paragraph(h.replace("\n", "<br/>"), head) for h in HEADERS]
    data = [header_row]
    for row in ROWS:
        year = row[0]
        cells = [Paragraph(str(year) if year != 2026 else "2026*", cell_b)]
        for v in row[1:]:
            cells.append(Paragraph(fmt(v), cell))
        data.append(cells)

    col_w = [22 * mm, 38 * mm, 36 * mm, 28 * mm, 32 * mm, 34 * mm, 36 * mm, 30 * mm]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5D8DC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("BACKGROUND", (0, 7), (-1, 7), colors.HexColor("#FDEBD0")),  # 2008
        ("BACKGROUND", (0, 24), (-1, 24), colors.HexColor("#D5F5E3")),  # 2025
        ("BACKGROUND", (0, 25), (-1, 25), colors.HexColor("#FCF3CF")),  # 2026
    ]
    for i in range(1, len(data)):
        if i not in (7, 24, 25) and i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#EBF5FB")))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "* 2026 = 1º semestre (Itaú Unibanco, lucro recorrente gerencial de R$ 24,68 bi). "
        "Linha 2008 em laranja: ano da fusão Itaú–Unibanco; o contábil 4,85 não cobre o ano inteiro "
        "dos dois bancos — o recorrente 10,57 é pro forma. Linha 2025 em verde: último ano civil fechado. "
        "Santander 2025 gerencial = R$ 15,6 bi. BTG Pactual 2025 ajustado = R$ 16,68 bi "
        "(a tabela usa o contábil 15,95).",
        body,
    ))

    story.append(Paragraph("Notas de perímetro e critério", h2))
    story.append(Paragraph(
        "<b>Contábil vs recorrente.</b> O lucro recorrente (ou gerencial) exclui itens extraordinários "
        "e é o número usado nas teleconferências. O contábil/IFRS vem das demonstrações. "
        "As duas colunas não devem ser somadas. Onde o recorrente não aparece, o release da época "
        "não foi compilado na mesma base.",
        body,
    ))
    story.append(Paragraph(
        "<b>Itaú e Unibanco.</b> Até 2007 são bancos distintos. Em 3/11/2008 anunciaram a associação. "
        "O Unibanco deixou de ter série própria. Em 2008, sozinho, o Itaú contribuiu com cerca de "
        "R$ 7,72 bi recorrentes e o Unibanco com R$ 2,85 bi.",
        body,
    ))
    story.append(Paragraph(
        "<b>Santander Brasil.</b> A série muda de tamanho em 2008–2009 com a incorporação do Banco Real "
        "(ABN Amro). Antes disso o perímetro é essencialmente Banespa/Santander. "
        "Não há 2002, 2004 e 2005 em IFRS comparável na mesma base usada aqui.",
        body,
    ))
    story.append(Paragraph(
        "<b>BTG / BTG Pactual.</b> Pactual foi vendido ao UBS em 2006. Em 2009 o grupo BTG recompra "
        "o banco e nasce o BTG Pactual (IPO em 2012). O valor de 2008 (2,65) é UBS Pactual. "
        "2015 (4,62) é o pico pré-crise; 2016–2017 caem após a prisão de André Esteves. "
        "A partir de 2019 a série volta a crescer de forma contínua até 16,7 bi ajustados em 2025.",
        body,
    ))
    story.append(Paragraph(
        "<b>Fontes.</b> Releases de relações com investidores do Itaú Unibanco, Bradesco, "
        "Santander Brasil e BTG Pactual; demonstrações IFRS/BRGAAP; séries ITUB, BBD e BSBR "
        "(VCP Scanner); Folha de S.Paulo, Valor Econômico e G1 para Unibanco e o pro forma de 2008. "
        "Compilação em 16/08/2026, sujeita a revisão das peças originais.",
        body,
    ))

    def footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Times-Roman", 8)
        canvas.setFillColor(colors.HexColor("#5D6D7E"))
        canvas.drawString(
            14 * mm,
            8 * mm,
            "Lucros líquidos em R$ bilhões | 2002–2026 | Compilação a partir de RI / IFRS",
        )
        canvas.drawRightString(
            landscape(A3)[0] - 14 * mm,
            8 * mm,
            f"Página {doc_.page}",
        )
        canvas.restoreState()

    path.parent.mkdir(parents=True, exist_ok=True)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


def main() -> None:
    out = Path("output")
    xlsx = out / "lucros_liquidos_bancos_2002_2026.xlsx"
    pdf = out / "lucros_liquidos_bancos_2002_2026.pdf"
    write_xlsx(xlsx)
    write_pdf(pdf)
    print(f"XLSX: {xlsx.resolve()}")
    print(f"PDF:  {pdf.resolve()}")


if __name__ == "__main__":
    main()
