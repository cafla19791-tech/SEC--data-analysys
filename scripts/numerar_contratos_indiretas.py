#!/usr/bin/env python3
"""Numera contratos de operações indiretas no formato N-AAAA.

Regra (a partir de 2002):
  - na planilha do ano 2002, a 1ª linha → 1-2002, a 2ª → 2-2002, …;
  - na planilha do ano 2003, a 1ª linha → 1-2003, e assim por diante.

Cada ano reinicia a sequência em 1. O número é único no formato ``{seq}-{ano}``.

Uso (ContAgil):
  python scripts/numerar_contratos_indiretas.py
  python scripts/numerar_contratos_indiretas.py --pasta-dados dados --saida saida/BNDES_INDIRETAS_NUMERADOS.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.gerar_fluxos import (  # noqa: E402
    CONTAGIL_PASTA_DADOS,
    CONTAGIL_PASTA_SAIDA,
    CONTAGIL_WINPYTHON,
    _mapear_colunas_contratos,
    parse_datas,
)

COL_NUMERO = "Número do contrato"
ANO_MIN_DEFAULT = 2002
_ANO_NO_NOME = re.compile(r"(19|20)\d{2}")


def ano_do_nome_arquivo(path: Path) -> Optional[int]:
    m = _ANO_NO_NOME.search(path.stem)
    return int(m.group(0)) if m else None


def atribuir_numero_contrato_anual(
    df: pd.DataFrame,
    *,
    col_data: str = "data_contratacao",
    anos: Optional[pd.Series] = None,
) -> pd.Series:
    """Gera série ``1-2002``, ``2-2002``, … reiniciando a cada ano.

    Preserva a ordem das linhas: a 1ª linha de cada ano recebe ``1-AAAA``.
    """
    if len(df) == 0:
        return pd.Series(dtype=str, index=df.index)

    if anos is None:
        if col_data not in df.columns:
            raise ValueError("Informe col_data com datas ou a série anos.")
        anos = parse_datas(df[col_data]).dt.year

    anos_i = pd.Series(anos, index=df.index).astype("Int64")
    if anos_i.isna().any():
        raise ValueError(
            f"Há {int(anos_i.isna().sum())} linha(s) sem ano para numerar o contrato."
        )

    seq = df.groupby(anos_i, sort=False).cumcount() + 1
    return seq.astype(str) + "-" + anos_i.astype(str)


def numerar_sequencial_ano(df: pd.DataFrame, ano: int) -> pd.DataFrame:
    """Numera 1-AAAA … N-AAAA na ordem atual das linhas (um único ano)."""
    out = df.copy()
    n = len(out)
    out["numero_contrato"] = [f"{i}-{int(ano)}" for i in range(1, n + 1)]
    out["_ano_num"] = int(ano)
    return out


def listar_arquivos_indiretas(pasta: Path) -> list[Path]:
    if not pasta.exists():
        return []
    arquivos: list[Path] = []
    for p in sorted(pasta.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        nome = p.name.upper()
        if "INDIRET" in nome or "BNDES" in nome:
            arquivos.append(p)
    if not arquivos:
        arquivos = sorted(
            p for p in pasta.glob("*.xlsx") if not p.name.startswith("~$")
        )
    return arquivos


def _detectar_header(path: Path) -> int:
    for h0 in range(0, 9):
        try:
            df = pd.read_excel(path, header=h0, nrows=3)
        except Exception:
            continue
        mapped, _ = _mapear_colunas_contratos(df)
        cols = {str(c).lower() for c in mapped.columns}
        if "data_contratacao" in cols:
            return h0
        raw_cols = " ".join(str(c).lower() for c in df.columns)
        if "contrat" in raw_cols and ("valor" in raw_cols or "desembolso" in raw_cols):
            return h0
    return 0


def carregar_indiretas_excel(path: Path) -> pd.DataFrame:
    """Carrega Excel com colunas mapeadas + auxiliares ``_data`` / ``_ano``."""
    h0 = _detectar_header(path)
    raw = pd.read_excel(path, header=h0)
    mapped, rename = _mapear_colunas_contratos(raw)
    out = mapped.copy()
    out.attrs["rename"] = rename
    out.attrs["header0"] = h0
    out.attrs["fonte"] = str(path)
    if "data_contratacao" in out.columns:
        out["_data"] = parse_datas(out["data_contratacao"])
        out["_ano"] = out["_data"].dt.year
    else:
        out["_data"] = pd.NaT
        out["_ano"] = pd.NA
    return out


def _colunas_saida(df: pd.DataFrame) -> pd.DataFrame:
    prefer = [
        ("numero_contrato", COL_NUMERO),
        ("cliente", "Cliente"),
        ("cnpj", "CNPJ"),
        ("uf", "UF"),
        ("data_contratacao", "Data da contratação"),
        ("valor_desembolsado", "Valor desembolsado R$"),
        ("valor_contratado", "Valor contratado R$"),
        ("agente", "Instituição Financeira Credenciada"),
        ("custo_financeiro", "Custo financeiro"),
        ("juros", "Juros"),
        ("prazo_carencia", "Prazo - Carência (meses)"),
        ("prazo_amortizacao", "Prazo - Amortização (meses)"),
        ("forma_de_apoio", "Forma de apoio"),
    ]
    data: dict[str, object] = {}
    cols: list[str] = []
    for canon, titulo in prefer:
        if canon in df.columns:
            data[titulo] = df[canon].values
            cols.append(titulo)

    skip = {
        "numero_contrato",
        "_data",
        "_ano",
        "_ano_num",
        *[c for c, _ in prefer],
    }
    for c in df.columns:
        if c in skip or str(c).startswith("_"):
            continue
        titulo = str(c)
        if titulo in data:
            continue
        data[titulo] = df[c].values
        cols.append(titulo)

    if COL_NUMERO not in data and "numero_contrato" in df.columns:
        data = {COL_NUMERO: df["numero_contrato"].values, **data}
        cols = [COL_NUMERO] + [c for c in cols if c != COL_NUMERO]

    return pd.DataFrame(data)[cols] if cols else pd.DataFrame()


def escrever_por_ano(
    partes: dict[int, pd.DataFrame],
    saida: Path,
    *,
    ano_min: int = ANO_MIN_DEFAULT,
) -> Path:
    """Grava um Excel com uma aba por ano (2002, 2003, …)."""
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True
    fill_h = PatternFill("solid", fgColor="1F4E79")
    font_h = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    anos = sorted(a for a in partes if a is not None and int(a) >= ano_min)
    if not anos:
        anos = sorted(int(a) for a in partes if a is not None)

    for ano in anos:
        df = _colunas_saida(partes[ano])
        nome = str(int(ano))[:31]
        if first:
            ws = wb.active
            ws.title = nome
            first = False
        else:
            ws = wb.create_sheet(nome)

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(list(row))

        for col in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(1, col)
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions["A"].width = 14
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions

    if first:
        wb.active.title = "vazio"

    wb.save(saida)
    return saida


def _acumular_por_ano(
    por_ano: dict[int, list[pd.DataFrame]], df: pd.DataFrame, ano_arquivo: Optional[int]
) -> None:
    """Reparte o DataFrame por ano e acumula (ainda sem numeração final)."""
    if df.empty:
        return

    if ano_arquivo is not None:
        anos_presentes = set(df["_ano"].dropna().astype(int).unique()) if "_ano" in df.columns else set()
        if not anos_presentes or anos_presentes == {ano_arquivo}:
            por_ano.setdefault(ano_arquivo, []).append(df.copy())
            return

    if "_ano" not in df.columns or df["_ano"].isna().all():
        if ano_arquivo is None:
            raise ValueError("Sem ano na data nem no nome do arquivo.")
        por_ano.setdefault(ano_arquivo, []).append(df.copy())
        return

    for ano, g in df.groupby(df["_ano"], sort=False):
        if pd.isna(ano):
            if ano_arquivo is not None:
                por_ano.setdefault(ano_arquivo, []).append(g.copy())
            continue
        por_ano.setdefault(int(ano), []).append(g.copy())


def processar_pasta(
    pasta_dados: Path,
    saida: Path,
    *,
    ano_min: int = ANO_MIN_DEFAULT,
) -> tuple[Path, dict[int, int]]:
    arquivos = listar_arquivos_indiretas(pasta_dados)
    if not arquivos:
        raise FileNotFoundError(f"Nenhum Excel de indiretas em: {pasta_dados}")

    por_ano: dict[int, list[pd.DataFrame]] = {}
    for arq in arquivos:
        print(f"[INFO] Lendo {arq.name} ...")
        df = carregar_indiretas_excel(arq)
        _acumular_por_ano(por_ano, df, ano_do_nome_arquivo(arq))

    finais: dict[int, pd.DataFrame] = {}
    contagem: dict[int, int] = {}
    for ano in sorted(por_ano):
        base = pd.concat(por_ano[ano], ignore_index=True)
        base = numerar_sequencial_ano(base, ano)
        finais[ano] = base
        contagem[ano] = len(base)
        print(
            f"[INFO] Ano {ano}: {len(base):,} contratos "
            f"(1-{ano} … {len(base)}-{ano})"
        )

    escrever_por_ano(finais, saida, ano_min=ano_min)
    print(f"[OK] {saida}")
    return saida, contagem


def processar_excel(
    excel: Path,
    saida: Path,
    *,
    ano_min: int = ANO_MIN_DEFAULT,
) -> Path:
    df = carregar_indiretas_excel(excel)
    por_ano: dict[int, list[pd.DataFrame]] = {}
    _acumular_por_ano(por_ano, df, ano_do_nome_arquivo(excel))
    finais = {ano: numerar_sequencial_ano(pd.concat(chunks, ignore_index=True), ano) for ano, chunks in por_ano.items()}
    for ano, g in sorted(finais.items()):
        print(f"[INFO] Ano {ano}: {len(g):,} contratos (1-{ano} … {len(g)}-{ano})")
    escrever_por_ano(finais, saida, ano_min=ano_min)
    print(f"[OK] {saida}")
    return saida


def resolver_pasta_dados(arg: Optional[Path]) -> Path:
    if arg is not None and arg.exists():
        return arg
    for cand in (
        Path.cwd() / "dados",
        CONTAGIL_PASTA_DADOS,
        CONTAGIL_WINPYTHON / "dados",
        ROOT / "data" / "contagil_winpython" / "dados",
    ):
        if cand.exists():
            return cand
    if arg is not None:
        return arg
    return Path.cwd() / "dados"


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--pasta-dados", type=Path, default=None)
    p.add_argument("--saida", type=Path, default=None)
    p.add_argument("--ano-min", type=int, default=ANO_MIN_DEFAULT)
    p.add_argument("--excel", type=Path, default=None, help="Um único Excel")
    args = p.parse_args(argv)

    if args.saida is None:
        base_saida = (
            CONTAGIL_PASTA_SAIDA if CONTAGIL_PASTA_SAIDA.exists() else Path.cwd() / "saida"
        )
        if not base_saida.exists():
            base_saida = ROOT / "output"
        args.saida = base_saida / "BNDES_INDIRETAS_NUMERADOS.xlsx"

    if args.excel is not None:
        processar_excel(args.excel, args.saida, ano_min=args.ano_min)
        return 0

    pasta = resolver_pasta_dados(args.pasta_dados)
    print(f"[INFO] Pasta de dados: {pasta}")
    processar_pasta(pasta, args.saida, ano_min=args.ano_min)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
