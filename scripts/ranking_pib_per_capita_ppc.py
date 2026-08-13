#!/usr/bin/env python3
"""Gera ranking da variação do PIB per capita em PPC (dólares internacionais
constantes de 2021) para vários períodos, a partir da API do Banco Mundial.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INDICATOR = "NY.GDP.PCAP.PP.KD"
INDICATOR_NAME = (
    "PIB per capita, PPC (dólares internacionais constantes de 2021)"
)
PERIODS = [
    (1995, 2002),
    (2003, 2016),
    (2003, 2018),
    (2019, 2022),
    (2023, 2025),
]
HIGHLIGHT = {
    "BRA": "Brasil",
    "CHN": "China",
    "USA": "Estados Unidos",
    "IND": "Índia",
}

# Cores
NAVY = "1B4F72"
NAVY2 = "154360"
GOLD = "B7950B"
WHITE = "FFFFFF"
LIGHT = "F4F6F7"
ALT = "EBF5FB"
GREEN = "D5F5E3"
RED = "FADBD8"
YELLOW = "FCF3CF"
ORANGE = "FDEBD0"
WORLD_BG = "D6EAF8"
THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def fetch_json(url: str) -> object:
    with urlopen(url, timeout=90) as resp:
        return json.loads(resp.read().decode("utf-8"))


def load_or_download(path: Path, url: str) -> object:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    data = fetch_json(url)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data), encoding="utf-8")
    return data


def country_map(raw: list) -> tuple[dict, set]:
    info, aggregates = {}, set()
    countries = raw[1] if isinstance(raw, list) and len(raw) > 1 else raw
    for c in countries:
        iso3 = c.get("id")
        region = ((c.get("region") or {}).get("value") or "").strip()
        if not iso3:
            continue
        if region == "Aggregates" or not region:
            aggregates.add(iso3)
            continue
        info[iso3] = {
            "name": c.get("name") or iso3,
            "region": region,
            "income": ((c.get("incomeLevel") or {}).get("value") or ""),
        }
    return info, aggregates


def series_map(raw: list) -> tuple[dict, str]:
    meta, rows = raw[0], raw[1]
    vals: dict[str, dict[int, float]] = {}
    for r in rows:
        iso3 = r.get("countryiso3code") or ""
        year = r.get("date")
        val = r.get("value")
        if iso3 and year and val is not None:
            vals.setdefault(iso3, {})[int(year)] = float(val)
    return vals, meta.get("lastupdated") or ""


def rank_period(vals: dict, info: dict, y0: int, y1: int) -> list[dict]:
    recs = []
    n = y1 - y0
    for iso3, meta in info.items():
        d = vals.get(iso3, {})
        a, b = d.get(y0), d.get(y1)
        if a is None or b is None or a <= 0:
            continue
        recs.append(
            {
                "iso3": iso3,
                "name": meta["name"],
                "region": meta["region"],
                "income": meta["income"],
                "start": a,
                "end": b,
                "abs": b - a,
                "pct": (b / a - 1) * 100,
                "cagr": ((b / a) ** (1 / n) - 1) * 100,
                "years": n,
            }
        )
    recs.sort(key=lambda x: x["pct"], reverse=True)
    for i, r in enumerate(recs, 1):
        r["rank"] = i
    return recs


def world_stats(vals: dict, y0: int, y1: int) -> dict | None:
    d = vals.get("WLD", {})
    a, b = d.get(y0), d.get(y1)
    if a is None or b is None or a <= 0:
        return None
    n = y1 - y0
    return {
        "start": a,
        "end": b,
        "abs": b - a,
        "pct": (b / a - 1) * 100,
        "cagr": ((b / a) ** (1 / n) - 1) * 100,
        "years": n,
    }


def style_title(ws, row, text, size=16, color=WHITE, fill=NAVY):
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Calibri", size=size, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def fill_range(ws, row, cols, color):
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = PatternFill("solid", fgColor=color)


def write_metodologia(ws, lastupdated: str, counts: dict):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")
    style_title(ws, 1, "Metodologia e notas")
    fill_range(ws, 1, 2, NAVY)
    ws.row_dimensions[1].height = 32

    lines = [
        ("Indicador", INDICATOR_NAME),
        ("Código Banco Mundial", INDICATOR),
        ("Fonte", "World Development Indicators — The World Bank"),
        ("Atualização da base", lastupdated),
        ("Unidade", "Dólares internacionais constantes de 2021"),
        (
            "O que mede",
            "Produto interno bruto por habitante ajustado pela paridade "
            "do poder de compra, em preços constantes de 2021. A variação "
            "reflete crescimento real do poder de compra médio, sem inflação.",
        ),
        (
            "Por que constantes de 2021",
            "A série corrente (NY.GDP.PCAP.PP.CD) mistura volume, inflação "
            "e mudança das paridades. Esta série (NY.GDP.PCAP.PP.KD) é a "
            "adequada para comparar evolução no tempo.",
        ),
        (
            "Universo",
            "Somente economias nacionais. Agregados regionais e faixas de "
            "renda do Banco Mundial foram excluídos. O Mundo (WLD) aparece "
            "como referência em cada aba, sem ocupar posição no ranking.",
        ),
        (
            "Critério de inclusão",
            "O país entra no ranking do período se houver valor positivo "
            "no ano inicial e valor no ano final.",
        ),
        (
            "Variação percentual",
            "((valor final / valor inicial) − 1) × 100. Compara o ano civil "
            "inicial com o ano civil final (não é estoque de 1º de janeiro "
            "nem de 31 de dezembro).",
        ),
        (
            "CAGR",
            "((valor final / valor inicial) ^ (1 / n) − 1) × 100, em que "
            "n = ano final − ano inicial.",
        ),
        (
            "Períodos",
            "1995–2002 (7 anos); 2003–2016 (13 anos); 2003–2018 (15 anos); "
            "2019–2022 (3 anos); 2023–2025 (2 anos).",
        ),
        (
            "Países por período",
            "; ".join(f"{a}–{b}: {counts[(a, b)]} países" for a, b in PERIODS),
        ),
        (
            "Destaques nas abas",
            "Brasil em amarelo; China em laranja. Topo e base do ranking "
            "recebem fundo verde e vermelho claros. A coluna de variação "
            "percentual tem escala de cores.",
        ),
        (
            "Limitações",
            "Valores de 2024–2025 podem ser estimativas e sujeitos a "
            "revisão. Países sem dado em um dos extremos do intervalo "
            "não entram naquele ranking. A série aplica preços relativos "
            "de 2021 a todos os anos (problema clássico de número-índice).",
        ),
    ]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    body_font = Font(name="Calibri", size=11)
    for i, (k, v) in enumerate(lines, start=3):
        ws.row_dimensions[i].height = 36 if len(v) > 80 else 22
        a, b = ws.cell(i, 1, k), ws.cell(i, 2, v)
        a.font = header_font
        a.fill = PatternFill("solid", fgColor=NAVY2 if i % 2 else NAVY)
        a.alignment = Alignment(vertical="center")
        b.font = body_font
        b.alignment = Alignment(wrap_text=True, vertical="center")
        b.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
        a.border = b.border = THIN


def write_ranking(ws, y0: int, y1: int, recs: list[dict], world: dict | None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:K{6 + len(recs)}"
    n = y1 - y0
    title = (
        f"Ranking da variação do PIB per capita em PPC — {y0} a {y1}"
    )
    ws.merge_cells("A1:K1")
    style_title(ws, 1, title, size=15)
    fill_range(ws, 1, 11, NAVY)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:K2")
    sub = (
        f"{INDICATOR_NAME}  |  {n} anos  |  {len(recs)} países  |  "
        f"Fonte: Banco Mundial ({INDICATOR})"
    )
    ws.cell(2, 1, sub).font = Font(name="Calibri", italic=True, size=10, color="2C3E50")
    fill_range(ws, 2, 11, "D4E6F1")

    if world:
        wr = sum(1 for r in recs if r["pct"] > world["pct"]) + 1
        ws.merge_cells("A3:K3")
        ws.cell(
            3,
            1,
            (
                f"Mundo (referência, fora do ranking): "
                f"{world['start']:,.1f} → {world['end']:,.1f}  |  "
                f"variação {world['pct']:+.2f}%  |  "
                f"CAGR {world['cagr']:+.2f}% a.a.  |  "
                f"equivaleria à {wr}ª posição entre {len(recs)} países"
            ),
        ).font = Font(name="Calibri", bold=True, size=10, color="1A5276")
        fill_range(ws, 3, 11, WORLD_BG)

    ws.merge_cells("A4:K4")
    ws.cell(
        4,
        1,
        "Brasil destacado em amarelo; China em laranja. "
        "Valores em dólares internacionais constantes de 2021.",
    ).font = Font(name="Calibri", size=9, color="5D6D7E")

    headers = [
        "Rank",
        "País",
        "ISO3",
        "Região",
        "Faixa de renda",
        f"PIB pc {y0}",
        f"PIB pc {y1}",
        "Variação absoluta",
        "Variação %",
        "CAGR % a.a.",
        "Vs. mundo",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(6, col, h)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[6].height = 28

    widths = [8, 34, 10, 32, 20, 14, 14, 18, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    world_pct = world["pct"] if world else None
    for i, r in enumerate(recs):
        row = 7 + i
        vs = ""
        if world_pct is not None:
            vs = "Acima" if r["pct"] > world_pct else (
                "Igual" if abs(r["pct"] - world_pct) < 1e-9 else "Abaixo"
            )
        values = [
            r["rank"],
            r["name"],
            r["iso3"],
            r["region"],
            r["income"],
            r["start"],
            r["end"],
            r["abs"],
            r["pct"] / 100,
            r["cagr"] / 100,
            vs,
        ]
        if r["iso3"] == "BRA":
            fill = YELLOW
        elif r["iso3"] == "CHN":
            fill = ORANGE
        elif r["rank"] <= 10:
            fill = GREEN
        elif r["rank"] > len(recs) - 10:
            fill = RED
        elif i % 2:
            fill = ALT
        else:
            fill = WHITE

        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            cell.font = Font(name="Calibri", size=10, bold=(r["iso3"] in HIGHLIGHT))
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if col in (6, 7, 8):
                cell.number_format = "#,##0.0"
            elif col in (9, 10):
                cell.number_format = "0.00%"
            if col in (1, 3, 11):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    last = 6 + len(recs)
    ws.conditional_formatting.add(
        f"I7:I{last}",
        ColorScaleRule(
            start_type="min",
            start_color="E74C3C",
            mid_type="percentile",
            mid_value=50,
            mid_color="F7F9F9",
            end_type="max",
            end_color="1E8449",
        ),
    )
    ws.auto_filter.ref = f"A6:K{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddHeader.left.text = f"Ranking PIB per capita PPC 2021 — {y0}-{y1}"
    ws.print_title_rows = "1:6"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def write_resumo(ws, rankings: dict, worlds: dict, lastupdated: str):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L1")
    style_title(ws, 1, "Resumo comparativo dos períodos")
    fill_range(ws, 1, 12, NAVY)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:L2")
    ws.cell(
        2,
        1,
        f"{INDICATOR_NAME}  |  Atualização: {lastupdated}",
    ).font = Font(name="Calibri", italic=True, size=10)
    fill_range(ws, 2, 12, "D4E6F1")

    headers = [
        "Período",
        "Anos",
        "Nº países",
        "Mundo var. %",
        "Mundo CAGR %",
        "Brasil rank",
        "Brasil var. %",
        "Brasil CAGR %",
        "China rank",
        "China var. %",
        "China CAGR %",
        "EUA rank",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[4].height = 30

    def find(recs, iso3):
        for r in recs:
            if r["iso3"] == iso3:
                return r
        return None

    for i, (y0, y1) in enumerate(PERIODS):
        recs = rankings[(y0, y1)]
        w = worlds[(y0, y1)]
        br, cn, us = find(recs, "BRA"), find(recs, "CHN"), find(recs, "USA")
        row = 5 + i
        vals = [
            f"{y0}–{y1}",
            y1 - y0,
            len(recs),
            (w["pct"] / 100) if w else None,
            (w["cagr"] / 100) if w else None,
            br["rank"] if br else None,
            (br["pct"] / 100) if br else None,
            (br["cagr"] / 100) if br else None,
            cn["rank"] if cn else None,
            (cn["pct"] / 100) if cn else None,
            (cn["cagr"] / 100) if cn else None,
            us["rank"] if us else None,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.font = Font(name="Calibri", size=11)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor=YELLOW if col in (6, 7, 8) else (ORANGE if col in (9, 10, 11) else (ALT if i % 2 else WHITE)))
            if col in (4, 5, 7, 8, 10, 11) and isinstance(val, float):
                cell.number_format = "0.00%"

    widths = [14, 8, 12, 14, 14, 13, 14, 14, 12, 14, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A12:E12")
    style_title(ws, 12, "Top 10 de cada período (variação %)", size=13)
    fill_range(ws, 12, 5, NAVY2)

    row = 14
    for y0, y1 in PERIODS:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        style_title(ws, row, f"Top 10 — {y0}–{y1}", size=12, fill=NAVY2)
        fill_range(ws, row, 5, NAVY2)
        row += 1
        for col, h in enumerate(["Rank", "País", "ISO3", "Variação %", "CAGR % a.a."], 1):
            cell = ws.cell(row, col, h)
            cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN
        row += 1
        for r in recs[:10]:
            vals = [r["rank"], r["name"], r["iso3"], r["pct"] / 100, r["cagr"] / 100]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row, col, val)
                cell.font = Font(name="Calibri", size=10, bold=(r["iso3"] in HIGHLIGHT))
                cell.border = THIN
                cell.fill = PatternFill(
                    "solid",
                    fgColor=YELLOW if r["iso3"] == "BRA" else (ORANGE if r["iso3"] == "CHN" else GREEN),
                )
                if col in (4, 5):
                    cell.number_format = "0.00%"
            row += 1
        row += 2


def build_workbook(vals, info, lastupdated: str, dest: Path) -> dict:
    rankings = {}
    worlds = {}
    counts = {}
    for y0, y1 in PERIODS:
        recs = rank_period(vals, info, y0, y1)
        rankings[(y0, y1)] = recs
        worlds[(y0, y1)] = world_stats(vals, y0, y1)
        counts[(y0, y1)] = len(recs)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Metodologia"
    write_metodologia(ws0, lastupdated, counts)

    ws_r = wb.create_sheet("Resumo", 1)
    write_resumo(ws_r, rankings, worlds, lastupdated)

    for y0, y1 in PERIODS:
        ws = wb.create_sheet(f"{y0}-{y1}")
        write_ranking(ws, y0, y1, rankings[(y0, y1)], worlds[(y0, y1)])

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return {"rankings": rankings, "worlds": worlds, "counts": counts}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        default="output/ranking_pib_per_capita_ppc_2021.xlsx",
        help="Caminho do arquivo Excel de saída",
    )
    parser.add_argument(
        "--cache-dir",
        default="data/worldbank",
        help="Pasta para cache JSON da API do Banco Mundial",
    )
    args = parser.parse_args()

    cache = Path(args.cache_dir)
    countries_raw = load_or_download(
        cache / "countries.json",
        "https://api.worldbank.org/v2/country?format=json&per_page=400",
    )
    pcap_raw = load_or_download(
        cache / "ny_gdp_pcap_pp_kd_1995_2025.json",
        (
            "https://api.worldbank.org/v2/country/all/indicator/"
            f"{INDICATOR}?date=1995:2025&format=json&per_page=20000"
        ),
    )
    info, _ = country_map(countries_raw)
    vals, lastupdated = series_map(pcap_raw)
    dest = Path(args.output)
    summary = build_workbook(vals, info, lastupdated, dest)
    print(f"Arquivo gravado: {dest.resolve()}")
    print(f"Atualização da base: {lastupdated}")
    for (y0, y1), n in summary["counts"].items():
        w = summary["worlds"][(y0, y1)]
        br = next((r for r in summary["rankings"][(y0, y1)] if r["iso3"] == "BRA"), None)
        print(
            f"  {y0}-{y1}: {n} países"
            + (f" | mundo {w['pct']:+.2f}%" if w else "")
            + (f" | Brasil #{br['rank']} {br['pct']:+.2f}%" if br else "")
        )


if __name__ == "__main__":
    main()
