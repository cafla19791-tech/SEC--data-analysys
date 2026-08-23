"""Resumo explicativo das atas do Copom (2001–hoje).

Fonte: Banco Central do Brasil
  https://www.bcb.gov.br/publicacoes/atascopom
  API: /api/servico/sitebcb/atascopom/ultimas e /principal
  Selic meta: SGS 432

Uso:
  python3 scripts/resumo_atas_copom.py
  python3 scripts/resumo_atas_copom.py --sem-download
"""

from __future__ import annotations

import argparse
import io
import re
import sys
import time
from datetime import datetime
from html import unescape
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import matplotlib.pyplot as plt
import pandas as pd
import requests
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None  # type: ignore

BCB = "https://www.bcb.gov.br"
API_LISTA = f"{BCB}/api/servico/sitebcb/atascopom/ultimas"
API_ATA = f"{BCB}/api/servico/sitebcb/atascopom/principal"
BCB_SGS_SOAP = "https://www3.bcb.gov.br/wssgs/services/FachadaWSSGS"
HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "User-Agent": "Mozilla/5.0",
    "Referer": f"{BCB}/publicacoes/atascopom",
}

ANO_INICIO = 2001
ANO_FIM = 2026
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"

ESTILO_TABELA = (
    "border-collapse:collapse;border-spacing:0;"
    "border:1.5px solid #1a1a1a;width:100%;"
)
ESTILO_CELULA = (
    "border:1px solid #1a1a1a;padding:6px 8px;"
    "font-family:Calibri,Arial,sans-serif;font-size:13px;"
)
ESTILO_TH = ESTILO_CELULA + "background:#1f4e79;color:#fff;font-weight:700;"
BORDA_CONTINUA = Border(
    left=Side(style="thin", color="1A1A1A"),
    right=Side(style="thin", color="1A1A1A"),
    top=Side(style="thin", color="1A1A1A"),
    bottom=Side(style="thin", color="1A1A1A"),
)

TEMAS = {
    "inflacao": r"infla[cç][aã]o|IPCA|metas? de infla",
    "cambio": r"c[aâ]mbio|desvaloriza|real\b|d[oó]lar",
    "fiscal": r"fiscal|prim[aá]rio|teto de gastos|arcabou[cç]o",
    "atividade": r"atividade|hiato|desemprego|PIB",
    "credito": r"cr[eé]dito|endividamento|inadimpl",
    "externo": r"externo|Fed\b|geopol|guerra|commodit",
    "expectativas": r"expectativa|ancoragem|desancor",
    "guidance": r"forward guidance|comunicado|sinaliza[cç]|guidance",
    "credibilidade": r"credibilidade|reputa[cç][aã]o",
    "pandemia": r"pandemia|covid|coronav[ií]rus",
}

CICLOS = [
    (2001, 2002, "Consolidação do regime e crise de 2002"),
    (2003, 2005, "Primeiro recuo e reversão (16% → 19,75%)"),
    (2006, 2008, "Cortes longos e reversão de 2008"),
    (2009, 2010, "Crise: corte a 8,75% e início da volta"),
    (2011, 2014, "Nova matriz: recuo a 7,25% e volta forçada"),
    (2015, 2016, "Recessão e aperto até 14,25%"),
    (2017, 2019, "Ciclo de cortes até a mínima histórica"),
    (2020, 2021, "Pandemia, piso de 2% e início do aperto"),
    (2022, 2023, "Selic em 13,75% e início do afrouxamento"),
    (2024, 2026, "Cortes, retomada do aperto e patamar alto"),
]


def tabela_html(cabecalhos: list[str], linhas: list[list[str]], aligns: list[str] | None = None) -> str:
    if aligns is None:
        aligns = ["center"] + ["right"] * (len(cabecalhos) - 1)
    ths = "".join(f'<th style="{ESTILO_TH};text-align:center">{h}</th>' for h in cabecalhos)
    body = []
    for row in linhas:
        tds = []
        for i, cell in enumerate(row):
            a = aligns[i] if i < len(aligns) else "right"
            tds.append(f'<td style="{ESTILO_CELULA};text-align:{a}">{cell}</td>')
        body.append("<tr>" + "".join(tds) + "</tr>")
    return f'<table style="{ESTILO_TABELA}"><thead><tr>{ths}</tr></thead><tbody>{"".join(body)}</tbody></table>'


def _fmt(valor: float | None, casas: int = 2) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    n = float(valor)
    txt = f"{n:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return txt


def _fmt_signed(valor: float | None, casas: int = 2) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    n = float(valor)
    if abs(n) < 5e-3:
        return "0,00"
    sinal = "+" if n > 0 else "−"
    return sinal + _fmt(abs(n), casas)


def html_para_texto(html: str) -> str:
    texto = unescape(html or "")
    texto = re.sub(r"(?is)<script.*?>.*?</script>", " ", texto)
    texto = re.sub(r"(?is)<style.*?>.*?</style>", " ", texto)
    texto = re.sub(r"(?i)<br\s*/?>", "\n", texto)
    texto = re.sub(r"(?i)</p>", "\n", texto)
    texto = re.sub(r"(?i)</div>", "\n", texto)
    texto = re.sub(r"(?i)</h[1-6]>", "\n", texto)
    texto = re.sub(r"<[^>]+>", " ", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    texto = re.sub(r"\n{3,}", "\n\n", texto)
    return texto.strip()


def numero_reuniao(titulo: str) -> int | None:
    m = re.search(r"(\d+)\s*ª", titulo or "")
    return int(m.group(1)) if m else None


def identificador(link: str) -> str:
    return (link or "").rstrip("/").split("/")[-1]


def _contar_nomes_voto(bloco: str) -> int:
    bloco = re.sub(r"\([^)]*\)", " ", bloco or "")
    partes = re.split(r",|\se\s", bloco)
    return sum(1 for p in partes if len(re.sub(r"[^A-Za-zÁÉÍÓÚÂÊÔÃÕáéíóúâêôãõçÇ]", "", p)) >= 6)


_NUM_PT = {
    "um": "1",
    "uma": "1",
    "dois": "2",
    "duas": "2",
    "três": "3",
    "tres": "3",
    "quatro": "4",
    "cinco": "5",
    "seis": "6",
    "sete": "7",
    "oito": "8",
    "nove": "9",
}


def _num_voto(token: str) -> str:
    t = (token or "").strip().lower()
    return _NUM_PT.get(t, t if t.isdigit() else "")


def extrair_voto(texto: str) -> str:
    t = texto or ""
    m = re.search(r"por\s+(\d+)\s+votos?\s+a\s+(\d+)", t, flags=re.I)
    if m:
        return f"{m.group(1)} a {m.group(2)}"
    m = re.search(
        r"por\s+(\d+|um|uma|dois|duas|tr[eê]s|quatro|cinco|seis|sete|oito|nove)\s+votos?\s+"
        r"a\s+favor\s+e\s+(\d+|um|uma|dois|duas|tr[eê]s|quatro|cinco|seis|sete|oito|nove)\s+contra",
        t,
        flags=re.I,
    )
    if m:
        a, b = _num_voto(m.group(1)), _num_voto(m.group(2))
        if a and b:
            return f"{a} a {b}"
    m = re.search(r"(\d+)\s+votos?\s+a\s+favor[^\n]{0,40}(\d+)\s+voto", t, flags=re.I)
    if m:
        return f"{m.group(1)} a {m.group(2)}"
    listas = re.findall(
        r"(?:Votaram por|membros votaram por)\s+[^:]{0,140}:\s*([A-ZÁÉÍÓÚÂÊÔÃÕ][^.]{12,500})",
        t,
        flags=re.I,
    )
    if len(listas) >= 2:
        contagens = sorted((_contar_nomes_voto(x) for x in listas), reverse=True)
        if contagens[1] > 0:
            return f"{contagens[0]} a {contagens[1]}"
    if re.search(
        r"por unanimidade|decidiu,?\s+unanimemente|decis[aã]o foi (tomada )?por unanimidade",
        t,
        flags=re.I,
    ):
        return "unânime"
    if listas:
        return "unânime"
    if re.search(r"unanimidade", t, flags=re.I):
        return "unânime"
    return "—"


def classificar_decisao(delta: float | None) -> str:
    if delta is None or pd.isna(delta):
        return "—"
    if delta > 0.04:
        return "alta"
    if delta < -0.04:
        return "corte"
    return "manutenção"


def trecho_decisao(texto: str, limite: int = 520) -> str:
    t = texto or ""
    marcas = [
        r"Diante disso[^\n]{0,40}o Copom decidiu",
        r"o Copom decidiu[^\n]{0,40}(reduzir|elevar|aumentar|manter|fixar)",
        r"Decis[aã]o de pol[ií]tica monet[aá]ria",
        r"O Copom (?:decidiu|avalia)",
        r"O Comit[eê] (?:decidiu|avalia)",
    ]
    pos = -1
    for marca in marcas:
        m = re.search(marca, t, flags=re.I)
        if m:
            pos = m.start()
            break
    bloco = t[pos:] if pos >= 0 else t
    bloco = re.sub(r"\s+", " ", bloco).strip()
    if len(bloco) > limite:
        bloco = bloco[: limite - 1].rsplit(" ", 1)[0] + "…"
    return bloco


def contar_temas(texto: str) -> dict[str, int]:
    t = texto or ""
    return {nome: len(re.findall(pat, t, flags=re.I)) for nome, pat in TEMAS.items()}


def _sessao() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def listar_atas(sessao: requests.Session | None = None) -> pd.DataFrame:
    sessao = sessao or _sessao()
    resp = sessao.get(API_LISTA, params={"quantidade": 1000, "filtro": ""}, timeout=90)
    resp.raise_for_status()
    itens = resp.json()["conteudo"]
    linhas = []
    for item in itens:
        data = pd.to_datetime(item.get("DataReferencia"), utc=True, errors="coerce")
        if pd.isna(data):
            continue
        data = data.tz_convert("America/Sao_Paulo").tz_localize(None)
        if data.year < ANO_INICIO:
            continue
        linhas.append(
            {
                "reuniao": numero_reuniao(item.get("Titulo") or ""),
                "data": data.normalize(),
                "titulo": item.get("Titulo"),
                "url_pdf": item.get("Url"),
                "link": item.get("LinkPagina"),
                "ident": identificador(item.get("LinkPagina") or ""),
            }
        )
    out = pd.DataFrame(linhas).dropna(subset=["data"]).sort_values("data")
    return out.drop_duplicates(subset=["ident"]).reset_index(drop=True)


def baixar_html_ata(ident: str, sessao: requests.Session | None = None) -> str:
    sessao = sessao or _sessao()
    resp = sessao.get(API_ATA, params={"filtro": f"IdentificadorUrl eq '{ident}'"}, timeout=90)
    resp.raise_for_status()
    conteudo = resp.json().get("conteudo") or []
    if not conteudo:
        return ""
    return html_para_texto(conteudo[0].get("OutrasInformacoes") or "")


def baixar_pdf_ata(url_rel: str, sessao: requests.Session | None = None) -> str:
    if not url_rel or PdfReader is None:
        return ""
    sessao = sessao or _sessao()
    url = url_rel if url_rel.startswith("http") else BCB + url_rel
    resp = sessao.get(url, timeout=120)
    resp.raise_for_status()
    reader = PdfReader(io.BytesIO(resp.content))
    partes = []
    for pagina in reader.pages:
        partes.append(pagina.extract_text() or "")
    return re.sub(r"[ \t]+", " ", "\n".join(partes)).strip()


def carregar_textos(
    catalogo: pd.DataFrame,
    cache_dir: Path,
    baixar: bool = True,
) -> pd.DataFrame:
    cache = cache_dir / "copom_atas_textos.csv"
    ja = {}
    if cache.exists():
        antigo = pd.read_csv(cache)
        ja = {str(r["ident"]): r["texto"] for r in antigo.to_dict("records") if r.get("texto")}
    sessao = _sessao()
    linhas = []
    for rec in catalogo.to_dict("records"):
        ident = str(rec["ident"])
        texto = ja.get(ident, "")
        if not texto and baixar:
            print(f"  ata {rec['reuniao']} ({ident})...", flush=True)
            texto = baixar_html_ata(ident, sessao)
            if len(texto) < 400 and rec.get("url_pdf"):
                try:
                    texto = baixar_pdf_ata(rec["url_pdf"], sessao)
                except Exception as exc:  # noqa: BLE001
                    print(f"    falha PDF {ident}: {exc}", flush=True)
            time.sleep(0.08)
        linhas.append({**rec, "texto": texto or "", "n_caracteres": len(texto or "")})
        if baixar and len(linhas) % 10 == 0:
            cache_dir.mkdir(parents=True, exist_ok=True)
            pd.DataFrame(linhas)[["ident", "reuniao", "data", "titulo", "n_caracteres", "texto"]].to_csv(
                cache, index=False
            )
    out = pd.DataFrame(linhas)
    cache_dir.mkdir(parents=True, exist_ok=True)
    out[["ident", "reuniao", "data", "titulo", "n_caracteres", "texto"]].to_csv(cache, index=False)
    return out


def _baixar_sgs_soap(cod: int, inicio: str, fim: str) -> pd.DataFrame:
    corpo = f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:xsd="http://www.w3.org/2001/XMLSchema"
                  xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <soapenv:Body>
    <ns1:getValoresSeriesXML soapenv:encodingStyle="http://schemas.xmlsoap.org/soap/encoding/"
        xmlns:ns1="https://www3.bcb.gov.br/wssgs/services/FachadaWSSGS">
      <in0 xmlns:soapenc="http://schemas.xmlsoap.org/soap/encoding/" soapenc:arrayType="xsd:long[1]">
        <item xsi:type="xsd:long">{int(cod)}</item>
      </in0>
      <in1 xsi:type="xsd:string">{inicio}</in1>
      <in2 xsi:type="xsd:string">{fim}</in2>
    </ns1:getValoresSeriesXML>
  </soapenv:Body>
</soapenv:Envelope>"""
    resp = requests.post(
        BCB_SGS_SOAP,
        data=corpo.encode("utf-8"),
        headers={"Content-Type": "text/xml; charset=utf-8", "SOAPAction": ""},
        timeout=120,
    )
    resp.raise_for_status()
    itens = re.findall(
        r"<ITEM>\s*<DATA>([^<]+)</DATA>\s*<VALOR>([^<]*)</VALOR>",
        unescape(resp.text),
        flags=re.I,
    )
    linhas = []
    for data_txt, valor_txt in itens:
        try:
            dt = datetime.strptime(data_txt.strip(), "%d/%m/%Y")
        except ValueError:
            try:
                dt = datetime.strptime(data_txt.strip(), "%m/%Y")
            except ValueError:
                continue
        try:
            valor = float(valor_txt.replace(",", "."))
        except ValueError:
            continue
        linhas.append({"data": pd.Timestamp(dt), "selic": valor})
    if not linhas:
        return pd.DataFrame(columns=["data", "selic"])
    return pd.DataFrame(linhas).drop_duplicates("data").sort_values("data")


def carregar_selic(cache_dir: Path, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / "sgs_432_selic_meta.csv"
    if cache.exists():
        return pd.read_csv(cache, parse_dates=["data"])
    if not baixar:
        raise FileNotFoundError(cache)
    print("Baixando SGS 432 (meta Selic do Copom)...", flush=True)
    try:
        df = _baixar_sgs_soap(432, "01/12/2000", datetime.now().strftime("%d/%m/%Y"))
    except Exception:
        df = pd.DataFrame(columns=["data", "selic"])
        cursor = pd.Timestamp("2000-12-01")
        fim = pd.Timestamp.now()
        while cursor <= fim:
            bloco = min(cursor + pd.DateOffset(years=1) - pd.DateOffset(days=1), fim)
            parte = _baixar_sgs_soap(432, cursor.strftime("%d/%m/%Y"), bloco.strftime("%d/%m/%Y"))
            if not parte.empty:
                df = pd.concat([df, parte], ignore_index=True)
            cursor = bloco + pd.DateOffset(days=1)
            time.sleep(0.08)
    cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache, index=False)
    return df


def agregar_reunioes(atas: pd.DataFrame, selic: pd.DataFrame) -> pd.DataFrame:
    out = atas.copy()
    out["data"] = pd.to_datetime(out["data"])
    sel = selic.copy()
    sel["data"] = pd.to_datetime(sel["data"])
    sel = sel.sort_values("data")
    # A meta nova entra no SGS no dia seguinte ao segundo dia da reunião.
    antes = pd.merge_asof(
        out[["data"]].sort_values("data"),
        sel.rename(columns={"selic": "selic_antes"}),
        on="data",
        direction="backward",
    )
    depois = pd.merge_asof(
        out.assign(data_depois=out["data"] + pd.Timedelta(days=1))
        .sort_values("data_depois")[["data", "data_depois"]],
        sel.rename(columns={"selic": "selic", "data": "data_depois"}),
        on="data_depois",
        direction="backward",
    )
    out = out.sort_values("data").reset_index(drop=True)
    out["selic_antes"] = antes["selic_antes"].to_numpy()
    out["selic"] = depois["selic"].to_numpy()
    out["delta_selic"] = out["selic"] - out["selic_antes"]
    out["decisao"] = out["delta_selic"].map(classificar_decisao)
    out["voto"] = out["texto"].map(extrair_voto)
    out["trecho"] = out["texto"].map(lambda t: trecho_decisao(t, 420))
    temas = out["texto"].map(contar_temas).apply(pd.Series)
    out = pd.concat([out, temas], axis=1)
    out["ano"] = out["data"].dt.year
    return out.reset_index(drop=True)


def resumo_anual(reunioes: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for ano, bloco in reunioes.groupby("ano"):
        linhas.append(
            {
                "ano": int(ano),
                "n": int(len(bloco)),
                "selic_ini": float(bloco["selic"].dropna().iloc[0]) if bloco["selic"].notna().any() else None,
                "selic_fim": float(bloco["selic"].dropna().iloc[-1]) if bloco["selic"].notna().any() else None,
                "altas": int((bloco["decisao"] == "alta").sum()),
                "cortes": int((bloco["decisao"] == "corte").sum()),
                "manutencoes": int((bloco["decisao"] == "manutenção").sum()),
                "unanimes": int((bloco["voto"] == "unânime").sum()),
                "tema_top": max(TEMAS, key=lambda k: int(bloco[k].sum())) if not bloco.empty else "",
            }
        )
    out = pd.DataFrame(linhas)
    out["delta"] = out["selic_fim"] - out["selic_ini"]
    return out


def resumo_ciclos(reunioes: pd.DataFrame) -> list[dict]:
    linhas = []
    for ini, fim, rotulo in CICLOS:
        bloco = reunioes[(reunioes["ano"] >= ini) & (reunioes["ano"] <= fim)]
        if bloco.empty:
            continue
        sel = bloco["selic"].dropna()
        linhas.append(
            {
                "periodo": f"{ini}–{fim}",
                "rotulo": rotulo,
                "n": int(len(bloco)),
                "selic_ini": float(sel.iloc[0]) if not sel.empty else None,
                "selic_fim": float(sel.iloc[-1]) if not sel.empty else None,
                "selic_max": float(sel.max()) if not sel.empty else None,
                "selic_min": float(sel.min()) if not sel.empty else None,
                "altas": int((bloco["decisao"] == "alta").sum()),
                "cortes": int((bloco["decisao"] == "corte").sum()),
                "manutencoes": int((bloco["decisao"] == "manutenção").sum()),
                "trecho": trecho_decisao(" ".join(bloco["trecho"].dropna().astype(str).tail(2)), 360),
            }
        )
    return linhas


def identificar_recuos(
    reunioes: pd.DataFrame,
    ano_ini: int = 2003,
    ano_fim: int = 2015,
) -> list[dict]:
    """Cortes seguidos de alta que devolve (ou tenta devolver) o patamar anterior."""
    df = reunioes.sort_values("data").reset_index(drop=True)
    rotulos = [
        "2003–05 — desinflação interrompida",
        "2005–08 — cortes longos e alta de 2008",
        "2009–11 — crise e quase-retorno a 13,75%",
        "2011–15 — nova matriz e retorno acima do ponto de partida",
    ]
    runs: list[dict] = []
    i = 0
    while i < len(df):
        if df.loc[i, "decisao"] != "corte":
            i += 1
            continue
        ano = int(df.loc[i, "ano"])
        if ano < ano_ini or ano > ano_fim:
            i += 1
            continue
        start = i
        if start > 0 and pd.notna(df.loc[start - 1, "selic"]):
            patamar = float(df.loc[start - 1, "selic"])
        else:
            patamar = float(df.loc[start, "selic_antes"])
        j = start
        while j < len(df) and df.loc[j, "decisao"] != "alta":
            j += 1
        if j >= len(df):
            break
        easing = df.loc[start : j - 1]
        cortes = easing[easing["decisao"] == "corte"]
        last_cut = cortes.iloc[-1]
        trough_idx = easing["selic"].idxmin()
        k = j
        while k < len(df) and df.loc[k, "decisao"] in ("alta", "manutenção"):
            k += 1
        hike = df.loc[j : k - 1]
        altas = hike[hike["decisao"] == "alta"]
        peak_idx = altas["selic"].idxmax()
        trough = float(df.loc[trough_idx, "selic"])
        peak = float(df.loc[peak_idx, "selic"])
        corte_pp = patamar - trough
        if corte_pp < 0.20:
            i = k
            continue
        frac = (peak - trough) / corte_pp
        if peak > patamar + 0.04:
            retorno = "retornou e superou"
        elif peak + 0.04 >= patamar:
            retorno = "retornou"
        elif frac >= 0.70:
            retorno = "quase retornou"
        else:
            retorno = "reversão parcial"
        meses = (pd.to_datetime(df.loc[j, "data"]) - pd.to_datetime(last_cut["data"])).days / 30.44
        n = len(runs)
        runs.append(
            {
                "episodio": n + 1,
                "rotulo": rotulos[n] if n < len(rotulos) else f"Recuo {n + 1}",
                "corte_ini": int(df.loc[start, "reuniao"]),
                "corte_ini_data": pd.to_datetime(df.loc[start, "data"]),
                "ultimo_corte": int(last_cut["reuniao"]),
                "ultimo_corte_data": pd.to_datetime(last_cut["data"]),
                "patamar": patamar,
                "piso": trough,
                "piso_reuniao": int(df.loc[trough_idx, "reuniao"]),
                "piso_data": pd.to_datetime(df.loc[trough_idx, "data"]),
                "corte_pp": corte_pp,
                "alta_ini": int(df.loc[j, "reuniao"]),
                "alta_ini_data": pd.to_datetime(df.loc[j, "data"]),
                "pico": peak,
                "pico_reuniao": int(df.loc[peak_idx, "reuniao"]),
                "pico_data": pd.to_datetime(df.loc[peak_idx, "data"]),
                "meses_ate_reverter": meses,
                "retorno": retorno,
                "n_cortes": int(len(cortes)),
                "n_altas": int(len(altas)),
                "fracao": frac,
            }
        )
        i = k
    return runs


def cabecalhos_recuos() -> list[str]:
    return [
        "Episódio",
        "Cortes (atas)",
        "Patamar antes",
        "Piso",
        "Corte p.p.",
        "1ª alta",
        "Pico depois",
        "Meses até reverter",
        "Retorno",
    ]


def linhas_recuos(recuos: list[dict]) -> list[list[str]]:
    linhas = []
    for r in recuos:
        cortes = f"{r['corte_ini']}ª–{r['ultimo_corte']}ª"
        alta = f"{r['alta_ini']}ª ({r['alta_ini_data'].strftime('%d/%m/%Y')})"
        pico = f"{_fmt(r['pico'])} ({r['pico_reuniao']}ª)"
        linhas.append(
            [
                f"{r['episodio']} — {r['rotulo'].split(' — ', 1)[-1] if ' — ' in r['rotulo'] else r['rotulo']}",
                cortes,
                _fmt(r["patamar"]),
                _fmt(r["piso"]),
                _fmt(r["corte_pp"]),
                alta,
                pico,
                _fmt(r["meses_ate_reverter"], 1),
                r["retorno"],
            ]
        )
    return linhas


def _narrativas_recuos() -> dict[int, str]:
    return {
        1: (
            "A 86ª (23/07/2003) fixa a Selic em 24,50% e abre o recuo desde "
            "26,50%. Os cortes seguem até 16,00% na 95ª (14/04/2004); a 94ª "
            "já havia sido dividida (6 a 3). Cinco meses depois a 100ª "
            "(15/09/2004) eleva a meta a 16,25% por 5 votos a 3 — o piso não "
            "seguiu. Oito altas levam a taxa a 19,75% (108ª, mai/2005). Não "
            "volta aos 26,50%, mas desfaz o tramo final do recuo e sobe "
            "3,75 p.p. acima do piso."
        ),
        2: (
            "A 112ª (14/09/2005) corta de 19,75% para 19,50% e diz que "
            "acompanhará o cenário antes de definir os próximos passos. O "
            "ciclo se alonga até 11,25% (129ª, set/2007). Sete meses depois "
            "a 134ª (16/04/2008) eleva a Selic a 11,75% e antecipa 'parte "
            "relevante do movimento' para reduzir o ajuste total. O pico "
            "fica em 13,75% (137ª, set/2008): Lehman corta a volta ao "
            "patamar de 19,75%, mas a direção já havia se invertido."
        ),
        3: (
            "A 140ª (21/01/2009) inicia o corte anticíclico desde 13,75%. "
            "A 144ª (22/07/2009) para em 8,75% e avalia que o patamar é "
            "'consistente com um cenário inflacionário benigno'. Nove meses "
            "depois a 150ª (28/04/2010) começa a volta; em julho de 2011 a "
            "160ª chega a 12,50% — quase o 13,75% de partida."
        ),
        4: (
            "O recuo mais curto entre pico e corte: a 160ª (20/07/2011) "
            "eleva a Selic a 12,50%; a 161ª, 42 dias depois, corta para "
            "12,00% por 5 votos a 2 (a minoria queria manter 12,50%). As "
            "atas da nova matriz seguem até 7,25% (170ª, 5 a 3). Seis "
            "meses após o piso, a 174ª (17/04/2013) inicia a alta (6 a 2). "
            "Desta vez a taxa não só retorna: em julho de 2015 a 192ª "
            "fixa 14,25% — acima dos 12,50% abandonados em 2011 — e fala "
            "em mantê-la 'por período suficientemente prolongado'."
        ),
    }


def _narrativas() -> dict[str, str]:
    """Leitura qualitativa ancorada no padrão recorrente das atas oficiais."""
    return {
        "2001–2002": (
            "O regime de metas (1999) ainda se consolidava. As atas de 2001 tratam do "
            "choque de energia, da desvalorização e da inflação corrente acima da meta; "
            "o Copom sobe a Selic para conter o repasse cambial. Em 2002 o risco "
            "eleitoral e a crise de confiança dominam o cenário externo/doméstico: "
            "o câmbio dispara, as expectativas se desancoram e o Comitê responde com "
            "aperto adicional no fim do ano."
        ),
        "2003–2005": (
            "As atas de 2003 enfatizam credibilidade e desinflação, e o Copom "
            "corta a Selic de 26,50% para 16,00% (86ª–95ª). O recuo não se "
            "sustenta: cinco meses após o último corte, a 100ª reunião (set/2004) "
            "já eleva a taxa, por 5 votos a 3, e oito altas levam o juro a "
            "19,75% em maio de 2005. A ata registra atividade e serviços mais "
            "fortes do que o cenário que justificara os cortes de 0,25 p.p. de "
            "março–abril de 2004 (a 94ª foi 6 a 3)."
        ),
        "2006–2008": (
            "O segundo recuo (112ª, set/2005) parte de 19,75% e chega a 11,25% "
            "em setembro de 2007. Sete meses depois a 134ª (abr/2008) eleva a "
            "Selic a 11,75% e diz que fará 'de imediato parte relevante do "
            "movimento' para reduzir o ajuste total. O pico fica em 13,75% "
            "(set/2008): a crise de Lehman interrompe a volta ao patamar de "
            "19,75%, mas o sentido da política já havia se invertido."
        ),
        "2009–2010": (
            "O terceiro recuo é a resposta à crise: 13,75% → 8,75% (140ª–144ª). "
            "A 144ª afirma que 8,75% seria 'consistente com um cenário "
            "inflacionário benigno'. Nove meses depois a 150ª (abr/2010) inicia "
            "a volta; em julho de 2011 a Selic está em 12,50%, quase o "
            "patamar de 13,75% de onde se partira."
        ),
        "2011–2014": (
            "O quarto recuo é o mais nítido. A 160ª (20/07/2011) leva a Selic a "
            "12,50%; 42 dias depois a 161ª corta para 12,00% (5 a 2). As atas "
            "da 'nova matriz' seguem cortando até 7,25% (170ª, 5 a 3). Seis "
            "meses após o piso, a 174ª (abr/2013) começa a volta (6 a 2). A "
            "Selic não só retorna aos 12,50%: em 2015 supera e vai a 14,25%."
        ),
        "2015–2016": (
            "As atas descrevem recessão, realinhamento de preços administrados e "
            "IPCA de dois dígitos. O Copom leva a Selic a 14,25% e a mantém. O "
            "vocabulário é de realinhamento, inércia e necessidade de política "
            "contracionista por período prolongado. Em 2016 o tom começa a mudar "
            "quando a atividade colapsa e a inflação recua."
        ),
        "2017–2019": (
            "Ciclo longo de cortes. As atas destacam reformas, teto de gastos, "
            "hiato negativo e convergência das expectativas. A Selic chega a 6,5% "
            "e depois a 4,5% em 2019. O Comitê discute, pela primeira vez de forma "
            "sistemática, o limite inferior da taxa estrutural e o papel da "
            "comunicação (forward guidance implícito)."
        ),
        "2020–2021": (
            "A pandemia entra nas atas de março de 2020: choque de demanda, "
            "crédito, linhas temporárias e Selic em 2%. O guidance de manutenção "
            "é explícito. Em 2021 o diagnóstico vira: choques de oferta, "
            "desancoragem e atividade mais forte do que o esperado. O Copom inicia "
            "o ciclo de altas mais intenso do regime de metas."
        ),
        "2022–2023": (
            "As atas de 2022 tratam de inflação persistente, guerra na Ucrânia e "
            "Selic em 13,75%, com o Comitê sinalizando que o juro permanecerá "
            "alto 'por tempo suficientemente prolongado'. Em 2023, com a "
            "desinflação de bens e a nova âncora fiscal, começa o ciclo de cortes "
            "(agosto), ainda com ênfase em expectativas e hiato."
        ),
        "2024–2026": (
            "2024 mistura cortes residuais e, depois, a retomada do aperto quando "
            "as expectativas de inflação de médio prazo se desancoram e o fiscal "
            "volta ao centro do risco. Em 2025 o Copom leva a Selic a 15,00% e a "
            "mantém. Em 2026 inicia um ciclo de cortes cauteloso (15,00% → 14,00% "
            "na 280ª reunião). As atas no formato A/B/C separam conjuntura, "
            "cenários e decisão, e reiteram política contracionista até a "
            "convergência da inflação à meta."
        ),
    }


def cabecalhos_ciclos() -> list[str]:
    return ["Período", "Contexto", "Reuniões", "Selic início", "Selic fim", "Máx", "Mín", "Altas", "Cortes", "Manut."]


def cabecalhos_anual() -> list[str]:
    return ["Ano", "Reuniões", "Selic início", "Selic fim", "Δ p.p.", "Altas", "Cortes", "Manut.", "Unânimes"]


def cabecalhos_reunioes() -> list[str]:
    return ["Reunião", "Data", "Selic", "Δ p.p.", "Decisão", "Voto"]


def linhas_ciclos(ciclos: list[dict]) -> list[list[str]]:
    return [
        [
            c["periodo"],
            c["rotulo"],
            str(c["n"]),
            _fmt(c["selic_ini"]),
            _fmt(c["selic_fim"]),
            _fmt(c["selic_max"]),
            _fmt(c["selic_min"]),
            str(c["altas"]),
            str(c["cortes"]),
            str(c["manutencoes"]),
        ]
        for c in ciclos
    ]


def linhas_anual(anual: pd.DataFrame) -> list[list[str]]:
    return [
        [
            str(int(r["ano"])),
            str(int(r["n"])),
            _fmt(r["selic_ini"]),
            _fmt(r["selic_fim"]),
            _fmt_signed(r["delta"]),
            str(int(r["altas"])),
            str(int(r["cortes"])),
            str(int(r["manutencoes"])),
            str(int(r["unanimes"])),
        ]
        for r in anual.to_dict("records")
    ]


def linhas_reunioes(reunioes: pd.DataFrame) -> list[list[str]]:
    return [
        [
            str(int(r["reuniao"])) if pd.notna(r.get("reuniao")) else "—",
            pd.to_datetime(r["data"]).strftime("%d/%m/%Y"),
            _fmt(r.get("selic")),
            _fmt_signed(r.get("delta_selic")),
            r.get("decisao") or "—",
            r.get("voto") or "—",
        ]
        for r in reunioes.to_dict("records")
    ]


def _escrever_aba(ws, cabecalhos, linhas) -> None:
    preench_cab = PatternFill("solid", fgColor="1F4E79")
    preench_alt = PatternFill("solid", fgColor="EEF3F8")
    fonte_cab = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fonte_cel = Font(name="Calibri", size=11)
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(1, col, cab)
        cell.border = BORDA_CONTINUA
        cell.fill = preench_cab
        cell.font = fonte_cab
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(linhas, start=2):
        for col, valor in enumerate(row, start=1):
            cell = ws.cell(i, col, valor)
            cell.border = BORDA_CONTINUA
            cell.font = fonte_cel
            cell.alignment = Alignment(horizontal="center" if col == 1 else "right")
            if i % 2 == 0:
                cell.fill = preench_alt
    for col in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def exportar_tabelas(
    reunioes: pd.DataFrame,
    anual: pd.DataFrame,
    ciclos: list[dict],
    recuos: list[dict],
    output_dir: Path,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_r = output_dir / "copom_atas_reunioes_2001_2026.csv"
    reunioes.drop(columns=["texto"], errors="ignore").to_csv(csv_r, index=False)
    csv_a = output_dir / "copom_atas_anual_2001_2026.csv"
    anual.to_csv(csv_a, index=False)
    csv_rec = output_dir / "copom_atas_recuos_2003_2015.csv"
    pd.DataFrame(recuos).to_csv(csv_rec, index=False)
    xlsx = output_dir / "resumo_atas_copom_2001_2026.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ciclos"
    _escrever_aba(ws1, cabecalhos_ciclos(), linhas_ciclos(ciclos))
    ws2 = wb.create_sheet("Recuos_2003_2015")
    _escrever_aba(ws2, cabecalhos_recuos(), linhas_recuos(recuos))
    ws3 = wb.create_sheet("Anual")
    _escrever_aba(ws3, cabecalhos_anual(), linhas_anual(anual))
    ws4 = wb.create_sheet("Reunioes")
    _escrever_aba(ws4, cabecalhos_reunioes(), linhas_reunioes(reunioes))
    wb.save(xlsx)
    return [csv_r, csv_a, csv_rec, xlsx]


def gerar_graficos(reunioes: pd.DataFrame, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(12, 5.2))
    ax.plot(reunioes["data"], reunioes["selic"], color="#1f4e79", linewidth=2.0)
    ax.set_title("Meta Selic nas reuniões do Copom (2001–2026)")
    ax.set_ylabel("% a.a.")
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p1 = output_dir / "grafico_copom_selic_2001_2026.png"
    fig.savefig(p1, dpi=140)
    plt.close(fig)

    anual = reunioes.groupby("ano").agg(altas=("decisao", lambda s: (s == "alta").sum()), cortes=("decisao", lambda s: (s == "corte").sum()), manut=("decisao", lambda s: (s == "manutenção").sum()))
    fig, ax = plt.subplots(figsize=(12, 5.2))
    ax.bar(anual.index, anual["altas"], color="#b54708", label="Altas")
    ax.bar(anual.index, -anual["cortes"], color="#1b7f4a", label="Cortes")
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_title("Número de altas e cortes da Selic por ano")
    ax.legend(frameon=False)
    ax.grid(axis="y", linestyle=":", alpha=0.4)
    fig.tight_layout()
    p2 = output_dir / "grafico_copom_decisoes_ano.png"
    fig.savefig(p2, dpi=140)
    plt.close(fig)

    recuos = identificar_recuos(reunioes)
    janela = reunioes[(reunioes["ano"] >= 2003) & (reunioes["ano"] <= 2015)]
    fig, ax = plt.subplots(figsize=(12, 5.2))
    ax.plot(janela["data"], janela["selic"], color="#1f4e79", linewidth=2.0)
    cores = ["#c6dbef", "#fde0dd", "#d9f0d3", "#fee6ce"]
    for i, rec in enumerate(recuos):
        ax.axvspan(rec["corte_ini_data"], rec["ultimo_corte_data"], color=cores[i % 4], alpha=0.55)
        ax.axvspan(rec["alta_ini_data"], rec["pico_data"], color="#f4cccc", alpha=0.35)
        ax.scatter([rec["piso_data"]], [rec["piso"]], color="#1b7f4a", zorder=5)
        ax.scatter([rec["pico_data"]], [rec["pico"]], color="#b54708", zorder=5)
    ax.set_title("Recuos e reversões da Selic nas atas (2003–2015)")
    ax.set_ylabel("% a.a.")
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    p3 = output_dir / "grafico_copom_recuos_2003_2015.png"
    fig.savefig(p3, dpi=140)
    plt.close(fig)
    return [p1, p2, p3]


def gerar_relatorio(
    reunioes: pd.DataFrame,
    anual: pd.DataFrame,
    ciclos: list[dict],
    recuos: list[dict],
    output_dir: Path,
) -> Path:
    narr = _narrativas()
    narr_rec = _narrativas_recuos()
    primeiro = reunioes.iloc[0]
    ultimo = reunioes.iloc[-1]
    html_ciclos = tabela_html(cabecalhos_ciclos(), linhas_ciclos(ciclos), ["center", "left"] + ["right"] * 8)
    html_recuos = tabela_html(cabecalhos_recuos(), linhas_recuos(recuos), ["left", "center"] + ["right"] * 7)
    html_anual = tabela_html(cabecalhos_anual(), linhas_anual(anual))
    html_reun = tabela_html(cabecalhos_reunioes(), linhas_reunioes(reunioes))
    recuo_blocos = []
    for rec in recuos:
        recuo_blocos.append(
            f"### Episódio {rec['episodio']} — {rec['rotulo']}\n\n"
            f"Patamar {_fmt(rec['patamar'])}% → piso {_fmt(rec['piso'])}% "
            f"({rec['n_cortes']} cortes). Primeira alta na {rec['alta_ini']}ª, "
            f"{_fmt(rec['meses_ate_reverter'], 1)} meses após o último corte. "
            f"Pico posterior {_fmt(rec['pico'])}% ({rec['retorno']}).\n\n"
            f"{narr_rec.get(rec['episodio'], '')}\n"
        )
    blocos = []
    for c in ciclos:
        texto = narr.get(c["periodo"], "")
        blocos.append(
            f"### {c['periodo']} — {c['rotulo']}\n\n"
            f"{c['n']} reuniões. Selic {_fmt(c['selic_ini'])}% → {_fmt(c['selic_fim'])}% "
            f"(máx. {_fmt(c['selic_max'])}%; mín. {_fmt(c['selic_min'])}%). "
            f"Altas {c['altas']}, cortes {c['cortes']}, manutenções {c['manutencoes']}.\n\n"
            f"{texto}\n"
        )
    recentes = reunioes.tail(8)
    rec_txt = []
    for r in recentes.to_dict("records"):
        rec_txt.append(
            f"- **{int(r['reuniao'])}ª** ({pd.to_datetime(r['data']).strftime('%d/%m/%Y')}): "
            f"Selic {_fmt(r.get('selic'))}% ({r.get('decisao')}, {r.get('voto')}). "
            f"{r.get('trecho') or ''}"
        )
    gerado = datetime.now().strftime("%Y-%m-%d")
    md = f"""# Resumo explicativo das atas do Copom (2001–2026)

**Fonte:** Banco Central do Brasil, [atas do Copom](https://www.bcb.gov.br/publicacoes/atascopom)
(API `atascopom/ultimas` e `atascopom/principal`). Meta Selic: SGS 432.
**Consulta:** {gerado}. Meta Selic: SGS 432 (vigente no dia seguinte à reunião).
Tabelas com **grade contínua**.

O Copom (presidente e diretores do Banco Central) define a **meta da Selic**
visando a inflação do IPCA na meta do CMN. A decisão sai no comunicado do
segundo dia; a **ata** sai em até quatro dias úteis e explica o diagnóstico
(atividade, inflação, expectativas, fiscal, cenário externo) e o balanço de
riscos. Desde 1999 o Brasil opera sob **metas de inflação**; as atas a partir
de 2001 já estão nesse regime.

Cobertura: **{int(primeiro.reuniao)}ª** reunião
({pd.to_datetime(primeiro.data).strftime('%d/%m/%Y')}) à **{int(ultimo.reuniao)}ª**
({pd.to_datetime(ultimo.data).strftime('%d/%m/%Y')}) — **{len(reunioes)}** atas.
Selic de {_fmt(primeiro.selic)}% para {_fmt(ultimo.selic)}%.
Atas 200–231 vieram só em PDF; as demais, em HTML da API oficial.

## Como ler as atas

Três blocos se repetem ao longo de 25 anos, ainda que o formato mude
(até 2016: evolução / perspectivas / implementação; depois: seções A/B/C):

1. **Conjuntura** — IPCA, atividade, crédito, câmbio, mundo.
2. **Cenário prospectivo** — expectativas, hiato, balanço de riscos.
3. **Decisão** — alta, corte ou manutenção; voto; e, cada vez mais, a
   *comunicação* sobre os próximos passos.

O fio condutor é a **âncora das expectativas**. Quando elas se afastam da
meta, o Copom aperta (2002, 2015, 2021–22, 2024–25). Quando o hiato está
negativo e as expectativas convergem, corta (2009, 2017–19, 2020, 2023).

Entre **2003 e 2015** esse fio aparece como **stop-and-go**: as atas
justificam um recuo da Selic; poucos meses após o último corte, o Comitê
é obrigado a reverter e, em três dos quatro episódios, a taxa volta perto
ou acima do patamar de partida.

## Recuos e reversões (2003–2015)

Quatro vezes o Copom cortou a Selic e, em 5 a 9 meses depois do último
corte, teve de iniciar a volta. O padrão está nas próprias atas: o texto
do recuo fala em cenário benigno ou em desaceleração mundial; o texto da
reversão fala em atividade mais forte, expectativas e, em 2011–15, perda
de âncora.

{html_recuos}

{"".join(recuo_blocos)}

## Ciclos de política nas atas

{html_ciclos}

{"".join(blocos)}

## Ano a ano

{html_anual}

## Atas mais recentes (trechos da decisão)

{chr(10).join(rec_txt)}

## Todas as reuniões (2001–2026)

{html_reun}

## Arquivos

- `resumo_atas_copom_2001_2026.md` / `.xlsx` / `.pdf`
- `copom_atas_reunioes_2001_2026.csv`
- `copom_atas_recuos_2003_2015.csv`
- `grafico_copom_selic_2001_2026.png`
- `grafico_copom_recuos_2003_2015.png`
"""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "resumo_atas_copom_2001_2026.md"
    path.write_text(md, encoding="utf-8")
    return path


def _pdf_txt(texto) -> str:
    return str(texto).replace("$", r"\$")


def gerar_pdf(
    reunioes: pd.DataFrame,
    anual: pd.DataFrame,
    ciclos: list[dict],
    recuos: list[dict],
    imagens: list[Path],
    output_dir: Path,
) -> Path:
    path = output_dir / "resumo_atas_copom_2001_2026.pdf"
    output_dir.mkdir(parents=True, exist_ok=True)
    a4 = (11.69, 8.27)

    def capa():
        fig = plt.figure(figsize=a4, facecolor="white")
        fig.text(0.07, 0.82, _pdf_txt("Atas do Copom — resumo explicativo (2001–2026)"), fontsize=18, fontweight="bold", color="#1f4e79")
        fig.text(0.07, 0.68, _pdf_txt("Fonte: Banco Central do Brasil — www.bcb.gov.br/publicacoes/atascopom"), fontsize=11)
        fig.text(0.07, 0.60, _pdf_txt(f"{len(reunioes)} atas, da {int(reunioes.iloc[0].reuniao)}ª à {int(reunioes.iloc[-1].reuniao)}ª reunião."), fontsize=11)
        fig.text(0.07, 0.52, _pdf_txt("Selic meta (SGS 432) vigente no dia seguinte a cada reunião."), fontsize=11)
        fig.text(0.07, 0.10, "Tabelas com grade contínua. Valores em % a.a.", fontsize=8, color="#555")
        return fig

    def pagina_tabela(titulo, cabs, lins, larg=None):
        fig = plt.figure(figsize=a4, facecolor="white")
        ax = fig.add_axes([0.03, 0.04, 0.94, 0.86])
        ax.set_axis_off()
        fig.suptitle(_pdf_txt(titulo), fontsize=12, x=0.04, ha="left", y=0.96, color="#1f4e79")
        fonte = 8 if len(cabs) <= 8 else 6.4
        tab = ax.table(
            cellText=[[_pdf_txt(c) for c in row] for row in lins],
            colLabels=[_pdf_txt(h) for h in cabs],
            loc="center",
            cellLoc="center",
            colWidths=larg,
        )
        tab.auto_set_font_size(False)
        tab.set_fontsize(fonte)
        tab.scale(1, 1.15)
        for (r, _c), cell in tab.get_celld().items():
            cell.set_edgecolor("#1a1a1a")
            cell.set_linewidth(0.6)
            cell.visible_edges = "BTRL"
            if r == 0:
                cell.set_facecolor("#1f4e79")
                cell.get_text().set_color("white")
                cell.get_text().set_fontweight("bold")
            elif r % 2 == 0:
                cell.set_facecolor("#eef3f8")
        return fig

    with PdfPages(path) as pdf:
        fig = capa()
        pdf.savefig(fig)
        plt.close(fig)
        fig = pagina_tabela("Ciclos de política nas atas", cabecalhos_ciclos(), linhas_ciclos(ciclos), [0.10, 0.28, 0.08, 0.09, 0.09, 0.07, 0.07, 0.07, 0.08, 0.07])
        pdf.savefig(fig)
        plt.close(fig)
        fig = pagina_tabela(
            "Recuos e reversões da Selic (2003–2015)",
            cabecalhos_recuos(),
            linhas_recuos(recuos),
            [0.22, 0.10, 0.09, 0.07, 0.08, 0.14, 0.12, 0.10, 0.08],
        )
        pdf.savefig(fig)
        plt.close(fig)
        fig = pagina_tabela("Ano a ano", cabecalhos_anual(), linhas_anual(anual))
        pdf.savefig(fig)
        plt.close(fig)
        recs = linhas_reunioes(reunioes)
        for i in range(0, len(recs), 28):
            fig = pagina_tabela(f"Reuniões {i + 1}–{min(i + 28, len(recs))}", cabecalhos_reunioes(), recs[i : i + 28])
            pdf.savefig(fig)
            plt.close(fig)
        for img in imagens:
            if not img.exists():
                continue
            fig = plt.figure(figsize=a4, facecolor="white")
            ax = fig.add_axes([0.04, 0.06, 0.92, 0.88])
            ax.imshow(plt.imread(img))
            ax.set_axis_off()
            pdf.savefig(fig)
            plt.close(fig)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--sem-download", action="store_true")
    args = parser.parse_args(argv)

    cache_cat = args.cache_dir / "copom_atas_catalogo.csv"
    if cache_cat.exists():
        catalogo = pd.read_csv(cache_cat, parse_dates=["data"])
    else:
        if args.sem_download:
            raise FileNotFoundError(cache_cat)
        print("Listando atas em bcb.gov.br/publicacoes/atascopom ...", flush=True)
        catalogo = listar_atas()
        args.cache_dir.mkdir(parents=True, exist_ok=True)
        catalogo.to_csv(cache_cat, index=False)
    print(f"Catálogo: {len(catalogo)} atas ({catalogo['data'].min().date()}–{catalogo['data'].max().date()})")
    atas = carregar_textos(catalogo, args.cache_dir, baixar=not args.sem_download)
    selic = carregar_selic(args.cache_dir, baixar=not args.sem_download)
    reunioes = agregar_reunioes(atas, selic)
    anual = resumo_anual(reunioes)
    ciclos = resumo_ciclos(reunioes)
    recuos = identificar_recuos(reunioes)
    caminhos = exportar_tabelas(reunioes, anual, ciclos, recuos, args.output_dir)
    caminhos.append(gerar_relatorio(reunioes, anual, ciclos, recuos, args.output_dir))
    imgs = gerar_graficos(reunioes, args.output_dir)
    caminhos.extend(imgs)
    caminhos.append(gerar_pdf(reunioes, anual, ciclos, recuos, imgs, args.output_dir))
    print("Recuos 2003–2015:")
    for rec in recuos:
        print(
            f"  {rec['episodio']}. {rec['rotulo']}: {_fmt(rec['patamar'])}% → "
            f"{_fmt(rec['piso'])}% → {_fmt(rec['pico'])}% ({rec['retorno']})"
        )
    print(anual.to_string(index=False))
    for p in caminhos:
        print(f"  {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
