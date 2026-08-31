#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Simulação da DBGG com Selic anual = IPCA do ano + spread (default 0,37 p.p.).

Usa a planilha especial do BCB (estoque, juros nominais e emissões líquidas
por indexador) e as séries SGS 433 (IPCA) e 4390 (Selic % a.m.).

Metodologia (contadorfactual contábil, sem 2ª ordem):
  1. Para cada ano civil, IPCA_ano = Π(1 + ipca_m) − 1 nos meses do ano.
     Selic_cf,ano = IPCA_ano + spread. Taxa mensal constante no ano:
     r_cf_m = (1 + Selic_cf,ano)^(1/n) − 1, n = meses do ano na amostra.
  2. Juros da parcela Selic no mês t:
     juros_cf = juros_act × (r_cf_m / r_selic_m) × (S_cf[t−1] / S_act[t−1])
  3. S_cf[t] = S_cf[t−1] + emissão_líquida_selic[t] + juros_cf[t] + resíduo_BCB[t]
     onde o resíduo replica o fechamento observado do estoque Selic.
  4. DBGG_cf[t] = DBGG_act[t] − (S_act[t] − S_cf[t])

A parcela Selic do BCB inclui LFT, LFT-A/B, compromissadas, dívida bancária
e securitizadas indexadas à Selic. Demais indexadores e as emissões líquidas
observadas permanecem iguais.

Uso::

  python scripts/simular_dbgg_selic_ipca.py
  python scripts/simular_dbgg_selic_ipca.py --spread 0.37 --saida-dir output
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.calcular_diretas_ipca_selic import (  # noqa: E402
    IPCA_COD,
    SELIC_MENSAL_COD,
    _baixar_sgs,
    carregar_ipca,
    carregar_selic_mensal,
)

URL_DBGG = (
    "https://www.bcb.gov.br/content/estatisticas/Documents/"
    "Tabelas_especiais/Dbggindexp.xlsx"
)
DBGG_PIB_COD = 4513  # DBGG / PIB (%) — SGS
PIB_12M_COD = 4382  # PIB acumulado em 12 meses (R$ milhões)
DPMFI_COD = 4181  # Dívida mobiliária — posição em carteira (R$ milhões)
DPMFI_SELIC_PCT_COD = 4177  # Participação Over/Selic na DPMFi (%)
SPREAD_DEFAULT = 0.37
MES_INICIO_DEFAULT = pd.Timestamp("2007-01-01")
MES_FIM_DEFAULT = pd.Timestamp("2026-06-01")
ANOS_OBSERVADOS_DEFAULT: tuple[int, ...] = ()
UA = "SEC-data-analysys/dbgg-selic-ipca"

MESES_PT = {
    "Jan": 1,
    "Fev": 2,
    "Feb": 2,
    "Mar": 3,
    "Abr": 4,
    "Apr": 4,
    "Mai": 5,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Ago": 8,
    "Aug": 8,
    "Set": 9,
    "Sep": 9,
    "Out": 10,
    "Oct": 10,
    "Nov": 11,
    "Dez": 12,
    "Dec": 12,
}

COLUNAS_INDEX = [
    "cambial_interna",
    "cambial_externa",
    "cambial_total",
    "igpm",
    "igpdi",
    "ipca",
    "indices_total",
    "selic",
    "tjlp_tlp",
    "tr",
    "prefixado",
    "outros",
    "total",
]


def baixar_planilha_dbgg(
    destino: Path,
    url: str = URL_DBGG,
    *,
    forcar: bool = False,
    tentativas: int = 5,
) -> Path:
    """Baixa Dbggindexp.xlsx do BCB (ou reutiliza cache local)."""
    if destino.exists() and not forcar:
        return destino
    destino.parent.mkdir(parents=True, exist_ok=True)
    last: Exception | None = None
    for i in range(tentativas):
        try:
            req = Request(url, headers={"User-Agent": UA})
            with urlopen(req, timeout=120) as resp:
                destino.write_bytes(resp.read())
            if destino.stat().st_size < 1000:
                raise RuntimeError("Planilha BCB vazia ou incompleta")
            return destino
        except (HTTPError, URLError, TimeoutError, RuntimeError, OSError) as exc:
            last = exc
            if i < tentativas - 1:
                import time

                time.sleep(2**i)
    raise RuntimeError(f"Falha ao baixar {url}: {last}") from last


def _primeira_linha_dados(ws) -> int:
    for i in range(1, min(40, int(ws.max_row or 1) + 1)):
        b = ws.cell(i, 2).value
        if b is not None and str(b).strip() in MESES_PT:
            return i
    raise ValueError("Nenhuma linha de dados (mês) encontrada na aba")


def ler_aba_indexadores(path: Path, aba: str) -> pd.DataFrame:
    """Lê estoque/juros/emissões por indexador (R$ milhões)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba {aba!r} ausente. Disponíveis: {wb.sheetnames}")
    ws = wb[aba]
    start = _primeira_linha_dados(ws)
    year: int | None = None
    rows: list[dict] = []
    for i in range(start, int(ws.max_row or start) + 1):
        a = ws.cell(i, 1).value
        b = ws.cell(i, 2).value
        if isinstance(a, (int, float)) and not isinstance(a, bool):
            year = int(a)
        if b is None or year is None:
            continue
        mes_n = MESES_PT.get(str(b).strip())
        if mes_n is None:
            continue
        vals = [ws.cell(i, j).value for j in range(3, 16)]
        if all(v is None for v in vals):
            continue
        rec: dict = {
            "mes": pd.Timestamp(year=year, month=mes_n, day=1),
            "ano": year,
            "mes_n": mes_n,
        }
        for nome, raw in zip(COLUNAS_INDEX, vals, strict=True):
            rec[nome] = float(raw) if raw is not None else np.nan
        rows.append(rec)
    wb.close()
    if not rows:
        raise ValueError(f"Aba {aba!r} sem observações numéricas")
    out = pd.DataFrame(rows).drop_duplicates("mes").sort_values("mes")
    return out.reset_index(drop=True)


def carregar_dbgg(path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Devolve (estoque, juros, emissões líquidas) em R$ milhões."""
    estoque = ler_aba_indexadores(path, "DividaR$")
    juros = ler_aba_indexadores(path, "JurosR$")
    emissoes = ler_aba_indexadores(path, "PrimarioR$")
    return estoque, juros, emissoes


def _mes_ts(serie: pd.Series) -> pd.Series:
    return pd.to_datetime(serie).dt.to_period("M").dt.to_timestamp()


def _linha_indexadores(mes: pd.Timestamp, selic: float, total: float) -> dict:
    rec = {c: 0.0 for c in COLUNAS_INDEX}
    rec.update(
        {
            "mes": pd.Timestamp(mes).to_period("M").to_timestamp(),
            "ano": int(pd.Timestamp(mes).year),
            "mes_n": int(pd.Timestamp(mes).month),
            "selic": float(selic),
            "total": float(total),
        }
    )
    return rec


def reconstruir_pre_oficial(
    estoque_oficial: pd.DataFrame,
    selic_m: pd.DataFrame,
    dpmfi: pd.DataFrame,
    share_selic: pd.DataFrame,
    pib_12m: pd.DataFrame,
    dbgg_pib: pd.DataFrame,
    mes_primeiro: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Estoque/juros/emissões Selic de ``mes_primeiro−1`` até o mês anterior ao oficial.

    O estoque Selic é o produto DPMFi (SGS 4181) × participação Over/Selic
    (SGS 4177), reescalonado para coincidir com o primeiro estoque oficial
    (dez/2006). A DBGG total usa 4513 × PIB 12 meses (4382), também
    reescalonada no mesmo mês de emenda. Juros = estoque anterior × Selic
    mensal (SGS 4390); emissão líquida fecha o estoque.
    """
    mes_primeiro = pd.Timestamp(mes_primeiro).to_period("M").to_timestamp()
    mes0 = mes_primeiro - pd.DateOffset(months=1)
    of_min = pd.Timestamp(estoque_oficial["mes"].min()).to_period("M").to_timestamp()
    if mes0 >= of_min:
        vazio = estoque_oficial.iloc[0:0].copy()
        return vazio, vazio.copy(), vazio.copy()

    def _prep(df: pd.DataFrame, nome: str) -> pd.DataFrame:
        out = df.rename(columns={"valor": nome}).copy()
        out["mes"] = _mes_ts(out["mes"])
        return out[["mes", nome]].drop_duplicates("mes")

    prox = _prep(dpmfi, "dpmfi").merge(_prep(share_selic, "share"), on="mes")
    prox["proxy"] = prox["dpmfi"] * prox["share"] / 100.0
    pib = _prep(pib_12m, "pib").merge(_prep(dbgg_pib, "pib_pct"), on="mes")
    pib["dbgg_imp"] = pib["pib"] * pib["pib_pct"] / 100.0

    of0 = estoque_oficial.loc[estoque_oficial["mes"] == of_min].iloc[0]
    if of_min not in set(prox["mes"]) or of_min not in set(pib["mes"]):
        raise ValueError(
            f"SGS 4181/4177/4382/4513 sem o mês de emenda {of_min.strftime('%Y-%m')}"
        )
    proxy0 = float(prox.loc[prox["mes"] == of_min, "proxy"].iloc[0])
    dbgg0 = float(pib.loc[pib["mes"] == of_min, "dbgg_imp"].iloc[0])
    if proxy0 <= 0 or dbgg0 <= 0:
        raise ValueError("Proxy DPMFi/Selic ou DBGG implícita nula no mês de emenda")
    escala_s = float(of0["selic"]) / proxy0
    escala_d = float(of0["total"]) / dbgg0

    meses = pd.date_range(mes0, of_min - pd.DateOffset(months=1), freq="MS")
    sel = _prep(selic_m, "selic_am")
    base = pd.DataFrame({"mes": meses}).merge(prox[["mes", "proxy"]], on="mes", how="left")
    base = base.merge(pib[["mes", "dbgg_imp"]], on="mes", how="left")
    base = base.merge(sel, on="mes", how="left")
    if base[["proxy", "dbgg_imp"]].isna().any().any():
        raise ValueError("Faltam DPMFi/PIB para recuar a DBGG antes da planilha oficial")

    base["selic"] = base["proxy"] * escala_s
    base["total"] = base["dbgg_imp"] * escala_d
    # Selic do mês 0 (estoque inicial) não entra em juros; preenche 0
    base["selic_am"] = base["selic_am"].fillna(0.0)

    est_rows = [_linha_indexadores(r.mes, r.selic, r.total) for r in base.itertuples()]
    ju_rows = []
    em_rows = []
    s_prev = float(base.iloc[0]["selic"])
    for i, r in enumerate(base.itertuples()):
        if i == 0:
            continue
        j = s_prev * float(r.selic_am) / 100.0
        e = float(r.selic) - s_prev - j
        ju_rows.append(_linha_indexadores(r.mes, j, j))
        em_rows.append(_linha_indexadores(r.mes, e, e))
        s_prev = float(r.selic)
    # Mês de emenda (dez/2006): estoque oficial; juros/emissão fecham com o recuo.
    r_emenda = float(sel.loc[sel["mes"] == of_min, "selic_am"].iloc[0]) if of_min in set(sel["mes"]) else 0.0
    j_emenda = s_prev * r_emenda / 100.0
    e_emenda = float(of0["selic"]) - s_prev - j_emenda
    ju_rows.append(_linha_indexadores(of_min, j_emenda, j_emenda))
    em_rows.append(_linha_indexadores(of_min, e_emenda, e_emenda))
    return (
        pd.DataFrame(est_rows),
        pd.DataFrame(ju_rows),
        pd.DataFrame(em_rows),
    )


def projetar_meses_apos_oficial(
    estoque: pd.DataFrame,
    juros: pd.DataFrame,
    emissoes: pd.DataFrame,
    selic_m: pd.DataFrame,
    mes_fim: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Prolonga estoque/juros/emissões até ``mes_fim`` com emissão líquida zero.

    Juros do mês = estoque Selic anterior × Selic SGS 4390. A DBGG total
    sobe só pelo juro da parcela Selic (demais indexadores ficam congelados).
    """
    mes_fim = pd.Timestamp(mes_fim).to_period("M").to_timestamp()
    est = estoque.sort_values("mes").copy()
    ju = juros.copy() if juros.empty or "mes" not in juros.columns else juros.sort_values("mes").copy()
    em = emissoes.copy() if emissoes.empty or "mes" not in emissoes.columns else emissoes.sort_values("mes").copy()
    last = est.iloc[-1]
    cursor = pd.Timestamp(last["mes"]).to_period("M").to_timestamp()
    sel = selic_m.rename(columns={"valor": "selic_am"}).copy()
    sel["mes"] = _mes_ts(sel["mes"])
    sel = sel.set_index("mes")["selic_am"]
    s = float(last["selic"])
    tot = float(last["total"])
    while cursor < mes_fim:
        nxt = cursor + pd.DateOffset(months=1)
        if nxt not in sel.index:
            raise ValueError(f"Selic SGS 4390 ausente para projetar {nxt.strftime('%Y-%m')}")
        r = float(sel.loc[nxt])
        j = s * r / 100.0
        s = s + j
        tot = tot + j
        est = pd.concat([est, pd.DataFrame([_linha_indexadores(nxt, s, tot)])], ignore_index=True)
        ju = pd.concat([ju, pd.DataFrame([_linha_indexadores(nxt, j, j)])], ignore_index=True)
        em = pd.concat([em, pd.DataFrame([_linha_indexadores(nxt, 0.0, 0.0)])], ignore_index=True)
        cursor = nxt
    return est.reset_index(drop=True), ju.reset_index(drop=True), em.reset_index(drop=True)


def fundir_pre_oficial(
    recon_est: pd.DataFrame,
    recon_ju: pd.DataFrame,
    recon_em: pd.DataFrame,
    est: pd.DataFrame,
    ju: pd.DataFrame,
    em: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Concatena o recuo com a planilha oficial, sem duplicar o mês de emenda."""
    if recon_est.empty:
        return est, ju, em
    of_min = pd.Timestamp(est["mes"].min())
    est2 = pd.concat([recon_est[recon_est["mes"] < of_min], est], ignore_index=True)
    ju2 = pd.concat([recon_ju[~recon_ju["mes"].isin(set(ju["mes"]))], ju], ignore_index=True)
    em2 = pd.concat([recon_em[~recon_em["mes"].isin(set(em["mes"]))], em], ignore_index=True)
    return (
        est2.sort_values("mes").drop_duplicates("mes").reset_index(drop=True),
        ju2.sort_values("mes").drop_duplicates("mes").reset_index(drop=True),
        em2.sort_values("mes").drop_duplicates("mes").reset_index(drop=True),
    )


def ipca_por_ano(ipca: pd.DataFrame, meses: pd.Series) -> pd.DataFrame:
    """IPCA acumulado e Selic contrafactual por ano civil da amostra.

    ``selic_cf_aa`` é IPCA_ano + spread aplicado depois, aqui só o IPCA.
    """
    base = ipca.rename(columns={"valor": "ipca_am"}).copy()
    base["mes"] = pd.to_datetime(base["mes"]).dt.to_period("M").dt.to_timestamp()
    alvo = pd.DataFrame({"mes": pd.to_datetime(meses)})
    alvo["mes"] = alvo["mes"].dt.to_period("M").dt.to_timestamp()
    alvo["ano"] = alvo["mes"].dt.year
    m = alvo.merge(base[["mes", "ipca_am"]], on="mes", how="left")
    if m["ipca_am"].isna().any():
        faltando = m.loc[m["ipca_am"].isna(), "mes"].dt.strftime("%Y-%m").tolist()
        raise ValueError(f"IPCA ausente em: {faltando[:8]}")

    def _agg(g: pd.DataFrame) -> pd.Series:
        fator = float((1.0 + g["ipca_am"] / 100.0).prod())
        n = int(len(g))
        return pd.Series(
            {
                "n_meses": n,
                "ipca_acum_pct": (fator - 1.0) * 100.0,
                "ipca_aa_equiv_pct": (fator ** (12.0 / n) - 1.0) * 100.0,
            }
        )

    return m.groupby("ano", sort=True).apply(_agg, include_groups=False).reset_index()


def taxas_mensais_cf(
    ipca_anos: pd.DataFrame,
    meses: pd.Series,
    spread_pp: float,
    selic: pd.DataFrame | None = None,
    anos_observados: tuple[int, ...] | list[int] = (),
) -> pd.DataFrame:
    """Taxa mensal contrafactual constante dentro de cada ano civil.

    Em ``anos_observados`` a Selic contrafactual é a observada (SGS 4390),
    não IPCA + spread.
    """
    anos = ipca_anos.copy()
    anos["selic_cf_acum_pct"] = anos["ipca_acum_pct"] + spread_pp * (
        anos["n_meses"] / 12.0
    )
    # (1 + acum)^(1/n) − 1
    anos["selic_cf_am"] = (1.0 + anos["selic_cf_acum_pct"] / 100.0) ** (
        1.0 / anos["n_meses"]
    ) - 1.0
    anos["selic_cf_am"] *= 100.0
    out = pd.DataFrame({"mes": pd.to_datetime(meses)})
    out["mes"] = out["mes"].dt.to_period("M").dt.to_timestamp()
    out["ano"] = out["mes"].dt.year
    out = out.merge(
        anos[
            [
                "ano",
                "n_meses",
                "ipca_acum_pct",
                "ipca_aa_equiv_pct",
                "selic_cf_acum_pct",
                "selic_cf_am",
            ]
        ],
        on="ano",
        how="left",
    )
    out["selic_alterada"] = ~out["ano"].isin(set(anos_observados))
    if selic is not None and anos_observados:
        base = selic.rename(columns={"valor": "selic_am"}).copy()
        base["mes"] = pd.to_datetime(base["mes"]).dt.to_period("M").dt.to_timestamp()
        out = out.merge(base[["mes", "selic_am"]], on="mes", how="left")
        mask = ~out["selic_alterada"]
        if mask.any() and out.loc[mask, "selic_am"].isna().any():
            faltando = out.loc[mask & out["selic_am"].isna(), "mes"]
            raise ValueError(
                "Selic observada ausente nos anos sem alteração: "
                + ", ".join(faltando.dt.strftime("%Y-%m").tolist()[:8])
            )
        out.loc[mask, "selic_cf_am"] = out.loc[mask, "selic_am"]
        for ano in set(anos_observados):
            m_ano = out["ano"] == ano
            if not m_ano.any():
                continue
            fator = float((1.0 + out.loc[m_ano, "selic_cf_am"] / 100.0).prod())
            out.loc[m_ano, "selic_cf_acum_pct"] = (fator - 1.0) * 100.0
        out = out.drop(columns=["selic_am"])
    return out


def selic_acumulada_por_ano(selic: pd.DataFrame, meses: pd.Series) -> pd.DataFrame:
    """Selic observada acumulada no ano (composta das taxas mensais)."""
    base = selic.rename(columns={"valor": "selic_am"}).copy()
    base["mes"] = pd.to_datetime(base["mes"]).dt.to_period("M").dt.to_timestamp()
    alvo = pd.DataFrame({"mes": pd.to_datetime(meses)})
    alvo["mes"] = alvo["mes"].dt.to_period("M").dt.to_timestamp()
    alvo["ano"] = alvo["mes"].dt.year
    m = alvo.merge(base[["mes", "selic_am"]], on="mes", how="left")
    if m["selic_am"].isna().any():
        faltando = m.loc[m["selic_am"].isna(), "mes"].dt.strftime("%Y-%m").tolist()
        raise ValueError(f"Selic ausente em: {faltando[:8]}")

    def _agg(g: pd.DataFrame) -> pd.Series:
        fator = float((1.0 + g["selic_am"] / 100.0).prod())
        n = int(len(g))
        return pd.Series(
            {
                "n_meses": n,
                "selic_acum_pct": (fator - 1.0) * 100.0,
                "selic_aa_equiv_pct": (fator ** (12.0 / n) - 1.0) * 100.0,
            }
        )

    return m.groupby("ano", sort=True).apply(_agg, include_groups=False).reset_index()


def simular_parcela_selic(
    estoque: pd.DataFrame,
    juros: pd.DataFrame,
    emissoes: pd.DataFrame,
    selic_m: pd.DataFrame,
    taxas_cf: pd.DataFrame,
    *,
    mes_inicio: pd.Timestamp,
    mes_fim: pd.Timestamp,
) -> pd.DataFrame:
    """Reconstrói o estoque Selic e a DBGG no cenário IPCA + spread."""
    mes_inicio = pd.Timestamp(mes_inicio).to_period("M").to_timestamp()
    mes_fim = pd.Timestamp(mes_fim).to_period("M").to_timestamp()
    mes0 = mes_inicio - pd.DateOffset(months=1)

    est = estoque.set_index("mes").sort_index()
    ju = juros.set_index("mes").sort_index()
    em = emissoes.set_index("mes").sort_index()
    sel = selic_m.rename(columns={"valor": "selic_am"}).copy()
    sel["mes"] = pd.to_datetime(sel["mes"]).dt.to_period("M").dt.to_timestamp()
    sel = sel.set_index("mes")["selic_am"]
    cf = taxas_cf.set_index("mes")

    if mes0 not in est.index:
        raise ValueError(
            f"Estoque inicial {mes0.strftime('%Y-%m')} ausente na aba DividaR$"
        )

    meses = pd.date_range(mes_inicio, mes_fim, freq="MS")
    s_act_prev = float(est.loc[mes0, "selic"])
    s_cf_prev = s_act_prev
    dbgg0 = float(est.loc[mes0, "total"])

    recs: list[dict] = [
        {
            "mes": mes0,
            "ano": int(mes0.year),
            "selic_am": np.nan,
            "selic_cf_am": np.nan,
            "ipca_acum_ano_pct": np.nan,
            "selic_cf_acum_ano_pct": np.nan,
            "estoque_selic_act": s_act_prev,
            "estoque_selic_cf": s_cf_prev,
            "juros_selic_act": 0.0,
            "juros_selic_cf": 0.0,
            "emissao_selic": 0.0,
            "residuo_selic": 0.0,
            "dbgg_act": dbgg0,
            "dbgg_cf": dbgg0,
            "delta_estoque_selic": 0.0,
            "delta_dbgg": 0.0,
        }
    ]

    for mes in meses:
        if mes not in est.index or mes not in ju.index or mes not in em.index:
            raise ValueError(f"Mês {mes.strftime('%Y-%m')} ausente na planilha BCB")
        if mes not in sel.index:
            raise ValueError(f"Selic SGS 4390 ausente em {mes.strftime('%Y-%m')}")
        if mes not in cf.index:
            raise ValueError(f"Taxa contrafactual ausente em {mes.strftime('%Y-%m')}")

        s_act = float(est.loc[mes, "selic"])
        j_act = float(ju.loc[mes, "selic"])
        e_act = float(em.loc[mes, "selic"])
        residuo = s_act - (s_act_prev + e_act + j_act)
        r_act = float(sel.loc[mes])
        r_cf = float(cf.loc[mes, "selic_cf_am"])
        if r_act <= 0 or s_act_prev == 0:
            j_cf = j_act * 0.0 if r_act <= 0 and r_cf <= 0 else (
                s_cf_prev * (r_cf / 100.0) if s_act_prev == 0 else j_act
            )
        else:
            j_cf = j_act * (r_cf / r_act) * (s_cf_prev / s_act_prev)

        s_cf = s_cf_prev + e_act + j_cf + residuo
        dbgg_act = float(est.loc[mes, "total"])
        dbgg_cf = dbgg_act - (s_act - s_cf)

        recs.append(
            {
                "mes": mes,
                "ano": int(mes.year),
                "selic_am": r_act,
                "selic_cf_am": r_cf,
                "ipca_acum_ano_pct": float(cf.loc[mes, "ipca_acum_pct"]),
                "selic_cf_acum_ano_pct": float(cf.loc[mes, "selic_cf_acum_pct"]),
                "estoque_selic_act": s_act,
                "estoque_selic_cf": s_cf,
                "juros_selic_act": j_act,
                "juros_selic_cf": j_cf,
                "emissao_selic": e_act,
                "residuo_selic": residuo,
                "dbgg_act": dbgg_act,
                "dbgg_cf": dbgg_cf,
                "delta_estoque_selic": s_act - s_cf,
                "delta_dbgg": dbgg_act - dbgg_cf,
            }
        )
        s_act_prev = s_act
        s_cf_prev = s_cf

    return pd.DataFrame(recs)


def anexar_pib(mensal: pd.DataFrame, dbgg_pib: pd.DataFrame | None) -> pd.DataFrame:
    """Anexa DBGG/PIB oficial e a razão contrafactual (mesmo PIB)."""
    out = mensal.copy()
    out["dbgg_pib_act"] = np.nan
    out["dbgg_pib_cf"] = np.nan
    if dbgg_pib is None or dbgg_pib.empty:
        return out
    pib = dbgg_pib.rename(columns={"valor": "dbgg_pib_act"}).copy()
    pib["mes"] = pd.to_datetime(pib["mes"]).dt.to_period("M").dt.to_timestamp()
    out = out.drop(columns=["dbgg_pib_act"], errors="ignore")
    out = out.merge(pib[["mes", "dbgg_pib_act"]], on="mes", how="left")
    ratio = out["dbgg_cf"] / out["dbgg_act"].replace(0, np.nan)
    out["dbgg_pib_cf"] = out["dbgg_pib_act"] * ratio
    return out


def agregar_anual(mensal: pd.DataFrame) -> pd.DataFrame:
    """Resumo por ano civil (último mês do ano na amostra)."""
    sim = mensal[mensal["selic_am"].notna()].copy()
    if sim.empty:
        return pd.DataFrame()

    def _agg(g: pd.DataFrame) -> pd.Series:
        last = g.iloc[-1]
        return pd.Series(
            {
                "n_meses": int(len(g)),
                "mes_final": last["mes"],
                "ipca_acum_pct": float(last["ipca_acum_ano_pct"]),
                "selic_cf_acum_pct": float(last["selic_cf_acum_ano_pct"]),
                "selic_acum_pct": float(
                    ((1.0 + g["selic_am"] / 100.0).prod() - 1.0) * 100.0
                ),
                "juros_selic_act": float(g["juros_selic_act"].sum()),
                "juros_selic_cf": float(g["juros_selic_cf"].sum()),
                "economia_juros": float(
                    g["juros_selic_act"].sum() - g["juros_selic_cf"].sum()
                ),
                "estoque_selic_act": float(last["estoque_selic_act"]),
                "estoque_selic_cf": float(last["estoque_selic_cf"]),
                "dbgg_act": float(last["dbgg_act"]),
                "dbgg_cf": float(last["dbgg_cf"]),
                "delta_dbgg": float(last["delta_dbgg"]),
                "dbgg_pib_act": float(last["dbgg_pib_act"])
                if pd.notna(last.get("dbgg_pib_act", np.nan))
                else np.nan,
                "dbgg_pib_cf": float(last["dbgg_pib_cf"])
                if pd.notna(last.get("dbgg_pib_cf", np.nan))
                else np.nan,
            }
        )

    return (
        sim.groupby("ano", sort=True)
        .apply(_agg, include_groups=False)
        .reset_index()
    )


def montar_discriminativo(anual: pd.DataFrame) -> pd.DataFrame:
    """Redução da DBGG em cada ano (fluxo) e acumulada (estoque).

    A redução do ano é a economia de juros nominais da parcela Selic
    (``economia_juros``), igual à variação da diferença de estoque
    ``ΔDBGG_t − ΔDBGG_{t−1}``.
    """
    if anual.empty:
        return anual.copy()
    out = anual.sort_values("ano").reset_index(drop=True).copy()
    out["reducao_ano"] = out["economia_juros"]
    out["reducao_acumulada"] = out["delta_dbgg"]
    prev = out["delta_dbgg"].shift(1).fillna(0.0)
    out["variacao_delta_dbgg"] = out["delta_dbgg"] - prev
    total = float(out["reducao_ano"].sum())
    if total == 0:
        out["participacao_pct"] = 0.0
    else:
        out["participacao_pct"] = 100.0 * out["reducao_ano"] / total
    out["spread_selic_menos_cf"] = out["selic_acum_pct"] - out["selic_cf_acum_pct"]
    return out


def _fmt_bi(valor: float) -> str:
    return (
        f"R$ {valor / 1000.0:,.1f} bi"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def _fmt_tri(valor: float) -> str:
    """``valor`` em R$ milhões → texto em R$ trilhões."""
    return (
        f"R$ {valor / 1e6:,.2f} tri"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def _fmt_pct(valor: float, casas: int = 2) -> str:
    return f"{valor:,.{casas}f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_num(valor: float, casas: int = 1) -> str:
    return f"{valor:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _nota_ano_incompleto(anual: pd.DataFrame, spread_pp: float) -> str:
    """Texto do último ano quando a amostra não fecha em dezembro."""
    if anual.empty:
        return ""
    last = anual.iloc[-1]
    n = int(last["n_meses"])
    if n >= 12:
        return (
            "A taxa mensal contrafactual é a equivalente composta, constante "
            "dentro do ano."
        )
    ano = int(last["ano"])
    mes_nome = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
    }.get(n, f"{n} meses")
    return (
        "A taxa mensal contrafactual é a equivalente composta, constante "
        f"dentro do ano. Em {ano} o ano está incompleto (janeiro–{mes_nome}): "
        "o IPCA e a Selic contrafactual são os acumulados do período, com "
        f"o spread de {_fmt_pct(spread_pp)} × {n}/12 = "
        f"{_fmt_pct(spread_pp * n / 12.0, 3)}."
    )


def escrever_markdown(
    anual: pd.DataFrame,
    mensal: pd.DataFrame,
    path: Path,
    *,
    spread_pp: float,
    gerado_em: str,
    fonte_planilha: str,
    periodo: str | None = None,
    anos_observados: tuple[int, ...] | list[int] = (),
) -> None:
    sim = mensal[mensal["selic_am"].notna()]
    last = sim.iloc[-1]
    first = sim.iloc[0]
    economia_juros = float(
        sim["juros_selic_act"].sum() - sim["juros_selic_cf"].sum()
    )
    periodo_txt = periodo or (
        f"{first['mes'].strftime('%b/%Y')} a {last['mes'].strftime('%b/%Y')}"
    )
    last_ano = anual.iloc[-1] if not anual.empty else None
    meses_nome = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }
    mes_final_txt = (
        meses_nome.get(int(pd.Timestamp(last_ano["mes_final"]).month), "dezembro")
        if last_ano is not None
        else "dezembro"
    )
    linhas = [
        "# Simulação da DBGG — Selic anual = IPCA do ano + "
        f"{_fmt_pct(spread_pp)}",
        "",
        f"**Período:** {periodo_txt}.",
        f"**Gerado em:** {gerado_em}",
        "",
        "## Resultado no último mês",
        "",
        f"- DBGG observada: **{_fmt_tri(float(last['dbgg_act']))}** "
        f"({_fmt_bi(float(last['dbgg_act']))}).",
        f"- DBGG simulada: **{_fmt_tri(float(last['dbgg_cf']))}** "
        f"({_fmt_bi(float(last['dbgg_cf']))}).",
        f"- Diferença (observada − simulada): **{_fmt_tri(float(last['delta_dbgg']))}** "
        f"({_fmt_pct(100.0 * float(last['delta_dbgg']) / float(last['dbgg_act']))} "
        "da DBGG observada).",
        f"- Estoque Selic observado: {_fmt_tri(float(last['estoque_selic_act']))}.",
        f"- Estoque Selic simulado: {_fmt_tri(float(last['estoque_selic_cf']))}.",
        f"- Juros nominais da parcela Selic no período: observados "
        f"{_fmt_tri(float(sim['juros_selic_act'].sum()))}; "
        f"simulados {_fmt_tri(float(sim['juros_selic_cf'].sum()))}; "
        f"economia {_fmt_tri(economia_juros)}.",
    ]
    if pd.notna(last.get("dbgg_pib_act", np.nan)):
        linhas.extend(
            [
                f"- DBGG/PIB observada (SGS 4513): **{_fmt_pct(float(last['dbgg_pib_act']))}**.",
                f"- DBGG/PIB simulada (mesmo PIB): **{_fmt_pct(float(last['dbgg_pib_cf']))}** "
                f"(−{_fmt_pct(float(last['dbgg_pib_act']) - float(last['dbgg_pib_cf']))}).",
            ]
        )
    linhas.extend(
        [
            "",
            "## Metodologia",
            "",
            "Fonte da dívida: planilha especial do Banco Central "
            "[Dbggindexp.xlsx](" + URL_DBGG + ") — abas `DividaR$` (estoques), "
            "`JurosR$` (juros nominais mensais) e `PrimarioR$` (emissões líquidas). "
            f"Arquivo usado: `{fonte_planilha}`.",
            "",
            "A coluna **Selic** do BCB reúne LFT, LFT-A, LFT-B, operações "
            "compromissadas (recompra e revenda), dívida bancária e securitizadas "
            "indexadas à Selic.",
            "",
            f"Hipótese: em cada ano civil a Selic acumulada no ano é igual ao "
            f"IPCA acumulado no mesmo conjunto de meses **mais "
            f"{_fmt_pct(spread_pp)} proporcionais** (`spread × n/12`). "
            + _nota_ano_incompleto(anual, spread_pp)
            + (
                " Em "
                + ", ".join(str(a) for a in anos_observados)
                + " a Selic permanece a **observada** (não se aplica IPCA + "
                f"{_fmt_pct(spread_pp)})."
                if anos_observados
                else ""
            ),
            "",
            "Os juros Selic observados (contabilidade BCB) são reescalonados pela "
            "razão entre as taxas mensais (SGS 4390 vs. contrafactual) e pela "
            "razão dos estoques Selic do mês anterior (efeito de estoque menor). "
            "As emissões líquidas por indexador permanecem as observadas. O "
            "resíduo de fechamento do BCB "
            "(`Δestoque − emissão − juros`) é replicado no cenário simulado, "
            "para não atribuir quebras estatísticas à Selic.",
            "",
            "A DBGG simulada é a observada menos a diferença de estoque da "
            "parcela Selic. Não há efeitos de segunda ordem (PIB, câmbio, "
            "cupons dos prefixados, NTN-B, resultado primário ou composição "
            "das emissões). Em um equilíbrio geral os prefixados novos também "
            "sairiam mais baratos; esta simulação é, portanto, um **piso** "
            "para a redução da dívida.",
            "",
            "Séries auxiliares: IPCA SGS 433 (% a.m.); Selic SGS 4390 (% a.m.); "
            "DBGG/PIB SGS 4513 (quando disponível).",
            "",
            "## Série anual",
            "",
            "Valores de estoque e DBGG no **último mês** de cada ano na amostra "
            f"(dezembro; no último ano, {mes_final_txt}). Juros e IPCA/Selic "
            "são acumulados no ano. Unidades: R$ bilhões (estoques e juros) e "
            "% a.a. (ou % no período, quando n < 12).",
            "",
            "| Ano | n | IPCA | Selic obs. | Selic cf. | "
            "DBGG obs. | DBGG cf. | Δ DBGG | Juros Selic obs. | Juros cf. |",
            "|----:|--:|-----:|-----------:|----------:|"
            "-----------:|---------:|-------:|-----------------:|----------:|",
        ]
    )
    for row in anual.itertuples(index=False):
        linhas.append(
            "| {ano} | {n} | {ipca} | {selic} | {cf} | {dobs} | {dcf} | {delta} | "
            "{jobs} | {jcf} |".format(
                ano=int(row.ano),
                n=int(row.n_meses),
                ipca=_fmt_pct(float(row.ipca_acum_pct)),
                selic=_fmt_pct(float(row.selic_acum_pct)),
                cf=_fmt_pct(float(row.selic_cf_acum_pct)),
                dobs=_fmt_num(float(row.dbgg_act) / 1000.0),
                dcf=_fmt_num(float(row.dbgg_cf) / 1000.0),
                delta=_fmt_num(float(row.delta_dbgg) / 1000.0),
                jobs=_fmt_num(float(row.juros_selic_act) / 1000.0),
                jcf=_fmt_num(float(row.juros_selic_cf) / 1000.0),
            )
        )
    linhas.extend(
        [
            "",
            "Δ DBGG = observada − simulada (R$ bilhões). Juros em R$ bilhões.",
            "",
            "## Discriminativo das reduções por ano",
            "",
        ]
    )
    disc = montar_discriminativo(anual)
    linhas.extend(
        [
            "A **redução no ano** é a economia de juros nominais da parcela "
            "Selic naquele exercício (fluxo). A **redução acumulada** é a "
            "diferença de estoque da DBGG no último mês do ano. Participação "
            "= redução do ano / soma das reduções do período.",
            "",
            "| Ano | Selic obs. | Selic cf. | "
            "Juros Selic obs. | Juros cf. | "
            "Redução no ano | Redução acum. | Part. |",
            "|----:|-----------:|----------:|"
            "-----------------:|----------:|"
            "---------------:|--------------:|------:|",
        ]
    )
    for row in disc.itertuples(index=False):
        linhas.append(
            "| {ano} | {selic} | {cf} | {jobs} | {jcf} | {red} | {acum} | {part} |".format(
                ano=int(row.ano),
                selic=_fmt_pct(float(row.selic_acum_pct)),
                cf=_fmt_pct(float(row.selic_cf_acum_pct)),
                jobs=_fmt_num(float(row.juros_selic_act) / 1000.0),
                jcf=_fmt_num(float(row.juros_selic_cf) / 1000.0),
                red=_fmt_num(float(row.reducao_ano) / 1000.0),
                acum=_fmt_num(float(row.reducao_acumulada) / 1000.0),
                part=_fmt_pct(float(row.participacao_pct)),
            )
        )
    tot_red = float(disc["reducao_ano"].sum())
    tot_acum = float(disc["reducao_acumulada"].iloc[-1]) if len(disc) else 0.0
    tot_jobs = float(disc["juros_selic_act"].sum())
    tot_jcf = float(disc["juros_selic_cf"].sum())
    linhas.append(
        "| **Total** | | | {jobs} | {jcf} | {red} | {acum} | {part} |".format(
            jobs=_fmt_num(tot_jobs / 1000.0),
            jcf=_fmt_num(tot_jcf / 1000.0),
            red=_fmt_num(tot_red / 1000.0),
            acum=_fmt_num(tot_acum / 1000.0),
            part=_fmt_pct(100.0),
        )
    )
    linhas.extend(
        [
            "",
            "Valores em R$ bilhões. Sinal negativo = a Selic observada ficou "
            "abaixo de IPCA + spread (a simulação *aumenta* os juros naquele ano).",
            "",
        ]
    )
    if anos_observados:
        linhas.extend(
            [
                "Em "
                + ", ".join(str(a) for a in anos_observados)
                + " a Selic contrafactual é a **observada**: a redução "
                "daqueles anos vem só do estoque Selic menor herdado dos "
                "anos anteriores (mesma taxa, base menor).",
                "",
            ]
        )
    else:
        linhas.extend(
            [
                "Em 2020–2021 a Selic observada ficou **abaixo** de IPCA + "
                f"{_fmt_pct(spread_pp)} (ciclo de juros reais negativos). "
                "Nesses anos a simulação *aumenta* os juros da parcela Selic "
                "e a diferença de estoque recua — o que confere o sinal do "
                "exercício.",
                "",
            ]
        )
    linhas.extend(
        [
            "## Premissas e limitações",
            "",
            "- Só a remuneração da **parcela indexada à Selic** muda.",
            "- Emissões líquidas (aba `PrimarioR$`) ficam iguais às históricas: "
            "o Tesouro não reduz a colocação de LFT/compromissadas além do "
            "efeito automático dos juros capitalizados.",
            "- Prefixados, IPCA (NTN-B), câmbio, TR e TJLP/TLP não são "
            "reprecificados.",
            "- O PIB usado na razão DBGG/PIB simulada é o mesmo da série "
            "oficial (SGS 4513).",
            "",
        ]
    )
    path.write_text("\n".join(linhas), encoding="utf-8")


def escrever_discriminativo(
    disc: pd.DataFrame,
    path: Path,
    *,
    spread_pp: float,
    gerado_em: str,
    periodo: str = "janeiro/2007 a junho/2026",
    anos_observados: tuple[int, ...] | list[int] = (),
) -> None:
    """Markdown só com o discriminativo anual das reduções."""
    if disc.empty:
        path.write_text("# Discriminativo das reduções da DBGG\n\n(sem dados)\n")
        return
    tot_red = float(disc["reducao_ano"].sum())
    tot_acum = float(disc["reducao_acumulada"].iloc[-1])
    linhas = [
        "# Discriminativo das reduções da DBGG",
        "",
        f"**Hipótese:** Selic anual = IPCA do ano + {_fmt_pct(spread_pp)}"
        + (
            f", exceto {', '.join(str(a) for a in anos_observados)} "
            "(Selic observada)."
            if anos_observados
            else "."
        ),
        f"**Período:** {periodo}.",
        f"**Gerado em:** {gerado_em}",
        "",
        f"Redução acumulada ao final: **{_fmt_tri(tot_acum)}** "
        f"({_fmt_bi(tot_acum)}).",
        f"Soma das reduções anuais (economia de juros da parcela Selic): "
        f"**{_fmt_tri(tot_red)}**.",
        "",
        "A redução **no ano** é o fluxo (juros Selic observados − simulados). "
        "A redução **acumulada** é o estoque (DBGG observada − simulada) no "
        "último mês do ano. Participação = redução do ano / soma do período. "
        "Sinal negativo: naquele ano a Selic observada ficou abaixo de "
        f"IPCA + {_fmt_pct(spread_pp)}.",
        "",
        "| Ano | n | IPCA | Selic obs. | Selic cf. | "
        "Juros obs. | Juros cf. | Redução no ano | Redução acum. | Part. |",
        "|----:|--:|-----:|-----------:|----------:|"
        "-----------:|----------:|---------------:|--------------:|------:|",
    ]
    for row in disc.itertuples(index=False):
        linhas.append(
            "| {ano} | {n} | {ipca} | {selic} | {cf} | "
            "{jobs} | {jcf} | {red} | {acum} | {part} |".format(
                ano=int(row.ano),
                n=int(row.n_meses),
                ipca=_fmt_pct(float(row.ipca_acum_pct)),
                selic=_fmt_pct(float(row.selic_acum_pct)),
                cf=_fmt_pct(float(row.selic_cf_acum_pct)),
                jobs=_fmt_num(float(row.juros_selic_act) / 1000.0),
                jcf=_fmt_num(float(row.juros_selic_cf) / 1000.0),
                red=_fmt_num(float(row.reducao_ano) / 1000.0),
                acum=_fmt_num(float(row.reducao_acumulada) / 1000.0),
                part=_fmt_pct(float(row.participacao_pct)),
            )
        )
    linhas.append(
        "| **Total** | | | | | {jobs} | {jcf} | {red} | {acum} | 100,00% |".format(
            jobs=_fmt_num(float(disc["juros_selic_act"].sum()) / 1000.0),
            jcf=_fmt_num(float(disc["juros_selic_cf"].sum()) / 1000.0),
            red=_fmt_num(tot_red / 1000.0),
            acum=_fmt_num(tot_acum / 1000.0),
        )
    )
    linhas.extend(
        [
            "",
            "Unidades: taxas em % no período de cada ano; juros e reduções "
            "em R$ bilhões.",
            "",
        ]
    )
    if anos_observados:
        linhas.extend(
            [
                "Em "
                + ", ".join(str(a) for a in anos_observados)
                + " a Selic contrafactual é a **observada**: a redução "
                "daqueles anos vem só do estoque Selic menor herdado dos "
                "anos anteriores (mesma taxa, base menor).",
                "",
            ]
        )
    linhas.extend(
        [
            "2003–nov/2006: estoque Selic recuado com DPMFi × participação "
            "Over/Selic (SGS 4181 e 4177), emendado no primeiro mês da "
            "planilha oficial. Julho/2026, se ainda não publicado pelo BCB, "
            "é projetado com emissão líquida zero e juro = estoque × Selic "
            "do mês.",
            "",
        ]
    )
    path.write_text("\n".join(linhas), encoding="utf-8")


def gravar_grafico_discriminativo(
    disc: pd.DataFrame, pasta: Path, stem: str = "dbgg_selic_ipca_2007_2026"
) -> Path:
    import matplotlib.pyplot as plt

    d = disc.sort_values("ano")
    cores = ["#c45911" if v >= 0 else "#1f4e79" for v in d["reducao_ano"]]
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(d["ano"].astype(int).astype(str), d["reducao_ano"] / 1000.0, color=cores)
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_ylabel("R$ bilhões")
    ax.set_title("Redução da DBGG em cada ano (economia de juros da parcela Selic)")
    ax.grid(True, axis="y", alpha=0.3)
    fig.autofmt_xdate(rotation=45)
    fig.tight_layout()
    png = pasta / f"{stem}_discriminativo.png"
    fig.savefig(png, dpi=140)
    plt.close(fig)
    return png


def gravar_graficos(
    mensal: pd.DataFrame,
    anual: pd.DataFrame,
    pasta: Path,
    stem: str = "dbgg_selic_ipca_2007_2026",
) -> list[Path]:
    import matplotlib.pyplot as plt

    sim = mensal[mensal["selic_am"].notna()].copy()
    sim["dbgg_act_bi"] = sim["dbgg_act"] / 1000.0
    sim["dbgg_cf_bi"] = sim["dbgg_cf"] / 1000.0
    sim["delta_bi"] = sim["delta_dbgg"] / 1000.0

    caminhos: list[Path] = []
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    fig.suptitle(
        "DBGG: Selic observada vs. IPCA do ano + 0,37 p.p.",
        fontsize=12,
    )

    ax = axes[0, 0]
    ax.plot(sim["mes"], sim["dbgg_act_bi"], label="Observada", color="#1f4e79")
    ax.plot(sim["mes"], sim["dbgg_cf_bi"], label="Simulada", color="#c45911")
    ax.set_title("DBGG (R$ bilhões)")
    ax.legend()
    ax.grid(True, alpha=0.3)

    ax = axes[0, 1]
    ax.fill_between(sim["mes"], 0, sim["delta_bi"], color="#c45911", alpha=0.35)
    ax.plot(sim["mes"], sim["delta_bi"], color="#c45911")
    ax.set_title("Diferença acumulada (obs. − sim., R$ bi)")
    ax.grid(True, alpha=0.3)

    ax = axes[1, 0]
    ax.plot(
        anual["mes_final"],
        anual["selic_acum_pct"],
        label="Selic observada (acum. no ano)",
        color="#1f4e79",
        marker="o",
        markersize=3,
    )
    ax.plot(
        anual["mes_final"],
        anual["selic_cf_acum_pct"],
        label="IPCA + 0,37 p.p.",
        color="#c45911",
        marker="o",
        markersize=3,
    )
    ax.plot(
        anual["mes_final"],
        anual["ipca_acum_pct"],
        label="IPCA",
        color="#548235",
        linestyle="--",
    )
    ax.set_title("Taxas acumuladas no ano (%)")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)

    ax = axes[1, 1]
    ax.plot(
        sim["mes"],
        sim["juros_selic_act"].cumsum() / 1000.0,
        label="Juros Selic obs.",
        color="#1f4e79",
    )
    ax.plot(
        sim["mes"],
        sim["juros_selic_cf"].cumsum() / 1000.0,
        label="Juros Selic sim.",
        color="#c45911",
    )
    ax.set_title("Juros nominais da parcela Selic (acum., R$ bi)")
    ax.legend()
    ax.grid(True, alpha=0.3)

    fig.tight_layout()
    png = pasta / f"{stem}.png"
    fig.savefig(png, dpi=140)
    plt.close(fig)
    caminhos.append(png)

    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        figp = make_subplots(
            rows=2,
            cols=2,
            subplot_titles=(
                "DBGG (R$ bilhões)",
                "Diferença acumulada (R$ bi)",
                "Taxas acumuladas no ano (%)",
                "Juros Selic acumulados (R$ bi)",
            ),
        )
        figp.add_trace(
            go.Scatter(x=sim["mes"], y=sim["dbgg_act_bi"], name="DBGG observada"),
            row=1,
            col=1,
        )
        figp.add_trace(
            go.Scatter(x=sim["mes"], y=sim["dbgg_cf_bi"], name="DBGG simulada"),
            row=1,
            col=1,
        )
        figp.add_trace(
            go.Scatter(
                x=sim["mes"],
                y=sim["delta_bi"],
                name="Δ DBGG",
                fill="tozeroy",
            ),
            row=1,
            col=2,
        )
        figp.add_trace(
            go.Scatter(
                x=anual["mes_final"],
                y=anual["selic_acum_pct"],
                name="Selic obs.",
            ),
            row=2,
            col=1,
        )
        figp.add_trace(
            go.Scatter(
                x=anual["mes_final"],
                y=anual["selic_cf_acum_pct"],
                name="Selic cf.",
            ),
            row=2,
            col=1,
        )
        figp.add_trace(
            go.Scatter(
                x=sim["mes"],
                y=sim["juros_selic_act"].cumsum() / 1000.0,
                name="Juros obs.",
            ),
            row=2,
            col=2,
        )
        figp.add_trace(
            go.Scatter(
                x=sim["mes"],
                y=sim["juros_selic_cf"].cumsum() / 1000.0,
                name="Juros cf.",
            ),
            row=2,
            col=2,
        )
        figp.update_layout(
            title="DBGG: Selic observada vs. IPCA do ano + 0,37 p.p.",
            height=720,
            showlegend=True,
        )
        html = pasta / f"{stem}.html"
        figp.write_html(html, include_plotlyjs="cdn")
        caminhos.append(html)
    except Exception as exc:  # pragma: no cover - plotly opcional na prática
        print(f"[AVISO] Gráfico HTML não gerado: {exc}", flush=True)
    return caminhos


def gravar_saidas(
    mensal: pd.DataFrame,
    anual: pd.DataFrame,
    saida_dir: Path,
    *,
    spread_pp: float,
    fonte_planilha: str,
    stem: str = "dbgg_selic_ipca_2007_2026",
    periodo: str = "janeiro/2007 a junho/2026",
    anos_observados: tuple[int, ...] | list[int] = (),
) -> dict[str, Path]:
    saida_dir.mkdir(parents=True, exist_ok=True)
    csv_m = saida_dir / f"{stem}.csv"
    csv_a = saida_dir / f"{stem}_anual.csv"
    xlsx = saida_dir / f"{stem}.xlsx"
    md = saida_dir / f"{stem}.md"
    disc = montar_discriminativo(anual)
    csv_d = saida_dir / f"{stem}_discriminativo.csv"
    md_d = saida_dir / f"{stem}_discriminativo.md"
    mensal.to_csv(csv_m, index=False, float_format="%.8f")
    anual.to_csv(csv_a, index=False, float_format="%.8f")
    disc.to_csv(csv_d, index=False, float_format="%.8f")
    with pd.ExcelWriter(xlsx, engine="xlsxwriter") as writer:
        mensal.to_excel(writer, sheet_name="Mensal", index=False)
        anual.to_excel(writer, sheet_name="Anual", index=False)
        disc.to_excel(writer, sheet_name="Discriminativo", index=False)
        meta = pd.DataFrame(
            {
                "campo": [
                    "spread_pp",
                    "fonte",
                    "url",
                    "periodo",
                    "anos_observados",
                    "metodologia",
                ],
                "valor": [
                    spread_pp,
                    fonte_planilha,
                    URL_DBGG,
                    periodo,
                    ",".join(str(a) for a in anos_observados) or "—",
                    "Selic cf = IPCA do ano + spread; só parcela Selic",
                ],
            }
        )
        meta.to_excel(writer, sheet_name="Notas", index=False)
    gerado = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    escrever_markdown(
        anual,
        mensal,
        md,
        spread_pp=spread_pp,
        gerado_em=gerado,
        fonte_planilha=fonte_planilha,
        periodo=periodo,
        anos_observados=anos_observados,
    )
    escrever_discriminativo(
        disc,
        md_d,
        spread_pp=spread_pp,
        gerado_em=gerado,
        periodo=periodo,
        anos_observados=anos_observados,
    )
    graficos = gravar_graficos(mensal, anual, saida_dir, stem=stem)
    if not disc.empty:
        graficos.append(gravar_grafico_discriminativo(disc, saida_dir, stem=stem))
    out = {
        "mensal_csv": csv_m,
        "anual_csv": csv_a,
        "discriminativo_csv": csv_d,
        "discriminativo_md": md_d,
        "xlsx": xlsx,
        "md": md,
    }
    for i, p in enumerate(graficos):
        out[f"grafico_{i}"] = p
    return out


def processar(
    saida_dir: Path,
    *,
    planilha: Path | None = None,
    ipca_path: Path | None = None,
    selic_path: Path | None = None,
    spread_pp: float = SPREAD_DEFAULT,
    mes_inicio: pd.Timestamp = MES_INICIO_DEFAULT,
    mes_fim: pd.Timestamp = MES_FIM_DEFAULT,
    forcar_download: bool = False,
    baixar_pib: bool = True,
    anos_observados: tuple[int, ...] | list[int] = ANOS_OBSERVADOS_DEFAULT,
    recuar_pre_oficial: bool = True,
    projetar_apos_oficial: bool = True,
    stem: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    cache = ROOT / "data" / "Dbggindexp.xlsx"
    if planilha is not None and planilha.exists():
        path = planilha
    else:
        print("[1/4] Planilha BCB Dbggindexp.xlsx …", flush=True)
        path = baixar_planilha_dbgg(cache, forcar=forcar_download)

    print(f"[2/4] Lendo estoque, juros e emissões ({path}) …", flush=True)
    estoque, juros, emissoes = carregar_dbgg(path)
    mes0 = pd.Timestamp(mes_inicio).to_period("M").to_timestamp() - pd.DateOffset(
        months=1
    )
    sgs_ini = min(mes0, pd.Timestamp(estoque["mes"].min())).strftime("%d/%m/%Y")
    sgs_fim = pd.Timestamp(mes_fim).strftime("%d/%m/%Y")

    print("[3/4] IPCA (SGS 433) e Selic mensal (SGS 4390) …", flush=True)
    if ipca_path is not None:
        ipca = carregar_ipca(ipca_path)
    else:
        ipca = _baixar_sgs(IPCA_COD, inicio=sgs_ini, fim=sgs_fim)
        ipca = ipca.sort_values("mes").drop_duplicates("mes").reset_index(drop=True)
        ipca["fator"] = (1.0 + ipca["valor"] / 100.0).cumprod()
    if selic_path is not None:
        selic = carregar_selic_mensal(selic_path)
    else:
        selic = _baixar_sgs(SELIC_MENSAL_COD, inicio=sgs_ini, fim=sgs_fim)

    if recuar_pre_oficial and mes0 < pd.Timestamp(estoque["mes"].min()):
        print("     Recuando 2003–2006 via DPMFi Over/Selic (SGS 4181/4177) …", flush=True)
        dpmfi = _baixar_sgs(DPMFI_COD, inicio=sgs_ini, fim=sgs_fim)
        share = _baixar_sgs(DPMFI_SELIC_PCT_COD, inicio=sgs_ini, fim=sgs_fim)
        pib = _baixar_sgs(PIB_12M_COD, inicio=sgs_ini, fim=sgs_fim)
        dbgg_pib_s = _baixar_sgs(DBGG_PIB_COD, inicio=sgs_ini, fim=sgs_fim)
        r_est, r_ju, r_em = reconstruir_pre_oficial(
            estoque, selic, dpmfi, share, pib, dbgg_pib_s, mes_inicio
        )
        estoque, juros, emissoes = fundir_pre_oficial(
            r_est, r_ju, r_em, estoque, juros, emissoes
        )

    if projetar_apos_oficial and pd.Timestamp(mes_fim) > pd.Timestamp(juros["mes"].max()):
        print(
            f"     Projetando até {pd.Timestamp(mes_fim).strftime('%Y-%m')} "
            "(emissão líquida zero) …",
            flush=True,
        )
        estoque, juros, emissoes = projetar_meses_apos_oficial(
            estoque, juros, emissoes, selic, mes_fim
        )

    meses = pd.date_range(mes_inicio, mes_fim, freq="MS")
    ipca_anos = ipca_por_ano(ipca, pd.Series(meses))
    taxas_cf = taxas_mensais_cf(
        ipca_anos,
        pd.Series(meses),
        spread_pp,
        selic=selic,
        anos_observados=tuple(anos_observados),
    )

    print("[4/4] Simulando estoque Selic e DBGG …", flush=True)
    mensal = simular_parcela_selic(
        estoque,
        juros,
        emissoes,
        selic,
        taxas_cf,
        mes_inicio=mes_inicio,
        mes_fim=mes_fim,
    )

    dbgg_pib = None
    if baixar_pib:
        try:
            dbgg_pib = _baixar_sgs(DBGG_PIB_COD, inicio=sgs_ini, fim=sgs_fim)
        except Exception as exc:
            print(f"[AVISO] SGS 4513 (DBGG/PIB) indisponível: {exc}", flush=True)
    mensal = anexar_pib(mensal, dbgg_pib)
    anual = agregar_anual(mensal)

    if stem is None:
        stem = (
            f"dbgg_selic_ipca_{mes_inicio.year}_{mes_fim.year}"
            f"{mes_fim.month:02d}"
        )
        if anos_observados:
            stem += "_exc" + "".join(str(a)[-2:] for a in anos_observados)
    periodo = (
        f"{pd.Timestamp(mes_inicio).strftime('%b/%Y')} a "
        f"{pd.Timestamp(mes_fim).strftime('%b/%Y')}"
    )
    caminhos = gravar_saidas(
        mensal,
        anual,
        saida_dir,
        spread_pp=spread_pp,
        fonte_planilha=str(path),
        stem=stem,
        periodo=periodo,
        anos_observados=tuple(anos_observados),
    )
    for nome, p in caminhos.items():
        print(f"[OK] {nome}: {p}", flush=True)

    last = mensal[mensal["selic_am"].notna()].iloc[-1]
    print(
        f"DBGG {last['mes'].strftime('%Y-%m')}: "
        f"obs {_fmt_tri(float(last['dbgg_act']))} | "
        f"cf {_fmt_tri(float(last['dbgg_cf']))} | "
        f"Δ {_fmt_tri(float(last['delta_dbgg']))}",
        flush=True,
    )
    return mensal, anual


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--planilha", type=Path, default=None, help="Dbggindexp.xlsx local.")
    p.add_argument("--ipca", type=Path, default=None, help="Excel IPCA opcional.")
    p.add_argument("--selic", type=Path, default=None, help="selic_mensal.xlsx.")
    p.add_argument(
        "--spread",
        type=float,
        default=SPREAD_DEFAULT,
        help="Pontos percentuais somados ao IPCA do ano (default 0,37).",
    )
    p.add_argument(
        "--inicio",
        type=str,
        default="2007-01",
        help="Primeiro mês da simulação (YYYY-MM).",
    )
    p.add_argument(
        "--fim",
        type=str,
        default="2026-06",
        help="Último mês da simulação (YYYY-MM).",
    )
    p.add_argument(
        "--anos-observados",
        type=str,
        default="",
        help="Anos civis em que a Selic permanece a observada (ex.: 2020,2021).",
    )
    p.add_argument(
        "--stem",
        type=str,
        default=None,
        help="Prefixo dos arquivos de saída.",
    )
    p.add_argument(
        "--saida-dir",
        type=Path,
        default=ROOT / "output",
        help="Pasta das saídas.",
    )
    p.add_argument(
        "--forcar-download",
        action="store_true",
        help="Baixa de novo a planilha do BCB.",
    )
    p.add_argument(
        "--sem-pib",
        action="store_true",
        help="Não baixa SGS 4513 (DBGG/PIB).",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    inicio = pd.Timestamp(args.inicio).to_period("M").to_timestamp()
    fim = pd.Timestamp(args.fim).to_period("M").to_timestamp()
    try:
        anos_obs = tuple(
            int(x) for x in args.anos_observados.split(",") if x.strip()
        )
        processar(
            args.saida_dir,
            planilha=args.planilha,
            ipca_path=args.ipca,
            selic_path=args.selic,
            spread_pp=float(args.spread),
            mes_inicio=inicio,
            mes_fim=fim,
            forcar_download=args.forcar_download,
            baixar_pib=not args.sem_pib,
            anos_observados=anos_obs,
            stem=args.stem,
        )
    except Exception as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
