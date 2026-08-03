"""Totais por item (linhas 6–177) da aba 1.2-A, um total por período.

Para cada item em A6:A177, soma as colunas mensais do período, como:

    jan/97–dez/02 → =SUM(B6:BU6), =SUM(B7:BU7), ...
    jan/03–mai/16 → =SUM(BV6:HZ6), ...
    etc.

Uso::

    python somar_aba_1_2A_periodos.py "serie_historica_mai26 (2).xlsx" --out saida/totais_1_2A_por_item.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

PERIODS: list[tuple[str, datetime, datetime]] = [
    ("jan/97–dez/02", datetime(1997, 1, 1), datetime(2002, 12, 1)),
    ("jan/03–mai/16", datetime(2003, 1, 1), datetime(2016, 5, 1)),
    ("jun/16–dez/18", datetime(2016, 6, 1), datetime(2018, 12, 1)),
    ("jan/19–dez/22", datetime(2019, 1, 1), datetime(2022, 12, 1)),
    ("jan/23–mai/26", datetime(2023, 1, 1), datetime(2026, 5, 1)),
]

SHEET = "1.2-A"
ROW_START = 6
ROW_END = 177


def _load(path: str | Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Instale openpyxl: pip install openpyxl") from exc
    return load_workbook(path, read_only=True, data_only=True)


def _col_letter(idx0: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(idx0 + 1)


def totals_por_item(path: str | Path) -> dict[str, Any]:
    """Retorna 172 itens × 5 períodos (soma mensal por linha)."""
    wb = _load(path)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Aba '{SHEET}' nao encontrada. Disponiveis: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[4]
    unit = rows[2][0] if len(rows) > 2 else None

    period_meta = []
    period_cols: list[list[int]] = []
    for lab, start, end in PERIODS:
        cols = [
            col
            for col, h in enumerate(headers)
            if col > 0 and isinstance(h, datetime) and start <= h <= end
        ]
        if not cols:
            raise ValueError(f"Nenhuma coluna mensal para periodo {lab}")
        period_cols.append(cols)
        period_meta.append(
            {
                "periodo": lab,
                "inicio": start.strftime("%Y-%m"),
                "fim": end.strftime("%Y-%m"),
                "meses": len(cols),
                "coluna_inicio": _col_letter(cols[0]),
                "coluna_fim": _col_letter(cols[-1]),
                "formula_exemplo_linha6": (
                    f"=SUM({_col_letter(cols[0])}6:{_col_letter(cols[-1])}6)"
                ),
            }
        )

    itens: list[dict[str, Any]] = []
    for ridx in range(ROW_START - 1, ROW_END):
        row = rows[ridx]
        label = row[0]
        rec: dict[str, Any] = {
            "linha": ridx + 1,
            "item": label,
        }
        for pi, (lab, _s, _e) in enumerate(PERIODS):
            s = 0.0
            for c in period_cols[pi]:
                v = row[c] if c < len(row) else None
                if v is None:
                    continue
                try:
                    s += float(v)
                except (TypeError, ValueError):
                    continue
            rec[lab] = round(s, 2)
            rec[f"{lab}_R$bi"] = round(s / 1000.0, 2)
        itens.append(rec)

    return {
        "path": str(path),
        "sheet": SHEET,
        "unit": unit,
        "row_start": ROW_START,
        "row_end": ROW_END,
        "n_itens": len(itens),
        "periodos": period_meta,
        "itens": itens,
        "explicacao": (
            "Cada celula = soma dos meses do periodo naquela linha "
            "(ex.: jan/97–dez/02 na linha 6 = SUM(B6:BU6))."
        ),
    }


# alias legado
sum_periods = totals_por_item


def write_outputs(table: dict[str, Any], out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    period_labels = [p[0] for p in PERIODS]

    if out.suffix.lower() == ".csv":
        fieldnames = ["linha", "item"] + period_labels + [f"{p}_R$bi" for p in period_labels]
        with out.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            w.writeheader()
            w.writerows(table["itens"])
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    xwb = Workbook()
    ws = xwb.active
    ws.title = "Totais_por_item"
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    ws["A1"] = "Aba 1.2-A — total por item (linhas 6–177) em cada período"
    ws["A1"].font = bold
    ws["A2"] = table.get("unit") or "R$ milhões — IPCA"
    ws["A3"] = table["explicacao"]
    ws["A4"] = f"Fonte: {table['path']}"

    # Period column map (for transparency)
    ws["A5"] = "Mapeamento das colunas na planilha original:"
    ws["A5"].font = bold
    for i, meta in enumerate(table["periodos"]):
        ws.cell(
            6 + i,
            1,
            (
                f"{i + 1}) {meta['periodo']}: "
                f"{meta['coluna_inicio']}..{meta['coluna_fim']} "
                f"({meta['meses']} meses)  ex. linha 6 → {meta['formula_exemplo_linha6']}"
            ),
        )

    header_row = 12
    headers = ["linha", "item"] + [f"{p} (R$ mi)" for p in period_labels] + [
        f"{p} (R$ bi)" for p in period_labels
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        cell.font = bold
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    for i, item in enumerate(table["itens"]):
        r = header_row + 1 + i
        ws.cell(r, 1, item["linha"]).border = thin
        ws.cell(r, 2, item["item"]).border = thin
        for j, p in enumerate(period_labels):
            cell = ws.cell(r, 3 + j, item[p])
            cell.border = thin
            cell.number_format = "#,##0.00"
        for j, p in enumerate(period_labels):
            cell = ws.cell(r, 3 + len(period_labels) + j, item[f"{p}_R$bi"])
            cell.border = thin
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 72
    for col_idx in range(3, 3 + 2 * len(period_labels)):
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.freeze_panes = "C13"
    ws.auto_filter.ref = f"A{header_row}:L{header_row + len(table['itens'])}"

    # Small preview sheet with first items
    ws2 = xwb.create_sheet("Periodos")
    ws2["A1"] = "Períodos e intervalos de colunas"
    ws2["A1"].font = bold
    for c, h in enumerate(
        ["#", "periodo", "inicio", "fim", "meses", "col_ini", "col_fim", "formula_linha6"], 1
    ):
        ws2.cell(3, c, h).font = bold
    for i, meta in enumerate(table["periodos"], 1):
        ws2.cell(3 + i, 1, i)
        ws2.cell(3 + i, 2, meta["periodo"])
        ws2.cell(3 + i, 3, meta["inicio"])
        ws2.cell(3 + i, 4, meta["fim"])
        ws2.cell(3 + i, 5, meta["meses"])
        ws2.cell(3 + i, 6, meta["coluna_inicio"])
        ws2.cell(3 + i, 7, meta["coluna_fim"])
        ws2.cell(3 + i, 8, meta["formula_exemplo_linha6"])
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["H"].width = 22

    xwb.save(out)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Totais por item (A6:A177) da aba 1.2-A em cada periodo"
    )
    p.add_argument("path", help="Caminho do serie_historica_*.xlsx")
    p.add_argument("--out", default="totais_1_2A_por_item.xlsx")
    args = p.parse_args(argv)
    path = Path(args.path)
    if not path.exists():
        print(json.dumps({"error": f"arquivo nao encontrado: {path}"}, ensure_ascii=False))
        return 1
    try:
        table = totals_por_item(path)
        out = Path(args.out)
        write_outputs(table, out)
        preview = table["itens"][:3]
        print(
            json.dumps(
                {
                    "ok": True,
                    "out": str(out.resolve()),
                    "unit": table["unit"],
                    "n_itens": table["n_itens"],
                    "periodos": table["periodos"],
                    "explicacao": table["explicacao"],
                    "preview_itens": preview,
                },
                indent=2,
                ensure_ascii=False,
            )
        )
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
