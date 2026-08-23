"""Taxas médias anuais de crédito (SGS) e 5 maiores bancos por modalidade (IF.data + Olinda).

Fontes do Banco Central:
  - SGS: taxa média das novas operações, ponderada pelas concessões (% a.a.)
  - IF.data: saldo da carteira ativa por instituição e modalidade
  - Olinda TaxasJuros: taxa média praticada por IF (ponderada pelo valor contratado)

Uso:
  python3 scripts/taxas_credito_bcb.py
  python3 scripts/taxas_credito_bcb.py --sem-download
"""

from __future__ import annotations

import argparse
import ast
import sys
import time
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DATA_DIR = ROOT / "data" / "credito_bcb"
OUTPUT_DIR = ROOT / "output"
SGS_REST = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{cod}/dados"
IFDATA = "https://olinda.bcb.gov.br/olinda/servico/IFDATA/versao/v1/odata"
OLINDA = "https://olinda.bcb.gov.br/olinda/servico/taxaJuros/versao/v2/odata"
ANO_INI = 2001
ANO_FIM = 2026

# Séries oficiais em % a.a. (Nota para Imprensa de crédito / SGS).
SERIES = [
    (20714, "Total do SFN", "Total", "Livre e direcionado"),
    (20715, "Total PJ", "PJ", "Livre e direcionado"),
    (20716, "Total PF", "PF", "Livre e direcionado"),
    (20717, "Recursos livres — total", "Total", "Livres"),
    (20718, "Livres PJ — total", "PJ", "Livres"),
    (20719, "PJ — desconto de duplicatas e recebíveis", "PJ", "Livres"),
    (20720, "PJ — desconto de cheques", "PJ", "Livres"),
    (20721, "PJ — antecipação de faturas de cartão", "PJ", "Livres"),
    (20722, "PJ — capital de giro até 365 dias", "PJ", "Livres"),
    (20723, "PJ — capital de giro acima de 365 dias", "PJ", "Livres"),
    (20724, "PJ — capital de giro rotativo", "PJ", "Livres"),
    (20725, "PJ — capital de giro total", "PJ", "Livres"),
    (20726, "PJ — conta garantida", "PJ", "Livres"),
    (20727, "PJ — cheque especial", "PJ", "Livres"),
    (20728, "PJ — aquisição de veículos", "PJ", "Livres"),
    (20740, "Livres PF — total", "PF", "Livres"),
    (20741, "PF — cheque especial", "PF", "Livres"),
    (20742, "PF — crédito pessoal não consignado", "PF", "Livres"),
    (20747, "PF — crédito pessoal consignado total", "PF", "Livres"),
    (20748, "PF — crédito pessoal total", "PF", "Livres"),
    (20749, "PF — aquisição de veículos", "PF", "Livres"),
    (20750, "PF — aquisição de outros bens", "PF", "Livres"),
    (20756, "Recursos direcionados — total", "Total", "Direcionados"),
    (20760, "PJ — crédito rural total", "PJ", "Direcionados"),
    (20763, "PJ — financiamento imobiliário total", "PJ", "Direcionados"),
    (20768, "Direcionados PF — total", "PF", "Direcionados"),
    (20771, "PF — crédito rural total", "PF", "Direcionados"),
    (20774, "PF — financiamento imobiliário total", "PF", "Direcionados"),
    (20782, "PF — microcrédito total", "PF", "Direcionados"),
    (22021, "PJ — cartão de crédito total", "PJ", "Livres"),
    (22022, "PF — cartão de crédito rotativo", "PF", "Livres"),
    (22023, "PF — cartão de crédito parcelado", "PF", "Livres"),
    (22024, "PF — cartão de crédito total", "PF", "Livres"),
]

MAPA_OLINDA = {
    "Cartão de Crédito": [
        "Cartão de crédito - rotativo total - Prefixado",
        "Cartão de crédito - parcelado - Prefixado",
    ],
    "Empréstimo com Consignação em Folha": [
        "Crédito pessoal consignado INSS - Prefixado",
        "Crédito pessoal consignado privado - Prefixado",
        "Crédito pessoal consignado público - Prefixado",
    ],
    "Empréstimo sem Consignação em Folha": [
        "Crédito pessoal não consignado - Prefixado",
        "Cheque especial - Prefixado",
    ],
    "Habitação": [
        "Financiamento imobiliário com taxas de mercado - Prefixado",
        "Financiamento imobiliário com taxas de mercado - Pós-fixado referenciado em TR",
    ],
    "Veículos": ["Aquisição de veículos - Prefixado"],
    "Outros Créditos": ["Aquisição de outros bens - Prefixado", "Desconto de cheques - Prefixado"],
    "Capital de Giro": [
        "Capital de giro com prazo até 365 dias - Prefixado",
        "Capital de giro com prazo superior a 365 dias - Prefixado",
    ],
    "Cheque Especial e Conta Garantida": [
        "Cheque especial - Prefixado",
        "Conta garantida - Prefixado",
    ],
    "Operações com Recebíveis": [
        "Desconto de duplicatas - Prefixado",
        "Antecipação de faturas de cartão de crédito - Prefixado",
    ],
    "Comércio Exterior": [
        "Adiantamento sobre contratos de câmbio (ACC) - Pós-fixado referenciado em moeda estrangeira"
    ],
}


def _borda() -> Border:
    lado = Side(style="thin", color="1A1A1A")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def media_anual(mensal: pd.Series) -> pd.Series:
    """Média aritmética das taxas mensais oficiais (% a.a.) em cada ano-calendário."""
    s = pd.to_numeric(mensal, errors="coerce")
    return s.groupby(s.index.year).mean()


def baixar_sgs(cod: int, inicio: str, fim: str) -> pd.DataFrame:
    resp = requests.get(
        SGS_REST.format(cod=cod),
        params={"formato": "json", "dataInicial": inicio, "dataFinal": fim},
        timeout=90,
    )
    resp.raise_for_status()
    dados = resp.json()
    if not dados:
        return pd.DataFrame(columns=["data", "taxa"])
    out = pd.DataFrame(dados)
    out["data"] = pd.to_datetime(out["data"], dayfirst=True)
    out["taxa"] = pd.to_numeric(out["valor"], errors="coerce")
    return out[["data", "taxa"]].dropna()


def carregar_series_sgs(cache_dir: Path, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / "sgs_taxas_modalidade.csv"
    if cache.exists():
        return pd.read_csv(cache, parse_dates=["data"])
    if not baixar:
        raise FileNotFoundError(cache)
    cache_dir.mkdir(parents=True, exist_ok=True)
    partes = []
    for cod, nome, segmento, origem in SERIES:
        print(f"  SGS {cod} {nome}", flush=True)
        try:
            df = baixar_sgs(cod, "01/01/2001", "01/12/2026")
        except Exception as exc:  # noqa: BLE001
            print(f"    falha {cod}: {exc}", flush=True)
            continue
        if df.empty:
            continue
        df["codigo"] = cod
        df["modalidade"] = nome
        df["segmento"] = segmento
        df["origem"] = origem
        partes.append(df)
        time.sleep(0.05)
    if not partes:
        raise RuntimeError("Nenhuma série SGS de crédito foi baixada.")
    out = pd.concat(partes, ignore_index=True)
    out.to_csv(cache, index=False)
    return out


def taxas_medias_anuais(mensal: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for (cod, nome, seg, origem), g in mensal.groupby(["codigo", "modalidade", "segmento", "origem"], sort=False):
        s = g.set_index("data")["taxa"].sort_index()
        anual = media_anual(s)
        rec = {
            "codigo": int(cod),
            "modalidade": nome,
            "segmento": seg,
            "origem": origem,
            "inicio": s.index.min().strftime("%m/%Y"),
            "fim": s.index.max().strftime("%m/%Y"),
        }
        for ano in range(ANO_INI, ANO_FIM + 1):
            rec[str(ano)] = float(anual[ano]) if ano in anual.index else float("nan")
        linhas.append(rec)
    return pd.DataFrame(linhas)


def _get_json(url: str, tentativas: int = 4) -> dict:
    ultimo = None
    for i in range(tentativas):
        try:
            resp = requests.get(url, timeout=180)
            resp.raise_for_status()
            return resp.json()
        except Exception as exc:  # noqa: BLE001
            ultimo = exc
            time.sleep(2 * (i + 1))
    raise RuntimeError(f"Falha ao consultar {url}: {ultimo}")


def ultimo_anomes_ifdata(relatorio: str = "11") -> int:
    for am in (202603, 202512, 202509, 202506, 202503):
        url = (
            f"{IFDATA}/IfDataValores(AnoMes=@AnoMes,TipoInstituicao=@TipoInstituicao,Relatorio=@Relatorio)"
            f"?@AnoMes={am}&@TipoInstituicao=1&@Relatorio='{relatorio}'&$top=1&$format=json"
        )
        try:
            if _get_json(url).get("value"):
                return am
        except Exception:
            continue
    raise RuntimeError("Nenhum trimestre IF.data disponível.")


def carregar_ifdata(cache_dir: Path, anomes: int, relatorio: str, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / f"ifdata_{relatorio}_{anomes}.csv"
    if cache.exists():
        return pd.read_csv(cache)
    if not baixar:
        raise FileNotFoundError(cache)
    url = (
        f"{IFDATA}/IfDataValores(AnoMes=@AnoMes,TipoInstituicao=@TipoInstituicao,Relatorio=@Relatorio)"
        f"?@AnoMes={anomes}&@TipoInstituicao=1&@Relatorio='{relatorio}'"
        f"&$filter=NomeColuna eq 'Total'&$format=json&$top=100000"
    )
    print(f"  IF.data rel {relatorio} {anomes}", flush=True)
    dados = _get_json(url).get("value", [])
    if not dados:
        url = (
            f"{IFDATA}/IfDataValores(AnoMes=@AnoMes,TipoInstituicao=@TipoInstituicao,Relatorio=@Relatorio)"
            f"?@AnoMes={anomes}&@TipoInstituicao=1&@Relatorio='{relatorio}'&$format=json&$top=100000"
        )
        dados = _get_json(url).get("value", [])
    df = pd.DataFrame(dados)
    cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache, index=False)
    return df


def carregar_cadastro(cache_dir: Path, anomes: int, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / f"ifdata_cadastro_{anomes}.csv"
    if cache.exists():
        return pd.read_csv(cache, dtype={"CodInst": str, "CnpjInstituicaoLider": str, "CodConglomeradoPrudencial": str})
    if not baixar:
        raise FileNotFoundError(cache)
    url = f"{IFDATA}/IfDataCadastro(AnoMes=@AnoMes)?@AnoMes={anomes}&$format=json&$top=20000"
    print(f"  IF.data cadastro {anomes}", flush=True)
    df = pd.DataFrame(_get_json(url).get("value", []))
    cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache, index=False)
    return df


def consolidadores(cadastro: pd.DataFrame) -> pd.DataFrame:
    cad = cadastro.copy()
    cad["CodInst"] = cad["CodInst"].astype(str)
    cad["CodConglomeradoPrudencial"] = cad["CodConglomeradoPrudencial"].astype(str)
    cad["CnpjInstituicaoLider"] = cad["CnpjInstituicaoLider"].fillna("").astype(str).str.replace(r"\D", "", regex=True).str.zfill(8)
    cong = cad[cad["CodInst"] == cad["CodConglomeradoPrudencial"]].copy()
    membros = (
        cad[cad["CodInst"].str.fullmatch(r"\d{8}")]
        .groupby("CodConglomeradoPrudencial")["CodInst"]
        .apply(lambda s: sorted(set(s.astype(str))))
        .to_dict()
    )
    cong["cnpjs"] = cong.apply(
        lambda r: sorted(set((membros.get(r["CodInst"], []) + [r["CnpjInstituicaoLider"]]))) ,
        axis=1,
    )
    return cong


def top5_por_grupo(ifdata: pd.DataFrame, cadastro: pd.DataFrame, segmento: str) -> pd.DataFrame:
    tot = ifdata.copy()
    tot = tot[tot["NomeColuna"].astype(str).str.lower() == "total"]
    tot = tot[tot["Grupo"].notna()]
    tot["Saldo"] = pd.to_numeric(tot["Saldo"], errors="coerce")
    tot["CodInst"] = tot["CodInst"].astype(str)
    cong = consolidadores(cadastro)
    tot = tot.merge(cong[["CodInst", "NomeInstituicao", "CnpjInstituicaoLider", "cnpjs"]], on="CodInst", how="inner")
    tot = tot.sort_values("Saldo", ascending=False).drop_duplicates(["Grupo", "CodInst"])
    linhas = []
    for grupo, g in tot.groupby("Grupo"):
        top = g.sort_values("Saldo", ascending=False).head(5)
        for i, rec in enumerate(top.itertuples(index=False), start=1):
            linhas.append(
                {
                    "segmento": segmento,
                    "modalidade_ifdata": grupo,
                    "rank": i,
                    "instituicao": rec.NomeInstituicao,
                    "cod_inst": rec.CodInst,
                    "cnpj_lider": rec.CnpjInstituicaoLider,
                    "cnpjs": rec.cnpjs,
                    "saldo": float(rec.Saldo),
                }
            )
    return pd.DataFrame(linhas)


def carregar_olinda_mes(cache_dir: Path, ano_mes: str, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / f"olinda_{ano_mes}.csv"
    if cache.exists():
        return pd.read_csv(cache, dtype={"cnpj8": str})
    if not baixar:
        raise FileNotFoundError(cache)
    url = (
        f"{OLINDA}/TaxasJurosMensalPorMes?$format=json&$top=10000"
        f"&$filter=anoMes eq '{ano_mes}'"
    )
    print(f"  Olinda mensal {ano_mes}", flush=True)
    df = pd.DataFrame(_get_json(url).get("value", []))
    if not df.empty:
        df["cnpj8"] = df["cnpj8"].astype(str).str.zfill(8)
    cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache, index=False)
    return df


def ultimo_mes_olinda() -> str:
    url = f"{OLINDA}/TaxasJurosMensalPorMes?$format=json&$top=1"
    return str(_get_json(url)["value"][0]["anoMes"])


def ultimo_periodo_diario_olinda() -> tuple[str, str]:
    url = f"{OLINDA}/ConsultaDatas?$format=json&$top=8000"
    df = pd.DataFrame(_get_json(url).get("value", []))
    d = df[df["tipoModalidade"] == "D"].sort_values("inicioPeriodo")
    ult = d.iloc[-1]
    return str(ult["inicioPeriodo"]), str(ult["fimPeriodo"])


def carregar_olinda_diario(cache_dir: Path, inicio: str, baixar: bool = True) -> pd.DataFrame:
    cache = cache_dir / f"olinda_diario_{inicio}.csv"
    if cache.exists():
        return pd.read_csv(cache, dtype={"cnpj8": str})
    if not baixar:
        raise FileNotFoundError(cache)
    url = (
        f"{OLINDA}/TaxasJurosDiariaPorInicioPeriodo?$format=json&$top=10000"
        f"&$filter=InicioPeriodo eq '{inicio}'"
    )
    print(f"  Olinda diário {inicio}", flush=True)
    df = pd.DataFrame(_get_json(url).get("value", []))
    if not df.empty:
        df["cnpj8"] = df["cnpj8"].astype(str).str.zfill(8)
    cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(cache, index=False)
    return df


def carregar_olinda(cache_dir: Path, baixar: bool = True) -> tuple[pd.DataFrame, str]:
    """Junta a tabela mensal (imobiliário) com a diária (demais modalidades)."""
    mes = ultimo_mes_olinda()
    ini, fim = ultimo_periodo_diario_olinda()
    mensal = carregar_olinda_mes(cache_dir, mes, baixar=baixar)
    diario = carregar_olinda_diario(cache_dir, ini, baixar=baixar)
    partes = [p for p in (mensal, diario) if not p.empty]
    out = pd.concat(partes, ignore_index=True)
    rotulo = f"{ini} a {fim} e {mes} (imobiliário)"
    return out, rotulo


_ALIASES = (
    ("ITAU", ("ITAU", "ITAÚ")),
    ("ITAÚ", ("ITAU", "ITAÚ")),
    ("BRADESCO", ("BRADESCO",)),
    ("BB -", ("BCO DO BRASIL", "BANCO DO BRASIL")),
    ("CAIXA", ("CAIXA ECONOMICA", "CAIXA ECONÔMICA")),
    ("SANTANDER", ("SANTANDER",)),
    ("NU PAGAMENTOS", ("NU FINANCEIRA", "NUBANK")),
    ("BTG", ("BTG",)),
    ("VOTORANTIM", ("VOTORANTIM", "BV")),
)


def cruzar_taxas(top5: pd.DataFrame, olinda: pd.DataFrame) -> pd.DataFrame:
    linhas = []
    for rec in top5.itertuples(index=False):
        mods = MAPA_OLINDA.get(rec.modalidade_ifdata, [])
        raw = rec.cnpjs
        if isinstance(raw, list):
            cnpjs = {str(x).zfill(8) for x in raw}
        elif isinstance(raw, str) and raw.startswith("["):
            try:
                cnpjs = {str(x).zfill(8) for x in ast.literal_eval(raw)}
            except (ValueError, SyntaxError):
                cnpjs = {str(rec.cnpj_lider).zfill(8)}
        else:
            cnpjs = {str(rec.cnpj_lider).zfill(8)}
        nome_if = str(rec.instituicao).upper()
        tokens = []
        for chave, alts in _ALIASES:
            if chave in nome_if:
                tokens.extend(alts)
        if not mods:
            linhas.append({**rec._asdict(), "modalidade_olinda": "—", "if_olinda": "—", "taxa_aa": float("nan")})
            continue
        achou = False
        for mod in mods:
            base = olinda[olinda["Modalidade"] == mod]
            if "Segmento" in base.columns:
                alvo = "FÍSICA" if rec.segmento == "PF" else "JURÍDICA"
                tem = base["Segmento"].notna()
                base = base[~tem | base["Segmento"].astype(str).str.contains(alvo, case=False, na=False)]
            cand = base[base["cnpj8"].isin(cnpjs)]
            if cand.empty and tokens:
                mask = False
                for tok in tokens:
                    mask = mask | base["InstituicaoFinanceira"].str.contains(tok, case=False, na=False)
                cand = base[mask]
            if cand.empty:
                continue
            # Se várias IFs do conglomerado reportam, usa a mediana das taxas.
            linhas.append(
                {
                    **{k: getattr(rec, k) for k in rec._fields if k != "cnpjs"},
                    "modalidade_olinda": mod,
                    "if_olinda": " / ".join(sorted(cand["InstituicaoFinanceira"].unique())),
                    "taxa_aa": float(cand["TaxaJurosAoAno"].median()),
                }
            )
            achou = True
        if not achou:
            linhas.append(
                {
                    **{k: getattr(rec, k) for k in rec._fields if k != "cnpjs"},
                    "modalidade_olinda": "sem publicação Olinda para o conglomerado",
                    "if_olinda": "—",
                    "taxa_aa": float("nan"),
                }
            )
    return pd.DataFrame(linhas)


def _fmt(valor: float, casas: int = 2) -> str:
    if valor is None or pd.isna(valor):
        return "—"
    txt = f"{float(valor):,.{casas}f}"
    return txt.replace(",", "X").replace(".", ",").replace("X", ".")


def _escrever_aba(ws, cabecalhos: list[str], linhas: list[list], larguras: list[int] | None = None) -> None:
    borda = _borda()
    fill_cab = PatternFill("solid", fgColor="E8E8E8")
    fill_alt = PatternFill("solid", fgColor="F4F4F4")
    fonte_cab = Font(name="Calibri", size=10, bold=True)
    fonte = Font(name="Calibri", size=9)
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(1, col, cab)
        cell.font = fonte_cab
        cell.fill = fill_cab
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, linha in enumerate(linhas):
        for col, valor in enumerate(linha, start=1):
            cell = ws.cell(i + 2, col, valor)
            cell.font = fonte
            cell.border = borda
            if isinstance(valor, float):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(valor, int):
                cell.alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                cell.fill = fill_alt
    for col in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[get_column_letter(col)].width = (larguras[col - 1] if larguras else 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(linhas) + 1}"


def gerar_planilha(anual: pd.DataFrame, top5_taxas: pd.DataFrame, anomes: int, mes_olinda: str, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Notas"
    ws0["A1"] = "Taxas de juros do crédito no Brasil — estatísticas do Banco Central (2001–2026)"
    ws0["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    notas = [
        "Aba Taxas_anuais: média aritmética, em cada ano-calendário, da taxa média mensal oficial "
        "das novas operações (SGS, % a.a., ponderada pelo valor das concessões).",
        "A metodologia atual das estatísticas de crédito começa em março de 2011 para o total do SFN. "
        "Algumas modalidades (cheque especial, veículos, consignado) têm série mais longa, desde 2000/2004.",
        "Aba Top5_bancos: os 5 conglomerados prudenciais de maior saldo da carteira ativa em cada "
        f"modalidade do IF.data (data-base {anomes}). O BCB não publica concessões por IF×modalidade; "
        "o ranking valorativo usa o estoque oficial da carteira.",
        f"As taxas dos 5 maiores vêm da publicação Olinda de taxas por instituição ({mes_olinda}), "
        "média ponderada pelo valor contratado. Quando várias IFs do conglomerado reportam, usa-se a mediana.",
        "Fontes: https://www.bcb.gov.br/estatisticas/estatisticasmonetariascredito  |  "
        "https://www3.bcb.gov.br/ifdata/  |  https://www.bcb.gov.br/estatisticas/txjuros",
    ]
    for i, txt in enumerate(notas, start=3):
        ws0[f"A{i}"] = txt
        ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws0[f"A{i}"].alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[i].height = 36
    ws0.column_dimensions["A"].width = 28

    anos = [str(a) for a in range(ANO_INI, ANO_FIM + 1)]
    cabs = ["Código SGS", "Modalidade", "Segmento", "Origem", "Início da série", "Fim da série"] + anos
    linhas = []
    for rec in anual.to_dict("records"):
        lin = [int(rec["codigo"]), rec["modalidade"], rec["segmento"], rec["origem"], rec["inicio"], rec["fim"]]
        for a in anos:
            v = rec.get(a)
            lin.append(None if v is None or pd.isna(v) else float(v))
        linhas.append(lin)
    ws1 = wb.create_sheet("Taxas_anuais")
    _escrever_aba(ws1, cabs, linhas, [12, 46, 10, 18, 14, 12] + [9] * len(anos))

    cabs2 = [
        "Segmento",
        "Modalidade (IF.data)",
        "Rank",
        "Conglomerado",
        "Saldo carteira (R$)",
        "Modalidade (Olinda)",
        "IF que reportou a taxa",
        f"Taxa média % a.a. ({mes_olinda})",
    ]
    linhas2 = []
    for rec in top5_taxas.itertuples(index=False):
        linhas2.append(
            [
                rec.segmento,
                rec.modalidade_ifdata,
                int(rec.rank),
                rec.instituicao,
                float(rec.saldo),
                rec.modalidade_olinda,
                rec.if_olinda,
                None if pd.isna(rec.taxa_aa) else float(rec.taxa_aa),
            ]
        )
    ws2 = wb.create_sheet("Top5_bancos")
    _escrever_aba(ws2, cabs2, linhas2, [10, 36, 8, 36, 20, 52, 40, 18])
    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR / "taxas_credito_bcb_2001_2026.xlsx")
    parser.add_argument("--sem-download", action="store_true")
    args = parser.parse_args(argv)

    print("Baixando séries SGS de taxas de crédito...", flush=True)
    mensal = carregar_series_sgs(args.cache_dir, baixar=not args.sem_download)
    anual = taxas_medias_anuais(mensal)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    anual.to_csv(args.output.parent / "taxas_credito_medias_anuais_2001_2026.csv", index=False)

    print("IF.data — ranking valorativo...", flush=True)
    anomes = ultimo_anomes_ifdata()
    cad = carregar_cadastro(args.cache_dir, anomes, baixar=not args.sem_download)
    pf = carregar_ifdata(args.cache_dir, anomes, "11", baixar=not args.sem_download)
    pj = carregar_ifdata(args.cache_dir, anomes, "13", baixar=not args.sem_download)
    top = pd.concat(
        [top5_por_grupo(pf, cad, "PF"), top5_por_grupo(pj, cad, "PJ")],
        ignore_index=True,
    )

    olinda, mes = carregar_olinda(args.cache_dir, baixar=not args.sem_download)
    cruzado = cruzar_taxas(top, olinda)
    cruzado.drop(columns=["cnpjs"], errors="ignore").to_csv(
        args.output.parent / "taxas_credito_top5_bancos.csv", index=False
    )

    path = gerar_planilha(anual, cruzado, anomes, mes, args.output)
    print(f"Planilha: {path}")
    print(anual[["modalidade", "inicio", "fim"]].to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
