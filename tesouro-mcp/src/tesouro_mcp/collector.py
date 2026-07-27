"""Annual fiscal/debt collector: RTN + DBGG + DPF emissions/redemptions + BNDES.

Optional merge of user-provided CSVs for DGT (renuncias) and FNO/FNE/FCO.
"""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import httpx

from . import bcb_client, providers

EMISSOES_XLSX_URL = (
    "https://www.tesourotransparente.gov.br/ckan/dataset/"
    "48a31b6b-034a-473b-be7d-228be30437a1/resource/"
    "bf69babd-ac07-40ce-90ff-c8e07ec8c8bf/download/"
    "emissoes-e-resgates---divida-publica-federal.xlsx"
)

COLUMN_ORDER = [
    "ano",
    "dbgg_01jan_R$bi",
    "dbgg_31dez_R$bi",
    "resultado_primario_R$bi",
    "emissoes_DPF_R$bi",
    "resgates_DPF_R$bi",
    "juros_nominais_R$bi",
    "resultado_nominal_R$bi",
    "renuncia_desenv_regional_R$bi",
    "renuncia_imunes_isentas_R$bi",
    "renuncia_automotivo_R$bi",
    "renuncia_cultura_audiovisual_R$bi",
    "renuncia_inovacao_R$bi",
    "desembolso_BNDES_R$bi",
    "financ_FNO_BASA_R$bi",
    "financ_FNE_BNB_R$bi",
    "financ_FCO_BB_R$bi",
]

DGT_COLUMNS = [
    "renuncia_desenv_regional_R$bi",
    "renuncia_imunes_isentas_R$bi",
    "renuncia_automotivo_R$bi",
    "renuncia_cultura_audiovisual_R$bi",
    "renuncia_inovacao_R$bi",
]

FUNDOS_COLUMNS = [
    "financ_FNO_BASA_R$bi",
    "financ_FNE_BNB_R$bi",
    "financ_FCO_BB_R$bi",
]

_MONTH = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}


def _round_or_none(v: float | None, nd: int = 2) -> float | None:
    if v is None:
        return None
    return round(float(v), nd)


def _mi_to_bi(v: float | None) -> float | None:
    if v is None:
        return None
    return float(v) / 1000.0


def _parse_br_month(label: Any) -> date | None:
    if not isinstance(label, str):
        return None
    m = re.fullmatch(r"([A-Za-zçÇ]{3})/(\d{2})", label.strip())
    if not m:
        return None
    mm = _MONTH.get(m.group(1).lower()[:3])
    if not mm:
        return None
    return date(2000 + int(m.group(2)), mm, 1)


def download_emissoes_xlsx(
    *,
    url: str = EMISSOES_XLSX_URL,
    timeout: float = 120.0,
) -> bytes:
    headers = {
        "User-Agent": providers.DEFAULT_UA,
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
    }
    with httpx.Client(headers=headers, timeout=timeout, follow_redirects=True) as client:
        resp = client.get(url)
        resp.raise_for_status()
        return resp.content


def parse_emissoes_resgates_xlsx(content: bytes) -> tuple[dict[int, float], dict[int, float]]:
    """Return annual totals in R$ milhoes for EMISSOES and RESGATES rows.

    Years with fewer than 12 months are dropped (incomplete).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Pacote openpyxl necessario para ler emissoes/resgates. "
            "Instale com: pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    months: list[tuple[int, date]] = []
    for col in range(2, ws.max_column + 1):
        d = _parse_br_month(ws.cell(5, col).value)
        if d:
            months.append((col, d))

    def row_annual(row_idx: int) -> dict[int, float]:
        totals: dict[int, float] = defaultdict(float)
        counts: dict[int, int] = defaultdict(int)
        for col, d in months:
            val = ws.cell(row_idx, col).value
            if val is None:
                continue
            try:
                totals[d.year] += float(val)
                counts[d.year] += 1
            except (TypeError, ValueError):
                continue
        return {y: v for y, v in totals.items() if counts[y] >= 12}

    # Fixed positions in Anexo 1.1
    return row_annual(7), row_annual(21)


def _rtn_annual(code: str, year_from: int, year_to: int) -> dict[int, float]:
    out = providers.get_serie(
        code,
        data_inicio=f"01/{year_from}",
        data_fim=f"12/{year_to}",
        correcao_ipca=False,
    )
    annual: dict[int, float] = defaultdict(float)
    for row in out.get("series") or []:
        if row.get("date") is None or row.get("value") is None:
            continue
        y = int(str(row["date"])[:4])
        if year_from <= y <= year_to:
            annual[y] += float(row["value"])
    return dict(annual)


def load_overlay_csv(
    path: str | Path,
    value_columns: list[str],
) -> dict[int, dict[str, float]]:
    """Load user CSV with column 'ano' + selected value columns (R$ bi)."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")
    out: dict[int, dict[str, float]] = {}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "ano" not in reader.fieldnames:
            raise ValueError(f"CSV precisa da coluna 'ano': {path}")
        for row in reader:
            raw_ano = (row.get("ano") or "").strip()
            if not raw_ano or raw_ano.startswith("#"):
                continue
            y = int(raw_ano)
            vals: dict[str, float] = {}
            for col in value_columns:
                raw = (row.get(col) or "").strip()
                if raw == "" or raw.lower() in {"n/d", "nd", "na", "-"}:
                    continue
                vals[col] = float(str(raw).replace(",", "."))
            if vals:
                out[y] = vals
    return out


def collect_annual_table(
    *,
    year_from: int = 2001,
    year_to: int = 2025,
    dgt_csv: str | Path | None = None,
    fundos_csv: str | Path | None = None,
    include_emissoes: bool = True,
) -> dict[str, Any]:
    """Build annual table (R$ bi, current prices)."""
    year_from = int(year_from)
    year_to = int(year_to)
    if year_from > year_to:
        year_from, year_to = year_to, year_from

    notes: list[str] = []
    sources: dict[str, str] = {
        "dbgg": "BCB SGS 13761 (R$ mi; 1/jan ~= estoque dez do ano anterior)",
        "resultado_primario": "Tesouro ARIA RTN 10.04.1",
        "juros_nominais": "Tesouro ARIA RTN 10.08.1",
        "resultado_nominal": "Tesouro ARIA RTN 10.09.1",
        "desembolso_BNDES": "BCB SGS 7415",
        "emissoes_resgates": "Tesouro Transparente XLSX Emissoes e Resgates da DPF",
    }

    # RTN
    prim = _rtn_annual("10.04.1", year_from, year_to)
    juros = _rtn_annual("10.08.1", year_from, year_to)
    nom = _rtn_annual("10.09.1", year_from, year_to)

    # DBGG monthly stocks
    dbgg_points = bcb_client.fetch_sgs_range(
        bcb_client.SERIES["dbgg_rs_mi"]["code"],
        date(max(2006, year_from - 1), 1, 1),
        date(year_to, 12, 31),
    )
    dbgg_dec = bcb_client.december_stocks(dbgg_points)

    # BNDES
    bndes_points = bcb_client.fetch_sgs_range(
        bcb_client.SERIES["bndes_desembolso"]["code"],
        date(year_from, 1, 1),
        date(year_to, 12, 31),
    )
    bndes_annual = bcb_client.annual_sum(bndes_points)

    # Emissoes / resgates
    emiss: dict[int, float] = {}
    resg: dict[int, float] = {}
    if include_emissoes:
        try:
            xlsx = download_emissoes_xlsx()
            emiss, resg = parse_emissoes_resgates_xlsx(xlsx)
            notes.append(
                "Emissoes/resgates: anos com <12 meses na planilha sao omitidos."
            )
        except Exception as exc:  # noqa: BLE001
            notes.append(f"Falha ao obter emissoes/resgates DPF: {exc}")

    # Overlays
    dgt_overlay: dict[int, dict[str, float]] = {}
    fundos_overlay: dict[int, dict[str, float]] = {}
    if dgt_csv:
        dgt_overlay = load_overlay_csv(dgt_csv, DGT_COLUMNS)
        sources["renuncias_dgt"] = str(dgt_csv)
    else:
        notes.append(
            "Renuncias DGT vazias: preencha data/templates/dgt_renuncias_anual.csv "
            "e passe --dgt."
        )
    if fundos_csv:
        fundos_overlay = load_overlay_csv(fundos_csv, FUNDOS_COLUMNS)
        sources["fundos"] = str(fundos_csv)
    else:
        notes.append(
            "Fundos FNO/FNE/FCO vazios: preencha "
            "data/templates/fundos_constitucionais_anual.csv e passe --fundos."
        )

    rows: list[dict[str, Any]] = []
    for y in range(year_from, year_to + 1):
        jan = dbgg_dec.get(y - 1)
        dez = dbgg_dec.get(y)
        row: dict[str, Any] = {
            "ano": y,
            "dbgg_01jan_R$bi": _round_or_none(_mi_to_bi(jan)),
            "dbgg_31dez_R$bi": _round_or_none(_mi_to_bi(dez)),
            "resultado_primario_R$bi": _round_or_none(_mi_to_bi(prim.get(y))),
            "emissoes_DPF_R$bi": _round_or_none(_mi_to_bi(emiss.get(y))),
            "resgates_DPF_R$bi": _round_or_none(_mi_to_bi(resg.get(y))),
            "juros_nominais_R$bi": _round_or_none(_mi_to_bi(juros.get(y))),
            "resultado_nominal_R$bi": _round_or_none(_mi_to_bi(nom.get(y))),
            "renuncia_desenv_regional_R$bi": None,
            "renuncia_imunes_isentas_R$bi": None,
            "renuncia_automotivo_R$bi": None,
            "renuncia_cultura_audiovisual_R$bi": None,
            "renuncia_inovacao_R$bi": None,
            "desembolso_BNDES_R$bi": _round_or_none(_mi_to_bi(bndes_annual.get(y))),
            "financ_FNO_BASA_R$bi": None,
            "financ_FNE_BNB_R$bi": None,
            "financ_FCO_BB_R$bi": None,
        }
        if y in dgt_overlay:
            row.update(dgt_overlay[y])
        if y in fundos_overlay:
            row.update(fundos_overlay[y])
        rows.append(row)

    return {
        "year_from": year_from,
        "year_to": year_to,
        "unit": "R$ bilhoes (valores correntes)",
        "count": len(rows),
        "columns": COLUMN_ORDER,
        "rows": rows,
        "sources": sources,
        "notes": notes,
        "provider": "tesouro-mcp/collector + BCB SGS",
    }


def rows_to_csv(rows: list[dict[str, Any]], *, columns: list[str] | None = None) -> str:
    cols = columns or COLUMN_ORDER
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    w.writeheader()
    for row in rows:
        out = {}
        for c in cols:
            v = row.get(c)
            out[c] = "" if v is None else v
        w.writerow(out)
    return buf.getvalue()


def write_annual_csv(
    path: str | Path,
    table: dict[str, Any],
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    text = rows_to_csv(table["rows"], columns=table.get("columns"))
    path.write_text(text, encoding="utf-8")
    return path


def template_dir() -> Path:
    # tesouro-mcp/data/templates relative to package parents
    return Path(__file__).resolve().parents[2] / "data" / "templates"
