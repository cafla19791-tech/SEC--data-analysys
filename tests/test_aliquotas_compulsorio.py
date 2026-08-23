"""Testes das alíquotas de recolhimento compulsório."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import pytest

from scripts.aliquotas_compulsorio import (
    DATA_DIR,
    aplicar_pos_xls,
    eh_aspas,
    gerar_planilha,
    ler_xls_oficial,
    para_pct,
    preencher,
    snapshot_anual,
)


def test_aspas_e_percentual() -> None:
    assert eh_aspas('"')
    assert abs(para_pct(0.45) - 45.0) < 1e-12
    assert abs(para_pct(21) - 21.0) < 1e-12
    assert preencher(45.0, '"') == 45.0
    assert preencher(45.0, 0.21) == 21.0
    assert preencher(8.0, "-") is None


def test_snapshot_anual_usa_ultima_regra_do_ano() -> None:
    alt = pd.DataFrame(
        [
            {"data": pd.Timestamp("2002-06-01"), "a_vista": 45.0, "a_prazo": 15.0, "poupanca_habitacional": 20.0,
             "poupanca_rural": 20.0, "adic_a_vista": None, "adic_a_prazo": None, "adic_poupanca": None,
             "norma": "xls", "nota": ""},
            {"data": pd.Timestamp("2003-02-01"), "a_vista": 60.0, "a_prazo": 15.0, "poupanca_habitacional": 20.0,
             "poupanca_rural": 20.0, "adic_a_vista": 8.0, "adic_a_prazo": 8.0, "adic_poupanca": 10.0,
             "norma": "xls", "nota": ""},
            {"data": pd.Timestamp("2003-08-01"), "a_vista": 45.0, "a_prazo": 15.0, "poupanca_habitacional": 20.0,
             "poupanca_rural": 20.0, "adic_a_vista": 8.0, "adic_a_prazo": 8.0, "adic_poupanca": 10.0,
             "norma": "xls", "nota": ""},
        ]
    )
    anual = snapshot_anual(alt, 2002, 2003)
    assert anual.loc[anual["ano"] == 2002, "a_vista"].iloc[0] == 45.0
    assert anual.loc[anual["ano"] == 2003, "a_vista"].iloc[0] == 45.0
    assert anual.loc[anual["ano"] == 2003, "desde"].iloc[0] == "01/08/2003"


def test_pos_xls_reduz_prazo_em_2020() -> None:
    base = pd.DataFrame(
        [
            {
                "data": pd.Timestamp("2019-07-01"),
                "a_vista": 21.0,
                "a_prazo": 31.0,
                "poupanca_habitacional": 20.0,
                "poupanca_rural": 20.0,
                "adic_a_vista": 0.0,
                "adic_a_prazo": 0.0,
                "adic_poupanca": 0.0,
                "norma": "xls",
                "nota": "",
            }
        ]
    )
    out = aplicar_pos_xls(base)
    assert out.iloc[-1]["a_prazo"] == 20.0
    assert out.iloc[-1]["a_vista"] == 21.0
    mar = out[out["data"] == pd.Timestamp("2020-03-30")].iloc[0]
    assert mar["a_prazo"] == 17.0


def test_gerar_planilha(tmp_path: Path) -> None:
    anual = pd.DataFrame(
        [
            {
                "ano": 2026,
                "a_vista": 21.0,
                "a_prazo": 20.0,
                "poupanca_habitacional": 20.0,
                "poupanca_rural": 20.0,
                "adic_a_vista": 0.0,
                "adic_a_prazo": 0.0,
                "adic_poupanca": 0.0,
                "desde": "29/11/2021",
                "norma": "Resolução BCB 145/2021",
            }
        ]
    )
    alt = pd.DataFrame(
        [
            {
                "data": pd.Timestamp("2021-11-29"),
                "a_vista": 21.0,
                "a_prazo": 20.0,
                "poupanca_habitacional": 20.0,
                "poupanca_rural": 20.0,
                "adic_a_vista": 0.0,
                "adic_a_prazo": 0.0,
                "adic_poupanca": 0.0,
                "norma": "Resolução BCB 145/2021",
                "nota": "20%",
            }
        ]
    )
    quadro = pd.DataFrame([{"Tipo": "Recursos à Vista", "Alíquota": "21%"}])
    path = tmp_path / "comp.xlsx"
    gerar_planilha(anual, alt, quadro, path)
    wb = load_workbook(path)
    assert wb.sheetnames[:3] == ["Notas", "Anual", "Alteracoes"]
    assert wb["Anual"]["B2"].value == 21.0
    assert wb["Vigente"]["B2"].value == "21%"


def test_xls_oficial_marcadores_de_fim_de_ano() -> None:
    xls = DATA_DIR / "compulsorios.xls"
    if not xls.exists():
        pytest.skip("cache do compulsorios.xls ausente")
    hist = ler_xls_oficial(xls)
    alt = aplicar_pos_xls(hist)
    anual = snapshot_anual(alt, 2002, 2026)
    por_ano = anual.set_index("ano")
    assert por_ano.loc[2002, "a_vista"] == 45.0
    assert por_ano.loc[2002, "a_prazo"] == 15.0
    assert por_ano.loc[2003, "a_vista"] == 45.0
    assert por_ano.loc[2018, ["a_vista", "a_prazo"]].tolist() == [21.0, 33.0]
    assert por_ano.loc[2019, "a_prazo"] == 31.0
    assert por_ano.loc[2020, "a_prazo"] == 17.0
    assert por_ano.loc[2021, "a_prazo"] == 20.0
    assert por_ano.loc[2026, ["a_vista", "a_prazo", "poupanca_habitacional"]].tolist() == [
        21.0,
        20.0,
        20.0,
    ]
