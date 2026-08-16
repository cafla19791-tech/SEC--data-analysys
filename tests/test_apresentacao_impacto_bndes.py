"""Testes do Excel de apresentação de impacto BNDES."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.apresentacao_impacto_bndes import (
    MARKER,
    construir_apresentacao,
)


def _seed(pasta: Path) -> None:
    pasta.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        [
            {
                "Ano": 2009,
                "Soma Subsídio Nominal (R$)": 100.0,
                "Impacto Fiscal 2026 (R$)": 1000.0,
                "Quantidade de Parcelas": 10,
            },
            {
                "Ano": 2010,
                "Soma Subsídio Nominal (R$)": 200.0,
                "Impacto Fiscal 2026 (R$)": 1500.0,
                "Quantidade de Parcelas": 20,
            },
        ]
    ).to_csv(pasta / "impacto_fiscal_por_ano.csv", index=False)
    pd.DataFrame(
        [
            {
                "Instituição Financeira": "BANCO A",
                "Qtd Contratos": 2,
                "Qtd Parcelas": 15,
                "Total Subsídio (R$)": 180.0,
                "Impacto Fiscal 2026 (R$)": 1800.0,
            },
            {
                "Instituição Financeira": "BANCO B",
                "Qtd Contratos": 1,
                "Qtd Parcelas": 15,
                "Total Subsídio (R$)": 120.0,
                "Impacto Fiscal 2026 (R$)": 700.0,
            },
        ]
    ).to_csv(pasta / "resumo_por_agente.csv", index=False)
    sub = pasta / "fluxos_por_ano_contrato"
    sub.mkdir(exist_ok=True)
    pd.DataFrame(
        [{"ano": 2009, "contratos": 2, "parcelas": 10, "ok": True}]
    ).to_csv(sub / "RESUMO.csv", index=False)


def test_marker():
    assert "apresentacao-impacto" in MARKER


def test_construir_apresentacao(tmp_path: Path):
    pasta = tmp_path / "saida"
    _seed(pasta)
    out = construir_apresentacao(pasta)
    assert out.exists()
    assert out.name == "APRESENTACAO_IMPACTO_BNDES_INDIRETAS.xlsx"
    wb = load_workbook(out)
    nomes = set(wb.sheetnames)
    assert {
        "Capa",
        "Sumario",
        "Por_Ano",
        "Top_20_Agentes",
        "Por_Agente",
        "Resumo_Geracao",
        "Notas",
    } <= nomes
    assert wb["Capa"]["A1"].value and "Impacto Fiscal" in str(wb["Capa"]["A1"].value)
    assert wb["Por_Ano"].cell(2, 1).value == 2009
    assert wb["Por_Agente"].cell(2, 2).value == "BANCO A"
    assert wb["Top_20_Agentes"].max_row == 3  # header + 2


def test_aceita_coluna_agente(tmp_path: Path):
    pasta = tmp_path / "saida"
    pasta.mkdir()
    pd.DataFrame(
        [
            {
                "Ano": 2015,
                "Soma Subsídio Nominal (R$)": 1.0,
                "Impacto Fiscal 2026 (R$)": 2.0,
                "Quantidade de Parcelas": 1,
            }
        ]
    ).to_csv(pasta / "impacto_fiscal_por_ano.csv", index=False)
    pd.DataFrame(
        [
            {
                "Agente": "CAIXA",
                "Total Subsídio (R$)": 1.0,
                "Impacto Fiscal 2026 (R$)": 2.0,
            }
        ]
    ).to_csv(pasta / "resumo_por_agente.csv", index=False)
    out = construir_apresentacao(pasta, pasta / "out.xlsx")
    wb = load_workbook(out)
    assert wb["Por_Agente"].cell(2, 2).value == "CAIXA"
