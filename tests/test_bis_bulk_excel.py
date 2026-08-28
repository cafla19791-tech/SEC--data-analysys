"""Testes do gerador de Excel por tema BIS (uma aba por país)."""

from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts.bis_bulk_excel import (
    TOPICOS,
    Pais,
    comparativo_paises,
    detectar_coluna_pais,
    eh_agregado,
    filtrar_anos,
    filtrar_frequencia,
    gerar_tema,
    indice_paises,
    limpar_colunas,
    montar_capa,
    partir_codigo_rotulo,
    preparar_aba_pais,
    sanitizar_aba,
    selecionar_topicos,
    escrever_excel,
)


def test_catalogo_tem_os_20_topicos_do_portal():
    ids = [t["id"] for t in TOPICOS]
    assert len(ids) == 20
    assert len(set(ids)) == 20
    assert ids[0] == "LBS"
    assert ids[-1] == "CPMI_FMI"
    assert "CBPOL" in ids
    assert "CPMI_VAREJO" in ids


def test_selecionar_topicos_filtra_e_rejeita_desconhecido():
    escolhidos = selecionar_topicos("cbpol,cpi")
    # ordem do portal (CPI aparece antes de CBPOL)
    assert [t["id"] for t in escolhidos] == ["CPI", "CBPOL"]
    with pytest.raises(SystemExit):
        selecionar_topicos("FOO")


def test_partir_codigo_rotulo_e_agregado():
    assert partir_codigo_rotulo("BR: Brazil") == ("BR", "Brazil")
    assert partir_codigo_rotulo("5A: All countries") == ("5A", "All countries")
    assert partir_codigo_rotulo("United States") == ("United States", "United States")
    assert eh_agregado("BR", "Brazil") is False
    assert eh_agregado("5A", "All countries") is True
    assert eh_agregado("5C", "Euro area") is True
    assert eh_agregado("XM", "Emerging market economies") is True


def test_detectar_coluna_pais_prioriza_ref_area():
    cols = [
        "FREQ:Frequency",
        "BORROWERS_CTY:Borrowers' country",
        "REF_AREA:Reference area",
        "TIME_PERIOD:Time period or range",
    ]
    assert detectar_coluna_pais(cols) == "REF_AREA:Reference area"
    assert detectar_coluna_pais(["L_REP_CTY:Reporting country", "L_CP_COUNTRY:Counterparty"]) == (
        "L_REP_CTY:Reporting country"
    )
    assert detectar_coluna_pais(["FOO:Bar"]) is None
    assert detectar_coluna_pais(
        ["XD_EXCHANGE:Location of trade (Exchange or country)", "ISSUE_CUR:Issue currency"]
    ) == "XD_EXCHANGE:Location of trade (Exchange or country)"


def test_detectar_coluna_pais_fallback_quando_declarante_so_tem_agregado():
    df = pd.DataFrame(
        {
            "DER_REP_CTY:Reporting country": ["5J: All countries"] * 4,
            "DER_CPC:Counterparty country": [
                "5J: All countries",
                "US: United States",
                "JP: Japan",
                "US: United States",
            ],
        }
    )
    assert detectar_coluna_pais(df.columns, df) == "DER_CPC:Counterparty country"

    lbs = pd.DataFrame(
        {
            "L_REP_CTY:Reporting country": ["BR: Brazil", "US: United States", "BR: Brazil"],
            "L_CP_COUNTRY:Counterparty country": [
                "DE: Germany",
                "FR: France",
                "JP: Japan",
            ],
        }
    )
    # mantém o país declarante mesmo se a contraparte tiver mais códigos
    assert detectar_coluna_pais(lbs.columns, lbs) == "L_REP_CTY:Reporting country"


def test_filtrar_frequencia_descarta_diario_quando_ha_mensal():
    df = pd.DataFrame(
        {
            "FREQ:Frequency": ["D: Daily", "D: Daily", "M: Monthly", "Q: Quarterly"],
            "OBS_VALUE": ["1", "2", "3", "4"],
        }
    )
    out, recorte = filtrar_frequencia(df, "FREQ:Frequency")
    assert recorte is not None
    assert set(out["FREQ:Frequency"]) == {"M: Monthly", "Q: Quarterly"}


def test_filtrar_anos_mantem_janela():
    df = pd.DataFrame(
        {
            "TIME_PERIOD:Time period or range": ["2001-Q1", "2020-Q1", "2024-12", "x"],
            "v": [1, 2, 3, 4],
        }
    )
    out, recorte = filtrar_anos(
        df, "TIME_PERIOD:Time period or range", anos=5, hoje=date(2024, 12, 1)
    )
    assert recorte == "períodos a partir de 2020"
    assert list(out["v"]) == [2, 3, 4]


def test_sanitizar_aba_respeita_limite_excel():
    assert sanitizar_aba("BR") == "BR"
    assert "/" not in sanitizar_aba("A/B")
    assert len(sanitizar_aba("X" * 80)) == 31


def test_preparar_aba_pais_filtra_e_converte_valor():
    df = pd.DataFrame(
        {
            "REF_AREA": ["BR: Brazil", "US: United States", "BR: Brazil"],
            "TIME_PERIOD": ["2020", "2020", "2021"],
            "OBS_VALUE": ["1.5", "9", "2.0"],
            "FREQ": ["A: Annual", "A: Annual", "A: Annual"],
        }
    )
    bloco = preparar_aba_pais(df, "REF_AREA", Pais("BR", "Brazil", False), max_linhas=10)
    assert list(bloco["OBS_VALUE"]) == [1.5, 2.0]
    assert "REF_AREA" not in bloco.columns
    assert len(bloco) == 2


def test_indice_e_comparativo(tmp_path: Path):
    df = pd.DataFrame(
        {
            "REF_AREA": ["BR: Brazil", "BR: Brazil", "5A: All countries"],
            "TIME_PERIOD": ["2023", "2024", "2024"],
            "OBS_VALUE": ["10", "12", "99"],
            "UNIT_MEASURE": ["Index"] * 3,
            "TITLE": ["CPI"] * 3,
        }
    )
    paises = {
        "BR": Pais("BR", "Brazil", False),
        "5A": Pais("5A", "All countries", True),
    }
    comp = comparativo_paises(df, "REF_AREA", paises)
    assert set(comp["codigo"]) == {"BR", "5A"}
    br = comp.loc[comp["codigo"] == "BR"].iloc[0]
    assert br["periodo"] == "2024"
    assert br["valor"] == 12.0
    idx = indice_paises(paises, comp)
    assert list(idx["tipo"]) == ["pais", "agregado"]


def test_limpar_colunas_remove_metadados():
    df = pd.DataFrame(
        {
            "STRUCTURE": ["dataflow"],
            "REF_AREA:Reference area": ["BR: Brazil"],
            "OBS_VALUE:Observation Value": ["1"],
            "OBS_CONF:Observation confidentiality": ["F"],
        }
    )
    out = limpar_colunas(df)
    assert "STRUCTURE" not in out.columns
    assert "OBS_CONF" not in out.columns
    assert "REF_AREA" in out.columns
    assert "OBS_VALUE" in out.columns


def test_escrever_excel_cria_aba_por_pais(tmp_path: Path):
    capa = montar_capa(
        {
            "titulo": "Teste",
            "titulo_en": "Test",
            "datasets": ["WS_X"],
        },
        ["recorte demo"],
        n_paises=1,
        n_agregados=0,
        n_linhas=2,
        gerado_em=datetime(2026, 8, 24, tzinfo=timezone.utc),
    )
    pais = Pais("BR", "Brazil", False)
    indice = pd.DataFrame([{"codigo": "BR", "pais": "Brazil", "tipo": "pais", "aba": "BR"}])
    comparativo = pd.DataFrame(
        [{"codigo": "BR", "pais": "Brazil", "periodo": "2024", "valor": 1.0}]
    )
    bloco = pd.DataFrame({"TIME_PERIOD": ["2023", "2024"], "OBS_VALUE": [1.0, 2.0]})
    path = tmp_path / "demo.xlsx"
    escrever_excel(path, capa, indice, comparativo, [(pais, bloco)])
    wb = load_workbook(path)
    assert wb.sheetnames == ["Capa", "Indice", "Comparativo", "BR"]
    ws = wb["BR"]
    assert ws["A1"].value == "TIME_PERIOD"
    assert ws["B2"].value == 1.0


def test_gerar_tema_com_zip_local(tmp_path: Path, monkeypatch):
    csv = (
        "FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,"
        "OBS_VALUE:Observation Value,UNIT_MEASURE:Unit of measure,TITLE:Title\n"
        "M: Monthly,BR: Brazil,2024-01,10.5,368: Per cent,Selic\n"
        "M: Monthly,BR: Brazil,2024-02,10.4,368: Per cent,Selic\n"
        "D: Daily,BR: Brazil,2024-02-01,10.4,368: Per cent,Selic\n"
        "M: Monthly,US: United States,2024-01,5.3,368: Per cent,Fed funds\n"
        "M: Monthly,5A: All countries,2024-01,3.1,368: Per cent,Avg\n"
    )
    zip_path = tmp_path / "cache" / "WS_CBPOL_csv_flat.zip"
    zip_path.parent.mkdir()
    import zipfile

    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("WS_CBPOL_csv_flat.csv", csv)

    def fake_download(url, destino, timeout=180):
        return destino

    monkeypatch.setattr("scripts.bis_bulk_excel.baixar_arquivo", fake_download)

    topico = {
        "id": "CBPOL",
        "titulo": "Taxas de política",
        "titulo_en": "Central bank policy rates",
        "datasets": ["WS_CBPOL"],
        "grande": False,
        "anos": None,
    }
    res = gerar_tema(topico, cache_dir=tmp_path / "cache", saida_dir=tmp_path / "out", max_linhas_aba=1000)
    wb = load_workbook(res.arquivo)
    assert "BR" in wb.sheetnames
    assert "US" in wb.sheetnames
    assert "5A" in wb.sheetnames
    assert res.n_paises == 2
    assert res.n_agregados == 1
    br = wb["BR"]
    headers = [br.cell(1, c).value for c in range(1, br.max_column + 1)]
    assert "OBS_VALUE" in headers
    col = headers.index("OBS_VALUE") + 1
    valores = [br.cell(r, col).value for r in range(2, br.max_row + 1)]
    # diário removido; 2 linhas mensais
    assert valores == [10.5, 10.4]
    assert br.max_row == 3
