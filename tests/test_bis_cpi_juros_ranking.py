"""Testes de CPI, juros em dias úteis e ranking Fisher."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.bis_cpi_juros_ranking import (
    calendario_uteis,
    extrair_cpi,
    fator_diario,
    gerar_cpi,
    gerar_ranking,
    juros_anual,
    juros_real_fisher,
    montar_juros_uteis,
    montar_ranking,
)


def test_fator_diario_constante_recupera_taxa_anual() -> None:
    n = 260
    f = fator_diario(10.0, n)
    acum = (f**n - 1.0) * 100.0
    assert abs(acum - 10.0) < 1e-10


def test_juros_real_fisher() -> None:
    # (1,10 / 1,05) − 1 = 4,7619...%
    assert abs(juros_real_fisher(10.0, 5.0) - (1.10 / 1.05 * 100 - 100)) < 1e-12


def test_calendario_so_segunda_a_sexta() -> None:
    dias = calendario_uteis(pd.Timestamp("1995-01-01"), pd.Timestamp("1995-01-08"))
    assert all(d.weekday() < 5 for d in dias)
    assert pd.Timestamp("1995-01-01") not in dias  # domingo
    assert pd.Timestamp("1995-01-02") in dias  # segunda


def test_montar_juros_uteis_ffill_e_acumulo() -> None:
    s = pd.Series(
        {
            "1995-01-02": 10.0,
            "1995-01-03": 10.0,
            # 04 sem cotação (útil) — deve herdar 10
            "1995-01-05": 10.0,
            "1995-01-06": 10.0,
        }
    )
    tab = montar_juros_uteis(s, pd.Timestamp("1995-01-01"))
    assert tab["data"].dt.weekday.max() < 5
    assert (tab["taxa_basica_aa"] == 10.0).all()
    jan = tab[tab["data"].dt.month == 1]
    assert abs(jan.iloc[-1]["taxa_acum_ano"] - jan.iloc[-1]["taxa_acum_mes"]) < 1e-12
    # ano só com janeiro: acumulado no ano = taxa oficial
    assert abs(jan.iloc[-1]["taxa_acum_ano"] - 10.0) < 1e-10


def test_extrair_cpi_mensal_e_anual() -> None:
    df = pd.DataFrame(
        [
            {
                "FREQ": "M",
                "REF_AREA": "BR",
                "Reference area": "Brazil",
                "UNIT_MEASURE": "628",
                "1995-01": "20",
                "1995": None,
            },
            {
                "FREQ": "M",
                "REF_AREA": "BR",
                "Reference area": "Brazil",
                "UNIT_MEASURE": "771",
                "1995-01": "6.5",
                "1995": None,
            },
            {
                "FREQ": "A",
                "REF_AREA": "BR",
                "Reference area": "Brazil",
                "UNIT_MEASURE": "628",
                "1995-01": None,
                "1995": "22",
            },
            {
                "FREQ": "A",
                "REF_AREA": "BR",
                "Reference area": "Brazil",
                "UNIT_MEASURE": "771",
                "1995-01": None,
                "1995": "66.0",
            },
        ]
    )
    out = extrair_cpi(df)
    assert "BR" in out
    assert out["BR"]["nome"] == "Brasil"
    assert abs(out["BR"]["mensal"].iloc[0]["indice_2010"] - 20) < 1e-9
    assert abs(out["BR"]["anual"].iloc[0]["var_12m"] - 66.0) < 1e-9


def test_gerar_cpi_e_ranking(tmp_path: Path) -> None:
    paises = {
        "BR": {
            "nome": "Brasil",
            "mensal": pd.DataFrame([{"periodo": "1995-01", "indice_2010": 20.0, "var_12m": 6.5}]),
            "anual": pd.DataFrame([{"periodo": "1995", "indice_2010": 22.0, "var_12m": 66.0}]),
        },
        "US": {
            "nome": "Estados Unidos",
            "mensal": pd.DataFrame([{"periodo": "1995-01", "indice_2010": 70.0, "var_12m": 2.8}]),
            "anual": pd.DataFrame([{"periodo": "1995", "indice_2010": 71.0, "var_12m": 2.8}]),
        },
    }
    path = tmp_path / "cpi.xlsx"
    gerar_cpi(paises, path)
    wb = load_workbook(path)
    assert any(s.startswith("BR") for s in wb.sheetnames)
    br = next(s for s in wb.sheetnames if s.startswith("BR"))
    assert wb[br]["B2"].value == 20.0
    assert wb[br]["G2"].value == 66.0

    tab_br = montar_juros_uteis(pd.Series({"1995-01-02": 40.0, "1995-12-29": 40.0}), pd.Timestamp("1995-01-01"))
    tab_us = montar_juros_uteis(pd.Series({"1995-01-02": 5.0, "1995-12-29": 5.0}), pd.Timestamp("1995-01-01"))
    juros = juros_anual({"BR": tab_br, "US": tab_us}, {"BR": "Brasil", "US": "Estados Unidos"})
    infla = pd.DataFrame(
        [
            {"codigo": "BR", "pais": "Brasil", "ano": 1995, "inflacao": 66.0},
            {"codigo": "US", "pais": "Estados Unidos", "ano": 1995, "inflacao": 2.8},
        ]
    )
    base = montar_ranking(juros, infla)
    dest = tmp_path / "rank.xlsx"
    gerar_ranking(base, dest)
    wb2 = load_workbook(dest)
    assert "1995" in wb2.sheetnames
    # primeiro do ranking de inflação deve ser Brasil (66%)
    assert wb2["1995"]["F5"].value == "Brasil"
