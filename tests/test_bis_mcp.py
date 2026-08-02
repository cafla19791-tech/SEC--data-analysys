"""Unit tests for bis-mcp providers (mocked + local CSV fixture)."""

from __future__ import annotations

from pathlib import Path

import pytest

from bis_mcp import providers


FIXTURE_CSV = """FREQ,REF_AREA,TIME_PERIOD,OBS_VALUE,TITLE
M,BR,2024-01,11.75,Brazil
M,BR,2024-02,11.25,Brazil
M,BR,2024-03,10.75,Brazil
M,US,2024-01,5.5,United States
M,US,2024-02,5.5,United States
M,US,2024-03,5.5,United States
D,BR,2024-03-01,10.75,Brazil daily
"""

# Same shape as data.bis.org/static/bulk/WS_CBPOL_csv_flat.zip
FLAT_FIXTURE_CSV = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-01,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-02,11.25,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-03,10.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,US: United States,2024-03,5.5,United States,A: Normal value
"""


def test_resolve_area_aliases():
    a = providers.resolve_area("brasil")
    assert a["code"] == "BR"
    b = providers.resolve_area("selic")
    assert b["code"] == "BR"
    c = providers.resolve_area("euro")
    assert c["code"] == "XM"
    d = providers.resolve_area("US")
    assert d["code"] == "US"


def test_resolve_area_unknown():
    with pytest.raises(ValueError, match="desconhecida"):
        providers.resolve_area("nao_existe_xyz")


def test_normalize_freq():
    assert providers._normalize_freq("mensal") == "M"
    assert providers._normalize_freq("D") == "D"
    with pytest.raises(ValueError):
        providers._normalize_freq("W")


def test_rows_from_csv_text():
    rows = providers._rows_from_csv_text(FIXTURE_CSV)
    assert len(rows) == 7
    assert rows[0]["ref_area"] == "BR"
    assert rows[0]["value"] == pytest.approx(11.75)


def test_read_local_series(tmp_path: Path):
    csv_path = tmp_path / "WS_CBPOL_csv_flat.csv"
    csv_path.write_text(FIXTURE_CSV, encoding="utf-8")
    out = providers.read_local_series(
        "BR,US",
        csv_path=csv_path,
        freq="M",
        last=2,
    )
    assert out["source"] == "local_csv"
    assert out["count"] == 4  # 2 per country
    br = [r for r in out["series"] if r["ref_area"] == "BR"]
    assert br[-1]["time_period"] == "2024-03"
    assert br[-1]["value"] == pytest.approx(10.75)


def test_read_local_series_flat_headers(tmp_path: Path):
    csv_path = tmp_path / "WS_CBPOL_csv_flat.csv"
    csv_path.write_text(FLAT_FIXTURE_CSV, encoding="utf-8")
    out = providers.read_local_series("brasil", csv_path=csv_path, freq="M", last=2)
    assert out["count"] == 2
    assert out["series"][0]["ref_area"] == "BR"
    assert out["series"][-1]["value"] == pytest.approx(10.75)
    assert out["series"][-1]["obs_status"] == "A"


def test_get_policy_rates_sdmx(monkeypatch):
    sample = (
        "FREQ,REF_AREA,TIME_PERIOD,OBS_VALUE\n"
        "M,BR,2025-05,14.75\n"
        "M,BR,2025-06,15\n"
        "M,BR,2025-07,15\n"
    )

    def fake_text(url, **kwargs):
        assert "WS_CBPOL/M.BR" in url
        assert "format=csv" in url
        return sample

    monkeypatch.setattr(providers, "_get_text", fake_text)
    out = providers.get_policy_rates("brasil", last=2)
    assert out["source"] == "sdmx"
    assert out["count"] == 2
    assert out["series"][-1]["value"] == pytest.approx(15.0)


def test_compare_latest(monkeypatch):
    sample = (
        "FREQ,REF_AREA,TIME_PERIOD,OBS_VALUE\n"
        "M,BR,2025-07,15\n"
        "M,US,2025-07,4.5\n"
        "M,XM,2025-07,2.15\n"
    )
    monkeypatch.setattr(providers, "_get_text", lambda url, **kwargs: sample)
    out = providers.compare_latest("BR,US,XM")
    assert len(out["latest"]) == 3
    by_code = {x["code"]: x["value"] for x in out["latest"]}
    assert by_code["BR"] == pytest.approx(15.0)
    assert by_code["US"] == pytest.approx(4.5)


def test_download_flat_csv(tmp_path: Path, monkeypatch):
    import io
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("WS_CBPOL_csv_flat.csv", FIXTURE_CSV)
    blob = buf.getvalue()

    monkeypatch.setattr(providers, "_get_bytes", lambda url, **kwargs: blob)
    out = providers.download_flat_csv(tmp_path)
    path = Path(out["path"])
    assert path.exists()
    assert path.name == "WS_CBPOL_csv_flat.csv"
    assert "11.75" in path.read_text(encoding="utf-8")


def test_extract_areas_csv(tmp_path: Path):
    src = tmp_path / "WS_CBPOL_csv_flat.csv"
    src.write_text(FLAT_FIXTURE_CSV, encoding="utf-8")
    out = tmp_path / "slim.csv"
    result = providers.extract_areas_csv("BR", out, csv_path=src, freq="M")
    assert result["rows"] == 3
    text = out.read_text(encoding="utf-8")
    assert "REF_AREA" in text
    assert "11.75" in text
    assert "US" not in text.splitlines()[1]


def test_taxa_diaria_e_acumulada_composta():
    from bis_mcp.excel_diario import build_country_rows, taxa_diaria_composta_aa

    # 14.5% a.a. -> % a.d. com 252 dias uteis
    ad = taxa_diaria_composta_aa(14.5)
    assert ad == pytest.approx((1.145) ** (1 / 252) - 1)
    rows = build_country_rows([("2024-01-01", 14.5), ("2024-01-02", 14.5)])
    assert rows[0]["Taxa (% a.d.)"] == pytest.approx(ad * 100)
    # compound 2 days
    expected = ((1 + ad) ** 2 - 1) * 100
    assert rows[1]["Taxa acumulada (%)"] == pytest.approx(expected)


def test_acumulado_mes_e_ano_so_no_ultimo_dia():
    from bis_mcp.excel_diario import build_country_rows, taxa_diaria_composta_aa

    ad = taxa_diaria_composta_aa(10.0)
    points = [
        ("2024-01-30", 10.0),
        ("2024-01-31", 10.0),
        ("2024-02-01", 10.0),
        ("2024-12-31", 10.0),
    ]
    rows = build_country_rows(points)
    assert rows[0]["Taxa acumulada mês (%)"] is None
    assert rows[1]["Taxa acumulada mês (%)"] == pytest.approx(((1 + ad) ** 2 - 1) * 100)
    assert rows[1]["Taxa acumulada ano (%)"] is None
    assert rows[2]["Taxa acumulada mês (%)"] == pytest.approx(ad * 100)  # so feb/1
    # ultimo dia do ano na serie
    assert rows[3]["Taxa acumulada ano (%)"] is not None
    # ano 2024: dias 30/01, 31/01, 01/02, 31/12 = 4 dias
    assert rows[3]["Taxa acumulada ano (%)"] == pytest.approx(((1 + ad) ** 4 - 1) * 100)
    assert rows[0]["Taxa acumulada ano (%)"] is None
    assert rows[2]["Taxa acumulada ano (%)"] is None


def test_gerar_excel_diario_from_flat(tmp_path: Path):
    from bis_mcp import excel_diario

    # Minimal daily flat rows
    flat = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2024-01-01,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2024-01-02,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,US: United States,2024-01-02,5.5,United States,A: Normal value
"""
    src = tmp_path / "WS_CBPOL_csv_flat.csv"
    src.write_text(flat, encoding="utf-8")
    out = tmp_path / "out.xlsx"
    meta = excel_diario.gerar_excel_diario(out, csv_path=src, areas="BR,US")
    assert meta["countries"] == 2
    assert out.exists()
    import pandas as pd

    xls = pd.ExcelFile(out)
    assert "00_Legenda" in xls.sheet_names
    assert "01_Indice" in xls.sheet_names
    br = [s for s in xls.sheet_names if s.startswith("BR")][0]
    df = pd.read_excel(out, sheet_name=br)
    assert list(df.columns) == [
        "Dia",
        "Taxa (% a.d.)",
        "Taxa acumulada (%)",
        "Taxa acumulada mês (%)",
        "Taxa acumulada ano (%)",
    ]
    assert len(df) == 2


def test_acumular_periodo_exclui_fim_de_semana():
    from datetime import date

    from bis_mcp.excel_periodos import acumular_periodo, is_weekday

    assert is_weekday(date(2024, 1, 5))  # sexta
    assert not is_weekday(date(2024, 1, 6))  # sabado
    assert not is_weekday(date(2024, 1, 7))  # domingo

    # Sex 5/jan, Sab 6/jan, Dom 7/jan, Seg 8/jan — so 2 dias uteis
    points = [
        ("2024-01-05", 10.0),
        ("2024-01-06", 10.0),
        ("2024-01-07", 10.0),
        ("2024-01-08", 10.0),
    ]
    out = acumular_periodo(points, date(2024, 1, 1), date(2024, 1, 31))
    assert out is not None
    assert out["n_obs"] == 2


def test_gerar_excel_periodos(tmp_path: Path):
    from bis_mcp import excel_periodos

    flat = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2003-01-02,25.0,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2003-01-03,25.0,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,US: United States,2003-01-02,1.25,United States,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,US: United States,2003-01-03,1.25,United States,A: Normal value
"""
    src = tmp_path / "WS_CBPOL_csv_flat.csv"
    src.write_text(flat, encoding="utf-8")
    out = tmp_path / "periodos.xlsx"
    meta = excel_periodos.gerar_excel_periodos(out, csv_path=src)
    assert meta["periodos"] == 6
    import pandas as pd

    xls = pd.ExcelFile(out)
    assert "02_2003_a_2016-04" in xls.sheet_names
    df = pd.read_excel(out, sheet_name="02_2003_a_2016-04", header=1)
    assert "Pais" in df.columns
    assert "Taxa acumulada (%)" in df.columns
    # ordenacao crescente
    vals = df["Taxa acumulada (%)"].astype(float).tolist()
    assert vals == sorted(vals)
    assert df.iloc[0]["Pais"] == "United States"


def test_taxa_mensal_e_periodos_mensal():
    from bis_mcp.excel_diario import taxa_mensal_composta_aa
    from bis_mcp.excel_mensal import build_country_rows_mensal
    from bis_mcp.excel_periodos import acumular_periodo_mensal
    from datetime import date

    am = taxa_mensal_composta_aa(12.0)
    assert am == pytest.approx((1.12) ** (1 / 12) - 1)
    rows = build_country_rows_mensal(
        [("2024-01", 12.0), ("2024-02", 12.0), ("2024-12", 12.0)]
    )
    assert rows[0]["Taxa acumulada ano (%)"] is None
    assert rows[2]["Taxa acumulada ano (%)"] == pytest.approx(((1 + am) ** 3 - 1) * 100)

    stats = acumular_periodo_mensal(
        [("2024-01", 12.0), ("2024-02", 12.0), ("2024-03", 12.0)],
        date(2024, 1, 1),
        date(2024, 2, 29),
    )
    assert stats is not None
    assert stats["n_obs"] == 2
    assert stats["fim_obs"] == "2024-02"


def test_gerar_excel_mensal_and_periodos(tmp_path: Path):
    from bis_mcp import excel_mensal, excel_periodos

    flat = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2003-01,25.0,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2003-02,25.0,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,US: United States,2003-01,1.25,United States,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,US: United States,2003-02,1.25,United States,A: Normal value
"""
    src = tmp_path / "WS_CBPOL_csv_flat.csv"
    src.write_text(flat, encoding="utf-8")
    out_m = tmp_path / "mensal.xlsx"
    meta = excel_mensal.gerar_excel_mensal(out_m, csv_path=src, areas="BR,US")
    assert meta["countries"] == 2
    import pandas as pd

    br = [s for s in pd.ExcelFile(out_m).sheet_names if s.startswith("BR")][0]
    df = pd.read_excel(out_m, sheet_name=br)
    assert list(df.columns) == [
        "Mês",
        "Taxa (% a.m.)",
        "Taxa acumulada (%)",
        "Taxa acumulada ano (%)",
    ]

    out_p = tmp_path / "periodos_m.xlsx"
    meta_p = excel_periodos.gerar_excel_periodos(out_p, csv_path=src, freq="M")
    assert meta_p["freq"] == "M"
    dfp = pd.read_excel(out_p, sheet_name="02_2003_a_2016-04", header=1)
    assert "N_meses" in dfp.columns
    assert dfp.iloc[0]["Pais"] == "United States"


def test_para_pdf_requires_soffice_or_converts(tmp_path: Path):
    from bis_mcp import pdf_export

    if not pdf_export.find_soffice():
        # Ambiente sem LibreOffice: apenas valida a mensagem.
        with pytest.raises(RuntimeError, match="LibreOffice"):
            pdf_export.xlsx_para_pdf(tmp_path / "x.xlsx")
        return

    # Mini workbook
    import pandas as pd

    xlsx = tmp_path / "mini.xlsx"
    pd.DataFrame({"A": [1, 2], "B": [3, 4]}).to_excel(xlsx, index=False)
    out = pdf_export.xlsx_para_pdf(xlsx, out_dir=tmp_path)
    assert Path(out["path"]).is_file()
    assert out["bytes"] > 0


def test_column_widths_fit_headers():
    from bis_mcp.excel_format import column_widths_for_frame

    cols = [
        "Dia",
        "Taxa (% a.d.)",
        "Taxa acumulada (%)",
        "Taxa acumulada mês (%)",
        "Taxa acumulada ano (%)",
    ]
    widths = column_widths_for_frame(cols, [["2024-01-31", 0.05, 12.3, 1.37, None]], padding=4)
    assert widths[0] >= len("Dia") + 4
    assert widths[3] >= len("Taxa acumulada mês (%)") + 4
    assert widths[4] >= len("Taxa acumulada ano (%)") + 4


def test_excel_diario_mensal_cells_centered(tmp_path: Path):
    """Cabecalhos e dados das abas de pais ficam centralizados."""
    from bis_mcp import excel_diario, excel_mensal
    from openpyxl import load_workbook

    flat_d = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2024-01-01,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2024-01-02,11.75,Brazil,A: Normal value
"""
    src_d = tmp_path / "daily.csv"
    src_d.write_text(flat_d, encoding="utf-8")
    out_d = tmp_path / "diario.xlsx"
    excel_diario.gerar_excel_diario(out_d, csv_path=src_d, areas="BR")

    wb_d = load_workbook(out_d)
    br_d = [s for s in wb_d.sheetnames if s.startswith("BR")][0]
    ws_d = wb_d[br_d]
    assert ws_d["A1"].alignment.horizontal == "center"
    assert ws_d["A1"].alignment.vertical == "center"
    assert ws_d["B1"].alignment.horizontal == "center"
    assert ws_d["A2"].alignment.horizontal == "center"
    assert ws_d["B2"].alignment.horizontal == "center"
    assert ws_d["C2"].alignment.horizontal == "center"

    flat_m = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-01,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-02,11.25,Brazil,A: Normal value
"""
    src_m = tmp_path / "mensal.csv"
    src_m.write_text(flat_m, encoding="utf-8")
    out_m = tmp_path / "mensal.xlsx"
    excel_mensal.gerar_excel_mensal(out_m, csv_path=src_m, areas="BR")

    wb_m = load_workbook(out_m)
    br_m = [s for s in wb_m.sheetnames if s.startswith("BR")][0]
    ws_m = wb_m[br_m]
    assert ws_m["A1"].alignment.horizontal == "center"
    assert ws_m["A2"].alignment.horizontal == "center"
    assert ws_m["B2"].alignment.horizontal == "center"
    assert ws_m["C2"].alignment.horizontal == "center"


def test_excel_print_layout_for_pdf(tmp_path: Path):
    """Planilhas diarias/mensais ficam prontas para PDF (paisagem + fit width)."""
    from bis_mcp import excel_diario, excel_mensal
    from openpyxl import load_workbook

    flat = """STRUCTURE,STRUCTURE_ID,ACTION,FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,TITLE:Title,OBS_STATUS:Observation Status
dataflow,BIS:WS_CBPOL(1.0),I,D: Daily,BR: Brazil,2024-01-01,11.75,Brazil,A: Normal value
dataflow,BIS:WS_CBPOL(1.0),I,M: Monthly,BR: Brazil,2024-01,11.75,Brazil,A: Normal value
"""
    src = tmp_path / "cbpol.csv"
    src.write_text(flat, encoding="utf-8")

    out_d = tmp_path / "diario.xlsx"
    excel_diario.gerar_excel_diario(out_d, csv_path=src, areas="BR")
    wb_d = load_workbook(out_d)
    ws_d = wb_d[[s for s in wb_d.sheetnames if s.startswith("BR")][0]]
    assert ws_d.page_setup.orientation == "landscape"
    assert ws_d.page_setup.paperSize == 9  # A4
    assert ws_d.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws_d.page_setup.fitToHeight == 0  # altura livre; largura = 1 pagina
    assert ws_d.print_title_rows == "$1:$1"
    assert ws_d.freeze_panes == "A2"
    assert ws_d.print_options.horizontalCentered is True

    out_m = tmp_path / "mensal.xlsx"
    excel_mensal.gerar_excel_mensal(out_m, csv_path=src, areas="BR")
    wb_m = load_workbook(out_m)
    ws_m = wb_m[[s for s in wb_m.sheetnames if s.startswith("BR")][0]]
    assert ws_m.page_setup.orientation == "landscape"
    assert ws_m.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws_m.print_title_rows == "$1:$1"


def test_cli_help():
    from bis_mcp.cli import build_parser

    help_text = build_parser().format_help()
    assert "serie" in help_text
    assert "compare" in help_text
    assert "download" in help_text
    assert "catalog" in help_text
    assert "extract" in help_text
    assert "excel-periodos" in help_text
