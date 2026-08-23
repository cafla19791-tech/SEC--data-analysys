"""Testes das planilhas BIS com uma aba por país."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.bis_paises_planilhas import (
    coluna_pais,
    gerar_planilha,
    nome_aba,
    nome_pais,
    partir_codigo_nome,
    preparar,
)


def test_partir_codigo_nome() -> None:
    assert partir_codigo_nome("BR: Brazil") == ("BR", "Brazil")
    assert partir_codigo_nome("5C: Euro area") == ("5C", "Euro area")
    assert partir_codigo_nome("US") == ("US", "US")


def test_nome_pais_portugues() -> None:
    assert nome_pais("BR", "Brazil") == "Brasil"
    assert nome_pais("ZZ", "Zedland") == "Zedland"


def test_nome_aba_limite_e_colisao() -> None:
    usados: set[str] = set()
    a = nome_aba("BR", "Brasil", usados)
    b = nome_aba("BR", "Brasil", usados)
    assert a != b
    assert len(a) <= 31 and len(b) <= 31
    assert ":" not in a
    assert "BR" in nome_aba("BR", "Nome / inválido:[x]", set())


def test_coluna_pais_ref_area_ou_borrowers() -> None:
    assert coluna_pais(["FREQ:Frequency", "REF_AREA:Reference area"]) == "REF_AREA:Reference area"
    assert coluna_pais(["BORROWERS_CTY:Borrowers' country", "X"]) == "BORROWERS_CTY:Borrowers' country"


def test_gerar_planilha_uma_aba_por_pais(tmp_path: Path) -> None:
    bruto = pd.DataFrame(
        {
            "STRUCTURE": ["X", "X", "X"],
            "REF_AREA:Reference area": ["BR: Brazil", "BR: Brazil", "US: United States"],
            "FREQ:Frequency": ["M: Monthly", "M: Monthly", "M: Monthly"],
            "TIME_PERIOD:Time period or range": ["2020-01", "2020-02", "2020-01"],
            "OBS_VALUE:Observation Value": ["10.5", "11.0", "0.25"],
            "TITLE:Title": ["Selic", "Selic", "Fed funds"],
        }
    )
    prep, col = preparar(bruto)
    assert col == "REF_AREA:Reference area"
    path = tmp_path / "bis.xlsx"
    gerar_planilha(prep, "Taxas de política", "https://data.bis.org/static/bulk/WS_CBPOL_csv_flat.zip", path)
    wb = load_workbook(path)
    assert wb.sheetnames[0] == "Notas"
    assert wb.sheetnames[1] == "Indice"
    assert any(s.startswith("BR") for s in wb.sheetnames)
    assert any(s.startswith("US") for s in wb.sheetnames)
    assert wb.sheetnames.index([s for s in wb.sheetnames if s.startswith("BR")][0]) < wb.sheetnames.index(
        [s for s in wb.sheetnames if s.startswith("US")][0]
    )
    br = next(s for s in wb.sheetnames if s.startswith("BR"))
    assert wb[br].max_row == 3
    assert wb["Indice"]["A2"].value == "BR"
    assert wb["Indice"]["D2"].value == 2
