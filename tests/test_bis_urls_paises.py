"""Testes das planilhas BIS (uma aba por país) a partir das URLs oficiais."""

from __future__ import annotations

import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.bis_urls_paises import (
    coluna_pais,
    e_periodo,
    fundir_cabecalho,
    gerar_planilha,
    nome_aba,
    nome_pais,
    partir_codigo_nome,
    partir_flat_em_arquivos,
    pivotar_flat,
    preparar_col,
)


def test_fundir_cabecalho_recompõe_virgulas() -> None:
    bruto = [
        "REF_AREA:Reference area",
        "STO:Stocks",
        " Transactions",
        " Other Flows",
        "EXPENDITURE:Expenditure (COFOG",
        " COICOP",
        " COPP or COPNI)",
        "TIME_PERIOD:Time period",
    ]
    out = fundir_cabecalho(bruto)
    assert out == [
        "REF_AREA:Reference area",
        "STO:Stocks, Transactions, Other Flows",
        "EXPENDITURE:Expenditure (COFOG, COICOP, COPP or COPNI)",
        "TIME_PERIOD:Time period",
    ]


def test_partir_e_periodo() -> None:
    assert partir_codigo_nome("BR: Brazil") == ("BR", "Brazil")
    assert e_periodo("2024-Q3")
    assert e_periodo("2024-S1")
    assert e_periodo("2024-01")
    assert e_periodo("1999")
    assert not e_periodo("FREQ")
    assert nome_pais("BR", "Brazil") == "Brasil"


def test_nome_aba_limite() -> None:
    usados: set[str] = set()
    a = nome_aba("BR", "Brasil", usados)
    b = nome_aba("BR", "Brasil", usados)
    assert a != b
    assert len(a) <= 31 and len(b) <= 31


def test_coluna_pais_candidatos() -> None:
    assert coluna_pais(["FREQ", "ISSUER_RES", "Issuer residence"], ("ISSUER_RES", "Issuer residence")) == "ISSUER_RES"
    assert coluna_pais(["REF_AREA:Reference area", "X"], ("REF_AREA:Reference area",)) == "REF_AREA:Reference area"


def test_preparar_col_usa_rotulo_vizinho() -> None:
    df = pd.DataFrame(
        {
            "FREQ": ["Q", "Q"],
            "Frequency": ["Quarterly", "Quarterly"],
            "ISSUER_RES": ["BR", "US"],
            "Issuer residence": ["Brazil", "United States"],
            "2024-Q1": ["1.5", "2.0"],
        }
    )
    prep, col = preparar_col(df, ("ISSUER_RES", "Issuer residence"))
    assert col == "ISSUER_RES"
    assert prep.loc[prep["_codigo"] == "BR", "_pais"].iloc[0] == "Brasil"
    assert prep.loc[prep["_codigo"] == "US", "2024-Q1"].iloc[0] == 2.0


def test_pivotar_flat() -> None:
    df = pd.DataFrame(
        {
            "REF_AREA:Reference area": ["BR: Brazil", "BR: Brazil"],
            "TITLE:Title": ["Dívida", "Dívida"],
            "TIME_PERIOD:Time period": ["2024-Q1", "2024-Q2"],
            "OBS_VALUE:Observation value": ["10", "12"],
        }
    )
    wide = pivotar_flat(df)
    assert "2024-Q1" in wide.columns and "2024-Q2" in wide.columns
    assert float(wide.iloc[0]["2024-Q1"]) == 10
    assert float(wide.iloc[0]["2024-Q2"]) == 12


def test_gerar_planilha_e_partir_flat(tmp_path: Path) -> None:
    csv_txt = (
        "REF_AREA:Reference area,STO:Stocks, Transactions, Other Flows,TIME_PERIOD:Time period,OBS_VALUE:Observation value\n"
        "BR: Brazil,LE: Stocks,2024-Q1,1.0\n"
        "BR: Brazil,LE: Stocks,2024-Q2,1.5\n"
        "US: United States,LE: Stocks,2024-Q1,2.0\n"
    )
    zip_path = tmp_path / "mini.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr("mini.csv", csv_txt)

    partes = partir_flat_em_arquivos(zip_path, "REF_AREA:Reference area", tmp_path / "split")
    codigos = {c for c, _n, _p in partes}
    assert codigos == {"BR", "US"}

    grupos = []
    for codigo, nome_en, path in partes:
        g = pd.read_csv(path, dtype=str)
        grupos.append((codigo, nome_pais(codigo, nome_en), pivotar_flat(g)))

    dest = tmp_path / "out.xlsx"
    gerar_planilha(grupos, "Teste DSS", "https://data.bis.org/static/bulk/x.zip", "país", dest)
    wb = load_workbook(dest)
    assert wb.sheetnames[0] == "Notas"
    assert wb.sheetnames[1] == "Indice"
    assert any(s.startswith("BR") for s in wb.sheetnames)
    br = next(s for s in wb.sheetnames if s.startswith("BR"))
    us = next(s for s in wb.sheetnames if s.startswith("US"))
    assert wb.sheetnames.index(br) < wb.sheetnames.index(us)
    assert wb[br].max_row == 2
    assert wb["Indice"]["A2"].value == "BR"
