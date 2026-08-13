"""Testes das colunas K–N (IPCA, SELIC, juros, diferença) em OPERACOES DIRETAS."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from scripts.calcular_diretas_ipca_selic import (
    HEADERS_KLMN,
    calcular_klmn,
    carregar_ipca,
    carregar_selic_mensal,
    fator_ipca_entre,
    gravar_colunas_klmn,
    juros_contrato_cap,
    main,
    preparar_contratos,
    selic_cap_mensal,
)
from scripts.gerar_fluxos import taxa_contrato_efetiva


def _serie_constante(inicio: str, n: int, taxa: float) -> pd.DataFrame:
    mes = pd.date_range(inicio, periods=n, freq="MS")
    fator = (1 + taxa / 100.0) ** np.arange(1, n + 1)
    return pd.DataFrame({"mes": mes, "valor": taxa, "fator": fator})


def test_fator_ipca_entre():
    ipca = _serie_constante("2009-01-01", 24, 0.5)  # 0,5% a.m.
    f = fator_ipca_entre(ipca, pd.Timestamp("2009-01-01"), pd.Timestamp("2009-12-01"))
    # cumprod: fator[0]=(1.005)^1 … fator[11]=(1.005)^12 → razão = (1.005)^11
    assert abs(f - (ipca.loc[11, "fator"] / ipca.loc[0, "fator"])) < 1e-9
    assert abs(f - (1.005**11)) < 1e-9


def test_selic_e_juros_cap():
    selic = _serie_constante("2009-01-01", 36, 0.9)  # 0,9% a.m.
    l = selic_cap_mensal(selic, pd.Timestamp("2009-01-01"), 12)
    assert abs(l - (1.009**12)) < 1e-9

    m = juros_contrato_cap("TAXA FIXA", 6.0, 12)
    taxa_m = taxa_contrato_efetiva("TAXA FIXA", 6.0)
    assert abs(m - (1 + taxa_m) ** 12) < 1e-9
    assert l > m  # SELIC 0,9% a.m. > ~0,486% a.m. de 6% a.a.


def test_calcular_klmn_e_gravar(tmp_path: Path):
    excel = tmp_path / "OPERACOES DIRETAS - 2002 a 2018.xlsx"
    pd.DataFrame(
        {
            "data_da_contratacao": ["2009-03-15", "2010-01-20"],
            "valor_contratado_reais": [100000.0, 200000.0],
            "valor_desembolsado_reais": [100000.0, 200000.0],
            "custo_financeiro": ["TAXA FIXA", "TJLP"],
            "juros": [6.0, 2.5],
            "prazo_carencia_meses": [0, 6],
            "prazo_amortizacao_meses": [12, 24],
            "forma_de_apoio": ["DIRETA", "DIRETA"],
            "cliente": ["A", "B"],
            "uf": ["SP", "RJ"],
        }
    ).to_excel(excel, index=False)

    # Séries sintéticas cobrindo 2009–2026
    ipca = _serie_constante("2009-01-01", 220, 0.4)
    selic = _serie_constante("2009-01-01", 220, 0.9)
    # grava caches locais para o main também
    ipca_path = tmp_path / "ipca.xlsx"
    selic_path = tmp_path / "selic_mensal.xlsx"
    pd.DataFrame({"Data": ipca["mes"], "IPCA % a.m.": ipca["valor"]}).to_excel(
        ipca_path, index=False
    )
    pd.DataFrame(
        {"Data": selic["mes"], "Taxa Selic mensal - % a.m.": selic["valor"]}
    ).to_excel(selic_path, index=False)

    df_raw = pd.read_excel(excel)
    contratos = preparar_contratos(df_raw)
    calc = calcular_klmn(
        contratos, carregar_ipca(ipca_path), carregar_selic_mensal(selic_path),
        data_ref=datetime(2026, 6, 30),
    )
    assert len(calc) == 2
    assert calc[HEADERS_KLMN["K"]].iloc[0] > 100000.0  # atualizado IPCA
    assert calc[HEADERS_KLMN["L"]].iloc[0] > 1.0
    assert calc[HEADERS_KLMN["M"]].iloc[0] > 1.0
    assert abs(
        calc[HEADERS_KLMN["N"]].iloc[0]
        - (calc[HEADERS_KLMN["L"]].iloc[0] - calc[HEADERS_KLMN["M"]].iloc[0])
    ) < 1e-9

    saida = tmp_path / "out_calculado.xlsx"
    gravar_colunas_klmn(excel, calc, saida, header_row=1)
    wb = load_workbook(saida)
    ws = wb.active
    assert ws["K1"].value == HEADERS_KLMN["K"]
    assert ws["L1"].value == HEADERS_KLMN["L"]
    assert ws["N2"].value == calc[HEADERS_KLMN["N"]].iloc[0]


def test_typos_contagil_operacoes_diretas(tmp_path: Path):
    """Cabeçalhos reais ContAgil: Prezo/amortizaca (typos) + Valor desembolsado R$."""
    excel = tmp_path / "OPERACOES DIRETAS - 2002 a 2018.xlsx"
    pd.DataFrame(
        {
            "Cliente": ["A"],
            "CNPJ": ["123"],
            "UF": ["SP"],
            "Número do contrato": ["99"],
            "Data da contratacao": ["15/03/2009"],
            "Valor desembolsado R$": [100000.0],
            "Custo financeiro": ["TJLP"],
            "Juros": [2.5],
            "Prezo - carencia (meses)": [6],
            "Prazo - amortizaca (meses)": [24],
        }
    ).to_excel(excel, index=False)

    ipca = _serie_constante("2009-01-01", 220, 0.4)
    selic = _serie_constante("2009-01-01", 220, 0.9)
    ipca_path = tmp_path / "ipca.xlsx"
    selic_path = tmp_path / "selic.xlsx"
    pd.DataFrame({"Data": ipca["mes"], "IPCA": ipca["valor"]}).to_excel(
        ipca_path, index=False
    )
    pd.DataFrame({"Data": selic["mes"], "SELIC": selic["valor"]}).to_excel(
        selic_path, index=False
    )

    saida = tmp_path / "calc.xlsx"
    rc = main(
        [
            "--excel",
            str(excel),
            "--saida",
            str(saida),
            "--ipca",
            str(ipca_path),
            "--selic",
            str(selic_path),
        ]
    )
    assert rc == 0
    wb = load_workbook(saida)
    assert wb.active["K2"].value is not None
    assert wb.active["L2"].value > 1.0


def test_main_cli(tmp_path: Path):
    excel = tmp_path / "OPERACOES DIRETAS - 2002 a 2018.xlsx"
    pd.DataFrame(
        {
            "Data da Contratação": ["15/03/2009"],
            "Valor desembolsado Reais": [50000.0],
            "Juros": [5.0],
            "Prazo de Carência (meses)": [3],
            "Prazo de Amortização (meses)": [12],
            "Custo Financeiro": ["TAXA FIXA"],
            "Forma de Apoio": ["DIRETA"],
        }
    ).to_excel(excel, index=False)

    ipca = _serie_constante("2009-01-01", 220, 0.4)
    selic = _serie_constante("2009-01-01", 220, 0.9)
    ipca_path = tmp_path / "ipca.xlsx"
    selic_path = tmp_path / "selic.xlsx"
    pd.DataFrame({"Data": ipca["mes"], "IPCA": ipca["valor"]}).to_excel(
        ipca_path, index=False
    )
    pd.DataFrame({"Data": selic["mes"], "SELIC": selic["valor"]}).to_excel(
        selic_path, index=False
    )

    saida = tmp_path / "calc.xlsx"
    rc = main(
        [
            "--excel",
            str(excel),
            "--saida",
            str(saida),
            "--ipca",
            str(ipca_path),
            "--selic",
            str(selic_path),
            "--data-ref",
            "2026-06-30",
        ]
    )
    assert rc == 0
    assert saida.exists()
    wb = load_workbook(saida)
    assert wb.active["K1"].value == "Valor_Desembolsado_IPCA"
    assert isinstance(wb.active["K2"].value, (int, float))
