"""Testes do cotejamento IPCA × CPI All Items."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.cotejamento_ipca_cpi import (
    acumulado,
    cpi_dez_dez,
    gravar,
    ipca_dez_dez,
    media_geometrica,
    montar_cotejamento,
    resumo_periodo,
)


def test_acumulado_e_geometrica():
    assert abs(acumulado([0.10, 0.10]) - 0.21) < 1e-12
    assert abs(media_geometrica([0.10, 0.10]) - 0.10) < 1e-12


def test_ipca_dezembro_oficial():
    df = pd.DataFrame(
        {"data": pd.to_datetime(["2021-12-01", "2022-12-01"]), "valor": [10.06, 5.79]}
    )
    assert abs(ipca_dez_dez(df, 2022) - 0.0579) < 1e-12


def test_cpi_dezembro_sobre_dezembro():
    df = pd.DataFrame(
        {"data": pd.to_datetime(["2021-12-01", "2022-12-01"]), "valor": [200.0, 213.0]}
    )
    assert abs(cpi_dez_dez(df, 2022) - 0.065) < 1e-12
    assert cpi_dez_dez(df, 2021) is None


def test_montar_e_resumo(tmp_path: Path):
    ipca = pd.DataFrame(
        {
            "data": pd.to_datetime(["2023-12-01", "2024-12-01"]),
            "valor": [4.62, 4.83],
        }
    )
    cpi = pd.DataFrame(
        {
            "data": pd.to_datetime(["2022-12-01", "2023-12-01", "2024-12-01"]),
            "valor": [100.0, 103.35, 106.34],
        }
    )
    tab = montar_cotejamento(ipca, cpi, [2023, 2024])
    assert list(tab["ano"]) == [2023, 2024]
    assert abs(tab.iloc[0]["ipca_pct"] - 4.62) < 1e-12
    assert tab.iloc[0]["maior_no_ano"] == "IPCA"
    assert tab.iloc[-1]["indice_ipca"] > 100
    sintet = resumo_periodo(tab, 2023, 2024, "2023–2024")
    assert sintet["anos_ipca_maior"] == 2
    caminhos = gravar(tab, pd.DataFrame([sintet]), tmp_path, stem="teste")
    wb = load_workbook(caminhos["xlsx"])
    assert "Anual_1990_2025" in wb.sheetnames or wb.sheetnames[0].startswith("Anual")
    assert "Sintese" in wb.sheetnames
    assert "IPCA" in caminhos["md"].read_text(encoding="utf-8")
