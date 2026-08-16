"""Testes do conversor CSV → XLSX fatiado."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.csv_fluxos_para_xlsx_fatiado import (
    MARKER,
    converter_csv_para_xlsx,
    listar_csvs_ano,
    processar_pasta,
)


def _csv(path: Path, n: int) -> None:
    rows = [
        {
            "contrato": f"{i}-2002",
            "data_fluxo": "2002-06-15",
            "subsidio": float(i),
            "impacto_fiscal": float(i) * 2,
        }
        for i in range(1, n + 1)
    ]
    pd.DataFrame(rows).to_csv(path, index=False)


def test_marker():
    assert "csv-para-xlsx-fatiado" in MARKER


def test_fatiar_em_varias_abas(tmp_path: Path):
    csv_p = tmp_path / "2002.csv"
    _csv(csv_p, 25)
    info = converter_csv_para_xlsx(csv_p, linhas_por_aba=10)
    assert info["linhas"] == 25
    assert info["n_abas"] == 3  # 10 + 10 + 5
    assert info["abas"] == ["2002_p01", "2002_p02", "2002_p03"]
    wb = load_workbook(info["xlsx"], read_only=True)
    assert wb.sheetnames == ["2002_p01", "2002_p02", "2002_p03"]
    # header + 10 dados na primeira
    ws1 = wb["2002_p01"]
    rows1 = list(ws1.iter_rows(values_only=True))
    assert rows1[0][0] == "contrato"
    assert len(rows1) == 11
    ws3 = wb["2002_p03"]
    rows3 = list(ws3.iter_rows(values_only=True))
    assert len(rows3) == 6  # header + 5
    wb.close()


def test_retomar_pula_existente(tmp_path: Path):
    _csv(tmp_path / "2003.csv", 5)
    _csv(tmp_path / "2004.csv", 5)
    # gera só 2003
    converter_csv_para_xlsx(tmp_path / "2003.csv", linhas_por_aba=100)
    assert (tmp_path / "2003.xlsx").exists()
    res = processar_pasta(tmp_path, retomar=True, linhas_por_aba=100)
    # só 2004 convertido nesta passagem
    assert len(res) == 1
    assert res[0]["ano"] == "2004"
    assert (tmp_path / "2004.xlsx").exists()


def test_listar_ignora_resumo(tmp_path: Path):
    _csv(tmp_path / "2010.csv", 2)
    (tmp_path / "RESUMO.csv").write_text("ano,ok\n2010,1\n", encoding="utf-8")
    csvs = listar_csvs_ano(tmp_path)
    assert [p.name for p in csvs] == ["2010.csv"]
