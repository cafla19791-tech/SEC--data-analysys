"""Testes do custo de captação (mercado e referencial BCB)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.custo_captacao_bcb import (
    PARES_SPREAD,
    _fatias_periodo,
    am_para_aa,
    codigo_spread,
    comparativo_anual,
    custo_referencial,
    custos_anuais,
    custos_mensais,
    gerar_planilha,
    medias_anuais_mercado,
)


def test_fatias_periodo_cobrem_o_intervalo() -> None:
    fatias = _fatias_periodo("01/01/2001", "01/12/2006", anos=4)
    assert fatias[0] == ("01/01/2001", "31/12/2004")
    assert fatias[-1][1] == "01/12/2006"


def test_am_para_aa_capitaliza() -> None:
    assert abs(am_para_aa(0.0) - 0.0) < 1e-12
    assert abs(am_para_aa(1.0) - ((1.01**12) - 1) * 100) < 1e-12


def test_custo_referencial_e_codigo_spread() -> None:
    assert abs(custo_referencial(33.44, 21.99) - 11.45) < 1e-12
    assert codigo_spread(20714) == 20783
    assert codigo_spread(20756) == 20825
    assert codigo_spread(22022) is None
    for taxa, spread, *_ in PARES_SPREAD:
        if 20714 <= taxa <= 20782:
            assert spread == codigo_spread(taxa)


def test_medias_anuais_mercado_anualiza_cdi() -> None:
    mensal = pd.DataFrame(
        {
            "data": pd.to_datetime(["2024-01-01", "2024-02-01", "2025-01-01"]),
            "taxa": [1.0, 1.0, 0.5],
            "codigo": [4391, 4391, 4391],
            "indicador": ["CDI"] * 3,
            "unidade_origem": ["am"] * 3,
            "nota": ["teste"] * 3,
        }
    )
    anual = medias_anuais_mercado(mensal)
    esperado = am_para_aa(1.0)
    assert abs(float(anual.loc[0, "2024"]) - esperado) < 1e-12
    assert abs(float(anual.loc[0, "2025"]) - am_para_aa(0.5)) < 1e-12


def test_custos_anuais_taxa_menos_spread() -> None:
    taxas = pd.DataFrame(
        {
            "data": pd.to_datetime(["2011-03-01", "2011-04-01", "2012-01-01"]),
            "taxa": [30.0, 32.0, 20.0],
            "codigo_taxa": [20714, 20714, 20714],
            "modalidade": ["Total do SFN"] * 3,
            "segmento": ["Total"] * 3,
            "origem": ["Livre e direcionado"] * 3,
        }
    )
    spreads = pd.DataFrame(
        {
            "data": pd.to_datetime(["2011-03-01", "2011-04-01", "2012-01-01"]),
            "spread": [20.0, 22.0, 12.0],
            "codigo_taxa": [20714, 20714, 20714],
            "codigo_spread": [20783, 20783, 20783],
            "modalidade": ["Total do SFN"] * 3,
            "segmento": ["Total"] * 3,
            "origem": ["Livre e direcionado"] * 3,
        }
    )
    mensal = custos_mensais(taxas, spreads)
    assert list(mensal["custo"]) == [10.0, 10.0, 8.0]
    anual = custos_anuais(mensal)
    assert abs(float(anual.loc[0, "2011"]) - 10.0) < 1e-12
    assert abs(float(anual.loc[0, "2012"]) - 8.0) < 1e-12
    assert pd.isna(anual.loc[0, "2010"])


def test_gerar_planilha_abas(tmp_path: Path) -> None:
    anos = {str(a): (14.0 if a == 2024 else float("nan")) for a in range(2001, 2027)}
    mercado = pd.DataFrame(
        [
            {
                "codigo": 4189,
                "indicador": "Selic over",
                "unidade": "% a.a.",
                "inicio": "01/2001",
                "fim": "08/2026",
                "nota": "teste",
                **anos,
            }
        ]
    )
    referencial = pd.DataFrame(
        [
            {
                "codigo_taxa": 20714,
                "codigo_spread": 20783,
                "modalidade": "Total do SFN",
                "segmento": "Total",
                "origem": "Livre e direcionado",
                "inicio": "03/2011",
                "fim": "06/2026",
                **{str(a): (11.0 if a == 2024 else float("nan")) for a in range(2001, 2027)},
            }
        ]
    )
    taxas_a = pd.DataFrame(
        [
            {
                "codigo": 20714,
                "modalidade": "Total do SFN",
                "segmento": "Total",
                "origem": "Livre e direcionado",
                "inicio": "03/2011",
                "fim": "06/2026",
                **{str(a): (30.0 if a == 2024 else float("nan")) for a in range(2001, 2027)},
            }
        ]
    )
    spreads_m = pd.DataFrame(
        {
            "data": pd.to_datetime(["2024-01-01"]),
            "spread": [19.0],
            "codigo": [20783],
            "codigo_spread": [20783],
            "modalidade": ["Total do SFN"],
            "segmento": ["Total"],
            "origem": ["Livre e direcionado"],
        }
    )
    comp = comparativo_anual(taxas_a, spreads_m, referencial)
    path = tmp_path / "captacao.xlsx"
    gerar_planilha(mercado, referencial, comp, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Notas", "Mercado", "Custo_referencial", "Taxa_spread_custo"]
    assert wb["Mercado"]["B2"].value == "Selic over"
    assert wb["Custo_referencial"]["C2"].value == "Total do SFN"
    assert abs(float(wb["Custo_referencial"].cell(2, 8 + (2024 - 2001)).value) - 11.0) < 1e-9
