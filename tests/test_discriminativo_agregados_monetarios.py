"""Testes do discriminativo M1–M4."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
import pytest

from scripts.discriminativo_agregados_monetarios import (
    AGREGADOS,
    CODIGO_AGREGADO,
    MENOS,
    carregar_painel,
    processar,
    tabela_anual,
    tabela_composicao,
    tabela_mensal,
)


def _painel_sintetico() -> pd.DataFrame:
    meses = pd.date_range("2001-12-01", "2003-07-01", freq="MS")
    n = len(meses)
    m1 = pd.Series([100_000.0 + 1_000 * i for i in range(n)], index=meses)
    m2 = m1 + 50_000
    m3 = m2 + 80_000
    m4 = m3 + 40_000
    pmpp = m1 * 0.35
    vista = m1 * 0.65
    return pd.DataFrame(
        {
            "M1": m1,
            "M2": m2,
            "M3": m3,
            "M4": m4,
            "Papel-moeda em poder do público": pmpp,
            "Depósitos à vista": vista,
        }
    )


def test_catalogo_oficial():
    assert CODIGO_AGREGADO == {"M1": 27791, "M2": 27810, "M3": 27813, "M4": 27815}
    assert AGREGADOS == ("M1", "M2", "M3", "M4")


def test_camadas_e_identidade():
    mensal = tabela_mensal(_painel_sintetico())
    assert (mensal["M1"] + mensal["M2 − M1"]).equals(mensal["M2"])
    assert (mensal["M2"] + mensal["M3 − M2"]).equals(mensal["M3"])
    assert (mensal["M3"] + mensal["M4 − M3"]).equals(mensal["M4"])
    assert mensal["M2 − M1"].iloc[0] == pytest.approx(50_000)
    assert "M1 Δ% 12m" in mensal.columns
    # 12 meses depois: +12.000 / 100.000 = 12%
    dez02 = mensal[pd.to_datetime(mensal["Mês"]) == pd.Timestamp("2002-12-01")].iloc[0]
    assert dez02["M1 Δ% 12m"] == pytest.approx(12_000 / 100_000)


def test_composicao_m1():
    comp = tabela_composicao(_painel_sintetico())
    assert comp["Resíduo M1"].abs().max() < 1e-6


def test_anual_pega_dezembro_e_ultimo():
    mensal = tabela_mensal(_painel_sintetico())
    anual = tabela_anual(mensal)
    anos = pd.to_datetime(anual["Mês"]).dt.year.tolist()
    assert anos == [2001, 2002, 2003]
    assert pd.to_datetime(anual.loc[anual.index[0], "Mês"]).month == 12
    assert pd.to_datetime(anual.loc[anual.index[-1], "Mês"]).month == 7


def _gravar(pasta: Path, painel: pd.DataFrame) -> dict[int, Path]:
    pasta.mkdir(parents=True, exist_ok=True)
    nomes = {
        "M1": 27791,
        "M2": 27810,
        "M3": 27813,
        "M4": 27815,
        "Papel-moeda em poder do público": 27789,
        "Depósitos à vista": 27790,
    }
    arquivos = {}
    for nome, codigo in nomes.items():
        dest = pasta / f"{codigo}.csv"
        pd.DataFrame({"mes": painel.index, "valor": painel[nome].to_numpy()}).to_csv(
            dest, index=False
        )
        arquivos[codigo] = dest
    return arquivos


def test_planilha_discriminativo(tmp_path: Path):
    painel = _painel_sintetico()
    arquivos = _gravar(tmp_path / "in", painel)
    saida = tmp_path / "m1m4.xlsx"
    path = processar(
        pasta_cache=tmp_path / "cache",
        saida=saida,
        usar_cache=False,
        arquivos=arquivos,
    )
    assert path.exists()
    nomes = pd.ExcelFile(path).sheet_names
    assert nomes[0] == "Metodologia"
    assert "Discriminativo" in nomes
    assert "Anual" in nomes
    assert "Composicao_M1" in nomes
    disc = pd.read_excel(path, sheet_name="Discriminativo", header=3)
    assert "M1" in disc.columns and "M4" in disc.columns
    assert "M2 − M1" in disc.columns or "M2 - M1" in disc.columns
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["Discriminativo"]
    headers = [c.value for c in ws[4]]
    assert headers[0] == "Mês"
    assert "M1" in headers and "M4" in headers
    assert any("M2" in str(h) and "M1" in str(h) for h in headers if h)
    # formato com sinal −
    cel = ws.cell(5, 2)
    assert MENOS in (cel.number_format or "")


def test_carregar_painel_local(tmp_path: Path):
    painel = _painel_sintetico()
    arquivos = _gravar(tmp_path / "in", painel)
    out = carregar_painel(tmp_path / "cache", arquivos=arquivos)
    assert set(AGREGADOS).issubset(out.columns)
    assert out.index.min().date() == date(2001, 12, 1)
