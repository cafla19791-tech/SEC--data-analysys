"""Testes do discriminativo de juros reais por país (BIS / Fisher)."""

from __future__ import annotations

import zipfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts.discriminativo_juros_reais_paises import (
    COL_INDICE,
    COL_INFLACAO,
    COL_MES,
    COL_NOMINAL,
    COL_REAL_ANO,
    COL_REAL_MES,
    COL_TIPO,
    MARKER,
    acumular_fator,
    carregar_cbpol_mensal,
    carregar_cpi_mensal,
    linhas_discriminativo,
    montar_serie_pais,
    nome_aba,
    pivot_anual,
    processar,
    resumo_pais,
    taxa_mensal_composta,
    taxa_real_fisher,
)


def test_taxa_mensal_composta():
    assert taxa_mensal_composta(0.12) == pytest.approx((1.12 ** (1 / 12)) - 1)


def test_fisher_e_acumulado_anual():
    i_aa = 0.12
    i_am = taxa_mensal_composta(i_aa)
    pi_am = 0.005
    r_am = taxa_real_fisher(i_am, pi_am)
    assert r_am == pytest.approx((1 + i_am) / 1.005 - 1)
    anual = acumular_fator(pd.Series([r_am] * 12))
    assert anual == pytest.approx((1 + r_am) ** 12 - 1)
    # real anual < aproximação linear 12% − 6%
    assert anual < 0.06


def test_acumular_ignora_nulos():
    assert acumular_fator(pd.Series([0.01, None, 0.01])) == pytest.approx(1.01 * 1.01 - 1)
    assert pd.isna(acumular_fator(pd.Series([None, None])))


def _serie_sintetica() -> tuple[pd.DataFrame, pd.DataFrame]:
    """24 meses: 2023–2024, taxa 12% a.a., IPC +0,5% a.m."""
    meses = pd.date_range("2022-12-01", "2024-12-01", freq="MS")
    indice0 = 100.0
    indices = [indice0 * (1.005 ** i) for i in range(len(meses))]
    cpi = pd.DataFrame(
        {
            "codigo": "BR",
            "pais_en": "Brazil",
            "mes": meses,
            "indice": indices,
        }
    )
    nom = pd.DataFrame(
        {
            "codigo": "BR",
            "pais_en": "Brazil",
            "mes": meses[1:],  # a partir de jan/2023
            "nominal_aa": 0.12,
        }
    )
    return nom, cpi


def test_montar_serie_e_linhas_dezembro():
    nom, cpi = _serie_sintetica()
    serie = montar_serie_pais(nom, cpi, codigo="BR")
    assert len(serie) == 24
    assert serie["inflacao_am"].iloc[0] == pytest.approx(0.005)
    i_am = taxa_mensal_composta(0.12)
    r_am = taxa_real_fisher(i_am, 0.005)
    assert serie["real_am"].iloc[0] == pytest.approx(r_am)

    linhas = linhas_discriminativo(serie)
    meses = linhas[linhas["tipo"] == "mes"]
    acums = linhas[linhas["tipo"] == "acumulado"]
    assert len(meses) == 24
    assert list(acums[COL_MES]) == ["ACUMULADO 2023", "ACUMULADO 2024"]
    # acumulado vem imediatamente após dezembro
    idx_dez_2023 = linhas.index[linhas[COL_MES] == "12/2023"][0]
    assert linhas.loc[idx_dez_2023 + 1, COL_MES] == "ACUMULADO 2023"
    esperado = acumular_fator(pd.Series([r_am] * 12))
    assert acums.iloc[0][COL_REAL_ANO] == pytest.approx(esperado)
    assert pd.isna(meses.iloc[0][COL_REAL_ANO])


def test_acumulado_parcial_ano_incompleto():
    nom, cpi = _serie_sintetica()
    nom = nom[nom["mes"] <= "2024-03-01"]
    cpi = cpi[cpi["mes"] <= "2024-03-01"]
    serie = montar_serie_pais(nom, cpi, codigo="BR")
    linhas = linhas_discriminativo(serie)
    parciais = linhas[linhas["tipo"] == "parcial"]
    assert len(parciais) == 1
    assert "ACUMULADO PARCIAL 2024" in parciais.iloc[0][COL_MES]
    assert linhas[linhas["tipo"] == "acumulado"].iloc[0][COL_MES] == "ACUMULADO 2023"


def test_recorte_ano_usa_ipc_anterior():
    nom, cpi = _serie_sintetica()
    serie = montar_serie_pais(nom, cpi, codigo="BR", ano_inicio=2024)
    assert serie["mes"].min() == pd.Timestamp("2024-01-01")
    assert serie["inflacao_am"].iloc[0] == pytest.approx(0.005)


def test_nome_aba_excel():
    assert nome_aba("BR") == "Brasil"
    assert nome_aba("US") == "Estados Unidos"
    assert "/" not in nome_aba("XX", "A/B*C")


def test_resumo_pais():
    nom, cpi = _serie_sintetica()
    serie = montar_serie_pais(nom, cpi, codigo="BR")
    linhas = linhas_discriminativo(serie)
    info = resumo_pais(serie, linhas)
    assert info["pais"] == "Brasil"
    assert info["ultimo_ano_completo"] == 2024
    assert info["n_meses"] == 24


def _zip_csv(path: Path, nome: str, csv_text: str) -> Path:
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(nome, csv_text)
    return path


def _cbpol_csv() -> str:
    header = (
        "FREQ:Frequency,REF_AREA:Reference area,TIME_PERIOD:Time period or range,"
        "OBS_VALUE:Observation Value,OBS_STATUS:Observation Status\n"
    )
    rows = ["D: Daily,BR: Brazil,2023-01-15,99,A: Normal value\n"]
    for ym, val in (("2023-01", "12.0"), ("2023-02", "12.0"), ("2023-12", "12.0")):
        rows.append(f"M: Monthly,BR: Brazil,{ym},{val},A: Normal value\n")
    rows.append("M: Monthly,US: United States,2023-01,5.25,A: Normal value\n")
    return header + "".join(rows)


def _cpi_csv() -> str:
    header = (
        "FREQ:Frequency,REF_AREA:Reference area,UNIT_MEASURE:Unit of measure,"
        "TIME_PERIOD:Time period or range,OBS_VALUE:Observation Value,"
        "OBS_STATUS:Observation Status\n"
    )
    rows = []
    idx = 100.0
    for ym in ("2022-12", "2023-01", "2023-02", "2023-12"):
        rows.append(
            f'M: Monthly,BR: Brazil,"628: Index, 2010 = 100",{ym},{idx:.4f},A: Normal value\n'
        )
        idx *= 1.005
        rows.append(
            f'A: Annual,BR: Brazil,"771: Year-on-year changes, in per cent",{ym[:4]},6.0,A: Normal value\n'
        )
    rows.append(
        'M: Monthly,US: United States,"628: Index, 2010 = 100",2023-01,120.0,A: Normal value\n'
    )
    return header + "".join(rows)


def test_carregar_zips_filtra_mensal(tmp_path: Path):
    cbpol = _zip_csv(tmp_path / "cbpol.zip", "WS_CBPOL_csv_flat.csv", _cbpol_csv())
    cpi = _zip_csv(tmp_path / "cpi.zip", "WS_LONG_CPI_csv_flat.csv", _cpi_csv())
    nom = carregar_cbpol_mensal(cbpol)
    ipc = carregar_cpi_mensal(cpi)
    assert set(nom["codigo"]) == {"BR", "US"}
    assert not (nom["mes"].dt.day != 1).any()
    assert (nom["codigo"] == "BR").sum() == 3
    assert ipc[ipc["codigo"] == "BR"]["indice"].iloc[0] == pytest.approx(100.0)


def test_processar_gera_abas(tmp_path: Path):
    cbpol = _zip_csv(tmp_path / "cbpol.zip", "WS_CBPOL.csv", _cbpol_csv())
    cpi = _zip_csv(tmp_path / "cpi.zip", "WS_CPI.csv", _cpi_csv())
    saida = tmp_path / "out.xlsx"
    processar(
        tmp_path,
        saida,
        baixar=False,
        cbpol_zip=cbpol,
        cpi_zip=cpi,
        paises={"BR"},
    )
    assert saida.exists()
    wb = load_workbook(saida)
    assert "Capa" in wb.sheetnames
    assert "Resumo" in wb.sheetnames
    assert "Anual" in wb.sheetnames
    assert "Brasil" in wb.sheetnames
    assert "Estados Unidos" not in wb.sheetnames
    capa = wb["Capa"]
    valores = [c.value for row in capa.iter_rows(max_col=2) for c in row]
    assert MARKER in valores
    br = wb["Brasil"]
    headers = [c.value for c in next(br.iter_rows(min_row=5, max_row=5))]
    assert headers[:6] == [
        COL_MES,
        COL_NOMINAL,
        COL_INDICE,
        COL_INFLACAO,
        COL_REAL_MES,
        COL_REAL_ANO,
    ]
    meses = []
    acum = None
    for row in br.iter_rows(min_row=6, values_only=True):
        if row[0] and str(row[0]).startswith("ACUMULADO"):
            acum = row
        elif row[0]:
            meses.append(row[0])
    assert "01/2023" in meses
    assert "12/2023" in meses
    assert acum is not None
    assert acum[5] is not None
    anual = wb["Anual"]
    assert anual["A3"].value == "País"
    anos = [c.value for c in anual[3] if isinstance(c.value, int)]
    assert anos
    assert min(anos) >= 1995
    assert anual["A4"].value == "Brasil"
    assert anual.column_dimensions["B"].width >= 12
    assert not anual["B4"].alignment.wrap_text


def test_pivot_anual_so_desde_1995():
    def _linhas(*anos: int) -> pd.DataFrame:
        return pd.DataFrame(
            {
                COL_MES: [f"ACUMULADO {a}" for a in anos],
                COL_REAL_ANO: [0.01 * (i + 1) for i in range(len(anos))],
                "ano": list(anos),
                COL_TIPO: ["acumulado"] * len(anos),
            }
        )

    p = pivot_anual({"BR": _linhas(1980, 1994, 1995, 2001), "US": _linhas(1990, 2000)})
    assert list(p.index) == [1995, 2000, 2001]
    assert 1980 not in p.index
    assert "Brasil" in p.columns
