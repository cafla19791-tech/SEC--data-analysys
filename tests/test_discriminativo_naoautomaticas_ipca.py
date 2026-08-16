"""Testes do discriminativo de operações não automáticas BNDES."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from scripts.discriminativo_naoautomaticas_ipca import (
    COL_CLIENTE,
    COL_DATA,
    COL_DESEMBOLSO,
    COL_IPCA,
    MARKER,
    aplicar_ipca,
    carregar_contratos,
    filtrar_periodo,
    montar_aba_periodo,
    montar_aba_por_cliente,
    processar,
)


def _fake_fonte(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SITE"
    ws.append(["Operações não automáticas"])
    ws.append(["Período"])
    ws.append(["Data"])
    ws.append([])
    ws.append(
        [
            "Cliente",
            "CNPJ",
            "Número do contrato",
            "Data da contratação",
            "Valor contratado  R$",
            "Valor desembolsado R$",
            "UF",
            "Situação do contrato",
            "Forma de apoio",
            "Produto",
        ]
    )
    rows = [
        ("CLIENTE ALTO", "1", "C1", datetime(2002, 6, 15), 1000, 1000, "SP", "ATIVO", "DIRETA", "P"),
        ("CLIENTE ALTO", "1", "C2", datetime(2002, 8, 1), 500, 500, "SP", "ATIVO", "DIRETA", "P"),
        ("CLIENTE BAIXO", "2", "C3", datetime(2004, 1, 10), 100, 100, "RJ", "ATIVO", "INDIRETA", "P"),
        ("CLIENTE BAIXO", "2", "C4", datetime(2016, 5, 11), 50, 50, "RJ", "ATIVO", "INDIRETA", "P"),
        ("CLIENTE MEIO", "3", "C5", datetime(2016, 5, 12), 200, 200, "MG", "ATIVO", "DIRETA", "P"),
        ("CLIENTE MEIO", "3", "C6", datetime(2018, 12, 31), 300, 300, "MG", "ATIVO", "DIRETA", "P"),
        ("CLIENTE NOVO", "4", "C7", datetime(2020, 3, 1), 80, 80, "PR", "ATIVO", "DIRETA", "P"),
        ("CLIENTE HOJE", "5", "C8", datetime(2024, 2, 1), 400, 400, "BA", "ATIVO", "DIRETA", "P"),
    ]
    for r in rows:
        ws.append(list(r))
    # outras abas exigidas pelo arquivo real
    wb.create_sheet("DISCLAIMER")
    wb.create_sheet("DE-PARA CNAE")
    wb.save(path)


def _ipca(path: Path) -> None:
    mes = pd.date_range("2002-01-01", periods=300, freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [0.4] * len(mes)}).to_excel(path, index=False)


def test_marker():
    assert "naoautomaticas" in MARKER


def test_periodos_e_ipca(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fake_fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = aplicar_ipca(carregar_contratos(fonte, hoje=datetime(2026, 8, 16)), carregar_ipca(ipca_path))
    assert COL_IPCA in df.columns
    assert float(df[COL_IPCA].sum()) > float(df[COL_DESEMBOLSO].sum())

    p2002 = filtrar_periodo(df, datetime(2002, 1, 1), datetime(2002, 12, 31), datetime(2026, 8, 16))
    assert len(p2002) == 2
    p2 = filtrar_periodo(df, datetime(2003, 1, 1), datetime(2016, 5, 11), datetime(2026, 8, 16))
    assert len(p2) == 2
    p3 = filtrar_periodo(df, datetime(2016, 5, 12), datetime(2018, 12, 31), datetime(2026, 8, 16))
    assert len(p3) == 2
    p4 = filtrar_periodo(df, datetime(2019, 1, 1), datetime(2022, 12, 31), datetime(2026, 8, 16))
    assert len(p4) == 1
    p5 = filtrar_periodo(df, datetime(2023, 1, 1), None, datetime(2026, 8, 16))
    assert len(p5) == 1

    detalhe, totais = montar_aba_periodo(p2002)
    assert list(totais["Ano"]) == [2002]
    assert int(totais["Qtd Contratos"].iloc[0]) == 2


def test_por_cliente_ordem_decrescente(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fake_fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = aplicar_ipca(carregar_contratos(fonte, hoje=datetime(2026, 8, 16)), carregar_ipca(ipca_path))
    out = montar_aba_por_cliente(df)
    # primeiro cliente no detalhe deve ser CLIENTE ALTO (1500)
    primeiros = [r for r in out[COL_CLIENTE].tolist() if r and not str(r).startswith("TOTAL")]
    assert primeiros[0] == "CLIENTE ALTO"
    assert any(str(x).startswith("TOTAL — CLIENTE ALTO") for x in out[COL_CLIENTE])


def test_processar_workbook(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fake_fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    saida = tmp_path / "out.xlsx"
    processar(
        fonte=fonte,
        saida=saida,
        ipca_path=ipca_path,
        hoje=datetime(2026, 8, 16),
        baixar=False,
    )
    wb = load_workbook(saida, read_only=True)
    names = wb.sheetnames
    assert "Capa" in names
    assert "1_Ano_2002" in names
    assert "2_2003_a_2016-05-11" in names
    assert "3_2016-05-12_a_2018" in names
    assert "4_2019_a_2022" in names
    assert "5_2023_ate_hoje" in names
    assert "6_Por_Cliente_2002_hoje" in names
    assert "Resumo" in names
    # IPCA na aba 1
    rows = list(wb["1_Ano_2002"].iter_rows(min_row=4, max_row=4, values_only=True))
    assert COL_IPCA in rows[0]
    # totais por ano presentes
    textos = [c[0] for c in wb["1_Ano_2002"].iter_rows(min_col=1, max_col=1, values_only=True) if c[0]]
    assert "TOTAIS POR ANO" in textos
    assert "TOTAL PERÍODO" in textos
