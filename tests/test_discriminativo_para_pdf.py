"""Testes da exportação Excel → PDF (marcadores) e HTML (abas)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

from scripts.discriminativo_para_pdf import MARKER, formatar_valor, processar


def _xlsx_minimo(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Capa"
    ws["A1"] = "Discriminativo teste"
    ws["A1"].font = Font(bold=True, color="1B4F72")
    ws = wb.create_sheet("2024")
    ws["A1"] = "Ranking 2024 — taxa básica de juros real acumulada no ano"
    ws.merge_cells("A1:C1")
    ws["A4"] = "Posição"
    ws["B4"] = "País"
    ws["C4"] = "Taxa básica real\nacumulada no ano (%)"
    ws["A4"].fill = PatternFill("solid", fgColor="1B4F72")
    ws["A4"].font = Font(bold=True, color="FFFFFF")
    ws["A5"] = 1
    ws["B5"] = "Brasil"
    ws["C5"] = 0.0584
    ws["C5"].number_format = "0.00%"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    wb.save(path)
    return path


def test_formatar_percentual():
    class _C:
        value = 0.0584
        number_format = "0.00%"

    assert formatar_valor(_C()) == "5,84%"


def test_pdf_marcadores_e_html_abas(tmp_path: Path):
    xlsx = _xlsx_minimo(tmp_path / "demo.xlsx")
    pdf, html = processar(
        xlsx,
        tmp_path / "demo.pdf",
        tmp_path / "demo.html",
        titulo="Discriminativo teste",
    )
    assert pdf.exists() and pdf.stat().st_size > 500
    reader = PdfReader(str(pdf))
    assert len(reader.pages) >= 2
    texto = "\n".join((p.extract_text() or "") for p in reader.pages)
    assert "Índice" in texto or "Indice" in texto or "abas" in texto.lower()
    assert "2024" in texto
    assert "Brasil" in texto
    outlines = reader.outline
    titulos = []

    def _walk(itens):
        for it in itens:
            if isinstance(it, list):
                _walk(it)
            else:
                titulos.append(getattr(it, "title", str(it)))

    _walk(outlines)
    assert any("Índice" in t or "Indice" in t or "abas" in t.lower() for t in titulos)
    assert any("2024" in t for t in titulos)
    assert any("Capa" in t for t in titulos)

    html_txt = html.read_text(encoding="utf-8")
    assert MARKER in html_txt
    assert "name='aba'" in html_txt or 'name="aba"' in html_txt
    assert "2024" in html_txt
    assert "Brasil" in html_txt
    assert "5,84%" in html_txt
    assert "colspan='3'" in html_txt or 'colspan="3"' in html_txt
    from scripts.discriminativo_para_pdf import ler_workbook

    abas = ler_workbook(xlsx)
    aba_2024 = next(a for a in abas if a["nome"] == "2024")
    assert (1, 1, 1, 3) in aba_2024["merges"]
