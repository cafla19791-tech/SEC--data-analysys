"""Testes do ranking anual de juros reais (uma aba por ano)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from scripts.discriminativo_juros_reais_paises import (
    estatisticas_anuais,
    montar_serie_pais,
    taxa_mensal_composta,
    taxa_real_fisher,
)
from scripts.discriminativo_ranking_juros_reais import (
    COL_COD,
    COL_PAIS,
    COL_POS,
    COL_REAL,
    MARKER,
    montar_rankings,
    processar,
    ranking_do_ano,
    resumo_rankings,
    escrever_planilha,
)


def _pais_constante(
    codigo: str,
    pais_en: str,
    inicio: str,
    n_meses: int,
    nominal_aa: float,
    inflacao_am: float,
    indice0: float = 100.0,
) -> pd.DataFrame:
    """Série mensal sintética (precisa de 1 mês de IPC anterior)."""
    meses_ipc = pd.date_range(pd.Timestamp(inicio) - pd.DateOffset(months=1), periods=n_meses + 1, freq="MS")
    meses = meses_ipc[1:]
    indices = [indice0 * ((1 + inflacao_am) ** i) for i in range(len(meses_ipc))]
    cpi = pd.DataFrame({"codigo": codigo, "pais_en": pais_en, "mes": meses_ipc, "indice": indices})
    nom = pd.DataFrame(
        {"codigo": codigo, "pais_en": pais_en, "mes": meses, "nominal_aa": nominal_aa}
    )
    return montar_serie_pais(nom, cpi, codigo=codigo)


def test_estatisticas_anuais_12_meses_completo():
    serie = _pais_constante("BR", "Brazil", "2024-01-01", 12, 0.12, 0.005)
    anuais = estatisticas_anuais(serie)
    assert list(anuais["ano"]) == [2024]
    assert bool(anuais.iloc[0]["completo"])
    assert anuais.iloc[0]["n_meses"] == 12
    r_am = taxa_real_fisher(taxa_mensal_composta(0.12), 0.005)
    assert anuais.iloc[0]["real_aa"] == pytest.approx((1 + r_am) ** 12 - 1)


def test_ranking_ordena_maior_real_primeiro():
    br = estatisticas_anuais(_pais_constante("BR", "Brazil", "2024-01-01", 12, 0.12, 0.005))
    us = estatisticas_anuais(_pais_constante("US", "United States", "2024-01-01", 12, 0.05, 0.005))
    anuais = pd.concat([br, us], ignore_index=True)
    rank = ranking_do_ano(anuais, 2024)
    assert list(rank[COL_COD]) == ["BR", "US"]
    assert list(rank[COL_POS]) == [1, 2]
    assert bool(rank.iloc[0]["oficial"])


def test_ranking_incompleto_sem_posicao_quando_ha_completos():
    br = estatisticas_anuais(_pais_constante("BR", "Brazil", "2024-01-01", 12, 0.12, 0.005))
    ar = estatisticas_anuais(_pais_constante("AR", "Argentina", "2024-01-01", 6, 0.40, 0.02))
    anuais = pd.concat([br, ar], ignore_index=True)
    rank = ranking_do_ano(anuais, 2024)
    assert rank.iloc[0][COL_COD] == "BR"
    extra = rank[rank[COL_COD] == "AR"].iloc[0]
    assert pd.isna(extra[COL_POS])
    assert not bool(extra["oficial"])
    assert "parcial" in extra["Cobertura"]


def test_ranking_2026_parcial_todos_ranqueados():
    br = estatisticas_anuais(_pais_constante("BR", "Brazil", "2026-01-01", 7, 0.1425, 0.004))
    us = estatisticas_anuais(_pais_constante("US", "United States", "2026-01-01", 7, 0.043, 0.002))
    anuais = pd.concat([br, us], ignore_index=True)
    assert not bool(anuais["completo"].any())
    rank = ranking_do_ano(anuais, 2026)
    assert list(rank[COL_COD]) == ["BR", "US"]
    assert list(rank[COL_POS]) == [1, 2]
    assert rank["oficial"].all()


def test_montar_rankings_e_resumo_abas(tmp_path: Path):
    por_serie = {
        "BR": _pais_constante("BR", "Brazil", "2023-01-01", 24, 0.12, 0.004),
        "US": _pais_constante("US", "United States", "2023-01-01", 24, 0.05, 0.004),
    }
    por_ano = montar_rankings(por_serie, [2023, 2024])
    assert set(por_ano) == {2023, 2024}
    resumo = resumo_rankings(por_ano)
    assert list(resumo["ano"]) == [2023, 2024]
    assert resumo.iloc[0]["1_pais"] == "Brasil"
    assert resumo.iloc[0]["brasil_pos"] == 1

    saida = tmp_path / "rank.xlsx"
    escrever_planilha(por_ano, saida, anos_pedido=[2023, 2024])
    wb = load_workbook(saida)
    assert wb.sheetnames[:3] == ["Capa", "Resumo", "Brasil"]
    assert "2023" in wb.sheetnames and "2024" in wb.sheetnames
    capa = [c.value for row in wb["Capa"].iter_rows(max_col=2) for c in row]
    assert MARKER in capa
    headers = [c.value for c in wb["2024"][4]]
    assert headers[0] == COL_POS
    assert headers[1] == COL_PAIS
    assert headers[3] == COL_REAL
    assert wb["2024"]["B5"].value == "Brasil"
    assert wb["2024"]["A5"].value == 1


def test_processar_cli_sintetico(tmp_path: Path):
    """Smoke: processar com ZIPs mínimos já cobertos no módulo-base; aqui só o recorte."""
    from tests.test_discriminativo_juros_reais_paises import _cbpol_csv, _cpi_csv, _zip_csv

    cbpol = _zip_csv(tmp_path / "cbpol.zip", "c.csv", _cbpol_csv())
    cpi = _zip_csv(tmp_path / "cpi.zip", "p.csv", _cpi_csv())
    saida = tmp_path / "rank.xlsx"
    processar(
        tmp_path,
        saida,
        baixar=False,
        ano_inicio=2023,
        ano_fim=2023,
        paises={"BR"},
        cbpol_zip=cbpol,
        cpi_zip=cpi,
    )
    wb = load_workbook(saida)
    assert "2023" in wb.sheetnames
    assert "Capa" in wb.sheetnames
