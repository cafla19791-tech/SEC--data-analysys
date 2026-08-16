"""Testes do discriminativo Sudam/Sudene 75% IRPJ."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from scripts.discriminativo_sudam_sudene_75_ipca import (
    COL_ANO,
    COL_CNPJ,
    COL_IPCA,
    COL_NOME,
    COL_VALOR,
    MARKER,
    aplicar_ipca,
    carregar_renuncias_75,
    processar,
    resumo_por_ano,
)


def _fonte(path: Path) -> None:
    wb = Workbook()
    # índice
    ws = wb.active
    ws.title = "Índice"
    ws.append(["idx"])
    # dados
    ws = wb.create_sheet("Renúncias Fiscais Sudam-Sudene")
    ws.append(
        [
            "Ano-Calendário",
            "CNPJ",
            "Beneficiário",
            "Nome Fantasia",
            "UF",
            "Município",
            "CNAE",
            "Tipo de Renúncia",
            "Benefício Fiscal",
            "Valor Renunciado(R$)",
            "Valor Renunciado atualizado pelo IPCA até 31/05/2026 (R$)",
        ]
    )
    rows = [
        (2018, "11.111.111/0001-11", "EMP A", "A", "PE", "RECIFE", "x", "Decl", "Sudam/Sudene - Redução 75% Projeto Setor Prioritário", 1000, 0),
        (2019, "11.111.111/0001-11", "EMP A", "A", "PE", "RECIFE", "x", "Decl", "Sudam/Sudene - Redução 75% Projeto Setor Prioritário", 500, 0),
        (2019, "22.222.222/0001-22", "EMP B", "B", "AM", "MANAUS", "x", "Decl", "Sudam/Sudene - Redução 75% Projeto Setor Prioritário", 200, 0),
        (2019, "33.333.333/0001-33", "EMP C", "C", "BA", "SSA", "x", "Decl", "Sudam/Sudene - Redução por Reinvestimento", 9999, 0),
    ]
    for r in rows:
        ws.append(list(r))
    wb.create_sheet("IPCA_Fatores")
    wb.create_sheet("Resumo_por_Empresa")
    wb.save(path)


def _ipca(path: Path) -> None:
    mes = pd.date_range("2015-01-01", periods=150, freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [0.4] * len(mes)}).to_excel(path, index=False)


def test_marker():
    assert "sudam" in MARKER


def test_filtra_apenas_75(tmp_path: Path):
    fonte = tmp_path / "r.xlsx"
    _fonte(fonte)
    df = carregar_renuncias_75(fonte)
    assert len(df) == 3
    assert set(df[COL_CNPJ]) == {"11.111.111/0001-11", "22.222.222/0001-22"}


def test_ipca_e_por_ano(tmp_path: Path):
    fonte = tmp_path / "r.xlsx"
    _fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = aplicar_ipca(carregar_renuncias_75(fonte), carregar_ipca(ipca_path))
    assert COL_IPCA in df.columns
    assert float(df[COL_IPCA].sum()) > float(df[COL_VALOR].sum())
    por = resumo_por_ano(df)
    assert list(por[COL_ANO]) == [2018, 2019]


def test_processar_workbook(tmp_path: Path):
    fonte = tmp_path / "r.xlsx"
    _fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    saida = tmp_path / "out.xlsx"
    info = processar(fonte=fonte, saida=saida, ipca_path=ipca_path)
    wb = load_workbook(saida, read_only=True)
    assert set(wb.sheetnames) >= {"Capa", "Empresas", "Por_Ano", "Empresa_Ano"}
    assert info["empresas"].iloc[0][COL_NOME] == "EMP A"
    assert len(info["empresas"]) == 2
