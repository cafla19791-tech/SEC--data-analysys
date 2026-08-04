"""Testes dos discriminativos OPERAÇÕES DIRETAS + IPCA."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.discriminativos_diretas_ipca import (
    COL_IPCA,
    COL_VALOR,
    atualizar_valores_ipca,
    filtrar_periodo,
    main,
    montar_linhas_aba,
    preparar_base,
    escrever_workbook,
)
from scripts.calcular_diretas_ipca_selic import carregar_ipca


def _ipca(path: Path, taxa: float = 0.5) -> None:
    mes = pd.date_range("2000-01-01", "2026-06-01", freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [taxa] * len(mes)}).to_excel(path, index=False)


def _base_xlsx(path: Path) -> None:
    rows = [
        {
            "Cliente": "BETA SA",
            "CNPJ": "2",
            "UF": "RJ",
            "Número do contrato": "B1",
            "Data da contratacao": "15/03/2002",
            "Valor desembolsado R$": 1000.0,
            "Custo financeiro": "TJLP",
            "Juros": 2.0,
            "Prazo - amortizaca (meses)": 24,
        },
        {
            "Cliente": "ALFA SA",
            "CNPJ": "1",
            "UF": "SP",
            "Número do contrato": "A1",
            "Data da contratacao": "10/01/2002",
            "Valor desembolsado R$": 2000.0,
            "Custo financeiro": "TJLP",
            "Juros": 2.0,
            "Prazo - amortizaca (meses)": 36,
        },
        {
            "Cliente": "ALFA SA",
            "CNPJ": "1",
            "UF": "SP",
            "Número do contrato": "A2",
            "Data da contratacao": "20/06/2002",
            "Valor desembolsado R$": 500.0,
            "Custo financeiro": "TJLP",
            "Juros": 2.0,
            "Prazo - amortizaca (meses)": 12,
        },
        {
            "Cliente": "GAMA SA",
            "CNPJ": "3",
            "UF": "MG",
            "Número do contrato": "G1",
            "Data da contratacao": "01/05/2010",
            "Valor desembolsado R$": 3000.0,
            "Custo financeiro": "TJLP",
            "Juros": 3.0,
            "Prazo - amortizaca (meses)": 48,
        },
        {
            "Cliente": "DELTA SA",
            "CNPJ": "4",
            "UF": "BA",
            "Número do contrato": "D1",
            "Data da contratacao": "01/01/2020",
            "Valor desembolsado R$": 4000.0,
            "Custo financeiro": "TLP",
            "Juros": 1.5,
            "Prazo - amortizaca (meses)": 60,
        },
        {
            "Cliente": "EPSILON SA",
            "CNPJ": "5",
            "UF": "PR",
            "Número do contrato": "E1",
            "Data da contratacao": "15/08/2024",
            "Valor desembolsado R$": 8000.0,
            "Custo financeiro": "TLP",
            "Juros": 1.0,
            "Prazo - amortizaca (meses)": 24,
        },
    ]
    pd.DataFrame(rows).to_excel(path, index=False)


def test_ipca_e_periodos(tmp_path: Path):
    excel = tmp_path / "ops.xlsx"
    _base_xlsx(excel)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)

    raw = pd.read_excel(excel)
    base = preparar_base(raw)
    ipca = carregar_ipca(ipca_path)
    base["_valor_ipca"] = atualizar_valores_ipca(
        base["_data"], base["_valor"], ipca, datetime(2026, 6, 30)
    )
    assert (base["_valor_ipca"] > base["_valor"]).all()

    a2002 = filtrar_periodo(base, 2002, 2002)
    assert list(a2002["_cliente"].unique()) == ["ALFA SA", "BETA SA"]
    assert len(a2002) == 3
    # ordenado por cliente + data
    assert list(a2002["numero_contrato"]) == ["A1", "A2", "B1"]

    header, linhas, quebras = montar_linhas_aba(a2002)
    assert COL_IPCA in header
    assert COL_VALOR in header
    # 3 contratos + 2 subtotais + 1 total
    assert len(linhas) == 6
    assert len(quebras) == 1  # quebra entre ALFA e BETA
    sub_alfa = next(r for r in linhas if str(r[0]).startswith("SUBTOTAL — ALFA"))
    assert abs(sub_alfa[header.index(COL_VALOR)] - 2500.0) < 1e-9


def test_main_workbook(tmp_path: Path):
    excel = tmp_path / "ops.xlsx"
    _base_xlsx(excel)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    saida = tmp_path / "out.xlsx"
    rc = main(
        [
            "--excel",
            str(excel),
            "--ipca",
            str(ipca_path),
            "--saida",
            str(saida),
            "--data-ref",
            "2026-06-30",
        ]
    )
    assert rc == 0
    wb = load_workbook(saida)
    assert wb.sheetnames == ["2002", "2003-2018", "2019-2022", "2023-atual"]
    ws = wb["2002"]
    assert ws.cell(1, 1).value == "Cliente"
    assert any("IPCA-30 DE JUNHO DE 2026" in str(c.value) for c in ws[1])
    # page breaks registrados
    assert len(ws.row_breaks) >= 1
    # aba 2023 tem EPSILON
    vals = [c.value for row in wb["2023-atual"].iter_rows(min_row=2, max_col=1) for c in row]
    assert any(v and "EPSILON" in str(v) for v in vals)
