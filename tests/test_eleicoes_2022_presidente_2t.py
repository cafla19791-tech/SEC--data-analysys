"""Testes da planilha do 2º turno presidencial de 2022."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from scripts.eleicoes_2022_presidente_2t import (
    gravar_xlsx,
    montar_tabelas,
    parse_uf,
    pct_oficial,
)


def _payload(uf: str, lula: int, bolso: int) -> dict:
    validos = lula + bolso
    return {
        "cdabr": uf,
        "vv": str(validos),
        "vb": "10",
        "vn": "20",
        "c": str(validos + 30),
        "a": "5",
        "e": str(validos + 35),
        "pst": "100,00",
        "cand": [
            {"n": "13", "vap": str(lula), "pvap": "50,20"},
            {"n": "22", "vap": str(bolso), "pvap": "49,80"},
        ],
    }


def test_pct_oficial_virgula():
    assert pct_oficial("50,90") == 50.90
    assert pct_oficial("1.620,97") == 1620.97


def test_parse_uf_vencedor_e_diferenca():
    row = parse_uf(_payload("mg", 6190960, 6141310))
    assert row["uf"] == "MG"
    assert row["unidade"] == "Minas Gerais"
    assert row["vencedor"] == "Lula"
    assert row["diferenca_votos"] == 49650
    assert row["lula_votos"] == 6190960


def test_montar_tabelas_separa_exterior():
    linhas = [
        parse_uf(_payload("sp", 100, 200)),
        parse_uf(_payload("df", 40, 60)),
        parse_uf(_payload("zz", 10, 8)),
    ]
    # ajustar percentuais oficiais já parseados
    ufs, exterior, brasil = montar_tabelas(linhas)
    assert list(ufs["uf"]) == ["SP", "DF"]
    assert list(exterior["uf"]) == ["ZZ"]
    assert int(brasil.iloc[0]["lula_votos"]) == 150
    assert int(brasil.iloc[0]["bolsonaro_votos"]) == 268


def test_gravar_xlsx(tmp_path: Path):
    linhas = [
        parse_uf(_payload("sp", 100, 200)),
        parse_uf(_payload("df", 40, 60)),
        parse_uf(_payload("zz", 10, 8)),
    ]
    ufs, exterior, brasil = montar_tabelas(linhas)
    path = tmp_path / "eleicoes.xlsx"
    gravar_xlsx(ufs, exterior, brasil, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Por_UF", "Brasil_e_exterior", "Fonte"]
    ws = wb["Por_UF"]
    assert ws["A1"].value == "UF"
    assert ws["D1"].value == "Lula (votos)"
    assert ws.max_row == 3
    assert {ws["A2"].value, ws["A3"].value} == {"SP", "DF"}
