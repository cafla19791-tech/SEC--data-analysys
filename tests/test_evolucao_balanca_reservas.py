"""Testes da agregação anual da balança comercial e das reservas."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from openpyxl import load_workbook

from scripts.evolucao_balanca_reservas import (
    ANO_FIM,
    ANO_INICIO,
    BORDA_CONTINUA,
    _sgs_para_mensal,
    agregar_anual,
    desenhar_tabela_png,
    exportar_excel_grade,
    exportar_tabelas,
    fases_historicas,
    gerar_graficos,
    gerar_relatorio,
    gerar_tabelas_png,
    parse_sgs_soap_xml,
    tabela_html,
)


def _mensal(ano_ini: int, ano_fim: int, valor_fn) -> pd.DataFrame:
    meses = pd.date_range(f"{ano_ini}-01-01", f"{ano_fim}-12-01", freq="MS")
    return pd.DataFrame({"mes": meses, "valor": [valor_fn(m) for m in meses]})


def _series_sinteticas() -> dict[str, pd.DataFrame]:
    # 1995: déficit 12 × 1 = 12; 1996: superávit 12 × 2 = 24
    # exportações 10/mês; importações −9 (1995) e −8 (1996) → volume 9 e 8
    def saldo(m: pd.Timestamp) -> float:
        return -1.0 if m.year == 1995 else 2.0

    def exp(m: pd.Timestamp) -> float:
        return 10.0

    def imp(m: pd.Timestamp) -> float:
        return -9.0 if m.year == 1995 else -8.0

    def reservas(m: pd.Timestamp) -> float:
        return 100.0 + (m.year - 1995) * 50 + m.month

    return {
        "saldo_comercial": _mensal(1995, 1996, saldo),
        "exportacoes": _mensal(1995, 1996, exp),
        "importacoes": _mensal(1995, 1996, imp),
        "reservas": _mensal(1995, 1996, reservas),
    }


def test_agregar_anual_soma_fluxos_e_estoque_dezembro():
    anual = agregar_anual(_series_sinteticas())
    assert list(anual["ano"]) == [1995, 1996]
    assert anual.loc[0, "saldo_comercial"] == -12.0
    assert anual.loc[1, "saldo_comercial"] == 24.0
    assert anual.loc[0, "exportacoes"] == 120.0
    assert anual.loc[1, "exportacoes"] == 120.0
    # débito BPM6 invertido para volume
    assert anual.loc[0, "importacoes"] == 108.0
    assert anual.loc[1, "importacoes"] == 96.0
    assert anual.loc[0, "saldo_reconstruido"] == 12.0
    # reservas = dezembro
    assert anual.loc[0, "reservas"] == 100.0 + 12
    assert anual.loc[1, "reservas"] == 150.0 + 12
    assert anual.loc[1, "var_reservas"] == 50.0


def test_importacoes_ja_positivas_nao_sao_invertidas():
    series = _series_sinteticas()
    series["importacoes"] = series["importacoes"].assign(valor=9.0)
    anual = agregar_anual(series)
    assert anual.loc[0, "importacoes"] == 108.0


def test_filtra_anos_fora_de_1995_2025():
    extra = pd.DataFrame(
        {
            "mes": pd.to_datetime(["1994-12-01", "2026-01-01"]),
            "valor": [999.0, 999.0],
        }
    )
    series = _series_sinteticas()
    series["saldo_comercial"] = pd.concat([extra, series["saldo_comercial"]], ignore_index=True)
    anual = agregar_anual(series)
    assert anual["ano"].min() >= ANO_INICIO
    assert anual["ano"].max() <= ANO_FIM
    assert 1994 not in set(anual["ano"])
    assert 2026 not in set(anual["ano"])


def test_fases_e_relatorio(tmp_path: Path):
    # Estende a série sintética para cobrir todos os recortes
    def saldo(m: pd.Timestamp) -> float:
        return 1.0 if m.year >= 1999 else -1.0

    def const(v: float):
        return lambda m: v

    def reservas(m: pd.Timestamp) -> float:
        return float(m.year) * 10 + m.month

    series = {
        "saldo_comercial": _mensal(1995, 2025, saldo),
        "exportacoes": _mensal(1995, 2025, const(5.0)),
        "importacoes": _mensal(1995, 2025, const(4.0)),
        "reservas": _mensal(1995, 2025, reservas),
    }
    anual = agregar_anual(series)
    assert len(anual) == 31
    fases = fases_historicas(anual)
    assert fases[0]["periodo"] == "1995–1998"
    assert fases[0]["saldo_acumulado"] == -48.0  # 4 anos × 12 × −1
    assert fases[-1]["periodo"] == "2022–2025"
    assert fases[-1]["reservas_fim"] == 2025 * 10 + 12

    rel = gerar_relatorio(anual, tmp_path)
    texto = rel.read_text(encoding="utf-8")
    assert "1995–2025" in texto
    assert "22707" in texto
    assert "3546" in texto
    assert "<table" in texto
    assert "border-collapse:collapse" in texto
    assert "border:1px solid #1a1a1a" in texto
    assert ">1995<" in texto
    assert ">2025<" in texto
    assert "US$ " in texto
    assert "," in texto  # separador decimal brasileiro

    saidas = exportar_tabelas(anual, tmp_path)
    assert all(p.exists() for p in saidas)
    lido = pd.read_csv(saidas[0])
    # CSV em US$ bilhões (÷ 1000)
    assert abs(lido.loc[lido["ano"] == 1995, "saldo_comercial"].iloc[0] - (-0.012)) < 1e-9
    assert saidas[2].suffix == ".xlsx"


def test_tabela_html_grade_continua():
    html = tabela_html(["A", "B"], [["1", "2"], ["3", "4"]])
    assert html.count("border:1px solid #1a1a1a") >= 6  # 2 th + 4 td
    assert "border-collapse:collapse" in html
    assert html.count("<tr>") == 3


def test_excel_borda_continua(tmp_path: Path):
    anual = agregar_anual(_series_sinteticas())
    path = exportar_excel_grade(anual, tmp_path)
    wb = load_workbook(path)
    ws = wb["Serie_anual"]
    assert ws["A1"].border.left.style == BORDA_CONTINUA.left.style
    assert ws["B2"].border.top.style == "thin"
    assert ws["B2"].border.bottom.style == "thin"
    assert ws["B2"].border.left.style == "thin"
    assert ws["B2"].border.right.style == "thin"


def test_tabelas_png(tmp_path: Path):
    anual = agregar_anual(_series_sinteticas())
    paths = gerar_tabelas_png(anual, tmp_path)
    assert len(paths) == 2
    assert all(p.exists() and p.stat().st_size > 0 for p in paths)
    extra = desenhar_tabela_png(["X"], [["1"]], tmp_path / "uma.png", "T")
    assert extra.exists()


def test_graficos(tmp_path: Path):
    anual = agregar_anual(_series_sinteticas())
    paths = gerar_graficos(anual, tmp_path)
    assert len(paths) == 4
    assert all(p.exists() and p.stat().st_size > 0 for p in paths)


def test_parse_sgs_soap_xml_e_fim_de_mes():
    xml = (
        "&lt;SERIE&gt;&lt;ITEM&gt;&lt;DATA&gt;1/2024&lt;/DATA&gt;"
        "&lt;VALOR&gt;10,5&lt;/VALOR&gt;&lt;/ITEM&gt;"
        "&lt;ITEM&gt;&lt;DATA&gt;15/01/2024&lt;/DATA&gt;"
        "&lt;VALOR&gt;11&lt;/VALOR&gt;&lt;/ITEM&gt;"
        "&lt;ITEM&gt;&lt;DATA&gt;02/2024&lt;/DATA&gt;"
        "&lt;VALOR&gt;12&lt;/VALOR&gt;&lt;/ITEM&gt;&lt;/SERIE&gt;"
    )
    bruto = parse_sgs_soap_xml(xml)
    assert len(bruto) == 3
    mensal = _sgs_para_mensal(bruto)
    assert list(mensal["mes"].dt.month) == [1, 2]
    assert mensal.loc[0, "valor"] == 11.0
    assert mensal.loc[1, "valor"] == 12.0
