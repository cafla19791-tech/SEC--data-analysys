"""Fluxos a partir de BNDES_INDIRETAS_NUMERADOS (uma aba por ano)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from scripts.fluxos_por_ano_contrato_numerados import listar_abas_ano, processar
from scripts.gerar_fluxos import gerar_fluxos_contrato, normalizar_colunas


def _make_numerados(path: Path) -> Path:
    header = [
        "Número do contrato",
        "Data da contratação",
        "Valor desembolsado R$",
        "Instituição Financeira Credenciada",
        "Custo financeiro",
        "Juros",
        "Prazo - Carência (meses)",
        "Prazo - Amortização (meses)",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "2002"
    ws.append(header)
    ws.append(["1-2002", "2002-01-10", 1000, "BANCO X", "TAXA FIXA", 5, 0, 3])
    ws.append(["2-2002", "2002-06-10", 2000, "BANCO X", "TAXA FIXA", 5, 0, 3])
    ws2 = wb.create_sheet("2003")
    ws2.append(header)
    ws2.append(["1-2003", "2003-03-15", 1500, "BANCO Y", "TAXA FIXA", 4, 0, 2])
    wb.save(path)
    return path


def test_listar_abas_ano(tmp_path: Path):
    p = _make_numerados(tmp_path / "num.xlsx")
    assert listar_abas_ano(p) == ["2002", "2003"]


def test_numero_contrato_no_fluxo():
    rows = gerar_fluxos_contrato(
        data_contr=pd.Timestamp("2022-12-12"),
        valor=3000.0,
        taxa_juros_aa=0.05,
        carencia=0,
        n=3,
        contrato_id=0,
        juros_pct=5.0,
        custo_financeiro="TAXA FIXA",
        numero_contrato="7-2022",
        selic_aa=0.145,
    )
    assert len(rows) == 3
    assert {r["numero_contrato"] for r in rows} == {"7-2022"}
    assert {r["ano_contrato"] for r in rows} == {2022}


def test_processar_cria_abas_e_csv(tmp_path: Path):
    numerados = _make_numerados(tmp_path / "BNDES_INDIRETAS_NUMERADOS.xlsx")
    saida = tmp_path / "saida"
    xlsx = processar(numerados, saida, fatores=0.145, lote=10)
    assert xlsx.exists()
    wb = load_workbook(xlsx)
    assert "2002" in wb.sheetnames
    assert "2003" in wb.sheetnames
    assert "RESUMO" in wb.sheetnames

    csv2002 = saida / "fluxos_por_ano_contrato" / "2002.csv"
    assert csv2002.exists()
    df = pd.read_csv(csv2002)
    assert "ano_contrato" in df.columns
    assert set(df["ano_contrato"]) == {2002}
    assert "numero_contrato" in df.columns
    assert set(df["numero_contrato"]) == {"1-2002", "2-2002"}
    # 2 contratos × 3 parcelas
    assert len(df) == 6

    df3 = pd.read_csv(saida / "fluxos_por_ano_contrato" / "2003.csv")
    assert len(df3) == 2
    assert set(df3["numero_contrato"]) == {"1-2003"}


def test_retomar_pula_ano_existente(tmp_path: Path):
    numerados = _make_numerados(tmp_path / "BNDES_INDIRETAS_NUMERADOS.xlsx")
    saida = tmp_path / "saida"
    processar(numerados, saida, fatores=0.145, lote=10)
    csv2002 = saida / "fluxos_por_ano_contrato" / "2002.csv"
    mtime = csv2002.stat().st_mtime
    processar(numerados, saida, fatores=0.145, lote=10, retomar=True)
    assert csv2002.stat().st_mtime == mtime
    resumo = pd.read_csv(saida / "fluxos_por_ano_contrato" / "RESUMO.csv")
    assert set(resumo["status"]) == {"retomado"}


def test_prints_cp1252_safe():
    """ContAgil WinPython usa cp1252; emoji no print derruba a massa."""
    import scripts.gerar_fluxos as gf
    import scripts.fluxos_por_ano_contrato_numerados as fa

    for path in (Path(gf.__file__), Path(fa.__file__)):
        text = path.read_text(encoding="utf-8")
        for i, line in enumerate(text.splitlines(), 1):
            if "print(" not in line:
                continue
            try:
                line.encode("cp1252")
            except UnicodeEncodeError as exc:
                raise AssertionError(f"{path.name}:{i} nao e cp1252-safe: {line!r}") from exc


def test_normalizar_preserva_numero():
    bruto = pd.DataFrame(
        {
            "Número do contrato": ["1-2002"],
            "Data da contratação": ["2002-01-10"],
            "Valor desembolsado R$": [1000],
            "Instituição Financeira Credenciada": ["BANCO X"],
            "Custo financeiro": ["TAXA FIXA"],
            "Juros": [5],
            "Prazo - Carência (meses)": [0],
            "Prazo - Amortização (meses)": [3],
        }
    )
    out = normalizar_colunas(bruto)
    assert "numero_contrato" in out.columns
    assert out.loc[0, "numero_contrato"] == "1-2002"


def test_selic_serie_duck_type_nao_vira_float():
    """Regressao ContAgil: duas classes SelicSerie via importlib → float(serie)."""
    import numpy as np

    from scripts.gerar_fluxos import SelicSerie, _eh_selic_serie, _resolver_segundo_arg, gerar_e_gravar_fluxos

    class FakeSerie:
        def __init__(self):
            self.datas = np.array(["2020-01-01", "2026-06-01"], dtype="datetime64[ns]")
            self.fatores = np.array([1.0, 1.5], dtype=float)
            self.fator_referencia = 1.5
            self.origem = "fake"

        def idx_proximo(self, data):
            return 0

        def capitalizar(self, valor, data_fluxo, data_impacto=None):
            return round(float(valor) * 1.1, 2)

    fake = FakeSerie()
    assert _eh_selic_serie(fake)
    taxa, serie, orig = _resolver_segundo_arg(fake, None)
    assert serie is fake
    assert orig is None
    assert isinstance(taxa, float)

    # Serie real tambem
    real = SelicSerie(
        np.array(["2000-01-01", "2026-06-01"], dtype="datetime64[ns]"),
        np.array([1.0, 2.0], dtype=float),
        origem="t",
        fator_referencia=2.0,
    )
    bruto = pd.DataFrame(
        {
            "Número do contrato": ["1-2002"],
            "Data da contratação": ["2002-01-10"],
            "Valor desembolsado R$": [1000],
            "Custo financeiro": ["TAXA FIXA"],
            "Juros": [5],
            "Prazo - Carência (meses)": [0],
            "Prazo - Amortização (meses)": [3],
        }
    )
    contratos = normalizar_colunas(bruto)
    stats = gerar_e_gravar_fluxos(
        contratos, real, saida_csv=Path("/tmp/test_duck_selic.csv"), gravar_excel=False
    )
    assert stats["parcelas"] > 0


def test_processar_com_serie_contagil(tmp_path: Path):
    """processar() + SelicSerie do contagil_fluxos_seguro (mesmo path ContAgil)."""
    import numpy as np

    from scripts.contagil_fluxos_seguro import carregar_fatores_mensais

    fatores = tmp_path / "fator_acumulado_SELIC_TJLP_TLP.xlsx"
    datas = pd.date_range("2000-01-01", "2026-07-01", freq="MS")
    pd.DataFrame(
        {"Data": datas, "Fator_Acumulado": np.cumprod(np.full(len(datas), 1.005))}
    ).to_excel(fatores, index=False)
    serie = carregar_fatores_mensais(fatores)

    numerados = _make_numerados(tmp_path / "BNDES_INDIRETAS_NUMERADOS.xlsx")
    saida = tmp_path / "saida"
    xlsx = processar(numerados, saida, fatores=serie, lote=10)
    assert xlsx.exists()
    resumo = pd.read_csv(saida / "fluxos_por_ano_contrato" / "RESUMO.csv")
    assert set(resumo["status"]) == {"ok"}
    assert int(resumo["qtd_parcelas"].sum()) > 0
