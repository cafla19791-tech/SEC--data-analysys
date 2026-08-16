"""Testes da impressão PDF de operações diretas/indiretas."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from scripts.imprimir_operacoes_pdf import MARKER, montar_pdf, processar, resumo_por_ano
from scripts.discriminativo_naoautomaticas_ipca import (
    COL_CLIENTE,
    COL_DATA,
    COL_DESEMBOLSO,
    COL_IPCA,
)


def _fonte(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SITE"
    ws.append(["ops"])
    ws.append(["p"])
    ws.append(["d"])
    ws.append([])
    ws.append(
        [
            "Cliente",
            "CNPJ",
            "Número do contrato",
            "Data da contratação",
            "Valor contratado  R$",
            "Valor desembolsado R$",
            "Forma de apoio",
            "UF",
            "Situação do contrato",
            "Produto",
        ]
    )
    ws.append(["A", "1", "1", datetime(2010, 1, 1), 100, 100, "DIRETA", "SP", "ATIVO", "P"])
    ws.append(["B", "2", "2", datetime(2015, 1, 1), 50, 50, "INDIRETA", "RJ", "ATIVO", "P"])
    ws.append(["A", "1", "3", datetime(2020, 1, 1), 200, 200, "DIRETA", "SP", "ATIVO", "P"])
    wb.create_sheet("DISCLAIMER")
    wb.create_sheet("DE-PARA CNAE")
    wb.save(path)


def _ipca(path: Path) -> None:
    mes = pd.date_range("2009-01-01", periods=220, freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [0.4] * len(mes)}).to_excel(path, index=False)


def test_marker():
    assert "pdf" in MARKER


def test_montar_pdf_minimo(tmp_path: Path):
    df = pd.DataFrame(
        {
            COL_CLIENTE: ["X", "Y"],
            COL_DATA: pd.to_datetime(["2010-01-01", "2011-01-01"]),
            COL_DESEMBOLSO: [10.0, 20.0],
            COL_IPCA: [15.0, 30.0],
        }
    )
    out = tmp_path / "t.pdf"
    montar_pdf(out, titulo="T", subtitulo="S", df=df)
    assert out.exists() and out.stat().st_size > 500
    assert list(resumo_por_ano(df)["ano"]) == [2010, 2011]


def test_processar_dois_pdfs(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fonte(fonte)
    ipca = tmp_path / "ipca.xlsx"
    _ipca(ipca)
    saida = tmp_path / "pdfs"
    paths = processar(
        fonte_naoauto=fonte,
        saida_dir=saida,
        automaticas_xlsx=None,
        ipca_path=ipca,
        hoje=datetime(2026, 8, 16),
        baixar=False,
    )
    assert paths["diretas"].exists()
    assert paths["indiretas"].exists()
    assert paths["diretas"].stat().st_size > 500
    assert paths["indiretas"].stat().st_size > 500
