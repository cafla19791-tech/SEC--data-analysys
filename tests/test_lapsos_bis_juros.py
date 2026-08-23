"""Testes da acumulação de taxas básicas do BIS por lapso Selic."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.lapsos_bis_juros import (
    acumulado_paises,
    gerar_planilha,
    pregao_do_lapso,
    taxa_acumulada,
    taxas_no_pregao,
)


def test_taxa_acumulada_um_ano_de_pregões() -> None:
    s = pd.Series([12.0] * 252)
    esperado = ((1.0 + 0.12 / 252.0) ** 252 - 1.0) * 100.0
    assert abs(taxa_acumulada(s) - esperado) < 1e-9


def test_taxa_acumulada_vazia() -> None:
    assert pd.isna(taxa_acumulada(pd.Series([float("nan"), float("nan")])))


def test_taxas_no_pregao_repete_ultima_vigente() -> None:
    serie = pd.Series(
        [10.0, 11.0],
        index=pd.to_datetime(["2015-01-05", "2015-01-08"]),
    )
    pregao = pd.DatetimeIndex(pd.to_datetime(["2015-01-06", "2015-01-07", "2015-01-09"]))
    vals = taxas_no_pregao(serie, pregao)
    assert list(vals) == [10.0, 10.0, 11.0]


def test_acumulado_paises_brasil_primeiro() -> None:
    idx = pd.to_datetime(["2015-01-05", "2015-01-06", "2015-01-07"])
    bis = pd.DataFrame({"US": [0.25, 0.25, 0.25], "BR": [11.75, 11.75, 12.25]}, index=idx)
    pregao = pd.DatetimeIndex(idx)
    tab = acumulado_paises(bis, pregao)
    assert list(tab["codigo"]) == ["BR", "US"]
    assert tab.iloc[0]["pais"] == "Brasil"
    assert tab.iloc[0]["n_pregao"] == 3
    assert tab.iloc[0]["taxa_ini"] == 11.75
    assert tab.iloc[0]["taxa_fim"] == 12.25
    br = taxa_acumulada(pd.Series([11.75, 11.75, 12.25]))
    assert abs(tab.iloc[0]["taxa_acumulada"] - br) < 1e-12


def test_pregao_do_lapso_inclusivo() -> None:
    pregao = pd.DatetimeIndex(pd.to_datetime(["2014-12-03", "2014-12-04", "2014-12-05", "2015-01-21", "2015-01-22"]))
    dias = pregao_do_lapso(pregao, "2014-12-04", "2015-01-21")
    assert list(dias.strftime("%Y-%m-%d")) == ["2014-12-04", "2014-12-05", "2015-01-21"]


def test_gerar_planilha_uma_aba_por_lapso(tmp_path: Path) -> None:
    lapsos = pd.DataFrame(
        {
            "ordem": [105, 106],
            "selic": [11.75, 12.25],
            "inicio": pd.to_datetime(["2014-12-04", "2015-01-22"]),
            "fim": pd.to_datetime(["2015-01-21", "2015-01-23"]),
            "n_pregao": [3, 2],
        }
    )
    idx = pd.date_range("2014-12-01", "2015-01-25", freq="D")
    bis = pd.DataFrame(
        {
            "BR": [11.75] * len(idx),
            "US": [0.125] * len(idx),
        },
        index=idx,
    )
    pregao = pd.DatetimeIndex(pd.to_datetime(["2014-12-04", "2014-12-05", "2015-01-21", "2015-01-22", "2015-01-23"]))
    path = tmp_path / "bis.xlsx"
    gerar_planilha(lapsos, bis, pregao, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Indice", "Lapso 105", "Lapso 106"]
    ws = wb["Lapso 105"]
    assert "11,75" in str(ws["A2"].value)
    assert ws["A7"].value == "País"
    assert ws["A8"].value == "Brasil"
    assert ws["F8"].value is not None
