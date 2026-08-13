"""Testes da numeração N-AAAA de operações indiretas."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.gerar_fluxos import _numerar_contratos_por_ano, load_from_excel
from scripts.numerar_contratos_indiretas import (
    atribuir_numero_contrato_anual,
    main,
    numerar_sequencial_ano,
    processar_pasta,
)


def _xlsx_ano(path: Path, ano: int, n: int) -> None:
    rows = []
    for i in range(n):
        rows.append(
            {
                "data_da_contratacao": f"{ano}-0{(i % 9) + 1}-15",
                "valor_desembolsado_reais": 1000.0 * (i + 1),
                "juros": 5.0,
                "prazo_carencia_meses": 3,
                "prazo_amortizacao_meses": 24,
                "instituicao_financeira_credenciada": "BANCO TESTE",
                "custo_financeiro": "TAXA FIXA",
            }
        )
    pd.DataFrame(rows).to_excel(path, index=False)


def test_atribuir_reinicia_por_ano():
    df = pd.DataFrame(
        {
            "data_contratacao": pd.to_datetime(
                ["2002-01-10", "2002-06-01", "2003-02-01", "2003-03-01", "2003-04-01"]
            )
        }
    )
    nums = atribuir_numero_contrato_anual(df)
    assert list(nums) == ["1-2002", "2-2002", "1-2003", "2-2003", "3-2003"]


def test_numerar_sequencial_ano():
    df = pd.DataFrame({"x": [1, 2, 3]})
    out = numerar_sequencial_ano(df, 2002)
    assert list(out["numero_contrato"]) == ["1-2002", "2-2002", "3-2002"]


def test_gerar_fluxos_helper():
    datas = pd.to_datetime(["2002-01-01", "2002-02-01", "2004-01-01"])
    assert list(_numerar_contratos_por_ano(datas)) == ["1-2002", "2-2002", "1-2004"]


def test_processar_pasta_abas(tmp_path: Path):
    dados = tmp_path / "dados"
    dados.mkdir()
    _xlsx_ano(dados / "BNDES INDIRETAS 2002.xlsx", 2002, 3)
    _xlsx_ano(dados / "BNDES INDIRETAS 2003.xlsx", 2003, 2)
    saida = tmp_path / "out.xlsx"
    processar_pasta(dados, saida, ano_min=2002)

    wb = load_workbook(saida)
    assert wb.sheetnames == ["2002", "2003"]
    ws = wb["2002"]
    assert ws.cell(1, 1).value == "Número do contrato"
    assert ws.cell(2, 1).value == "1-2002"
    assert ws.cell(3, 1).value == "2-2002"
    assert ws.cell(4, 1).value == "3-2002"
    assert wb["2003"].cell(2, 1).value == "1-2003"
    assert wb["2003"].cell(3, 1).value == "2-2003"


def test_load_from_excel_numera(tmp_path: Path):
    path = tmp_path / "BNDES INDIRETAS 2002.xlsx"
    _xlsx_ano(path, 2002, 2)
    df = load_from_excel(path)
    assert list(df["numero_contrato"]) == ["1-2002", "2-2002"]


def test_main_excel(tmp_path: Path):
    excel = tmp_path / "BNDES INDIRETAS 2002.xlsx"
    _xlsx_ano(excel, 2002, 2)
    saida = tmp_path / "num.xlsx"
    assert main(["--excel", str(excel), "--saida", str(saida)]) == 0
    assert saida.exists()
