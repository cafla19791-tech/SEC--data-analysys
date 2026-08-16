"""Testes da planilha de indiretas automáticas com IPCA até jul/2026."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.planilha_indiretas_automaticas_ipca import (
    COL_AGENTE,
    COL_CLIENTE,
    COL_DATA,
    COL_IPCA,
    COL_VALOR,
    MARKER,
    aplicar_ipca_vetorizado,
    carregar_operacoes,
    detectar_header,
    escrever_planilha,
    ordenar,
    totais_por_agente,
)


def _fake_bndes_xlsx(path: Path) -> None:
    """Excel no formato BNDES (linhas de preâmbulo + cabeçalho)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Listagem das operações automáticas contratadas. Período considerado: "])
    ws.append(["01/01/2009 até 31/12/2009"])
    ws.append(["Nota"])
    ws.append(["Valores desembolsados ..."])
    ws.append([""])
    ws.append(
        [
            "Cliente",
            "CPF/CNPJ",
            "UF",
            "Data da contratação",
            "Valor Desembolsado R$ (*)",
            "Prazo - Carência (meses)",
            "Prazo - Amortização (meses)",
            "Instituição Financeira Credenciada",
            "Porte do cliente",
        ]
    )
    ws.append(
        [
            "CLIENTE B",
            "**.*",
            "SP",
            "15/06/2009",
            100.0,
            3,
            12,
            "BANCO ZETA",
            "MICRO",
        ]
    )
    ws.append(
        [
            "CLIENTE A",
            "**.*",
            "RJ",
            "10/01/2009",
            200.0,
            0,
            24,
            "BANCO ALFA",
            "PEQUENA",
        ]
    )
    ws.append(
        [
            "CLIENTE C",
            "**.*",
            "MG",
            "20/02/2009",
            50.0,
            6,
            36,
            "BANCO ALFA",
            "MEDIA",
        ]
    )
    ws.append(
        [
            "CLIENTE D",
            "**.*",
            "PR",
            "01/03/2010",
            80.0,
            0,
            12,
            "BANCO ALFA",
            "MICRO",
        ]
    )
    wb.save(path)


def _ipca_constante(path: Path, inicio: str = "2009-01-01", n: int = 220, taxa: float = 0.5) -> None:
    mes = pd.date_range(inicio, periods=n, freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [taxa] * len(mes)}).to_excel(path, index=False)


def test_marker():
    assert "ipca" in MARKER


def test_detectar_e_carregar(tmp_path: Path):
    xlsx = tmp_path / "fake.xlsx"
    _fake_bndes_xlsx(xlsx)
    assert detectar_header(xlsx) == 5
    df = carregar_operacoes(xlsx)
    assert set(df["_ano"]) == {2009, 2010}
    assert COL_AGENTE in df.columns
    assert df[COL_CLIENTE].tolist()[0]  # não vazio
    # não pegou "Porte do cliente" como cliente
    assert "MICRO" not in df[COL_CLIENTE].tolist()


def test_ordenar_e_totais(tmp_path: Path):
    xlsx = tmp_path / "fake.xlsx"
    _fake_bndes_xlsx(xlsx)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca_constante(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = aplicar_ipca_vetorizado(
        carregar_operacoes(xlsx),
        carregar_ipca(ipca_path),
        datetime(2026, 7, 31),
    )
    y2009 = df[df["_ano"] == 2009].drop(columns=["_ano"])
    ord_ = ordenar(y2009)
    agentes = ord_[COL_AGENTE].tolist()
    assert agentes == sorted(agentes)
    # dentro de BANCO ALFA, datas crescentes
    alfa = ord_[ord_[COL_AGENTE] == "BANCO ALFA"]
    assert list(alfa[COL_DATA]) == sorted(alfa[COL_DATA])

    tot = totais_por_agente(ord_)
    assert set(tot[COL_AGENTE]) == {"BANCO ALFA", "BANCO ZETA"}
    assert float(tot.loc[tot[COL_AGENTE] == "BANCO ALFA", COL_VALOR].iloc[0]) == 250.0
    assert float(tot[COL_IPCA].sum()) > float(tot[COL_VALOR].sum())


def test_escrever_planilha_abas(tmp_path: Path):
    xlsx = tmp_path / "fake.xlsx"
    _fake_bndes_xlsx(xlsx)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca_constante(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = aplicar_ipca_vetorizado(
        carregar_operacoes(xlsx),
        carregar_ipca(ipca_path),
        datetime(2026, 7, 31),
    )
    por_ano = {
        int(a): g.drop(columns=["_ano"]) for a, g in df.groupby("_ano", sort=True)
    }
    out = tmp_path / "out.xlsx"
    escrever_planilha(por_ano, out, data_ref=datetime(2026, 7, 31), mes_ipca_ref=pd.Timestamp("2026-07-01"))
    wb = load_workbook(out, read_only=True)
    assert "Capa" in wb.sheetnames
    assert "2009" in wb.sheetnames
    assert "2010" in wb.sheetnames
    assert "Resumo_Anual" in wb.sheetnames
    rows = list(wb["2009"].iter_rows(values_only=True))
    assert rows[0][0] == COL_AGENTE
    assert rows[0][6] == COL_IPCA
    # ordenado: ALFA antes de ZETA
    assert rows[1][0] == "BANCO ALFA"
    # totais no rodapé
    textos = [r[0] for r in rows if r[0]]
    assert any(isinstance(t, str) and t.startswith("TOTAIS POR AGENTE") for t in textos)
    assert "TOTAL GERAL" in textos
