"""Testes da simulação DBGG com Selic = IPCA do ano + spread."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from scripts.simular_dbgg_selic_ipca import (
    agregar_anual,
    anexar_pib,
    carregar_dbgg,
    escrever_discriminativo,
    escrever_markdown,
    gravar_saidas,
    ipca_por_ano,
    ler_aba_indexadores,
    montar_discriminativo,
    processar,
    projetar_meses_apos_oficial,
    reconstruir_pre_oficial,
    simular_parcela_selic,
    taxas_mensais_cf,
)


def _aba_indexadores(ws, linhas: list[tuple]) -> None:
    """Cabeçalho solto + linhas (ano, mês, selic, total, demais=0)."""
    ws["A1"] = "Dívida Bruta do Governo Geral por indexador"
    ws["B7"] = None
    ws["C7"] = "Cambial"
    # linha 9 = nomes como no BCB
    ws["B9"] = None
    start = 12
    for i, (ano, mes, selic, total, juros_ou_emissao) in enumerate(linhas):
        r = start + i
        ws.cell(r, 1, ano)
        ws.cell(r, 2, mes)
        # C–I e K–N = 0; J = selic; O = total
        for c in range(3, 16):
            ws.cell(r, c, 0.0)
        ws.cell(r, 10, selic)  # J
        ws.cell(r, 15, total)  # O
        if juros_ou_emissao is not None:
            pass  # selic já é o valor da coluna J


def _planilha_minima(path: Path) -> Path:
    """Dez/2006 + jan–fev/2007 com identidade estoque = prev + emissão + juros."""
    # Estoque: Dez 1000/2000; Jan 1060/2070; Fev 1100/2120
    # Jan: emissão 40, juros 20 → 1000+40+20=1060
    # Fev: emissão 10, juros 30 → 1060+10+30=1100
    wb = Workbook()
    # openpyxl cria Sheet; renomeamos e criamos as 3 abas
    ws0 = wb.active
    ws0.title = "DividaR$"
    for nome in ("JurosR$", "PrimarioR$"):
        wb.create_sheet(nome)

    _aba_indexadores(
        wb["DividaR$"],
        [
            (2006, "Dez", 1000.0, 2000.0, None),
            (2007, "Jan", 1060.0, 2070.0, None),
            (None, "Fev", 1100.0, 2120.0, None),
        ],
    )
    _aba_indexadores(
        wb["JurosR$"],
        [
            (2007, "Jan", 20.0, 25.0, None),
            (None, "Fev", 30.0, 35.0, None),
        ],
    )
    _aba_indexadores(
        wb["PrimarioR$"],
        [
            (2007, "Jan", 40.0, 45.0, None),
            (None, "Fev", 10.0, 15.0, None),
        ],
    )
    wb.save(path)
    return path


def test_ler_aba_propaga_ano_e_coluna_selic(tmp_path: Path):
    path = _planilha_minima(tmp_path / "dbgg.xlsx")
    est = ler_aba_indexadores(path, "DividaR$")
    assert list(est["mes"]) == [
        pd.Timestamp("2006-12-01"),
        pd.Timestamp("2007-01-01"),
        pd.Timestamp("2007-02-01"),
    ]
    assert est.loc[est["mes"] == "2007-02-01", "ano"].iloc[0] == 2007
    assert float(est.loc[est["mes"] == "2006-12-01", "selic"].iloc[0]) == 1000.0
    assert float(est.loc[est["mes"] == "2007-02-01", "total"].iloc[0]) == 2120.0


def test_ler_aba_aceita_dec_ingles():
    from scripts.simular_dbgg_selic_ipca import MESES_PT

    assert MESES_PT["Dec"] == 12
    assert MESES_PT["Dez"] == 12


def test_carregar_dbgg_tres_abas(tmp_path: Path):
    path = _planilha_minima(tmp_path / "dbgg.xlsx")
    est, ju, em = carregar_dbgg(path)
    assert len(est) == 3
    assert len(ju) == 2
    assert float(ju.iloc[0]["selic"]) == 20.0
    assert float(em.iloc[1]["selic"]) == 10.0


def test_ipca_por_ano_e_taxa_cf_prorrata():
    ipca = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [1.0, 1.0],  # 1% a.m.
        }
    )
    meses = pd.Series(pd.to_datetime(["2007-01-01", "2007-02-01"]))
    anos = ipca_por_ano(ipca, meses)
    fator = 1.01 * 1.01
    assert abs(float(anos.iloc[0]["ipca_acum_pct"]) - (fator - 1) * 100) < 1e-9
    assert int(anos.iloc[0]["n_meses"]) == 2

    cf = taxas_mensais_cf(anos, meses, spread_pp=0.37)
    # spread no semestre de 2 meses: 0.37 * 2/12
    esperado_acum = (fator - 1) * 100 + 0.37 * (2 / 12)
    assert abs(float(cf.iloc[0]["selic_cf_acum_pct"]) - esperado_acum) < 1e-9
    r_m = ((1 + esperado_acum / 100) ** (1 / 2) - 1) * 100
    assert abs(float(cf.iloc[0]["selic_cf_am"]) - r_m) < 1e-9
    assert float(cf.iloc[0]["selic_cf_am"]) == float(cf.iloc[1]["selic_cf_am"])


def test_anos_observados_usam_selic_oficial():
    ipca = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2020-01-01", "2020-02-01"]),
            "valor": [0.5, 0.5],
        }
    )
    selic = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2020-01-01", "2020-02-01"]),
            "valor": [0.3, 0.4],
        }
    )
    meses = pd.Series(selic["mes"])
    cf = taxas_mensais_cf(
        ipca_por_ano(ipca, meses),
        meses,
        0.37,
        selic=selic,
        anos_observados=(2020,),
    )
    assert list(cf["selic_cf_am"].round(8)) == [0.3, 0.4]
    assert not bool(cf["selic_alterada"].iloc[0])
    acum = ((1.003 * 1.004) - 1) * 100
    assert abs(float(cf.iloc[0]["selic_cf_acum_pct"]) - acum) < 1e-9


def test_reconstruir_pre_oficial_emenda_no_estoque():
    of = pd.DataFrame(
        [
            {
                "mes": pd.Timestamp("2006-12-01"),
                "ano": 2006,
                "mes_n": 12,
                "selic": 100.0,
                "total": 200.0,
            }
        ]
    )
    selic = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2006-11-01", "2006-12-01"]),
            "valor": [1.0, 1.0],
        }
    )
    dpmfi = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2006-11-01", "2006-12-01"]),
            "valor": [80.0, 80.0],
        }
    )
    share = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2006-11-01", "2006-12-01"]),
            "valor": [50.0, 50.0],
        }
    )
    pib = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2006-11-01", "2006-12-01"]),
            "valor": [1000.0, 1000.0],
        }
    )
    dbgg_pib = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2006-11-01", "2006-12-01"]),
            "valor": [20.0, 20.0],
        }
    )
    est, ju, em = reconstruir_pre_oficial(
        of, selic, dpmfi, share, pib, dbgg_pib, pd.Timestamp("2006-12-01")
    )
    # proxy dez = 40; escala = 100/40 = 2,5 → nov = 40*2,5 = 100
    assert len(est) == 1
    assert abs(float(est.iloc[0]["selic"]) - 100.0) < 1e-9
    # juros/emissão do mês de emenda (dez/2006)
    assert len(ju) == 1
    assert ju.iloc[0]["mes"] == pd.Timestamp("2006-12-01")


def test_projetar_mes_seguinte_emissao_zero():
    est = pd.DataFrame(
        [_linha("2006-12-01", 100.0, 200.0)]
    )
    ju = pd.DataFrame()
    em = pd.DataFrame()
    selic = pd.DataFrame(
        {"mes": pd.to_datetime(["2007-01-01"]), "valor": [2.0]}
    )
    est2, ju2, em2 = projetar_meses_apos_oficial(
        est, ju, em, selic, pd.Timestamp("2007-01-01")
    )
    assert abs(float(est2.iloc[-1]["selic"]) - 102.0) < 1e-9
    assert abs(float(ju2.iloc[-1]["selic"]) - 2.0) < 1e-9
    assert float(em2.iloc[-1]["selic"]) == 0.0


def _linha(mes: str, selic: float, total: float) -> dict:
    ts = pd.Timestamp(mes)
    return {
        "mes": ts,
        "ano": int(ts.year),
        "mes_n": int(ts.month),
        "selic": selic,
        "total": total,
        "cambial_interna": 0.0,
        "cambial_externa": 0.0,
        "cambial_total": 0.0,
        "igpm": 0.0,
        "igpdi": 0.0,
        "ipca": 0.0,
        "indices_total": 0.0,
        "tjlp_tlp": 0.0,
        "tr": 0.0,
        "prefixado": 0.0,
        "outros": 0.0,
    }


def test_simulacao_reescala_juros_e_composto_estoque(tmp_path: Path):
    path = _planilha_minima(tmp_path / "dbgg.xlsx")
    est, ju, em = carregar_dbgg(path)
    selic = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [2.0, 2.0],  # 2% a.m. observada
        }
    )
    ipca = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [0.5, 0.5],
        }
    )
    meses = pd.Series(selic["mes"])
    cf = taxas_mensais_cf(ipca_por_ano(ipca, meses), meses, spread_pp=0.37)
    r_cf = float(cf.iloc[0]["selic_cf_am"])
    ratio = r_cf / 2.0

    out = simular_parcela_selic(
        est,
        ju,
        em,
        selic,
        cf,
        mes_inicio=pd.Timestamp("2007-01-01"),
        mes_fim=pd.Timestamp("2007-02-01"),
    )
    sim = out[out["selic_am"].notna()].reset_index(drop=True)

    # Jan: S_prev=1000, juros_cf = 20 * ratio * 1
    j1 = 20.0 * ratio
    s1 = 1000.0 + 40.0 + j1  # resíduo 0
    assert abs(float(sim.iloc[0]["juros_selic_cf"]) - j1) < 1e-9
    assert abs(float(sim.iloc[0]["estoque_selic_cf"]) - s1) < 1e-9
    assert abs(float(sim.iloc[0]["residuo_selic"])) < 1e-9
    assert abs(float(sim.iloc[0]["dbgg_cf"]) - (2070.0 - (1060.0 - s1))) < 1e-9

    # Fev: escala também pelo estoque menor
    j2 = 30.0 * ratio * (s1 / 1060.0)
    s2 = s1 + 10.0 + j2
    assert abs(float(sim.iloc[1]["juros_selic_cf"]) - j2) < 1e-9
    assert abs(float(sim.iloc[1]["estoque_selic_cf"]) - s2) < 1e-9
    assert float(sim.iloc[1]["delta_dbgg"]) == float(sim.iloc[1]["delta_estoque_selic"])
    assert float(sim.iloc[1]["dbgg_cf"]) < float(sim.iloc[1]["dbgg_act"])


def test_agregar_anual_ultimo_mes_e_soma_juros(tmp_path: Path):
    path = _planilha_minima(tmp_path / "dbgg.xlsx")
    est, ju, em = carregar_dbgg(path)
    selic = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [2.0, 2.0],
        }
    )
    ipca = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [0.5, 0.5],
        }
    )
    meses = pd.Series(selic["mes"])
    cf = taxas_mensais_cf(ipca_por_ano(ipca, meses), meses, 0.37)
    mensal = simular_parcela_selic(
        est, ju, em, selic, cf,
        mes_inicio=pd.Timestamp("2007-01-01"),
        mes_fim=pd.Timestamp("2007-02-01"),
    )
    mensal = anexar_pib(
        mensal,
        pd.DataFrame(
            {
                "mes": pd.to_datetime(["2006-12-01", "2007-01-01", "2007-02-01"]),
                "valor": [50.0, 51.0, 52.0],
            }
        ),
    )
    anual = agregar_anual(mensal)
    assert len(anual) == 1
    assert int(anual.iloc[0]["n_meses"]) == 2
    assert anual.iloc[0]["mes_final"] == pd.Timestamp("2007-02-01")
    assert abs(float(anual.iloc[0]["juros_selic_act"]) - 50.0) < 1e-9
    assert float(anual.iloc[0]["dbgg_pib_cf"]) < float(anual.iloc[0]["dbgg_pib_act"])


def test_montar_discriminativo_fluxo_e_acumulado():
    anual = pd.DataFrame(
        {
            "ano": [2007, 2008, 2009],
            "n_meses": [12, 12, 12],
            "ipca_acum_pct": [4.0, 5.0, 4.5],
            "selic_acum_pct": [12.0, 13.0, 10.0],
            "selic_cf_acum_pct": [4.37, 5.37, 4.87],
            "juros_selic_act": [100.0, 120.0, 80.0],
            "juros_selic_cf": [40.0, 50.0, 90.0],
            "economia_juros": [60.0, 70.0, -10.0],
            "delta_dbgg": [60.0, 130.0, 120.0],
            "dbgg_act": [1000.0, 1100.0, 1200.0],
            "dbgg_cf": [940.0, 970.0, 1080.0],
        }
    )
    disc = montar_discriminativo(anual)
    assert list(disc["reducao_ano"]) == [60.0, 70.0, -10.0]
    assert list(disc["reducao_acumulada"]) == [60.0, 130.0, 120.0]
    assert abs(float(disc["variacao_delta_dbgg"].iloc[1]) - 70.0) < 1e-9
    assert abs(float(disc["participacao_pct"].sum()) - 100.0) < 1e-9
    assert float(disc.loc[disc["ano"] == 2009, "participacao_pct"].iloc[0]) < 0


def test_escrever_discriminativo_contem_anos_e_total(tmp_path: Path):
    anual = pd.DataFrame(
        {
            "ano": [2007, 2008],
            "n_meses": [12, 12],
            "ipca_acum_pct": [4.46, 5.90],
            "selic_acum_pct": [11.85, 12.48],
            "selic_cf_acum_pct": [4.83, 6.27],
            "juros_selic_act": [63620.0, 81637.0],
            "juros_selic_cf": [25947.0, 38490.0],
            "economia_juros": [37673.0, 43147.0],
            "delta_dbgg": [37673.0, 80820.0],
            "dbgg_act": [1.5e6, 1.7e6],
            "dbgg_cf": [1.46e6, 1.62e6],
        }
    )
    disc = montar_discriminativo(anual)
    path = tmp_path / "disc.md"
    escrever_discriminativo(disc, path, spread_pp=0.37, gerado_em="2026-08-31")
    texto = path.read_text(encoding="utf-8")
    assert "Discriminativo das reduções" in texto
    assert "2007" in texto
    assert "2008" in texto
    assert "Total" in texto
    assert "0,37" in texto


def test_markdown_e_saidas(tmp_path: Path):
    path = _planilha_minima(tmp_path / "dbgg.xlsx")
    est, ju, em = carregar_dbgg(path)
    selic = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [2.0, 2.0],
        }
    )
    ipca = pd.DataFrame(
        {
            "mes": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "valor": [0.5, 0.5],
        }
    )
    meses = pd.Series(selic["mes"])
    cf = taxas_mensais_cf(ipca_por_ano(ipca, meses), meses, 0.37)
    mensal = anexar_pib(
        simular_parcela_selic(
            est, ju, em, selic, cf,
            mes_inicio=pd.Timestamp("2007-01-01"),
            mes_fim=pd.Timestamp("2007-02-01"),
        ),
        None,
    )
    anual = agregar_anual(mensal)
    md = tmp_path / "r.md"
    escrever_markdown(
        anual,
        mensal,
        md,
        spread_pp=0.37,
        gerado_em="2026-08-31",
        fonte_planilha=str(path),
    )
    texto = md.read_text(encoding="utf-8")
    assert "Selic anual = IPCA do ano" in texto
    assert "0,37" in texto
    assert "Metodologia" in texto
    assert "2007" in texto
    assert "Discriminativo das reduções" in texto
    assert "2020–2021" in texto

    md_obs = tmp_path / "r_obs.md"
    escrever_markdown(
        anual,
        mensal,
        md_obs,
        spread_pp=0.37,
        gerado_em="2026-08-31",
        fonte_planilha=str(path),
        periodo="Jan/2003 a Jul/2026",
        anos_observados=(2020, 2021),
    )
    texto_obs = md_obs.read_text(encoding="utf-8")
    assert "Jan/2003 a Jul/2026" in texto_obs
    assert "2020, 2021" in texto_obs
    assert "permanece a **observada**" in texto_obs
    assert "ciclo de juros reais negativos" not in texto_obs

    saidas = gravar_saidas(
        mensal, anual, tmp_path / "out", spread_pp=0.37, fonte_planilha=str(path)
    )
    assert saidas["xlsx"].exists()
    assert saidas["mensal_csv"].exists()
    assert saidas["md"].exists()
    assert saidas["discriminativo_csv"].exists()
    assert saidas["discriminativo_md"].exists()
    wb = load_workbook(saidas["xlsx"])
    assert "Discriminativo" in wb.sheetnames


def test_processar_offline_com_planilha_e_series(tmp_path: Path):
    planilha = _planilha_minima(tmp_path / "dbgg.xlsx")
    ipca_x = tmp_path / "ipca.xlsx"
    selic_x = tmp_path / "selic.xlsx"
    pd.DataFrame(
        {"Data": pd.to_datetime(["2007-01-01", "2007-02-01"]), "IPCA": [0.5, 0.5]}
    ).to_excel(ipca_x, index=False)
    pd.DataFrame(
        {
            "Data": pd.to_datetime(["2007-01-01", "2007-02-01"]),
            "Taxa Selic mensal - % a.m.": [2.0, 2.0],
        }
    ).to_excel(selic_x, index=False)
    saida = tmp_path / "out"
    mensal, anual = processar(
        saida,
        planilha=planilha,
        ipca_path=ipca_x,
        selic_path=selic_x,
        spread_pp=0.37,
        mes_inicio=pd.Timestamp("2007-01-01"),
        mes_fim=pd.Timestamp("2007-02-01"),
        baixar_pib=False,
        recuar_pre_oficial=False,
        projetar_apos_oficial=False,
        stem="dbgg_selic_ipca_2007_2026",
    )
    assert (saida / "dbgg_selic_ipca_2007_2026.md").exists()
    assert (saida / "dbgg_selic_ipca_2007_2026_discriminativo.md").exists()
    assert float(mensal[mensal["selic_am"].notna()].iloc[-1]["dbgg_cf"]) < 2120.0
    assert int(anual.iloc[0]["ano"]) == 2007
