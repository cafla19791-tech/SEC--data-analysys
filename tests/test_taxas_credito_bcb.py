"""Testes das taxas médias anuais de crédito e do ranking dos 5 maiores bancos."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.taxas_credito_bcb import (
    cruzar_taxas,
    gerar_planilha,
    media_anual,
    taxas_medias_anuais,
    top5_por_grupo,
)


def test_media_anual_simples() -> None:
    s = pd.Series(
        [10.0, 14.0, 12.0],
        index=pd.to_datetime(["2020-01-01", "2020-02-01", "2021-01-01"]),
    )
    anual = media_anual(s)
    assert abs(float(anual.loc[2020]) - 12.0) < 1e-12
    assert abs(float(anual.loc[2021]) - 12.0) < 1e-12


def test_taxas_medias_anuais_colunas_de_ano() -> None:
    mensal = pd.DataFrame(
        {
            "data": pd.to_datetime(["2001-01-01", "2001-02-01", "2011-03-01"]),
            "taxa": [40.0, 50.0, 25.0],
            "codigo": [20741, 20741, 20741],
            "modalidade": ["PF — cheque especial"] * 3,
            "segmento": ["PF"] * 3,
            "origem": ["Livres"] * 3,
        }
    )
    anual = taxas_medias_anuais(mensal)
    assert abs(float(anual.loc[0, "2001"]) - 45.0) < 1e-12
    assert abs(float(anual.loc[0, "2011"]) - 25.0) < 1e-12
    assert pd.isna(anual.loc[0, "2002"])


def test_top5_por_grupo_so_conglomerado_prudencial() -> None:
    ifdata = pd.DataFrame(
        {
            "CodInst": ["C1", "C1", "X9", "C2"],
            "Grupo": ["Veículos"] * 4,
            "NomeColuna": ["Total"] * 4,
            "Saldo": [100.0, 5.0, 80.0, 90.0],
        }
    )
    cadastro = pd.DataFrame(
        {
            "CodInst": ["C1", "C2", "X9", "12345678"],
            "NomeInstituicao": ["ITAU - PRUDENCIAL", "BB - PRUDENCIAL", "Filial", "ITAU SA"],
            "CodConglomeradoPrudencial": ["C1", "C2", "C1", "C1"],
            "CnpjInstituicaoLider": ["60701190", "00000000", "60701190", "60701190"],
        }
    )
    top = top5_por_grupo(ifdata, cadastro, "PF")
    assert list(top["instituicao"]) == ["ITAU - PRUDENCIAL", "BB - PRUDENCIAL"]
    assert list(top["rank"]) == [1, 2]


def test_cruzar_taxas_mediana_do_conglomerado() -> None:
    top5 = pd.DataFrame(
        [
            {
                "segmento": "PF",
                "modalidade_ifdata": "Veículos",
                "rank": 1,
                "instituicao": "ITAU - PRUDENCIAL",
                "cod_inst": "C1",
                "cnpj_lider": "60701190",
                "cnpjs": ["60701190", "17192451"],
                "saldo": 10.0,
            }
        ]
    )
    olinda = pd.DataFrame(
        {
            "Modalidade": ["Aquisição de veículos - Prefixado"] * 2,
            "cnpj8": ["60701190", "17192451"],
            "InstituicaoFinanceira": ["ITAU UNIBANCO", "ITAU CDC"],
            "TaxaJurosAoAno": [20.0, 30.0],
        }
    )
    out = cruzar_taxas(top5, olinda)
    assert abs(float(out.iloc[0]["taxa_aa"]) - 25.0) < 1e-12


def test_gerar_planilha(tmp_path: Path) -> None:
    anual = pd.DataFrame(
        [
            {
                "codigo": 20741,
                "modalidade": "PF — cheque especial",
                "segmento": "PF",
                "origem": "Livres",
                "inicio": "01/2001",
                "fim": "07/2026",
                **{str(a): (148.0 if a == 2001 else float("nan")) for a in range(2001, 2027)},
            }
        ]
    )
    top = pd.DataFrame(
        [
            {
                "segmento": "PF",
                "modalidade_ifdata": "Veículos",
                "rank": 1,
                "instituicao": "ITAU - PRUDENCIAL",
                "saldo": 1e9,
                "modalidade_olinda": "Aquisição de veículos - Prefixado",
                "if_olinda": "ITAU UNIBANCO",
                "taxa_aa": 22.5,
            }
        ]
    )
    path = tmp_path / "credito.xlsx"
    gerar_planilha(anual, top, 202603, "2026-07", path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Notas", "Taxas_anuais", "Top5_bancos"]
    assert wb["Taxas_anuais"]["B2"].value == "PF — cheque especial"
    assert wb["Top5_bancos"]["D2"].value == "ITAU - PRUDENCIAL"
