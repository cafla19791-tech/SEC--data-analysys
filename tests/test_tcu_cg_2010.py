"""Testes das tabelas TCU Contas do Governo 2010 e da atualização IPCA."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.build_tcu_cg_2010 import (
    build,
    df_beneficios,
    df_indicadores,
    df_pac_deson,
    df_pac_eixo,
    df_resumo_ipca,
    fator_dez2010_ref,
)
from scripts.analisar_base_monetaria_tcu import df_detalhe_2009_2010, df_fatores
from scripts.analisar_reservas_agregados_tcu import df_agregados, df_quadro, df_reservas
from scripts.cotejar_selic_base_tcu import df_cotejamento
from scripts.tcu_cg_2010_dados import (
    FONTE_URL,
    agregados_monetarios_dezembro,
    autorizacoes_legais,
    beneficios_fin_cred,
    creditos_dlsp,
    fatores_base_monetaria,
    indicadores,
    pac_desoneracoes,
    pac_subsidios_eixo,
    renuncia_previdenciaria,
    renuncia_regional,
    renuncia_tributaria,
    reservas_internacionais_liquidez,
    selic_anual,
    selic_na_data,
)


def _ipca_constante(taxa: float = 0.5) -> pd.DataFrame:
    mes = pd.date_range("2003-01-01", "2026-06-01", freq="MS")
    fator = (1.0 + taxa / 100.0) ** pd.Series(range(1, len(mes) + 1))
    return pd.DataFrame({"mes": mes, "valor": taxa, "fator": fator.to_numpy()})


def test_totais_oficiais():
    trib = next(r for r in renuncia_tributaria() if r["tributo"] == "Total")
    assert abs(trib["y2010"] - 105_843.31) < 1e-6
    assert abs(trib["y2006"] - 65_397.52) < 1e-6

    ir = next(r for r in renuncia_tributaria() if r["tributo"] == "IR — total")
    partes = [
        r
        for r in renuncia_tributaria()
        if r["tributo"] in {"IR — Pessoa Física", "IR — Pessoa Jurídica", "IR — Retido na Fonte"}
    ]
    assert abs(sum(r["y2010"] for r in partes) - ir["y2010"]) < 0.02

    prev = next(r for r in renuncia_previdenciaria() if r["item"] == "Total")
    assert abs(prev["y2010"] - 19_246.0) < 1e-6

    fin = next(r for r in beneficios_fin_cred() if r["item"].startswith("Total"))
    assert abs(fin["y2010"] - 18_877.65) < 1e-6
    assert abs(fin["y2009"] - 16_901.39) < 1e-6

    reg = next(r for r in renuncia_regional() if r["regiao"] == "Total")
    assert abs(reg["total"] - 143.97) < 1e-6
    assert abs(reg["tributarios"] + reg["trib_prev"] + reg["fin_cred"] - 143.97) < 0.02

    pac = next(r for r in pac_desoneracoes() if r["medida"] == "Total")
    partes_pac = [r for r in pac_desoneracoes() if r["medida"] != "Total"]
    assert abs(sum(r["projecao_2010"] for r in partes_pac) - pac["projecao_2010"]) < 0.01
    assert abs(pac["projecao_2010"] - 23_318.0) < 1e-6

    eixo = next(r for r in pac_subsidios_eixo() if r["eixo"] == "Total")
    partes_eixo = [r for r in pac_subsidios_eixo() if r["eixo"] != "Total"]
    assert abs(sum(r["desembolsos"] for r in partes_eixo) - eixo["desembolsos"]) < 0.02
    assert abs(sum(r["contratacoes"] for r in partes_eixo) - eixo["contratacoes"]) < 0.02


def test_creditos_bndes_e_autorizacoes():
    bndes = next(r for r in creditos_dlsp() if r["item"] == "Créditos junto ao BNDES")
    assert bndes["v2009"] == 129_237
    assert bndes["v2010"] == 236_723
    assert abs(bndes["pib2010"] - 6.47) < 1e-9

    ifo = next(
        r for r in creditos_dlsp() if r["item"].startswith("Créditos concedidos a inst")
    )
    assert ifo["v2010"] - ifo["v2009"] == 256_602 - 144_787

    tetos = {r["norma"]: r["limite_r_bi"] for r in autorizacoes_legais()}
    assert tetos["Lei 11.948/2009, art. 1º"] == 100.0
    assert tetos["Lei 12.249/2010, art. 44"] == 180.0
    assert tetos["MP 505/2010 (Lei 12.397/2011)"] == 30.0

    custo = next(
        r for r in indicadores() if "Selic − TJLP" in r["indicador"] or "Selic - TJLP" in r["indicador"]
    )
    assert custo["valor_2010"] == 14_200.0


def test_fator_e_ipca_aplicado():
    ipca = _ipca_constante(0.5)
    fator = fator_dez2010_ref(ipca, data_ref=datetime(2026, 6, 30))
    assert fator > 1.0
    # dez/2010 → jun/2026 = 186 meses de variação após o mês-base
    # fator_ipca_entre usa razão dos fatores acumulados
    resumo = df_resumo_ipca(fator, datetime(2026, 6, 30))
    linha = resumo.loc[
        resumo["indicador"].str.contains("Créditos União"), "ipca_jun2026_r_mi"
    ].iloc[0]
    assert abs(linha - 236_723.0 * fator) < 1e-6

    ind = df_indicadores(fator)
    cred = ind.loc[ind["indicador"].str.contains("Créditos da União junto ao BNDES")].iloc[0]
    assert abs(cred["valor_2010_ipca_jun2026_r_mi"] - 236_723.0 * fator) < 1e-6


def test_build_gera_xlsx_e_md(tmp_path: Path):
    ipca = _ipca_constante(0.4)
    xlsx = tmp_path / "TCU_CG_2010.xlsx"
    md = tmp_path / "TCU_CG_2010_RELATORIO.md"
    p_xlsx, p_md, fator = build(
        ipca,
        data_ref=datetime(2026, 6, 30),
        xlsx=xlsx,
        md=md,
    )
    assert p_xlsx.exists()
    assert p_md.exists()
    assert fator > 1.0

    wb = load_workbook(p_xlsx)
    esperadas = {
        "Fonte",
        "Indicadores",
        "Creditos_DLSP",
        "Autorizacoes_Legais",
        "PAC_Subsidios_Eixo",
        "Resumo_IPCA",
        "Beneficios_Fin_Cred",
        "Base_Monetaria",
        "Base_Monetaria_Acum",
        "Selic_Copom",
        "Selic_Anual",
        "Cotejamento_Selic_Base",
        "Reservas_Internacionais",
        "Agregados_M1_M4",
        "Selic_IPCA_2003_2016",
    }
    assert esperadas <= set(wb.sheetnames)
    assert (tmp_path / "TCU_CG_2010_BASE_MONETARIA.md").exists()
    assert (tmp_path / "grafico_base_monetaria_2003_2010.png").exists()
    assert (tmp_path / "TCU_CG_2010_SELIC_BASE.md").exists()
    assert (tmp_path / "grafico_selic_base_monetaria_2003_2010.png").exists()
    assert (tmp_path / "TCU_CG_2010_RESERVAS_M1M4.md").exists()
    assert (tmp_path / "grafico_reservas_agregados_2002_2010.png").exists()
    assert (tmp_path / "TCU_CG_2010_SELIC_BP_2003_2016.md").exists()
    assert (tmp_path / "grafico_selic_bp_1999_2016.png").exists()

    texto = p_md.read_text(encoding="utf-8")
    assert "236,72" in texto
    assert "14,20" in texto
    assert "tcu.gov.br" in texto
    assert FONTE_URL in texto
    assert "Lei 11.948" in texto


def test_fatores_base_monetaria_identidade():
    rows = fatores_base_monetaria()
    assert [r["ano"] for r in rows] == list(range(2003, 2011))
    assert rows[0]["ano"] == 2003
    y2010 = next(r for r in rows if r["ano"] == 2010)
    assert y2010["titulos_publicos"] == 249_513
    assert y2010["demais_operacoes"] == -233_082
    assert y2010["var_base"] == 40_780

    df = df_fatores()
    assert df["identidade_ok"].all()
    assert abs(df["residuo"]).max() <= 1.0
    # Tesouro é contracionista em todos os anos
    assert (df["tesouro_nacional"] < 0).all()
    # 2008 é o único ano de contração do setor externo
    assert list(df.loc[df["setor_externo"] < 0, "ano"]) == [2008]

    det = df_detalhe_2009_2010()
    d2010 = det.loc[det["ano"] == 2010].iloc[0]
    demais = (
        d2010["depositos_inst_financ"]
        + d2010["derivativos_ajustes"]
        + d2010["outras_contas_ajustes"]
    )
    assert abs(demais - (-233_082)) < 1e-6


def test_selic_copom_e_cotejamento():
    assert selic_na_data(date(2003, 1, 1)) == 25.00
    assert selic_na_data(date(2003, 2, 20)) == 26.50
    assert selic_na_data(date(2009, 7, 23)) == 8.75
    assert selic_na_data(date(2010, 4, 28)) == 8.75
    assert selic_na_data(date(2010, 6, 10)) == 10.25
    assert selic_na_data(date(2010, 12, 31)) == 10.75
    assert selic_na_data(date(2011, 1, 20)) == 11.25

    anual = {r["ano"]: r for r in selic_anual()}
    assert anual[2003]["selic_fim"] == 16.50
    assert anual[2003]["delta_pp"] == -8.50
    assert anual[2003]["selic_max"] == 26.50
    assert anual[2005]["selic_max"] == 19.75
    assert anual[2009]["selic_min"] == 8.75
    assert anual[2009]["delta_pp"] == -5.00
    assert anual[2010]["selic_ini"] == 8.75
    assert anual[2010]["selic_fim"] == 10.75
    assert anual[2010]["delta_pp"] == 2.00
    assert anual[2010]["sentido"] == "alta"
    assert abs(anual[2010]["selic_media"] - 9.90) < 0.05

    cruz = df_cotejamento()
    y2010 = cruz.loc[cruz["ano"] == 2010].iloc[0]
    assert y2010["titulos_publicos"] == 249_513
    assert y2010["demais_operacoes"] == -233_082
    assert y2010["var_base"] == 40_780
    assert "compulsório" in y2010["instrumento_dominante"].lower()
    y2007 = cruz.loc[cruz["ano"] == 2007].iloc[0]
    assert y2007["titulos_publicos"] == -73_974
    assert y2007["setor_externo"] == 155_390
    assert y2007["sentido_selic"] == "queda"


def test_reservas_e_agregados_m1_m4():
    res = {r["ano"]: r for r in reservas_internacionais_liquidez()}
    assert res[2002]["reservas_usd_mi"] == 37_823
    assert res[2003]["reservas_usd_mi"] == 49_296
    assert res[2007]["reservas_usd_mi"] == 180_334
    assert res[2010]["reservas_usd_mi"] == 288_575
    assert res[2007]["reservas_usd_mi"] / res[2003]["reservas_usd_mi"] > 3.0

    agg = {r["ano"]: r for r in agregados_monetarios_dezembro()}
    assert agg[2003]["base"] - agg[2002]["base"] == -83
    assert agg[2010]["base"] - agg[2009]["base"] == 40_780
    assert abs((agg[2007]["base"] - agg[2006]["base"]) - 25_516) <= 1
    assert agg[2002]["m1"] < agg[2002]["m2"] < agg[2002]["m3"] < agg[2002]["m4"]
    assert agg[2010]["m1"] == 287_739
    assert agg[2010]["m4"] == 2_976_783
    # M1 cai em 2008; M2 e M3 continuam a subir
    assert agg[2008]["m1"] < agg[2007]["m1"]
    assert agg[2008]["m2"] > agg[2007]["m2"]

    df_r = df_reservas()
    y10 = df_r.loc[df_r["ano"] == 2010].iloc[0]
    assert abs(y10["reservas_r_mi"] - 288_575 * 1.6662) < 1.0

    cruz = df_quadro()
    y07 = cruz.loc[cruz["ano"] == 2007].iloc[0]
    assert y07["setor_externo"] == 155_390
    assert y07["titulos_publicos"] == -73_974
    assert y07["compromissadas"] == 165_813
    assert df_agregados()["m3"].is_monotonic_increasing


def test_selic_bp_2003_2016():
    from scripts.selic_bp_2003_2016_dados import (
        balanca_comercial,
        reservas_dezembro,
        selic_ipca_anual,
        transacoes_correntes,
    )

    bc = {r["ano"]: r["usd_mi"] for r in balanca_comercial()}
    assert bc[2000] < 0
    assert bc[2001] > 0
    assert sum(bc[y] for y in range(2001, 2007)) > 150_000
    assert bc[2006] > bc[2001]

    tc = {r["ano"]: r["usd_mi"] for r in transacoes_correntes()}
    assert tc[2003] > 0
    assert tc[2006] > 0
    assert tc[2008] < 0
    assert tc[2014] < -100_000

    rs = {r["ano"]: r["usd_mi"] for r in reservas_dezembro()}
    assert rs[2000] == 33_011
    assert rs[2007] == 180_334
    assert rs[2010] == 288_575

    si = {r["ano"]: r for r in selic_ipca_anual()}
    assert abs(si[2016]["selic_acum_pct"] - 459.74) < 0.01
    assert abs(si[2016]["ipca_acum_pct"] - 134.13) < 0.01
    assert si[2016]["selic_acum_pct"] > si[2016]["ipca_acum_pct"]
    assert si[2010]["ipca_pct"] == 5.91


def test_dataframes_auxiliares():
    fator = 2.0
    assert abs(df_beneficios(fator)["y2010_ipca_jun2026"].iloc[-1] - 18_877.65 * 2) < 1e-6
    assert abs(df_pac_deson(fator)["projecao_2010_ipca_jun2026"].iloc[-1] - 46_636.0) < 1e-6
    assert abs(df_pac_eixo(fator)["desembolsos"].iloc[-1] - 4_444.24) < 1e-6


def test_simulacao_dbgg_conta_unica():
    from scripts.sim_dbgg_conta_unica_dados import (
        DBGG_DEZ_R_MI,
        INDIRETAS_R_MI,
        PARTICIPACOES_R_MI,
        RENUNCIA_2015_EFETIVA,
        TESOURO_CAPTACAO_R_MI,
        renuncia_reconstruida,
    )
    from scripts.simular_dbgg_conta_unica import df_fluxos, df_simulacao

    assert abs(INDIRETAS_R_MI[2010] - 81_538.0) < 1e-6
    assert abs(sum(INDIRETAS_R_MI[a] for a in range(2003, 2016)) - 658_375.0) < 1.0
    assert PARTICIPACOES_R_MI[2003] is None
    assert abs(PARTICIPACOES_R_MI[2010] - 25_429.091) < 1e-3
    assert TESOURO_CAPTACAO_R_MI[2009] == 105_000.0
    assert abs(TESOURO_CAPTACAO_R_MI[2010] - 107_100.0) < 1e-6
    assert abs(DBGG_DEZ_R_MI[2009] - 1_973_423.68) < 0.1
    assert abs(DBGG_DEZ_R_MI[2010] - 2_011_521.66) < 0.1

    r15 = renuncia_reconstruida(2015)
    assert abs(r15["desenvolvimento_regional"] - RENUNCIA_2015_EFETIVA["desenvolvimento_regional"]) < 1e-6
    assert abs(r15["inovacao"] - RENUNCIA_2015_EFETIVA["inovacao_lei_bem"]) < 1e-6
    assert abs(r15["imunes_isentas"] - RENUNCIA_2015_EFETIVA["imunes_isentas"]) < 1e-6
    assert r15["regional_ampla"] == r15["desenvolvimento_regional"] + r15["zfm_alc"]
    assert renuncia_reconstruida(2004)["inovacao"] == 0.0
    assert renuncia_reconstruida(2006)["inovacao"] > 0.0

    fluxos = df_fluxos()
    sim = df_simulacao(fluxos)
    y10 = fluxos.loc[fluxos["ano"] == 2010].iloc[0]
    assert abs(y10["fluxo_total"] - (y10["indiretas"] + y10["participacoes_uso"] + y10["renuncia"])) < 1e-6
    y15 = sim.loc[sim["ano"] == 2015].iloc[0]
    assert y15["dbgg_cf_estoque"] < y15["dbgg_oficial"]
    assert y15["dbgg_cf_selic"] < y15["dbgg_cf_estoque"]
    assert abs(y15["acum_fluxo"] - fluxos["fluxo_total"].sum()) < 1e-6
    assert y15["emissao_evitada"] == -y15["fluxo"]
    y03 = sim.loc[sim["ano"] == 2003].iloc[0]
    assert y03["dbgg_oficial"] is None or pd.isna(y03["dbgg_oficial"])
    assert y03["juros_evitados"] == 0.0


def test_bloco_macrometrico():
    from scripts.macrometrica_conta_unica import (
        CARGA_FEDERAL_PIB,
        ELASTICIDADE_RECEITA,
        MU_BNDES,
        MU_TRIBUTO,
        df_dbgg_macrometrica,
        df_macrometrica,
    )
    from scripts.simular_dbgg_conta_unica import df_fluxos, df_simulacao

    mm = df_macrometrica()
    y15 = mm.loc[mm["ano"] == 2015].iloc[0]
    assert y15["delta_pib"] < 0
    assert y15["receita_ciclo"] < 0
    assert y15["renuncia_liquida"] < y15["renuncia_estatica"]
    assert y15["fluxo_macrometrico"] < y15["fluxo_estatico"]
    assert abs(y15["vazamento"] - (y15["fluxo_estatico"] - y15["fluxo_macrometrico"])) < 1e-6
    esperado_dy = -MU_TRIBUTO * y15["renuncia_estatica"] - MU_BNDES * y15["bndes"]
    assert abs(y15["delta_pib"] - esperado_dy) < 1e-6
    esperado_ciclo = ELASTICIDADE_RECEITA * CARGA_FEDERAL_PIB * esperado_dy
    assert abs(y15["receita_ciclo"] - esperado_ciclo) < 1e-6

    est = df_simulacao(df_fluxos())
    mm_d = df_dbgg_macrometrica(mm)
    e15 = est.loc[est["ano"] == 2015].iloc[0]
    d15 = mm_d.loc[mm_d["ano"] == 2015].iloc[0]
    assert d15["dbgg_mm_selic"] > e15["dbgg_cf_selic"]
    assert d15["dbgg_mm_selic"] < d15["dbgg_oficial"]
