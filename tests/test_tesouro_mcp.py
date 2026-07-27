"""Unit tests for tesouro-mcp providers (mocked)."""

from __future__ import annotations

import pytest

from tesouro_mcp import providers


def test_parse_month_formats():
    assert providers._parse_month("01/2024") == "01/2024"
    assert providers._parse_month("2024-01") == "01/2024"
    assert providers._parse_month("2024-01-15") == "01/2024"


def test_resolve_serie_alias_and_code():
    a = providers.resolve_serie("resultado_primario")
    assert a["codigo_serie"] == "10.04.1"
    assert a["tema"] == "10"

    b = providers.resolve_serie("10.01.1")
    assert b["codigo_serie"] == "10.01.1"
    assert b["tema"] == "10"


def test_resolve_serie_unknown():
    with pytest.raises(ValueError, match="desconhecida"):
        providers.resolve_serie("nao_existe_xyz")


def test_normalize_row():
    row = providers._normalize_row(
        {
            "data": "2024-03-01T00:00:00.000Z",
            "valor": 199226.89,
            "codigoTema": "10",
            "nomeSubtema": "Receitas",
            "codigoSerie": "10.01.1",
            "nomeSerie": "Receita Total",
        }
    )
    assert row["date"] == "2024-03-01"
    assert row["value"] == pytest.approx(199226.89)
    assert row["unit"] == "R$ milhoes"


def test_get_serie_mocked(monkeypatch):
    def fake_resultado_fiscal(**kwargs):
        assert kwargs["tema"] == "10"
        assert kwargs["codigo_serie"] == "10.04.1"
        return {
            "tema": "10",
            "count": 1,
            "series": [{"date": "2024-01-01", "value": -10.0}],
        }

    monkeypatch.setattr(providers, "get_resultado_fiscal", fake_resultado_fiscal)
    out = providers.get_serie("resultado_primario", data_inicio="2024-01", data_fim="2024-01")
    assert out["alias"] == "resultado_primario"
    assert out["count"] == 1


def test_paginate_follows_next(monkeypatch):
    calls = []

    def fake_json(url, **kwargs):
        calls.append(url)
        if "page=2" in url:
            return {"registros": [{"codigoSerie": "b"}], "next": None, "status": "ok"}
        return {
            "registros": [{"codigoSerie": "a"}],
            "next": "https://apiapex.tesouro.gov.br/aria//v1/series-temporais/custom/series?page=2",
            "status": "ok",
        }

    monkeypatch.setattr(providers, "_get_json", fake_json)
    rows = providers._paginate("https://example/series")
    assert len(rows) == 2
    assert "aria/v1" in calls[1]
    assert "aria//v1" not in calls[1]


def test_cli_help():
    from tesouro_mcp.cli import build_parser

    help_text = build_parser().format_help()
    assert "serie" in help_text
    assert "headline" in help_text
    assert "ckan-show" in help_text
    assert "coletar-anual" in help_text


def test_load_overlay_csv(tmp_path):
    from tesouro_mcp import collector

    p = tmp_path / "dgt.csv"
    p.write_text(
        "ano,renuncia_desenv_regional_R$bi,renuncia_imunes_isentas_R$bi,"
        "renuncia_automotivo_R$bi,renuncia_cultura_audiovisual_R$bi,"
        "renuncia_inovacao_R$bi\n"
        "2020,45.2,12.1,,,8.3\n"
        "2021,,,,,n/d\n"
        "2022,1.5,,,,\n",
        encoding="utf-8",
    )
    out = collector.load_overlay_csv(p, collector.DGT_COLUMNS)
    assert out[2020]["renuncia_desenv_regional_R$bi"] == pytest.approx(45.2)
    assert out[2020]["renuncia_inovacao_R$bi"] == pytest.approx(8.3)
    assert 2021 not in out
    assert out[2022]["renuncia_desenv_regional_R$bi"] == pytest.approx(1.5)


def test_rows_to_csv_empty_cells():
    from tesouro_mcp import collector

    text = collector.rows_to_csv(
        [
            {
                "ano": 2020,
                "dbgg_01jan_R$bi": 1.5,
                "resultado_primario_R$bi": None,
            }
        ]
    )
    assert "ano,dbgg_01jan" in text
    assert "2020,1.5," in text


def test_collect_annual_table_mocked(monkeypatch, tmp_path):
    from datetime import date

    from tesouro_mcp import bcb_client, collector

    monkeypatch.setattr(collector, "_rtn_annual", lambda code, a, b: {2020: 1500.0})
    monkeypatch.setattr(
        bcb_client,
        "fetch_sgs_range",
        lambda code, start, end, **kw: {
            date(2019, 12, 1): 5000000.0,
            date(2020, 12, 1): 5500000.0,
            date(2020, 6, 1): 100.0,
        },
    )
    monkeypatch.setattr(
        bcb_client,
        "december_stocks",
        lambda points: {d.year: v for d, v in points.items() if d.month == 12},
    )
    monkeypatch.setattr(
        bcb_client,
        "annual_sum",
        lambda points: {2020: 1200.0},
    )
    monkeypatch.setattr(collector, "download_emissoes_xlsx", lambda: b"fake")
    monkeypatch.setattr(
        collector,
        "parse_emissoes_resgates_xlsx",
        lambda content: ({2020: 800000.0}, {2020: 700000.0}),
    )

    dgt = tmp_path / "dgt.csv"
    dgt.write_text(
        "ano,renuncia_desenv_regional_R$bi,renuncia_imunes_isentas_R$bi,"
        "renuncia_automotivo_R$bi,renuncia_cultura_audiovisual_R$bi,"
        "renuncia_inovacao_R$bi\n"
        "2020,10,,,,\n",
        encoding="utf-8",
    )
    fundos = tmp_path / "fundos.csv"
    fundos.write_text(
        "ano,financ_FNO_BASA_R$bi,financ_FNE_BNB_R$bi,financ_FCO_BB_R$bi\n"
        "2020,1.1,2.2,3.3\n",
        encoding="utf-8",
    )

    table = collector.collect_annual_table(
        year_from=2020,
        year_to=2020,
        dgt_csv=dgt,
        fundos_csv=fundos,
        include_emissoes=True,
    )
    assert table["count"] == 1
    row = table["rows"][0]
    assert row["ano"] == 2020
    assert row["dbgg_01jan_R$bi"] == pytest.approx(5000.0)
    assert row["dbgg_31dez_R$bi"] == pytest.approx(5500.0)
    assert row["resultado_primario_R$bi"] == pytest.approx(1.5)
    assert row["emissoes_DPF_R$bi"] == pytest.approx(800.0)
    assert row["resgates_DPF_R$bi"] == pytest.approx(700.0)
    assert row["desembolso_BNDES_R$bi"] == pytest.approx(1.2)
    assert row["renuncia_desenv_regional_R$bi"] == pytest.approx(10.0)
    assert row["financ_FCO_BB_R$bi"] == pytest.approx(3.3)


def test_parse_br_month():
    from datetime import date

    from tesouro_mcp.collector import _parse_br_month

    assert _parse_br_month("dez/20") == date(2020, 12, 1)
    assert _parse_br_month("jan/01") == date(2001, 1, 1)
    assert _parse_br_month("invalid") is None


def test_rtn_xlsx_extract_annual(tmp_path):
    from datetime import datetime

    from openpyxl import Workbook

    from tesouro_mcp import rtn_xlsx

    wb = Workbook()
    ws = wb.active
    ws.title = "1.1"
    ws["A2"] = "Tabela 1.1"
    ws["A3"] = "R$ Milhoes - Valores Correntes"
    # header row 5
    ws.cell(5, 1, "Discriminacao")
    for i, (y, m) in enumerate([(2020, 1), (2020, 2), (2020, 3)] + [(2020, m) for m in range(4, 13)], start=2):
        ws.cell(5, i, datetime(y, m, 1))
    # 12 months of 10 each -> annual 120
    ws.cell(6, 1, "1. RECEITA TOTAL 1/")
    ws.cell(66, 1, "5. RESULTADO PRIMÁRIO GOVERNO CENTRAL - ACIMA DA LINHA (3 - 4)")
    ws.cell(74, 1, "9. JUROS NOMINAIS 7/")
    ws.cell(75, 1, "10. RESULTADO NOMINAL DO GOVERNO CENTRAL (8 + 9) 8/")
    for col in range(2, 14):
        ws.cell(6, col, 10.0)
        ws.cell(66, col, -5.0)
        ws.cell(74, col, -2.0)
        ws.cell(75, col, -7.0)

    # minimal 1.1-A clone
    ws_a = wb.create_sheet("1.1-A")
    ws_a["A2"] = "Tabela 1.1-A"
    ws_a["A3"] = "R$ Milhoes - Valores de Mai/2026 - IPCA"
    ws_a.cell(5, 1, "Discriminacao")
    for col in range(2, 14):
        ws_a.cell(5, col, datetime(2020, col - 1, 1))
        ws_a.cell(6, col, 20.0)
        ws_a.cell(66, col, -10.0)
        ws_a.cell(74, col, -4.0)
        ws_a.cell(75, col, -14.0)
    ws_a.cell(6, 1, "1. RECEITA TOTAL 1/")
    ws_a.cell(66, 1, "5. RESULTADO PRIMÁRIO GOVERNO CENTRAL - ACIMA DA LINHA (3 - 4)")
    ws_a.cell(74, 1, "9. JUROS NOMINAIS 7/")
    ws_a.cell(75, 1, "10. RESULTADO NOMINAL DO GOVERNO CENTRAL (8 + 9) 8/")

    path = tmp_path / "serie.xlsx"
    wb.save(path)

    out = rtn_xlsx.extract_annual_rtn(path, year_from=2020, year_to=2020, include_fundos=False)
    assert out["count"] == 1
    assert out["rows"][0]["resultado_primario_R$mi"] == pytest.approx(-60.0)
    assert out["rows"][0]["resultado_primario_R$bi"] == pytest.approx(-0.06)

    out_ipca = rtn_xlsx.extract_annual_rtn(
        path, year_from=2020, year_to=2020, constantes_ipca=True, include_fundos=False
    )
    assert out_ipca["sheet"] == "1.1-A"
    assert out_ipca["rows"][0]["resultado_primario_R$mi"] == pytest.approx(-120.0)
