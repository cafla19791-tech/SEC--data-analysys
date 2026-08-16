"""Testes do top 100 diretas 2003–2018 com IPCA."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from scripts.top100_diretas_2003_2018_ipca import (
    COL_CNPJ,
    COL_NOME,
    COL_SOMA,
    MARKER,
    atualizar_ipca,
    carregar_diretas_periodo,
    processar,
    top100_empresas,
)


def _fonte(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SITE"
    ws.append(["t"])
    ws.append(["p"])
    ws.append(["d"])
    ws.append([])
    ws.append(
        [
            "Cliente",
            "CNPJ",
            "Data da contratação",
            "Valor contratado  R$",
            "Valor desembolsado R$",
            "Forma de apoio",
        ]
    )
    # dentro do período, DIRETA — empresa A maior
    ws.append(["EMPRESA A SA", "11222333000181", datetime(2010, 1, 15), 1000, 1000, "DIRETA"])
    ws.append(["EMPRESA A SA", "11222333000181", datetime(2012, 6, 1), 500, 500, "DIRETA"])
    ws.append(["EMPRESA B LTDA", "99888777000166", datetime(2005, 3, 1), 200, 200, "DIRETA"])
    # fora: indireta, fora do período, desembolso zero usa contratado
    ws.append(["EMPRESA C", "11111111000111", datetime(2015, 1, 1), 9000, 9000, "INDIRETA"])
    ws.append(["EMPRESA D", "22222222000122", datetime(2002, 1, 1), 9000, 9000, "DIRETA"])
    ws.append(["EMPRESA E", "33333333000133", datetime(2016, 1, 1), 300, 0, "DIRETA"])
    wb.create_sheet("DISCLAIMER")
    wb.create_sheet("DE-PARA CNAE")
    wb.save(path)


def _ipca(path: Path) -> None:
    mes = pd.date_range("2002-01-01", periods=300, freq="MS")
    pd.DataFrame({"Data": mes, "IPCA": [0.5] * len(mes)}).to_excel(path, index=False)


def test_marker():
    assert "top100" in MARKER


def test_filtro_periodo_e_forma(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fonte(fonte)
    df = carregar_diretas_periodo(fonte)
    assert set(df["cnpj"]) == {"11222333000181", "99888777000166", "33333333000133"}
    # E usa contratado (300) pois desembolsado=0
    assert float(df.loc[df["cnpj"] == "33333333000133", "emprestimo"].iloc[0]) == 300.0


def test_top100_ordem(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    from scripts.calcular_diretas_ipca_selic import carregar_ipca

    df = atualizar_ipca(carregar_diretas_periodo(fonte), carregar_ipca(ipca_path))
    ranking = top100_empresas(df, n=100)
    assert list(ranking.columns) == [COL_CNPJ, COL_NOME, COL_SOMA]
    assert ranking.iloc[0][COL_NOME] == "EMPRESA A SA"
    assert ranking.iloc[0][COL_SOMA] > ranking.iloc[1][COL_SOMA]


def test_processar_excel(tmp_path: Path):
    fonte = tmp_path / "nao.xlsx"
    _fonte(fonte)
    ipca_path = tmp_path / "ipca.xlsx"
    _ipca(ipca_path)
    saida = tmp_path / "out.xlsx"
    ranking = processar(
        fonte=fonte,
        saida=saida,
        ipca_path=ipca_path,
        baixar=False,
    )
    assert saida.exists()
    wb = load_workbook(saida, read_only=True)
    assert "Top100" in wb.sheetnames
    assert "Capa" in wb.sheetnames
    header = next(wb["Top100"].iter_rows(min_row=1, max_row=1, values_only=True))
    assert header == (COL_CNPJ, COL_NOME, COL_SOMA)
    assert len(ranking) == 3
